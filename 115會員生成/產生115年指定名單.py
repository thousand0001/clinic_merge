# -*- coding: utf-8 -*-
"""
115年指定名單生成工具

輸入：選會員 xlsx（醫生看分頁 AH欄「114會員名單」打勾的列）
樣板：115指定會員模板.xlsx（放在本程式同一資料夾）
輸出：{診所名}115年指定名單{mmdd_hhmm}.xlsx
      → 存放到輸入檔案上一層的「115會員生成」資料夾
"""
from __future__ import annotations

import datetime
import re
import sys
import tkinter as tk
from copy import copy
from pathlib import Path
from tkinter import filedialog, messagebox
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter

# ── 常數 ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
DOCTOR_SHEET_KEYWORD = "醫生看"
AH_COL_IDX = 33          # 0-based（AH = 第 34 欄）
DATA_START_ROW = 4        # 醫生看資料從第 4 列開始（前 3 列為表頭）
AH_CHECK_VALUES = {"✔", "✓"}
OUTPUT_SUBDIR = "115會員生成"
DATE_NUMBER_FORMAT = r"[$-409]yyyy\-mm\-dd"
TZ_TW = datetime.timezone(datetime.timedelta(hours=8))

# 模板欄（1-based）→ 醫生看欄（0-based）
# 醫生看：A=0身份證, C=2生日, AW=48個案類別, AX=49論質名單,
#          AY=50 65歲多重慢性病, AZ=51高診次, BA=52慢性病, BB=53非慢性病,
#          BC=54與前一年相同, BD=55疾病樣態, BE=56 ASCVD,
#          BF=57三高, BG=58高血壓, BH=59高血脂, BI=60高血糖
COL_MAP: dict[int, int] = {
    2:  0,   # B  ID          ← A  身份證號碼
    3:  2,   # C  BIRTHDAY    ← C  生日
    4:  48,  # D  個案類別    ← AW
    5:  49,  # E  論質名單    ← AX
    6:  50,  # F  65歲以上多重慢性病 ← AY
    7:  51,  # G  高診次      ← AZ
    8:  52,  # H  慢性病      ← BA
    9:  53,  # I  非慢性病    ← BB
    10: 54,  # J  與前一年相同 ← BC
    11: 55,  # K  疾病樣態    ← BD
    12: 56,  # L  ASCVD       ← BE
    13: 57,  # M  三高        ← BF
    14: 58,  # N  高血壓      ← BG
    15: 59,  # O  高血脂      ← BH
    16: 60,  # P  高血糖      ← BI
}


# ── 工具函式 ──────────────────────────────────────────────────────────────────
def pick_file(title: str, filetypes: list) -> Optional[Path]:
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    root.update()
    path = filedialog.askopenfilename(title=title, filetypes=filetypes)
    root.destroy()
    return Path(path) if path else None


def find_template() -> Optional[Path]:
    """在程式所在資料夾尋找含「模板」的 xlsx，找不到就讓使用者選。"""
    candidates = list(SCRIPT_DIR.glob("*模板*.xlsx")) + list(SCRIPT_DIR.glob("*template*.xlsx"))
    if candidates:
        return candidates[0]
    return pick_file(
        title="選擇 115 指定會員模板 Excel 檔案",
        filetypes=[("Excel 檔案", "*.xlsx"), ("所有檔案", "*.*")],
    )


def extract_clinic_name(input_path: Path) -> str:
    """從輸入檔名提取診所名稱（移除「選會員_dddd_dddd」後綴）。"""
    name = re.sub(r"選會員.*$", "", input_path.stem).strip()
    return name if name else input_path.stem


def normalize_date(val) -> Optional[datetime.datetime]:
    """將各種日期格式統一成 datetime.datetime，非日期回傳 None。"""
    if isinstance(val, datetime.datetime):
        return val
    if isinstance(val, datetime.date):
        return datetime.datetime(val.year, val.month, val.day)
    if isinstance(val, str):
        val = val.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
            try:
                return datetime.datetime.strptime(val[:10], fmt[:len(fmt)])
            except ValueError:
                continue
    return None


def copy_cell_style(src_cell, dst_cell):
    """複製儲存格樣式（字型、填色、對齊、框線、數字格式）。"""
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.border = copy(src_cell.border)
        dst_cell.number_format = src_cell.number_format


# ── 主流程 ────────────────────────────────────────────────────────────────────
def main() -> None:
    # 1. 選擇輸入檔（選會員 xlsx）
    input_path = pick_file(
        title="選擇選會員 Excel 檔案",
        filetypes=[("Excel 檔案", "*.xlsx *.xlsm"), ("所有檔案", "*.*")],
    )
    if not input_path:
        print("已取消，未選擇輸入檔案。")
        return

    # 2. 尋找模板
    template_path = find_template()
    if not template_path:
        print("已取消，未選擇模板檔案。")
        return

    print(f"輸入檔：{input_path.name}")
    print(f"模板  ：{template_path.name}")

    # 3. 讀取模板（取得院所ID、欄寬、表頭格式、資料列格式參考）
    tmpl_wb = openpyxl.load_workbook(template_path)
    tmpl_ws = tmpl_wb.worksheets[0]
    clinic_id = tmpl_ws.title  # 院所ID = sheet 名稱
    n_cols = tmpl_ws.max_column

    # 儲存欄寬
    col_widths = {
        col_letter: dim.width
        for col_letter, dim in tmpl_ws.column_dimensions.items()
        if dim.width
    }

    # 儲存第 1 列（表頭）格式，以便清空後還原
    header_formats = []
    for cell in tmpl_ws[1]:
        header_formats.append({
            "font":          copy(cell.font)      if cell.has_style else None,
            "fill":          copy(cell.fill)      if cell.has_style else None,
            "alignment":     copy(cell.alignment) if cell.has_style else None,
            "border":        copy(cell.border)    if cell.has_style else None,
            "number_format": cell.number_format,
        })

    # 儲存第 2 列（資料列）格式參考（用於新寫入資料的欄位格式）
    data_row_formats = []
    if tmpl_ws.max_row >= 2:
        for cell in tmpl_ws[2]:
            data_row_formats.append({
                "font":          copy(cell.font)      if cell.has_style else None,
                "fill":          copy(cell.fill)      if cell.has_style else None,
                "alignment":     copy(cell.alignment) if cell.has_style else None,
                "border":        copy(cell.border)    if cell.has_style else None,
                "number_format": cell.number_format,
            })

    # 清除所有資料列（保留表頭第 1 列）
    if tmpl_ws.max_row >= 2:
        tmpl_ws.delete_rows(2, tmpl_ws.max_row - 1)

    # 還原表頭格式（delete_rows 有時會影響格式）
    for col_idx, fmt in enumerate(header_formats, start=1):
        cell = tmpl_ws.cell(1, col_idx)
        if fmt["font"]:          cell.font          = fmt["font"]
        if fmt["fill"]:          cell.fill          = fmt["fill"]
        if fmt["alignment"]:     cell.alignment     = fmt["alignment"]
        if fmt["border"]:        cell.border        = fmt["border"]
        if fmt["number_format"]: cell.number_format = fmt["number_format"]

    # 4. 讀取選會員檔案，收集 AH 打勾的列
    print("讀取選會員檔案中…")
    src_wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)

    doctor_ws = None
    for name in src_wb.sheetnames:
        if DOCTOR_SHEET_KEYWORD in name:
            doctor_ws = src_wb[name]
            break

    if doctor_ws is None:
        messagebox.showerror("錯誤", f"找不到含「{DOCTOR_SHEET_KEYWORD}」的分頁。")
        src_wb.close()
        tmpl_wb.close()
        return

    selected_rows: list[tuple] = []
    for row in doctor_ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        ah_val = row[AH_COL_IDX] if len(row) > AH_COL_IDX else None
        if ah_val and str(ah_val).strip() in AH_CHECK_VALUES:
            selected_rows.append(row)

    src_wb.close()
    print(f"AH 打勾筆數：{len(selected_rows)}")

    # 5. 將資料寫入模板
    for row_idx, src_row in enumerate(selected_rows, start=2):
        # 逐欄寫值
        for tmpl_col in range(1, n_cols + 1):
            cell = tmpl_ws.cell(row_idx, tmpl_col)

            if tmpl_col == 1:
                # A 欄：院所ID（固定）
                cell.value = clinic_id

            elif tmpl_col == 3:
                # C 欄：BIRTHDAY（日期格式）
                src_idx = COL_MAP.get(tmpl_col)
                raw_val = src_row[src_idx] if src_idx is not None and src_idx < len(src_row) else None
                dt = normalize_date(raw_val)
                cell.value = dt
                cell.number_format = DATE_NUMBER_FORMAT

            elif tmpl_col in COL_MAP:
                src_idx = COL_MAP[tmpl_col]
                cell.value = src_row[src_idx] if src_idx < len(src_row) else None

            # 套用資料列格式（日期欄已單獨處理）
            if data_row_formats and tmpl_col <= len(data_row_formats):
                fmt = data_row_formats[tmpl_col - 1]
                if fmt["font"]:      cell.font      = copy(fmt["font"])
                if fmt["fill"]:      cell.fill      = copy(fmt["fill"])
                if fmt["alignment"]: cell.alignment = copy(fmt["alignment"])
                if fmt["border"]:    cell.border    = copy(fmt["border"])
                # 日期欄格式已設，其餘欄套用模板格式
                if tmpl_col != 3 and fmt["number_format"]:
                    cell.number_format = fmt["number_format"]

    # 6. 還原欄寬
    for col_letter, width in col_widths.items():
        tmpl_ws.column_dimensions[col_letter].width = width

    # 7. 決定輸出路徑
    now = datetime.datetime.now(TZ_TW)
    clinic_name = extract_clinic_name(input_path)
    timestamp = now.strftime("%m%d_%H%M")
    output_filename = f"{clinic_name}115年指定名單{timestamp}.xlsx"

    output_dir = input_path.parent.parent / OUTPUT_SUBDIR
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / output_filename

    # 8. 儲存
    tmpl_wb.save(output_path)
    tmpl_wb.close()

    msg = (
        f"完成！\n"
        f"診所：{clinic_name}\n"
        f"院所ID：{clinic_id}\n"
        f"筆數：{len(selected_rows)}\n"
        f"\n輸出檔案：\n{output_path}"
    )
    print(msg)
    messagebox.showinfo("完成", msg)

    # 點確定後自動開啟輸出檔案
    import subprocess, sys
    if sys.platform == "win32":
        subprocess.Popen(["start", "", str(output_path)], shell=True)
    elif sys.platform == "darwin":
        subprocess.Popen(["open", str(output_path)])
    else:
        subprocess.Popen(["xdg-open", str(output_path)])


if __name__ == "__main__":
    main()
