# -*- coding: utf-8 -*-
"""
耀聖系統前置清洗 + 共用核心包裝（0623）

用途：
- 耀聖 HIS 的 最後就診日.CSV 只有病歷號/姓名/生日，無身分證號。
  本程式掃描來源資料夾建立 姓名+生日 → 身分證 對照表，
  將最後就診日轉為含 ID 的標準格式，供共用核心填入 K 欄（最後就診日）。
- 次數/ 資料夾的月份 CSV 同樣用對照表補齊 ID。
- 預防保健名單.CSV / 預防保健.CSV：
  - 結束日/就診日作為各類預防保健的檢查日期。
  - 預防保健/卡序依 IC 代碼分流至成健/BC肝/子抹/糞篩/老流。
  - 主次代碼/ICD 彙整為主次診斷代碼。
- R11440 資料夾等已含 ID 的原始檔案直接交給共用核心，不另行處理。
- 自動偵測日期最新的選會員共用核心與模板完成輸出。
- 不包「新耀聖」（另有 前置清洗_新耀聖_*.py）。
"""

from __future__ import annotations

import csv
import datetime
import gc
import hashlib
import importlib.util
import getpass
import os
import re
import shutil
import sys
import tempfile
import time
import traceback
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
LAST_RUN_NOTICE = ""

SCREENING_CODE_MAP = {
    "IC3E": ("成人健檢",),
    "IC3D": ("成人健檢",),
    "IC21": ("成人健檢",),
    "IC22": ("成人健檢",),
    "IC23": ("成人健檢",),
    "IC24": ("成人健檢",),
    "IC29": ("肝炎篩檢",),
    "IC31": ("子宮抹片",),
    "IC37": ("子宮抹片",),
    "IC85": ("糞便潛血",),
    "IC01": ("老人流感",),
    "ICL1001": ("成人健檢", "肝炎篩檢"),
    "ICL1002": ("成人健檢", "肝炎篩檢"),
}
SCREENING_SHEETS = ("成人健檢", "子宮抹片", "糞便潛血", "老人流感", "肝炎篩檢")
ZONGBEI_PASSWORD_HINT = "宗北加密 Excel 密碼"


def _safe_move_excel(temp_output: Path, final_output: Path) -> Path:
    """
    Use copy instead of shutil.move on Windows.

    shutil.move may copy successfully and then fail at os.unlink(src) when the
    temp file is still locked, which raises WinError 32 even though the formal
    output already exists. Copy first, then best-effort delete the temp file.
    """
    gc.collect()
    final_output.parent.mkdir(parents=True, exist_ok=True)
    target = final_output
    if target.exists():
        try:
            target.unlink()
        except PermissionError:
            timestamp = datetime.datetime.now().strftime("%H%M%S")
            target = target.with_name(f"{target.stem}_{timestamp}{target.suffix}")

    for attempt in range(10):
        try:
            shutil.copy2(str(temp_output), str(target))
            try:
                temp_output.unlink()
            except PermissionError:
                print(f"  暫存檔仍被 Windows 鎖住，正式輸出已完成：{target}", flush=True)
            return target
        except PermissionError:
            if attempt < 9:
                time.sleep(0.8)
            else:
                raise
HEPATITIS_LAB_CODES = {"L1001C", "L1002C"}


# ─── 共用核心自動偵測 ─────────────────────────────────────────────────────────

def _find_generic_script(script_dir: Path) -> Path:
    def _key(p: Path) -> Tuple[int, str]:
        m = re.search(r"(\d{4})(?=\.py$)", p.name)
        return (int(m.group(1)) if m else -1, p.name)

    candidates = sorted(script_dir.glob("選會員_共用核心_*.py"), key=_key, reverse=True)
    if not candidates:
        raise RuntimeError("找不到共用核心 選會員_共用核心_*.py")
    return candidates[0]


GENERIC_SCRIPT = _find_generic_script(SCRIPT_DIR)


def _load_generic_module():
    spec = importlib.util.spec_from_file_location("member_merge_core_ys", GENERIC_SCRIPT)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"無法載入共用核心：{GENERIC_SCRIPT}")
    mod = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = mod
    spec.loader.exec_module(mod)
    return mod


# ─── 工具函數 ────────────────────────────────────────────────────────────────

def _normalize_name(v: Any) -> str:
    return re.sub(r"\s+", "", str(v or "").strip())


def _parse_date_iso(v: Any) -> str:
    """回傳 ISO 格式日期字串 (YYYY-MM-DD)，解析失敗回傳空字串。"""
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    s = str(v).strip()
    # 民國 7 碼：1141206
    m = re.match(r"^(\d{3})(\d{2})(\d{2})$", s)
    if m:
        try:
            return datetime.date(int(m.group(1)) + 1911, int(m.group(2)), int(m.group(3))).isoformat()
        except ValueError:
            return ""
    # ISO / 短橫 / 斜線 / 點：2025-12-06、114-12-06、075.06.02
    m = re.match(r"^(\d{2,4})[./\-](\d{1,2})[./\-](\d{1,2})$", s)
    if m:
        yr = int(m.group(1))
        if yr < 1911:
            yr += 1911
        try:
            return datetime.date(yr, int(m.group(2)), int(m.group(3))).isoformat()
        except ValueError:
            return ""
    return ""


def _read_csv_rows(path: Path) -> List[List[str]]:
    data = path.read_bytes()
    for enc in ("utf-8-sig", "big5", "cp950", "utf-8"):
        try:
            text = data.decode(enc)
            return list(csv.reader(text.splitlines()))
        except UnicodeDecodeError:
            continue
    return list(csv.reader(data.decode("cp950", errors="replace").splitlines()))


def _find_col(header: List, candidates: Tuple[str, ...]) -> Optional[int]:
    normalized = [str(v or "").strip() for v in header]
    for c in candidates:
        if c in normalized:
            return normalized.index(c)
    return None


def _find_header_row(
    rows: List[List],
    required: Tuple[Tuple[str, ...], ...],
    max_rows: int = 20,
) -> Optional[int]:
    """找到同時包含所有 required 群組至少一個欄位的第一列。"""
    for i, row in enumerate(rows[:max_rows]):
        cells = {str(v or "").strip() for v in row}
        if all(any(c in cells for c in group) for group in required):
            return i
    return None


def _find_prevention_csv(source_dir: Path) -> Optional[Path]:
    """回傳耀聖預防保健 CSV，同時支援新舊檔名。"""
    for name in (
        "預防保健名單.CSV",
        "預防保健名單.csv",
        "預防保健.CSV",
        "預防保健.csv",
    ):
        path = source_dir / name
        if path.is_file():
            return path
    return None


def _is_zongbei_source(source_dir: Path) -> bool:
    name = source_dir.name
    return "3501102622" in name or "宗北" in name


def _is_ole_encrypted_excel(path: Path) -> bool:
    try:
        data = path.read_bytes()[:4096]
    except OSError:
        return False
    if not data.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        return False
    return b"encryption" in data.lower()


def _find_unreadable_excel_like_files(source_dir: Path) -> List[Path]:
    files: List[Path] = []
    for path in sorted(source_dir.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        if _is_ole_encrypted_excel(path):
            files.append(path)
    return files


def _find_unreadable_prevention_files(source_dir: Path) -> List[Path]:
    return [
        path
        for path in _find_unreadable_excel_like_files(source_dir)
        if "預防保健" in path.name
    ]


def _write_empty_screening_workbook(out_dir: Path) -> Path:
    """建立必要篩檢分頁，讓缺可讀預防保健檔的宗北流程可先完成。"""
    output_path = out_dir / "耀聖_預防保健_空白.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name in SCREENING_SHEETS:
        ws = wb.create_sheet(sheet_name)
        ws.append(["指標名稱", "ID", "生日", "姓名", "最後篩檢日期"])
    wb.save(output_path)
    wb.close()
    return output_path


def _write_internal_member_roster(
    out_dir: Path,
    identity_map: Dict[Tuple[str, str], str],
    internal_ids: set[str],
) -> int:
    """宗北次數檔無 ID 時，補一份中間會員名單讓純次數人員不被漏掉。"""
    if not internal_ids:
        return 0
    output_path = out_dir / "耀聖_內部ID會員補充.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "會員名單"
    ws.append(["ID", "姓名", "BRITHDAY", "疾病樣態"])
    count = 0
    for (name, bday), pid in sorted(identity_map.items()):
        if pid not in internal_ids:
            continue
        ws.append([pid, name, bday, ""])
        count += 1
    wb.save(output_path)
    wb.close()
    return count


def _build_roster_birthday_map(source_dir: Path) -> Dict[str, str]:
    """以家醫/較需要照護名單為準，建立 ID -> 生日。"""
    result: Dict[str, str] = {}
    id_cands = ("ID", "身分證號", "身份證號", "身分證號碼", "身份證號碼", "家醫收案會員ID")
    bday_cands = ("BRITHDAY", "BIRTHDAY", "生日", "出生日期")

    for path in sorted(source_dir.glob("*")):
        if not path.is_file() or path.name.startswith("~$"):
            continue
        if "家醫" not in path.name and "照護名單" not in path.name and "會員名單" not in path.name:
            continue

        if path.suffix.lower() == ".csv":
            rows = _read_csv_rows(path)
            h = _find_header_row(rows, (id_cands, bday_cands), max_rows=10)
            if h is None:
                continue
            id_col = _find_col(rows[h], id_cands)
            bday_col = _find_col(rows[h], bday_cands)
            if id_col is None or bday_col is None:
                continue
            for row in rows[h + 1:]:
                pid = _valid_pid(row[id_col] if id_col < len(row) else "")
                bday = _parse_date_iso(row[bday_col] if bday_col < len(row) else "")
                if pid and bday:
                    result[pid] = bday

        elif path.suffix.lower() == ".xlsx":
            try:
                wb = load_workbook(path, read_only=True, data_only=True)
            except Exception:
                continue
            try:
                for ws in wb.worksheets:
                    rows = [list(r) for r in ws.iter_rows(max_row=500, values_only=True)]
                    h = _find_header_row(rows, (id_cands, bday_cands), max_rows=10)
                    if h is None:
                        continue
                    id_col = _find_col(rows[h], id_cands)
                    bday_col = _find_col(rows[h], bday_cands)
                    if id_col is None or bday_col is None:
                        continue
                    for row in rows[h + 1:]:
                        pid = _valid_pid(row[id_col] if id_col < len(row) else "")
                        bday = _parse_date_iso(row[bday_col] if bday_col < len(row) else "")
                        if pid and bday:
                            result[pid] = bday
            finally:
                wb.close()
    return result


def _write_workbook_first_sheet_as_csv(xlsx_path: Path, csv_path: Path) -> int:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    rows_written = 0
    try:
        ws = wb[wb.sheetnames[0]]
        with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if v is None else v for v in row])
                rows_written += 1
        return rows_written
    finally:
        wb.close()


def _decrypt_office_excel(path: Path, password: str, output_path: Path) -> None:
    try:
        import msoffcrypto
    except ImportError as exc:
        raise RuntimeError(
            "宗北加密 Excel 需要 msoffcrypto-tool，請先安裝：python -m pip install msoffcrypto-tool"
        ) from exc

    with path.open("rb") as src, output_path.open("wb") as dst:
        office_file = msoffcrypto.OfficeFile(src)
        office_file.load_key(password=password)
        office_file.decrypt(dst)


def _ask_password_dialog(title: str, prompt: str) -> Optional[str]:
    try:
        import tkinter as tk
        from tkinter import simpledialog
    except Exception:
        return None

    root = tk.Tk()
    root.withdraw()
    try:
        root.attributes("-topmost", True)
        password = simpledialog.askstring(title, prompt, show="*", parent=root)
        return password
    finally:
        root.destroy()


def _get_encrypted_excel_password(
    encrypted_files: List[Path],
    provided_password: Optional[str] = None,
    failed_file: Optional[Path] = None,
) -> str:
    if provided_password:
        return provided_password
    if failed_file is not None:
        prompt = f"上一個密碼無法解開此檔案，請重新輸入密碼：\n\n{failed_file.name}"
    else:
        file_list = "\n".join(path.name for path in encrypted_files[:8])
        extra = "\n..." if len(encrypted_files) > 8 else ""
        prompt = (
            "偵測到宗北加密 Excel，請輸入密碼：\n\n"
            f"{file_list}{extra}"
        )
    password = _ask_password_dialog(ZONGBEI_PASSWORD_HINT, prompt)
    if password:
        return password
    if sys.stdin is not None and sys.stdin.isatty():
        if failed_file is not None:
            password = getpass.getpass(f"請重新輸入 {failed_file.name} 的密碼：")
        else:
            password = getpass.getpass("請輸入宗北加密 Excel 密碼：")
        if password:
            return password
    raise RuntimeError("未輸入宗北加密 Excel 密碼，無法解密預防保健/篩檢檔。")


def _decrypt_zongbei_excel_files(
    source_dir: Path,
    provided_password: Optional[str] = None,
) -> Tuple[List[Path], List[str]]:
    decrypted: List[Path] = []
    messages: List[str] = []
    encrypted_files = _find_unreadable_excel_like_files(source_dir)
    if not encrypted_files:
        return decrypted, messages

    password = _get_encrypted_excel_password(encrypted_files, provided_password)
    for path in encrypted_files:
        decrypted_path = path.with_name(f"{path.stem}_解密.xlsx")
        for attempt in range(2):
            try:
                _decrypt_office_excel(path, password, decrypted_path)
                break
            except Exception as exc:
                decrypted_path.unlink(missing_ok=True)
                if attempt == 0 and provided_password is None:
                    messages.append(f"{path.name} 使用前一個密碼解密失敗，改請使用者重新輸入。")
                    password = _get_encrypted_excel_password(
                        encrypted_files,
                        provided_password=None,
                        failed_file=path,
                    )
                    continue
                messages.append(f"{path.name} 解密失敗：{exc}")
                break
        if not decrypted_path.exists():
            continue

        decrypted.append(decrypted_path)
        messages.append(f"{path.name} → {decrypted_path.name}")

        if "預防保健" in path.name:
            csv_path = source_dir / "預防保健.CSV"
            rows = _write_workbook_first_sheet_as_csv(decrypted_path, csv_path)
            messages.append(f"{decrypted_path.name} → {csv_path.name}（{rows} 列）")
            decrypted_path.unlink(missing_ok=True)

        path.rename(path.with_name(f"{path.name}.encrypted"))
    return decrypted, messages


_ID_RE = re.compile(r"^[A-Z]{1,2}\d{8,9}$")


def _valid_pid(value: Any) -> str:
    pid = re.sub(r"\s+", "", str(value or "").strip().upper())
    return pid if _ID_RE.fullmatch(pid) else ""


def _internal_id(seed: str) -> str:
    number = int(hashlib.sha1(seed.encode("utf-8")).hexdigest()[:12], 16) % 1_000_000_000
    return f"Z{number:09d}"


# ─── 身分證對照表 ─────────────────────────────────────────────────────────────

def _infer_name_col(rows: List[List], h: int, id_col: int, bday_col: int) -> Optional[int]:
    """
    當 header 中找不到姓名欄（耀聖家醫名單姓名欄 header 為 None）時，
    掃描前幾筆資料列，找出內容為中文姓名（非 ID、非日期）的欄位。
    """
    _name_like = re.compile(r"^[一-鿿㐀-䶿]{2,6}$")
    candidates: Dict[int, int] = {}
    for row in rows[h + 1: h + 20]:
        if not row:
            continue
        for c, v in enumerate(row):
            if c in (id_col, bday_col):
                continue
            s = str(v or "").strip()
            if _name_like.match(s):
                candidates[c] = candidates.get(c, 0) + 1
    if not candidates:
        return None
    return max(candidates, key=lambda c: candidates[c])


def build_identity_map(source_dir: Path) -> Dict[Tuple[str, str], str]:
    """掃描 source_dir 根目錄所有 XLSX，建立 (姓名, 生日ISO) → 身分證號 對照表。"""
    mapping: Dict[Tuple[str, str], str] = {}
    id_cands   = ("ID", "身分證號", "身份證號", "身份證號碼", "身分證號碼", "家醫收案會員ID")
    name_cands = ("姓名", "會員姓名", "個案姓名")
    bday_cands = ("生日", "出生日期", "BIRTHDAY")

    for path in sorted(source_dir.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            for ws in wb.worksheets:
                rows = [list(r) for r in ws.iter_rows(max_row=500, values_only=True)]
                # 優先找含 ID + 姓名 的 header；耀聖家醫名單姓名欄 header 為 None，改用只找 ID
                h = _find_header_row(rows, (id_cands, name_cands), max_rows=20)
                if h is None:
                    h = _find_header_row(rows, (id_cands, bday_cands), max_rows=20)
                if h is None:
                    continue
                id_col   = _find_col(rows[h], id_cands)
                name_col = _find_col(rows[h], name_cands)
                bday_col = _find_col(rows[h], bday_cands)
                if id_col is None or bday_col is None:
                    continue
                if name_col is None:
                    name_col = _infer_name_col(rows, h, id_col, bday_col)
                if name_col is None:
                    continue
                max_col = max(id_col, name_col, bday_col)
                for row in rows[h + 1:]:
                    if not row or len(row) <= max_col:
                        continue
                    pid  = str(row[id_col] or "").strip().upper()
                    name = _normalize_name(row[name_col])
                    bday = _parse_date_iso(row[bday_col])
                    if pid and name and bday and _ID_RE.match(pid):
                        mapping[(name, bday)] = pid
            wb.close()
        except Exception:
            continue
    return mapping


# ─── 預防保健 CSV 轉換 ───────────────────────────────────────────────────────

def _merge_prevention_identity_map(
    source_dir: Path,
    identity_map: Dict[Tuple[str, str], str],
) -> int:
    path = _find_prevention_csv(source_dir)
    if path is None:
        return 0

    rows = _read_csv_rows(path)
    h = _find_header_row(
        rows,
        (
            ("結束日", "就診日"),
            ("姓名",),
            ("生日",),
            ("身份證號", "身分證號"),
            ("預防保健", "卡序"),
            ("主次代碼", "ICD"),
        ),
    )
    if h is None:
        raise ValueError("預防保健名單找不到結束日/姓名/生日/身分證/預防保健/主次代碼必要欄位")
    header = rows[h]
    name_col = _find_col(header, ("姓名",))
    bday_col = _find_col(header, ("生日",))
    id_col = _find_col(header, ("身份證號", "身分證號"))
    added = 0
    for row in rows[h + 1:]:
        name = _normalize_name(row[name_col] if name_col is not None and name_col < len(row) else "")
        bday = _parse_date_iso(row[bday_col] if bday_col is not None and bday_col < len(row) else "")
        pid = _valid_pid(row[id_col] if id_col is not None and id_col < len(row) else "")
        if not name or not bday or not pid:
            continue
        key = (name, bday)
        existing = identity_map.get(key)
        if existing and existing != pid:
            raise ValueError(f"預防保健姓名生日對應多個 ID：{name} / {bday}")
        if not existing:
            identity_map[key] = pid
            added += 1
    return added


def _extend_identity_map_with_internal_ids(
    source_dir: Path,
    identity_map: Dict[Tuple[str, str], str],
) -> set[str]:
    keys: set[Tuple[str, str]] = set()
    csv_paths = (
        sorted(source_dir.glob("最後就診日*.CSV"))
        + sorted(source_dir.glob("最後就診日*.csv"))
    )
    count_dir = source_dir / "次數"
    if count_dir.is_dir():
        csv_paths.extend(sorted(count_dir.glob("*.CSV")))
        csv_paths.extend(sorted(count_dir.glob("*.csv")))

    for path in csv_paths:
        rows = _read_csv_rows(path)
        h = _find_header_row(rows, (("姓名", "姓    名", "病患姓名"), ("生日", "生   日")))
        if h is None:
            continue
        name_col = _find_col(rows[h], ("姓名", "病患姓名", "姓    名"))
        bday_col = _find_col(rows[h], ("生日", "生   日"))
        for row in rows[h + 1:]:
            name = _normalize_name(row[name_col] if name_col is not None and name_col < len(row) else "")
            bday = _parse_date_iso(row[bday_col] if bday_col is not None and bday_col < len(row) else "")
            if name and bday:
                keys.add((name, bday))

    internal_ids: set[str] = set()
    used_ids = set(identity_map.values())
    seed_base = source_dir.name
    for name, bday in sorted(keys):
        key = (name, bday)
        if key in identity_map:
            continue
        salt = 0
        while True:
            pid = _internal_id(f"{seed_base}|{name}|{bday}|{salt}")
            if pid not in used_ids:
                break
            salt += 1
        identity_map[key] = pid
        used_ids.add(pid)
        internal_ids.add(pid)
    return internal_ids


def _convert_prevention_csv(source_dir: Path, out_dir: Path) -> Dict[str, int]:
    path = _find_prevention_csv(source_dir)
    if path is None:
        raise FileNotFoundError("找不到預防保健名單.CSV 或 預防保健.CSV")

    rows = _read_csv_rows(path)
    h = _find_header_row(
        rows,
        (
            ("結束日", "就診日"),
            ("姓名",),
            ("生日",),
            ("身份證號", "身分證號"),
            ("預防保健", "卡序"),
            ("主次代碼", "ICD"),
        ),
    )
    if h is None:
        raise ValueError("預防保健名單找不到結束日/姓名/生日/身分證/預防保健/主次代碼必要欄位")
    header = rows[h]
    end_date_col = _find_col(header, ("結束日", "就診日"))
    name_col = _find_col(header, ("姓名",))
    bday_col = _find_col(header, ("生日",))
    id_col = _find_col(header, ("身份證號", "身分證號"))
    prevention_col = _find_col(header, ("預防保健", "卡序"))
    primary_secondary_col = _find_col(header, ("主次代碼", "ICD"))
    lab_code_col = _find_col(header, ("檢驗代碼",))

    screening: Dict[str, Dict[str, Tuple[str, str, str]]] = {
        sheet_name: {} for sheet_name in SCREENING_SHEETS
    }
    dx_codes: Dict[str, Dict[str, Any]] = {}
    selected_rows = 0
    skipped_codes = 0

    for row in rows[h + 1:]:
        pid = _valid_pid(row[id_col] if id_col is not None and id_col < len(row) else "")
        if not pid:
            continue
        name = str(row[name_col] if name_col is not None and name_col < len(row) else "").strip()
        bday = _parse_date_iso(row[bday_col] if bday_col is not None and bday_col < len(row) else "")
        screening_date = _parse_date_iso(
            row[end_date_col] if end_date_col is not None and end_date_col < len(row) else ""
        )
        prevention_code = re.sub(
            r"\s+",
            "",
            str(
                row[prevention_col]
                if prevention_col is not None and prevention_col < len(row)
                else ""
            ),
        ).upper()
        primary_secondary_code = str(
            row[primary_secondary_col]
            if primary_secondary_col is not None and primary_secondary_col < len(row)
            else ""
        ).strip()
        lab_code = re.sub(
            r"\s+",
            "",
            str(row[lab_code_col] if lab_code_col is not None and lab_code_col < len(row) else ""),
        ).upper()

        if primary_secondary_code:
            rec = dx_codes.setdefault(pid, {"name": name, "birth": bday, "codes": []})
            if primary_secondary_code not in rec["codes"]:
                rec["codes"].append(primary_secondary_code)

        kinds = list(SCREENING_CODE_MAP.get(prevention_code, ()))
        if lab_code in HEPATITIS_LAB_CODES and "肝炎篩檢" not in kinds:
            kinds.append("肝炎篩檢")
        if not kinds:
            skipped_codes += 1
            continue
        if not screening_date:
            continue
        selected_rows += 1
        for kind in kinds:
            existing = screening[kind].get(pid)
            if existing is None or screening_date > existing[0]:
                screening[kind][pid] = (screening_date, name, bday)

    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name in SCREENING_SHEETS:
        ws = wb.create_sheet(sheet_name)
        ws.append(["指標名稱", "ID", "生日", "姓名", "最後篩檢日期"])
        for pid, (last_date, name, bday) in sorted(screening[sheet_name].items()):
            ws.append([sheet_name, pid, bday, name, last_date])
    wb.save(out_dir / "耀聖_預防保健_補正.xlsx")
    wb.close()

    wb_dx = Workbook()
    ws_dx = wb_dx.active
    ws_dx.title = "主次診斷"
    ws_dx.append(["ID", "姓名", "生日", "診斷代碼"])
    for pid, rec in sorted(dx_codes.items()):
        ws_dx.append([pid, rec["name"], rec["birth"], ",".join(rec["codes"])])
    wb_dx.save(out_dir / "耀聖_主次診斷_補正.xlsx")
    wb_dx.close()

    wb_health = Workbook()
    ws_health = wb_health.active
    ws_health.title = "HealthCase"
    ws_health.append([
        "家醫收案會員ID", "姓名", "生日",
        "最近一次HbA1c檢查結果(%)", "最近一次HbA1c檢查日期",
        "最近一次LDL檢查結果(mg/dL)", "最近一次LDL檢查日期",
        "最近一次UACR檢查結果(mg/gm)", "最近一次UACR檢查日期",
    ])
    wb_health.save(out_dir / "耀聖_HealthCase_空白.xlsx")
    wb_health.close()

    path.unlink(missing_ok=True)
    return {
        "source_rows": max(len(rows) - h - 1, 0),
        "selected_rows": selected_rows,
        "selected_people": len({pid for values in screening.values() for pid in values}),
        "dx_people": len(dx_codes),
        "skipped_codes": skipped_codes,
        **{sheet: len(values) for sheet, values in screening.items()},
    }


# ─── 最後就診日.CSV 轉換 ──────────────────────────────────────────────────────

def _convert_last_visit_csv(
    source_dir: Path,
    identity_map: Dict[Tuple[str, str], str],
    out_dir: Path,
) -> Tuple[int, Dict[str, str]]:
    """
    讀取 source_dir 的 最後就診日.CSV，用 identity_map 補齊 ID，
    輸出 耀聖_最後就診日_補正.xlsx（sheet「門診次數費用」含日期欄，count=0）
    供共用核心填最後就診日。原始 CSV 會被移除，避免共用核心重複解析。
    回傳 (成功匹配筆數, pid→date_iso 字典)；後者用於 post-processing 補填年份 114/115 外的日期。
    """
    candidates = sorted(
        list(source_dir.glob("最後就診日*.CSV")) + list(source_dir.glob("最後就診日*.csv"))
    )
    if not candidates:
        return 0, {}

    latest: Dict[str, Tuple[str, str]] = {}  # pid → (name, date_iso)
    unmatched = 0

    for csv_path in candidates:
        raw = _read_csv_rows(csv_path)
        if len(raw) < 2:
            continue
        h = _find_header_row(raw, (("姓名", "病患姓名"), ("生日", "生   日"), ("最後回診日", "最後就診日", "最後看診日")))
        if h is None:
            print(f"  跳過 {csv_path.name}：缺姓名/生日/最後回診日欄", flush=True)
            continue
        name_col = _find_col(raw[h], ("姓名", "病患姓名"))
        bday_col = _find_col(raw[h], ("生日", "生   日"))
        date_col = _find_col(raw[h], ("最後回診日", "最後就診日", "最後看診日"))
        for row in raw[h + 1:]:
            name = _normalize_name(row[name_col] if name_col is not None and name_col < len(row) else "")
            bday = _parse_date_iso(row[bday_col] if bday_col is not None and bday_col < len(row) else "")
            date = _parse_date_iso(row[date_col] if date_col is not None and date_col < len(row) else "")
            if not name or not bday or not date:
                continue
            pid = identity_map.get((name, bday), "")
            if not pid:
                unmatched += 1
                continue
            existing = latest.get(pid)
            if existing is None or date > existing[1]:
                latest[pid] = (name, date)

    # 刪除原始 CSV，避免共用核心誤判
    for csv_path in candidates:
        csv_path.unlink(missing_ok=True)

    if not latest:
        if unmatched:
            print(f"  最後就診日：0 筆命中（{unmatched} 筆找不到 ID）", flush=True)
        return 0, {}

    wb = Workbook()
    ws = wb.active
    ws.title = "門診次數費用"
    ws.append(["身分證號", "姓名", "日期", "次數", "總額"])
    for pid, (name, date) in sorted(latest.items()):
        ws.append([pid, name, date, 0, 0])
    wb.save(out_dir / "耀聖_最後就診日_補正.xlsx")
    wb.close()
    print(f"  最後就診日：{len(latest)} 筆已補 ID，{unmatched} 筆找不到 ID", flush=True)
    last_visit_dict = {pid: date_iso for pid, (_, date_iso) in latest.items()}
    return len(latest), last_visit_dict


# ─── 次數 CSV 轉換 ────────────────────────────────────────────────────────────

def _convert_count_csvs(
    source_dir: Path,
    identity_map: Dict[Tuple[str, str], str],
    out_dir: Path,
) -> int:
    """
    讀取 source_dir/次數/ 內的 YYYYMM.CSV，補齊 ID 後輸出月份補正 xlsx。
    sheet 名稱 = 月份代碼（讓共用核心識別年月），不含日期欄（避免汙染最後就診日）。
    """
    count_dir = source_dir / "次數"
    if not count_dir.is_dir():
        return 0

    total = 0
    # 優先選 UTF-8 版（耀聖 HIS 產出帶 BOM 的 UTF-8 檔，命名含 _Big5）
    seen_months: set = set()
    all_csvs = sorted(
        list(count_dir.glob("*.CSV")) + list(count_dir.glob("*.csv"))
    )
    # 先加入 YYYYMM_Big5 版（UTF-8-BOM），再加入純 YYYYMM 版作後備
    ordered: list = []
    for csv_path in all_csvs:
        m = re.fullmatch(r"(1(?:14|15)\d{2})_Big5", csv_path.stem, re.IGNORECASE)
        if m:
            ordered.append((m.group(1), csv_path))
    for csv_path in all_csvs:
        if re.fullmatch(r"1(?:14|15)\d{2}", csv_path.stem):
            ordered.append((csv_path.stem, csv_path))

    for month_code, csv_path in ordered:
        if month_code in seen_months:
            continue
        seen_months.add(month_code)
        raw = _read_csv_rows(csv_path)
        if len(raw) < 2:
            continue
        h = _find_header_row(raw, (("姓名", "姓    名", "病患姓名"), ("生日", "生   日"), ("次數",)))
        if h is None:
            continue
        name_col  = _find_col(raw[h], ("姓名", "姓    名", "病患姓名"))
        bday_col  = _find_col(raw[h], ("生日", "生   日"))
        count_col = _find_col(raw[h], ("次數",))
        if None in (name_col, bday_col, count_col):
            continue

        rows_out: List[List] = []
        for row in raw[h + 1:]:
            name = _normalize_name(row[name_col] if name_col is not None and name_col < len(row) else "")
            bday = _parse_date_iso(row[bday_col] if bday_col is not None and bday_col < len(row) else "")
            cnt_raw = row[count_col] if count_col is not None and count_col < len(row) else 0
            try:
                cnt = float(str(cnt_raw).replace(",", "") or 0)
            except (ValueError, TypeError):
                cnt = 0.0
            if not name or not bday:
                continue
            pid = identity_map.get((name, bday), "")
            if not pid:
                continue
            rows_out.append([pid, name, cnt, 0])

        if rows_out:
            wb = Workbook()
            ws = wb.active
            ws.title = month_code
            ws.append(["身分證號", "姓名", "次數", "總額"])
            for r in rows_out:
                ws.append(r)
            wb.save(out_dir / f"耀聖_次數_{month_code}_補正.xlsx")
            wb.close()
            total += len(rows_out)
            print(f"  次數 {month_code}：{len(rows_out)} 筆已補 ID", flush=True)

    return total


# ─── 後處理：補填共用核心因年份過濾漏掉的最後就診日 ──────────────────────────────

_ID_CANDS_OUT = ("身分證號", "身份證號", "身分證號碼", "身份證號碼", "ID", "家醫收案會員ID")


def _post_fill_last_visit(output_path: Path, last_visit_dict: Dict[str, str]) -> None:
    """
    共用核心只處理 ROC 114-115 的日期，ROC 113（2024）就診日會被略過。
    輸出後再掃一次總表，對 K 欄（最後就診日）仍為空的會員補填日期。
    """
    if not last_visit_dict or not output_path.exists():
        return

    wb = load_workbook(output_path)
    changed = 0
    try:
        for ws in wb.worksheets:
            id_col: Optional[int] = None
            lv_col: Optional[int] = None
            header_row: Optional[int] = None

            max_scan = min(10, ws.max_row or 0)
            for row_idx in range(1, max_scan + 1):
                cells = [
                    str(ws.cell(row_idx, c).value or "").strip()
                    for c in range(1, (ws.max_column or 0) + 1)
                ]
                if any(c in cells for c in _ID_CANDS_OUT) and "最後就診日" in cells:
                    header_row = row_idx
                    for col_idx, cell_val in enumerate(cells, start=1):
                        if cell_val in _ID_CANDS_OUT and id_col is None:
                            id_col = col_idx
                        if cell_val == "最後就診日" and lv_col is None:
                            lv_col = col_idx
                    break

            if id_col is None or lv_col is None or header_row is None:
                continue

            for row_idx in range(header_row + 1, (ws.max_row or 0) + 1):
                pid = str(ws.cell(row_idx, id_col).value or "").strip().upper()
                if not pid or not _ID_RE.match(pid):
                    continue
                if pid not in last_visit_dict:
                    continue
                lv_cell = ws.cell(row_idx, lv_col)
                if lv_cell.value is not None:
                    continue
                try:
                    lv_cell.value = datetime.date.fromisoformat(last_visit_dict[pid])
                    changed += 1
                except ValueError:
                    pass

        if changed:
            wb.save(output_path)
            print(f"  後處理補填最後就診日：{changed} 格（包含 ROC 113 / 2024 年份）", flush=True)
    finally:
        wb.close()


def _blank_internal_ids(output_path: Path, internal_ids: set[str]) -> int:
    if not internal_ids:
        return 0
    wb = load_workbook(output_path)
    changed = 0
    try:
        id_headers = {
            "ID", "身份證號", "身分證號", "身份證號碼", "身分證號碼", "家醫收案會員ID"
        }
        for ws in wb.worksheets:
            id_columns: set[int] = set()
            header_limit = min(10, ws.max_row)
            for row in range(1, header_limit + 1):
                for col in range(1, ws.max_column + 1):
                    if str(ws.cell(row, col).value or "").strip() in id_headers:
                        id_columns.add(col)
            for col in id_columns:
                for row in range(1, ws.max_row + 1):
                    pid = str(ws.cell(row, col).value or "").strip().upper()
                    if pid in internal_ids:
                        ws.cell(row, col).value = None
                        changed += 1
        if changed:
            wb.save(output_path)
        return changed
    finally:
        wb.close()


def _post_fill_main_dx(output_path: Path, source_dir: Path) -> int:
    path = _find_prevention_csv(source_dir)
    if path is None:
        return 0
    rows = _read_csv_rows(path)
    h = _find_header_row(rows, (("身份證號", "身分證號"), ("ICD", "主次代碼")))
    if h is None:
        return 0
    id_col = _find_col(rows[h], ("身份證號", "身分證號"))
    dx_col = _find_col(rows[h], ("ICD", "主次代碼"))
    dx_map: Dict[str, List[str]] = {}
    for row in rows[h + 1:]:
        pid = _valid_pid(row[id_col] if id_col is not None and id_col < len(row) else "")
        dx = str(row[dx_col] if dx_col is not None and dx_col < len(row) else "").strip()
        if not pid or not dx:
            continue
        bucket = dx_map.setdefault(pid, [])
        if dx not in bucket:
            bucket.append(dx)

    wb = load_workbook(output_path)
    changed = 0
    try:
        if "會員總表" not in wb.sheetnames:
            return 0
        ws = wb["會員總表"]
        id_col_out = None
        dx_col_out = None
        for col in range(1, ws.max_column + 1):
            header = str(ws.cell(1, col).value or "").replace("\n", "").strip()
            if header in ("身份證號碼", "身分證號碼", "身份證號", "身分證號", "ID"):
                id_col_out = col
            if "病123" in header:
                dx_col_out = col
        if id_col_out is None or dx_col_out is None:
            return 0
        for row in range(3, ws.max_row + 1):
            pid = str(ws.cell(row, id_col_out).value or "").strip().upper()
            codes = dx_map.get(pid)
            if not codes:
                continue
            value = ",".join(codes)
            if ws.cell(row, dx_col_out).value != value:
                ws.cell(row, dx_col_out).value = value
                changed += 1
        if changed:
            wb.save(output_path)
        return changed
    finally:
        wb.close()


def _collect_count_totals_by_person(source_dir: Path) -> Dict[Tuple[str, str], List[float]]:
    count_dir = source_dir / "次數"
    totals: Dict[Tuple[str, str], List[float]] = {}
    if not count_dir.is_dir():
        return totals

    for csv_path in sorted(list(count_dir.glob("*.CSV")) + list(count_dir.glob("*.csv"))):
        m = re.fullmatch(r"(1(?:14|15)\d{2})(?:_Big5)?", csv_path.stem, re.IGNORECASE)
        if not m:
            continue
        month_code = m.group(1)
        year_idx = 0 if month_code.startswith("114") else 1
        rows = _read_csv_rows(csv_path)
        h = _find_header_row(rows, (("姓名", "姓    名", "病患姓名"), ("生日", "生   日"), ("次數",)))
        if h is None:
            continue
        name_col = _find_col(rows[h], ("姓名", "姓    名", "病患姓名"))
        bday_col = _find_col(rows[h], ("生日", "生   日"))
        count_col = _find_col(rows[h], ("次數",))
        if name_col is None or bday_col is None or count_col is None:
            continue
        for row in rows[h + 1:]:
            name = _normalize_name(row[name_col] if name_col < len(row) else "")
            bday = _parse_date_iso(row[bday_col] if bday_col < len(row) else "")
            if not name or not bday:
                continue
            try:
                cnt = float(str(row[count_col] if count_col < len(row) else 0).replace(",", "") or 0)
            except (ValueError, TypeError):
                cnt = 0.0
            bucket = totals.setdefault((name, bday), [0.0, 0.0])
            bucket[year_idx] += cnt
    return totals


def _collect_output_count_totals_by_person(output_path: Path) -> Dict[Tuple[str, str], List[float]]:
    totals: Dict[Tuple[str, str], List[float]] = {}
    wb = load_workbook(output_path, read_only=True, data_only=True)
    try:
        if "會員總表" not in wb.sheetnames:
            return totals
        ws = wb["會員總表"]
        for row in ws.iter_rows(min_row=3, min_col=1, max_col=14, values_only=True):
            if not any(v not in (None, "") for v in row):
                continue
            name = _normalize_name(row[0] if len(row) > 0 else "")
            bday = _parse_date_iso(row[6] if len(row) > 6 else "")
            if not name or not bday:
                continue
            bucket = totals.setdefault((name, bday), [0.0, 0.0])
            for idx, pos in enumerate((11, 13)):
                try:
                    bucket[idx] += float(str(row[pos] if len(row) > pos else 0).replace(",", "") or 0)
                except (ValueError, TypeError):
                    pass
        return totals
    finally:
        wb.close()


def _build_zongbei_missing_count_message(output_path: Path, source_dir: Path) -> str:
    expected = _collect_count_totals_by_person(source_dir)
    actual = _collect_output_count_totals_by_person(output_path)
    missing: List[Tuple[str, str, float, float, str]] = []

    for key, counts in sorted(expected.items()):
        actual_counts = actual.get(key, [0.0, 0.0])
        if actual_counts == counts:
            continue
        diff_114 = max(counts[0] - actual_counts[0], 0.0)
        diff_115 = max(counts[1] - actual_counts[1], 0.0)
        if diff_114 == 0 and diff_115 == 0:
            continue
        name, bday = key
        reason = "次數檔生日與會員總表生日不一致，不併入總表"
        missing.append((name, bday, diff_114, diff_115, reason))

    if not missing:
        return ""

    lines = [
        "宗北未入總表資料：",
        f"  未入總表人數：{len(missing)}",
        f"  114年未入次數合計：{sum(row[2] for row in missing):g}",
        f"  115年未入次數合計：{sum(row[3] for row in missing):g}",
        "  明細：",
    ]
    for name, bday, cnt_114, cnt_115, reason in missing:
        lines.append(f"    {name} / {bday} / 114={cnt_114:g} / 115={cnt_115:g} / {reason}")
    return "\n".join(lines)


# ─── 主流程 ──────────────────────────────────────────────────────────────────

def process_excel(
    source_path: str,
    template_path: Optional[str] = None,
    encrypted_excel_password: Optional[str] = None,
) -> str:
    global LAST_RUN_NOTICE
    LAST_RUN_NOTICE = ""
    generic    = _load_generic_module()
    source_dir = Path(source_path).resolve()
    if not source_dir.is_dir():
        raise ValueError("請選擇資料夾。")

    template = template_path or generic._find_template(str(SCRIPT_DIR))
    if not os.path.exists(template):
        raise ValueError(f"找不到模板檔：{template}")

    temp_root   = Path(tempfile.mkdtemp(prefix="ysheng_clean_"))
    temp_source = temp_root / source_dir.name

    try:
        shutil.copytree(source_dir, temp_source)

        is_zongbei = _is_zongbei_source(source_dir)
        if is_zongbei:
            decrypted_paths, decrypt_messages = _decrypt_zongbei_excel_files(
                temp_source,
                encrypted_excel_password,
            )
            if decrypt_messages:
                print("  宗北加密 Excel 解密：", flush=True)
                for message in decrypt_messages:
                    print(f"    {message}", flush=True)
            birthday_map = _build_roster_birthday_map(temp_source)
            if birthday_map:
                print(f"  宗北家醫/會員名單生日基準：{len(birthday_map)} 筆", flush=True)

        has_last_visit = bool(
            list(temp_source.glob("最後就診日*.CSV")) or
            list(temp_source.glob("最後就診日*.csv"))
        )
        has_count_dir = (temp_source / "次數").is_dir()
        has_prevention = _find_prevention_csv(temp_source) is not None
        unreadable_excel_files = _find_unreadable_excel_like_files(temp_source)
        unreadable_prevention_files = _find_unreadable_prevention_files(temp_source)

        matched = 0
        last_visit_dict: Dict[str, str] = {}
        internal_ids: set[str] = set()
        if has_last_visit or has_count_dir or has_prevention:
            print("建立姓名+生日 → ID 對照表…", flush=True)
            identity_map = build_identity_map(temp_source)
            if has_prevention:
                prevention_ids = _merge_prevention_identity_map(temp_source, identity_map)
                print(f"  預防保健新增 ID 對照：{prevention_ids} 筆", flush=True)
            internal_ids = _extend_identity_map_with_internal_ids(temp_source, identity_map)
            if internal_ids:
                print(f"  無真實 ID，暫以內部 ID 串接：{len(internal_ids)} 人", flush=True)
                if is_zongbei:
                    roster_count = _write_internal_member_roster(temp_source, identity_map, internal_ids)
                    print(f"  宗北純次數會員補充：{roster_count} 人", flush=True)
            print(f"  對照表：{len(identity_map)} 筆", flush=True)
            if has_prevention:
                prevention_stats = _convert_prevention_csv(temp_source, temp_source)
                print(
                    "  預防保健："
                    f"來源 {prevention_stats['source_rows']} 筆；"
                    f"指定代碼 {prevention_stats['selected_rows']} 筆／"
                    f"{prevention_stats['selected_people']} 人；"
                    f"成健 {prevention_stats['成人健檢']}、"
                    f"BC肝 {prevention_stats['肝炎篩檢']}、"
                    f"子抹 {prevention_stats['子宮抹片']}、"
                    f"糞篩 {prevention_stats['糞便潛血']}、"
                    f"老流 {prevention_stats['老人流感']}；"
                    f"主次診斷 {prevention_stats['dx_people']} 人",
                    flush=True,
                )
            if has_last_visit:
                count, last_visit_dict = _convert_last_visit_csv(temp_source, identity_map, temp_source)
                matched += count
            if has_count_dir:
                matched += _convert_count_csvs(temp_source, identity_map, temp_source)
        if is_zongbei and not has_prevention and unreadable_prevention_files:
            _write_empty_screening_workbook(temp_source)
            print(
                "  宗北預防保健檔無法由 Python 讀取，已建立空白篩檢分頁讓流程先完成；"
                "下列預防保健資料未吃到，需用 Excel 另存成標準 xlsx/csv 後重跑："
                + "、".join(path.name for path in unreadable_prevention_files),
                flush=True,
            )
        if unreadable_excel_files:
            print(
                "  偵測到加密或非標準 xlsx，已略過原始加密檔："
                + "、".join(path.name for path in unreadable_excel_files),
                flush=True,
            )
        print(f"前置清洗完成，補正 {matched} 筆", flush=True)

        temp_output = Path(generic.process_excel(str(temp_source), template))
        _post_fill_last_visit(temp_output, last_visit_dict)
        dx_count = _post_fill_main_dx(temp_output, source_dir)
        if dx_count:
            print(f"  主次診斷原始欄位回填：{dx_count} 人", flush=True)
        blanked_ids = _blank_internal_ids(temp_output, internal_ids)
        if blanked_ids:
            print(f"  已清空暫時內部 ID：{blanked_ids} 格", flush=True)
        final_output = source_dir.parent / temp_output.name
        final_output = _safe_move_excel(temp_output, final_output)
        if is_zongbei:
            LAST_RUN_NOTICE = _build_zongbei_missing_count_message(final_output, source_dir)
            if LAST_RUN_NOTICE:
                print(LAST_RUN_NOTICE, flush=True)
        return str(final_output)
    finally:
        shutil.rmtree(temp_root, ignore_errors=True)


# ─── GUI 入口 ─────────────────────────────────────────────────────────────────

def main() -> None:
    generic = _load_generic_module()
    if len(sys.argv) > 1:
        src = sys.argv[1]
        template = sys.argv[2] if len(sys.argv) > 2 else generic._find_template(str(SCRIPT_DIR))
        password = sys.argv[3] if len(sys.argv) > 3 else None
        out = process_excel(src, template, password)
        print(f"已輸出：{out}", flush=True)
        return

    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()

    src = filedialog.askdirectory(title="選擇耀聖來源資料夾")
    if not src:
        return

    template = generic._find_template(str(SCRIPT_DIR))

    try:
        out = process_excel(src, template)
        notice = f"\n\n{LAST_RUN_NOTICE}" if LAST_RUN_NOTICE else ""
        messagebox.showinfo("完成", f"已輸出：\n{out}{notice}")
        generic.open_file_cross_platform(out)
    except Exception as e:
        traceback.print_exc()
        messagebox.showerror("錯誤", str(e))


if __name__ == "__main__":
    main()
