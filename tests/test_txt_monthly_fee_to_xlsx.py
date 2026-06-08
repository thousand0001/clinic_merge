import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import load_workbook

from tools.醫聖月次數費用txt轉檔V2 import build_workbook, parse_txt_bytes


ROOT = Path(__file__).resolve().parents[1]


class TxtMonthlyFeeToXlsxTests(unittest.TestCase):
    def test_parse_existing_zip_fixture(self) -> None:
        fixture = ROOT / "backup" / "tmp_11501_test.zip"
        with zipfile.ZipFile(fixture) as zf:
            data = zf.read("nested/115-01.txt")

        month, records = parse_txt_bytes(data, f"{fixture}!nested/115-01.txt")

        self.assertEqual(month, "11501")
        self.assertEqual(len(records), 1840)
        self.assertEqual(
            records[0],
            {
                "病歷號": "064897",
                "看診日": "115.01.01",
                "姓名": "陳雋文",
                "身份證號": "A126358477",
                "生日": "74.05.18",
                "住址": "台北市內湖區行善路25巷40弄3號3樓",
                "電話": "27927448",
                "天數": 3,
                "申請金額": 466,
                "最後就診日": "1150321",
                "診斷代碼": "J309,L509,L308,T07XXXD,L0390",
                "件數": 1,
            },
        )

    def test_build_workbook_existing_fixture_shape(self) -> None:
        fixture = ROOT / "backup" / "tmp_11501_test.zip"
        with zipfile.ZipFile(fixture) as zf:
            data = zf.read("nested/115-01.txt")
        month, records = parse_txt_bytes(data, f"{fixture}!nested/115-01.txt")

        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "11501.xlsx"
            build_workbook(month, records, output_path)
            wb = load_workbook(output_path, data_only=True)

        self.assertEqual(wb.sheetnames, ["11501", "主次診斷", "行動電話"])

        ws_month = wb["11501"]
        self.assertEqual(
            [cell.value for cell in ws_month[1]],
            ["身份證號", "姓名", "件數", "申請金額"],
        )
        self.assertEqual(ws_month.max_row, 948)
        self.assertEqual(
            [ws_month.cell(2, idx).value for idx in range(1, 5)],
            ["A126358477", "陳雋文", 3, 1510],
        )

        ws_diag = wb["主次診斷"]
        self.assertEqual(
            [cell.value for cell in ws_diag[1]],
            ["姓名", "身份證號", "最後就診日(日期)\n(以最新的日期為主)", "診斷代碼(病1,病23)"],
        )
        self.assertEqual(ws_diag.max_row, 948)

        ws_phone = wb["行動電話"]
        self.assertEqual([cell.value for cell in ws_phone[1]], ["姓名", "ID", "電話", "地址"])
        self.assertEqual(ws_phone.max_row, 948)

    def test_parse_multiline_comma_format(self) -> None:
        sample = (
            "                                費用年月:11503                                  \r\n"
            "病歷號,   姓名   ,  身分證  , 看診日  , 自費 ,                                      主訴                                      , 安養院代 ,\r\n"
            "018393,呂雪惠    ,A221613073,115.03.12,     0,after covid-19 with skin                                                        ,26        ,\r\n"
            "      ,          ,          ,         ,      ,\x191150312/19:28陳冠宇  06\x1aA09A00C02UFC00A00A00C00C00A00C00                       ,          ,\r\n"
            "005943,林瑞燕    ,J201539903,115.03.02,     0,1140909 AC 86, a!C 6, TG 116, TC 183,                                           ,26        ,\r\n"
            "      ,          ,          ,         ,      ,DLL 110, Cr 0.6, egFR 104, GOT 22/20,                                           ,          ,\r\n"
            "      ,          ,          ,         ,      ,\x191150302/08:33簡志龍  04\x1aA09A00C02C1C00A00A00C00C00A00C00                       ,          ,\r\n"
        ).encode("cp950", errors="replace")

        month, records = parse_txt_bytes(sample, "P11503.txt")

        self.assertEqual(month, "11503")
        self.assertEqual(len(records), 2)
        self.assertEqual(
            records[0],
            {
                "病歷號": "018393",
                "看診日": "115.03.12",
                "姓名": "呂雪惠",
                "身份證號": "A221613073",
                "生日": "",
                "住址": "",
                "電話": "",
                "天數": 1,
                "申請金額": 0,
                "最後就診日": "1150312",
                "診斷代碼": "A09A00C02UFC00A00A00C00C00A00C00",
                "件數": 1,
            },
        )
        self.assertEqual(records[1]["最後就診日"], "1150302")
        self.assertEqual(records[1]["診斷代碼"], "A09A00C02C1C00A00A00C00C00A00C00")


if __name__ == "__main__":
    unittest.main()
