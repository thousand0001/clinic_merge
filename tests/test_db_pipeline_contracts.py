import datetime as dt
import json
import tempfile
import unittest
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook

from db_pipeline.config.models import load_clinic_config
from db_pipeline.datasets.models import DatasetBundle, MonthlyClaimRecord, SourceTrace
from db_pipeline.detection.detector import detect_source_system
from db_pipeline.parsers.sm import SmParser
from db_pipeline.validation.validator import validate_bundle


class DbPipelineContractTests(unittest.TestCase):
    def test_load_config_and_detect_by_file_name(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            config_path = root / "clinic.json"
            config_path.write_text(
                json.dumps(
                    {
                        "clinic_code": "3531142947",
                        "clinic_name": "本一診所",
                        "source_system": "sm",
                        "detection": {"file_name_contains": ["cliScores"]},
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            (root / "sm_3531142947_cliScores_001.xlsx").touch()

            config = load_clinic_config(config_path)
            result = detect_source_system(root, [config])

            self.assertEqual(result.source_system, "sm")
            self.assertEqual(result.confidence, "low")

    def test_clinic_code_in_path_has_priority(self):
        with tempfile.TemporaryDirectory(prefix="3531142947_") as tmp:
            root = Path(tmp)
            config_path = root / "clinic.json"
            config_path.write_text(
                json.dumps(
                    {
                        "clinic_code": "3531142947",
                        "clinic_name": "本一診所",
                        "source_system": "sm",
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            (root / "一般名稱.xlsx").touch()

            result = detect_source_system(root, [load_clinic_config(config_path)])

            self.assertEqual(result.source_system, "sm")
            self.assertEqual(result.confidence, "high")

    def test_validate_monthly_claim(self):
        trace = SourceTrace(
            clinic_code="3531142947",
            batch_id="batch-test",
            source_system="sm",
            source_file="11501.xlsx",
            source_sheet="11501",
            source_row=2,
            raw_row_hash="abc",
        )
        bundle = DatasetBundle(
            monthly_claims=[
                MonthlyClaimRecord(
                    trace=trace,
                    person_id="A123456789",
                    roc_year=115,
                    month=1,
                    visit_count=Decimal("2"),
                    amount=Decimal("300"),
                    last_visit_date=dt.date(2026, 1, 2),
                )
            ]
        )

        report = validate_bundle(bundle)

        self.assertTrue(report.is_valid)
        self.assertEqual(report.dataset_counts["monthly_claims"], 1)

    def test_invalid_month_is_reported(self):
        trace = SourceTrace(
            clinic_code="3531142947",
            batch_id="batch-test",
            source_system="sm",
            source_file="11513.xlsx",
            source_sheet="11513",
            source_row=2,
            raw_row_hash="abc",
        )
        bundle = DatasetBundle(
            monthly_claims=[
                MonthlyClaimRecord(
                    trace=trace,
                    person_id="A123456789",
                    roc_year=115,
                    month=13,
                )
            ]
        )

        report = validate_bundle(bundle)

        self.assertFalse(report.is_valid)
        self.assertEqual(report.issues[0].code, "invalid_month")

    def test_sm_parser_infers_unlabelled_member_columns(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            roster_path = root / "3531142947照護名單.xlsx"
            roster = Workbook()
            roster_ws = roster.active
            roster_ws.title = "照護名單"
            for _ in range(3):
                roster_ws.append([])
            roster_ws.append(
                [
                    "院所ID",
                    "ID",
                    "BIRTHDAY",
                    None,
                    None,
                    None,
                    None,
                    None,
                    "個案類別",
                ]
            )
            roster_ws.append(
                [
                    "3531142947",
                    "A123456789",
                    "115/01/02",
                    "A123456789",
                    "19450102",
                    "王小明",
                    "台北市中正區忠孝東路1號",
                    "0912345678",
                    "一般",
                ]
            )
            roster.save(roster_path)

            claims_path = root / "R11440門診次數費用.xlsx"
            claims = Workbook()
            claims_ws = claims.active
            claims_ws.title = "11501"
            claims_ws.append(
                ["病歷號", "姓名", "生日", "來診次數", "最後回診日", "申報總金額"]
            )
            claims_ws.append([1, "王小明", "115/01/02", 2, "115/01/20", 300])
            claims.save(claims_path)

            selected_path = root / "115年自選會員.xlsx"
            selected = Workbook()
            selected_ws = selected.active
            selected_ws.append(["姓名", "電話", "地址"])
            selected_ws.append(
                ["王小明", "0912345678", "台北市中正區忠孝東路1號"]
            )
            selected.save(selected_path)

            config_path = root / "clinic.json"
            config_path.write_text(
                json.dumps(
                    {
                        "clinic_code": "3531142947",
                        "clinic_name": "本一診所",
                        "source_system": "sm",
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            result = SmParser().parse(
                root,
                load_clinic_config(config_path),
                "batch-test",
            )

            self.assertEqual(result.bundle.members[0].name, "王小明")
            self.assertEqual(result.bundle.members[0].phone, "0912345678")
            self.assertEqual(len(result.bundle.monthly_claims), 1)
            self.assertEqual(
                [
                    item.selection_type
                    for item in result.bundle.member_selections
                ],
                ["designated_114", "self_selected_115"],
            )
            self.assertFalse(
                any(issue.severity == "error" for issue in result.issues)
            )


if __name__ == "__main__":
    unittest.main()
