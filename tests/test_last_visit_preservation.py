import datetime
import importlib.util
import sys
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
CORE_PATH = ROOT / "選會員_共用核心_0610.py"


def _load_core():
    spec = importlib.util.spec_from_file_location("member_merge_core_0610", CORE_PATH)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def test_monthly_counts_without_date_preserve_existing_last_visit():
    core = _load_core()
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "ID"
    ws["B1"] = "最後就診日"
    ws["C1"] = "114年就診次數"
    ws["D1"] = "115年就診次數"
    ws["E1"] = "114年實際申報總額"
    ws["F1"] = "115年實際申報總額"
    ws["A2"] = "A123456789"
    ws["B2"] = datetime.date(2024, 12, 6)

    cols = {
        "id": 1,
        "last_visit": 2,
        "m_count_114": 3,
        "n_count_115": 4,
        "r_amount_114": 5,
        "s_amount_115": 6,
    }
    claim_sums = {
        "A123456789": {
            **core._empty_claim_bucket(),
            "114_cnt_full": 3.0,
        }
    }

    core.fill_monthly_claim_summary_columns(
        ws,
        data_start=2,
        last_row=2,
        cols=cols,
        claim_sums=claim_sums,
    )

    assert ws["B2"].value == datetime.date(2024, 12, 6)
    assert ws["C2"].value == 3


def test_missing_monthly_record_preserves_existing_last_visit():
    core = _load_core()
    wb = Workbook()
    ws = wb.active
    ws["A2"] = "A123456789"
    ws["B2"] = datetime.date(2024, 12, 6)
    cols = {
        "id": 1,
        "last_visit": 2,
        "m_count_114": 3,
        "n_count_115": 4,
        "r_amount_114": 5,
        "s_amount_115": 6,
    }

    core.fill_monthly_claim_summary_columns(
        ws,
        data_start=2,
        last_row=2,
        cols=cols,
        claim_sums={},
    )

    assert ws["B2"].value == datetime.date(2024, 12, 6)


def test_ascvd_two_row_header_sheet_is_detected_as_ascvd():
    core = _load_core()
    source = Workbook().active
    source["A1"] = "ID"
    source["C1"] = "最後就診日"
    source["B2"] = "ASCVD"
    assert core._looks_like_member_roster_with_ascvd_sheet(source)


def test_ascvd_two_row_header_fills_last_visit():
    core = _load_core()
    source_wb = Workbook()
    source = source_wb.active
    source["A1"] = "ID"
    source["C1"] = "最後就診日"
    source["B2"] = "ASCVD"
    source["A3"] = "A123456789"
    source["B3"] = "b"
    source["C3"] = 1131206

    output_wb = Workbook()
    output = output_wb.active
    output["A2"] = "A123456789"
    cols = {"ascvd": 2, "last_visit": 3}
    meta = {2: core.MemberMeta(row=2, pid="A123456789")}

    core._fill_ascvd(
        output,
        source,
        cols,
        {"A123456789": [2]},
        meta,
    )

    assert output["B2"].value == "b"
    assert output["C2"].value == datetime.date(2024, 12, 6)
