import csv
import importlib.util
import sys
from pathlib import Path

from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parents[1]


def _load_module():
    path = PROJECT_DIR / "前置清洗_健佑_0611.py"
    spec = importlib.util.spec_from_file_location("jianyou_prevention_test", path)
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def test_prevention_uses_column_e_and_date_from_column_a(tmp_path):
    module = _load_module()
    source = tmp_path / "預防保健名單.CSV"
    rows = [
        ["結束日", "姓名", "生日", "身份證號", "預防保健", "醫師", "科別", "主次代碼"],
        ["1150302", "王小明", "0600101", "A123456789", "IC85", "", "", "IC29"],
        ["1150303", "李小華", "0610202", "B223456789", "A000", "", "", "IC29"],
        ["1150304", "陳美麗", "0620303", "C223456789", "ICL1001", "", "", "Z000"],
    ]
    with source.open("w", encoding="cp950", newline="") as handle:
        csv.writer(handle).writerows(rows)

    stats = module._convert_prevention_csv(tmp_path, tmp_path)

    assert stats["糞便潛血"] == 1
    assert stats["成人健檢"] == 1
    assert stats["肝炎篩檢"] == 1
    assert stats["skipped_codes"] == 1

    workbook = load_workbook(
        tmp_path / "健佑_預防保健_補正.xlsx",
        read_only=True,
        data_only=True,
    )
    try:
        fit_rows = list(workbook["糞便潛血"].iter_rows(min_row=2, values_only=True))
        adult_rows = list(workbook["成人健檢"].iter_rows(min_row=2, values_only=True))
        hepatitis_rows = list(workbook["肝炎篩檢"].iter_rows(min_row=2, values_only=True))
    finally:
        workbook.close()

    assert fit_rows == [("A123456789", "王小明", "1971-01-01", "2026-03-02")]
    assert adult_rows == [("C223456789", "陳美麗", "1973-03-03", "2026-03-04")]
    assert hepatitis_rows == [("C223456789", "陳美麗", "1973-03-03", "2026-03-04")]
