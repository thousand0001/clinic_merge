# -*- coding: utf-8 -*-
"""
新耀聖前置清洗 + 共用核心包裝（0610）

用途：
- 將含「需照護名單」的雙工作表主檔拆成乾淨的 ascvd / 會員名單
- 避免百分位計算分頁誤判成會員名單
- 自動偵測日期最新的選會員共用核心與模板完成輸出
"""

from __future__ import annotations

import importlib.util
import csv
import os
import re
import shutil
import sys
import tempfile
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
import xlrd


SCRIPT_DIR = Path(__file__).resolve().parent


def _find_generic_script(script_dir: Path) -> Path:
    def date_key(path: Path) -> tuple[int, str]:
        match = re.search(r"(\d{4})(?=\.py$)", path.name)
        return (int(match.group(1)) if match else -1, path.name)

    candidates = sorted(
        script_dir.glob("選會員_共用核心_*.py"),
        key=date_key,
        reverse=True,
    )
    if not candidates:
        raise RuntimeError("找不到共用核心 選會員_共用核心_*.py")
    return candidates[0]

GENERIC_SCRIPT = _find_generic_script(SCRIPT_DIR)


def _load_generic_module():
    spec = importlib.util.spec_from_file_location("member_merge_core", GENERIC_SCRIPT)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"無法載入共用核心：{GENERIC_SCRIPT}")
    mod = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = mod
    spec.loader.exec_module(mod)
    return mod


def _copy_sheet_values(src_path: Path, sheet_name: str, out_path: Path, target_title: str) -> None:
    wb_src = load_workbook(src_path, data_only=True, read_only=True)
    try:
        ws_src = wb_src[sheet_name]
        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = target_title
        for row in ws_src.iter_rows(values_only=True):
            ws_out.append(list(row))
        wb_out.save(out_path)
    finally:
        wb_src.close()


def _sanitize_roster_workbook(source_dir: Path) -> bool:
    candidates = sorted({
        *source_dir.glob("*需照護名單*.xlsx"),
        *source_dir.glob("*需要照護名單*.xlsx"),
    })
    if not candidates:
        return False

    roster_path = candidates[0]
    wb = load_workbook(roster_path, data_only=True, read_only=True)
    try:
        if not wb.sheetnames:
            return False
        sheetnames = list(wb.sheetnames)
        # 馥齡這類多分頁 1/2/3/a/b 的主檔，本身就能被新版通用邏輯吃到，
        # 不應在前置器硬拆成單一工作表。
        if any(name in {"1", "2", "3", "a", "b"} for name in sheetnames):
            return False
        first_sheet = wb.sheetnames[0]
    finally:
        wb.close()

    _copy_sheet_values(roster_path, first_sheet, source_dir / "新耀聖_ascvd_補正.xlsx", "ascvd")
    _copy_sheet_values(roster_path, first_sheet, source_dir / "新耀聖_會員名單_補正.xlsx", "會員名單")
    roster_path.unlink(missing_ok=True)
    return True


def _extract_month_code(text: str) -> Optional[str]:
    match = re.search(r"(?<!\d)(1(?:14|15)\d{2})(?!\d)", str(text))
    return match.group(1) if match else None


def _find_first_index(headers: list[object], candidates: tuple[str, ...]) -> Optional[int]:
    normalized = [str(v or "").strip() for v in headers]
    for cand in candidates:
        if cand in normalized:
            return normalized.index(cand)
    return None


def _write_month_csv(output_dir: Path, month_code: str, rows: list[list[object]]) -> None:
    out_path = output_dir / f"{month_code}.csv"
    with out_path.open("w", encoding="utf-8-sig", newline="") as f:
        csv.writer(f).writerows(rows)


def _convert_xlsx_month_workbook(book_path: Path, output_dir: Path) -> int:
    wb = load_workbook(book_path, data_only=True, read_only=True)
    written = 0
    try:
        for sheet_name in wb.sheetnames:
            month_code = _extract_month_code(sheet_name)
            if not month_code:
                continue
            ws = wb[sheet_name]
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            id_idx = _find_first_index(headers, ("身分證號", "身份證號", "ID", "身份證號碼", "身分證號碼"))
            name_idx = _find_first_index(headers, ("姓名", "病患姓名", "會員姓名"))
            date_idx = _find_first_index(headers, ("最後看診日期", "最後回診日", "日期", "最後就診日", "看診日"))
            count_idx = _find_first_index(headers, ("看診次數", "件數", "就診次數", "次數"))
            amount_idx = _find_first_index(headers, ("申報總金額", "申請金額", "總金額", "總額"))
            if id_idx is None or name_idx is None or count_idx is None:
                continue
            out_rows = [["ID", "姓名", "日期", "次數", "申請金額"]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                pid = str(row[id_idx] or "").strip()
                name = str(row[name_idx] or "").strip()
                if not pid:
                    continue
                dt = row[date_idx] if date_idx is not None else ""
                amt = row[amount_idx] if amount_idx is not None else ""
                cnt = row[count_idx]
                out_rows.append([pid, name, dt, cnt, amt])
            if len(out_rows) > 1:
                _write_month_csv(output_dir, month_code, out_rows)
                written += 1
    finally:
        wb.close()
    return written


def _convert_xls_month_workbook(book_path: Path, output_dir: Path) -> int:
    book = xlrd.open_workbook(str(book_path))
    written = 0
    for sheet_name in book.sheet_names():
        month_code = _extract_month_code(sheet_name) or _extract_month_code(book_path.stem)
        if not month_code:
            continue
        sh = book.sheet_by_name(sheet_name)
        if sh.nrows < 2:
            continue
        headers = sh.row_values(0)
        id_idx = _find_first_index(headers, ("身分證號", "身份證號", "ID", "身份證號碼", "身分證號碼"))
        name_idx = _find_first_index(headers, ("姓名", "病患姓名", "會員姓名"))
        date_idx = _find_first_index(headers, ("最後看診日期", "最後回診日", "日期", "最後就診日", "看診日"))
        count_idx = _find_first_index(headers, ("看診次數", "件數", "就診次數", "次數"))
        amount_idx = _find_first_index(headers, ("申報總金額", "申請金額", "總金額", "總額"))
        if id_idx is None or name_idx is None or count_idx is None:
            continue
        out_rows = [["ID", "姓名", "日期", "次數", "申請金額"]]
        for r in range(1, sh.nrows):
            row = sh.row_values(r)
            pid = str(row[id_idx] or "").strip()
            name = str(row[name_idx] or "").strip()
            if not pid or pid.startswith("="):
                continue
            dt = row[date_idx] if date_idx is not None else ""
            amt = row[amount_idx] if amount_idx is not None else ""
            cnt = row[count_idx]
            out_rows.append([pid, name, dt, cnt, amt])
        if len(out_rows) > 1:
            _write_month_csv(output_dir, month_code, out_rows)
            written += 1
    return written


def _sanitize_month_workbooks(source_dir: Path) -> int:
    written = 0
    targets = list(source_dir.glob("*月報表*.xlsx")) + list(source_dir.glob("*就診次數*.xlsx")) + list(source_dir.glob("*看診次數統計名單*.xls"))
    for path in sorted(targets):
        file_written = 0
        if path.suffix.lower() == ".xlsx":
            file_written = _convert_xlsx_month_workbook(path, source_dir)
        elif path.suffix.lower() == ".xls":
            file_written = _convert_xls_month_workbook(path, source_dir)
        written += file_written
        if file_written > 0:
            path.unlink(missing_ok=True)
    return written


def process_excel(source_path: str, template_path: Optional[str] = None) -> str:
    generic = _load_generic_module()
    source_dir = Path(source_path).resolve()
    if not source_dir.is_dir():
        raise ValueError("請選擇資料夾。")

    template = template_path or generic._find_template(str(SCRIPT_DIR))
    if not os.path.exists(template):
        raise ValueError(f"找不到模板檔：{template}")

    temp_root = Path(tempfile.mkdtemp(prefix="xys_clean_"))
    temp_source = temp_root / source_dir.name

    try:
        shutil.copytree(source_dir, temp_source)
        _sanitize_roster_workbook(temp_source)
        written = _sanitize_month_workbooks(temp_source)
        print(f"已產生清洗後月份檔 {written} 個", flush=True)
        temp_output = Path(generic.process_excel(str(temp_source), template))
        final_output = source_dir.parent / temp_output.name
        if final_output.exists():
            final_output.unlink()
        shutil.move(str(temp_output), str(final_output))
        return str(final_output)
    finally:
        shutil.rmtree(temp_root, ignore_errors=True)


def main() -> None:
    generic = _load_generic_module()
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()

    src = filedialog.askdirectory(title="選擇新耀聖來源資料夾")
    if not src:
        return

    template = generic._find_template(str(SCRIPT_DIR))

    try:
        out = process_excel(src, template)
        messagebox.showinfo("完成", f"已輸出：\n{out}")
        generic.open_file_cross_platform(out)
    except Exception as e:
        messagebox.showerror("錯誤", str(e))


if __name__ == "__main__":
    main()
