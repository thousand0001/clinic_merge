# -*- coding: utf-8 -*-
"""
耀聖系統前置清洗 + 共用核心包裝（0611）

用途：
- 耀聖 HIS 的 最後就診日.CSV 只有病歷號/姓名/生日，無身分證號。
  本程式掃描來源資料夾建立 姓名+生日 → 身分證 對照表，
  將最後就診日轉為含 ID 的標準格式，供共用核心填入 K 欄（最後就診日）。
- 次數/ 資料夾的月份 CSV 同樣用對照表補齊 ID。
- R11440 資料夾等已含 ID 的原始檔案直接交給共用核心，不另行處理。
- 自動偵測日期最新的選會員共用核心與模板完成輸出。
- 不包「新耀聖」（另有 前置清洗_新耀聖_*.py）。
"""

from __future__ import annotations

import csv
import datetime
import importlib.util
import os
import re
import shutil
import sys
import tempfile
import traceback
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent


# ─── 共用核心自動偵測 ─────────────────────────────────────────────────────────

def _find_generic_script(script_dir: Path) -> Path:
    def _key(p: Path) -> Tuple[int, str]:
        m = re.search(r"(\d{4})(?=\.py$)", p.name)
        return (int(m.group(1)) if m else -1, p.name)

    candidates = sorted(script_dir.glob("選會員_共用核心_*.py"), key=_key, reverse=True)
    if not candidates:
        raise RuntimeError("找不到共用核心 選會員_共用核心_*.py")
    return candidates[0]


GENERIC_SCRIPT = _find_generic_script(SCRIPT_DIR)


def _load_generic_module():
    spec = importlib.util.spec_from_file_location("member_merge_core_ys", GENERIC_SCRIPT)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"無法載入共用核心：{GENERIC_SCRIPT}")
    mod = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = mod
    spec.loader.exec_module(mod)
    return mod


# ─── 工具函數 ────────────────────────────────────────────────────────────────

def _normalize_name(v: Any) -> str:
    return re.sub(r"\s+", "", str(v or "").strip())


def _parse_date_iso(v: Any) -> str:
    """回傳 ISO 格式日期字串 (YYYY-MM-DD)，解析失敗回傳空字串。"""
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    s = str(v).strip()
    # 民國 7 碼：1141206
    m = re.match(r"^(\d{3})(\d{2})(\d{2})$", s)
    if m:
        try:
            return datetime.date(int(m.group(1)) + 1911, int(m.group(2)), int(m.group(3))).isoformat()
        except ValueError:
            return ""
    # ISO / 短橫 / 斜線 / 點：2025-12-06、114-12-06、075.06.02
    m = re.match(r"^(\d{2,4})[./\-](\d{1,2})[./\-](\d{1,2})$", s)
    if m:
        yr = int(m.group(1))
        if yr < 1911:
            yr += 1911
        try:
            return datetime.date(yr, int(m.group(2)), int(m.group(3))).isoformat()
        except ValueError:
            return ""
    return ""


def _read_csv_rows(path: Path) -> List[List[str]]:
    data = path.read_bytes()
    for enc in ("utf-8-sig", "big5", "cp950", "utf-8"):
        try:
            text = data.decode(enc)
            return list(csv.reader(text.splitlines()))
        except UnicodeDecodeError:
            continue
    return list(csv.reader(data.decode("cp950", errors="replace").splitlines()))


def _find_col(header: List, candidates: Tuple[str, ...]) -> Optional[int]:
    normalized = [str(v or "").strip() for v in header]
    for c in candidates:
        if c in normalized:
            return normalized.index(c)
    return None


def _find_header_row(
    rows: List[List],
    required: Tuple[Tuple[str, ...], ...],
    max_rows: int = 20,
) -> Optional[int]:
    """找到同時包含所有 required 群組至少一個欄位的第一列。"""
    for i, row in enumerate(rows[:max_rows]):
        cells = {str(v or "").strip() for v in row}
        if all(any(c in cells for c in group) for group in required):
            return i
    return None


_ID_RE = re.compile(r"^[A-Z]{1,2}\d{8,9}$")


# ─── 身分證對照表 ─────────────────────────────────────────────────────────────

def _infer_name_col(rows: List[List], h: int, id_col: int, bday_col: int) -> Optional[int]:
    """
    當 header 中找不到姓名欄（耀聖家醫名單姓名欄 header 為 None）時，
    掃描前幾筆資料列，找出內容為中文姓名（非 ID、非日期）的欄位。
    """
    _name_like = re.compile(r"^[一-鿿㐀-䶿]{2,6}$")
    candidates: Dict[int, int] = {}
    for row in rows[h + 1: h + 20]:
        if not row:
            continue
        for c, v in enumerate(row):
            if c in (id_col, bday_col):
                continue
            s = str(v or "").strip()
            if _name_like.match(s):
                candidates[c] = candidates.get(c, 0) + 1
    if not candidates:
        return None
    return max(candidates, key=lambda c: candidates[c])


def build_identity_map(source_dir: Path) -> Dict[Tuple[str, str], str]:
    """掃描 source_dir 根目錄所有 XLSX，建立 (姓名, 生日ISO) → 身分證號 對照表。"""
    mapping: Dict[Tuple[str, str], str] = {}
    id_cands   = ("ID", "身分證號", "身份證號", "身份證號碼", "身分證號碼", "家醫收案會員ID")
    name_cands = ("姓名", "會員姓名", "個案姓名")
    bday_cands = ("生日", "出生日期", "BIRTHDAY")

    for path in sorted(source_dir.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            for ws in wb.worksheets:
                rows = [list(r) for r in ws.iter_rows(max_row=500, values_only=True)]
                # 優先找含 ID + 姓名 的 header；耀聖家醫名單姓名欄 header 為 None，改用只找 ID
                h = _find_header_row(rows, (id_cands, name_cands), max_rows=20)
                if h is None:
                    h = _find_header_row(rows, (id_cands, bday_cands), max_rows=20)
                if h is None:
                    continue
                id_col   = _find_col(rows[h], id_cands)
                name_col = _find_col(rows[h], name_cands)
                bday_col = _find_col(rows[h], bday_cands)
                if id_col is None or bday_col is None:
                    continue
                if name_col is None:
                    name_col = _infer_name_col(rows, h, id_col, bday_col)
                if name_col is None:
                    continue
                max_col = max(id_col, name_col, bday_col)
                for row in rows[h + 1:]:
                    if not row or len(row) <= max_col:
                        continue
                    pid  = str(row[id_col] or "").strip().upper()
                    name = _normalize_name(row[name_col])
                    bday = _parse_date_iso(row[bday_col])
                    if pid and name and bday and _ID_RE.match(pid):
                        mapping[(name, bday)] = pid
            wb.close()
        except Exception:
            continue
    return mapping


# ─── 最後就診日.CSV 轉換 ──────────────────────────────────────────────────────

def _convert_last_visit_csv(
    source_dir: Path,
    identity_map: Dict[Tuple[str, str], str],
    out_dir: Path,
) -> Tuple[int, Dict[str, str]]:
    """
    讀取 source_dir 的 最後就診日.CSV，用 identity_map 補齊 ID，
    輸出 耀聖_最後就診日_補正.xlsx（sheet「門診次數費用」含日期欄，count=0）
    供共用核心填最後就診日。原始 CSV 會被移除，避免共用核心重複解析。
    回傳 (成功匹配筆數, pid→date_iso 字典)；後者用於 post-processing 補填年份 114/115 外的日期。
    """
    candidates = sorted(
        list(source_dir.glob("最後就診日*.CSV")) + list(source_dir.glob("最後就診日*.csv"))
    )
    if not candidates:
        return 0, {}

    latest: Dict[str, Tuple[str, str]] = {}  # pid → (name, date_iso)
    unmatched = 0

    for csv_path in candidates:
        raw = _read_csv_rows(csv_path)
        if len(raw) < 2:
            continue
        h = _find_header_row(raw, (("姓名", "病患姓名"), ("生日", "生   日"), ("最後回診日", "最後就診日", "最後看診日")))
        if h is None:
            print(f"  跳過 {csv_path.name}：缺姓名/生日/最後回診日欄", flush=True)
            continue
        name_col = _find_col(raw[h], ("姓名", "病患姓名"))
        bday_col = _find_col(raw[h], ("生日", "生   日"))
        date_col = _find_col(raw[h], ("最後回診日", "最後就診日", "最後看診日"))
        for row in raw[h + 1:]:
            name = _normalize_name(row[name_col] if name_col is not None and name_col < len(row) else "")
            bday = _parse_date_iso(row[bday_col] if bday_col is not None and bday_col < len(row) else "")
            date = _parse_date_iso(row[date_col] if date_col is not None and date_col < len(row) else "")
            if not name or not bday or not date:
                continue
            pid = identity_map.get((name, bday), "")
            if not pid:
                unmatched += 1
                continue
            existing = latest.get(pid)
            if existing is None or date > existing[1]:
                latest[pid] = (name, date)

    # 刪除原始 CSV，避免共用核心誤判
    for csv_path in candidates:
        csv_path.unlink(missing_ok=True)

    if not latest:
        if unmatched:
            print(f"  最後就診日：0 筆命中（{unmatched} 筆找不到 ID）", flush=True)
        return 0, {}

    wb = Workbook()
    ws = wb.active
    ws.title = "門診次數費用"
    ws.append(["身分證號", "姓名", "日期", "次數", "總額"])
    for pid, (name, date) in sorted(latest.items()):
        ws.append([pid, name, date, 0, 0])
    wb.save(out_dir / "耀聖_最後就診日_補正.xlsx")
    print(f"  最後就診日：{len(latest)} 筆已補 ID，{unmatched} 筆找不到 ID", flush=True)
    last_visit_dict = {pid: date_iso for pid, (_, date_iso) in latest.items()}
    return len(latest), last_visit_dict


# ─── 次數 CSV 轉換 ────────────────────────────────────────────────────────────

def _convert_count_csvs(
    source_dir: Path,
    identity_map: Dict[Tuple[str, str], str],
    out_dir: Path,
) -> int:
    """
    讀取 source_dir/次數/ 內的 YYYYMM.CSV，補齊 ID 後輸出月份補正 xlsx。
    sheet 名稱 = 月份代碼（讓共用核心識別年月），不含日期欄（避免汙染最後就診日）。
    """
    count_dir = source_dir / "次數"
    if not count_dir.is_dir():
        return 0

    total = 0
    # 優先選 UTF-8 版（耀聖 HIS 產出帶 BOM 的 UTF-8 檔，命名含 _Big5）
    seen_months: set = set()
    all_csvs = sorted(
        list(count_dir.glob("*.CSV")) + list(count_dir.glob("*.csv"))
    )
    # 先加入 YYYYMM_Big5 版（UTF-8-BOM），再加入純 YYYYMM 版作後備
    ordered: list = []
    for csv_path in all_csvs:
        m = re.fullmatch(r"(1(?:14|15)\d{2})_Big5", csv_path.stem, re.IGNORECASE)
        if m:
            ordered.append((m.group(1), csv_path))
    for csv_path in all_csvs:
        if re.fullmatch(r"1(?:14|15)\d{2}", csv_path.stem):
            ordered.append((csv_path.stem, csv_path))

    for month_code, csv_path in ordered:
        if month_code in seen_months:
            continue
        seen_months.add(month_code)
        raw = _read_csv_rows(csv_path)
        if len(raw) < 2:
            continue
        h = _find_header_row(raw, (("姓名", "姓    名", "病患姓名"), ("生日", "生   日"), ("次數",)))
        if h is None:
            continue
        name_col  = _find_col(raw[h], ("姓名", "姓    名", "病患姓名"))
        bday_col  = _find_col(raw[h], ("生日", "生   日"))
        count_col = _find_col(raw[h], ("次數",))
        if None in (name_col, bday_col, count_col):
            continue

        rows_out: List[List] = []
        for row in raw[h + 1:]:
            name = _normalize_name(row[name_col] if name_col is not None and name_col < len(row) else "")
            bday = _parse_date_iso(row[bday_col] if bday_col is not None and bday_col < len(row) else "")
            cnt_raw = row[count_col] if count_col is not None and count_col < len(row) else 0
            try:
                cnt = float(str(cnt_raw).replace(",", "") or 0)
            except (ValueError, TypeError):
                cnt = 0.0
            if not name or not bday:
                continue
            pid = identity_map.get((name, bday), "")
            if not pid:
                continue
            rows_out.append([pid, name, cnt, 0])

        if rows_out:
            wb = Workbook()
            ws = wb.active
            ws.title = month_code
            ws.append(["身分證號", "姓名", "次數", "總額"])
            for r in rows_out:
                ws.append(r)
            wb.save(out_dir / f"耀聖_次數_{month_code}_補正.xlsx")
            total += len(rows_out)
            print(f"  次數 {month_code}：{len(rows_out)} 筆已補 ID", flush=True)

    return total


# ─── 後處理：補填共用核心因年份過濾漏掉的最後就診日 ──────────────────────────────

_ID_CANDS_OUT = ("身分證號", "身份證號", "ID", "家醫收案會員ID")


def _post_fill_last_visit(output_path: Path, last_visit_dict: Dict[str, str]) -> None:
    """
    共用核心只處理 ROC 114-115 的日期，ROC 113（2024）就診日會被略過。
    輸出後再掃一次總表，對 K 欄（最後就診日）仍為空的會員補填日期。
    """
    if not last_visit_dict or not output_path.exists():
        return

    wb = load_workbook(output_path)
    changed = 0

    for ws in wb.worksheets:
        id_col: Optional[int] = None
        lv_col: Optional[int] = None
        header_row: Optional[int] = None

        max_scan = min(10, ws.max_row or 0)
        for row_idx in range(1, max_scan + 1):
            cells = [
                str(ws.cell(row_idx, c).value or "").strip()
                for c in range(1, (ws.max_column or 0) + 1)
            ]
            if any(c in cells for c in _ID_CANDS_OUT) and "最後就診日" in cells:
                header_row = row_idx
                for col_idx, cell_val in enumerate(cells, start=1):
                    if cell_val in _ID_CANDS_OUT and id_col is None:
                        id_col = col_idx
                    if cell_val == "最後就診日" and lv_col is None:
                        lv_col = col_idx
                break

        if id_col is None or lv_col is None or header_row is None:
            continue

        for row_idx in range(header_row + 1, (ws.max_row or 0) + 1):
            pid = str(ws.cell(row_idx, id_col).value or "").strip().upper()
            if not pid or not _ID_RE.match(pid):
                continue
            if pid not in last_visit_dict:
                continue
            lv_cell = ws.cell(row_idx, lv_col)
            if lv_cell.value is not None:
                continue
            try:
                lv_cell.value = datetime.date.fromisoformat(last_visit_dict[pid])
                changed += 1
            except ValueError:
                pass

    if changed:
        wb.save(output_path)
        print(f"  後處理補填最後就診日：{changed} 格（包含 ROC 113 / 2024 年份）", flush=True)
    wb.close()


# ─── 主流程 ──────────────────────────────────────────────────────────────────

def process_excel(source_path: str, template_path: Optional[str] = None) -> str:
    generic    = _load_generic_module()
    source_dir = Path(source_path).resolve()
    if not source_dir.is_dir():
        raise ValueError("請選擇資料夾。")

    template = template_path or generic._find_template(str(SCRIPT_DIR))
    if not os.path.exists(template):
        raise ValueError(f"找不到模板檔：{template}")

    temp_root   = Path(tempfile.mkdtemp(prefix="ysheng_clean_"))
    temp_source = temp_root / source_dir.name

    try:
        shutil.copytree(source_dir, temp_source)

        has_last_visit = bool(
            list(temp_source.glob("最後就診日*.CSV")) or
            list(temp_source.glob("最後就診日*.csv"))
        )
        has_count_dir = (temp_source / "次數").is_dir()

        matched = 0
        last_visit_dict: Dict[str, str] = {}
        if has_last_visit or has_count_dir:
            print("建立姓名+生日 → ID 對照表…", flush=True)
            identity_map = build_identity_map(temp_source)
            print(f"  對照表：{len(identity_map)} 筆", flush=True)
            if has_last_visit:
                count, last_visit_dict = _convert_last_visit_csv(temp_source, identity_map, temp_source)
                matched += count
            if has_count_dir:
                matched += _convert_count_csvs(temp_source, identity_map, temp_source)
        print(f"前置清洗完成，補正 {matched} 筆", flush=True)

        temp_output = Path(generic.process_excel(str(temp_source), template))
        _post_fill_last_visit(temp_output, last_visit_dict)
        final_output = source_dir.parent / temp_output.name
        if final_output.exists():
            final_output.unlink()
        shutil.move(str(temp_output), str(final_output))
        return str(final_output)
    finally:
        shutil.rmtree(temp_root, ignore_errors=True)


# ─── GUI 入口 ─────────────────────────────────────────────────────────────────

def main() -> None:
    generic = _load_generic_module()
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()

    src = filedialog.askdirectory(title="選擇耀聖來源資料夾")
    if not src:
        return

    template = generic._find_template(str(SCRIPT_DIR))

    try:
        out = process_excel(src, template)
        messagebox.showinfo("完成", f"已輸出：\n{out}")
        generic.open_file_cross_platform(out)
    except Exception as e:
        traceback.print_exc()
        messagebox.showerror("錯誤", str(e))


if __name__ == "__main__":
    main()
