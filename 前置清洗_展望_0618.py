# -*- coding: utf-8 -*-
"""
展望系統前置清洗 + 共用核心包裝（0618）

用途：
- 不修改共用核心。
- 直接呼叫日期最新的「選會員_共用核心_*.py」產生正式 Excel。
- 額外從來源資料中可辨識的標準欄位彙整 AW~BI，
  在共用核心產生「醫生看」後補入，並補算 AR / AS / BJ。
- 額外支援展望新診所提供的「預防保健-成健/BC肝/子抹/糞篩/老流/成健＋BC肝」格式，
  將「看診日期」轉成共用核心可讀的「最後篩檢日期」。

注意：
- AT 欄位先不處理，保留共用核心原邏輯。
- BJ 由醫生看 BC=1 且 BG=1 判斷打勾。
"""

from __future__ import annotations

import argparse
import datetime as dt
import importlib.util
import os
import re
import shutil
import sys
import tempfile
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook

SCRIPT_DIR = Path(__file__).resolve().parent


DETAIL_TO_DOCTOR_COL = {
    "case_category": "AW",
    "quality_roster": "AX",
    "multi_chronic_65": "AY",
    "high_visit": "AZ",
    "chronic_mark": "BA",
    "non_chronic_mark": "BB",
    "same_clinic_previous_year": "BC",
    "dmk_raw": "BD",
    "ascvd": "BE",
    "three_highs": "BF",
    "hypertension": "BG",
    "hyperlipidemia": "BH",
    "hyperglycemia": "BI",
}

NEW_PREVENTION_LABELS = {
    "adult": ("成人健檢", "成人預防保健檢查"),
    "hep": ("肝炎篩檢", "B、C肝炎篩檢率"),
    "pap": ("子宮抹片", "子宮頸抹片檢查"),
    "fit": ("糞便潛血", "糞便潛血檢查"),
    "flu": ("老人流感", "65歲以上老人流感注射"),
}

def _find_common_core(script_dir: Path) -> Path:
    def version_key(path: Path) -> tuple[int, str]:
        match = re.search(r"(\d{4})(?=\.py$)", path.name)
        return (int(match.group(1)) if match else -1, path.name)

    candidates = sorted(script_dir.glob("選會員_共用核心_*.py"), key=version_key, reverse=True)
    if not candidates:
        raise RuntimeError("找不到共用核心：選會員_共用核心_*.py")
    return candidates[0]


COMMON_CORE = _find_common_core(SCRIPT_DIR)


def _load_common_core():
    spec = importlib.util.spec_from_file_location("prospect_common_core", COMMON_CORE)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"無法載入共用核心：{COMMON_CORE}")
    mod = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = mod
    spec.loader.exec_module(mod)
    return mod


GENERIC = _load_common_core()


def _normalize_filename_token(value: Any) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[\s\-_()（）\[\]{}]+", "", text)


def _is_generated_output_sheet(sheet_name: str) -> bool:
    token = _normalize_filename_token(sheet_name)
    return "從會員指標內容key過來" in token


def _iter_excel_files(source_dir: Path) -> List[Path]:
    files: List[Path] = []
    for path in source_dir.rglob("*"):
        if not path.is_file() or path.name.startswith("~$"):
            continue
        if path.suffix.lower() in {".xlsx", ".xls", ".xlsm", ".ods"}:
            files.append(path)
    return sorted(files)


def _parse_date(value: Any) -> Optional[dt.date]:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = str(value).strip()
    if text in {"", "-", "—", "–", "None"}:
        return None
    text = text.split()[0]
    match = re.fullmatch(r"(\d{2,4})[./\-](\d{1,2})[./\-](\d{1,2})", text)
    if match:
        year = int(match.group(1))
        if year < 1911:
            year += 1911
        try:
            return dt.date(year, int(match.group(2)), int(match.group(3)))
        except ValueError:
            return None
    digits = re.sub(r"\D", "", text)
    try:
        if len(digits) == 7:
            return dt.date(int(digits[:3]) + 1911, int(digits[3:5]), int(digits[5:7]))
        if len(digits) == 8:
            year = int(digits[:4])
            if year < 1911:
                year += 1911
            return dt.date(year, int(digits[4:6]), int(digits[6:8]))
        if len(digits) == 6:
            return dt.date(int(digits[:2]) + 1911, int(digits[2:4]), int(digits[4:6]))
    except ValueError:
        return None
    return None


def _format_date(value: dt.date) -> str:
    return value.strftime("%Y/%m/%d")


def _compact_header(value: Any) -> str:
    return re.sub(r"[\s　]+", "", str(value or "").strip()).upper()


def _find_col(header: Iterable[Any], aliases: Iterable[str]) -> Optional[int]:
    compact = [_compact_header(value) for value in header]
    for alias in aliases:
        target = _compact_header(alias)
        for index, value in enumerate(compact):
            if value == target:
                return index
    return None


def _classify_new_prevention_file(path: Path) -> Tuple[str, ...]:
    token = _normalize_filename_token(path.stem)
    if "cliscores" in token:
        return ()
    kinds: List[str] = []
    if "成健" in token or "成人預防保健" in token:
        kinds.append("adult")
    if "bc肝" in token or "b肝c肝" in token or "肝" in token:
        kinds.append("hep")
    if "子抹" in token or "抹片" in token:
        kinds.append("pap")
    if "糞篩" in token or "糞便潛血" in token or "潛血" in token:
        kinds.append("fit")
    if "老流" in token or "老人流感" in token or "流感" in token:
        kinds.append("flu")
    return tuple(dict.fromkeys(kinds))


def _collect_new_prevention_rows(source_dir: Path) -> Dict[str, Dict[str, Tuple[dt.date, str, Any]]]:
    collected: Dict[str, Dict[str, Tuple[dt.date, str, Any]]] = {
        key: {} for key in NEW_PREVENTION_LABELS
    }
    source_counts: Dict[str, int] = {}

    for path in _iter_excel_files(source_dir):
        kinds = _classify_new_prevention_file(path)
        if not kinds:
            continue
        try:
            wb = GENERIC.DEFAULT_PROFILE.load_source_workbook(str(path))
        except Exception:
            continue
        try:
            for ws in wb.worksheets:
                rows = ws.iter_rows(values_only=True)
                try:
                    header = list(next(rows))
                except StopIteration:
                    continue
                id_col = _find_col(header, ("身分證", "身分證號", "身分證號碼", "身份證", "身份證號", "身份證號碼", "ID"))
                name_col = _find_col(header, ("姓名", "病患姓名", "會員姓名"))
                bday_col = _find_col(header, ("生日", "出生日期", "BIRTHDAY"))
                date_col = _find_col(header, ("看診日期", "看診日", "就診日", "日期"))
                if id_col is None or date_col is None:
                    continue
                max_col = max(col for col in (id_col, name_col, bday_col, date_col) if col is not None)
                for row in rows:
                    if len(row) <= max_col:
                        continue
                    pid = GENERIC.normalize_id(row[id_col])
                    visit_date = _parse_date(row[date_col])
                    if not pid or visit_date is None:
                        continue
                    name = str(row[name_col] or "").strip() if name_col is not None and name_col < len(row) else ""
                    bday = row[bday_col] if bday_col is not None and bday_col < len(row) else ""
                    for kind in kinds:
                        existing = collected[kind].get(pid)
                        if existing is None or visit_date > existing[0]:
                            collected[kind][pid] = (visit_date, name, bday)
                        source_counts[kind] = source_counts.get(kind, 0) + 1
        finally:
            try:
                wb.close()
            except Exception:
                pass

    if source_counts:
        print(
            "展望新預防保健來源："
            + "；".join(
                f"{NEW_PREVENTION_LABELS[key][0]} {count} 筆"
                for key, count in sorted(source_counts.items())
            ),
            flush=True,
        )
    return collected


def _write_new_prevention_standard_workbook(source_dir: Path) -> Optional[Path]:
    collected = _collect_new_prevention_rows(source_dir)
    if not any(collected.values()):
        return None
    path = source_dir / "展望_新預防保健_標準.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    for kind, (sheet_name, indicator) in NEW_PREVENTION_LABELS.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(["指標名稱", "ID", "生日", "姓名", "最後篩檢日期"])
        for pid, (last_date, name, bday) in sorted(collected[kind].items()):
            ws.append([indicator, pid, bday, name, _format_date(last_date)])
    wb.save(path)
    wb.close()
    print(f"展望新預防保健標準檔：{path.name}", flush=True)
    return path


def _merge_detail(target: Dict[str, Any], detail: Dict[str, Any]) -> None:
    for key in DETAIL_TO_DOCTOR_COL:
        value = detail.get(key)
        if value not in (None, "") and not target.get(key):
            target[key] = value


def collect_member_extra_details(source_dir: Path) -> Dict[str, Dict[str, Any]]:
    """
    從展望來源資料抽出醫生看 AW~BI 所需標準欄位。

    只讀來源檔，不寫檔、不移動檔案。
    """
    id_aliases = ["ID", "身份證號", "身份證號碼", "身分證號", "身分證號碼", "身份証號", "身分証號"]
    details: Dict[str, Dict[str, Any]] = {}
    source_sheets: List[str] = []

    for path in _iter_excel_files(source_dir):
        try:
            wb = GENERIC.DEFAULT_PROFILE.load_source_workbook(str(path))
        except Exception:
            continue
        try:
            for ws in wb.worksheets:
                if _is_generated_output_sheet(ws.title):
                    continue
                partial_map = GENERIC._extract_member_partial_map(ws, id_aliases)
                useful_count = 0
                for pid, detail in partial_map.items():
                    useful = {key: detail.get(key) for key in DETAIL_TO_DOCTOR_COL if detail.get(key) not in (None, "")}
                    if not useful:
                        continue
                    _merge_detail(details.setdefault(pid, {}), detail)
                    useful_count += 1
                if useful_count:
                    source_sheets.append(f"{path.name}/{ws.title}={useful_count}")
        finally:
            try:
                wb.close()
            except Exception:
                pass

    if source_sheets:
        print("展望 AW~BI 標準欄位來源：" + "；".join(source_sheets), flush=True)
    else:
        print("展望 AW~BI 標準欄位來源：未找到可補醫生看的標準欄位", flush=True)
    return details


def _score_value(value: Any) -> Optional[int]:
    if value in (None, ""):
        return None
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def _sum_scores(*values: Any) -> Optional[int]:
    nums = [_score_value(value) for value in values]
    valid = [num for num in nums if num is not None]
    return sum(valid) if valid else None


def _clean_doctor_detail_value(value: Any) -> Optional[Any]:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.replace("\u3000", " ").strip()
        return text or None
    return value


def _is_one(value: Any) -> bool:
    return GENERIC.normalize_text(value) == "1"


def patch_doctor_sheet(wb_tpl, details: Dict[str, Dict[str, Any]]) -> None:
    if GENERIC.DOCTOR_SHEET_NAME not in wb_tpl.sheetnames:
        return

    ws_doc = wb_tpl[GENERIC.DOCTOR_SHEET_NAME]
    extra_filled_rows = 0
    for row in range(4, ws_doc.max_row + 1):
        ws_doc[f"AR{row}"] = _sum_scores(
            ws_doc[f"AM{row}"].value,
            ws_doc[f"AK{row}"].value,
            ws_doc[f"AL{row}"].value,
        )
        ws_doc[f"AS{row}"] = _sum_scores(
            ws_doc[f"AM{row}"].value,
            ws_doc[f"AL{row}"].value,
        )

        pid = GENERIC.normalize_id(ws_doc[f"A{row}"].value)
        detail = details.get(pid) if pid else None
        if detail:
            row_filled = False
            for key, col_letter in DETAIL_TO_DOCTOR_COL.items():
                value = _clean_doctor_detail_value(detail.get(key))
                if value not in (None, ""):
                    ws_doc[f"{col_letter}{row}"] = value
                    row_filled = True
            if row_filled:
                extra_filled_rows += 1

        ws_doc[f"BJ{row}"] = "✔" if _is_one(ws_doc[f"BC{row}"].value) and _is_one(ws_doc[f"BG{row}"].value) else None

    print(f"醫生看 AW~BI：依標準欄位補入 {extra_filled_rows} 列；AR/AS/BJ 已補算", flush=True)


def process_excel(source_path: str, template_path: Optional[str] = None) -> str:
    source_dir = Path(source_path).expanduser().resolve()
    if not source_dir.is_dir():
        raise ValueError("請選擇展望來源資料夾。")

    template = template_path or GENERIC._find_template(str(SCRIPT_DIR))
    if not os.path.exists(template):
        raise ValueError(f"找不到模板檔：{template}")

    with tempfile.TemporaryDirectory(prefix="prospect_clean_") as tmp:
        temp_source = Path(tmp) / source_dir.name
        shutil.copytree(source_dir, temp_source)
        _write_new_prevention_standard_workbook(temp_source)

        member_extra_details = collect_member_extra_details(temp_source)
        original_populate_doctor_sheet = GENERIC.populate_doctor_sheet

        def populate_doctor_sheet_with_prospect_patch(*args: Any, **kwargs: Any) -> None:
            original_populate_doctor_sheet(*args, **kwargs)
            wb_tpl = args[0]
            patch_doctor_sheet(wb_tpl, member_extra_details)

        GENERIC.populate_doctor_sheet = populate_doctor_sheet_with_prospect_patch
        try:
            temp_output = Path(GENERIC.process_excel(str(temp_source), str(template)))
        finally:
            GENERIC.populate_doctor_sheet = original_populate_doctor_sheet

        final_output = source_dir.parent / temp_output.name
        if final_output.exists():
            final_output.unlink()
        shutil.move(str(temp_output), str(final_output))
        return str(final_output)


def main(argv: Optional[List[str]] = None) -> int:
    import tkinter as tk
    from tkinter import filedialog, messagebox

    parser = argparse.ArgumentParser(description="展望前置清洗與選會員輸出")
    parser.add_argument("source_dir", nargs="?", help="來源資料夾；未指定時開啟選擇視窗")
    parser.add_argument("--no-open", action="store_true", help="完成後不自動開啟 Excel")
    parser.add_argument("--no-dialog", action="store_true", help="完成後不顯示完成視窗")
    args = parser.parse_args(argv)

    root = None
    src = args.source_dir
    if not src:
        root = tk.Tk()
        root.withdraw()
        src = filedialog.askdirectory(title="選擇展望來源資料夾")
    if not src:
        if root is not None:
            root.destroy()
        print("未選擇來源資料夾，程式已取消。")
        return 0

    try:
        out = process_excel(src)
        print(f"完成：{out}")
        if not args.no_dialog:
            if root is None:
                root = tk.Tk()
                root.withdraw()
            messagebox.showinfo("完成", f"已輸出：\n{out}", parent=root)
        if not args.no_open:
            GENERIC.open_file_cross_platform(out)
    except Exception as exc:
        print(f"錯誤：{exc}")
        if not args.no_dialog:
            if root is None:
                root = tk.Tk()
                root.withdraw()
            messagebox.showerror("錯誤", str(exc), parent=root)
        return 2
    finally:
        if root is not None:
            root.destroy()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
