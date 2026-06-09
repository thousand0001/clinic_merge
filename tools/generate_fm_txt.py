# -*- coding: utf-8 -*-
"""
產生家庭醫師整合性照護計畫名單上傳檔（FM.txt）

格式規範：MHB5 系統，定長 208 BYTES，CP950 編碼
檔名規則：業務組別(1)+醫事機構代號(10)+上傳月份(2)+流水號(2)+FM.txt
每檔最多 9999 筆，超出自動拆檔（serial 遞增）

資料來源：
  - Excel 輸入檔   → 醫事代碼、會員 ID、生日、個案類別（有就用）
  - DB staging     → 會員生日、個案類別、姓名、電話（Excel 沒有時補充）
  - DB meta.clinics → 診所地址、診所電話、業務組別

支援格式：
  A. 指定會員模板：欄 A=院所ID, B=ID, C=BIRTHDAY, D=個案類別
  B. 選會員 Excel：sheet「醫生看(從會員指標內容Key過來)」, 欄 A=ID

用法（CLI）：
    python tools/generate_fm_txt.py --input "路徑/模板或選會員.xlsx" --dest "輸出/"
    python tools/generate_fm_txt.py --hosp-id 3501013059 --input "..." --dest "..."

GUI 模式（不帶任何參數直接執行）：
    python tools/generate_fm_txt.py
"""
from __future__ import annotations

import argparse
import datetime
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from db_pipeline.normalization import branch_from_address
from db_pipeline.storage import _run_query


# ── 常數 ──────────────────────────────────────────────────────────────────────
DEFAULT_PLAN_NO = "01"
DEFAULT_PRSN_ID = "A123456789"
TODAY = datetime.date.today().strftime("%Y%m%d")


# ── 工具函式 ──────────────────────────────────────────────────────────────────
def sex_from_id(pid: str) -> str:
    if not pid or len(pid) < 2:
        return " "
    g = str(pid)[1].upper()
    if g in ("1", "8", "A", "C"):
        return "1"
    if g in ("2", "9", "B", "D"):
        return "2"
    return " "


def to_yyyymmdd(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y%m%d")
    s = str(v).strip()
    if len(s) == 8 and s.isdigit():
        return s
    for sep in ("/", "-"):
        parts = s.split(sep)
        if len(parts) == 3 and all(p.isdigit() for p in parts):
            return f"{parts[0]}{parts[1].zfill(2)}{parts[2].zfill(2)}"
    return ""


def case_abc(v) -> str:
    if v is None:
        return "A"
    s = str(v).strip().lower()
    if s == "6":
        return "C"
    if "b" in s:
        return "B"
    return "A"


def fb(value, byte_len: int, encoding: str = "cp950") -> bytes:
    b = ("" if value is None else str(value)).strip().encode(encoding, errors="replace")
    while len(b) > byte_len:
        b = b[:-1]
    return b + b" " * (byte_len - len(b))


def make_record(
    plan_no: str, branch: int, hosp_id: str,
    pid: str, birthday: str, name: str, sex: str,
    addr: str, tel: str, prsn_id: str,
    case_type: str, case_date: str,
) -> bytes:
    rec = (
        b"A"
        + plan_no.encode()
        + str(branch).encode()
        + fb(hosp_id, 10)
        + fb(pid, 10)
        + fb(birthday, 8)
        + fb(name, 12)
        + sex.encode()
        + fb(addr, 120)
        + fb(tel, 15)
        + fb(prsn_id, 10)
        + case_type.encode()
        + fb(case_date, 8)
        + fb("", 8)
        + b" "
    )
    assert len(rec) == 208, f"記錄長度錯誤：{len(rec)}"
    return rec


# ── DB 查詢 ────────────────────────────────────────────────────────────────────
def _clinic_info(hosp_id: str) -> tuple[str, str]:
    """回傳 (institution_address, institution_phone)"""
    r = _run_query(
        f"SELECT institution_address, institution_phone "
        f"FROM meta.clinics WHERE clinic_code='{hosp_id}' LIMIT 1;"
    )
    if not r:
        raise ValueError(f"DB 找不到診所代碼：{hosp_id}")
    parts = r.split("|")
    return parts[0] if parts else "", parts[1] if len(parts) > 1 else ""


def _build_db_lookup(hosp_id: str) -> dict:
    """從最新批次的 staging.members 建立 {patient_id: {...}} lookup。"""
    r = _run_query(f"""
SELECT DISTINCT ON (m.patient_id_normalized)
    m.patient_id_normalized,
    COALESCE(m.name, ''),
    COALESCE(to_char(m.birth_date, 'YYYYMMDD'), ''),
    COALESCE(NULLIF(m.phone, ''), NULLIF(m.mobile, ''), ''),
    COALESCE(m.member_type::text, ''),
    COALESCE(m.address, '')
FROM staging.members m
JOIN meta.import_batches b ON b.batch_id = m.batch_id
JOIN meta.clinics c ON c.clinic_id = b.clinic_id
WHERE c.clinic_code = '{hosp_id}'
ORDER BY m.patient_id_normalized, b.started_at DESC;
""")
    lookup: dict = {}
    if not r:
        return lookup
    for line in r.splitlines():
        parts = line.split("|")
        if len(parts) < 5:
            continue
        lookup[parts[0]] = {
            "name":        parts[1],
            "birthday":    parts[2],
            "tel":         parts[3],
            "member_type": parts[4],
            "address":     parts[5] if len(parts) > 5 else "",
        }
    return lookup


# ── 選會員 Excel 補充查找（名字/電話） ────────────────────────────────────────────
def _find_xuanhui_excel(input_path: Path, hosp_id: str) -> Path | None:
    """在輸入檔同層或上一層找對應診所的 *選會員_*.xlsx。"""
    # 從 DB 取診所名稱，用來比對檔名
    clinic_name = _run_query(
        f"SELECT clinic_name FROM meta.clinics WHERE clinic_code='{hosp_id}' LIMIT 1;"
    ) or ""
    for d in [input_path.parent, input_path.parent.parent]:
        # 先找含診所名稱的
        for f in sorted(d.glob("*選會員_*.xlsx")):
            if clinic_name and clinic_name in f.name:
                return f
        # 找不到就取第一個
        for f in sorted(d.glob("*選會員_*.xlsx")):
            return f
    return None


def _build_xuanhui_lookup(excel_path: Path) -> dict:
    """從選會員 Excel 的「醫生看」sheet 讀 pid→{name, tel}。"""
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    sheet_name = next((s for s in wb.sheetnames if "醫生看" in s), None)
    if not sheet_name:
        return {}
    ws = wb[sheet_name]
    lookup: dict = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        pid = str(row[0]).strip() if row[0] else None
        if not pid or pid == "None":
            continue
        tel = str(row[4]).strip() if len(row) > 4 and row[4] else (
              str(row[5]).strip() if len(row) > 5 and row[5] else "")
        lookup[pid] = {
            "name": str(row[1]).strip() if len(row) > 1 and row[1] else "",
            "tel":  tel,
        }
    print(f"  選會員 Excel：{excel_path.name}（{len(lookup)} 筆）")
    return lookup


# ── Excel 讀取 ────────────────────────────────────────────────────────────────
def _read_input(input_path: Path) -> tuple[str | None, list[dict]]:
    """
    讀取輸入 Excel，回傳 (hosp_id_from_file, rows)。
    每個 row dict 含 pid, name, birthday, tel, member_type（沒有的欄位為空字串）。

    支援：
      - 指定會員模板：欄 A=院所ID, B=ID, C=BIRTHDAY, D=個案類別（無姓名/電話欄）
      - 選會員 Excel：sheet 醫生看..., 欄 A=ID, B=姓名, C=生日, E/F=電話, AW=個案類別
    """
    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)

    ws = wb.active
    first_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)

    rows = []
    hosp_id_from_file = None

    if first_row and str(first_row[0] or "").strip().isdigit() and len(str(first_row[0] or "").strip()) == 10:
        # 格式 A：指定會員模板（欄 A = 院所ID，無姓名/電話）
        hosp_id_from_file = str(first_row[0]).strip()
        for row in ws.iter_rows(min_row=2, values_only=True):
            pid = str(row[1]).strip() if row[1] else None
            if not pid or pid == "None":
                continue
            rows.append({
                "pid":         pid,
                "name":        "",
                "birthday":    to_yyyymmdd(row[2]) if len(row) > 2 else "",
                "tel":         "",
                "member_type": str(row[3]).strip() if len(row) > 3 and row[3] is not None else "",
            })
    else:
        # 格式 B：選會員 Excel，找「醫生看」sheet
        sheet_name = next((s for s in wb.sheetnames if "醫生看" in s), None)
        if sheet_name is None:
            raise ValueError(
                f"無法識別 Excel 格式：{input_path.name}\n"
                "請選擇「指定會員模板」或「選會員 Excel」。"
            )
        # 從檔名推斷診所名稱（例：書田泌尿科眼科診所選會員_0609_1548.xlsx → 書田泌尿科眼科診所）
        import re as _re
        m = _re.match(r"(.+?)選會員", input_path.stem)
        if m:
            clinic_name_hint = m.group(1).strip()
            matches = search_clinics(clinic_name_hint)
            if len(matches) == 1:
                hosp_id_from_file = matches[0]["code"]
            elif len(matches) > 1:
                # 取名稱最接近的
                hosp_id_from_file = matches[0]["code"]

        ws2 = wb[sheet_name]
        for row in ws2.iter_rows(min_row=4, values_only=True):
            pid = str(row[0]).strip() if row[0] else None
            if not pid or pid == "None":
                continue
            tel = str(row[4]).strip() if len(row) > 4 and row[4] else (
                  str(row[5]).strip() if len(row) > 5 and row[5] else "")
            rows.append({
                "pid":         pid,
                "name":        str(row[1]).strip() if len(row) > 1 and row[1] else "",
                "birthday":    to_yyyymmdd(row[2]) if len(row) > 2 else "",
                "tel":         tel,
                "member_type": str(row[48]).strip() if len(row) > 48 and row[48] is not None else "",
            })

    return hosp_id_from_file, rows


# ── 診所搜尋 ──────────────────────────────────────────────────────────────────
def search_clinics(keyword: str) -> list[dict]:
    kw = keyword.strip().replace("'", "''")
    r = _run_query(
        f"SELECT clinic_code, clinic_name FROM meta.clinics "
        f"WHERE clinic_name LIKE '%{kw}%' OR official_name LIKE '%{kw}%' "
        f"ORDER BY clinic_name LIMIT 20;"
    )
    if not r:
        return []
    results = []
    for line in r.splitlines():
        parts = line.split("|")
        if len(parts) >= 2:
            results.append({"code": parts[0], "name": parts[1]})
    return results


# ── 主邏輯 ────────────────────────────────────────────────────────────────────
def generate(
    input_path: Path,
    dest_dir: Path,
    hosp_id: str | None = None,
    plan_no: str = DEFAULT_PLAN_NO,
    serial: str = "01",
    prsn_id: str = DEFAULT_PRSN_ID,
    upload_month: str | None = None,
    confirm_overwrite=None,  # callable(Path) -> bool；None = 直接覆蓋
) -> list[Path]:
    # 讀 Excel
    hosp_id_from_file, rows = _read_input(input_path)
    hosp_id = hosp_id or hosp_id_from_file
    if not hosp_id:
        raise ValueError("無法從檔案取得醫事代碼，請用 --hosp-id 指定。")

    print(f"  醫事代碼：{hosp_id}，Excel 會員筆數：{len(rows)}")

    # 診所資料
    clinic_addr, clinic_phone = _clinic_info(hosp_id)
    dest_dir.mkdir(parents=True, exist_ok=True)
    branch = branch_from_address(clinic_addr) or 1
    month  = upload_month or datetime.date.today().strftime("%m")

    # DB 會員補充
    db = _build_db_lookup(hosp_id)
    print(f"  DB 會員筆數：{len(db)}")

    # 選會員 Excel 補充（自動偵測，用於名字/電話）
    xh_path = _find_xuanhui_excel(input_path, hosp_id)
    xh = _build_xuanhui_lookup(xh_path) if xh_path else {}

    # 組 records（優先順序：輸入 Excel > DB > 選會員 Excel > 診所電話）
    records = []
    for r in rows:
        pid = r["pid"]
        m   = db.get(pid, {})
        x   = xh.get(pid, {})
        birthday    = r["birthday"]    or m.get("birthday", "")
        name        = r["name"]        or m.get("name", "") or x.get("name", "")
        tel         = r["tel"]         or m.get("tel", "") or x.get("tel", "") or clinic_phone
        addr        = r.get("address", "") or m.get("address", "")
        member_type = r["member_type"] or m.get("member_type", "")
        records.append(make_record(
            plan_no   = plan_no,
            branch    = branch,
            hosp_id   = hosp_id,
            pid       = pid,
            birthday  = birthday or "        ",
            name      = name,
            sex       = sex_from_id(pid),
            addr      = addr,
            tel       = tel,
            prsn_id   = prsn_id,
            case_type = case_abc(member_type),
            case_date = TODAY,
        ))

    # 超過 9999 自動拆檔
    MAX_PER_FILE = 9999
    chunks = [records[i:i + MAX_PER_FILE] for i in range(0, len(records), MAX_PER_FILE)]
    out_paths = []
    serial_int = int(serial)
    for chunk in chunks:
        sn = f"{serial_int:02d}"
        fname = f"{branch}{hosp_id}{month}{sn}FM.txt"
        out_path = dest_dir / fname
        if out_path.exists() and confirm_overwrite is not None:
            if not confirm_overwrite(out_path):
                print(f"  跳過：{fname}")
                serial_int += 1
                continue
        with open(out_path, "wb") as fh:
            for rec in chunk:
                fh.write(rec + b"\r\n")
        print(f"✓ {fname}：{len(chunk)} 筆，業務組別={branch}")
        out_paths.append(out_path)
        serial_int += 1

    return out_paths


# ── GUI 模式 ──────────────────────────────────────────────────────────────────
def _gui():
    try:
        import tkinter as tk
        from tkinter import filedialog, messagebox
    except ImportError:
        print("無法使用 GUI，請用 CLI 參數模式。")
        return

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    # 選 Excel 輸入檔
    input_file = filedialog.askopenfilename(
        title="選擇「指定會員模板」或「選會員 Excel」",
        filetypes=[("Excel", "*.xlsx"), ("所有檔案", "*.*")],
    )
    if not input_file:
        return

    dest = Path(input_file).parent / "fm-txt"

    def _ask_overwrite(path: Path) -> bool:
        return messagebox.askyesno("檔案已存在", f"{path.name}\n\n要覆蓋嗎？")

    try:
        out = generate(
            input_path        = Path(input_file),
            dest_dir          = dest,
            confirm_overwrite = _ask_overwrite,
        )
        messagebox.showinfo("完成", "已產生：\n" + "\n".join(str(p) for p in out))
        import subprocess
        subprocess.Popen(["open", str(dest)])
    except Exception as e:
        messagebox.showerror("錯誤", str(e))

    root.destroy()


# ── CLI 入口 ──────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) == 1:
        _gui()
        return

    parser = argparse.ArgumentParser(description="產生家醫計畫名單上傳 FM.txt")
    parser.add_argument("--input",    required=True, type=Path,
                        help="指定會員模板 或 選會員 Excel 路徑")
    parser.add_argument("--dest",     required=True, type=Path, help="輸出資料夾")
    parser.add_argument("--hosp-id",  default=None,  help="醫事機構代碼（Excel 已含時可省略）")
    parser.add_argument("--plan-no",  default=DEFAULT_PLAN_NO)
    parser.add_argument("--serial",   default="01")
    parser.add_argument("--prsn-id",  default=DEFAULT_PRSN_ID)
    parser.add_argument("--month",    default=None, help="上傳月份 MM（預設本月）")
    args = parser.parse_args()

    generate(
        input_path   = args.input,
        dest_dir     = args.dest,
        hosp_id      = args.hosp_id,
        plan_no      = args.plan_no,
        serial       = args.serial,
        prsn_id      = args.prsn_id,
        upload_month = args.month,
    )


if __name__ == "__main__":
    main()
