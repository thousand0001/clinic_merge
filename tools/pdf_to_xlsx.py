# -*- coding: utf-8 -*-
"""
PDF 轉 XLSX 工具

用途：
- 將「看診次數統計名單」PDF 文字列轉成 Excel。
- 支援單一 PDF 或資料夾批次處理。
- 輸出檔名自動加時間戳，預設存到輸入檔同資料夾。
"""

from __future__ import annotations

import argparse
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover
    PdfReader = None


DEFAULT_INPUT = Path.home()  # PDF 路徑由命令列參數或執行時指定

HEADERS = ["來源檔", "頁碼", "病歷號", "姓名", "生日", "電話", "次數", "說明", "地址", "原始列"]
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True)

LINE_PATTERN = re.compile(
    r"^\s*"
    r"(?P<chart_no>\d+)\s+"
    r"(?P<name>\S+)\s+"
    r"(?P<birthday>\d{6,7})\s+"
    r"(?P<phone>[0-9\-]+)\s+"
    r"(?P<count>\d+)"
    r"(?P<remark>\S*)"
    r"(?:\s+(?P<address>.*?))?"
    r"\s*$"
)


def parse_pdf_rows(pdf_path: Path) -> List[Dict[str, object]]:
    if PdfReader is None:
        raise RuntimeError("缺少 pypdf 套件，請先執行：.venv/bin/pip install pypdf==6.10.1")

    reader = PdfReader(str(pdf_path))
    rows: List[Dict[str, object]] = []
    for page_index, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        for raw_line in text.splitlines():
            row = parse_text_line(raw_line, pdf_path.name, page_index)
            if row is not None:
                rows.append(row)
    return rows


def parse_text_line(raw_line: str, source_name: str, page_no: int) -> Optional[Dict[str, object]]:
    line = str(raw_line or "").strip()
    if not line or line.startswith("*") or line.startswith("="):
        return None
    if "病歷號" in line and "姓名" in line:
        return None

    match = LINE_PATTERN.match(line)
    if not match:
        return None

    count_text = match.group("count")
    try:
        count_value = int(count_text)
    except ValueError:
        return None

    return {
        "來源檔": source_name,
        "頁碼": page_no,
        "病歷號": normalize_chart_no(match.group("chart_no")),
        "姓名": match.group("name").strip(),
        "生日": match.group("birthday").strip(),
        "電話": match.group("phone").strip(),
        "次數": count_value,
        "說明": (match.group("remark") or "").strip(),
        "地址": (match.group("address") or "").strip(),
        "原始列": line,
    }


def normalize_chart_no(value: object) -> str:
    text = str(value or "").strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def collect_pdf_paths(input_path: Path) -> List[Path]:
    if input_path.is_file():
        if input_path.suffix.lower() != ".pdf":
            raise ValueError(f"輸入檔不是 PDF：{input_path}")
        return [input_path]
    if input_path.is_dir():
        return [path for path in sorted(input_path.glob("*.pdf")) if not path.name.startswith("~$")]
    raise FileNotFoundError(f"找不到輸入路徑：{input_path}")


def make_output_path(input_path: Path, output_dir: Optional[Path]) -> Path:
    timestamp = datetime.now().strftime("%m%d_%H%M")
    target_dir = output_dir or (input_path if input_path.is_dir() else input_path.parent)
    stem = input_path.stem if input_path.is_file() else input_path.name
    return target_dir / f"{stem}_轉xlsx_{timestamp}.xlsx"


def write_xlsx(rows: Sequence[Dict[str, object]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("PDF轉出")
    ws.freeze_panes = "A2"

    header_cells = []
    for value in HEADERS:
        cell = WriteOnlyCell(ws, value=value)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        header_cells.append(cell)
    ws.append(header_cells)

    for row in rows:
        ws.append([row.get(header, "") for header in HEADERS])

    widths = {
        "A": 34,
        "B": 8,
        "C": 12,
        "D": 12,
        "E": 12,
        "F": 14,
        "G": 8,
        "H": 12,
        "I": 48,
        "J": 80,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    summary = wb.create_sheet("轉檔摘要")
    summary.append(["項目", "數值"])
    summary.append(["轉出筆數", len(rows)])
    summary.append(["產生時間", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    summary.append(["欄位", "、".join(HEADERS)])
    for col_idx in range(1, 3):
        summary.column_dimensions[get_column_letter(col_idx)].width = 24 if col_idx == 1 else 80

    wb.save(output_path)


def convert_pdf_to_xlsx(input_path: Path, output_dir: Optional[Path] = None) -> Path:
    pdf_paths = collect_pdf_paths(input_path)
    all_rows: List[Dict[str, object]] = []
    for pdf_path in pdf_paths:
        rows = parse_pdf_rows(pdf_path)
        all_rows.extend(rows)
        print(f"讀取完成：{pdf_path}，轉出 {len(rows)} 筆")

    if not all_rows:
        raise RuntimeError("沒有解析到任何資料列，請確認 PDF 是否為可選取文字格式。")

    output_path = make_output_path(input_path, output_dir)
    write_xlsx(all_rows, output_path)
    return output_path


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="將看診次數統計 PDF 轉成 XLSX。")
    parser.add_argument(
        "input",
        nargs="?",
        default=str(DEFAULT_INPUT),
        help="PDF 檔案或 PDF 資料夾。未指定時使用目前 0603 看診次數 PDF。",
    )
    parser.add_argument(
        "-o",
        "--output-dir",
        default="",
        help="輸出資料夾。未指定時輸出到 PDF 同資料夾。",
    )
    return parser


def main(argv: Optional[Iterable[str]] = None) -> int:
    parser = build_parser()
    args = parser.parse_args(list(argv) if argv is not None else None)

    input_path = Path(args.input).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve() if args.output_dir else None

    try:
        output_path = convert_pdf_to_xlsx(input_path, output_dir)
    except Exception as exc:
        print(f"轉檔失敗：{exc}")
        return 1

    print("轉檔成功")
    print(f"輸出檔：{output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
