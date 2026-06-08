# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import openpyxl


def norm(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.endswith(" 00:00:00"):
        return text[:-9]
    return text


def load_by_id(
    path: Path,
    sheet_name: str,
    header_row: int,
    data_start: int,
    id_col: int,
    max_col: int,
    wanted_ids: Optional[set] = None,
    max_needed: Optional[int] = None,
) -> Tuple[List[str], Dict[str, List[Any]], List[str]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        headers: List[str] = []
        rows: Dict[str, List[Any]] = {}
        order: List[str] = []
        for ridx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max_col, values_only=True),
            start=1,
        ):
            if ridx == header_row:
                headers = [norm(value) or f"欄{i}" for i, value in enumerate(row, start=1)]
                continue
            if ridx < data_start:
                continue
            pid = norm(row[id_col - 1]).upper()
            if not pid:
                continue
            if wanted_ids is not None and pid not in wanted_ids:
                continue
            if pid not in rows:
                rows[pid] = list(row)
                order.append(pid)
                if max_needed and len(rows) >= max_needed:
                    break
        return headers, rows, order
    finally:
        wb.close()


def compare_sheet(
    old_path: Path,
    new_path: Path,
    sheet_name: str,
    header_row: int,
    data_start: int,
    id_col: int,
    max_col: int,
    sample_count: int,
) -> None:
    old_headers, old_rows, old_order = load_by_id(
        old_path, sheet_name, header_row, data_start, id_col, max_col, max_needed=max(sample_count * 3, 80)
    )
    wanted = set(old_order[: max(sample_count * 3, 80)])
    new_headers, new_rows, _ = load_by_id(
        new_path, sheet_name, header_row, data_start, id_col, max_col, wanted_ids=wanted
    )
    common = [pid for pid in old_order if pid in new_rows][:sample_count]
    print(f"\n[{sheet_name}] common={len(common)} old_loaded={len(old_rows)} new_matched={len(new_rows)}")
    print("IDs:", ", ".join(common))

    agg: Dict[Tuple[int, str, str], int] = {}
    examples = []
    for pid in common:
        diffs = []
        for idx in range(max_col):
            old_value = norm(old_rows[pid][idx])
            new_value = norm(new_rows[pid][idx])
            if old_value != new_value:
                key = (idx + 1, old_headers[idx], new_headers[idx])
                agg[key] = agg.get(key, 0) + 1
                diffs.append((idx + 1, old_headers[idx], new_headers[idx], old_value, new_value))
        if len(examples) < 3:
            examples.append((pid, diffs[:30], len(diffs)))

    for pid, diffs, count in examples:
        print(f"\nExample {pid}: diff_count={count}")
        for diff in diffs:
            print(" ", diff)

    print("\nDiff columns:")
    for key, count in sorted(agg.items(), key=lambda item: (-item[1], item[0][0])):
        print(count, key)


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="比對舊 run_merge 輸出與通用輸出共同欄位")
    parser.add_argument("old")
    parser.add_argument("new")
    parser.add_argument("--count", type=int, default=30)
    args = parser.parse_args(argv)

    old_path = Path(args.old)
    new_path = Path(args.new)
    compare_sheet(old_path, new_path, "會員總表", 1, 3, 5, 50, args.count)
    compare_sheet(old_path, new_path, "醫生看(從會員指標內容Key過來)", 1, 4, 1, 48, args.count)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
