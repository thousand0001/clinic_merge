#!/usr/bin/env python3
from __future__ import annotations

import argparse
import tkinter as tk
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

import openpyxl


class ProgressDialog:
    def __init__(self, root: tk.Tk) -> None:
        self.window = tk.Toplevel(root)
        self.window.title("ODS 轉 XLSX 進度")
        self.window.geometry("560x190")
        self.window.resizable(False, False)
        self.window.attributes("-topmost", True)
        self.cancelled = False

        self.title_var = tk.StringVar(value="準備開始")
        self.detail_var = tk.StringVar(value="")
        self.count_var = tk.StringVar(value="0 / 0")

        frame = tk.Frame(self.window, padx=16, pady=16)
        frame.pack(fill="both", expand=True)

        tk.Label(frame, textvariable=self.title_var, anchor="w", justify="left").pack(fill="x")
        tk.Label(frame, textvariable=self.detail_var, anchor="w", justify="left", wraplength=520).pack(fill="x", pady=(8, 12))

        self.progress = ttk.Progressbar(frame, orient="horizontal", mode="determinate", length=520)
        self.progress.pack(fill="x")

        tk.Label(frame, textvariable=self.count_var, anchor="e", justify="right").pack(fill="x", pady=(8, 0))
        tk.Button(frame, text="終止", width=10, command=self.cancel).pack(anchor="e", pady=(10, 0))

        self.window.update_idletasks()

    def update(self, current: int, total: int, title: str, detail: str) -> None:
        self.title_var.set(title)
        self.detail_var.set(detail)
        self.count_var.set(f"{current} / {total}")
        self.progress["maximum"] = max(total, 1)
        self.progress["value"] = current
        self.window.update_idletasks()

    def cancel(self) -> None:
        self.cancelled = True
        self.title_var.set("正在終止")
        self.window.update_idletasks()

    def raise_if_cancelled(self) -> None:
        if self.cancelled:
            raise InterruptedError("使用者已終止轉換。")

    def close(self) -> None:
        if self.window.winfo_exists():
            self.window.destroy()


def load_ods_as_workbook(ods_path: Path) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    office_ns = "urn:oasis:names:tc:opendocument:xmlns:office:1.0"
    table_ns = "urn:oasis:names:tc:opendocument:xmlns:table:1.0"
    text_ns = "urn:oasis:names:tc:opendocument:xmlns:text:1.0"
    q_spreadsheet = f"{{{office_ns}}}spreadsheet"
    q_table = f"{{{table_ns}}}table"
    q_row = f"{{{table_ns}}}table-row"
    q_cell = f"{{{table_ns}}}table-cell"
    q_covered_cell = f"{{{table_ns}}}covered-table-cell"
    q_name = f"{{{table_ns}}}name"
    q_col_repeat = f"{{{table_ns}}}number-columns-repeated"
    q_row_repeat = f"{{{table_ns}}}number-rows-repeated"
    q_paragraph = f"{{{text_ns}}}p"

    with zipfile.ZipFile(ods_path, "r") as zf:
        root = ET.fromstring(zf.read("content.xml"))

    spreadsheet = root.find(f".//{q_spreadsheet}")
    if spreadsheet is None:
        wb.create_sheet(title="Sheet1")
        return wb

    table_count = 0
    for idx, table in enumerate(spreadsheet, start=1):
        if table.tag != q_table:
            continue

        table_count += 1
        title = str(table.attrib.get(q_name) or f"Sheet{idx}")
        ws = wb.create_sheet(title=title[:31] or f"Sheet{idx}")

        for row in table:
            if row.tag != q_row:
                continue

            values: list[str] = []
            pending_blank_repeat = 0

            for cell in row:
                if cell.tag not in (q_cell, q_covered_cell):
                    continue

                repeat = int(cell.attrib.get(q_col_repeat, "1") or "1")
                text_parts = [
                    "".join(paragraph.itertext()).strip()
                    for paragraph in cell.iter(q_paragraph)
                ]
                text_parts = [part for part in text_parts if part]
                text = "\n".join(text_parts)

                if text:
                    if pending_blank_repeat:
                        values.extend([""] * pending_blank_repeat)
                        pending_blank_repeat = 0
                    values.extend([text] * repeat)
                else:
                    pending_blank_repeat += repeat

            if not values:
                continue

            row_repeat = int(row.attrib.get(q_row_repeat, "1") or "1")
            ws.append(values)
            if row_repeat > 1:
                row_copy = list(values)
                for _ in range(row_repeat - 1):
                    ws.append(row_copy)

    if table_count == 0:
        wb.create_sheet(title="Sheet1")
    return wb


def confirm_overwrite(root: tk.Tk, xlsx_path: Path) -> bool:
    return messagebox.askyesno(
        "檔案已存在",
        f"已存在同名檔案：\n{xlsx_path}\n\n是否覆蓋？",
        parent=root,
    )


def build_output_folder(folder: Path) -> Path:
    return folder.parent / f"{folder.name}_x"


def convert_folder(
    folder: Path,
    root: tk.Tk,
    progress: ProgressDialog,
) -> tuple[Path, list[Path], list[Path]]:
    if not folder.exists() or not folder.is_dir():
        raise NotADirectoryError(f"資料夾不存在: {folder}")

    output_folder = build_output_folder(folder)
    output_folder.mkdir(parents=True, exist_ok=True)

    converted_files: list[Path] = []
    skipped_files: list[Path] = []
    ods_files = sorted(folder.glob("**/*.ods"))
    total = len(ods_files)

    progress.update(0, total, "掃描完成", f"來源資料夾：{folder}")
    progress.raise_if_cancelled()

    for index, ods_path in enumerate(ods_files, start=1):
        progress.raise_if_cancelled()
        relative_path = ods_path.relative_to(folder)
        xlsx_path = (output_folder / relative_path).with_suffix(".xlsx")
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)

        progress.update(
            index - 1,
            total,
            f"準備轉換第 {index} 個檔案",
            f"{relative_path}",
        )
        progress.raise_if_cancelled()

        if xlsx_path.exists() and not confirm_overwrite(root, xlsx_path):
            skipped_files.append(xlsx_path)
            progress.update(
                index,
                total,
                f"略過第 {index} 個檔案",
                f"{relative_path}",
            )
            continue

        workbook = load_ods_as_workbook(ods_path)
        progress.raise_if_cancelled()
        workbook.save(xlsx_path)
        converted_files.append(xlsx_path)
        progress.update(
            index,
            total,
            f"已完成第 {index} 個檔案",
            f"{relative_path}",
        )

    return output_folder, converted_files, skipped_files


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="將資料夾內的 ODS 批次轉成 XLSX，輸出到同層的 _x 資料夾。")
    parser.add_argument("folder", nargs="?", help="ODS 所在資料夾路徑")
    return parser


def create_root() -> tk.Tk:
    root = tk.Tk()
    root.title("")
    root.geometry("0x0+0+0")
    root.overrideredirect(True)
    root.attributes("-alpha", 0)
    root.withdraw()
    root.attributes("-topmost", True)
    return root


def choose_folder_with_dialog(root: tk.Tk) -> Path | None:
    root.withdraw()
    selected = filedialog.askdirectory(title="選擇要轉換 ODS 的資料夾")
    root.withdraw()
    return Path(selected).resolve() if selected else None


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    root = create_root()
    progress: ProgressDialog | None = None
    try:
        if args.folder:
            folder = Path(args.folder).expanduser().resolve()
        else:
            folder = choose_folder_with_dialog(root)
            if folder is None:
                print("未選擇資料夾，已取消。")
                return 1

        progress = ProgressDialog(root)
        output_folder, converted_files, skipped_files = convert_folder(folder, root, progress)

        if not converted_files and not skipped_files:
            progress.update(0, 0, "完成", f"輸出資料夾：{output_folder}")
            print("這個資料夾和子資料夾內沒有找到 ODS 檔案。")
            return 0

        progress.update(
            len(converted_files) + len(skipped_files),
            len(converted_files) + len(skipped_files),
            "轉換完成",
            f"輸出資料夾：{output_folder}",
        )
        print(f"輸出資料夾: {output_folder}")
        print(f"已轉換 {len(converted_files)} 個檔案:")
        for path in converted_files:
            print(path)
        if skipped_files:
            print(f"略過 {len(skipped_files)} 個檔案:")
            for path in skipped_files:
                print(path)
        return 0
    except InterruptedError as exc:
        print(str(exc))
        return 1
    except Exception as exc:
        if progress is not None:
            progress.update(0, 1, "轉換失敗", str(exc))
        messagebox.showerror("轉換失敗", str(exc), parent=root)
        return 1
    finally:
        if progress is not None:
            progress.close()
        root.destroy()


if __name__ == "__main__":
    raise SystemExit(main())
