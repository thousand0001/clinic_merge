# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import datetime as dt
import re
import shutil
import tempfile
import zipfile
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence
from xml.etree import ElementTree as ET
from xml.sax.saxutils import escape

import openpyxl
from openpyxl.utils.datetime import to_excel


DEFAULT_TEMPLATE = Path("115年6月指定名單格式") / "115指定名單格式.xlsx"
DEFAULT_MOHW_FILE = Path("醫療群_衛福部資料.xlsx")
OFFICIAL_HEADERS = [
    "院所ID",
    "ID",
    "BIRTHDAY",
    "個案類別",
    "論質名單",
    "65歲以上多重慢性病註記",
    "高診次註記",
    "慢性病註記",
    "非慢性病註記",
    "與前一年家醫收案診所相同",
    "疾病樣態",
    "ASCVD",
    "三高",
    "高血壓",
    "高血脂",
    "高血糖",
]
DATE_COLUMNS = {3}
NUMBER_COLUMNS = {5, 6, 7, 8, 9, 10, 11, 13, 14, 15, 16}
NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
ET.register_namespace("", NS_MAIN)
ET.register_namespace("r", NS_REL)


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value).strip())


def normalize_id(value: Any) -> str:
    text = normalize_text(value).upper()
    if text.endswith(".0"):
        text = text[:-2]
    return text.replace(" ", "")


def safe_filename_part(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r'[\\/:*?"<>|]+', "", text)
    text = re.sub(r"\s+", "", text)
    return text or "院所"


def clinic_display_name(name: str) -> str:
    name = safe_filename_part(name)
    return re.sub(r"(診所|醫院)$", "", name) or name


def find_header_row(ws, required: Sequence[str], max_scan_rows: int = 10) -> Optional[int]:
    required_keys = {normalize_text(name) for name in required}
    for row_idx in range(1, min(ws.max_row, max_scan_rows) + 1):
        row_keys = {normalize_text(cell.value) for cell in ws[row_idx] if normalize_text(cell.value)}
        if required_keys.issubset(row_keys):
            return row_idx
    return None


def build_header_map(ws, row_idx: int) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for cell in ws[row_idx]:
        key = normalize_text(cell.value)
        if key and key not in out:
            out[key] = cell.column
    return out


def lookup_clinic_name_from_mohw(mohw_path: Path, clinic_code: str) -> Optional[str]:
    if not clinic_code or not mohw_path.exists():
        return None

    wb = openpyxl.load_workbook(mohw_path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            header_row = find_header_row(ws, ["醫事機構代碼", "醫事機構名稱"], max_scan_rows=10)
            if header_row is None:
                continue
            header_map = build_header_map(ws, header_row)
            code_col = header_map[normalize_text("醫事機構代碼")]
            name_col = header_map[normalize_text("醫事機構名稱")]
            for row in ws.iter_rows(min_row=header_row + 1, max_col=max(code_col, name_col), values_only=True):
                row_code = normalize_text(row[code_col - 1] if code_col <= len(row) else "")
                if row_code == clinic_code:
                    return clinic_display_name(row[name_col - 1] if name_col <= len(row) else "")
    finally:
        wb.close()
    return None


def column_index_from_letters(letters: str) -> int:
    index = 0
    for char in letters:
        index = index * 26 + ord(char.upper()) - ord("A") + 1
    return index


def cell_column_index(cell_ref: str) -> int:
    match = re.match(r"([A-Z]+)", cell_ref)
    if not match:
        return 0
    return column_index_from_letters(match.group(1))


def column_letters(index: int) -> str:
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(ord("A") + remainder) + letters
    return letters


def qname(name: str) -> str:
    return f"{{{NS_MAIN}}}{name}"


def find_designated_sheet(wb, clinic_code: Optional[str] = None):
    if clinic_code and clinic_code in wb.sheetnames:
        return wb[clinic_code]
    for ws in wb.worksheets:
        header_row = find_header_row(ws, ["院所ID", "ID", "BIRTHDAY"], max_scan_rows=8)
        if header_row:
            return ws
    raise ValueError("來源檔找不到 115 指定名單格式欄位：院所ID / ID / BIRTHDAY")


def iter_designated_rows(ws) -> Iterable[List[Any]]:
    header_row = find_header_row(ws, ["院所ID", "ID", "BIRTHDAY"], max_scan_rows=8)
    if header_row is None:
        raise ValueError(f"工作表「{ws.title}」找不到指定名單表頭")
    source_cols = build_header_map(ws, header_row)
    missing = [header for header in OFFICIAL_HEADERS if normalize_text(header) not in source_cols]
    if missing:
        raise ValueError(f"來源工作表「{ws.title}」缺少欄位：" + "、".join(missing))

    column_indexes = [source_cols[normalize_text(header)] for header in OFFICIAL_HEADERS]
    max_source_col = max(column_indexes)
    for row in ws.iter_rows(min_row=header_row + 1, max_col=max_source_col, values_only=True):
        values = [row[col_idx - 1] if col_idx <= len(row) else None for col_idx in column_indexes]
        clinic_code = normalize_text(values[0])
        pid = normalize_id(values[1])
        if not clinic_code or not pid:
            continue
        values[1] = pid
        yield values


def make_inline_cell(row_idx: int, col_idx: int, value: Any, style_id: Optional[str]) -> ET.Element:
    cell = ET.Element(qname("c"), {"r": f"{column_letters(col_idx)}{row_idx}", "t": "inlineStr"})
    if style_id:
        cell.set("s", style_id)
    inline = ET.SubElement(cell, qname("is"))
    text = ET.SubElement(inline, qname("t"))
    text.text = "" if value is None else str(value)
    return cell


def make_inline_cell_xml(row_idx: int, col_idx: int, value: Any, style_id: Optional[str]) -> bytes:
    cell_ref = f"{column_letters(col_idx)}{row_idx}"
    style_attr = f' s="{style_id}"' if style_id else ""
    if value is None or value == "":
        return f'<c r="{cell_ref}"{style_attr}/>'.encode("utf-8")

    if col_idx in DATE_COLUMNS and isinstance(value, (dt.datetime, dt.date)):
        serial = to_excel(value)
        serial_text = str(int(serial)) if float(serial).is_integer() else str(serial)
        return f'<c r="{cell_ref}"{style_attr}><v>{serial_text}</v></c>'.encode("utf-8")

    if col_idx in NUMBER_COLUMNS:
        text_value = str(value).strip()
        if re.fullmatch(r"-?\d+(?:\.\d+)?", text_value):
            return f'<c r="{cell_ref}"{style_attr}><v>{text_value}</v></c>'.encode("utf-8")

    attrs = f' r="{column_letters(col_idx)}{row_idx}" t="inlineStr"'
    if style_id:
        attrs += f' s="{style_id}"'
    text = "" if value is None else escape(str(value), {'"': "&quot;"})
    return f"<c{attrs}><is><t>{text}</t></is></c>".encode("utf-8")


def cell_sort_key(cell: ET.Element) -> int:
    return cell_column_index(cell.get("r", ""))


def cell_xml_column_index(cell_xml: bytes) -> int:
    match = re.search(rb'\br="([A-Z]+)\d+"', cell_xml)
    if not match:
        return 0
    return column_index_from_letters(match.group(1).decode("ascii"))


def extract_row_xml(sheet_xml: bytes, row_idx: int) -> Optional[bytes]:
    match = re.search(rb'<row\b(?=[^>]*\br="' + str(row_idx).encode("ascii") + rb'")[^>]*>.*?</row>', sheet_xml)
    return match.group(0) if match else None


def extract_a_to_p_styles(sheet_xml: bytes, template_row_idx: int = 3) -> Dict[int, Optional[str]]:
    row_xml = extract_row_xml(sheet_xml, template_row_idx)
    if not row_xml:
        return {}

    styles: Dict[int, Optional[str]] = {}
    for cell_match in re.finditer(rb'<c\b[^>]*\br="([A-Z]+)' + str(template_row_idx).encode("ascii") + rb'"[^>]*(?:/>|>.*?</c>)', row_xml):
        col_idx = column_index_from_letters(cell_match.group(1).decode("ascii"))
        if 1 <= col_idx <= len(OFFICIAL_HEADERS):
            style_match = re.search(rb'\bs="([^"]*)"', cell_match.group(0))
            styles[col_idx] = style_match.group(1).decode("utf-8") if style_match else None
    return styles


def remove_a_to_p_cells(row_xml: bytes) -> bytes:
    row_open_end = row_xml.find(b">") + 1
    pos = row_open_end
    cell_pattern = re.compile(rb'<c\b[^>]*\br="[A-Z]+\d+"[^>]*(?:/>|>.*?</c>)')

    while True:
        match = cell_pattern.match(row_xml, pos)
        if not match:
            break
        if cell_xml_column_index(match.group(0)) > len(OFFICIAL_HEADERS):
            break
        pos = match.end()

    return row_xml[:row_open_end] + row_xml[pos:]


def prepend_cells_to_row(row_xml: bytes, cells_xml: bytes) -> bytes:
    insert_at = row_xml.find(b">") + 1
    return row_xml[:insert_at] + cells_xml + row_xml[insert_at:]


def get_row_index(row_xml: bytes) -> Optional[int]:
    row_open_end = row_xml.find(b">")
    if row_open_end < 0:
        return None
    match = re.search(rb'\br="(\d+)"', row_xml[:row_open_end])
    return int(match.group(1)) if match else None


def make_row_xml(row_idx: int, values: Sequence[Any], styles: Dict[int, Optional[str]]) -> bytes:
    cells = b"".join(make_inline_cell_xml(row_idx, col_idx, value, styles.get(col_idx)) for col_idx, value in enumerate(values, start=1))
    return f'<row r="{row_idx}">'.encode("ascii") + cells + b"</row>"


def patch_template_sheet_xml(sheet_xml: bytes, rows: Sequence[Sequence[Any]], header_row: int = 2) -> bytes:
    data_start = header_row + 1
    output_last_row = data_start + len(rows) - 1
    style_by_col = extract_a_to_p_styles(sheet_xml, template_row_idx=data_start)

    sheet_data_match = re.search(rb"<sheetData>.*?</sheetData>", sheet_xml)
    if not sheet_data_match:
        raise ValueError("115指定名單樣板 sheet1.xml 缺少 sheetData")

    seen_rows = set()

    def patch_row(row_xml: bytes) -> bytes:
        row_idx = get_row_index(row_xml)
        if row_idx is None:
            return row_xml
        if row_idx < data_start:
            return row_xml

        seen_rows.add(row_idx)
        row_xml = remove_a_to_p_cells(row_xml)
        if row_idx > output_last_row:
            return row_xml

        cells_xml = b"".join(
            make_inline_cell_xml(row_idx, col_idx, value, style_by_col.get(col_idx))
            for col_idx, value in enumerate(rows[row_idx - data_start], start=1)
        )
        return prepend_cells_to_row(row_xml, cells_xml)

    sheet_data = sheet_data_match.group(0)
    chunks: List[bytes] = []
    pos = 0
    while True:
        row_start = sheet_data.find(b"<row", pos)
        if row_start < 0:
            chunks.append(sheet_data[pos:])
            break
        row_end = sheet_data.find(b"</row>", row_start)
        if row_end < 0:
            chunks.append(sheet_data[pos:])
            break
        row_end += len(b"</row>")
        chunks.append(sheet_data[pos:row_start])
        chunks.append(patch_row(sheet_data[row_start:row_end]))
        pos = row_end
    patched_sheet_data = b"".join(chunks)

    missing_rows = [
        make_row_xml(row_idx, rows[row_idx - data_start], style_by_col)
        for row_idx in range(data_start, output_last_row + 1)
        if row_idx not in seen_rows
    ]
    if missing_rows:
        patched_sheet_data = patched_sheet_data.replace(b"</sheetData>", b"".join(missing_rows) + b"</sheetData>")

    return sheet_xml[: sheet_data_match.start()] + patched_sheet_data + sheet_xml[sheet_data_match.end() :]


def write_patched_workbook(template_path: Path, output_path: Path, rows: Sequence[Sequence[Any]]) -> None:
    with zipfile.ZipFile(template_path, "r") as zin:
        with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "xl/worksheets/sheet1.xml":
                    data = patch_template_sheet_xml(data, rows)
                zout.writestr(item, data)


def export_designated_list(
    source_path: Path,
    template_path: Path,
    output_dir: Optional[Path] = None,
    clinic_code: Optional[str] = None,
    mohw_path: Optional[Path] = None,
) -> Path:
    wb_src = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    try:
        ws_src = find_designated_sheet(wb_src, clinic_code=clinic_code)
        rows = list(iter_designated_rows(ws_src))
    finally:
        wb_src.close()

    if not rows:
        raise ValueError("來源檔沒有可輸出的指定名單資料")

    out_dir = output_dir or source_path.parent
    out_dir.mkdir(parents=True, exist_ok=True)
    row_clinic_code = normalize_text(rows[0][0])
    clinic = lookup_clinic_name_from_mohw(mohw_path, row_clinic_code) if mohw_path else None
    clinic = clinic or row_clinic_code or "院所"
    stamp = dt.datetime.now().strftime("%m%d_%H%M")
    output_path = out_dir / f"{clinic}_115指定名單_{stamp}.xlsx"

    with tempfile.NamedTemporaryFile(dir=out_dir, suffix=".xlsx", delete=False) as tmp:
        temp_path = Path(tmp.name)
    try:
        write_patched_workbook(template_path, temp_path, rows)
        shutil.move(str(temp_path), output_path)
    finally:
        if temp_path.exists():
            temp_path.unlink()
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="輸出 115 年指定名單官方格式")
    parser.add_argument("source", help="含 115 指定名單欄位的來源 Excel")
    parser.add_argument("--template", default=str(DEFAULT_TEMPLATE), help="115指定名單格式.xlsx")
    parser.add_argument("--output-dir", default=None, help="輸出資料夾；預設同來源檔")
    parser.add_argument("--clinic-code", default=None, help="指定來源工作表名稱/院所代碼")
    parser.add_argument("--mohw-file", default=str(DEFAULT_MOHW_FILE), help="醫療群_衛福部資料.xlsx，用於院所代碼轉診所名稱")
    args = parser.parse_args()

    output = export_designated_list(
        source_path=Path(args.source).expanduser().resolve(),
        template_path=Path(args.template).expanduser().resolve(),
        output_dir=Path(args.output_dir).expanduser().resolve() if args.output_dir else None,
        clinic_code=args.clinic_code,
        mohw_path=Path(args.mohw_file).expanduser().resolve() if args.mohw_file else None,
    )
    print(output)


if __name__ == "__main__":
    main()
