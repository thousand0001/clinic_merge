from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


def main() -> int:
    if len(sys.argv) != 3:
        print("usage: compare_xlsx_hidden.py <old.xlsx> <new.xlsx>")
        return 2

    p1 = Path(sys.argv[1])
    p2 = Path(sys.argv[2])
    wb1 = openpyxl.load_workbook(p1, data_only=False)
    wb2 = openpyxl.load_workbook(p2, data_only=False)

    diffs = []
    all_sheets = list(dict.fromkeys(wb1.sheetnames + wb2.sheetnames))
    for sheet_name in all_sheets:
        if sheet_name not in wb1.sheetnames or sheet_name not in wb2.sheetnames:
            diffs.append((sheet_name, "SHEET_MISSING", sheet_name in wb1.sheetnames, sheet_name in wb2.sheetnames))
            continue
        ws1 = wb1[sheet_name]
        ws2 = wb2[sheet_name]
        max_col = max(ws1.max_column, ws2.max_column)
        max_row = max(ws1.max_row, ws2.max_row)
        for col in range(1, max_col + 1):
            h1 = bool(ws1.column_dimensions[get_column_letter(col)].hidden)
            h2 = bool(ws2.column_dimensions[get_column_letter(col)].hidden)
            if h1 != h2:
                diffs.append((sheet_name, "COL", get_column_letter(col), h1, h2))
        for row in range(1, max_row + 1):
            h1 = bool(ws1.row_dimensions[row].hidden)
            h2 = bool(ws2.row_dimensions[row].hidden)
            if h1 != h2:
                diffs.append((sheet_name, "ROW", row, h1, h2))

    print(f"HIDDEN_DIFF_COUNT {len(diffs)}")
    for item in diffs[:200]:
        print(repr(item))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
