#!/usr/bin/env python3
"""
撈月份_0319v7.py

從「門診診療次數月報表」產生「撈出資料」格式的 Excel 檔
輸出檔案名稱：前6字元_撈出_mmddhhmm.xlsx

changelog：
  v4  標題列自動偵測；身分證清洗驗證；欄位模糊比對；電話正規化；病1+病23合併
  v5  process_monthly() 統一用 clean_id()+is_valid_id()；phone sheet 含姓名欄
  v6  輸出檔名加時間戳記；完成後對話框可直接開啟檔案
  v7  tkinter root 統一管理，不再開第二個 Tk instance；datetime 移至頂層 import；
      build_phone() 無資料時空 DataFrame 欄位與 header 一致
"""

import pandas as pd
import re
import os
import sys
import datetime
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────
# 欄位名稱關鍵字（模糊比對用）
KW_ID         = '身分證號'
KW_ID_ALT     = '身份證號'
KW_NAME       = '姓名'
KW_DATE       = '日期'
KW_DX1        = '病1'
KW_DX23       = '病23'
KW_COUNT      = '次數'
KW_TOTAL      = '總額'
KW_PHONE_KEYS = ['電話', '電話號碼', '聯絡電話', '手機', '行動電話', '聯絡手機']
KW_DATE_KEYS  = ['日期', '就醫日期', '門診日期', '看診日期']
KW_COUNT_EXCLUDES = ['次數0', '次數01', '次數1', '門診次數0', '門診次數01']


def normalize_text(v) -> str:
    if pd.isna(v):
        return ''
    return str(v).replace('\n', '').replace('\r', '').replace('\u3000', '').strip()


def clean_id(v) -> str:
    """
    身分證號清洗：
    - 去空白 / 前導單引號 / 前後特殊字元
    - 轉大寫
    """
    if pd.isna(v):
        return ''
    s = str(v).strip()
    s = s.lstrip("'").strip()
    s = s.upper()
    s = re.sub(r'^[^A-Z0-9]+|[^A-Z0-9]+$', '', s)
    s = re.sub(r'\s+', '', s)
    return s


def is_valid_id(v) -> bool:
    s = clean_id(v)
    return bool(re.match(r'^[A-Z]{1,2}\d{8,9}$', s))


def find_header_row(df: pd.DataFrame) -> int:
    """
    搜尋標題列：
    1. 優先找同列同時含「身分證號/身份證號 + 姓名 + 日期」
    2. 其次找同列同時含「身分證號/身份證號 + 姓名」
    3. 最後才退回任何含「身分證號/身份證號」的列
    """
    fallback = 3

    for i, row in df.iterrows():
        vals = [normalize_text(v) for v in row if pd.notna(v)]
        has_id = any((KW_ID in v) or (KW_ID_ALT in v) for v in vals)
        has_name = any(KW_NAME in v for v in vals)
        has_date = any(any(k in v for k in KW_DATE_KEYS) for v in vals)
        if has_id and has_name and has_date:
            return i

    for i, row in df.iterrows():
        vals = [normalize_text(v) for v in row if pd.notna(v)]
        has_id = any((KW_ID in v) or (KW_ID_ALT in v) for v in vals)
        has_name = any(KW_NAME in v for v in vals)
        if has_id and has_name:
            return i

    for i, row in df.iterrows():
        vals = [normalize_text(v) for v in row if pd.notna(v)]
        if any((KW_ID in v) or (KW_ID_ALT in v) for v in vals):
            return i

    return fallback


def get_col(header_row: pd.Series, keyword: str, exclude: str = None) -> int:
    """
    從標題列找含 keyword 的欄位 index（0-based）
    exclude：排除含此字的欄位
    找不到回傳 -1
    """
    for idx, val in enumerate(header_row):
        s = normalize_text(val)
        if keyword in s:
            if exclude and exclude in s:
                continue
            return idx
    return -1


def get_col_by_keys(header_row: pd.Series, include_keys, exclude_keys=None) -> int:
    """
    多關鍵字找欄位 index
    - 先精準 match
    - 再模糊 match
    """
    exclude_keys = exclude_keys or []

    for idx, val in enumerate(header_row):
        s = normalize_text(val)
        if s in include_keys:
            if any(ex in s for ex in exclude_keys):
                continue
            return idx

    for idx, val in enumerate(header_row):
        s = normalize_text(val)
        if any(k in s for k in include_keys):
            if any(ex in s for ex in exclude_keys):
                continue
            return idx

    return -1

# ──────────────────────────────────────────────
def load_input(path: str) -> dict:
    """讀取所有工作表，回傳 {sheet_name: DataFrame}"""
    print(f"[讀取] {path}")
    all_sheets = pd.read_excel(path, sheet_name=None, header=None)
    print(f"  → 共 {len(all_sheets)} 個工作表：{list(all_sheets.keys())}")
    return all_sheets

# ──────────────────────────────────────────────
def normalize_phone(phone) -> str:
    """
    電話號碼正規化：
    - 去掉非數字字元
    - 886 開頭去除國碼補 0（如 886912345678 → 0912345678）
    - 空值回傳空字串
    """
    if pd.isna(phone):
        return ''
    p = ''.join(ch for ch in str(phone) if ch.isdigit())
    if p.startswith('886'):
        p = '0' + p[3:]
    return p

# ──────────────────────────────────────────────
def process_monthly(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    """
    將單月工作表整理為：身份證號 | 姓名 | 件數| 申請金額
    同一人可能多筆，加總次數與總額
    """
    hdr_row_idx = find_header_row(df)
    header_row  = df.iloc[hdr_row_idx]
    data        = df.iloc[hdr_row_idx + 1:].copy()

    col_id    = get_col_by_keys(header_row, [KW_ID, KW_ID_ALT])
    col_name  = get_col(header_row, KW_NAME)
    col_count = get_col_by_keys(header_row, [KW_COUNT], exclude_keys=KW_COUNT_EXCLUDES)
    col_total = get_col(header_row, KW_TOTAL)

    if any(c == -1 for c in [col_id, col_name, col_count, col_total]):
        print(f"  [{sheet_name}] 找不到必要欄位，略過")
        return pd.DataFrame(columns=['身份證號', '姓名', '件數', '申請金額'])

    # 過濾有效身份證：先清洗再驗證
    data = data[data[col_id].notna()].copy()
    data[col_id] = data[col_id].apply(clean_id)
    data = data[data[col_id].apply(is_valid_id)]

    if data.empty:
        print(f"  [{sheet_name}] 無有效資料，略過")
        return pd.DataFrame(columns=['身份證號', '姓名', '件數', '申請金額'])

    subset = data[[col_id, col_name, col_count, col_total]].copy()
    subset.columns = ['身份證號', '姓名', '件數', '申請金額']
    subset['件數']    = pd.to_numeric(subset['件數'],    errors='coerce').fillna(0)
    subset['申請金額'] = pd.to_numeric(subset['申請金額'], errors='coerce').fillna(0)

    grouped = (
        subset.groupby(['身份證號', '姓名'], as_index=False, sort=False)
              .agg({'件數': 'sum', '申請金額': 'sum'})
    )
    grouped['件數']    = grouped['件數'].astype(int)
    grouped['申請金額'] = grouped['申請金額'].astype(int)

    print(f"  [{sheet_name}] {len(grouped)} 人，次數合計 {grouped['件數'].sum()}，"
          f"總額合計 {grouped['申請金額'].sum():,}")
    return grouped

# ──────────────────────────────────────────────
def merge_dx(dx1, dx23) -> str:
    """
    合併病1與病23：
    - 去掉尾巴多餘的逗號與空白
    - 中間用 , 連接
    - 任一為空則只輸出有值的部分
    """
    d1  = str(dx1).strip().rstrip(',').strip()  if pd.notna(dx1)  and str(dx1).strip()  not in ('', 'nan') else ''
    d23 = str(dx23).strip().rstrip(',').strip() if pd.notna(dx23) and str(dx23).strip() not in ('', 'nan') else ''
    parts = [p for p in [d1, d23] if p]
    return ','.join(parts)

def build_main_diag(all_sheets: dict) -> pd.DataFrame:
    """
    從所有月份資料找每位病患的最後就診日及主次診斷
    輸出：姓名 | 身份證號 | 最後就診日(日期)\n(以最新的日期為主) | 診斷代碼(病1,病23)
    """
    print("\n[處理] 建立「主次診斷」工作表...")
    all_records = []

    for sheet_name, df in all_sheets.items():
        hdr_row_idx = find_header_row(df)
        header_row  = df.iloc[hdr_row_idx]
        data        = df.iloc[hdr_row_idx + 1:].copy()

        col_id   = get_col_by_keys(header_row, [KW_ID, KW_ID_ALT])
        col_name = get_col(header_row, KW_NAME)
        col_date = get_col_by_keys(header_row, KW_DATE_KEYS)
        col_dx1  = get_col(header_row, KW_DX1)
        col_dx23 = get_col(header_row, KW_DX23)

        if any(c == -1 for c in [col_id, col_name, col_date, col_dx1]):
            continue

        data = data[data[col_id].notna()].copy()
        data[col_id] = data[col_id].apply(clean_id)
        data = data[data[col_id].apply(is_valid_id)]
        if data.empty:
            continue

        cols = [col_name, col_id, col_date, col_dx1]
        names = ['姓名', '身份證號', '日期', '病1']
        if col_dx23 != -1:
            cols.append(col_dx23)
            names.append('病23')

        records = data[cols].copy()
        records.columns = names
        if '病23' not in records.columns:
            records['病23'] = ''
        all_records.append(records)

    if not all_records:
        return pd.DataFrame(columns=['姓名', '身份證號',
                                     '最後就診日(日期)\n(以最新的日期為主)',
                                     '診斷代碼(病1,病23)'])

    all_df = pd.concat(all_records, ignore_index=True)
    all_df['日期'] = pd.to_numeric(all_df['日期'], errors='coerce')
    all_df = all_df.dropna(subset=['日期'])

    # 取每人最新就診那一筆
    latest = (
        all_df.sort_values('日期', ascending=False)
              .drop_duplicates('身份證號')
              .reset_index(drop=True)
    )

    # 合併病1 + 病23
    latest['診斷代碼(病1,病23)'] = latest.apply(
        lambda r: merge_dx(r['病1'], r['病23']), axis=1
    )
    latest['日期'] = latest['日期'].astype(int)

    result = latest[['姓名', '身份證號', '日期', '診斷代碼(病1,病23)']].rename(
        columns={'日期': '最後就診日(日期)\n(以最新的日期為主)'}
    )

    # 依身份證號倒序排列（符合原始輸出格式）
    result = result.sort_values('身份證號', ascending=False).reset_index(drop=True)
    print(f"  → 主次診斷共 {len(result)} 位病患")
    return result

# ──────────────────────────────────────────────
def build_phone(all_sheets: dict) -> pd.DataFrame:
    """
    從所有月份資料整理每位病患的電話號碼
    規則：
      1. 優先保留 09 開頭的行動電話
      2. 同優先層級下，以較新日期的那筆為準
    輸出：ID | 電話
    """
    print("\n[處理] 建立「phone」工作表...")
    all_records = []

    for sheet_name, df in all_sheets.items():
        hdr_row_idx = find_header_row(df)
        header_row  = df.iloc[hdr_row_idx]
        data        = df.iloc[hdr_row_idx + 1:].copy()

        col_id    = get_col_by_keys(header_row, [KW_ID, KW_ID_ALT])
        col_name  = get_col(header_row, KW_NAME)
        col_phone = get_col_by_keys(header_row, KW_PHONE_KEYS)
        col_date  = get_col_by_keys(header_row, KW_DATE_KEYS)

        if any(c == -1 for c in [col_id, col_phone, col_date]):
            continue

        data = data[data[col_id].notna()].copy()
        data[col_id] = data[col_id].apply(clean_id)
        data = data[data[col_id].apply(is_valid_id)]
        if data.empty:
            continue

        cols  = [col_id, col_phone, col_date]
        names = ['ID', '電話', '日期']
        if col_name != -1:
            cols.append(col_name)
            names.append('姓名')

        records = data[cols].copy()
        records.columns = names
        if '姓名' not in records.columns:
            records['姓名'] = ''
        all_records.append(records)

    if not all_records:
        print("  → 無電話資料")
        return pd.DataFrame(columns=['姓名', 'ID', '電話'])

    all_df = pd.concat(all_records, ignore_index=True)

    # 電話正規化
    all_df['電話'] = all_df['電話'].apply(normalize_phone)
    all_df = all_df[all_df['電話'] != '']

    # 日期轉數字（越大越新）
    all_df['日期'] = pd.to_numeric(all_df['日期'], errors='coerce').fillna(0)

    # 標記是否為行動電話（09開頭）
    all_df['是行動電話'] = all_df['電話'].str.startswith('09').astype(int)

    # 排序：行動電話優先 → 日期越新越優先
    all_df = all_df.sort_values(
        ['ID', '是行動電話', '日期'],
        ascending=[True, False, False]
    )

    result = (
        all_df.drop_duplicates('ID', keep='first')
              .reset_index(drop=True)[['姓名', 'ID', '電話']]
    )

    print(f"  → phone 共 {len(result)} 位病患")
    return result

# ──────────────────────────────────────────────
def apply_header_style(ws):
    """套用標題列樣式（粗體、置中）"""
    font = Font(bold=True, name='Arial')
    for cell in ws[1]:
        cell.font      = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def write_sheet(ws, df: pd.DataFrame, header: list):
    """寫入標題與資料，並自動調整欄寬"""
    ws.append(header)
    for row in df.itertuples(index=False):
        ws.append(list(row))
    apply_header_style(ws)
    for col_idx, _ in enumerate(header, 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(header[col_idx - 1])),
            *(len(str(ws.cell(r, col_idx).value or '')) for r in range(2, ws.max_row + 1))
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

# ──────────────────────────────────────────────
def pick_input_file():
    """彈出檔案選擇視窗，回傳 (path, root)；取消則結束程式"""
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    path = filedialog.askopenfilename(
        title='請選擇門診診療次數月報表',
        filetypes=[('Excel 檔案', '*.xlsx *.xls *.XLSX *.XLS'), ('所有檔案', '*.*')]
    )
    if not path:
        print("❌ 未選擇檔案，程式結束。")
        root.destroy()
        sys.exit(0)
    return path, root

# ──────────────────────────────────────────────
def main():
    input_path, root = pick_input_file()

    base_name   = os.path.basename(input_path)
    name_no_ext = os.path.splitext(base_name)[0]
    prefix6     = name_no_ext[:6]
    now_str     = datetime.datetime.now().strftime('%m%d%H%M')
    output_filename = f"{prefix6}_撈出_{now_str}.xlsx"
    output_dir  = os.path.dirname(input_path)
    output_path = os.path.join(output_dir, output_filename)
    print("=" * 55)
    print(f"輸入檔案：{base_name}")
    print(f"輸出目錄：{output_dir}")
    print(f"輸出檔案：{output_filename}")
    print("=" * 55)

    all_sheets = load_input(input_path)

    wb = Workbook()
    wb.remove(wb.active)

    monthly_header = ['身份證號', '姓名', '件數', '申請金額']
    diag_header    = ['姓名', '身份證號',
                      '最後就診日(日期)\n(以最新的日期為主)',
                      '診斷代碼(病1,病23)']
    phone_header   = ['姓名', 'ID', '電話']

    print("\n[處理] 各月份工作表...")
    for sheet_name, df in all_sheets.items():
        result = process_monthly(df, sheet_name)
        ws = wb.create_sheet(sheet_name)
        write_sheet(ws, result, monthly_header)

    # 空工作表（11503, 11504），放在月份工作表之後、主次診斷之前
    existing = set(all_sheets.keys())
    for extra in ['11503', '11504']:
        if extra not in existing:
            ws_extra = wb.create_sheet(extra)
            ws_extra.append(monthly_header)
            apply_header_style(ws_extra)
            print(f"  [建立] 空工作表 {extra}")

    # 主次診斷
    diag_df = build_main_diag(all_sheets)
    ws_diag = wb.create_sheet('主次診斷')
    write_sheet(ws_diag, diag_df, diag_header)

    # phone（最後一個工作表）
    phone_df = build_phone(all_sheets)
    ws_phone = wb.create_sheet('phone')
    write_sheet(ws_phone, phone_df, phone_header)

    wb.save(output_path)
    print("\n" + "=" * 55)
    print(f"✅ 輸出完成：{output_path}")
    print(f"   工作表數量：{len(wb.sheetnames)}")
    print(f"   工作表清單：{wb.sheetnames}")
    print("=" * 55)

    # 完成後彈出視窗，點確定則開啟檔案
    import subprocess, platform
    from tkinter import messagebox
    answer = messagebox.askokcancel(
        '完成',
        f'輸出完成！\n\n{output_filename}\n\n點選「確定」開啟檔案。',
        parent=root
    )
    root.destroy()
    if answer:
        system = platform.system()
        if system == 'Windows':
            os.startfile(output_path)
        elif system == 'Darwin':
            subprocess.call(['open', output_path])
        else:
            subprocess.call(['xdg-open', output_path])

if __name__ == '__main__':
    main()
