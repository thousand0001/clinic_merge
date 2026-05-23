#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import sys
import tkinter as tk
import zipfile
from dataclasses import dataclass
from pathlib import Path
from tkinter import filedialog
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

SOURCE_ENCODING = "cp950"
MONTHLY_HEADER = ["身份證號", "姓名", "件數", "申請金額"]
DIAG_HEADER = ["姓名", "身份證號", "最後就診日(日期)\n(以最新的日期為主)", "診斷代碼(病1,病23)"]
PHONE_SHEET_NAME = "行動電話"
PHONE_HEADER = ["姓名", "ID", "電話", "地址"]
OUTPUT_FOLDER_NAME = "醫聖月份費用xlsx"


@dataclass
class SourceItem:
    source_label: str
    output_stem: str
    month: str
    records: list[dict]


@dataclass
class CollectResult:
    items: list[SourceItem]
    output_dir: Path
    skipped: list[str]


def normalize(value: str) -> str:
    return str(value).strip().replace("\u3000", "")


def parse_int(value: str) -> int:
    digits = normalize(value).replace(",", "")
    if not digits:
        return 0
    return int(float(digits))


def looks_like_date_token(value: str) -> bool:
    text = normalize(value)
    return bool(re.fullmatch(r"\d{2,3}\.\d{2}\.\d{2}", text))


def looks_like_id_token(value: str) -> bool:
    text = normalize(value).upper()
    return bool(re.fullmatch(r"[A-Z]{1,2}\d{8,9}", text))


def split_dx_fields(fields: list[str]) -> tuple[list[str], list[str]]:
    cleaned = [normalize(x) for x in fields]
    while cleaned and not cleaned[-1]:
        cleaned.pop()
    half = len(cleaned) // 2
    return cleaned[:half], cleaned[half:]


def parse_comma_records(lines: list[str]) -> list[dict]:
    records: list[dict] = []
    for line in lines[2:]:
        parts = [part.strip() for part in line.split(",")]
        while parts and not normalize(parts[-1]):
            parts.pop()
        if len(parts) < 12:
            continue
        id_idx = next((idx for idx, part in enumerate(parts) if looks_like_id_token(part)), None)
        if id_idx is None or id_idx < 5:
            continue
        chart_no = normalize(parts[id_idx - 5])
        name = normalize(parts[id_idx - 4])
        visit_date = normalize(parts[id_idx - 3])
        bday = normalize(parts[id_idx - 2])
        if not chart_no or not name or not looks_like_date_token(visit_date) or not looks_like_date_token(bday):
            continue

        last_visit_idx = next(
            (idx for idx in range(id_idx + 1, len(parts)) if looks_like_date_token(parts[idx])),
            None,
        )
        if last_visit_idx is None:
            continue

        icd10_list = [
            normalize(code)
            for code in parts[id_idx + 1:last_visit_idx]
            if normalize(code)
        ]
        if len(parts) < 3:
            continue
        apply_amount = parse_int(parts[-3])

        record = {
            "病歷號": chart_no,
            "看診日": visit_date,
            "姓名": name,
            "身份證號": normalize(parts[id_idx]),
            "生日": bday,
            "住址": "",
            "電話": "",
            "天數": 1,
            "申請金額": apply_amount,
            "最後就診日": normalize(parts[last_visit_idx]).replace(".", ""),
            "診斷代碼": ",".join(code for code in icd10_list if code),
            "件數": 1,
        }
        if record["身份證號"] and record["姓名"]:
            records.append(record)
    return records


def parse_fixed_width_records(lines: list[str]) -> list[dict]:
    records: list[dict] = []
    row_re = re.compile(r"^(\d{6})\s+(1\d{2}\.\d{2}\.\d{2})\s+(.+?)\s{2,}([A-Z0-9]{3})\s{2,}(.*)$")
    id_bday_re = re.compile(r"([A-Z]{1,2}\d{8,9})\s+(\d{2,3}\.\d{2}\.\d{2})\s+")
    fee_re = re.compile(
        r"^(.*?)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d{2,3}\.\d{2}\.\d{2})\s*$"
    )
    phone_block_re = re.compile(r"^\s*(\S+)(.*?)\s{2,}(\d{7,10})\s{2,}(\d*)\s{2,}(.*)$")

    for line in lines[2:]:
        if not re.match(r"^\d{6}\s+1\d{2}\.\d{2}\.\d{2}\s+", line):
            continue
        if set(line.strip()) == {"="}:
            continue

        row_match = row_re.match(line)
        if not row_match:
            continue
        chart_no, visit_date, name, _, tail = row_match.groups()

        id_match = id_bday_re.search(tail)
        if not id_match:
            continue
        id_no, bday = id_match.groups()
        after_bday = tail[id_match.end():]

        phone_match = phone_block_re.match(after_bday)
        if not phone_match:
            continue
        _, address, phone, _, fee_tail = phone_match.groups()

        fee_match = fee_re.match(fee_tail)
        if not fee_match:
            continue
        dx_part = fee_match.group(1)
        apply_amount = parse_int(fee_match.group(10))
        last_visit = normalize(fee_match.group(11)).replace(".", "")

        dx_codes = re.findall(r"\b[A-Z][0-9A-Z]{2,7}\b", dx_part)
        record = {
            "病歷號": normalize(chart_no),
            "看診日": normalize(visit_date),
            "姓名": normalize(name),
            "身份證號": normalize(id_no),
            "生日": normalize(bday),
            "住址": normalize(address),
            "電話": normalize(phone),
            "天數": parse_int(fee_match.group(2)),
            "申請金額": apply_amount,
            "最後就診日": last_visit,
            "診斷代碼": ",".join(dx_codes),
            "件數": 1,
        }
        if record["身份證號"] and record["姓名"]:
            records.append(record)

    return records


def looks_like_multiline_comma_format(lines: list[str]) -> bool:
    if len(lines) < 2:
        return False
    header = normalize(lines[1])
    return all(token in header for token in ("病歷號", "姓名", "身分證", "看診日", "自費", "主訴"))


def parse_multiline_comma_records(lines: list[str]) -> list[dict]:
    records: list[dict] = []
    current: dict | None = None

    for line in lines[2:]:
        parts = [part.strip() for part in line.split(",")]
        if len(parts) >= 5 and normalize(parts[0]).isdigit() and normalize(parts[2]):
            if current:
                records.append(current)

            current = {
                "病歷號": normalize(parts[0]),
                "看診日": normalize(parts[3]),
                "姓名": normalize(parts[1]),
                "身份證號": normalize(parts[2]),
                "生日": "",
                "住址": "",
                "電話": "",
                "天數": 1,
                "申請金額": parse_int(parts[4]),
                "最後就診日": normalize(parts[3]).replace(".", ""),
                "診斷代碼": "",
                "件數": 1,
            }
            continue

        if current is None:
            continue

        payload = ",".join(parts[5:]).strip() if len(parts) > 5 else ""
        if "\x19" in payload:
            match = re.search(r"\x19(\d{7})/\d{2}:\d{2}.*?(?:\x1a(.*?))?\s*$", payload)
            if match:
                current["最後就診日"] = match.group(1)
                current["診斷代碼"] = normalize((match.group(2) or "").rstrip(", "))

    if current:
        records.append(current)

    return [
        record
        for record in records
        if record["身份證號"] and record["姓名"]
    ]


def parse_txt_bytes(data: bytes, source_name: str) -> tuple[str, list[dict]]:
    text = data.decode(SOURCE_ENCODING, errors="replace")
    lines = [line.rstrip("\r\n") for line in text.splitlines() if line.strip()]
    if len(lines) < 3:
        raise ValueError(f"{source_name}: 資料列不足，無法轉換")

    title = normalize(lines[0])
    month = title.split(":", 1)[-1] if ":" in title else Path(source_name).stem.replace("-", "")
    month = month.strip()
    records = parse_comma_records(lines)
    if not records:
        records = parse_fixed_width_records(lines)
    if not records and looks_like_multiline_comma_format(lines):
        records = parse_multiline_comma_records(lines)

    if not records:
        raise ValueError(f"{source_name}: 沒有成功解析出任何資料列")

    return month, records


def parse_txt_file(path: Path) -> tuple[str, list[dict]]:
    return parse_txt_bytes(path.read_bytes(), str(path))


def autosize(ws) -> None:
    for idx in range(1, ws.max_column + 1):
        letter = get_column_letter(idx)
        width = max(
            len(str(ws.cell(row=row, column=idx).value or ""))
            for row in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[letter].width = min(width + 4, 40)


def style_header(ws) -> None:
    font = Font(bold=True, name="Arial")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.font = font
        cell.alignment = align


def append_rows(ws, header: list[str], rows: Iterable[Iterable]) -> None:
    ws.append(header)
    for row in rows:
        ws.append(list(row))
    style_header(ws)
    autosize(ws)


def normalize_month_sheet_name(month: str) -> str:
    text = normalize(month)
    match = re.match(r"^(1(?:14|15))\.(\d{2})\.\d{2}～\1\.\2\.\d{2}$", text)
    if match:
        return f"{match.group(1)}{match.group(2)}"
    match = re.match(r"^(1(?:14|15))(\d{2})$", text)
    if match:
        return text
    return text


def build_workbook(month: str, records: list[dict], output_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    ws_month = wb.create_sheet(normalize_month_sheet_name(month))
    month_map: dict[tuple[str, str], dict] = {}
    for record in records:
        key = (record["身份證號"], record["姓名"])
        row = month_map.setdefault(key, {"件數": 0, "申請金額": 0})
        row["件數"] += record["件數"]
        row["申請金額"] += record["申請金額"]
    month_rows = (
        (id_no, name, values["件數"], values["申請金額"])
        for (id_no, name), values in month_map.items()
    )
    append_rows(ws_month, MONTHLY_HEADER, month_rows)

    ws_diag = wb.create_sheet("主次診斷")
    diag_map: dict[str, dict] = {}
    for record in records:
        current = diag_map.get(record["身份證號"])
        if current is None or record["最後就診日"] >= current["最後就診日"]:
            diag_map[record["身份證號"]] = record
    diag_rows = (
        (r["姓名"], r["身份證號"], r["最後就診日"], r["診斷代碼"])
        for r in diag_map.values()
    )
    append_rows(ws_diag, DIAG_HEADER, diag_rows)

    ws_phone = wb.create_sheet(PHONE_SHEET_NAME)
    phone_map: dict[str, dict] = {}
    for record in records:
        current = phone_map.get(record["身份證號"])
        phone = record["電話"]
        is_mobile = phone.startswith("09")
        if current is None:
            phone_map[record["身份證號"]] = record
            continue
        current_mobile = current["電話"].startswith("09")
        if (is_mobile and not current_mobile) or (
            is_mobile == current_mobile and record["最後就診日"] >= current["最後就診日"]
        ):
            phone_map[record["身份證號"]] = record
    phone_rows = (
        (r["姓名"], r["身份證號"], r["電話"], r["住址"])
        for r in phone_map.values()
    )
    append_rows(ws_phone, PHONE_HEADER, phone_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def slugify_stem(name: str) -> str:
    stem = re.sub(r"[\\/:*?\"<>|]+", "_", name).strip()
    return stem or "output"


def choose_input_path() -> Path | None:
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    selected_dir = filedialog.askdirectory(
        title="選擇要批次轉換的資料夾"
    )
    if selected_dir:
        root.destroy()
        return Path(selected_dir).resolve()

    selected = filedialog.askopenfilename(
        title="或選擇單一 TXT / ZIP 檔案",
        filetypes=[("TXT/ZIP", "*.txt *.zip"), ("All files", "*.*")],
    )
    root.destroy()
    return Path(selected).resolve() if selected else None


def iter_folder_sources(folder: Path) -> tuple[list[SourceItem], list[str]]:
    items: list[SourceItem] = []
    skipped: list[str] = []

    for txt_path in sorted(folder.rglob("*.txt")):
        try:
            month, records = parse_txt_file(txt_path)
            items.append(
                SourceItem(
                    source_label=str(txt_path),
                    output_stem=slugify_stem(txt_path.stem),
                    month=month,
                    records=records,
                )
            )
        except Exception as exc:
            skipped.append(f"{txt_path} | {exc}")

    for zip_path in sorted(folder.rglob("*.zip")):
        zip_items, zip_skipped = iter_zip_sources(zip_path)
        items.extend(zip_items)
        skipped.extend(zip_skipped)

    return items, skipped


def iter_zip_sources(zip_path: Path) -> tuple[list[SourceItem], list[str]]:
    items: list[SourceItem] = []
    skipped: list[str] = []
    with zipfile.ZipFile(zip_path) as zf:
        for member in sorted(zf.namelist()):
            if member.endswith("/") or not member.lower().endswith(".txt"):
                continue
            source_label = f"{zip_path}!{member}"
            try:
                month, records = parse_txt_bytes(zf.read(member), source_label)
                inner_stem = slugify_stem(Path(member).stem)
                output_stem = slugify_stem(f"{zip_path.stem}_{inner_stem}")
                items.append(
                    SourceItem(
                        source_label=source_label,
                        output_stem=output_stem,
                        month=month,
                        records=records,
                    )
                )
            except Exception as exc:
                skipped.append(f"{source_label} | {exc}")
    return items, skipped


def collect_sources(input_path: Path) -> CollectResult:
    if input_path.is_dir():
        output_dir = input_path.parent / OUTPUT_FOLDER_NAME
        items, skipped = iter_folder_sources(input_path)
        return CollectResult(items=items, output_dir=output_dir, skipped=skipped)

    suffix = input_path.suffix.lower()
    if suffix == ".txt":
        month, records = parse_txt_file(input_path)
        output_dir = input_path.parent / OUTPUT_FOLDER_NAME
        return CollectResult(
            items=[
                SourceItem(
                    source_label=str(input_path),
                    output_stem=slugify_stem(input_path.stem),
                    month=month,
                    records=records,
                )
            ],
            output_dir=output_dir,
            skipped=[],
        )

    if suffix == ".zip":
        output_dir = input_path.parent / OUTPUT_FOLDER_NAME
        items, skipped = iter_zip_sources(input_path)
        return CollectResult(items=items, output_dir=output_dir, skipped=skipped)

    raise ValueError(f"不支援的輸入類型：{input_path}")


def resolve_output_path(output_dir: Path, item: SourceItem, used_names: set[str]) -> Path:
    preferred = slugify_stem(normalize_month_sheet_name(item.month)) or item.output_stem
    candidate = preferred
    index = 2
    while candidate.lower() in used_names:
        candidate = f"{preferred}_{index}"
        index += 1
    used_names.add(candidate.lower())
    return output_dir / f"{candidate}.xlsx"


def main() -> int:
    parser = argparse.ArgumentParser(description="將醫聖費用 TXT/ZIP 批次轉成 Excel 費用表單")
    parser.add_argument("input_path", nargs="?", type=Path, help="來源資料夾、TXT 或 ZIP 路徑")
    args = parser.parse_args()

    input_path = args.input_path.resolve() if args.input_path else choose_input_path()
    if input_path is None:
        print("未選擇任何輸入來源，程式結束。")
        return 1
    if not input_path.exists():
        raise FileNotFoundError(f"找不到輸入路徑：{input_path}")

    result = collect_sources(input_path)
    items = result.items
    output_dir = result.output_dir
    if not items:
        print(f"找不到可轉換的 TXT 資料：{input_path}")
        if result.skipped:
            print("以下檔案已略過：")
            for msg in result.skipped:
                print(f"  - {msg}")
        return 1

    output_dir.mkdir(parents=True, exist_ok=True)
    used_names: set[str] = set()
    converted = 0

    print(f"輸入來源：{input_path}")
    print(f"輸出資料夾：{output_dir}")
    for item in items:
        output_path = resolve_output_path(output_dir, item, used_names)
        build_workbook(item.month, item.records, output_path)
        converted += 1
        print(f"[OK] {item.source_label}")
        print(f"     -> {output_path}")

    print(f"完成，共轉出 {converted} 個費用表單。")
    if result.skipped:
        print(f"略過 {len(result.skipped)} 個無法解析的檔案：")
        for msg in result.skipped:
            print(f"  - {msg}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
