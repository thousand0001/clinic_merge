from __future__ import annotations

import sys
from pathlib import Path

import openpyxl


def main() -> int:
    if len(sys.argv) != 3:
        print("usage: compare_xlsx_values.py <old.xlsx> <new.xlsx>")
        return 2

    p1 = Path(sys.argv[1])
    p2 = Path(sys.argv[2])
    wb1 = openpyxl.load_workbook(p1, data_only=False)
    wb2 = openpyxl.load_workbook(p2, data_only=False)

    all_sheets = list(dict.fromkeys(wb1.sheetnames + wb2.sheetnames))
    diffs = []
    for sheet_name in all_sheets:
        if sheet_name not in wb1.sheetnames or sheet_name not in wb2.sheetnames:
            diffs.append((sheet_name, "SHEET_MISSING", sheet_name in wb1.sheetnames, sheet_name in wb2.sheetnames))
            continue
        ws1 = wb1[sheet_name]
        ws2 = wb2[sheet_name]
        max_row = max(ws1.max_row, ws2.max_row)
        max_col = max(ws1.max_column, ws2.max_column)
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                v1 = ws1.cell(row, col).value
                v2 = ws2.cell(row, col).value
                if v1 != v2:
                    diffs.append((sheet_name, row, col, v1, v2))

    print(f"DIFF_COUNT {len(diffs)}")
    for item in diffs[:200]:
        print(repr(item))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
