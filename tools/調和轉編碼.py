#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import csv
import os
import re
import shutil
import tkinter as tk
from pathlib import Path
from tkinter import filedialog
from typing import Any

import xlrd
from openpyxl import Workbook, load_workbook


DEFAULT_FILE_SUFFIX = "_修復xlsx"
DEFAULT_FOLDER_SUFFIX = "_轉檔OK"
CSV_TARGET_ENCODING = "utf-8-sig"
CSV_ENCODING_CANDIDATES = (
    "utf-16",
    "utf-16le",
    "utf-8-sig",
    "utf-8",
    "cp950",
    "big5",
    "gbk",
    "gb18030",
)
REPAIRABLE_WORKBOOK_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}
MONTH_SHEET_RE = re.compile(r"^1(?:14|15)\d{2}\.xlsx$", re.IGNORECASE)
COMMON_TEXT_TOKENS = (
    "姓名",
    "身份證",
    "身分證",
    "電話",
    "手機",
    "地址",
    "列印",
    "日期",
    "診所",
    "條件",
    "性別",
    "生日",
    "號碼",
    "來診日",
    "掛號",
    "巷",
    "路",
    "街",
    "段",
    "號",
    "樓",
    "區",
)


def _try_redecode(text: str, src_encoding: str, dst_encoding: str) -> str | None:
    try:
        return text.encode(src_encoding).decode(dst_encoding)
    except Exception:
        return None


def _text_quality_score(text: str) -> int:
    score = 0
    for ch in text:
        code = ord(ch)
        if "\u4e00" <= ch <= "\u9fff":
            score += 3
        elif ch.isascii() and (ch.isalnum() or ch in ",./:-_ ()[]"):
            score += 1
        elif 0xE000 <= code <= 0xF8FF:
            score -= 8
        elif 0x2500 <= code <= 0x257F:
            score -= 5
        elif 0x3040 <= code <= 0x30FF:
            score -= 4
        elif 0x3100 <= code <= 0x312F:
            score -= 4
        elif ch in "﹎﹏┮┦┑⌒":
            score -= 6

    for token in COMMON_TEXT_TOKENS:
        score += text.count(token) * 12
    return score


def repair_text(value: Any) -> Any:
    """優先挑選品質較好的修復結果，兼容舊式 cp950/Big5 與 UTF-8 亂碼。"""
    if not isinstance(value, str) or not value:
        return value

    candidates = [value]
    for src_encoding, dst_encoding in (
        ("latin1", "cp950"),
        ("gb18030", "cp950"),
        ("gbk", "cp950"),
    ):
        repaired = _try_redecode(value, src_encoding, dst_encoding)
        if repaired and repaired not in candidates:
            candidates.append(repaired)

    best = max(candidates, key=_text_quality_score)
    return best


def copy_cell_value(cell: xlrd.sheet.Cell) -> Any:
    value = cell.value
    if cell.ctype == xlrd.XL_CELL_TEXT:
        return repair_text(value)
    return value


def detect_text_encoding(file_path: Path) -> str:
    raw = file_path.read_bytes()
    if raw.startswith(b"\xff\xfe") or raw.startswith(b"\xfe\xff"):
        return "utf-16"
    if raw.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"

    last_error: Exception | None = None
    for encoding in CSV_ENCODING_CANDIDATES:
        try:
            sample = raw.decode(encoding)
            if "\ufffd" in sample:
                continue
            return encoding
        except UnicodeDecodeError as exc:
            last_error = exc
    raise RuntimeError(f"無法判斷編碼：{file_path.name} ({last_error})") from last_error


def convert_xls_to_xlsx(src_path: Path, dst_path: Path) -> None:
    book = xlrd.open_workbook(str(src_path))
    wb = Workbook()
    wb.remove(wb.active)

    for idx, sheet in enumerate(book.sheets(), start=1):
        title = repair_text(str(sheet.name or f"Sheet{idx}"))[:31] or f"Sheet{idx}"
        ws = wb.create_sheet(title=title)
        for row_idx in range(sheet.nrows):
            row_values = [copy_cell_value(sheet.cell(row_idx, col_idx)) for col_idx in range(sheet.ncols)]
            ws.append(row_values)

    dst_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dst_path)


def repair_workbook_file(src_path: Path, dst_path: Path) -> None:
    keep_vba = src_path.suffix.lower() == ".xlsm"
    wb = load_workbook(str(src_path), keep_vba=keep_vba)

    for ws in wb.worksheets:
        repaired_title = repair_text(ws.title)
        if repaired_title != ws.title:
            ws.title = repaired_title[:31] or ws.title

        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    fixed = repair_text(cell.value)
                    if fixed != cell.value:
                        cell.value = fixed

    dst_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dst_path)


def repair_csv_file(src_path: Path, dst_path: Path) -> None:
    raw = src_path.read_bytes()
    repaired_parts: list[str] = []

    for raw_line in raw.splitlines(keepends=True):
        line_candidates: list[str] = []
        for encoding in CSV_ENCODING_CANDIDATES:
            try:
                decoded = raw_line.decode(encoding, errors="replace")
            except Exception:
                continue
            if decoded not in line_candidates:
                line_candidates.append(decoded)
            repaired = repair_text(decoded)
            if repaired not in line_candidates:
                line_candidates.append(repaired)

        if line_candidates:
            repaired_parts.append(max(line_candidates, key=_text_quality_score))
        else:
            repaired_parts.append(raw_line.decode("latin1", errors="replace"))

    repaired = "".join(repaired_parts)

    dst_path.parent.mkdir(parents=True, exist_ok=True)
    dst_path.write_text(repaired, encoding=CSV_TARGET_ENCODING, newline="")


def iter_files(root: Path) -> list[Path]:
    files: list[Path] = []
    for dirpath, dirnames, filenames in os.walk(root):
        dirnames.sort()
        for name in sorted(filenames):
            files.append(Path(dirpath) / name)
    return files


def normalized_xlsx_name(src_path: Path) -> str:
    """
    只在 .xls 轉檔時標準化月份檔名：
    - 11401.xls -> 11401.xlsx
    - 11503PAITEM.xls -> 11503.xlsx
    其他檔名則維持原 stem。
    """
    stem = src_path.stem
    match = re.match(r"^(1(?:14|15)\d{2})", stem)
    if match:
        return f"{match.group(1)}.xlsx"
    return f"{stem}.xlsx"


def _copy_worksheet_values(src_ws, dst_ws) -> None:
    for row in src_ws.iter_rows(values_only=True):
        dst_ws.append(list(row))


def build_monthly_sheet_workbook(folder: Path) -> Path | None:
    month_files = sorted(
        p for p in folder.iterdir()
        if p.is_file() and MONTH_SHEET_RE.fullmatch(p.name)
    )
    if not month_files:
        return None

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    for month_file in month_files:
        wb_src = load_workbook(month_file, data_only=False)
        src_ws = wb_src[wb_src.sheetnames[0]]
        dst_ws = wb_out.create_sheet(title=month_file.stem)
        _copy_worksheet_values(src_ws, dst_ws)

    out_path = folder.with_suffix(".xlsx")
    wb_out.save(out_path)

    for month_file in month_files:
        try:
            month_file.unlink()
        except OSError:
            pass

    try:
        folder.rmdir()
    except OSError:
        pass

    return out_path


def build_monthly_sheet_workbooks(dst_root: Path) -> list[Path]:
    created: list[Path] = []
    for dirpath, dirnames, _filenames in os.walk(dst_root):
        dirnames.sort()
        folder = Path(dirpath)
        out_path = build_monthly_sheet_workbook(folder)
        if out_path is not None:
            created.append(out_path)
    return created


def convert_folder(src_root: Path, dst_root: Path) -> tuple[list[Path], list[tuple[Path, str]]]:
    converted: list[Path] = []
    skipped: list[tuple[Path, str]] = []

    for src_path in iter_files(src_root):
        rel = src_path.relative_to(src_root)
        suffix = src_path.suffix.lower()
        if suffix == ".xls":
            dst_path = (dst_root / rel).with_name(normalized_xlsx_name(src_path))
            try:
                convert_xls_to_xlsx(src_path, dst_path)
                converted.append(dst_path)
            except Exception as exc:
                skipped.append((src_path, str(exc)))
            continue

        dst_path = dst_root / rel
        try:
            if suffix in REPAIRABLE_WORKBOOK_SUFFIXES:
                repair_workbook_file(src_path, dst_path)
                converted.append(dst_path)
            elif suffix == ".csv":
                repair_csv_file(src_path, dst_path)
                converted.append(dst_path)
            else:
                dst_path.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(src_path, dst_path)
        except Exception as exc:
            skipped.append((src_path, str(exc)))

    monthly_books = build_monthly_sheet_workbooks(dst_root)
    if monthly_books:
        converted = [
            path for path in converted
            if not (path.parent.name == "R11440次數" and MONTH_SHEET_RE.fullmatch(path.name))
        ]
        converted.extend(monthly_books)
    return converted, skipped


def build_output_path(src_path: Path, output: str | None) -> Path:
    if output:
        return Path(output).expanduser().resolve()
    if src_path.is_file():
        return src_path.with_name(f"{src_path.stem}{DEFAULT_FILE_SUFFIX}.xlsx")
    return src_path.parent / f"{src_path.name}{DEFAULT_FOLDER_SUFFIX}"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="修復資料夾內 Excel/CSV 中文亂碼，並將 .xls 轉成 .xlsx。")
    parser.add_argument("source", nargs="?", help="來源檔案或資料夾")
    parser.add_argument("-o", "--output", help="輸出檔案或資料夾；未填則自動產生")
    return parser.parse_args()


def choose_source_path() -> Path | None:
    root = tk.Tk()
    root.withdraw()
    root.update()
    selected = filedialog.askdirectory(title="選擇要修復的資料夾")
    if not selected:
        selected = filedialog.askopenfilename(
            title="或選擇單一 XLS 檔案",
            filetypes=[("Excel 97-2003", "*.xls"), ("All files", "*.*")],
        )
    root.destroy()
    if not selected:
        return None
    return Path(selected).expanduser().resolve()


def main() -> int:
    args = parse_args()
    if args.source:
        src_path = Path(args.source).expanduser().resolve()
    else:
        chosen = choose_source_path()
        if chosen is None:
            print("未選擇來源，已取消。")
            return 1
        src_path = chosen

    if not src_path.exists():
        print(f"找不到來源：{src_path}")
        return 1

    dst_path = build_output_path(src_path, args.output)

    if src_path.is_file():
        suffix = src_path.suffix.lower()
        if suffix == ".xls":
            if dst_path.suffix.lower() != ".xlsx":
                dst_path = dst_path.with_name(normalized_xlsx_name(src_path))
            convert_xls_to_xlsx(src_path, dst_path)
        elif suffix in REPAIRABLE_WORKBOOK_SUFFIXES:
            if dst_path.suffix.lower() != suffix:
                dst_path = dst_path.with_suffix(suffix)
            repair_workbook_file(src_path, dst_path)
        elif suffix == ".csv":
            if dst_path.suffix.lower() != ".csv":
                dst_path = dst_path.with_suffix(".csv")
            repair_csv_file(src_path, dst_path)
        else:
            print(f"目前不支援單檔處理此類型：{src_path}")
            return 1
        print(f"已處理：{dst_path}")
        return 0

    converted, skipped = convert_folder(src_path, dst_path)
    print(f"輸出資料夾：{dst_path}")
    print(f"成功處理檔案：{len(converted)}")
    for path in converted:
        print(path)

    if skipped:
        print(f"略過/失敗：{len(skipped)}")
        for path, reason in skipped:
            print(f"{path} | {reason}")
        return 2

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
