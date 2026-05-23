#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
展望雲端專用選會員產檔。

輸出比照「選會員模板」的會員總表，並額外加上：
- 病歷號
- 備註

展望費用檔沒有身份證號碼，因此本程式會：
1. 優先用「病歷號 -> ID」對費用歸戶
2. 病歷號對不到時，用「姓名唯一對到 ID」補歸戶
3. 還是沒有 ID 的資料仍輸出，身份證號碼填入 "-"
"""

from __future__ import annotations

import argparse
import copy
import datetime as dt
import os
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter

import run_merge_通用_0430_1 as generic


TEMPLATE_GLOB = "選會員模板*.xlsx"
TARGET_SHEET = "會員總表"
PERCENTILE_SHEET_NAME = "百分位名單"
DOCTOR_SHEET_NAME = "醫生看(從會員指標內容Key過來)"
SELF_SELECT_SHEET_NAME = "自選名單(從會員指標內容Key過來)"
SCREENING_COLUMNS = {
    "成人健檢": "成人健檢\n最後篩檢日",
    "子宮抹片": "子宮抹片最後篩檢日",
    "老人流感": "老人流感最後注射日",
    "糞便潛血": "糞便潛血最後篩檢日",
    "BC肝炎": "BC肝炎最篩檢日期",
}


class TableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.rows: List[List[str]] = []
        self._row: Optional[List[str]] = None
        self._cell: Optional[List[str]] = None
        self._in_cell = False

    def handle_starttag(self, tag: str, attrs: Sequence[Tuple[str, Optional[str]]]) -> None:
        tag = tag.lower()
        if tag == "tr":
            self._row = []
        elif tag in ("td", "th") and self._row is not None:
            self._cell = []
            self._in_cell = True

    def handle_data(self, data: str) -> None:
        if self._in_cell and self._cell is not None:
            self._cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag in ("td", "th") and self._in_cell:
            text = " ".join("".join(self._cell or []).split())
            if self._row is not None:
                self._row.append(text)
            self._cell = None
            self._in_cell = False
        elif tag == "tr" and self._row is not None:
            if any(str(value).strip() for value in self._row):
                self.rows.append(self._row)
            self._row = None


@dataclass
class MemberRecord:
    name: str = ""
    pid: str = ""
    bday: Optional[dt.date] = None
    charts: Set[str] = field(default_factory=set)
    sex: str = ""
    clinic: str = ""
    ascvd: Any = None
    dmk_raw: Any = None
    health: Dict[str, Any] = field(default_factory=dict)
    p4p: Dict[str, Any] = field(default_factory=dict)
    screenings: Dict[str, Any] = field(default_factory=dict)
    count_114: float = 0.0
    count_114_q1: float = 0.0
    count_115: float = 0.0
    count_115_q1: float = 0.0
    amount_114: float = 0.0
    amount_114_q1: float = 0.0
    amount_115: float = 0.0
    amount_115_q1: float = 0.0
    last_visit: Optional[dt.date] = None
    notes: Set[str] = field(default_factory=set)
    sources: Set[str] = field(default_factory=set)
    is_self_select: bool = False


def normalize_name(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).upper()


def display_text(value: Any) -> str:
    return str(value or "").strip()


def normalize_id(value: Any) -> str:
    return str(value or "").strip().upper().replace(" ", "")


def is_valid_id(value: Any) -> bool:
    text = normalize_id(value)
    return bool(
        re.fullmatch(r"[A-Z][12]\d{8}", text)
        or re.fullmatch(r"[A-Z][A-Z0-9]\d{8}", text)
        or re.fullmatch(r"[A-Z]{1,2}\d{8,10}", text)
    )


def normalize_chart(value: Any) -> str:
    text = str(value or "").strip()
    return text.zfill(7) if text.isdigit() else text


def parse_number(value: Any) -> float:
    try:
        return float(str(value or "0").replace(",", "").strip() or 0)
    except Exception:
        return 0.0


def parse_roc_or_ad_date(value: Any) -> Optional[dt.date]:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = str(value).strip()
    if not text or text == "-":
        return None
    text = text.split(" ")[0].replace("-", "/")
    match = re.fullmatch(r"(\d{2,4})/(\d{1,2})/(\d{1,2})", text)
    if not match:
        return None
    year = int(match.group(1))
    if year < 1911:
        year += 1911
    try:
        return dt.date(year, int(match.group(2)), int(match.group(3)))
    except ValueError:
        return None


def parse_roc_yyyymmdd(value: Any) -> Optional[dt.date]:
    text = str(value or "").strip()
    match = re.fullmatch(r"(1(?:14|15))(\d{2})(\d{2})", text)
    if not match:
        return None
    try:
        return dt.date(int(match.group(1)) + 1911, int(match.group(2)), int(match.group(3)))
    except ValueError:
        return None


def format_date(value: Any) -> Any:
    parsed = parse_roc_or_ad_date(value)
    return parsed if parsed else None


def read_html_table(path: Path) -> List[List[str]]:
    parser = TableParser()
    parser.feed(path.read_text(encoding="utf-8-sig", errors="replace"))
    return parser.rows


def iter_workbook_rows(path: Path) -> Iterable[Tuple[str, List[List[Any]]]]:
    if path.suffix.lower() == ".ods":
        wb = generic._load_ods_as_workbook(str(path))
    else:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    for ws in wb.worksheets:
        rows: List[List[Any]] = []
        for row in ws.iter_rows(values_only=True):
            values = list(row)
            if any(value is not None and str(value).strip() != "" for value in values):
                rows.append(values)
        yield ws.title, rows


def compact_header(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).upper()


def find_col(header: Sequence[Any], aliases: Sequence[str]) -> Optional[int]:
    targets = {compact_header(alias) for alias in aliases}
    for index, value in enumerate(header):
        if compact_header(value) in targets:
            return index
    return None


def find_header(rows: Sequence[Sequence[Any]], required_groups: Sequence[Sequence[str]], scan_rows: int = 30) -> Optional[int]:
    for index, row in enumerate(rows[:scan_rows]):
        if all(find_col(row, group) is not None for group in required_groups):
            return index
    return None


def find_template() -> Path:
    candidates = sorted(Path(__file__).resolve().parent.glob(TEMPLATE_GLOB))
    candidates = [path for path in candidates if path.suffix.lower() == ".xlsx"]
    if not candidates:
        raise FileNotFoundError(f"找不到模板：{TEMPLATE_GLOB}")
    return candidates[-1]


def infer_gender(pid: str) -> str:
    text = normalize_id(pid)
    if len(text) >= 2:
        if text[1] == "1":
            return "男"
        if text[1] == "2":
            return "女"
    return ""


def age_at_today(bday: Optional[dt.date], today: dt.date) -> Optional[int]:
    if not bday:
        return None
    age = today.year - bday.year - ((today.month, today.day) < (bday.month, bday.day))
    return age if age >= 0 else None


def sheet_header_map(ws) -> Dict[str, int]:
    return {compact_header(ws.cell(1, col).value): col for col in range(1, ws.max_column + 1)}


def get_or_create_by_id(records: Dict[str, MemberRecord], pid: str) -> Optional[MemberRecord]:
    pid = normalize_id(pid)
    if not is_valid_id(pid):
        return None
    key = f"ID:{pid}"
    rec = records.setdefault(key, MemberRecord(pid=pid))
    rec.pid = pid
    return rec


def get_or_create_by_name(records: Dict[str, MemberRecord], name: str) -> Optional[MemberRecord]:
    norm = normalize_name(name)
    if not norm:
        return None
    key = f"NAME:{norm}"
    rec = records.setdefault(key, MemberRecord(name=display_text(name)))
    if not rec.name:
        rec.name = display_text(name)
    return rec


def add_identity(
    records: Dict[str, MemberRecord],
    name_to_ids: Dict[str, Set[str]],
    chart_to_ids: Dict[str, Set[str]],
    *,
    name: Any = "",
    pid: Any = "",
    chart: Any = "",
    bday: Any = "",
    source: str = "",
) -> Optional[MemberRecord]:
    text_pid = normalize_id(pid)
    rec = get_or_create_by_id(records, text_pid) if is_valid_id(text_pid) else get_or_create_by_name(records, display_text(name))
    if rec is None:
        return None

    if display_text(name) and not rec.name:
        rec.name = display_text(name)
    if is_valid_id(text_pid):
        rec.pid = text_pid
        if normalize_name(name):
            name_to_ids[normalize_name(name)].add(text_pid)
    parsed_bday = parse_roc_or_ad_date(bday)
    if parsed_bday and not rec.bday:
        rec.bday = parsed_bday
    chart_text = normalize_chart(chart)
    if chart_text:
        rec.charts.add(chart_text)
        if is_valid_id(text_pid):
            chart_to_ids[chart_text].add(text_pid)
    if source:
        rec.sources.add(source)
    return rec


def merge_name_records(records: Dict[str, MemberRecord], name_to_ids: Dict[str, Set[str]]) -> None:
    for key in [key for key in records if key.startswith("NAME:")]:
        rec = records[key]
        ids = name_to_ids.get(normalize_name(rec.name), set())
        if len(ids) != 1:
            continue
        target = get_or_create_by_id(records, next(iter(ids)))
        if target is None or target is rec:
            continue
        if not target.name:
            target.name = rec.name
        if not target.bday:
            target.bday = rec.bday
        target.charts.update(rec.charts)
        target.screenings.update({k: v for k, v in rec.screenings.items() if k not in target.screenings})
        target.health.update({k: v for k, v in rec.health.items() if k not in target.health})
        target.p4p.update({k: v for k, v in rec.p4p.items() if k not in target.p4p})
        target.count_114 += rec.count_114
        target.count_114_q1 += rec.count_114_q1
        target.count_115 += rec.count_115
        target.count_115_q1 += rec.count_115_q1
        target.amount_114 += rec.amount_114
        target.amount_114_q1 += rec.amount_114_q1
        target.amount_115 += rec.amount_115
        target.amount_115_q1 += rec.amount_115_q1
        if rec.last_visit and (target.last_visit is None or rec.last_visit > target.last_visit):
            target.last_visit = rec.last_visit
        target.notes.update(rec.notes)
        target.sources.update(rec.sources)
        del records[key]


def scan_root_files(
    source_dir: Path,
    records: Dict[str, MemberRecord],
    name_to_ids: Dict[str, Set[str]],
    chart_to_ids: Dict[str, Set[str]],
) -> None:
    for path in sorted(source_dir.iterdir()):
        if not path.is_file() or path.name.startswith("~$") or path.suffix.lower() not in (".xlsx", ".ods"):
            continue
        source_is_self_select = is_self_select_source(path.stem)
        for sheet_name, rows in iter_workbook_rows(path):
            sheet_is_self_select = source_is_self_select or is_self_select_source(sheet_name)
            if not rows:
                continue
            header_index = find_header(rows, [["姓名", "會員姓名", "病患姓名"], ["ID", "身分證", "身分證號", "身份證號", "家醫收案會員ID"]])
            if header_index is None:
                continue
            header = rows[header_index]
            name_col = find_col(header, ["姓名", "會員姓名", "病患姓名"])
            id_col = find_col(header, ["ID", "身分證", "身分證號", "身分證號碼", "身份證", "身份證號", "身份證號碼", "家醫收案會員ID"])
            bday_col = find_col(header, ["生日", "BIRTHDAY"])
            indicator_col = find_col(header, ["指標名稱"])
            ascvd_col = find_col(header, ["ASCVD", "ascvd"])
            disease_col = find_col(header, ["疾病樣態"])
            p4p_plan_col = find_col(header, ["P4P收案計畫"])
            p4p_status_col = find_col(header, ["收案狀態"])
            hba_val_col = find_col(header, ["最近一次HbA1c檢查結果(%)"])
            hba_dt_col = find_col(header, ["最近一次HbA1c檢查日期"])
            ldl_val_col = find_col(header, ["最近一次LDL檢查結果(mg/dL)"])
            ldl_dt_col = find_col(header, ["最近一次LDL檢查日期"])
            uacr_val_col = find_col(header, ["最近一次UACR檢查結果(mg/gm)", "最近一次UACR檢查結果"])
            uacr_dt_col = find_col(header, ["最近一次UACR檢查日期"])
            last_screen_col = find_col(header, ["最後篩檢日期"])
            if name_col is None or id_col is None:
                continue

            for row in rows[header_index + 1:]:
                if len(row) <= max(name_col, id_col):
                    continue
                name = row[name_col]
                pid = row[id_col]
                if not display_text(name) and not is_valid_id(pid):
                    continue
                bday = row[bday_col] if bday_col is not None and len(row) > bday_col else None
                rec = add_identity(
                    records,
                    name_to_ids,
                    chart_to_ids,
                    name=name,
                    pid=pid,
                    bday=bday,
                    source=path.name,
                )
                if rec is None:
                    continue
                if sheet_is_self_select:
                    rec.is_self_select = True
                if ascvd_col is not None and len(row) > ascvd_col and row[ascvd_col] not in (None, ""):
                    rec.ascvd = row[ascvd_col]
                if disease_col is not None and len(row) > disease_col and row[disease_col] not in (None, ""):
                    rec.dmk_raw = row[disease_col]
                if p4p_plan_col is not None and len(row) > p4p_plan_col:
                    rec.p4p["plan"] = row[p4p_plan_col]
                if p4p_status_col is not None and len(row) > p4p_status_col:
                    rec.p4p["status"] = row[p4p_status_col]
                if hba_val_col is not None and len(row) > hba_val_col:
                    rec.health["hba_val"] = row[hba_val_col]
                if hba_dt_col is not None and len(row) > hba_dt_col:
                    rec.health["hba_dt"] = parse_roc_or_ad_date(row[hba_dt_col])
                if ldl_val_col is not None and len(row) > ldl_val_col:
                    rec.health["ldl_val"] = row[ldl_val_col]
                if ldl_dt_col is not None and len(row) > ldl_dt_col:
                    rec.health["ldl_dt"] = parse_roc_or_ad_date(row[ldl_dt_col])
                if uacr_val_col is not None and len(row) > uacr_val_col:
                    rec.health["uacr_val"] = row[uacr_val_col]
                if uacr_dt_col is not None and len(row) > uacr_dt_col:
                    rec.health["uacr_dt"] = parse_roc_or_ad_date(row[uacr_dt_col])
                if indicator_col is not None and len(row) > indicator_col:
                    indicator = str(row[indicator_col] or "")
                    last_dt = row[last_screen_col] if last_screen_col is not None and len(row) > last_screen_col else None
                    if "成人預防保健" in indicator or "成人健檢" in indicator:
                        rec.screenings["成人健檢"] = parse_roc_or_ad_date(last_dt) or "-"
                    elif "子宮頸抹片" in indicator or "子宮抹片" in indicator:
                        rec.screenings["子宮抹片"] = parse_roc_or_ad_date(last_dt) or "-"
                    elif "流感" in indicator:
                        rec.screenings["老人流感"] = parse_roc_or_ad_date(last_dt) or "-"
                    elif "糞便潛血" in indicator or "潛血" in indicator:
                        rec.screenings["糞便潛血"] = parse_roc_or_ad_date(last_dt) or "-"
                    elif "肝炎" in indicator or "B、C肝" in indicator or "BC肝" in indicator.upper():
                        rec.screenings["BC肝炎"] = parse_roc_or_ad_date(last_dt) or "-"


def scan_count_files(
    source_dir: Path,
    records: Dict[str, MemberRecord],
    name_to_ids: Dict[str, Set[str]],
    chart_to_ids: Dict[str, Set[str]],
) -> None:
    count_dir = source_dir / "次數"
    if not count_dir.is_dir():
        return
    for path in sorted(count_dir.glob("*.xls")):
        month_match = re.fullmatch(r"1(14|15)(\d{2})", path.stem)
        if not month_match:
            continue
        year_bucket = int(path.stem[:3])
        month = int(path.stem[3:5])
        rows = read_html_table(path)
        header_index = find_header(rows, [["病歷號"], ["身分證", "身分證號", "身份證號"], ["看診次數", "次數"]])
        if header_index is None:
            continue
        header = rows[header_index]
        chart_col = find_col(header, ["病歷號", "病歷號碼"])
        name_col = find_col(header, ["姓名", "病患姓名"])
        id_col = find_col(header, ["身分證", "身分證號", "身份證號", "ID"])
        bday_col = find_col(header, ["生日"])
        count_col = find_col(header, ["看診次數", "次數", "就診次數"])
        if None in (chart_col, name_col, id_col, count_col):
            continue
        for row in rows[header_index + 1:]:
            if len(row) <= max(chart_col, name_col, id_col, count_col):
                continue
            rec = add_identity(
                records,
                name_to_ids,
                chart_to_ids,
                chart=row[chart_col],
                name=row[name_col],
                pid=row[id_col],
                bday=row[bday_col] if bday_col is not None and len(row) > bday_col else None,
                source=f"次數/{path.name}",
            )
            if rec is None:
                continue
            count = parse_number(row[count_col])
            if year_bucket == 114:
                rec.count_114 += count
                if month <= 4:
                    rec.count_114_q1 += count
            elif year_bucket == 115:
                rec.count_115 += count
                if month <= 4:
                    rec.count_115_q1 += count
            rec.notes.add("病歷號對到ID")


def attach_fee_record(
    records: Dict[str, MemberRecord],
    name_to_ids: Dict[str, Set[str]],
    chart_to_ids: Dict[str, Set[str]],
    *,
    chart: str,
    name: str,
    visit_date: Optional[dt.date],
    amount: float,
    source: str,
) -> MemberRecord:
    chart_ids = chart_to_ids.get(chart, set())
    rec: Optional[MemberRecord] = None
    if len(chart_ids) == 1:
        rec = get_or_create_by_id(records, next(iter(chart_ids)))
        if rec:
            rec.notes.add("病歷號對到ID")
    elif len(chart_ids) > 1:
        rec = get_or_create_by_name(records, name)
        if rec:
            rec.notes.add("病歷號對到多個ID")
    else:
        name_ids = name_to_ids.get(normalize_name(name), set())
        if len(name_ids) == 1:
            rec = get_or_create_by_id(records, next(iter(name_ids)))
            if rec:
                rec.notes.add("姓名唯一對到ID")
        elif len(name_ids) > 1:
            rec = get_or_create_by_name(records, name)
            if rec:
                rec.notes.add("姓名同名多人，未自動帶ID")
        else:
            rec = get_or_create_by_name(records, name)
            if rec:
                rec.notes.add("未對到ID")
    if rec is None:
        rec = MemberRecord(name=name)
        records[f"FEE:{source}:{chart}:{name}:{len(records)}"] = rec
        rec.notes.add("未對到ID")

    rec.name = rec.name or name
    if chart:
        rec.charts.add(chart)
    if visit_date and (rec.last_visit is None or visit_date > rec.last_visit):
        rec.last_visit = visit_date
    if visit_date:
        year_bucket = visit_date.year - 1911
        month = visit_date.month
        if year_bucket == 114:
            rec.amount_114 += amount
            if month <= 4:
                rec.amount_114_q1 += amount
        elif year_bucket == 115:
            rec.amount_115 += amount
            if month <= 4:
                rec.amount_115_q1 += amount
    rec.sources.add(source)
    return rec


def scan_fee_files(
    source_dir: Path,
    records: Dict[str, MemberRecord],
    name_to_ids: Dict[str, Set[str]],
    chart_to_ids: Dict[str, Set[str]],
) -> None:
    fee_dir = source_dir / "費用"
    if not fee_dir.is_dir():
        return
    for path in sorted(fee_dir.glob("*.xls")):
        rows = read_html_table(path)
        header_index = find_header(rows, [["病歷號"], ["姓名"], ["掛帳費"]])
        if header_index is None:
            continue
        header = rows[header_index]
        chart_col = find_col(header, ["病歷號", "病歷號碼"])
        name_col = find_col(header, ["姓名", "病患姓名"])
        date_col = find_col(header, ["日期"])
        amount_col = find_col(header, ["掛帳費"])
        if None in (chart_col, name_col, date_col, amount_col):
            continue
        for row in rows[header_index + 1:]:
            if len(row) <= max(chart_col, name_col, date_col, amount_col):
                continue
            raw_date = str(row[date_col] or "").strip()
            if not re.fullmatch(r"1(?:14|15)\d{4}", raw_date):
                continue
            chart = normalize_chart(row[chart_col])
            name = display_text(row[name_col])
            visit_date = parse_roc_yyyymmdd(raw_date)
            amount = parse_number(row[amount_col])
            attach_fee_record(
                records,
                name_to_ids,
                chart_to_ids,
                chart=chart,
                name=name,
                visit_date=visit_date,
                amount=amount,
                source=f"費用/{path.name}",
            )


def collect_data(source_dir: Path) -> Dict[str, MemberRecord]:
    records: Dict[str, MemberRecord] = {}
    name_to_ids: Dict[str, Set[str]] = defaultdict(set)
    chart_to_ids: Dict[str, Set[str]] = defaultdict(set)
    scan_root_files(source_dir, records, name_to_ids, chart_to_ids)
    scan_count_files(source_dir, records, name_to_ids, chart_to_ids)
    scan_fee_files(source_dir, records, name_to_ids, chart_to_ids)
    merge_name_records(records, name_to_ids)
    for rec in records.values():
        if not rec.pid:
            rec.notes.add("ID欄填-")
        if not rec.bday:
            rec.notes.add("缺生日")
        if rec.pid:
            rec.sex = infer_gender(rec.pid)
    return records


def is_self_select_source(text: Any) -> bool:
    compact = re.sub(r"[\s\-_()（）\[\]{}]+", "", str(text or "").strip()).lower()
    if not compact:
        return False
    if "115x" in compact or "不選" in compact or "不要" in compact:
        return False
    return any(token in compact for token in ("自選名單", "自選會員", "115自選", "a115"))


def style_like_previous_row(ws, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy.copy(src._style)
        dst.number_format = src.number_format
        dst.font = copy.copy(src.font)
        dst.fill = copy.copy(src.fill)
        dst.border = copy.copy(src.border)
        dst.alignment = copy.copy(src.alignment)
        dst.protection = copy.copy(src.protection)


def ensure_extra_columns(ws) -> Tuple[int, int]:
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    compact = {compact_header(value): index + 1 for index, value in enumerate(headers)}
    max_col = ws.max_column

    chart_col = compact.get(compact_header("病歷號"))
    if chart_col is None:
        max_col += 1
        chart_col = max_col
        ws.cell(1, chart_col).value = "病歷號"
        ws.cell(2, chart_col).value = ""

    note_col = compact.get(compact_header("備註"))
    if note_col is None:
        max_col += 1
        note_col = max_col
        ws.cell(1, note_col).value = "備註"
        ws.cell(2, note_col).value = ""

    for col in (chart_col, note_col):
        src = ws.cell(1, ws.max_column if ws.max_column < col else col - 1)
        ws.cell(1, col).font = Font(bold=True)
        ws.cell(1, col).fill = PatternFill("solid", fgColor="D9EAF7")
        ws.cell(1, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = 28 if col == note_col else 18
    return chart_col, note_col


def find_output_col(ws, aliases: Sequence[str]) -> Optional[int]:
    for row in (1, 2):
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row, col).value
            if any(compact_header(value) == compact_header(alias) for alias in aliases):
                return col
    return None


def set_if_col(ws, row: int, col: Optional[int], value: Any) -> None:
    if col:
        ws.cell(row, col).value = value


def number_or_blank(value: float) -> Any:
    return int(value) if value else None


def date_or_blank(value: Optional[dt.date]) -> Any:
    return value if value else None


def screening_value(rec: MemberRecord, key: str) -> Any:
    value = rec.screenings.get(key)
    if value in (None, "", "-"):
        return None
    return value


def display_pid(rec: MemberRecord) -> str:
    return rec.pid or "-"


def avg_amount(total: float, months: int) -> Any:
    if not total or months <= 0:
        return None
    return round(total / months, 2)


def is_roc_year(value: Any, roc_year: int) -> bool:
    parsed = parse_roc_or_ad_date(value)
    return bool(parsed and parsed.year - 1911 == roc_year)


def has_screening_done(rec: MemberRecord, key: str) -> bool:
    value = rec.screenings.get(key)
    return value not in (None, "", "-")


def score_record(rec: MemberRecord) -> Tuple[int, str, str, Tuple[int, int, int, int]]:
    visit_score = 10 if (rec.count_115 >= 4 or rec.count_114 >= 6) else 0

    fee_score = 0
    avg_114_q1 = rec.amount_114_q1 / 4 if rec.amount_114_q1 else 0
    avg_115_q1 = rec.amount_115_q1 / 4 if rec.amount_115_q1 else 0
    if avg_114_q1 and avg_115_q1 and avg_115_q1 < avg_114_q1:
        fee_score = 6

    hba_2026 = is_roc_year(rec.health.get("hba_dt"), 115)
    ldl_2026 = is_roc_year(rec.health.get("ldl_dt"), 115)
    uacr_2026 = is_roc_year(rec.health.get("uacr_dt"), 115)
    exam_score = 10 if (hba_2026 or ldl_2026 or uacr_2026) else 0

    prevention_items = [
        ("成人預防保健", 6),
        ("子宮抹片", 6),
        ("老人流感", 4),
        ("糞便潛血", 6),
        ("BC肝炎", 6),
    ]
    prevention_score = sum(points for key, points in prevention_items if has_screening_done(rec, key))
    missing = [key for key, _ in prevention_items if not has_screening_done(rec, key)]
    reminder = "；".join(f"今年需檢測{key}" for key in missing) if missing else "-"

    total = visit_score + fee_score + exam_score + prevention_score
    breakdown = "\n".join([
        f"1. 固定就診次數：{visit_score} 分",
        f"2. 醫療費用：{fee_score} 分",
        f"3. 糖心腎管理：{exam_score} 分",
        f"4. 預防保健：{prevention_score} 分",
    ])
    return total, breakdown, reminder, (visit_score, fee_score, exam_score, prevention_score)


def clear_sheet_data(ws, start_row: int) -> None:
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def write_row_values(ws, row_index: int, values_by_col: Dict[int, Any], style_row: int) -> None:
    max_col = max(ws.max_column, max(values_by_col.keys(), default=1))
    style_like_previous_row(ws, style_row, row_index, max_col)
    for col, value in values_by_col.items():
        ws.cell(row_index, col).value = value


def populate_doctor_sheet(wb, records: List[MemberRecord], today: dt.date) -> None:
    if DOCTOR_SHEET_NAME not in wb.sheetnames:
        return
    ws = wb[DOCTOR_SHEET_NAME]
    clear_sheet_data(ws, 4)
    for offset, rec in enumerate(records, start=4):
        total_score, breakdown, reminder, scores = score_record(rec)
        write_row_values(
            ws,
            offset,
            {
                1: display_pid(rec),
                2: rec.name or "-",
                3: date_or_blank(rec.bday),
                4: age_at_today(rec.bday, today),
                8: rec.dmk_raw,
                9: rec.ascvd,
                10: date_or_blank(rec.last_visit),
                11: number_or_blank(rec.count_114),
                12: number_or_blank(rec.count_114_q1),
                13: number_or_blank(rec.count_115),
                14: avg_amount(rec.amount_114, 12),
                15: avg_amount(rec.amount_115, 4),
                16: screening_value(rec, "成人健檢"),
                17: screening_value(rec, "子宮抹片"),
                18: screening_value(rec, "老人流感"),
                19: screening_value(rec, "糞便潛血"),
                20: screening_value(rec, "BC肝炎"),
                21: rec.health.get("hba_val"),
                22: rec.health.get("hba_dt"),
                23: rec.health.get("ldl_val"),
                24: rec.health.get("ldl_dt"),
                25: rec.health.get("uacr_val"),
                26: rec.health.get("uacr_dt"),
                28: rec.p4p.get("status"),
                35: total_score,
                36: scores[0],
                37: scores[1],
                38: scores[2],
                39: scores[3],
                40: breakdown,
                41: reminder,
            },
            style_row=3,
        )
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A4"


def populate_self_select_sheet(wb, records: List[MemberRecord]) -> None:
    if SELF_SELECT_SHEET_NAME not in wb.sheetnames:
        return
    ws = wb[SELF_SELECT_SHEET_NAME]
    clear_sheet_data(ws, 3)
    selected_records = [rec for rec in records if rec.is_self_select]
    for offset, rec in enumerate(selected_records, start=3):
        total_score, _breakdown, _reminder, scores = score_record(rec)
        write_row_values(
            ws,
            offset,
            {
                1: rec.name or "-",
                2: display_pid(rec),
                3: number_or_blank(rec.count_114),
                4: number_or_blank(rec.count_114_q1),
                5: number_or_blank(rec.count_115),
                6: number_or_blank(rec.amount_114),
                7: number_or_blank(rec.amount_115),
                8: screening_value(rec, "成人健檢"),
                9: screening_value(rec, "子宮抹片"),
                10: screening_value(rec, "老人流感"),
                11: screening_value(rec, "糞便潛血"),
                12: screening_value(rec, "BC肝炎"),
                13: rec.health.get("hba_val"),
                14: rec.health.get("hba_dt"),
                15: rec.health.get("ldl_val"),
                16: rec.health.get("ldl_dt"),
                17: rec.p4p.get("status"),
                21: "✔" if rec.ascvd or rec.dmk_raw else None,
                22: total_score,
                23: scores[0],
                24: scores[1],
                25: scores[2],
                26: scores[3],
            },
            style_row=2,
        )
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A3"


def parse_lab_number(value: Any) -> Optional[float]:
    if value in (None, "", "-"):
        return None
    try:
        return float(str(value).strip())
    except Exception:
        return None


def populate_percentile_sheet(wb, records: List[MemberRecord]) -> None:
    if PERCENTILE_SHEET_NAME not in wb.sheetnames:
        return
    ws = wb[PERCENTILE_SHEET_NAME]
    clear_sheet_data(ws, 5)
    ldl_rows = [
        rec for rec in records
        if parse_lab_number(rec.health.get("ldl_val")) is not None
    ]
    hba_rows = [
        rec for rec in records
        if parse_lab_number(rec.health.get("hba_val")) is not None
    ]
    ldl_rows.sort(key=lambda rec: parse_lab_number(rec.health.get("ldl_val")) or 0, reverse=True)
    hba_rows.sort(key=lambda rec: parse_lab_number(rec.health.get("hba_val")) or 0, reverse=True)
    ws["A1"] = f"LDL名單({len(ldl_rows)}人)"
    ws["N1"] = f"HBA1C名單({len(hba_rows)}人)"
    max_len = max(len(ldl_rows), len(hba_rows))
    for index in range(max_len):
        row_index = 5 + index
        values: Dict[int, Any] = {}
        if index < len(ldl_rows):
            rec = ldl_rows[index]
            values.update({
                1: rec.name or "-",
                2: date_or_blank(rec.bday),
                3: display_pid(rec),
                5: "；".join(sorted(rec.notes)) if rec.notes else None,
                6: rec.health.get("ldl_val"),
                7: rec.health.get("ldl_dt"),
            })
        if index < len(hba_rows):
            rec = hba_rows[index]
            values.update({
                14: rec.name or "-",
                15: date_or_blank(rec.bday),
                16: display_pid(rec),
                18: "；".join(sorted(rec.notes)) if rec.notes else None,
                19: rec.health.get("hba_val"),
                20: rec.health.get("hba_dt"),
            })
        write_row_values(ws, row_index, values, style_row=4)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A5"


def create_output(source_dir: Path, output_dir: Optional[Path] = None) -> Path:
    records = collect_data(source_dir)
    template_path = find_template()
    wb = openpyxl.load_workbook(template_path)
    ws = wb[TARGET_SHEET]
    generic.prepare_template_layout(ws)
    chart_col, note_col = ensure_extra_columns(ws)

    cols = {
        "name": find_output_col(ws, ["姓名"]),
        "member": find_output_col(ws, ["會員"]),
        "id": find_output_col(ws, ["身份證號碼", "身分證號碼", "ID"]),
        "sex": find_output_col(ws, ["性別"]),
        "bday": find_output_col(ws, ["生日"]),
        "age": find_output_col(ws, ["年齡"]),
        "last_visit": find_output_col(ws, ["最後就診日"]),
        "count_114": find_output_col(ws, ["114年就診次數"]),
        "amount_114": find_output_col(ws, ["114年實際申報總額", "114年申報金額/月"]),
        "count_115": find_output_col(ws, ["115年就診次數"]),
        "amount_115": find_output_col(ws, ["115年實際申報總額", "115年申報金額/月"]),
        "ascvd": find_output_col(ws, ["ASCVD"]),
        "dmk": find_output_col(ws, ["疾病樣態分類(7類)", "疾病樣態編號"]),
        "hba_val": find_output_col(ws, ["最近一次HbA1c檢查結果(%)"]),
        "hba_dt": find_output_col(ws, ["最近一次HbA1c檢查日期"]),
        "ldl_val": find_output_col(ws, ["最近一次LDL\n檢查結果(mg/dL)", "最近一次LDL檢查結果(mg/dL)"]),
        "ldl_dt": find_output_col(ws, ["最近一次LDL\n檢查日期", "最近一次LDL檢查日期"]),
        "uacr_val": find_output_col(ws, ["最近一次UACR檢查結果", "最近一次UACR檢查結果(mg/gm)"]),
        "uacr_dt": find_output_col(ws, ["最近一次UACR\n檢查日期", "最近一次UACR檢查日期"]),
        "p4p_plan": find_output_col(ws, ["P4P收案計畫"]),
        "p4p_status": find_output_col(ws, ["收案狀態"]),
        "score": find_output_col(ws, ["分數"]),
        "breakdown": find_output_col(ws, ["分數說明"]),
        "prevention_note": find_output_col(ws, ["預防保健提醒", "備註"]),
        "is_self_select": find_output_col(ws, ["是否為自選會員"]),
    }
    screening_cols = {key: find_output_col(ws, [header]) for key, header in SCREENING_COLUMNS.items()}

    data_start = 3
    if ws.max_row >= data_start:
        ws.delete_rows(data_start, ws.max_row - data_start + 1)

    today = dt.date.today()

    def sort_key(item: MemberRecord) -> Tuple[int, str, str]:
        return (0 if item.pid else 1, normalize_name(item.name), item.pid)

    sorted_records = sorted(records.values(), key=sort_key)

    row_index = data_start
    max_col = ws.max_column
    for rec in sorted_records:
        style_like_previous_row(ws, 2, row_index, max_col)
        age = age_at_today(rec.bday, today)
        total_score, breakdown, reminder, _scores = score_record(rec)
        set_if_col(ws, row_index, cols["name"], rec.name or "-")
        set_if_col(ws, row_index, cols["member"], None)
        set_if_col(ws, row_index, cols["id"], rec.pid or "-")
        set_if_col(ws, row_index, cols["sex"], rec.sex or None)
        set_if_col(ws, row_index, cols["bday"], rec.bday)
        set_if_col(ws, row_index, cols["age"], age)
        set_if_col(ws, row_index, cols["last_visit"], rec.last_visit)
        set_if_col(ws, row_index, cols["count_114"], number_or_blank(rec.count_114))
        set_if_col(ws, row_index, cols["amount_114"], number_or_blank(rec.amount_114))
        set_if_col(ws, row_index, cols["count_115"], number_or_blank(rec.count_115))
        set_if_col(ws, row_index, cols["amount_115"], number_or_blank(rec.amount_115))
        set_if_col(ws, row_index, cols["ascvd"], rec.ascvd)
        set_if_col(ws, row_index, cols["dmk"], rec.dmk_raw)
        set_if_col(ws, row_index, cols["hba_val"], rec.health.get("hba_val"))
        set_if_col(ws, row_index, cols["hba_dt"], rec.health.get("hba_dt"))
        set_if_col(ws, row_index, cols["ldl_val"], rec.health.get("ldl_val"))
        set_if_col(ws, row_index, cols["ldl_dt"], rec.health.get("ldl_dt"))
        set_if_col(ws, row_index, cols["uacr_val"], rec.health.get("uacr_val"))
        set_if_col(ws, row_index, cols["uacr_dt"], rec.health.get("uacr_dt"))
        set_if_col(ws, row_index, cols["p4p_plan"], rec.p4p.get("plan"))
        set_if_col(ws, row_index, cols["p4p_status"], rec.p4p.get("status"))
        set_if_col(ws, row_index, cols["score"], total_score)
        set_if_col(ws, row_index, cols["breakdown"], breakdown)
        set_if_col(ws, row_index, cols["prevention_note"], reminder)
        set_if_col(ws, row_index, cols["is_self_select"], "✔" if rec.is_self_select else None)
        for key, col in screening_cols.items():
            value = rec.screenings.get(key)
            set_if_col(ws, row_index, col, value if value not in (None, "") else None)
        ws.cell(row_index, chart_col).value = "、".join(sorted(rec.charts)) if rec.charts else "-"
        ws.cell(row_index, note_col).value = "；".join(sorted(rec.notes)) if rec.notes else "-"
        row_index += 1

    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        if col in (chart_col, note_col):
            continue
        if ws.column_dimensions[letter].width is None:
            ws.column_dimensions[letter].width = 12

    for row in ws.iter_rows(min_row=data_start, max_row=row_index - 1):
        for cell in row:
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal,
                vertical="center",
                wrap_text=True,
            )
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A3"
    populate_doctor_sheet(wb, sorted_records, today)
    populate_self_select_sheet(wb, sorted_records)
    populate_percentile_sheet(wb, sorted_records)

    if output_dir is None:
        output_dir = source_dir.parent
    output_dir.mkdir(parents=True, exist_ok=True)
    clinic_name = extract_clinic_name(source_dir) or source_dir.name
    timestamp = dt.datetime.now().strftime("%m%d_%H%M")
    out_path = output_dir / f"{sanitize_filename(clinic_name)}展望雲端選會員_{timestamp}.xlsx"
    wb.save(out_path)
    return out_path


def extract_clinic_name(source_dir: Path) -> str:
    text = source_dir.name
    match = re.search(r"\d{10}(.+?)(?:雲端)?展望?$", text)
    if match:
        return match.group(1)
    return ""


def sanitize_filename(text: str) -> str:
    text = re.sub(r'[\\/:*?"<>|]+', "_", text).strip()
    return text or "展望雲端"


def choose_folder_gui() -> Optional[str]:
    try:
        import tkinter as tk
        from tkinter import filedialog
    except Exception:
        return None
    root = tk.Tk()
    root.withdraw()
    folder = filedialog.askdirectory(title="請選擇展望雲端資料夾")
    root.destroy()
    return folder or None


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="展望雲端專用選會員產檔")
    parser.add_argument("source_dir", nargs="?", help="展望雲端資料夾")
    parser.add_argument("-o", "--output-dir", help="輸出資料夾；預設為來源底下的展望雲端整理")
    args = parser.parse_args(argv)

    source = args.source_dir or choose_folder_gui()
    if not source:
        print("未選擇資料夾。")
        return 1
    source_dir = Path(source).expanduser().resolve()
    if not source_dir.is_dir():
        print(f"找不到資料夾：{source_dir}")
        return 1
    output_dir = Path(args.output_dir).expanduser().resolve() if args.output_dir else None
    out_path = create_output(source_dir, output_dir)
    print(f"已輸出：{out_path}")
    return 0


def gui_main() -> None:
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()

    source = filedialog.askdirectory(title="選擇來源資料夾（展望雲端資料夾）")
    if not source:
        return

    try:
        out_path = create_output(Path(source).expanduser().resolve())
        messagebox.showinfo("完成", f"已輸出：\n{out_path}")
        generic.open_file_cross_platform(str(out_path))
    except Exception as exc:
        messagebox.showerror("錯誤", str(exc))


if __name__ == "__main__":
    if len(sys.argv) > 1:
        raise SystemExit(main())
    gui_main()
