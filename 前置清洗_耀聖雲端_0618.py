# -*- coding: utf-8 -*-
"""
耀聖雲端前置清洗 + 共用核心包裝（0618）

適用資料夾範例：
- 耀聖雲端/3501105767愛鄰耀聖

用途：
- 不修改共用核心。
- 保留來源根目錄既有 xlsx 資料：自選會員、114 家醫名單、HealthCase、預防保健 cliScores。
- 將「費用/*.xlsx」收入明細轉成共用核心可讀的月份費用中間檔。
- 費用明細以病歷號優先對 ID，姓名唯一對 ID 為輔；未對到 ID 的費用彙總後追加到會員總表，ID 留空。

注意：
- 未提供正式次數檔時，由費用明細以「同一會員同一天算一次」反推看診次數，避免用明細列數高估。
- 若之後取得正式次數檔，應優先使用正式次數檔，不要和費用反推次數重複計算。
"""

from __future__ import annotations

import argparse
import copy
import datetime as dt
import importlib.util
import os
import re
import shutil
import subprocess
import sys
import tempfile
from collections import defaultdict
from pathlib import Path
from typing import Any, DefaultDict, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
VERSION_TAG = "0618"


ID_ALIASES = (
    "ID", "身分證", "身分證號", "身分證號碼", "身份證", "身份證號", "身份證號碼",
    "家醫收案會員ID",
)
NAME_ALIASES = ("姓名", "會員姓名", "病患姓名")
BIRTH_ALIASES = ("生日", "出生日期", "BIRTHDAY")
CHART_ALIASES = ("病歷號", "病歷號碼")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def normalize_name(value: Any) -> str:
    return re.sub(r"[\s　]+", "", normalize_text(value)).upper()


def normalize_id(value: Any) -> str:
    return re.sub(r"\s+", "", normalize_text(value).upper().strip("'"))


def normalize_chart(value: Any) -> str:
    text = normalize_text(value)
    return text.zfill(7) if text.isdigit() else text


def is_valid_id(value: Any) -> bool:
    pid = normalize_id(value)
    return bool(re.fullmatch(r"(?:[A-Z][1289]\d{8}|[A-Z][A-D]\d{8}|[A-Z]{1,2}\d{8,10})", pid))


def parse_number(value: Any) -> float:
    text = normalize_text(value).replace(",", "")
    if text in {"", "-", "—", "–"}:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def parse_date(value: Any) -> Optional[dt.date]:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = normalize_text(value)
    if text in {"", "-", "—", "–"}:
        return None
    text = text.split()[0]
    match = re.fullmatch(r"(\d{2,4})[./\-](\d{1,2})[./\-](\d{1,2})", text)
    if match:
        year = int(match.group(1))
        if year < 1911:
            year += 1911
        try:
            return dt.date(year, int(match.group(2)), int(match.group(3)))
        except ValueError:
            return None
    digits = re.sub(r"\D", "", text)
    try:
        if len(digits) == 7:
            return dt.date(int(digits[:3]) + 1911, int(digits[3:5]), int(digits[5:7]))
        if len(digits) == 8:
            year = int(digits[:4])
            if year < 1911:
                year += 1911
            return dt.date(year, int(digits[4:6]), int(digits[6:8]))
        if len(digits) == 6:
            return dt.date(int(digits[:2]) + 1911, int(digits[2:4]), int(digits[4:6]))
    except ValueError:
        return None
    return None


def compact_header(value: Any) -> str:
    return re.sub(r"[\s　]+", "", normalize_text(value)).upper()


def find_col(header: Sequence[Any], aliases: Iterable[str]) -> Optional[int]:
    compact = [compact_header(value) for value in header]
    for alias in aliases:
        target = compact_header(alias)
        for index, value in enumerate(compact):
            if value == target:
                return index
    return None


def find_header_row(
    rows: Sequence[Sequence[Any]],
    required_groups: Sequence[Sequence[str]],
    max_rows: int = 20,
) -> Optional[int]:
    groups = [{compact_header(alias) for alias in group} for group in required_groups]
    for index, row in enumerate(rows[:max_rows]):
        values = {compact_header(value) for value in row}
        if all(group & values for group in groups):
            return index
    return None


def iter_workbook_rows(path: Path, max_rows: Optional[int] = None) -> Iterable[Tuple[str, List[List[Any]]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            rows: List[List[Any]] = []
            for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
                values = list(row)
                if any(value not in (None, "") for value in values):
                    rows.append(values)
                if max_rows is not None and row_index >= max_rows:
                    break
            yield ws.title, rows
    finally:
        wb.close()


def find_common_core() -> Path:
    def key(path: Path) -> Tuple[int, str]:
        match = re.search(r"(\d{4})(?=\.py$)", path.name)
        return (int(match.group(1)) if match else -1, path.name)

    candidates = sorted(SCRIPT_DIR.glob("選會員_共用核心_*.py"), key=key, reverse=True)
    if not candidates:
        raise RuntimeError("找不到共用核心：選會員_共用核心_*.py")
    return candidates[0]


def find_template() -> Path:
    def key(path: Path) -> Tuple[int, str]:
        match = re.search(r"(\d{4})(?=\.xlsx$)", path.name)
        return (int(match.group(1)) if match else -1, path.name)

    candidates = sorted(SCRIPT_DIR.glob("選會員模板*.xlsx"), key=key, reverse=True)
    if not candidates:
        raise RuntimeError("找不到模板：選會員模板*.xlsx")
    return candidates[0]


def load_common_core():
    core_path = find_common_core()
    spec = importlib.util.spec_from_file_location("ys_cloud_common_core", core_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"無法載入共用核心：{core_path}")
    mod = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = mod
    spec.loader.exec_module(mod)
    return mod


def is_source_workbook(path: Path) -> bool:
    if not path.is_file() or path.name.startswith("~$"):
        return False
    if path.suffix.lower() not in {".xlsx", ".xlsm", ".xls", ".ods"}:
        return False
    if path.name == ".DS_Store" or path.name.lower() == "desktop.ini":
        return False
    if "未對到ID" in path.stem:
        return False
    if re.search(r"選會員_\d{4}_\d{4}", path.stem):
        return False
    return True


def is_fee_workbook(path: Path) -> bool:
    return path.parent.name == "費用" and path.suffix.lower() in {".xlsx", ".xlsm"}


def build_identity_indexes(
    source_dir: Path,
) -> Tuple[DefaultDict[str, set[str]], DefaultDict[str, set[str]], Dict[str, str]]:
    name_to_ids: DefaultDict[str, set[str]] = defaultdict(set)
    chart_to_ids: DefaultDict[str, set[str]] = defaultdict(set)
    id_to_name: Dict[str, str] = {}

    for path in sorted(source_dir.glob("*.xlsx")):
        if not is_source_workbook(path):
            continue
        for _sheet_name, rows in iter_workbook_rows(path, max_rows=None):
            header_row = find_header_row(rows, (ID_ALIASES,), max_rows=12)
            if header_row is None:
                continue
            header = rows[header_row]
            id_col = find_col(header, ID_ALIASES)
            name_col = find_col(header, NAME_ALIASES)
            chart_col = find_col(header, CHART_ALIASES)
            if id_col is None:
                continue
            max_col = max(col for col in (id_col, name_col, chart_col) if col is not None)
            for row in rows[header_row + 1:]:
                if len(row) <= max_col:
                    continue
                pid = normalize_id(row[id_col])
                if not is_valid_id(pid):
                    continue
                if name_col is not None:
                    name = normalize_name(row[name_col])
                    if name:
                        name_to_ids[name].add(pid)
                        id_to_name.setdefault(pid, normalize_text(row[name_col]))
                if chart_col is not None:
                    chart = normalize_chart(row[chart_col])
                    if chart:
                        chart_to_ids[chart].add(pid)

    return name_to_ids, chart_to_ids, id_to_name


def read_fee_rows(
    source_dir: Path,
    name_to_ids: Dict[str, set[str]],
    chart_to_ids: Dict[str, set[str]],
    id_to_name: Dict[str, str],
) -> Tuple[Dict[Tuple[int, int], List[List[Any]]], List[Dict[str, Any]], Dict[str, Any]]:
    fee_dir = source_dir / "費用"
    monthly_rows: Dict[Tuple[int, int], List[List[Any]]] = defaultdict(list)
    unmatched: List[Dict[str, Any]] = []
    stats: Dict[str, Any] = {
        "files": 0,
        "raw_rows": 0,
        "matched_chart": 0,
        "matched_name": 0,
        "unmatched": 0,
        "matched_amount_114": 0.0,
        "matched_amount_115": 0.0,
        "unmatched_amount_114": 0.0,
        "unmatched_amount_115": 0.0,
        "matched_detail_rows_114": 0,
        "matched_detail_rows_115": 0,
        "unmatched_detail_rows_114": 0,
        "unmatched_detail_rows_115": 0,
    }
    if not fee_dir.is_dir():
        return monthly_rows, unmatched, stats

    for path in sorted(fee_dir.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        stats["files"] += 1
        for _sheet_name, rows in iter_workbook_rows(path, max_rows=None):
            header_row = find_header_row(rows, (("病歷號",), ("姓名",), ("日期",), ("應付",)), max_rows=10)
            if header_row is None:
                continue
            header = rows[header_row]
            chart_col = find_col(header, ("病歷號", "病歷號碼"))
            name_col = find_col(header, ("姓名", "病患姓名"))
            date_col = find_col(header, ("日期",))
            amount_col = find_col(header, ("應付", "已付", "總額", "金額"))
            if None in (chart_col, name_col, date_col, amount_col):
                continue
            max_col = max(chart_col, name_col, date_col, amount_col)  # type: ignore[arg-type]
            for row_no, row in enumerate(rows[header_row + 1:], start=header_row + 2):
                if len(row) <= max_col:
                    continue
                visit_date = parse_date(row[date_col])  # type: ignore[index]
                if visit_date is None:
                    continue
                roc_year = visit_date.year - 1911
                if roc_year not in (114, 115):
                    continue
                chart = normalize_chart(row[chart_col])  # type: ignore[index]
                name = normalize_text(row[name_col])  # type: ignore[index]
                amount = parse_number(row[amount_col])  # type: ignore[index]
                if not chart and not name:
                    continue
                stats["raw_rows"] += 1

                pid = ""
                match_kind = ""
                chart_ids = chart_to_ids.get(chart, set()) if chart else set()
                if len(chart_ids) == 1:
                    pid = next(iter(chart_ids))
                    match_kind = "病歷號"
                    stats["matched_chart"] += 1
                else:
                    name_ids = name_to_ids.get(normalize_name(name), set()) if name else set()
                    if len(name_ids) == 1:
                        pid = next(iter(name_ids))
                        match_kind = "姓名唯一"
                        stats["matched_name"] += 1

                amount_key = "matched_amount_115" if roc_year == 115 else "matched_amount_114"
                unmatched_amount_key = "unmatched_amount_115" if roc_year == 115 else "unmatched_amount_114"
                if pid:
                    stats[amount_key] += amount
                    stats[f"matched_detail_rows_{roc_year}"] += 1
                    monthly_rows[(roc_year, visit_date.month)].append([
                        pid,
                        id_to_name.get(pid) or name,
                        visit_date,
                        0,
                        amount,
                        match_kind,
                        path.name,
                        row_no,
                    ])
                else:
                    stats["unmatched"] += 1
                    stats[unmatched_amount_key] += amount
                    stats[f"unmatched_detail_rows_{roc_year}"] += 1
                    unmatched.append({
                        "file": path.name,
                        "row": row_no,
                        "chart": chart,
                        "name": name,
                        "date": visit_date,
                        "amount": amount,
                        "reason": "病歷號與姓名都無法唯一對到 ID",
                    })
    return monthly_rows, unmatched, stats


def aggregate_matched_fee_visits(
    monthly_rows: Dict[Tuple[int, int], List[List[Any]]],
) -> Tuple[Dict[Tuple[int, int], List[List[Any]]], Dict[str, int]]:
    aggregated: Dict[Tuple[int, int], List[List[Any]]] = defaultdict(list)
    visit_counter = {"matched_visit_count_114": 0, "matched_visit_count_115": 0}
    for year_month, rows in monthly_rows.items():
        year, _month = year_month
        groups: Dict[Tuple[str, dt.date], Dict[str, Any]] = {}
        for row in rows:
            pid, name, visit_date, _count, amount, match_kind, source_file, source_row = row
            if not isinstance(visit_date, dt.date):
                continue
            key = (pid, visit_date)
            group = groups.setdefault(key, {
                "pid": pid,
                "name": name,
                "date": visit_date,
                "amount": 0.0,
                "match_kinds": set(),
                "source_files": set(),
                "source_rows": [],
            })
            if not group["name"] and name:
                group["name"] = name
            group["amount"] += float(amount or 0)
            if match_kind:
                group["match_kinds"].add(str(match_kind))
            if source_file:
                group["source_files"].add(str(source_file))
            if source_row:
                group["source_rows"].append(str(source_row))

        visit_counter[f"matched_visit_count_{year}"] += len(groups)
        for group in sorted(groups.values(), key=lambda item: (item["pid"], item["date"])):
            aggregated[year_month].append([
                group["pid"],
                group["name"],
                group["date"],
                1,
                group["amount"],
                "+".join(sorted(group["match_kinds"])),
                "、".join(sorted(group["source_files"])),
                ",".join(group["source_rows"][:5]) + ("..." if len(group["source_rows"]) > 5 else ""),
            ])
    return aggregated, visit_counter


def write_fee_claims(path: Path, monthly_rows: Dict[Tuple[int, int], List[List[Any]]]) -> None:
    wb = Workbook()
    ws = wb.active
    wb.remove(ws)
    headers = ["身分證號", "姓名", "日期", "次數", "總額", "對應方式", "來源檔", "來源列"]
    for year, month in sorted(monthly_rows):
        title = f"{year}{month:02d}"
        ws = wb.create_sheet(title)
        ws.append(headers)
        for row in monthly_rows[(year, month)]:
            ws.append(row)
    if not wb.sheetnames:
        ws = wb.create_sheet("費用未對到")
        ws.append(headers)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()


def aggregate_unmatched_fees(unmatched: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    groups: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for item in unmatched:
        chart = normalize_chart(item.get("chart"))
        name = normalize_text(item.get("name"))
        key = (chart, normalize_name(name))
        if not chart and not key[1]:
            continue
        group = groups.setdefault(key, {
            "chart": chart,
            "name": name,
            "last_visit": None,
            "amount_114": 0.0,
            "amount_115": 0.0,
            "visit_dates_114": set(),
            "visit_dates_115": set(),
            "rows": 0,
        })
        if not group["name"] and name:
            group["name"] = name
        visit_date = item.get("date")
        if isinstance(visit_date, dt.date):
            existing = group.get("last_visit")
            if existing is None or visit_date > existing:
                group["last_visit"] = visit_date
            roc_year = visit_date.year - 1911
            if roc_year == 114:
                group["amount_114"] += float(item.get("amount") or 0)
                group["visit_dates_114"].add(visit_date)
            elif roc_year == 115:
                group["amount_115"] += float(item.get("amount") or 0)
                group["visit_dates_115"].add(visit_date)
        group["rows"] += 1
    result: List[Dict[str, Any]] = []
    for group in groups.values():
        group["count_114"] = len(group.pop("visit_dates_114"))
        group["count_115"] = len(group.pop("visit_dates_115"))
        result.append(group)
    return sorted(result, key=lambda g: (normalize_name(g.get("name")), g.get("chart") or ""))


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    if source_row < 1:
        return
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy.copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy.copy(src.alignment)


def set_workbook_zoom(wb: Any, scale: int = 75) -> None:
    for ws in wb.worksheets:
        ws.sheet_view.zoomScale = scale
        ws.sheet_view.zoomScaleNormal = scale


def append_unmatched_fees_to_member_total(
    output_path: Path,
    unmatched_groups: List[Dict[str, Any]],
    core: Any,
) -> int:
    if not unmatched_groups:
        return 0
    wb = load_workbook(output_path)
    try:
        sheet_name = core.Rules.SHEET_TARGET
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f"輸出檔缺少工作表：{sheet_name}")
        ws = wb[sheet_name]
        cols = core.detect_template_columns(ws, core.Rules.DATA_START_ROW)
        template_row = ws.max_row if ws.max_row >= core.Rules.DATA_START_ROW else core.Rules.DATA_START_ROW
        added = 0
        for group in unmatched_groups:
            row_no = ws.max_row + 1
            copy_row_style(ws, template_row, row_no)
            name = group.get("name") or f"病歷號{group.get('chart')}"
            ws.cell(row_no, cols["name"]).value = name
            ws.cell(row_no, cols["id"]).value = None
            if cols.get("last_visit") and group.get("last_visit"):
                ws.cell(row_no, cols["last_visit"]).value = group["last_visit"]
            if cols.get("m_count_114"):
                ws.cell(row_no, cols["m_count_114"]).value = group.get("count_114", 0)
            if cols.get("n_count_115"):
                ws.cell(row_no, cols["n_count_115"]).value = group.get("count_115", 0)
            if cols.get("r_amount_114"):
                ws.cell(row_no, cols["r_amount_114"]).value = group["amount_114"]
            if cols.get("s_amount_115"):
                ws.cell(row_no, cols["s_amount_115"]).value = group["amount_115"]
            if cols.get("note"):
                chart = group.get("chart") or ""
                ws.cell(row_no, cols["note"]).value = f"費用未對到ID，待補；病歷號：{chart}；明細列數：{int(group.get('rows') or 0)}"
            added += 1
        set_workbook_zoom(wb, 75)
        wb.save(output_path)
        return added
    finally:
        wb.close()


def prepare_clean_dir(
    source_dir: Path,
    tmp_root: Path,
) -> Tuple[Path, Dict[str, Any], List[Dict[str, Any]]]:
    clean_dir = tmp_root / f"{source_dir.name}_耀聖雲端清洗"
    clean_dir.mkdir(parents=True, exist_ok=True)

    for path in sorted(source_dir.iterdir()):
        if path.is_file() and is_source_workbook(path):
            shutil.copy2(path, clean_dir / path.name)

    name_to_ids, chart_to_ids, id_to_name = build_identity_indexes(source_dir)
    monthly_rows, unmatched, stats = read_fee_rows(source_dir, name_to_ids, chart_to_ids, id_to_name)
    monthly_rows, visit_counter = aggregate_matched_fee_visits(monthly_rows)
    stats.update(visit_counter)
    unmatched_groups = aggregate_unmatched_fees(unmatched)
    write_fee_claims(clean_dir / "耀聖雲端費用_標準.xlsx", monthly_rows)

    stats["name_keys"] = len(name_to_ids)
    stats["chart_keys"] = len(chart_to_ids)
    stats["claim_sheets"] = len(monthly_rows)
    stats["unmatched_groups"] = len(unmatched_groups)
    stats["unmatched_visit_count_114"] = sum(int(group.get("count_114") or 0) for group in unmatched_groups)
    stats["unmatched_visit_count_115"] = sum(int(group.get("count_115") or 0) for group in unmatched_groups)
    return clean_dir, stats, unmatched_groups


def run_common_core(clean_dir: Path) -> Tuple[Path, Any]:
    core = load_common_core()
    generated = Path(core.process_excel(str(clean_dir), str(find_template())))
    return generated, core


def process_excel(source_path: str) -> Tuple[Path, Dict[str, Any]]:
    source_dir = Path(source_path).expanduser().resolve()
    if not source_dir.is_dir():
        raise ValueError("請選擇耀聖雲端來源資料夾。")

    with tempfile.TemporaryDirectory(prefix="ys_cloud_") as tmp:
        clean_dir, stats, unmatched_groups = prepare_clean_dir(source_dir, Path(tmp))
        generated, core = run_common_core(clean_dir)
        stats["unmatched_added_to_total"] = append_unmatched_fees_to_member_total(
            generated,
            unmatched_groups,
            core,
        )
        out_dir = source_dir.parent
        out_dir.mkdir(parents=True, exist_ok=True)
        target = out_dir / generated.name
        if target.exists():
            stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
            target = out_dir / f"{generated.stem}_{stamp}{generated.suffix}"
        shutil.move(str(generated), target)
    return target, stats


def choose_source_dir() -> Optional[Path]:
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    try:
        selected = filedialog.askdirectory(title="選擇耀聖雲端來源資料夾")
    finally:
        root.destroy()
    return Path(selected) if selected else None


def open_file(path: Path) -> None:
    try:
        if sys.platform.startswith("darwin"):
            subprocess.call(("open", str(path)))
        elif os.name == "nt":
            os.startfile(str(path))  # type: ignore[attr-defined]
        else:
            subprocess.call(("xdg-open", str(path)))
    except Exception:
        pass


def show_message(output: Path, stats: Dict[str, Any]) -> None:
    try:
        import tkinter as tk
        from tkinter import messagebox

        root = tk.Tk()
        root.withdraw()
        try:
            msg = (
                f"正式 Excel 已產生：\n{output}\n\n"
                f"費用明細共有 {int(stats['raw_rows'])} 筆。\n"
                f"其中 {int(stats['matched_chart']) + int(stats['matched_name'])} 筆有對到 ID"
                f"（病歷號 {int(stats['matched_chart'])}、姓名唯一 {int(stats['matched_name'])}）。\n"
                f"另外 {int(stats['unmatched'])} 筆無法靠病歷號或姓名唯一對到 ID。\n"
                f"這 {int(stats['unmatched'])} 筆已彙總成 "
                f"{int(stats.get('unmatched_added_to_total', 0))} 個會員，放進正式 Excel 的會員總表，ID 欄留空。\n\n"
                f"看診次數：未提供正式次數檔時，已用費用明細反推，同一會員同一天算一次。"
            )
            messagebox.showinfo("完成", msg, parent=root)
        finally:
            root.destroy()
    except Exception:
        pass


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="耀聖雲端前置清洗與選會員輸出")
    parser.add_argument("source_dir", nargs="?", help="來源資料夾；未指定時開啟選擇視窗")
    parser.add_argument("--no-open", action="store_true", help="完成後不自動開啟 Excel")
    parser.add_argument("--no-dialog", action="store_true", help="完成後不顯示提示視窗")
    args = parser.parse_args(argv)

    source_dir = Path(args.source_dir).expanduser().resolve() if args.source_dir else choose_source_dir()
    if source_dir is None:
        print("未選擇來源資料夾，程式已取消。")
        return 0

    try:
        output, stats = process_excel(str(source_dir))
    except Exception as exc:
        print(f"錯誤：{exc}")
        return 2

    print(f"完成：{output}")
    print(
        "費用整理："
        f"檔案={int(stats['files'])}，原始列={int(stats['raw_rows'])}，"
        f"病歷號對到={int(stats['matched_chart'])}，姓名唯一對到={int(stats['matched_name'])}，"
        f"未對到={int(stats['unmatched'])}，"
        f"追加會員總表={int(stats.get('unmatched_added_to_total', 0))}。"
    )
    print(
        "費用反推次數："
        f"已對到ID 114={int(stats.get('matched_visit_count_114', 0))}，"
        f"115={int(stats.get('matched_visit_count_115', 0))}；"
        f"ID空白 114={int(stats.get('unmatched_visit_count_114', 0))}，"
        f"115={int(stats.get('unmatched_visit_count_115', 0))}。"
    )
    print(
        "對到費用："
        f"114={stats['matched_amount_114']:.0f}，115={stats['matched_amount_115']:.0f}；"
        "未對到費用："
        f"114={stats['unmatched_amount_114']:.0f}，115={stats['unmatched_amount_115']:.0f}。"
    )
    if not args.no_dialog:
        show_message(output, stats)
    if not args.no_open:
        open_file(output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
