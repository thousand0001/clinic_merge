# -*- coding: utf-8 -*-
"""
汐止衛生所展望前置清洗 + 通用主程式包裝

用途：
- 將「115自選會員.xlsx」轉成通用版可辨識的自選名單，並帶入生日
- 將「診斷代碼.xlsx」轉成通用版可辨識的主次診斷
- 將「114年家醫會員就診次數.xlsx」年度總表轉成 11401.csv，供 114 全年件數回填
- 避免「百分位計算」分頁被誤判為會員名單
"""

from __future__ import annotations

import csv
import importlib.util
import os
import shutil
import sys
import tempfile
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook, load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent


def _find_generic_script(script_dir: Path) -> Path:
    candidates = sorted(script_dir.glob("run_merge_通用_*.py"), reverse=True)
    if not candidates:
        raise RuntimeError("找不到通用程式 run_merge_通用_*.py")
    return candidates[0]


GENERIC_SCRIPT = _find_generic_script(SCRIPT_DIR)


def _load_generic_module():
    spec = importlib.util.spec_from_file_location("run_merge_generic", GENERIC_SCRIPT)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"無法載入通用程式：{GENERIC_SCRIPT}")
    mod = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = mod
    spec.loader.exec_module(mod)
    return mod


def _norm(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _header_index_map(header: List[object]) -> dict[str, int]:
    out: dict[str, int] = {}
    for idx, value in enumerate(header):
        key = _norm(value).replace(" ", "").replace("\n", "")
        if key:
            out[key] = idx
    return out


def _pick(row: List[object], mapping: dict[str, int], *candidates: str) -> str:
    for candidate in candidates:
        idx = mapping.get(candidate)
        if idx is not None and idx < len(row):
            text = _norm(row[idx])
            if text:
                return text
    return ""


def _pick_value(row: List[object], mapping: dict[str, int], *candidates: str) -> object:
    for candidate in candidates:
        idx = mapping.get(candidate)
        if idx is not None and idx < len(row):
            value = row[idx]
            if _norm(value):
                return value
    return None


def _disable_original(path: Path) -> None:
    if path.exists():
        path.rename(path.with_suffix(path.suffix + ".原始備份"))


def _convert_self_select(source_dir: Path) -> Optional[Path]:
    candidates = sorted(source_dir.glob("*自選會員*.xlsx"))
    if not candidates:
        return None

    src = candidates[0]
    wb = load_workbook(src, data_only=True, read_only=True)
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "自選名單"
    out_ws.append(["ID", "姓名", "生日"])

    try:
        ws = wb[wb.sheetnames[0]]
        mapping: dict[str, int] = {}
        for row in ws.iter_rows(values_only=True):
            values = list(row)
            current = _header_index_map(values)
            if any(k in current for k in ("身分證號", "身份證號", "ID")) and any(k in current for k in ("姓名", "會員姓名")):
                mapping = current
                continue

            if mapping:
                pid = _pick(values, mapping, "身分證號", "身份證號", "ID")
                name = _pick(values, mapping, "姓名", "會員姓名")
                bday = _pick_value(values, mapping, "生日", "出生日期", "出生年月日")
            else:
                pid = _norm(values[0] if len(values) > 0 else "")
                name = _norm(values[1] if len(values) > 1 else "")
                bday = values[2] if len(values) > 2 and _norm(values[2]) else None

            if pid and name:
                out_ws.append([pid, name, bday])
    finally:
        wb.close()

    if out_ws.max_row <= 1:
        return None

    out_path = source_dir / "汐止_自選名單_補正.xlsx"
    out_wb.save(out_path)
    _disable_original(src)
    return out_path


def _convert_diagnosis(source_dir: Path) -> Optional[Path]:
    candidates = sorted(source_dir.glob("*診斷代碼*.xlsx"))
    if not candidates:
        return None

    src = candidates[0]
    wb = load_workbook(src, data_only=True, read_only=True)
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "主次診斷"
    out_ws.append(["姓名", "身份證號", "生日", "最後就診日(日期)\n(以最新的日期為主)", "診斷代碼(病1,病23)"])

    try:
        ws = wb[wb.sheetnames[0]]
        header_row = None
        mapping: dict[str, int] = {}
        for row_no, row in enumerate(ws.iter_rows(values_only=True), start=1):
            current = _header_index_map(list(row))
            if any(k in current for k in ("身分證號", "身份證號", "ID")) and any(k in current for k in ("診斷碼", "診斷代碼")):
                header_row = row_no
                mapping = current
                break
        if header_row is None:
            return None

        latest_by_id: dict[str, dict[str, str]] = {}
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            values = list(row)
            pid = _pick(values, mapping, "身分證號", "身份證號", "ID")
            dx = _pick(values, mapping, "診斷碼", "診斷代碼")
            if not pid or not dx:
                continue
            visit_date = _pick(values, mapping, "看診日期", "日期")
            name = _pick(values, mapping, "患者姓名", "姓名", "會員姓名")
            bday = _pick(values, mapping, "生日", "出生日期", "出生年月日")
            existing = latest_by_id.get(pid)
            if existing is None or visit_date >= existing.get("date", ""):
                latest_by_id[pid] = {"name": name, "bday": bday, "date": visit_date, "dx": dx}

        for pid, rec in sorted(latest_by_id.items()):
            out_ws.append([rec["name"], pid, rec["bday"], rec["date"], rec["dx"]])
    finally:
        wb.close()

    if out_ws.max_row <= 1:
        return None

    out_path = source_dir / "汐止_主次診斷_補正.xlsx"
    out_wb.save(out_path)
    _disable_original(src)
    return out_path


def _convert_annual_114_counts(source_dir: Path) -> Optional[Path]:
    candidates = sorted(source_dir.glob("*114*就診次數*.xlsx"))
    if not candidates:
        return None

    src = candidates[0]
    wb = load_workbook(src, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = list(next(rows, ()))
        mapping = _header_index_map(header)
        if not any(k in mapping for k in ("身分證", "身分證號", "身份證號", "ID")) or "次數" not in mapping:
            return None

        out_path = source_dir / "11401.csv"
        with out_path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["ID", "次數", "申請金額"])
            for row in rows:
                values = list(row)
                pid = _pick(values, mapping, "身分證", "身分證號", "身份證號", "ID")
                count = _pick(values, mapping, "次數")
                if pid and count:
                    writer.writerow([pid, count, 0])
        return out_path
    finally:
        wb.close()


def _preclean(source_dir: Path) -> List[Path]:
    generated: List[Path] = []
    for item in (
        _convert_self_select(source_dir),
        _convert_diagnosis(source_dir),
        _convert_annual_114_counts(source_dir),
    ):
        if item is not None:
            generated.append(item)
    return generated


def process_excel(source_path: str, template_path: Optional[str] = None) -> str:
    generic = _load_generic_module()
    source_dir = Path(source_path).resolve()
    if not source_dir.is_dir():
        raise ValueError("請選擇資料夾。")

    template = template_path or generic._find_template(str(SCRIPT_DIR))
    if not os.path.exists(template):
        raise ValueError(f"找不到模板檔：{template}")

    class XizhiProfile(generic.ProcessingProfile):
        def canonical_source_sheet_name(self, sheet_name, file_path, single_sheet, src_ws=None):
            if "百分位" in str(sheet_name):
                return str(sheet_name)
            return super().canonical_source_sheet_name(sheet_name, file_path, single_sheet, src_ws)

    temp_root = Path(tempfile.mkdtemp(prefix="xizhi_clean_"))
    temp_source = temp_root / source_dir.name

    try:
        shutil.copytree(source_dir, temp_source)
        generated = _preclean(temp_source)
        print(f"已產生汐止清洗檔 {len(generated)} 個", flush=True)

        temp_output = Path(generic.process_excel(str(temp_source), template, profile=XizhiProfile()))
        final_output = source_dir.parent / temp_output.name
        if final_output.exists():
            final_output.unlink()
        shutil.move(str(temp_output), str(final_output))
        return str(final_output)
    finally:
        shutil.rmtree(temp_root, ignore_errors=True)


def main() -> None:
    generic = _load_generic_module()
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()

    src = filedialog.askdirectory(title="選擇汐止衛生所展望來源資料夾")
    if not src:
        return

    template = generic._find_template(str(SCRIPT_DIR))

    try:
        out = process_excel(src, template)
        messagebox.showinfo("完成", f"已輸出：\n{out}")
        generic.open_file_cross_platform(out)
    except Exception as e:
        messagebox.showerror("錯誤", str(e))


if __name__ == "__main__":
    main()
