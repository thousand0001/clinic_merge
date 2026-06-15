# -*- coding: utf-8 -*-
"""
選會員產檔工具 — 共用核心 0610

功能概要
- 整合所有來源 sheet 人員名單（會員名單 / ascvd / 115指定會員 / 自選名單 / 115X / P4P / 月份分頁）
- 套用 0325 樣板，寫入「會員指標」分頁，同步輸出：
  - 百分位名單（LDL / HbA1c 百分位）
  - 醫生看
  - 自選名單
- 填入欄位：疾病樣態 / 主次診斷 / 分數 / 備註 / 追蹤提醒 / 電話分流 / 隱藏輔助欄
- KPI：HbA1c（BG/BH）、LDL（BJ/BK）百分位

注意
- Python 3.9 相容；不拆模組，單一檔案維護
- V4_2：改為直接選「資料夾」，自動掃描資料夾內所有 Excel / CSV 並合併分析
- 模板：自動選取檔名日期最新的 選會員模板*.xlsx
"""

from __future__ import annotations

import datetime
import csv
import math
import os
import re
import subprocess
import sys
import tempfile
import traceback
import warnings
import zipfile
import xml.etree.ElementTree as ET
from copy import copy
from dataclasses import dataclass
from enum import Enum, auto
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Any, Tuple

import openpyxl
import xlrd
from odf.opendocument import load as load_ods_document
from odf.table import Table, TableRow, TableCell
from odf.text import P
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style, apply openpyxl's default",
    category=UserWarning,
)


# ============================================================
# KPI 儲存格地址（沿用既有樣板固定位置，直接寫死）
# BG = HbA1c 主, BH = HbA1c 73.8% 目標
# BJ = LDL 主,   BK = LDL 73.8% 目標
# ============================================================
# ============================================================
# 會員總表資料列固定樣式
# AM(39)~AO(41), AV(48), AW(49): 靠左、字型14
# 其餘欄位: 置中、字型24
# ============================================================
_LEFT_COLS  = frozenset([39, 40, 41, 48, 49])   # AM, AN, AO, AV, AW

_FONT_14    = Font(size=14)
_FONT_24    = Font(size=24)
_TZ_TW = datetime.timezone(datetime.timedelta(hours=8))  # 台灣時區，模組層級共用
_FILL_SCREENING_NOT_NEEDED = PatternFill(fill_type="solid", fgColor="FFF200")
_FILL_SCREENING_PENDING = PatternFill(fill_type="solid", fgColor="F4CCCC")
_FILL_SCREENING_EXCLUDED = PatternFill(fill_type="solid", fgColor="D9EAF7")
_FILL_NONE = PatternFill(fill_type=None)
_FILL_DOCTOR_LAB_FAIL = PatternFill(fill_type="solid", fgColor="FFC000")
_FILL_DOCTOR_DATE_DONE = PatternFill(fill_type="solid", fgColor="C6E0B4")
_BORDER_THIN_GRAY = Border(
    left=Side(style="thin", color="B7B7B7"),
    right=Side(style="thin", color="B7B7B7"),
    top=Side(style="thin", color="B7B7B7"),
    bottom=Side(style="thin", color="B7B7B7"),
)
_ALIGNMENT_CACHE: Dict[Tuple[str, bool], Alignment] = {}


def _get_alignment(horizontal: str, wrap_text: bool = False) -> Alignment:
    key = (horizontal, wrap_text)
    cached = _ALIGNMENT_CACHE.get(key)
    if cached is None:
        cached = Alignment(horizontal=horizontal, vertical="center", wrap_text=wrap_text)
        _ALIGNMENT_CACHE[key] = cached
    return cached


def _apply_member_row_style(ws, row: int, max_col: int) -> None:
    for c in range(1, max_col + 1):
        cell = ws.cell(row, c)
        if c in _LEFT_COLS:
            cell.font      = _FONT_14
        else:
            cell.font      = _FONT_24


# ============================================================
# 模板欄位 alias（標題偵測優先，不再依賴單一名稱）
# ============================================================
TEMPLATE_ALIASES: Dict[str, List[List[str]]] = {
    "clinic": [
        ["診所名稱或機構代碼"],
        ["診所名稱"],
        ["機構代碼"],
        ["院所名稱"],
    ],
    "name": [
        ["姓名"],
        ["個案姓名"],
        ["會員姓名"],
    ],
    "id": [
        ["身份證號碼"],
        ["身分證號碼"],
        ["身份證號"],
        ["身分證號"],
        ["身份証號"],
        ["身分証號"],
        ["ID"],
        ["會員身份證"],
        ["會員身分證"],
    ],
    "bday": [
        ["生日"],
        ["出生日期"],
        ["出生年月日"],
    ],
    "age": [
        ["年齡"],
        ["歲"],
    ],
    "tel": [
        ["電話"],
        ["聯絡電話"],
    ],
    "mobile": [
        ["手機"],
        ["行動電話"],
        ["手機號碼"],
    ],
    "cnt": [
        ["次數"],
        ["件數"],
        ["就診次數"],
    ],
    "sex": [
        ["性別"],
        ["男女"],
    ],
    "abc": [
        ["A/B/C"],
        ["ABC"],
        ["會員"],
    ],
    "dmk_code": [
        ["DM/CKD/DKD"],
        ["疾病樣態編號"],
        ["DM", "CKD", "DKD"],
        ["疾病樣態"],
    ],
    "ascvd": [
        ["ASCVD"],
    ],
    "last_visit": [
        ["最後就診日"],
        ["最近就診日"],
        ["最後看診日"],
    ],
}



# ============================================================
# 業務規則常數（每年調整這裡即可）
# ============================================================
class Rules:
    # 檢驗日期：哪些年份視為「有效」
    VALID_YEARS: Tuple[int, ...] = (2025, 2026)

    # 篩檢計分：哪一年的篩檢才給分
    SCREEN_YEAR: int = 2026

    # 2026 年：只要有檢查就給滿分（無條件給分）
    # 2025 年：需達標才給分
    YEAR_UNCONDITIONAL: int = 2026
    YEAR_VALUE_CHECK: int = 2025

    # 各項分數
    SCORE_HBA: int = 5
    SCORE_LDL: int = 5
    SCORE_ADULT: int = 6
    SCORE_PAP: int = 6
    SCORE_FLU: int = 4
    SCORE_FIT: int = 6
    SCORE_HEP: int = 6

    # 達標值（2025年）
    HBA_PASS_2025: float = 7.0
    LDL_PASS_2025: float = 120.0

    # KPI（HbA1c）
    HBA_CONTROL_THRESHOLD: float = 7.0  # AY8：HbA1c <= 7 的比例
    HBA_TARGET_PERCENT: float = 0.738   # AZ：要找 >=73.8% 的切點

    # KPI（LDL）
    LDL_KPI_THRESHOLDS: Tuple[float, float] = (100.0, 110.0)  # BB8/BC8
    LDL_TARGET_PERCENT: float = 0.738   # BC：要找 >=73.8% 的切點（與 HbA1c 目前相同，但獨立維護）

    # 回診追蹤
    AU_DAYS: int = 28
    AV_OFFSET_DAYS: int = 28
    AW_OFFSET_DAYS: int = 56

    # 模板設定
    TEMPLATE_GLOB: str = "選會員模板*.xlsx"
    SHEET_TARGET: str = "會員總表"
    DATA_START_ROW: int = 3

    # 月份申請統計輸出欄位
    # L=114全年件數, M=114實際申報總額(總額), N=115件數(有效月份總次數), O=115實際申報總額(總額)
    COL_114_COUNT: str = "L"
    COL_115_COUNT: str = "N"
    COL_114_AMOUNT: str = "M"
    COL_115_AMOUNT: str = "O"
    COL_115_COUNT_Q1_HIDDEN: str = "BC"  # 隱藏輔助欄：115年1-4月就診次數
    COL_114_COUNT_Q1_HIDDEN: str = "BG"  # 隱藏輔助欄：114年1-4月就診次數
    COL_ADDRESS_HIDDEN: str = "BF"    # 隱藏輔助欄：地址
    COL_IS_114: str = "AZ"
    COL_IS_SELF_SELECT: str = "BA"
    COL_IS_115X: str = "BB"
    COL_NOTE: str = "AY"





# ============================================================
# Enum / Dataclass
# ============================================================
class DiseaseCode(Enum):
    DM = 1
    CKD = 2
    DKD = 3   # DM + CKD
    OTHER = 4 # 非 DM/CKD，可能有 ASCVD


class AscvdCategory(Enum):
    NONE = auto()  # 無 ASCVD
    A = auto()     # 極高風險
    B = auto()     # 非常高風險


@dataclass
class MemberMeta:
    """每位會員的衍生資訊，供後續計算用"""
    row: int
    pid: str = ""
    bday: Optional[datetime.date] = None
    age: int = -1
    e_code: Optional[DiseaseCode] = None
    ascvd: AscvdCategory = AscvdCategory.NONE


@dataclass
class SourceContext:
    wb_src: Any
    sh_member: Any
    sh_ascvd: Any
    sh_health: Any
    sh_main_sub_dx: Any
    sh_phone: Any
    sh_self_select: Any
    sh_115x: Any
    sh_115_designated: Any
    sh_p4p_enroll: Any
    sh_p4p_track: Any
    screening_sheets: Dict[str, Any]
    claim_sums: Dict[str, Dict[str, float]]
    claim_months_115: List[int]
    all_members: Dict[str, Dict[str, Any]]
    p4p_map: Dict[str, Dict[str, Any]]
    designated_115_source_count: int
    designated_115_ids: set[str]
    designated_115_details: Dict[str, Dict[str, Any]]


@dataclass
class MonthlyClaimSheetScan:
    sheet_name: str
    header_row: int
    id_col: int
    date_col: Optional[int]
    count_col: int
    amount_col: int


@dataclass
class HisbCountSheetScan:
    sheet_name: str
    header_row: int
    id_col: Optional[int]
    name_col: int
    bday_col: Optional[int]
    count_col: int
    date_col: Optional[int]
    amount_col: Optional[int]
    year_bucket: int
    month: int


@dataclass
class SourceSheetScanCache:
    monthly_claim_sheets: Dict[str, MonthlyClaimSheetScan]
    hisb_count_sheets: Dict[str, HisbCountSheetScan]
    partial_maps: Dict[str, Dict[str, Dict[str, Any]]]


class ProcessingProfile:
    """可由診所前置器覆寫的資料讀取與辨識規則。"""

    def parse_date(self, value: Any) -> Optional[datetime.date]:
        return parse_date(value)

    def load_source_workbook(self, source_path: str):
        ext = os.path.splitext(source_path)[1].lower()
        if ext == ".csv":
            return _load_csv_as_workbook(source_path)
        if ext == ".ods":
            return _load_ods_as_workbook(source_path)
        if ext == ".xls":
            return self.load_xls_as_workbook(source_path)
        return _load_xlsx_as_workbook(source_path)

    def load_xls_as_workbook(self, xls_path: str):
        return _load_xls_as_workbook(xls_path)

    def canonical_source_sheet_name(
        self,
        sheet_name: str,
        file_path: str,
        single_sheet: bool,
        src_ws: Any = None,
    ) -> str:
        return _canonical_source_sheet_name(sheet_name, file_path, single_sheet, src_ws)

    def sheet_year_bucket(self, title: str) -> Optional[int]:
        return _sheet_year_bucket(title)

    def sheet_month(self, title: str) -> Optional[int]:
        return _sheet_month(title)

    def find_monthly_claim_header_row(self, sheet: Any, search_rows: int = 30) -> Optional[int]:
        return _find_monthly_claim_header_row(sheet, search_rows=search_rows)

    def scan_monthly_claim_sheet(self, sheet_name: str, sheet: Any) -> Optional[MonthlyClaimSheetScan]:
        return _scan_monthly_claim_sheet(sheet_name, sheet)

    def scan_hisb_count_sheet(self, sheet_name: str, sheet: Any) -> Optional[HisbCountSheetScan]:
        return _scan_hisb_count_sheet(sheet_name, sheet, profile=self)

    def collect_monthly_claim_summaries(
        self,
        wb_src: Any,
        monthly_scans: Optional[Dict[str, MonthlyClaimSheetScan]] = None,
    ) -> Tuple[Dict[str, Dict[str, float]], List[int]]:
        return collect_monthly_claim_summaries(wb_src, monthly_scans=monthly_scans, profile=self)


DEFAULT_PROFILE = ProcessingProfile()


def _resolve_profile(profile: Optional[ProcessingProfile]) -> ProcessingProfile:
    return profile or DEFAULT_PROFILE


def _find_template(script_dir: str) -> str:
    def template_date_key(path: Path) -> Tuple[int, str]:
        match = re.search(r"(\d{4})(?=\.xlsx$)", path.name, re.IGNORECASE)
        return (int(match.group(1)) if match else -1, path.name)

    candidates = sorted(
        Path(script_dir).glob(Rules.TEMPLATE_GLOB),
        key=template_date_key,
        reverse=True,
    )
    if not candidates:
        raise RuntimeError(f"找不到模板檔 {Rules.TEMPLATE_GLOB}，請確認模板放在程式同資料夾。")
    return str(candidates[0])


@dataclass
class TemplateContext:
    wb_tpl: Any
    ws: Any
    cols: Dict[str, Optional[int]]
    data_start: int


@dataclass
class RuntimeContext:
    id_to_rows: Dict[str, List[int]]
    meta: Dict[int, MemberMeta]
    last_row: int
    hba_candidates: Optional[List[Tuple[int, float]]] = None
    ldl_candidates: Optional[List[Tuple[int, float]]] = None
    hba_main_summary: str = ""
    hba_target_summary: str = ""
    ldl_main_summary: str = ""
    ldl_target_summary: str = ""


@dataclass
class ContactInfo:
    phone: Optional[str] = None
    mobile: Optional[str] = None
    address: Optional[str] = None


# ============================================================
# 工具函數
# ============================================================
def open_file_cross_platform(path: str) -> None:
    """跨平台開啟檔案（盡量不拋錯）"""
    try:
        if sys.platform.startswith("darwin"):
            subprocess.call(("open", path))
        elif os.name == "nt":
            os.startfile(path)  # type: ignore[attr-defined]
        else:
            subprocess.call(("xdg-open", path))
    except Exception:
        pass


def normalize_text(v: Any) -> str:
    if v is None:
        return ""
    return str(v).replace("\t", "").replace("　", "").strip().lstrip("'")


def normalize_header(v: Any) -> str:
    s = normalize_text(v)
    return s.replace(" ", "").replace("\n", "")


def clean_spaces(v: Any) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", "", str(v))


def is_phone_header(text: str) -> bool:
    s = normalize_header(text)
    return ("電話" in s) or ("手機" in s) or ("行動電話" in s)


def is_address_header(text: str) -> bool:
    s = normalize_header(text)
    return any(k in s for k in ("地址", "住址", "聯絡地址", "會員地址", "戶籍地址"))


def normalize_phone_value(v: Any) -> Optional[str]:
    if v is None:
        return None
    digits = re.sub(r"\D+", "", str(v))
    if len(digits) == 9 and digits.startswith(("9", "2")):
        digits = f"0{digits}"
    elif len(digits) == 8:
        digits = f"02{digits}"
    if len(digits) < 8:
        return None
    return digits


def pick_contact_from_values(values: List[Any]) -> ContactInfo:
    info = ContactInfo()
    for raw in values:
        phone = normalize_phone_value(raw)
        if not phone:
            continue
        if phone.startswith("09"):
            if not info.mobile:
                info.mobile = phone
        elif not info.phone:
            info.phone = phone
    return info


def merge_contact_info(primary: Optional[ContactInfo], secondary: Optional[ContactInfo]) -> ContactInfo:
    p = primary or ContactInfo()
    s = secondary or ContactInfo()
    return ContactInfo(
        phone=p.phone or s.phone,
        mobile=p.mobile or s.mobile,
        address=p.address or s.address,
    )


HIS_MONTHLY_FEE_FOLDER_NAME = "醫聖月份費用xlsx"


def safe_set(ws, row: int, col: Optional[int], value: Any) -> None:
    if col:
        ws.cell(row, col).value = value


def safe_set_check(ws, row: int, col: Optional[int], value: Any) -> None:
    """寫入打勾符號"""
    if col:
        cell = ws.cell(row, col)
        normalized_value = "✔" if value == "v" else value
        cell.value = normalized_value


def calc_age(bday: Optional[datetime.date], ref: datetime.date) -> int:
    if not isinstance(bday, datetime.date):
        return -1
    y = ref.year - bday.year
    if (ref.month, ref.day) < (bday.month, bday.day):
        y -= 1
    return y


def add_months(d: datetime.date, months: int) -> datetime.date:
    y = d.year + (d.month - 1 + months) // 12
    m = (d.month - 1 + months) % 12 + 1
    is_leap = (y % 4 == 0 and (y % 100 != 0 or y % 400 == 0))
    mdays = [31, 29 if is_leap else 28, 31, 30, 31, 30,
             31, 31, 30, 31, 30, 31][m - 1]
    return datetime.date(y, m, min(d.day, mdays))


def infer_gender_from_id(id_value: Any) -> str:
    """由台灣身分證/居留證第2碼推性別（1/8/A/C→男，2/9/B/D→女）"""
    if not id_value:
        return ""
    s = str(id_value).strip().upper()
    if len(s) < 2:
        return ""
    g = s[1]
    if g in ("1", "8", "A", "C"):
        return "男"
    if g in ("2", "9", "B", "D"):
        return "女"
    return ""


# ============================================================
# 解析函數
# ============================================================
def parse_date(value: Any) -> Optional[datetime.date]:
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value

    s = str(value).strip()
    if not s or s in ("-", "—", "–"):
        return None

    patterns = [
        (r"^(\d{4})[/-](\d{1,2})[/-](\d{1,2})$", False),
        (r"^(\d{2,3})(\d{2})(\d{2})$", True),
        (r"^(\d{2,3})[/-](\d{1,2})[/-](\d{1,2})$", True),
        (r"^(\d{2,3})\.(\d{1,2})\.(\d{1,2})$", True),
    ]
    for pat, is_roc in patterns:
        m = re.match(pat, s)
        if m:
            a, b, c = map(int, m.groups())
            year = a + 1911 if is_roc else a
            try:
                return datetime.date(year, b, c)
            except ValueError:
                return None
    return None


def parse_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    s = str(value).strip().lstrip("'").replace(",", "")
    if not s or s in ("-", "—", "–"):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def parse_disease_code(v: Any) -> Optional[DiseaseCode]:
    s = clean_spaces(v)
    if not s:
        return None
    if re.fullmatch(r"[1-4]\.0+", s):
        s = s.split(".", 1)[0]
    if s.isdigit():
        mapping = {
            1: DiseaseCode.DM,
            2: DiseaseCode.CKD,
            3: DiseaseCode.DKD,
            4: DiseaseCode.OTHER,
        }
        return mapping.get(int(s))
    text_map = {
        "DM":     DiseaseCode.DM,
        "CKD":    DiseaseCode.CKD,
        "DKD":    DiseaseCode.DKD,
        "DM+CKD": DiseaseCode.DKD,
        "DMCKD":  DiseaseCode.DKD,
    }
    return text_map.get(s.upper())




# ============================================================
# 主次診斷代碼（DM / CKD）
# ============================================================
DM_CODE_PREFIXES: Tuple[str, ...] = ("E08", "E09", "E10", "E11", "E12", "E13")

# 依使用者提供之「主次診斷代碼.docx」整理
CKD_CODES = {
    'A18.11',
    'A52.75',
    'C64.1',
    'C64.2',
    'C64.9',
    'C7A.093',
    'D59.3',
    'E08.21',
    'E08.22',
    'E08.29',
    'E08.65',
    'E09.21',
    'E09.22',
    'E09.29',
    'E09.65',
    'E10.21',
    'E10.22',
    'E10.29',
    'E10.65',
    'E12.21',
    'E12.22',
    'E12.29',
    'E12.65',
    'E11.21',
    'E11.22',
    'E11.29',
    'E11.65',
    'E13.21',
    'E13.22',
    'E13.29',
    'E13.65',
    'E74.8',
    'I70.1',
    'I72.2',
    'I75.81',
    'I77.3',
    'I77.73',
    'K76.7',
    'M10.30',
    'M10.311',
    'M10.312',
    'M10.319',
    'M10.321',
    'M10.322',
    'M10.329',
    'M10.331',
    'M10.332',
    'M10.339',
    'M10.341',
    'M10.342',
    'M10.349',
    'M10.351',
    'M10.352',
    'M10.359',
    'M10.361',
    'M10.362',
    'M10.369',
    'M10.371',
    'M10.372',
    'M10.379',
    'M10.38',
    'M10.39',
    'N00.0',
    'N00.1',
    'N00.2',
    'N00.3',
    'N00.4',
    'N00.5',
    'N00.6',
    'N00.7',
    'N00.8',
    'N00.9',
    'N01.0',
    'N01.1',
    'N01.2',
    'N01.3',
    'N01.4',
    'N01.5',
    'N01.6',
    'N01.7',
    'N01.8',
    'N01.9',
    'N02.0',
    'N02.1',
    'N02.2',
    'N02.3',
    'N02.4',
    'N02.5',
    'N02.6',
    'N02.7',
    'N02.8',
    'N02.9',
    'N03.0',
    'N03.1',
    'N03.2',
    'N03.3',
    'N03.4',
    'N03.5',
    'N03.6',
    'N03.7',
    'N03.8',
    'N03.9',
    'N04.0',
    'N04.1',
    'N04.2',
    'N04.3',
    'N04.4',
    'N04.5',
    'N04.6',
    'N04.7',
    'N04.8',
    'N04.9',
    'N05.0',
    'N05.1',
    'N05.2',
    'N05.3',
    'N05.4',
    'N05.5',
    'N05.6',
    'N05.7',
    'N05.8',
    'N05.9',
    'N06.0',
    'N06.1',
    'N06.2',
    'N06.3',
    'N06.4',
    'N06.5',
    'N06.6',
    'N06.7',
    'N06.8',
    'N06.9',
    'N07.0',
    'N07.1',
    'N07.2',
    'N07.3',
    'N07.4',
    'N07.5',
    'N07.6',
    'N07.7',
    'N07.8',
    'N07.9',
    'N08',
    'N14.0',
    'N14.1',
    'N14.2',
    'N14.3',
    'N14.4',
    'N15.0',
    'N15.8',
    'N15.9',
    'N16',
    'N17.0',
    'N17.1',
    'N17.2',
    'N17.8',
    'N17.9',
    'N18.1',
    'N18.2',
    'N18.3',
    'N20.0',
    'N25.0',
    'N25.1',
    'N25.81',
    'N25.89',
    'N25.9',
    'N26.9',
    'O10.211',
    'O10.212',
    'O10.213',
    'O10.219',
    'O10.22',
    'O10.23',
    'O10.311',
    'O10.312',
    'O10.313',
    'O10.319',
    'O10.32',
    'O10.33',
    'O10.411',
    'O10.412',
    'O10.413',
    'O10.419',
    'O10.42',
    'O10.43',
    'Q61.01',
    'Q61.02',
    'Q61.11',
    'Q61.19',
    'Q61.2',
    'Q61.3',
    'Q61.4',
    'Q61.5',
    'Q61.8',
    'Q62.0',
    'Q62.10',
    'Q62.11',
    'Q62.12',
    'Q62.2',
    'Q62.31',
    'Q62.32',
    'Q62.39',
    'R94.4'
}


def normalize_id(v: Any) -> str:
    return normalize_text(v).upper()


_ID_RE = re.compile(r'^[A-Z]{1,2}\d{8,9}$')

def is_valid_tw_id(v: Any) -> bool:
    """台灣身分證/居留證格式：1英+9數 或 2英+8數"""
    s = normalize_text(v).upper().replace(" ", "")
    return bool(_ID_RE.match(s))


def find_id_col_by_content(sheet, header_row: int, id_col_candidate: Optional[int]) -> Optional[int]:
    """
    先靠 header 找到候選 ID 欄，再往下掃內容驗證格式。
    若 header 找不到則逐欄掃所有欄位內容。
    """
    def _col_has_valid_id(col: int) -> bool:
        for r in range(header_row + 1, min(header_row + 30, sheet.max_row + 1)):
            v = sheet.cell(r, col).value
            if v and is_valid_tw_id(v):
                return True
        return False

    if id_col_candidate and _col_has_valid_id(id_col_candidate):
        return id_col_candidate

    # fallback：只接受身分證類欄位、會員ID，或欄位名精確等於 ID
    for c in range(1, sheet.max_column + 1):
        hdr = normalize_header(sheet.cell(header_row, c).value)
        hdr_upper = hdr.upper()
        is_id_header = (
            ("身分" in hdr) or
            ("身份" in hdr) or
            ("會員ID" in hdr_upper) or
            (hdr_upper == "ID")
        )
        if is_id_header and _col_has_valid_id(c):
            return c
    return None


def _first_valid_id_row(sheet, id_col: Optional[int], header_row: int) -> int:
    if id_col is None:
        return header_row + 1
    for r in range(header_row + 1, sheet.max_row + 1):
        if is_valid_tw_id(sheet.cell(r, id_col).value):
            return r
    return header_row + 1


def normalize_icd_code(v: Any) -> str:
    s = normalize_text(v).upper().replace(" ", "")
    if not s:
        return ""
    # 容忍 E8.9 / E9 / N8 這種少一個 0 的寫法
    s = re.sub(r"^([A-Z])(\d)(?=\.|$)", r"\g<1>0\g<2>", s)
    return s


def extract_icd_codes_from_cell(v: Any) -> List[str]:
    s = normalize_text(v).upper()
    if not s:
        return []
    # 正確使用 word boundary，擷取像 E11.21 / N18.3 / N08 這類 ICD 代碼
    # ICD-10 格式：標準(E11.21/N18.3/J069) + 特殊(C7A.093)
    found = re.findall(r"\b[A-Z]\d{2,3}(?:\.\d+)?\b|\b[A-Z]\d[A-Z](?:\.\d+)?\b", s)
    out: List[str] = []
    seen = set()
    for raw in found:
        code = normalize_icd_code(raw)
        if code and code not in seen:
            seen.add(code)
            out.append(code)
    return out

def is_dm_icd_code(code: str) -> bool:
    c = normalize_icd_code(code)
    return any(c.startswith(p) for p in DM_CODE_PREFIXES)


_DM_CKD_RENAL_RE = re.compile(r'^E(08|09|10|11|12|13)\.2')

def is_ckd_icd_code(code: str) -> bool:
    c = normalize_icd_code(code)
    # E08~E13 小數點後第一位為 2 → 腎臟併發症，直接算 CKD
    if _DM_CKD_RENAL_RE.match(c):
        return True
    return c in CKD_CODES


def classify_main_sub_dx(codes: List[str]) -> str:
    has_dm = any(is_dm_icd_code(c) for c in codes)
    has_ckd = any(is_ckd_icd_code(c) for c in codes)
    if has_dm and has_ckd:
        return "DM+CKD"
    if has_dm:
        return "DM"
    if has_ckd:
        return "CKD"
    return ""


def _build_main_sub_dx_map(sh_dx: Any) -> Dict[str, str]:
    id_aliases = [
        "ID", "id", "身份證號", "身份證號碼", "身分證號", "身分證號碼",
        "會員身份証", "會員身份證", "會員身分證",
    ]
    # 只用 id_aliases 找 header row
    header_row = _find_header_row_contains_any(sh_dx, [id_aliases], search_rows=20)
    if header_row is None:
        raise ValueError("原始檔「主次診斷」找不到 ID 欄位")

    hmap = build_header_map(sh_dx, header_row)
    id_col = find_id_col_by_content(sh_dx, header_row, find_column_exact(hmap, id_aliases))
    if id_col is None:
        raise ValueError("原始檔「主次診斷」找不到 ID 欄位")

    # 包含以下任一關鍵字即視為診斷碼欄
    _DX_KW = ["主診斷", "次診斷", "診斷碼", "診斷代碼", "診斷", "代碼", "病1", "病23", "icd", "疾病碼"]
    dx_cols: List[int] = []
    for hdr, c in hmap.items():
        if c == id_col:
            continue
        if any(kw.lower() in hdr.lower() for kw in _DX_KW):
            dx_cols.append(c)
    dx_cols = sorted(set(dx_cols))
    if not dx_cols:
        raise ValueError("原始檔「主次診斷」找不到診斷碼相關欄位")

    # 找最後就診日欄（可選）—— 模糊比對「就診日」
    last_visit_col = next(
        (c for hdr, c in hmap.items() if "就診日" in hdr and c != id_col),
        None
    )

    id_to_codes: Dict[str, List[str]] = {}
    id_to_last_visit: Dict[str, Optional[datetime.date]] = {}
    for r in range(header_row + 1, sh_dx.max_row + 1):
        pid = normalize_id(sh_dx.cell(r, id_col).value)
        if not pid:
            continue
        codes_here: List[str] = []
        for c in dx_cols:
            codes_here.extend(extract_icd_codes_from_cell(sh_dx.cell(r, c).value))
        if codes_here:
            bucket = id_to_codes.setdefault(pid, [])
            for code in codes_here:
                if code not in bucket:
                    bucket.append(code)
        # 最後就診日（取最新）
        if last_visit_col:
            dt = parse_date(sh_dx.cell(r, last_visit_col).value)
            if dt:
                existing = id_to_last_visit.get(pid)
                if existing is None or dt > existing:
                    id_to_last_visit[pid] = dt

    result: Dict[str, Dict[str, Any]] = {}
    all_pids = set(id_to_codes.keys()) | set(id_to_last_visit.keys())
    for pid in all_pids:
        codes   = id_to_codes.get(pid, [])
        label   = classify_main_sub_dx(codes)
        raw_str = ",".join(codes)
        lv      = id_to_last_visit.get(pid)
        if label or raw_str or lv:
            result[pid] = {"label": label, "raw": raw_str, "last_visit": lv}
    return result


def _build_main_sub_dx_map_from_monthly_sheets(wb_src: Any) -> Dict[str, Dict[str, Any]]:
    id_aliases = [
        "ID", "id", "身份證號", "身份證號碼", "身分證號", "身分證號碼",
        "會員身份証", "會員身份證", "會員身分證",
    ]
    result_codes: Dict[str, List[str]] = {}
    result_last_visit: Dict[str, datetime.date] = {}

    for sheet_name in wb_src.sheetnames:
        sh = wb_src[sheet_name]
        header_row = _find_header_row_contains_any(
            sh,
            [id_aliases, ["日期"], ["病1"]],
            search_rows=30,
        )
        if header_row is None:
            continue

        hmap = build_header_map(sh, header_row)
        id_col = find_id_col_by_content(sh, header_row, find_column_exact(hmap, id_aliases))
        date_col = find_column_exact(hmap, ["日期"])
        dx_cols = [c for hdr, c in hmap.items() if hdr in ("病1", "病23")]
        if id_col is None or date_col is None or not dx_cols:
            continue

        for r in range(header_row + 1, sh.max_row + 1):
            pid = normalize_id(sh.cell(r, id_col).value)
            if not pid:
                continue

            codes_here: List[str] = []
            for c in dx_cols:
                codes_here.extend(extract_icd_codes_from_cell(sh.cell(r, c).value))
            if codes_here:
                bucket = result_codes.setdefault(pid, [])
                for code in codes_here:
                    if code not in bucket:
                        bucket.append(code)

            dt = parse_date(sh.cell(r, date_col).value)
            if dt:
                existing = result_last_visit.get(pid)
                if existing is None or dt > existing:
                    result_last_visit[pid] = dt

    result: Dict[str, Dict[str, Any]] = {}
    all_pids = set(result_codes.keys()) | set(result_last_visit.keys())
    for pid in all_pids:
        codes = result_codes.get(pid, [])
        label = classify_main_sub_dx(codes)
        raw_str = ",".join(codes)
        lv = result_last_visit.get(pid)
        if label or raw_str or lv:
            result[pid] = {"label": label, "raw": raw_str, "last_visit": lv}
    return result


def _fill_main_sub_dx(
    ws,
    sh_dx: Any,
    wb_src: Any,
    cols: Dict[str, Optional[int]],
    id_to_rows: Dict[str, List[int]],
) -> None:
    dx_map = _build_main_sub_dx_map(sh_dx) if sh_dx is not None else _build_main_sub_dx_map_from_monthly_sheets(wb_src)
    col_ad = cols.get("main_sub_dx")   # AD：分類結果
    col_q  = cols.get("dx_raw")        # Q：原始診斷碼
    col_l  = cols.get("last_visit")    # L：最後就診日

    for pid, rows in id_to_rows.items():
        info = dx_map.get(normalize_id(pid))
        if not info:
            continue
        for rr in rows:
            if col_ad and info.get("label"):
                safe_set(ws, rr, col_ad, info["label"])
            if col_q and info.get("raw"):
                safe_set(ws, rr, col_q, info["raw"])
            # 最後就診日：只在目前空白時才填（ascvd 已填過的優先）
            if col_l and info.get("last_visit"):
                existing = parse_date(ws.cell(rr, col_l).value)
                new_dt   = info["last_visit"]
                if existing is None or new_dt > existing:
                    safe_set(ws, rr, col_l, new_dt)

def parse_ascvd(v: Any) -> AscvdCategory:
    if v is None:
        return AscvdCategory.NONE
    s = clean_spaces(v).lower()
    if not s or s == "0":
        return AscvdCategory.NONE
    if s in ("a", "1") or "極高" in s:
        return AscvdCategory.A
    if s in ("b", "2") or "非常高" in s:
        return AscvdCategory.B
    if "a" in s:
        return AscvdCategory.A
    if "b" in s:
        return AscvdCategory.B
    return AscvdCategory.NONE


def is_zero_like(v: Any) -> bool:
    """把 None / 空白 / '0' 視為零值"""
    if v is None:
        return True
    return str(v).strip() in ("", "0", "0.0")


# ============================================================
# 月份申請統計（114xx / 115xx）
# ============================================================
def _find_header_row_contains_any(sheet, alias_groups: List[List[str]], search_rows: int = 30) -> Optional[int]:
    # 正規化 alias 一次，避免每列每組重複計算
    norm_groups = [[a.replace(" ", "").replace("\n", "") for a in grp] for grp in alias_groups]
    for r in range(1, min(search_rows, sheet.max_row) + 1):
        row_headers = {normalize_header(sheet.cell(r, c).value) for c in range(1, sheet.max_column + 1)}
        if all(any(a in row_headers for a in grp) for grp in norm_groups):
            return r
    return None


def _sheet_year_bucket(title: str) -> Optional[int]:
    s = str(title).strip()
    if not re.fullmatch(r"\d{5}", s):
        return None
    if s.startswith("114"):
        return 114
    if s.startswith("115"):
        return 115
    return None


def _sheet_month(title: str) -> Optional[int]:
    """從分頁名稱取月份數字（11401→1, 11412→12），不符合則 None"""
    s = str(title).strip()
    if not re.fullmatch(r"\d{5}", s):
        return None
    return int(s[3:5])


def _roc_year_from_date(dt: Optional[datetime.date]) -> Optional[int]:
    if dt is None:
        return None
    return dt.year - 1911


def _find_monthly_claim_header_row(sheet, search_rows: int = 30) -> Optional[int]:
    """
    月報表頭可能只有一列，也可能混有多列說明。
    這裡採用和主流程接近的邏輯：
    1. 先找出疑似 ID 欄位
    2. 再驗證下一段資料列中是否有合法身分證字號
    3. 同時確認該列能找到次數 / 金額欄
       日期欄若缺少，但工作表名稱本身是 11401/11502 這類月份名稱，也視為可用
    """
    id_aliases = ["ID", "身分證號", "身分證號碼", "身份證號", "身份證號碼", "身分證字號", "身份證字號"]
    date_aliases = ["日期", "最後看診日期", "最後就診日", "最後看診日"]
    amount_aliases = ["申報總金額", "總金額", "總額", "申請金額"]
    count_aliases = ["次數", "件數"]
    has_sheet_month = (_sheet_year_bucket(sheet.title) in (114, 115) and _sheet_month(sheet.title) is not None)

    for r in range(1, min(search_rows, sheet.max_row) + 1):
        hmap = build_header_map(sheet, r)
        id_col_cand = find_column_exact(hmap, id_aliases)
        id_col = find_id_col_by_content(sheet, r, id_col_cand)
        if id_col is None:
            continue

        date_col = find_column_exact(hmap, date_aliases)
        count_col = find_column_exact(hmap, count_aliases)
        amount_col = None
        for alias in amount_aliases:
            amount_col = find_col_by_keywords_any_row(sheet, r, [alias])
            if amount_col:
                break
        if count_col and amount_col and (date_col or has_sheet_month):
            return r

    return None


def _find_hisb_count_header_row(sheet, search_rows: int = 10) -> Optional[int]:
    """
    HIS B 次數 CSV 格式範例：
    病歷號 / 姓名 / 生日 / 電話 / 次數 / 地址
    或：
    病歷號 / 姓名 / 性別 / 身分證號 / 電話 / 次數
    有身分證號時直接比對，沒有時才退回姓名+生日。
    """
    for r in range(1, min(search_rows, sheet.max_row) + 1):
        hmap = build_header_map(sheet, r)
        chart_col = find_column_exact(hmap, ["病歷號", "病歷號碼"])
        name_col = find_column_exact(hmap, ["姓名", "病患姓名", "會員姓名"])
        id_col = find_column_exact(hmap, ["ID", "身分證號", "身分證號碼", "身份證號", "身份證號碼"])
        bday_col = find_bday_column(hmap)
        count_col = find_column_exact(hmap, ["次數", "就診次數", "門診次數", "來診次數"])
        if chart_col and name_col and count_col and (bday_col or id_col):
            return r
    return None


def _scan_monthly_claim_sheet(sheet_name: str, sheet: Any) -> Optional[MonthlyClaimSheetScan]:
    id_aliases = ["ID", "身分證號", "身分證號碼", "身份證號", "身份證號碼", "身分證字號", "身份證字號"]
    date_aliases = ["日期", "最後看診日期", "最後就診日", "最後看診日"]
    amount_aliases = ["申報總金額", "總金額", "總額", "申請金額"]
    count_aliases = ["次數", "件數"]
    header_row = _find_monthly_claim_header_row(sheet, search_rows=30)
    if header_row is None:
        return None
    hmap = build_header_map(sheet, header_row)
    id_col = find_id_col_by_content(sheet, header_row, find_column_exact(hmap, id_aliases))
    date_col = find_column_exact(hmap, date_aliases)
    count_col = find_column_exact(hmap, count_aliases)
    amount_col = None
    for alias in amount_aliases:
        amount_col = find_col_by_keywords_any_row(sheet, header_row, [alias])
        if amount_col:
            break
    has_sheet_month = (_sheet_year_bucket(sheet_name) in (114, 115) and _sheet_month(sheet_name) is not None)
    if id_col is None or count_col is None or amount_col is None or (date_col is None and not has_sheet_month):
        return None
    return MonthlyClaimSheetScan(
        sheet_name=sheet_name,
        header_row=header_row,
        id_col=id_col,
        date_col=date_col,
        count_col=count_col,
        amount_col=amount_col,
    )


def _scan_hisb_count_sheet(
    sheet_name: str,
    sheet: Any,
    profile: Optional[ProcessingProfile] = None,
) -> Optional[HisbCountSheetScan]:
    profile = _resolve_profile(profile)
    year_bucket = profile.sheet_year_bucket(sheet_name)
    month = profile.sheet_month(sheet_name)
    if year_bucket not in (114, 115) or month is None:
        return None
    if profile.find_monthly_claim_header_row(sheet, search_rows=10) is not None:
        return None
    header_row = _find_hisb_count_header_row(sheet, search_rows=10)
    if header_row is None:
        return None
    hmap = build_header_map(sheet, header_row)
    id_col = find_column_exact(hmap, ["ID", "身分證號", "身分證號碼", "身份證號", "身份證號碼"])
    name_col = find_column_exact(hmap, ["姓名", "病患姓名", "會員姓名"])
    bday_col = find_bday_column(hmap)
    count_col = find_column_exact(hmap, ["次數", "就診次數", "門診次數", "來診次數"])
    date_col = find_column_exact(hmap, ["日期", "最後看診日期", "最後就診日", "最後看診日", "最後回診日"])
    amount_col = find_column_exact(hmap, ["申報總金額", "總金額", "總額", "申請金額"])
    if not name_col or not count_col or (not id_col and not bday_col):
        return None
    return HisbCountSheetScan(
        sheet_name=sheet_name,
        header_row=header_row,
        id_col=id_col,
        name_col=name_col,
        bday_col=bday_col,
        count_col=count_col,
        date_col=date_col,
        amount_col=amount_col,
        year_bucket=year_bucket,
        month=month,
    )


def _scan_source_sheets(
    wb_src,
    profile: Optional[ProcessingProfile] = None,
) -> SourceSheetScanCache:
    profile = _resolve_profile(profile)
    id_aliases = ["身份證號", "身份證號碼", "身分證號", "身分證號碼", "身份証號", "身分証號",
                  "會員身份証", "會員身份證", "會員身分證", "ID", "家醫收案會員ID"]
    monthly_claim_sheets: Dict[str, MonthlyClaimSheetScan] = {}
    hisb_count_sheets: Dict[str, HisbCountSheetScan] = {}
    partial_maps: Dict[str, Dict[str, Dict[str, Any]]] = {}

    for sheet_name in wb_src.sheetnames:
        sh = wb_src[sheet_name]
        monthly_scan = profile.scan_monthly_claim_sheet(sheet_name, sh)
        if monthly_scan is not None:
            monthly_claim_sheets[sheet_name] = monthly_scan
            partial_maps[sheet_name] = _extract_member_partial_map(sh, id_aliases, search_rows=5)
            continue

        hisb_scan = profile.scan_hisb_count_sheet(sheet_name, sh)
        if hisb_scan is not None:
            hisb_count_sheets[sheet_name] = hisb_scan

        if sheet_name != "會員名單":
            partial_maps[sheet_name] = _extract_member_partial_map(sh, id_aliases)

    return SourceSheetScanCache(
        monthly_claim_sheets=monthly_claim_sheets,
        hisb_count_sheets=hisb_count_sheets,
        partial_maps=partial_maps,
    )


def _empty_claim_bucket() -> Dict[str, float]:
    return {
        "114_cnt": 0.0,
        "114_cnt_full": 0.0,
        "115_cnt": 0.0,
        "115_cnt_q1": 0.0,
        "114_amt": 0.0,
        "115_amt": 0.0,
        "114_amt_total": 0.0,
        "115_amt_total": 0.0,
        "115_months": 0.0,
        "last_visit_ord": 0.0,
    }


def _claim_month_key(year: int, field: str, month: int) -> str:
    return f"{year}_{field}_m{month:02d}"


def _refresh_effective_month_claim_values(
    claim_sums: Dict[str, Dict[str, float]],
    months_115: Iterable[int],
) -> None:
    months = sorted({month for month in months_115 if 1 <= month <= 12})
    month_count = float(len(months))
    for bucket in claim_sums.values():
        bucket["114_cnt"] = sum(
            bucket.get(_claim_month_key(114, "cnt", month), 0.0)
            for month in months
        )
        bucket["115_cnt_q1"] = sum(
            bucket.get(_claim_month_key(115, "cnt", month), 0.0)
            for month in months
        )
        bucket["114_amt"] = sum(
            bucket.get(_claim_month_key(114, "amt", month), 0.0)
            for month in months
        )
        bucket["115_amt"] = sum(
            bucket.get(_claim_month_key(115, "amt", month), 0.0)
            for month in months
        )
        bucket["115_months"] = month_count


def collect_monthly_claim_summaries(
    wb_src,
    monthly_scans: Optional[Dict[str, MonthlyClaimSheetScan]] = None,
    profile: Optional[ProcessingProfile] = None,
) -> Tuple[Dict[str, Dict[str, float]], List[int]]:
    """
    掃描 R11440 類明細表，
    依 ID 彙總 R11440 明細格式：
      - 日期欄 -> 最後就診日
      - 次數欄 -> 114/115 就診次數
      - 總額欄 -> 114/115 實際申報總額
    可兼容月份分頁與單一總表，只要列內有日期可判斷年度與月份即可。
    若某 ID 完全沒有資料，後續保持空白，不填 0。
    """
    profile = _resolve_profile(profile)
    out: Dict[str, Dict[str, float]] = {}
    seen_115_months: set = set()
    scans = monthly_scans or {
        sheet_name: scan
        for sheet_name in wb_src.sheetnames
        if (scan := profile.scan_monthly_claim_sheet(sheet_name, wb_src[sheet_name])) is not None
    }

    for sheet_name, scan in scans.items():
        sh = wb_src[sheet_name]
        for r in range(scan.header_row + 1, sh.max_row + 1):
            pid_raw = sh.cell(r, scan.id_col).value
            pid = normalize_text(pid_raw).upper()
            if not pid or not is_valid_tw_id(pid):
                continue

            dt = profile.parse_date(sh.cell(r, scan.date_col).value) if scan.date_col else None
            cnt = parse_float(sh.cell(r, scan.count_col).value)
            amt = parse_float(sh.cell(r, scan.amount_col).value)
            if dt is None and cnt is None and amt is None:
                continue

            sheet_year_bucket = profile.sheet_year_bucket(sheet_name)
            sheet_month = profile.sheet_month(sheet_name)
            if sheet_year_bucket in (114, 115) and sheet_month is not None:
                # R11440 月份分頁的「日期」是最後就診日，不是申報年月。
                # 分頁名稱已明確標示年月時，必須優先依分頁歸類。
                year_bucket = sheet_year_bucket
                month = sheet_month
            else:
                year_bucket = _roc_year_from_date(dt)
                month = dt.month if dt is not None else None
            if year_bucket not in (114, 115):
                continue

            if month is None:
                continue

            has_claim_value = (cnt not in (None, 0)) or (amt not in (None, 0))
            if year_bucket == 115 and has_claim_value:
                seen_115_months.add(month)

            bucket = out.setdefault(pid, _empty_claim_bucket())

            if dt is not None:
                bucket["last_visit_ord"] = max(bucket.get("last_visit_ord", 0.0), float(dt.toordinal()))

            prefix = str(year_bucket)
            if cnt is not None:
                month_key = _claim_month_key(year_bucket, "cnt", month)
                bucket[month_key] = bucket.get(month_key, 0.0) + cnt
                if year_bucket == 115:
                    bucket["115_cnt"] += cnt
                if year_bucket == 114:
                    bucket["114_cnt_full"] += cnt
            if amt is not None:
                month_key = _claim_month_key(year_bucket, "amt", month)
                bucket[month_key] = bucket.get(month_key, 0.0) + amt
                bucket[f"{prefix}_amt_total"] += amt
    _refresh_effective_month_claim_values(out, seen_115_months)

    return out, sorted(seen_115_months)


def _supplement_claim_counts_from_hisb(
    wb_src,
    all_members: Dict[str, Dict[str, Any]],
    claim_sums: Dict[str, Dict[str, float]],
    claim_months_115: List[int],
    hisb_scans: Optional[Dict[str, HisbCountSheetScan]] = None,
    profile: Optional[ProcessingProfile] = None,
) -> Tuple[Dict[str, Dict[str, float]], List[int]]:
    """
    補吃 HIS B 的「次數 CSV」：
    - 來源通常是 11402.CSV / 11501.CSV 這類檔名
    - 只有次數，沒有身份證號/日期/金額
    - 用 姓名 + 生日 唯一比對到會員 ID 後，將次數累加到月份統計
    """
    profile = _resolve_profile(profile)
    member_key_to_ids: Dict[Tuple[str, datetime.date], List[str]] = {}
    for pid, info in all_members.items():
        name = normalize_text(info.get("name"))
        bday = info.get("bday")
        if not name or not isinstance(bday, datetime.date):
            continue
        member_key_to_ids.setdefault((name, bday), []).append(pid)

    seen_115_months = set(claim_months_115)

    scans = hisb_scans or {
        sheet_name: scan
        for sheet_name in wb_src.sheetnames
        if (scan := profile.scan_hisb_count_sheet(sheet_name, wb_src[sheet_name])) is not None
    }

    for sheet_name, scan in scans.items():
        sh = wb_src[sheet_name]
        has_any_data = False
        matched_any_115 = False
        for r in range(scan.header_row + 1, sh.max_row + 1):
            cnt = parse_float(sh.cell(r, scan.count_col).value)
            dt = profile.parse_date(sh.cell(r, scan.date_col).value) if scan.date_col else None
            amt = parse_float(sh.cell(r, scan.amount_col).value) if scan.amount_col else None
            if cnt is None and dt is None and amt is None:
                continue
            has_any_data = True

            matched_ids: List[str] = []
            if scan.id_col:
                pid = normalize_id(sh.cell(r, scan.id_col).value)
                if pid and is_valid_tw_id(pid):
                    matched_ids = [pid]

            if not matched_ids and scan.bday_col:
                name = normalize_text(sh.cell(r, scan.name_col).value)
                bday = profile.parse_date(sh.cell(r, scan.bday_col).value)
                if not name or not isinstance(bday, datetime.date):
                    continue
                matched_ids = member_key_to_ids.get((name, bday), [])

            if len(matched_ids) != 1:
                continue

            pid = matched_ids[0]
            bucket = claim_sums.setdefault(pid, _empty_claim_bucket())

            if dt is not None:
                bucket["last_visit_ord"] = max(bucket.get("last_visit_ord", 0.0), float(dt.toordinal()))

            if cnt is not None:
                month_key = _claim_month_key(scan.year_bucket, "cnt", scan.month)
                bucket[month_key] = bucket.get(month_key, 0.0) + cnt
                if scan.year_bucket == 115:
                    bucket["115_cnt"] += cnt
                else:
                    bucket["114_cnt_full"] += cnt

            if amt is not None:
                prefix = str(scan.year_bucket)
                month_key = _claim_month_key(scan.year_bucket, "amt", scan.month)
                bucket[month_key] = bucket.get(month_key, 0.0) + amt
                bucket[f"{prefix}_amt_total"] += amt

            if scan.year_bucket == 115 and ((cnt not in (None, 0)) or (amt not in (None, 0))):
                matched_any_115 = True

        if scan.year_bucket == 115 and matched_any_115:
            seen_115_months.add(scan.month)

    _refresh_effective_month_claim_values(claim_sums, seen_115_months)

    return claim_sums, sorted(seen_115_months)


def _to_excel_number(v: Optional[float]) -> Optional[float]:
    if v is None:
        return None
    if abs(v - round(v)) < 1e-9:
        return int(round(v))
    return round(v, 2)


def _to_excel_int(v: Optional[float]) -> Optional[int]:
    if v is None:
        return None
    return int(round(v))


def _fmt_percent(v: float, denom: int) -> str:
    if denom <= 0:
        return "0.00%"
    return f"{v * 100:.2f}%"


def fill_monthly_claim_summary_columns(
    ws,
    data_start: int,
    last_row: int,
    cols: Dict[str, Optional[int]],
    claim_sums: Dict[str, Dict[str, float]],
    meta: Optional[Dict[int, MemberMeta]] = None,
) -> None:
    col_last = cols.get("last_visit")            # K：最後就診日
    col_m    = cols.get("m_count_114")            # L：114全年件數
    col_n_q1 = cols.get("m_count_114_q1")         # BG：114件數（1-4月，輔助）
    col_o_q1 = cols.get("n_count_115_q1")         # BC：115件數（1-4月，輔助）
    col_o    = cols.get("n_count_115")            # N：115件數（有效月份總次數）
    col_s    = cols.get("r_amount_114")           # M：114實際申報總額（總額）
    col_t    = cols.get("s_amount_115")           # O：115實際申報總額（總額）
    col_bd   = cols.get("avg_amount_114_hidden")  # BD：114年月平均（隱藏輔助欄）
    col_be   = cols.get("avg_amount_115_hidden")  # BE：115年月平均（隱藏輔助欄）
    if not all([col_m, col_o, col_s, col_t, cols.get("id")]):
        raise ValueError("模板找不到 L/M/N/O 或 ID 欄位，無法填入月份申請統計")

    for rr in range(data_start, last_row + 1):
        pid = ""
        if meta is not None:
            pid = normalize_text(meta.get(rr, MemberMeta(row=rr)).pid).upper()
        if not pid:
            pid = normalize_text(ws.cell(rr, cols["id"]).value).upper()  # type: ignore[index]
        data = claim_sums.get(pid)

        if not data:
            ws.cell(rr, col_m).value = None
            if col_n_q1:
                ws.cell(rr, col_n_q1).value = None
            if col_o_q1:
                ws.cell(rr, col_o_q1).value = None
            ws.cell(rr, col_o).value = None
            ws.cell(rr, col_s).value = None
            ws.cell(rr, col_t).value = None
            if col_bd:
                ws.cell(rr, col_bd).value = None
            if col_be:
                ws.cell(rr, col_be).value = None
            continue

        v114c    = data.get("114_cnt", 0.0)
        v114c_fy = data.get("114_cnt_full", 0.0)
        v115c_q1 = data.get("115_cnt_q1", 0.0)
        v115c    = data.get("115_cnt", 0.0)
        v114a    = data.get("114_amt", 0.0)
        v115a    = data.get("115_amt", 0.0)
        v114a_total = data.get("114_amt_total", 0.0)
        v115a_total = data.get("115_amt_total", 0.0)
        v115_months = max(int(data.get("115_months", 0.0)), 1)
        last_visit_ord = int(data.get("last_visit_ord", 0.0) or 0)

        if col_last and last_visit_ord > 0:
            monthly_last_visit = datetime.date.fromordinal(last_visit_ord)
            existing_last_visit = parse_date(ws.cell(rr, col_last).value)
            if existing_last_visit is None or monthly_last_visit > existing_last_visit:
                ws.cell(rr, col_last).value = monthly_last_visit
        ws.cell(rr, col_m).value = _to_excel_number(v114c_fy) if v114c_fy != 0 else None
        if col_n_q1:
            ws.cell(rr, col_n_q1).value = _to_excel_number(v114c) if v114c != 0 else None
        if col_o_q1:
            ws.cell(rr, col_o_q1).value = _to_excel_number(v115c_q1) if v115c_q1 != 0 else None
        ws.cell(rr, col_o).value = _to_excel_number(v115c) if v115c != 0 else None
        ws.cell(rr, col_s).value = _to_excel_int(v114a_total) if v114a_total != 0 else None
        ws.cell(rr, col_t).value = _to_excel_int(v115a_total) if v115a_total != 0 else None
        if col_bd:
            avg_114 = (v114a_total / 12.0) if v114a_total != 0 else None
            ws.cell(rr, col_bd).value = _to_excel_int(avg_114)
        if col_be:
            avg_115 = (v115a / float(v115_months)) if v115a != 0 else None
            ws.cell(rr, col_be).value = _to_excel_int(avg_115)


# ============================================================
# 篩檢需求判斷
# ============================================================
def adult_check_interval_years(age: int, e_code: Optional[DiseaseCode] = None) -> Optional[int]:
    if e_code in (DiseaseCode.DM, DiseaseCode.CKD, DiseaseCode.DKD):
        return None
    if 30 <= age <= 39:
        return 5
    if 40 <= age <= 64:
        return 3
    if age >= 65:
        return 1
    return None


def adult_screening_status(
    age: int,
    e_code: Optional[DiseaseCode] = None,
) -> Tuple[bool, Optional[str], PatternFill]:
    """
    成人健檢狀態單一規則來源：
    - DM/CKD/DKD：不需檢查，且顯示排除個案文字
    - 其餘：依 adult_check_interval_years() 判斷是否需要
    """
    interval = adult_check_interval_years(age, e_code)
    if interval is None:
        if e_code in (DiseaseCode.DM, DiseaseCode.CKD, DiseaseCode.DKD):
            return False, "為DM/CKD/DKD排除個案", _FILL_SCREENING_EXCLUDED
        return False, "不需受檢", _FILL_NONE
    return True, "待受檢", _FILL_SCREENING_PENDING


def pap_check_interval_years(age: int, sex: str) -> Optional[int]:
    if sex != "女":
        return None
    if 25 <= age <= 29:
        return 3
    if age >= 30:
        return 1
    return None


def need_flu(age: int) -> bool:
    return age >= 65


def need_fit(age: int) -> bool:
    return 45 <= age <= 75


def need_bc_hep(age: int) -> bool:
    return 45 <= age < 80


# ============================================================
# 欄位偵測
# ============================================================
def build_header_map(sheet, header_row: int) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for c in range(1, sheet.max_column + 1):
        k = normalize_header(sheet.cell(header_row, c).value)
        if k:
            m[k] = c
    return m


def find_column_exact(hmap: Dict[str, int], aliases: List[str]) -> Optional[int]:
    for a in aliases:
        if a in hmap:
            return hmap[a]
    return None


def find_bday_column(hmap: Dict[str, int]) -> Optional[int]:
    for hdr, col in hmap.items():
        hdr_upper = hdr.upper()
        if "生日" in hdr or "出生" in hdr or "BIRTHDAY" in hdr_upper:
            return col
    return None


def _get_merged_header_text(ws, header_row: int, c: int) -> str:
    upper = normalize_header(ws.cell(header_row - 1, c).value) if header_row > 1 else ""
    lower = normalize_header(ws.cell(header_row, c).value)
    return (upper + " " + lower).strip()


def find_col_by_keywords(ws, header_row: int, keywords: List[str]) -> Optional[int]:
    """模糊比對：合併上下兩列表頭，同時包含所有 keywords"""
    keys = [k.replace(" ", "").replace("\n", "") for k in keywords if k]
    for c in range(1, ws.max_column + 1):
        h = _get_merged_header_text(ws, header_row, c).replace(" ", "")
        if h and all(k in h for k in keys):
            return c
    return None


def find_col_by_keywords_any_row(ws, max_row: int, keywords: List[str]) -> Optional[int]:
    """跨多列找欄位：同欄所有儲存格合併後包含全部 keywords"""
    keys = [k.replace(" ", "").replace("\n", "") for k in keywords if k]
    for c in range(1, ws.max_column + 1):
        blob = "".join(
            normalize_header(ws.cell(r, c).value)
            for r in range(1, max_row + 1)
        )
        if blob and all(k in blob for k in keys):
            return c
    return None


def find_header_row_contains(sheet, must_have: List[str], search_rows: int = 250) -> Optional[int]:
    must = [m.replace(" ", "").replace("\n", "") for m in must_have]
    for r in range(1, min(search_rows, sheet.max_row) + 1):
        found = set()
        for c in range(1, sheet.max_column + 1):
            v = normalize_header(sheet.cell(r, c).value)
            if v in must:
                found.add(v)
        if all(m in found for m in must):
            return r
    return None


def find_col_by_alias_groups(ws, header_row: int, alias_groups: List[List[str]]) -> Optional[int]:
    """
    依序嘗試多組 alias。
    每組 alias 內的字串需同時出現在同一欄標題中；先命中的先用。
    """
    for group in alias_groups:
        col = find_col_by_keywords(ws, header_row, group)
        if col:
            return col
    return None


def find_header_row_contains_alias_groups(
    ws,
    required_fields: List[str],
    search_rows: int = 250,
) -> Optional[int]:
    """
    以 alias 規則找最像模板主表頭的列。
    至少要命中所有 required_fields 對應欄位。
    """
    max_r = min(search_rows, ws.max_row)
    for r in range(1, max_r + 1):
        ok = True
        for field in required_fields:
            alias_groups = TEMPLATE_ALIASES.get(field, [])
            if not alias_groups:
                ok = False
                break
            if not find_col_by_alias_groups(ws, r, alias_groups):
                ok = False
                break
        if ok:
            return r
    return None


def _detect_basic_cols(ws, header_row: int) -> Dict[str, Optional[int]]:
    alias = find_col_by_alias_groups
    return {
        "clinic":     alias(ws, header_row, TEMPLATE_ALIASES["clinic"]),
        "name":       alias(ws, header_row, TEMPLATE_ALIASES["name"]),
        "id":         alias(ws, header_row, TEMPLATE_ALIASES["id"]),
        "bday":       alias(ws, header_row, TEMPLATE_ALIASES["bday"]),
        "age":        alias(ws, header_row, TEMPLATE_ALIASES["age"]),
        "tel":        alias(ws, header_row, TEMPLATE_ALIASES["tel"]),
        "mobile":     alias(ws, header_row, TEMPLATE_ALIASES["mobile"]),
        "cnt":        alias(ws, header_row, TEMPLATE_ALIASES["cnt"]),
        "sex":        alias(ws, header_row, TEMPLATE_ALIASES["sex"]),
        "abc":        alias(ws, header_row, TEMPLATE_ALIASES["abc"]),
        "dmk_code":   alias(ws, header_row, TEMPLATE_ALIASES["dmk_code"]),
        "ascvd":      alias(ws, header_row, TEMPLATE_ALIASES["ascvd"]),
        "last_visit": alias(ws, header_row, TEMPLATE_ALIASES["last_visit"]),
    }


def _detect_screening_cols(ws, header_row: int) -> Dict[str, Optional[int]]:
    """
    5大篩檢欄：只依表頭關鍵字偵測，不依賴固定欄位位置。
    """
    kw     = find_col_by_keywords
    kw_any = find_col_by_keywords_any_row
    max_scan = header_row + 3  # 多掃幾列以防標題跨列

    def _find_screening(keywords: List[str]) -> Optional[int]:
        return (
            kw(ws, header_row, keywords)
            or kw_any(ws, max_scan, keywords)
        )

    return {
        "adult": _find_screening(["成人", "健檢"]),
        "pap":   _find_screening(["子宮", "抹片"]),
        "flu":   _find_screening(["老人", "流感"]),
        "fit":   _find_screening(["糞便", "潛血"]),
        "hep":   _find_screening(["BC肝炎"]),
    }


def _detect_lab_cols(ws, header_row: int) -> Dict[str, Optional[int]]:
    kw = find_col_by_keywords
    return {
        "hba":     kw(ws, header_row, ["HbA1c"]),
        "hba_dt":  kw(ws, header_row, ["HbA1c", "日期"]),
        "ldl":     kw(ws, header_row, ["LDL"]),
        "ldl_dt":  kw(ws, header_row, ["LDL", "日期"]),
        "uacr":    kw(ws, header_row, ["UACR"]),
        "uacr_dt": kw(ws, header_row, ["UACR", "日期"]),
    }


def _detect_output_cols(ws, header_row: int, max_scan_row: int) -> Dict[str, Optional[int]]:
    kw = find_col_by_keywords
    kw_any = find_col_by_keywords_any_row
    return {
        "disease_text": (kw(ws, header_row, ["DM/CKD/DKD/ASCVD"])
                         or kw(ws, header_row, ["疾病樣態", "ASCVD"])
                         or kw(ws, header_row, ["疾病樣態", "分類"])),
        "score": (kw(ws, header_row, ["分數"])),
        "breakdown": (kw(ws, header_row, ["分數說明"])
                        or kw_any(ws, max_scan_row, ["分數說明"])),
        "note":  (kw(ws, header_row, ["預防保健提醒"])
                  or kw(ws, header_row, ["備註"])),
        # AK：合併大區塊，用 1~data_start-1 掃描
        "ak":    (kw_any(ws, max_scan_row, ["打勾"])
                  or kw_any(ws, max_scan_row, ["HbA1c", "打勾"])
                  or kw(ws, header_row, ["打勾"])
                  or kw(ws, header_row, ["HbA1c", "合格"])),
        "ldl_pass": (kw(ws, header_row, ["LDL", "合格"])
                     or kw_any(ws, max_scan_row, ["LDL", "合格"])),
        "uacr_pass": (kw(ws, header_row, ["UACR", "合格"])
                      or kw_any(ws, max_scan_row, ["UACR", "合格"])),
        "metabolic_enroll": (kw(ws, header_row, ["可納入新陳代謝收案"])
                             or kw_any(ws, max_scan_row, ["可納入新陳代謝收案"])),
        # AX：漏檢項目
        "ax":    (kw_any(ws, max_scan_row, ["漏檢項目"])
                  or kw_any(ws, max_scan_row, ["漏檢"])
                  or kw(ws, header_row, ["漏檢項目"])
                  or kw(ws, header_row, ["漏檢"])),
        "main_sub_dx": (kw(ws, header_row, ["主次診斷", "分類"])
                        or kw(ws, header_row, ["主次診斷"])
                        or kw(ws, header_row, ["DM/CKD"])),
        "dx_raw":      (kw(ws, header_row, ["病123"])
                        or kw(ws, header_row, ["英文字照抄"])),
        # 後段輔助欄（旗標 / 統計 / 隱藏地址）
        "p4p_plan":       kw(ws, header_row, ["P4P收案計畫"]),
        "p4p_status":     (kw(ws, header_row, ["P4P收案狀態"])
                           or kw(ws, header_row, ["收案狀態"])),
        "p4p_enroll_dt":  kw(ws, header_row, ["收案日期"]),
        "p4p_last_dt":    kw(ws, header_row, ["最後追蹤日"]),
        "p4p_next_dt":    (kw(ws, header_row, ["下次應追蹤日"])
                           or kw(ws, header_row, ["下次追蹤日"])),
        "p4p_overdue":    kw(ws, header_row, ["逾期未追蹤"]),
        "is_114":         kw(ws, header_row, ["是否為114會員名單"]),
        "is_self_select": kw(ws, header_row, ["是否為自選會員"]),
        "is_115x":        kw(ws, header_row, ["是否為115X"]),
        "m_count_114":    kw(ws, header_row, ["114年", "就診次數"]),
        "m_count_114_q1": kw(ws, header_row, ["114年1-4月就診次數"]),
        "n_count_115_q1": kw(ws, header_row, ["115年1-4月就診次數"]),
        "n_count_115":    kw(ws, header_row, ["115年", "就診次數"]),
        "r_amount_114":   (kw(ws, header_row, ["114年", "實際申報總額"])
                           or kw(ws, header_row, ["114年", "申報總額"])
                           or kw(ws, header_row, ["114年", "申報金額", "月"])),
        "s_amount_115":   (kw(ws, header_row, ["115年", "實際申報總額"])
                           or kw(ws, header_row, ["115年", "申報總額"])
                           or kw(ws, header_row, ["115年", "申報金額", "月"])),
        "avg_amount_114_hidden": (kw(ws, header_row, ["114年月平均"])
                                  or kw(ws, header_row, ["114年", "月平均"])),
        "avg_amount_115_hidden": (kw(ws, header_row, ["115年月平均"])
                                  or kw(ws, header_row, ["115年", "月平均"])),
        "address_hidden": kw(ws, header_row, ["地址"]),
    }


def _detect_kpi_col(ws, max_scan_row: int, keywords: List[str]) -> Optional[int]:
    """找 KPI 百分位標記欄，只依表頭文字判斷。"""
    keys = [normalize_header(k).lower() for k in keywords if k]
    for c in range(1, ws.max_column + 1):
        blob = "".join(
            normalize_header(ws.cell(r, c).value)
            for r in range(1, max_scan_row + 1)
        ).lower()
        if blob and all(k in blob for k in keys):
            return c
    return None


def _detect_followup_cols(ws, max_scan_row: int) -> Dict[str, Optional[int]]:
    kw = find_col_by_keywords_any_row
    return {
        "au": (kw(ws, max_scan_row, ["第一次提醒", "28天"])
               or kw(ws, max_scan_row, ["回診追蹤", "28天"])
               or kw(ws, max_scan_row, ["28天"])),
        "av": (kw(ws, max_scan_row, ["第二次提醒", "56天"])
               or kw(ws, max_scan_row, ["回診追蹤", "56天"])),
        "aw": (kw(ws, max_scan_row, ["第三次提醒", "84天"])
               or kw(ws, max_scan_row, ["回診追蹤", "84天"])),
    }


def detect_template_columns(ws, data_start: int) -> Dict[str, Optional[int]]:
    header_row = (
        find_header_row_contains_alias_groups(ws, ["name", "id"], 250)
        or find_header_row_contains(ws, ["姓名", "身份證號碼"], 250)
        or 1
    )
    max_scan_row = data_start - 1

    cols: Dict[str, Optional[int]] = {}
    cols.update(_detect_basic_cols(ws, header_row))
    cols.update(_detect_screening_cols(ws, header_row))
    cols.update(_detect_lab_cols(ws, header_row))

    followup_cols = _detect_followup_cols(ws, max_scan_row)
    cols.update(followup_cols)
    cols.update(_detect_output_cols(ws, header_row, max_scan_row))

    # KPI 標記欄（依表頭文字偵測）
    cols["ay_mark"] = (
        _detect_kpi_col(ws, max_scan_row, ["HbA1c<=7"])
        or _detect_kpi_col(ws, max_scan_row, ["HAb1c<=7"])
    )
    cols["az_mark"] = (
        _detect_kpi_col(ws, max_scan_row, ["HbA1c百分位", "73.8"])
        or _detect_kpi_col(ws, max_scan_row, ["HAb1c百分位", "73.8"])
    )
    cols["bb_mark"] = _detect_kpi_col(ws, max_scan_row, ["LDL百分位<=100"])
    cols["bc_mark"] = _detect_kpi_col(ws, max_scan_row, ["LDL百分位", "73.8"])

    # 必填欄位驗證：集中在最後檢查，避免偵測流程夾雜太多錯誤處理
    _require_cols(cols, [
        "name", "id", "bday", "tel", "abc",
        "mobile",
        "dmk_code", "ascvd", "sex",
        "hba", "hba_dt", "ldl", "ldl_dt", "uacr", "uacr_dt",
        "disease_text", "score", "breakdown", "note",
        "au", "av", "aw",
        "ak", "ldl_pass", "uacr_pass", "ax",
        "m_count_114", "n_count_115", "r_amount_114", "s_amount_115",
        "ay_mark", "az_mark", "bb_mark", "bc_mark",
    ])
    return cols


def _require_cols(cols: Dict[str, Optional[int]], required: List[str]) -> None:
    missing = [k for k in required if not cols.get(k)]
    if missing:
        raise ValueError(f"新模板欄位找不到：{'、'.join(missing)}")


# ============================================================
# 業務邏輯：疾病樣態文字（✅ DM+CKD 後面不加 (DKD)）
# ============================================================
def disease_group_text(
    e_code: Optional[DiseaseCode], ascvd: AscvdCategory
) -> Optional[str]:
    has_dm    = e_code in (DiseaseCode.DM, DiseaseCode.DKD)
    has_ckd   = e_code in (DiseaseCode.CKD, DiseaseCode.DKD)
    has_ascvd = ascvd != AscvdCategory.NONE

    table = {
        (True,  False, False): "DM",
        (False, True,  False): "CKD",
        (True,  True,  False): "DM+CKD",
        (False, False, True):  "ASCVD",
        (False, True,  True):  "CKD+ASCVD",
        (True,  False, True):  "DM+ASCVD",
        (True,  True,  True):  "DM+CKD+ASCVD",
    }
    return table.get((has_dm, has_ckd, has_ascvd))


def _is_normal_person_by_group(
    e_code: Optional[DiseaseCode],
    ascvd: AscvdCategory,
) -> bool:
    """
    正常人判定規則：
    - 以總表「疾病樣態分類(7類)」是否空白為準
    - 只要 disease_group_text(...) 會回傳空白，就視為正常人
    - 不再單獨依 e_code=4、ASCVD=0 文字值另外判斷

    正常人合格標準：
    - HbA1c <= 7.0
    - LDL < 100
    """
    return disease_group_text(e_code, ascvd) is None


# ============================================================
# 業務邏輯：AK（HbA1c 打勾 ✔）
# ============================================================
def should_check_ak(
    *,
    e_code: Optional[DiseaseCode],
    ascvd: AscvdCategory,
    age: int,
    hba_val: Any,
) -> bool:
    """
    AK 打勾規則（✔）：
    - 所有含 DM 的組合：
        - 年齡 >= 80：HbA1c < 8
        - 其餘：HbA1c < 7
    - 正常人（疾病樣態分類空白）：HbA1c <= 7.0
    其他疾病：不勾
    """
    has_dm  = e_code in (DiseaseCode.DM, DiseaseCode.DKD)
    is_normal = _is_normal_person_by_group(e_code, ascvd)

    if not has_dm and not is_normal:
        return False

    v = parse_float(hba_val)
    if v is None:
        return False

    if is_normal:
        return v <= 7.0

    if age < 0:
        return False

    if age >= 80:
        return v < 8.0
    return v < 7.0


def should_check_ldl_pass(
    *,
    e_code: Optional[DiseaseCode],
    ascvd_raw: Any,
    ldl_val: Any,
) -> bool:
    """
    LDL 合格規則：
    - 正常人（疾病樣態分類空白）：LDL <= 130
    - ASCVD a：LDL < 55
    - ASCVD b：LDL < 70
    - DM / DKD：LDL < 100
    - CKD：LDL < 130
    """
    v = parse_float(ldl_val)
    if v is None:
        return False

    has_dm = e_code in (DiseaseCode.DM, DiseaseCode.DKD)
    has_ckd = e_code in (DiseaseCode.CKD, DiseaseCode.DKD)
    token = _ascvd_token(ascvd_raw)
    is_normal = _is_normal_person_by_group(e_code, parse_ascvd(ascvd_raw))

    if token == "a":
        return v < 55.0
    if token == "b":
        return v < 70.0
    if is_normal:
        return v < 130.0
    if has_dm:
        return v < 100.0
    if has_ckd:
        return v < 130.0
    return False


def should_check_uacr_pass(
    *,
    e_code: Optional[DiseaseCode],
    uacr_val: Any,
) -> bool:
    if e_code not in (DiseaseCode.CKD, DiseaseCode.DKD):
        return False
    v = parse_float(uacr_val)
    if v is None:
        return False
    return v < 30.0


# ============================================================
# 業務邏輯：AX（漏檢項目）
# ============================================================
def _ascvd_token(v: Any) -> str:
    """AX 用：ASCVD 欄只接受 0/a/b（0 也算有值）"""
    s = clean_spaces(v).lower()
    if s in ("0", "a", "b"):
        return s
    return ""


def build_ax_leak_item(
    *,
    d_code: Optional[DiseaseCode],  # D欄：1/2/3/4
    ascvd_raw: Any,                 # E欄：0/a/b
    hba_dt: Optional[datetime.date],  # AD：HbA1c 檢查日期（程式內用 hba_dt）
    ldl_dt: Optional[datetime.date],  # AF：LDL 檢查日期（程式內用 ldl_dt）
) -> Optional[str]:
    """
    條件（同時成立）：
    - D(疾病樣態)=1/2/3/4
    - E(ascvd)=0/a/b（0 也算有值）
    - 排除：D=4 且 E=0
    - 只有 D=1 或 D=3 判斷漏檢
    - HbA1c日期 與 LDL日期 皆有值且不同天

    顯示：
    - HbA1c日 > LDL日 → LDL漏檢
    - LDL日 > HbA1c日 → HbA1c漏檢
    - 相同日 → 不顯示
    """
    if d_code not in (DiseaseCode.DM, DiseaseCode.CKD, DiseaseCode.DKD, DiseaseCode.OTHER):
        return None

    token = _ascvd_token(ascvd_raw)
    if not token:  # 必須是 0/a/b
        return None

    # 排除：D=4 且 E=0
    if d_code == DiseaseCode.OTHER and token == "0":
        return None

    # 只有 D=1 或 D=3 判斷漏檢
    if d_code not in (DiseaseCode.DM, DiseaseCode.DKD):
        return None

    if not isinstance(hba_dt, datetime.date) or not isinstance(ldl_dt, datetime.date):
        return None

    if hba_dt == ldl_dt:
        return None

    if hba_dt > ldl_dt:
        return "LDL漏檢"
    if ldl_dt > hba_dt:
        return "HbA1c漏檢"
    return None


# ============================================================
# 業務邏輯：備註（需要篩檢提示）
# ============================================================
def build_screening_note(
    *,
    age:      int,
    e_code:   Optional[DiseaseCode] = None,
    sex:      str,
    hep_dt:   Optional[datetime.date],
    fit_dt:   Optional[datetime.date],
    pap_dt:   Optional[datetime.date],
    adult_dt: Optional[datetime.date],
    flu_dt:   Optional[datetime.date],
    today:    datetime.date,
) -> str:
    msgs: List[str] = []

    if need_bc_hep(age) and hep_dt is None:
        msgs.append("今年需檢測BC肝")

    if need_fit(age):
        if fit_dt is None or (today.year - fit_dt.year) >= 2:
            msgs.append("今年需檢測糞便")

    pap_interval = pap_check_interval_years(age, sex)
    if pap_interval:
        if pap_dt is None or (today.year - pap_dt.year) >= pap_interval:
            msgs.append("今年需檢測子抹")

    adult_interval = adult_check_interval_years(age, e_code)
    if adult_interval:
        if adult_dt is None or (today.year - adult_dt.year) >= adult_interval:
            msgs.append("今年需檢測成健")

    if need_flu(age):
        if flu_dt is None or flu_dt.year < today.year:
            msgs.append("今年需檢測老流")

    return "\n".join(msgs)


# ============================================================
# 業務邏輯：分數計算
# ============================================================
def _score_hba(hba_val: Any, hba_dt: Optional[datetime.date]) -> int:
    if not isinstance(hba_dt, datetime.date):
        return 0
    if hba_dt.year == Rules.YEAR_UNCONDITIONAL:
        return Rules.SCORE_HBA
    if hba_dt.year == Rules.YEAR_VALUE_CHECK:
        v = parse_float(hba_val)
        if v is not None and v <= Rules.HBA_PASS_2025:
            return Rules.SCORE_HBA
    return 0


def _score_ldl(ldl_val: Any, ldl_dt: Optional[datetime.date]) -> int:
    if not isinstance(ldl_dt, datetime.date):
        return 0
    if ldl_dt.year == Rules.YEAR_UNCONDITIONAL:
        return Rules.SCORE_LDL
    if ldl_dt.year == Rules.YEAR_VALUE_CHECK:
        v = parse_float(ldl_val)
        if v is not None and v <= Rules.LDL_PASS_2025:
            return Rules.SCORE_LDL
    return 0


def _score_screening(dt: Optional[datetime.date], pts: int) -> int:
    if isinstance(dt, datetime.date) and dt.year == Rules.SCREEN_YEAR:
        return pts
    return 0


def _safe_int(value: Any) -> int:
    if value is None:
        return 0
    try:
        v = int(float(str(value).strip().lstrip("'").replace(",", "")))
        return max(v, 0)
    except Exception:
        return 0


def _score_visit_count(count_115: Any, count_114: Any) -> int:
    n115 = _safe_int(count_115)
    if n115 >= 4:
        return 10
    if n115 >= 2:
        return 8
    if n115 == 1:
        return 6

    n114 = _safe_int(count_114)
    if 8 <= n114 <= 20:
        return 4
    if 4 <= n114 <= 7:
        return 2
    return 1 if n114 > 0 else 0


def _score_fee(claim_amount_114: Any, claim_amount_115: Any) -> int:
    amt114 = parse_float(claim_amount_114)
    amt115 = parse_float(claim_amount_115)
    if amt114 is None or amt115 is None:
        return 0
    return 6 if amt115 < amt114 else 0


def _done_in_years(dt: Optional[datetime.date], years: set[int]) -> bool:
    return isinstance(dt, datetime.date) and dt.year in years


def _ever_done(dt: Optional[datetime.date]) -> bool:
    return isinstance(dt, datetime.date)


def _score_prevention(
    e_code: Optional[DiseaseCode],
    age: int,
    sex: str,
    adult_dt: Optional[datetime.date],
    adult_in_list: bool,
    pap_dt: Optional[datetime.date],
    pap_in_list: bool,
    flu_dt: Optional[datetime.date],
    flu_in_list: bool,
    fit_dt: Optional[datetime.date],
    fit_in_list: bool,
    hep_dt: Optional[datetime.date],
    hep_in_list: bool,
    today: datetime.date,
) -> int:
    sex = normalize_text(sex)
    if age < 0:
        return 0

    score = 28
    years_110_115 = {2021, 2022, 2023, 2024, 2025, 2026}
    years_112_115 = {2023, 2024, 2025, 2026}
    years_113_115 = {2024, 2025, 2026}
    year_115 = {2026}

    def deduct(done: bool, points: int) -> None:
        nonlocal score
        if not done:
            score -= points

    adult_interval = adult_check_interval_years(age, e_code)
    adult_ok = _screening_status_by_rule(
        "adult", in_screening_list=adult_in_list, dt=adult_dt,
        age=age, e_code=e_code, sex=sex, today=today,
    ) in ("not_needed", "done")
    pap_ok = _screening_status_by_rule(
        "pap", in_screening_list=pap_in_list, dt=pap_dt,
        age=age, e_code=e_code, sex=sex, today=today,
    ) in ("not_needed", "done")
    flu_ok = _screening_status_by_rule(
        "flu", in_screening_list=flu_in_list, dt=flu_dt,
        age=age, e_code=e_code, sex=sex, today=today,
    ) in ("not_needed", "done")
    fit_ok = _screening_status_by_rule(
        "fit", in_screening_list=fit_in_list, dt=fit_dt,
        age=age, e_code=e_code, sex=sex, today=today,
    ) in ("not_needed", "done")
    hep_ok = _screening_status_by_rule(
        "hep", in_screening_list=hep_in_list, dt=hep_dt,
        age=age, e_code=e_code, sex=sex, today=today,
    ) in ("not_needed", "done")

    if sex == "男":
        if 30 <= age <= 39:
            if adult_interval:
                deduct(adult_ok, 6)
        elif 40 <= age <= 64:
            if adult_interval:
                deduct(adult_ok, 6)
        elif age >= 65:
            if adult_interval:
                deduct(adult_ok, 6)

        if 45 <= age <= 75:
            deduct(fit_ok, 6)
        if 45 <= age < 80:
            deduct(hep_ok, 6)
        if age >= 65:
            deduct(flu_ok, 4)

    elif sex == "女":
        if 25 <= age <= 29:
            deduct(pap_ok, 6)
            if adult_interval:
                deduct(adult_ok, 6)
        elif 30 <= age <= 39:
            deduct(pap_ok, 6)
            if adult_interval:
                deduct(adult_ok, 6)
        elif 40 <= age <= 64:
            deduct(pap_ok, 6)
            if adult_interval:
                deduct(adult_ok, 6)
        elif age >= 65:
            deduct(pap_ok, 6)
            if adult_interval:
                deduct(adult_ok, 6)

        if 45 <= age <= 75:
            deduct(fit_ok, 6)
        if 45 <= age < 80:
            deduct(hep_ok, 6)
        if age >= 65:
            deduct(flu_ok, 4)

    return max(score, 0)


def _score_hba_management(
    e_code: Optional[DiseaseCode],
    age: int,
    hba_val: Any,
    hba_dt: Optional[datetime.date],
) -> int:
    if e_code not in (DiseaseCode.DM, DiseaseCode.DKD):
        return 5 if e_code in (DiseaseCode.CKD, DiseaseCode.OTHER) else 0
    if not isinstance(hba_dt, datetime.date) or hba_dt.year < 2026:
        return 0
    v = parse_float(hba_val)
    if v is None or age < 0:
        return 0
    return 5 if (v < 8.0 if age >= 80 else v < 7.0) else 0


def _score_ldl_management(
    e_code: Optional[DiseaseCode],
    ascvd_raw: Any,
    ldl_val: Any,
    ldl_dt: Optional[datetime.date],
) -> int:
    if not isinstance(ldl_dt, datetime.date) or ldl_dt.year < 2026:
        return 0
    v = parse_float(ldl_val)
    if v is None:
        return 0

    token = _ascvd_token(ascvd_raw)
    if token == "a":
        return 5 if v < 55.0 else 0
    if token == "b":
        return 5 if v < 70.0 else 0
    if e_code in (DiseaseCode.DM, DiseaseCode.DKD):
        return 5 if v < 100.0 else 0
    if e_code == DiseaseCode.CKD:
        return 5 if v < 130.0 else 0
    return 0


def _score_management_total(
    e_code: Optional[DiseaseCode],
    ascvd_raw: Any,
    age: int,
    hba_val: Any,
    hba_dt: Optional[datetime.date],
    ldl_val: Any,
    ldl_dt: Optional[datetime.date],
) -> int:
    token = _ascvd_token(ascvd_raw)

    if e_code == DiseaseCode.OTHER and token == "0":
        return 10
    if e_code is None and not token:
        return 0
    if e_code == DiseaseCode.OTHER and not token:
        return 0

    hba_score = _score_hba_management(e_code, age, hba_val, hba_dt)
    ldl_score = _score_ldl_management(e_code, ascvd_raw, ldl_val, ldl_dt)
    return hba_score + ldl_score

def calc_score(
    *,
    e_code:   Optional[DiseaseCode],
    ascvd_raw: Any,
    hba_val:  Any,
    hba_dt:   Optional[datetime.date],
    ldl_val:  Any,
    ldl_dt:   Optional[datetime.date],
    adult_dt: Optional[datetime.date],
    adult_in_list: bool,
    pap_dt:   Optional[datetime.date],
    pap_in_list: bool,
    flu_dt:   Optional[datetime.date],
    flu_in_list: bool,
    fit_dt:   Optional[datetime.date],
    fit_in_list: bool,
    hep_dt:   Optional[datetime.date],
    hep_in_list: bool,
    age:      int,
    sex:      str,
    today:    datetime.date,
    claim_amount_114: Any,
    claim_amount_115: Any,
    visit_count_114: Any,
    visit_count_115: Any,
) -> Tuple[int, str]:
    prevention_total = _score_prevention(
        e_code, age, sex,
        adult_dt, adult_in_list,
        pap_dt, pap_in_list,
        flu_dt, flu_in_list,
        fit_dt, fit_in_list,
        hep_dt, hep_in_list,
        today,
    )
    fee_score = _score_fee(claim_amount_114, claim_amount_115)
    exam_score = _score_management_total(e_code, ascvd_raw, age, hba_val, hba_dt, ldl_val, ldl_dt)
    visit_score = _score_visit_count(visit_count_115, visit_count_114)

    total = prevention_total + fee_score + exam_score + visit_score
    breakdown = "\n".join([
        f"1. 固定就診次數：{visit_score} 分",
        f"2. 醫療費用：{fee_score} 分",
        f"3. 糖心腎管理：{exam_score} 分",
        f"4. 預防保健：{prevention_total} 分",
    ])
    return total, breakdown


# ============================================================
# 業務邏輯：AU 第一次提醒（28天）
# ============================================================
def _is_valid_year(dt: Optional[datetime.date]) -> bool:
    return isinstance(dt, datetime.date) and dt.year in Rules.VALID_YEARS


def _au_item_line(item: str, last_dt: Optional[datetime.date], today: datetime.date) -> str:
    if not _is_valid_year(last_dt):
        return f"{item}:超過2年未檢查"
    due = last_dt + datetime.timedelta(days=Rules.AU_DAYS)  # type: ignore[operator]
    if due <= today:
        return f"{item}:可立刻通知回診"
    return f"{item}:{due.strftime('%Y-%m-%d')}需回診"


def build_au_note(
    *,
    e_code: Optional[DiseaseCode],
    ascvd:  AscvdCategory,
    hba_dt: Optional[datetime.date],
    ldl_dt: Optional[datetime.date],
    today:  datetime.date,
) -> str:
    has_dm    = e_code in (DiseaseCode.DM, DiseaseCode.DKD)
    has_ckd   = e_code in (DiseaseCode.CKD, DiseaseCode.DKD)
    has_ascvd = ascvd != AscvdCategory.NONE

    need_hba = has_dm
    need_ldl = has_dm or has_ckd or has_ascvd

    if not need_hba and not need_ldl:
        return "非疾病不需回診"

    any_valid = (
        (need_hba and _is_valid_year(hba_dt))
        or (need_ldl and _is_valid_year(ldl_dt))
    )
    lines: List[str] = []
    if not any_valid:
        if need_hba:
            lines.append("HbA1c:超過2年未檢查")
        if need_ldl:
            lines.append("LDL:超過2年未檢查")
        return "\n".join(lines)

    if need_hba:
        lines.append(_au_item_line("HbA1c", hba_dt, today))
    if need_ldl:
        lines.append(_au_item_line("LDL", ldl_dt, today))
    return "\n".join(lines)


# ============================================================
# 業務邏輯：AV/AW 從 AU 文字延後
# ============================================================
_DATE_RE = re.compile(r"(\d{4}-\d{2}-\d{2})")


def _should_skip_followup(au_txt: str) -> bool:
    s = (au_txt or "").replace(" ", "")
    return (not s) or ("非疾病" in s)


def _followup_item_line(
    item: str, base_dt: Optional[datetime.date], today: datetime.date, offset: int
) -> str:
    base = base_dt if isinstance(base_dt, datetime.date) else today
    due  = base + datetime.timedelta(days=offset)
    return f"{item}:{due.strftime('%Y-%m-%d')}需回診"


def build_followup_note(
    au_txt: str, today: datetime.date, offset_days: int
) -> Optional[str]:
    if not au_txt:
        return None

    out: List[str] = []
    for ln in (line.strip() for line in au_txt.splitlines() if line.strip()):
        if ln.startswith("HbA1c:"):
            item = "HbA1c"
        elif ln.startswith("LDL:"):
            item = "LDL"
        else:
            continue

        if "可立刻通知" in ln:
            out.append(_followup_item_line(item, None, today, offset_days))
            continue

        m = _DATE_RE.search(ln)
        if m:
            try:
                base = datetime.datetime.strptime(m.group(1), "%Y-%m-%d").date()
                out.append(_followup_item_line(item, base, today, offset_days))
            except ValueError:
                pass

    return "\n".join(out) if out else None


# ============================================================
# KPI：HbA1c（AY8 / AZ7 / AZ8 / AZ9）
# ============================================================
def _make_summary_text(label: Optional[str], numer: int, denom: int) -> str:
    label_txt = label or ""
    ratio_txt = _fmt_percent((numer / denom) if denom > 0 else 0.0, denom)
    fraction_txt = f"{numer}/{denom}" if denom > 0 else "0/0"
    return f"{label_txt}，{ratio_txt}，{fraction_txt}"


def _write_legacy_kpi_summary_cells(
    ws,
    *,
    hba_main_summary: str,
    hba_target_summary: str,
    ldl_main_summary: str,
    ldl_target_summary: str,
) -> None:
    """
    相容既有模板：保留 AQ2:AT2 的 KPI 摘要文字，
    避免舊版依賴這 4 格的輸出內容時出現差異。
    """
    ws["AQ2"] = hba_main_summary
    ws["AR2"] = hba_target_summary
    ws["AS2"] = ldl_main_summary
    ws["AT2"] = ldl_target_summary


def _collect_hba_candidates(
    ws, cols: Dict[str, Optional[int]], data_start: int, last_row: int
) -> List[Tuple[int, float]]:
    candidates: List[Tuple[int, float]] = []
    for r in range(data_start, last_row + 1):
        e = parse_disease_code(ws.cell(r, cols["dmk_code"]).value)  # type: ignore[index]
        if e not in (DiseaseCode.DM, DiseaseCode.DKD):
            continue
        if not is_zero_like(ws.cell(r, cols["ascvd"]).value):  # type: ignore[index]
            continue
        hba_dt = parse_date(ws.cell(r, cols["hba_dt"]).value)  # type: ignore[index]
        if not (isinstance(hba_dt, datetime.date) and hba_dt.year in Rules.VALID_YEARS):
            continue
        hba = parse_float(ws.cell(r, cols["hba"]).value)  # type: ignore[index]
        if hba is None:
            continue
        candidates.append((r, hba))
    return candidates


def _collect_ldl_candidates(
    ws, cols: Dict[str, Optional[int]], data_start: int, last_row: int
) -> List[Tuple[int, float]]:
    candidates: List[Tuple[int, float]] = []
    for r in range(data_start, last_row + 1):
        e = parse_disease_code(ws.cell(r, cols["dmk_code"]).value)  # type: ignore[index]
        if e not in (DiseaseCode.DM, DiseaseCode.CKD, DiseaseCode.DKD, DiseaseCode.OTHER):
            continue

        ascvd = parse_ascvd(ws.cell(r, cols["ascvd"]).value)  # type: ignore[index]
        if e == DiseaseCode.OTHER and ascvd == AscvdCategory.NONE:
            continue

        ldl_dt = parse_date(ws.cell(r, cols["ldl_dt"]).value)  # type: ignore[index]
        if not (isinstance(ldl_dt, datetime.date) and ldl_dt.year in Rules.VALID_YEARS):
            continue

        ldl = parse_float(ws.cell(r, cols["ldl"]).value)  # type: ignore[index]
        if ldl is None:
            continue

        candidates.append((r, ldl))
    return candidates


def collect_kpi_mark_sets(
    ws, cols: Dict[str, Optional[int]], data_start: int, last_row: int,
    hba_candidates: Optional[List[Tuple[int, float]]] = None,
    ldl_candidates: Optional[List[Tuple[int, float]]] = None,
) -> Dict[str, set]:
    if hba_candidates is None:
        hba_candidates = _collect_hba_candidates(ws, cols, data_start, last_row)
    if ldl_candidates is None:
        ldl_candidates = _collect_ldl_candidates(ws, cols, data_start, last_row)

    ay_rows = {r for r, v in hba_candidates if v <= Rules.HBA_CONTROL_THRESHOLD}

    hba_sorted = sorted(hba_candidates, key=lambda x: (x[1], x[0]))
    hba_k = int(math.ceil(Rules.HBA_TARGET_PERCENT * len(hba_sorted))) if hba_sorted else 0
    az_rows = {r for r, _ in hba_sorted[:hba_k]}

    ldl_threshold = Rules.LDL_KPI_THRESHOLDS[0]
    bb_rows = {r for r, v in ldl_candidates if v <= ldl_threshold}

    ldl_sorted = sorted(ldl_candidates, key=lambda x: (x[1], x[0]))
    ldl_k = int(math.ceil(Rules.LDL_TARGET_PERCENT * len(ldl_sorted))) if ldl_sorted else 0
    bc_rows = {r for r, _ in ldl_sorted[:ldl_k]}

    return {
        "ay": ay_rows,
        "az": az_rows,
        "bb": bb_rows,
        "bc": bc_rows,
    }


def calc_hba_kpi_ay_az(
    ws, cols: Dict[str, Optional[int]], data_start: int, last_row: int,
    hba_candidates: Optional[List[Tuple[int, float]]] = None,
) -> Tuple[str, str]:
    if hba_candidates is None:
        hba_candidates = _collect_hba_candidates(ws, cols, data_start, last_row)
    hba_values = [v for _, v in hba_candidates]
    denom = len(hba_values)

    numer_ay = sum(1 for v in hba_values if v <= Rules.HBA_CONTROL_THRESHOLD)
    main_summary = _make_summary_text("<=7", numer_ay, denom)

    if denom <= 0:
        target_summary = _make_summary_text("", 0, 0)
        print("AZ 分母=0，分子=0，比例=0.00%，切點=None")
        return main_summary, target_summary

    hba_values.sort()
    k = int(math.ceil(Rules.HBA_TARGET_PERCENT * denom))
    k = max(1, min(k, denom))
    cutoff = hba_values[k - 1]

    target_summary = _make_summary_text(f"<={cutoff:.2f}", k, denom)

    ratio = k / denom
    print(f"AZ 分母={denom}，分子={k}，比例={ratio*100:.2f}%，切點={cutoff:.2f}")
    return main_summary, target_summary


# ============================================================
# KPI：LDL（BB8/BC8）
# ============================================================
def calc_ldl_percentiles(
    ws, cols: Dict[str, Optional[int]], data_start: int, last_row: int,
    ldl_candidates: Optional[List[Tuple[int, float]]] = None,
) -> Tuple[str, str]:
    if ldl_candidates is None:
        ldl_candidates = _collect_ldl_candidates(ws, cols, data_start, last_row)
    ldl_values = [v for _, v in ldl_candidates]
    th_control = Rules.LDL_KPI_THRESHOLDS[0]  # 100
    denom = len(ldl_values)

    numer_bb = sum(1 for v in ldl_values if v <= th_control)
    main_summary = _make_summary_text("<=100", numer_bb, denom)

    if denom <= 0:
        target_summary = _make_summary_text("", 0, 0)
        print("BC 分母=0，分子=0，比例=0.00%，切點=None")
        return main_summary, target_summary

    ldl_values.sort()
    k = int(math.ceil(Rules.LDL_TARGET_PERCENT * denom))
    k = max(1, min(k, denom))
    cutoff = ldl_values[k - 1]

    target_summary = _make_summary_text(f"<={cutoff:.1f}".replace(".0", ""), k, denom)

    ratio = k / denom
    print(f"BC 分母={denom}，分子={k}，比例={ratio*100:.2f}%，切點={cutoff:.0f}")
    return main_summary, target_summary


# ============================================================
# 格式工具
# ============================================================
def apply_full_grid(ws, max_row: int, max_col: int) -> None:
    thin = Side(style="thin")
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            if cell.value is not None or r == 1:
                cell.border = grid


def apply_date_format(
    ws, cols: Dict[str, Optional[int]], data_start: int, last_row: int
) -> None:
    date_col_keys = ["bday", "adult", "pap", "flu", "fit", "hep",
                     "hba_dt", "ldl_dt", "uacr_dt",
                     "last_visit", "p4p_enroll_dt", "p4p_last_dt", "p4p_next_dt"]
    date_cols = [cols.get(k) for k in date_col_keys if cols.get(k)]
    for r in range(data_start, last_row + 1):
        for c in date_cols:
            cell = ws.cell(r, c)  # type: ignore[arg-type]
            if isinstance(cell.value, (datetime.date, datetime.datetime)):
                cell.number_format = "yyyy-mm-dd"


def apply_amount_format(
    ws, cols: Dict[str, Optional[int]], data_start: int, last_row: int
) -> None:
    amount_col_keys = [
        "r_amount_114", "s_amount_115",
    ]
    amount_cols = [cols.get(k) for k in amount_col_keys if cols.get(k)]
    for r in range(data_start, last_row + 1):
        for c in amount_cols:
            cell = ws.cell(r, c)  # type: ignore[arg-type]
            if cell.value is not None:
                cell.number_format = "#,##0"



# ============================================================
# 醫生看 sheet（從會員指標 Key 過來）
# ============================================================

# ============================================================
def _build_src_col_map(
    col_map: List[Tuple[str, str]],
    cols: Dict[str, Optional[int]],
) -> Dict[str, Optional[int]]:
    """
    根據 col_map（來源欄位 key, 目標欄字母）和 cols dict，
    建立來源欄位 key → 實際欄號的對照表。
    """
    result: Dict[str, Optional[int]] = {}
    for src_key, _ in col_map:
        result[src_key] = cols.get(src_key)
    return result


def _copy_sheet_rows(
    ws_main,
    ws_out,
    col_map: List[Tuple[str, str]],
    src_col_map: Dict[str, Optional[int]],
    data_start: int,
    last_row: int,
    dst_data_start: int = 3,
    filter_ids: Optional[set] = None,
    id_main_col: Optional[int] = None,
    filter_check_col: Optional[int] = None,
) -> int:
    """
    通用：把 ws_main 的資料按 col_map 逐列 copy 到 ws_out。
    filter_ids: 若指定，只 copy ID 在此 set 內的列。
    filter_check_col: 若指定，只 copy 該欄值為勾選的列。
    回傳實際寫入列數。
    """
    # 清空目標舊資料
    for r in range(dst_data_start, ws_out.max_row + 1):
        for c in range(1, ws_out.max_column + 1):
            ws_out.cell(r, c).value = None

    # 預先把 (src欄號, dst欄號) 轉換好，避免每列重複呼叫 column_index_from_string
    _col_pairs: List[Tuple[Optional[int], int]] = [
        (src_col_map.get(src_key), column_index_from_string(dst_letter))
        for src_key, dst_letter in col_map
    ]
    dst_row = dst_data_start
    for src_row in range(data_start, last_row + 1):
        if filter_check_col:
            if normalize_text(ws_main.cell(src_row, filter_check_col).value) not in ("✔", "v", "V"):
                continue
        # 過濾 ID
        if filter_ids is not None:
            pid = normalize_id(ws_main.cell(src_row, id_main_col).value) if id_main_col else ""
            if pid not in filter_ids:
                continue

        for src_c, dst_c in _col_pairs:
            if src_c:
                val = ws_main.cell(src_row, src_c).value
                normalized_val = "✔" if val == "v" else val
                cell = ws_out.cell(dst_row, dst_c)
                cell.value = normalized_val
        dst_row += 1

    ws_out.sheet_view.showGridLines = True
    return dst_row - dst_data_start


def _apply_vertical_center_to_sheet(
    ws,
    *,
    start_row: int = 1,
    end_row: Optional[int] = None,
    start_col: int = 1,
    end_col: Optional[int] = None,
) -> None:
    align_cache: Dict[int, Alignment] = {}
    end_row = end_row or ws.max_row
    end_col = end_col or ws.max_column
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cached_alignment = align_cache.get(cell.style_id)
            if cached_alignment is None:
                cached_alignment = copy(cell.alignment)
                cached_alignment.vertical = "center"
                align_cache[cell.style_id] = cached_alignment
            cell.alignment = cached_alignment


SELF_SELECT_SHEET_NAME = "自選名單(從會員指標內容Key過來)"

# ============================================================
# 醫生看 sheet
# ============================================================
DOCTOR_SHEET_NAME = "醫生看(從會員指標內容Key過來)"

_DOCTOR_COL_MAP: List[Tuple[str, str]] = [
    ("id", "A"),
    ("name", "B"),
    ("bday", "C"),
    ("tel", "E"),
    ("mobile", "F"),
    ("dx_raw", "G"),
    ("dmk_code", "H"),
    ("ascvd", "I"),
    ("last_visit", "J"),
    # 醫生看 K-O 對應：
    # K = 總表 L = 114全年就診次數
    # L = 總表隱藏輔助欄 BG = 114年1-4月就診次數
    # M = 總表 N = 115年有效月份總就診次數
    # N = 總表隱藏輔助欄 BD = 114年月平均費用
    # O = 總表隱藏輔助欄 BE = 115年有效月平均費用
    ("m_count_114", "K"),
    ("m_count_114_q1", "L"),
    ("n_count_115", "M"),
    ("avg_amount_114_hidden", "N"),
    ("avg_amount_115_hidden", "O"),
    ("adult", "P"),
    ("pap", "Q"),
    ("flu", "R"),
    ("fit", "S"),
    ("hep", "T"),
    ("hba", "U"),
    ("hba_dt", "V"),
    ("ldl", "W"),
    ("ldl_dt", "X"),
    ("uacr", "Y"),
    ("uacr_dt", "Z"),
    ("main_sub_dx", "AA"),
    ("p4p_status", "AB"),
    ("p4p_enroll_dt", "AC"),
    ("p4p_last_dt", "AD"),
    ("p4p_next_dt", "AE"),
    ("is_115x", "AF"),
    ("is_self_select", "AG"),
    ("is_114", "AH"),
    ("score", "AI"),
    ("breakdown", "AN"),
    ("metabolic_enroll", "AP"),
    ("note", "AO"),
]


def _extract_score_components(breakdown: Any) -> Tuple[Optional[int], Optional[int], Optional[int], Optional[int]]:
    text = normalize_text(breakdown)
    if not text:
        return None, None, None, None

    patterns = [
        ("固定就診次數", None),
        ("醫療費用", None),
        ("糖心腎管理", None),
        ("預防保健", None),
    ]
    values: List[Optional[int]] = []
    for label, _ in patterns:
        m = re.search(rf"{re.escape(label)}：\s*(\d+)\s*分", text)
        values.append(int(m.group(1)) if m else None)
    return values[0], values[1], values[2], values[3]


def _doctor_screening_status(
    kind: str,
    *,
    age: int,
    e_code: Optional[DiseaseCode] = None,
    sex: str,
) -> Tuple[str, PatternFill]:
    if kind == "adult":
        _, status_text, status_fill = adult_screening_status(age, e_code)
        return status_text or "不需受檢", status_fill
    elif kind == "pap":
        needed = pap_check_interval_years(age, sex) is not None
    elif kind == "flu":
        needed = need_flu(age)
    elif kind == "fit":
        needed = need_fit(age)
    elif kind == "hep":
        needed = need_bc_hep(age)
    else:
        needed = False

    if needed:
        return "待受檢", _FILL_SCREENING_PENDING
    return "不需受檢", _FILL_NONE


def _screening_needed(
    kind: str,
    *,
    age: int,
    e_code: Optional[DiseaseCode] = None,
    sex: str,
) -> bool:
    if kind == "adult":
        return adult_check_interval_years(age, e_code) is not None
    if kind == "pap":
        return pap_check_interval_years(age, sex) is not None
    if kind == "flu":
        return need_flu(age)
    if kind == "fit":
        return need_fit(age)
    if kind == "hep":
        return need_bc_hep(age)
    return False


def _screening_status_by_rule(
    kind: str,
    *,
    in_screening_list: bool,
    dt: Optional[datetime.date],
    age: int,
    e_code: Optional[DiseaseCode] = None,
    sex: str,
    today: datetime.date,
) -> str:
    if not _screening_needed(kind, age=age, e_code=e_code, sex=sex):
        return "not_needed"
    if not in_screening_list:
        return "uncertain"
    if dt is None:
        return "pending"
    if _doctor_screening_is_overdue(kind, dt=dt, age=age, e_code=e_code, sex=sex, today=today):
        return "overdue"
    return "done"


def _doctor_screening_display_from_status(
    *,
    status: str,
    dt: Optional[datetime.date],
) -> Tuple[str, PatternFill]:
    if status == "not_needed":
        return "不需受檢", _FILL_NONE
    if status == "uncertain":
        return "不確定(主動確認+補做機會)", _FILL_NONE
    if status == "done" and dt is not None:
        return dt.strftime("%Y-%m-%d"), _FILL_DOCTOR_DATE_DONE
    if status == "overdue":
        return "過期需受檢", _FILL_SCREENING_PENDING
    return "待受檢", _FILL_SCREENING_PENDING


def _doctor_unknown_age_display() -> Tuple[str, PatternFill]:
    return "年齡未知", _FILL_NONE


def _doctor_screening_is_overdue(
    kind: str,
    *,
    dt: Optional[datetime.date],
    age: int,
    e_code: Optional[DiseaseCode] = None,
    sex: str,
    today: datetime.date,
    ) -> bool:
    if not isinstance(dt, datetime.date):
        return False
    if kind == "adult":
        interval = adult_check_interval_years(age, e_code)
        return bool(interval and (today.year - dt.year) >= interval)
    if kind == "pap":
        interval = pap_check_interval_years(age, sex)
        return bool(interval and (today.year - dt.year) >= interval)
    if kind == "flu":
        return need_flu(age) and dt.year < today.year
    if kind == "fit":
        return need_fit(age) and (today.year - dt.year) >= 2
    if kind == "hep":
        return False
    return False


def _doctor_lab_status(
    kind: str,
    *,
    e_code: Optional[DiseaseCode],
    ascvd_raw: Any,
) -> Tuple[str, PatternFill]:
    if e_code is None and normalize_text(ascvd_raw) == "":
        return "", _FILL_NONE

    has_dm = e_code in (DiseaseCode.DM, DiseaseCode.DKD)
    has_ckd = e_code in (DiseaseCode.CKD, DiseaseCode.DKD)
    has_ascvd = parse_ascvd(ascvd_raw) != AscvdCategory.NONE

    if kind == "hba":
        needed = has_dm
    elif kind == "ldl":
        needed = has_dm or has_ckd or has_ascvd
    else:
        needed = False

    if needed:
        return "待受檢", _FILL_SCREENING_PENDING
    return "不需受檢", _FILL_NONE


def _doctor_lab_needed_by_disease(
    kind: str,
    *,
    e_code: Optional[DiseaseCode],
    ascvd_raw: Any,
) -> bool:
    has_dm = e_code in (DiseaseCode.DM, DiseaseCode.DKD)
    has_ckd = e_code in (DiseaseCode.CKD, DiseaseCode.DKD)
    has_ascvd = parse_ascvd(ascvd_raw) != AscvdCategory.NONE

    if kind == "hba":
        return has_dm
    if kind == "ldl":
        return has_dm or has_ckd or has_ascvd
    return False


def _doctor_lab_has_target(
    kind: str,
    *,
    e_code: Optional[DiseaseCode],
    ascvd_raw: Any,
) -> bool:
    if kind == "hba":
        return e_code in (DiseaseCode.DM, DiseaseCode.DKD) or _is_normal_person_by_group(
            e_code, parse_ascvd(ascvd_raw)
        )
    if kind == "ldl":
        return (
            e_code in (DiseaseCode.DM, DiseaseCode.CKD, DiseaseCode.DKD)
            or _is_normal_person_by_group(e_code, parse_ascvd(ascvd_raw))
            or _ascvd_token(ascvd_raw) in ("a", "b")
        )
    return False


def _doctor_date_text(dt_value: Any) -> Optional[str]:
    dt = parse_date(dt_value)
    if dt is None:
        return None
    return dt.strftime("%Y-%m-%d")


def _doctor_dmk_display(e_code: Optional[DiseaseCode], raw_value: Any) -> Any:
    if e_code == DiseaseCode.DM:
        return "DM"
    if e_code == DiseaseCode.CKD:
        return "CKD"
    if e_code == DiseaseCode.DKD:
        return "DKD"
    if e_code == DiseaseCode.OTHER:
        return "None"
    return raw_value


def _doctor_ascvd_display(raw_value: Any) -> Any:
    value = clean_spaces(raw_value).lower()
    if value == "0":
        return "0"
    if value in ("1", "a"):
        return "ASCVD-a"
    if value == "b":
        return "ASCVD-b"
    return raw_value


def _doctor_cell_alignment(col_letter: str, value: Any) -> Alignment:
    if col_letter in {"G", "AN", "AO"}:
        return _get_alignment("left", wrap_text=True)

    text = "" if value is None else str(value)
    if col_letter in {"P", "Q", "R", "S", "T", "Y"}:
        return _get_alignment("center", wrap_text=bool(text))
    if col_letter in {"U", "W"}:
        return _get_alignment("center", wrap_text="\n" in text)
    if col_letter in {"V", "X"}:
        return _get_alignment("center", wrap_text=text in ("待受檢", "不需受檢", "為DM/CKD/DKD排除個案"))
    return _get_alignment("center", wrap_text=False)


def _finalize_doctor_sheet_alignment(ws_doc) -> None:
    end_col = min(ws_doc.max_column, column_index_from_string("BJ"))
    for row in range(4, ws_doc.max_row + 1):
        for col in range(1, end_col + 1):
            cell = ws_doc.cell(row, col)
            col_letter = get_column_letter(col)
            cell.alignment = _doctor_cell_alignment(col_letter, cell.value)


def populate_doctor_sheet(
    wb_tpl,
    ws_main,
    cols: Dict[str, Optional[int]],
    data_start: int,
    last_row: int,
    today: Optional[datetime.date] = None,
    screening_member_ids: Optional[Dict[str, set[str]]] = None,
    p4p_map: Optional[Dict[str, Dict[str, Any]]] = None,
    claim_months_115: Optional[List[int]] = None,
    all_members: Optional[Dict[str, Dict[str, Any]]] = None,
    designated_115_ids: Optional[set[str]] = None,
    designated_115_details: Optional[Dict[str, Dict[str, Any]]] = None,
) -> None:
    if DOCTOR_SHEET_NAME not in wb_tpl.sheetnames:
        return
    src_col_map = _build_src_col_map(_DOCTOR_COL_MAP, cols)
    ws_doc = wb_tpl[DOCTOR_SHEET_NAME]
    months = sorted({month for month in (claim_months_115 or []) if 1 <= month <= 12})
    if months:
        month_text = _format_month_span(months)
        month_count = len(months)
        ws_doc["L2"] = f"✅ 條件\n114年{month_text}\n同期就診次數合計"
        ws_doc["M2"] = f"✅ 條件\n115年{month_text}\n就診次數合計"
        ws_doc["O2"] = (
            f"115年{month_text}\n月平均費用\n"
            f"（有效月份總額 ÷ {month_count}）"
        )
    else:
        ws_doc["L2"] = "114年同期就診次數\n（未偵測到115年有效月份）"
        ws_doc["M2"] = "115年就診次數\n（未偵測到有效月份）"
        ws_doc["O2"] = "115年月平均費用\n（未偵測到有效月份）"
    n = _copy_sheet_rows(
        ws_main,
        ws_doc,
        _DOCTOR_COL_MAP,
        src_col_map,
        data_start,
        last_row,
        dst_data_start=4,
    )
    today = today or datetime.date.today()
    ws_doc["P3"] = (
        "各項預防保健計分原則：\n"
        "老人流感：4 分、其餘各項：6 分\n"
        "醫生看 P-T 顯示：\n"
        "年齡未知 → 年齡未知\n"
        "不符合篩檢條件 → 不需受檢\n"
        "符合條件且在名單內：\n"
        "有日期且未過期 → 顯示日期\n"
        "已過期 → 過期需受檢\n"
        "沒日期 → 待受檢\n"
        "符合條件但不在預防保健名單內：\n"
        "不確定(主動確認+補做機會)\n"
        "給分條件：\n"
        "不符合篩檢條件、有日期且沒過期 → 給分\n"
        "沒篩檢過、過期需受檢、不在預防保健名單內 → 不給分"
    )
    for col_letter in ("P", "Q", "R", "S", "T"):
        ws_doc.column_dimensions[col_letter].width = max(
            ws_doc.column_dimensions[col_letter].width or 0,
            18,
        )

    breakdown_col = cols.get("breakdown")
    age_col = cols.get("age")
    sex_col = cols.get("sex")
    id_col = cols.get("id")
    abc_col = cols.get("abc")
    disease_text_col = cols.get("disease_text")
    hba_pass_col = cols.get("ak")
    ldl_pass_col = cols.get("ldl_pass")
    uacr_pass_col = cols.get("uacr_pass")
    leak_item_col = cols.get("ax")
    has_0526_extra_columns = normalize_text(ws_doc["AW1"].value) == "個案類別"
    doctor_extra_map = {
        "AW": "case_category",
        "AX": "quality_roster",
        "AY": "multi_chronic_65",
        "AZ": "high_visit",
        "BA": "chronic_mark",
        "BB": "non_chronic_mark",
        "BC": "same_clinic_previous_year",
        "BD": "dmk_raw",
        "BE": "ascvd",
        "BF": "three_highs",
        "BG": "hypertension",
        "BH": "hyperlipidemia",
        "BI": "hyperglycemia",
    }
    doctor_alert_font = Font(bold=True, color="FF0000")
    for offset, src_row in enumerate(range(data_start, last_row + 1), start=4):
        visit_score = fee_score = exam_score = prevention_score = None
        if breakdown_col:
            visit_score, fee_score, exam_score, prevention_score = _extract_score_components(
                ws_main.cell(src_row, breakdown_col).value
            )
        raw_dmk_value = ws_main.cell(src_row, cols["dmk_code"]).value
        e_code = parse_disease_code(raw_dmk_value)
        ascvd_raw = ws_main.cell(src_row, cols["ascvd"]).value
        age_val = ws_main.cell(src_row, age_col).value if age_col else None
        sex_val = normalize_text(ws_main.cell(src_row, sex_col).value) if sex_col else ""
        pid_val = normalize_id(ws_main.cell(src_row, id_col).value) if id_col else ""
        abc_val = normalize_text(ws_main.cell(src_row, abc_col).value) if abc_col else ""
        disease_text_val = normalize_text(ws_main.cell(src_row, disease_text_col).value) if disease_text_col else ""
        leak_item_val = normalize_text(ws_main.cell(src_row, leak_item_col).value) if leak_item_col else ""
        p4p_records = (p4p_map or {}).get(pid_val, {}).get("records", []) if pid_val else []
        ws_doc[f"D{offset}"] = age_val
        ws_doc[f"H{offset}"] = _doctor_dmk_display(e_code, raw_dmk_value)
        ws_doc[f"I{offset}"] = _doctor_ascvd_display(ascvd_raw)
        if p4p_records:
            ws_doc[f"AB{offset}"] = format_p4p_record_display(p4p_records[0])
            for extra_idx, record in enumerate(p4p_records[1:], start=0):
                ws_doc.cell(offset, column_index_from_string("AT") + extra_idx).value = format_p4p_record_display(record)
        else:
            p4p_status = normalize_text(ws_doc[f"AB{offset}"].value)
            p4p_plan = normalize_text(ws_main.cell(src_row, cols["p4p_plan"]).value) if cols.get("p4p_plan") else ""
            if p4p_status and p4p_plan:
                ws_doc[f"AB{offset}"] = f"{p4p_status}({p4p_plan})"
        ws_doc[f"AJ{offset}"] = visit_score
        ws_doc[f"AK{offset}"] = fee_score
        ws_doc[f"AL{offset}"] = exam_score
        ws_doc[f"AM{offset}"] = prevention_score
        ws_doc[f"AQ{offset}"] = "✔" if pid_val and pid_val in (designated_115_ids or set()) else None
        if e_code in (DiseaseCode.DM, DiseaseCode.CKD, DiseaseCode.DKD):
            ws_doc[f"AP{offset}"] = None
        if has_0526_extra_columns and pid_val:
            member_source = (designated_115_details or {}).get(pid_val, {})
            for col_letter, source_key in doctor_extra_map.items():
                ws_doc[f"{col_letter}{offset}"] = member_source.get(source_key)
            same_clinic = normalize_text(member_source.get("same_clinic_previous_year"))
            hypertension = normalize_text(member_source.get("hypertension"))
            ws_doc[f"BJ{offset}"] = "✔" if same_clinic == "1" and hypertension == "1" else None

        age_num = _safe_int(age_val)
        for col_letter, kind in (
            ("P", "adult"),
            ("Q", "pap"),
            ("R", "flu"),
            ("S", "fit"),
            ("T", "hep"),
        ):
            cell = ws_doc[f"{col_letter}{offset}"]
            cell.font = copy(cell.font)
            dt = parse_date(cell.value)
            if age_val in (None, ""):
                status_text, status_fill = _doctor_unknown_age_display()
                cell.value = status_text
                cell.fill = status_fill
                cell.font = doctor_alert_font
                cell.border = _BORDER_THIN_GRAY
                continue
            in_screening_list = bool(
                pid_val and screening_member_ids and pid_val in screening_member_ids.get(kind, set())
            )
            status = _screening_status_by_rule(
                kind,
                in_screening_list=in_screening_list,
                dt=dt,
                age=age_num,
                e_code=e_code,
                sex=sex_val,
                today=today,
            )
            status_text, status_fill = _doctor_screening_display_from_status(status=status, dt=dt)
            cell.value = status_text
            cell.fill = status_fill
            cell.border = _BORDER_THIN_GRAY

        for col_letter in ("P", "Q", "R", "S", "T"):
            ws_doc[f"{col_letter}{offset}"].border = _BORDER_THIN_GRAY

        for col_letter, kind in (("V", "hba"), ("X", "ldl")):
            cell = ws_doc[f"{col_letter}{offset}"]
            dt = parse_date(cell.value)
            status_text, status_fill = _doctor_lab_status(kind, e_code=e_code, ascvd_raw=ascvd_raw)
            if dt is not None:
                if status_text == "不需受檢":
                    cell.value = f"{dt.strftime('%Y-%m-%d')}\n(不需受檢)"
                if dt.year == Rules.SCREEN_YEAR:
                    cell.fill = _FILL_DOCTOR_DATE_DONE
                cell.border = _BORDER_THIN_GRAY
                continue
            if cell.value not in (None, ""):
                cell.border = _BORDER_THIN_GRAY
                continue
            cell.value = status_text
            cell.fill = status_fill
            cell.border = _BORDER_THIN_GRAY

        for col_letter, kind, pass_col in (("U", "hba", hba_pass_col), ("W", "ldl", ldl_pass_col)):
            cell = ws_doc[f"{col_letter}{offset}"]
            cell.border = _BORDER_THIN_GRAY
            cell.font = copy(cell.font)
            date_col_letter = "V" if col_letter == "U" else "X"
            date_cell_value = ws_doc[f"{date_col_letter}{offset}"].value
            date_val = parse_date(date_cell_value)
            needs_check = _doctor_lab_has_target(kind, e_code=e_code, ascvd_raw=ascvd_raw)
            needs_check_by_disease = _doctor_lab_needed_by_disease(kind, e_code=e_code, ascvd_raw=ascvd_raw)
            is_unknown_group = e_code is None and normalize_text(ascvd_raw) == ""
            is_other_without_ascvd = e_code == DiseaseCode.OTHER and _ascvd_token(ascvd_raw) == "0"

            if (col_letter == "U" and leak_item_val == "HbA1c漏檢") or (col_letter == "W" and leak_item_val == "LDL漏檢"):
                display_val = normalize_text(cell.value)
                if display_val:
                    cell.value = f"{display_val}\n(2026年漏檢)"
                else:
                    cell.value = "(2026年漏檢)"
                cell.fill = _FILL_NONE
                cell.font = doctor_alert_font
                continue

            if needs_check_by_disease and (date_val is None or date_val.year != Rules.SCREEN_YEAR):
                display_val = normalize_text(cell.value)
                if display_val and display_val not in ("待受檢", "不需受檢"):
                    cell.value = f"{display_val}\n(2026需受檢)"
                else:
                    cell.value = "(2026需受檢)"
                cell.fill = _FILL_NONE
                cell.font = doctor_alert_font
                continue

            lab_val = parse_float(cell.value)
            if lab_val is None:
                continue

            if not needs_check:
                cell.fill = _FILL_NONE
                continue

            pass_mark = normalize_text(ws_main.cell(src_row, pass_col).value) if pass_col else ""
            if pass_mark != "✔":
                cell.fill = _FILL_NONE
                if is_unknown_group or is_other_without_ascvd:
                    display_val = str(int(lab_val)) if float(lab_val).is_integer() else str(lab_val)
                    cell.value = f"{display_val}\n(已受檢未達控制)"
                    cell.font = doctor_alert_font
                elif date_val is not None and date_val.year == Rules.SCREEN_YEAR:
                    display_val = str(int(lab_val)) if float(lab_val).is_integer() else str(lab_val)
                    cell.value = f"{display_val}\n(已受檢未達控制)"
                    cell.font = doctor_alert_font

        uacr_cell = ws_doc[f"Y{offset}"]
        uacr_cell.font = copy(uacr_cell.font)
        uacr_val_num = parse_float(uacr_cell.value)
        uacr_dt = parse_date(ws_doc[f"Z{offset}"].value)
        uacr_pass_mark = normalize_text(ws_main.cell(src_row, uacr_pass_col).value) if uacr_pass_col else ""
        needs_uacr_by_disease = e_code in (DiseaseCode.CKD, DiseaseCode.DKD)
        if (
            uacr_val_num is not None
            and uacr_pass_mark != "✔"
            and uacr_dt is not None
            and uacr_dt.year == Rules.SCREEN_YEAR
        ):
            display_val = str(int(uacr_val_num)) if float(uacr_val_num).is_integer() else str(uacr_val_num)
            uacr_cell.value = f"{display_val}\n(已受檢未達控制)"
            uacr_cell.font = doctor_alert_font
        elif (
            uacr_val_num is not None
            and uacr_val_num > 30
            and uacr_dt is not None
            and uacr_dt.year == Rules.SCREEN_YEAR
        ):
            display_val = str(int(uacr_val_num)) if float(uacr_val_num).is_integer() else str(uacr_val_num)
            uacr_cell.value = f"{display_val}\n(已受檢未達控制)"
            uacr_cell.font = doctor_alert_font
        elif needs_uacr_by_disease and (uacr_dt is None or uacr_dt.year != Rules.SCREEN_YEAR):
            display_val = normalize_text(uacr_cell.value)
            if display_val:
                uacr_cell.value = f"{display_val}\n(2026需受檢)"
            else:
                uacr_cell.value = "(2026需受檢)"
            uacr_cell.font = doctor_alert_font

    print(f"醫生看 sheet 已寫入 {n} 列")


# ============================================================
# 自選名單 sheet
# ============================================================
_SELF_SELECT_COL_MAP: List[Tuple[str, str]] = [
    ("name", "A"),
    ("id", "B"),
    ("m_count_114", "C"),
    ("m_count_114_q1", "D"),
    ("n_count_115", "E"),
    ("r_amount_114", "F"),
    ("s_amount_115", "G"),
    ("adult", "H"),
    ("pap", "I"),
    ("flu", "J"),
    ("fit", "K"),
    ("hep", "L"),
    ("hba", "M"),
    ("hba_dt", "N"),
    ("ldl", "O"),
    ("ldl_dt", "P"),
    ("p4p_status", "Q"),
    ("p4p_enroll_dt", "R"),
    ("p4p_last_dt", "S"),
    ("p4p_next_dt", "T"),
    ("is_114", "U"),
]

_SELF_SELECT_HIDDEN_SCORE_COLS: List[Tuple[str, str]] = [
    ("total_score", "V"),
    ("visit_score", "W"),
    ("fee_score", "X"),
    ("exam_score", "Y"),
    ("prevention_score", "Z"),
]


def populate_self_select_sheet(
    wb_tpl,
    ws_main,
    cols: Dict[str, Optional[int]],
    data_start: int,
    last_row: int,
    sh_self_select: Any,
) -> None:
    if SELF_SELECT_SHEET_NAME not in wb_tpl.sheetnames:
        return

    self_select_col = cols.get("is_self_select")
    if not self_select_col:
        print("自選名單 sheet 找不到主表自選欄，略過")
        return

    src_col_map = _build_src_col_map(_SELF_SELECT_COL_MAP, cols)
    ws_out = wb_tpl[SELF_SELECT_SHEET_NAME]
    n = _copy_sheet_rows(
        ws_main, ws_out, _SELF_SELECT_COL_MAP, src_col_map,
        data_start, last_row,
        filter_check_col=self_select_col,
    )

    score_col = cols.get("score")
    breakdown_col = cols.get("breakdown")
    hidden_col_indexes = {
        key: column_index_from_string(letter)
        for key, letter in _SELF_SELECT_HIDDEN_SCORE_COLS
    }

    for key, letter in _SELF_SELECT_HIDDEN_SCORE_COLS:
        ws_out.column_dimensions[letter].hidden = True

    ws_out.cell(2, hidden_col_indexes["total_score"]).value = "分數"
    ws_out.cell(2, hidden_col_indexes["visit_score"]).value = "固定就診次數"
    ws_out.cell(2, hidden_col_indexes["fee_score"]).value = "醫療費用"
    ws_out.cell(2, hidden_col_indexes["exam_score"]).value = "糖心腎管理"
    ws_out.cell(2, hidden_col_indexes["prevention_score"]).value = "預防保健"

    dst_row = 3
    for src_row in range(data_start, last_row + 1):
        if normalize_text(ws_main.cell(src_row, self_select_col).value) not in ("✔", "v", "V"):
            continue

        visit_score = fee_score = exam_score = prevention_score = None
        if breakdown_col:
            visit_score, fee_score, exam_score, prevention_score = _extract_score_components(
                ws_main.cell(src_row, breakdown_col).value
            )

        ws_out.cell(dst_row, hidden_col_indexes["total_score"]).value = ws_main.cell(src_row, score_col).value if score_col else None
        ws_out.cell(dst_row, hidden_col_indexes["visit_score"]).value = visit_score
        ws_out.cell(dst_row, hidden_col_indexes["fee_score"]).value = fee_score
        ws_out.cell(dst_row, hidden_col_indexes["exam_score"]).value = exam_score
        ws_out.cell(dst_row, hidden_col_indexes["prevention_score"]).value = prevention_score
        dst_row += 1

    print(f"自選名單 sheet 已寫入 {n} 列")


def _finalize_main_sheet_alignment(ws, data_start: int, last_row: int) -> None:
    end_col = _main_alignment_end_col(ws)
    for row in range(data_start, last_row + 1):
        for col in range(1, end_col + 1):
            cell = ws.cell(row, col)
            if col in _LEFT_COLS:
                cell.alignment = _get_alignment("left", wrap_text=True)
            else:
                cell.alignment = _get_alignment("center", wrap_text=True)


def _finalize_self_select_sheet_alignment(ws) -> None:
    end_col = column_index_from_string("Z")
    for row in range(3, ws.max_row + 1):
        for col in range(1, end_col + 1):
            ws.cell(row, col).alignment = _get_alignment("center", wrap_text=True)

# ============================================================
# 百分位名單 sheet（sheet2）
# ============================================================
PERCENTILE_SHEET_NAME = "百分位名單"
PINK_FILL = openpyxl.styles.PatternFill(fill_type="solid", fgColor="EAC0C0")
BLUE_FILL = openpyxl.styles.PatternFill(fill_type="solid", fgColor="A9C2D9")
NO_FILL = openpyxl.styles.PatternFill(fill_type=None)
PERCENTILE_FONT = Font(size=14)
PERCENTILE_LEFT_OFFSETS = frozenset([0, 4, 7, 8, 9])


def _copy_row_style(ws, src_row: int, dst_row: int, start_col: int = 1, end_col: int = 24) -> None:
    """直接複製 _style 索引，不建立任何新物件，比逐屬性 copy() 快約 20 倍。"""
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for c in range(start_col, end_col + 1):
        ws.cell(dst_row, c)._style = ws.cell(src_row, c)._style


def _cache_row_styles(ws, src_row: int, start_col: int, end_col: int) -> dict:
    """
    快取模板列每欄的 _style 索引與列高。
    openpyxl 內部以整數索引參照共享樣式表，直接複製索引比 copy() 物件快約 20 倍，
    且框線、底色、字型、格式均完整保留。
    """
    cache: dict = {
        "_height": ws.row_dimensions[src_row].height,
        "_styles": [ws.cell(src_row, c)._style for c in range(start_col, end_col + 1)],
        "_start":  start_col,
    }
    return cache


def _apply_cached_style(ws, dst_row: int, style_cache: dict, start_col: int, end_col: int) -> None:
    """直接寫入 _style 索引，不建立任何新物件，速度遠快於逐屬性 copy()。"""
    ws.row_dimensions[dst_row].height = style_cache["_height"]
    styles = style_cache["_styles"]
    base   = style_cache["_start"]
    for c in range(start_col, end_col + 1):
        ws.cell(dst_row, c)._style = styles[c - base]


def _clear_percentile_data_area(ws, start_row: int = 5, end_col: int = 24) -> None:
    for r in range(start_row, ws.max_row + 1):
        for c in range(1, end_col + 1):
            ws.cell(r, c).value = None


def _filter_followup_by_metric(text: Any, metric: str) -> Optional[str]:
    """
    AU/AV/AW 欄位可能同時包含 HbA1c 與 LDL 兩行（用 \\n 分隔）。
    百分位名單第二頁只顯示對應指標那一行，避免 LDL 側出現 HbA1c 文字或反之。
    """
    if text is None:
        return None
    prefix = "LDL:" if metric == "ldl" else "HbA1c:"
    lines = [ln for ln in str(text).splitlines() if ln.strip().startswith(prefix)]
    return "\n".join(lines) if lines else None


def _make_percentile_record(ws_main, row: int, cols: Dict[str, Optional[int]], metric: str, name_fill: str) -> Dict[str, Any]:
    value_col = cols["ldl"] if metric == "ldl" else cols["hba"]
    date_col = cols["ldl_dt"] if metric == "ldl" else cols["hba_dt"]

    raw_au = ws_main.cell(row, cols["au"]).value  # type: ignore[index]
    raw_av = ws_main.cell(row, cols["av"]).value  # type: ignore[index]
    raw_aw = ws_main.cell(row, cols["aw"]).value  # type: ignore[index]

    return {
        "row": row,
        "name": ws_main.cell(row, cols["name"]).value,  # type: ignore[index]
        "bday": ws_main.cell(row, cols["bday"]).value,  # type: ignore[index]
        "id": ws_main.cell(row, cols["id"]).value,      # type: ignore[index]
        "score": ws_main.cell(row, cols["score"]).value,  # type: ignore[index]
        "note": ws_main.cell(row, cols["note"]).value,    # type: ignore[index]
        "value": ws_main.cell(row, value_col).value if value_col else None,
        "last_dt": ws_main.cell(row, date_col).value if date_col else None,
        "au": _filter_followup_by_metric(raw_au, metric),
        "av": _filter_followup_by_metric(raw_av, metric),
        "aw": _filter_followup_by_metric(raw_aw, metric),
        "fill": name_fill,
    }


def _collect_percentile_records(
    ws_main,
    cols: Dict[str, Optional[int]],
    data_start: int,
    last_row: int,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    一次掃描會員總表，同時收集 LDL 和 HbA1c 的百分位名單。
    原本兩次 _collect_side_records_sorted 共掃 2 遍，現合併為 1 遍。
    """
    ldl_primary:   List[int] = []
    ldl_secondary: List[int] = []
    hba_primary:   List[int] = []
    hba_secondary: List[int] = []

    c_bb = cols.get("bb_mark")
    c_bc = cols.get("bc_mark")
    c_ay = cols.get("ay_mark")
    c_az = cols.get("az_mark")

    for row in range(data_start, last_row + 1):
        if c_bb and normalize_text(ws_main.cell(row, c_bb).value) == "✔":
            ldl_primary.append(row)
        elif c_bc and normalize_text(ws_main.cell(row, c_bc).value) == "✔":
            ldl_secondary.append(row)

        if c_ay and normalize_text(ws_main.cell(row, c_ay).value) == "✔":
            hba_primary.append(row)
        elif c_az and normalize_text(ws_main.cell(row, c_az).value) == "✔":
            hba_secondary.append(row)

    ldl_records = (
        [_make_percentile_record(ws_main, r, cols, "ldl", "pink") for r in ldl_primary] +
        [_make_percentile_record(ws_main, r, cols, "ldl", "blue") for r in ldl_secondary]
    )
    hba_records = (
        [_make_percentile_record(ws_main, r, cols, "hba", "pink") for r in hba_primary] +
        [_make_percentile_record(ws_main, r, cols, "hba", "blue") for r in hba_secondary]
    )
    return ldl_records, hba_records




def _ratio_numerator_text(value: Any) -> str:
    s = normalize_text(value)
    if not s:
        return "0"
    return s.split("/", 1)[0].strip() or "0"


def _percentile_summary_text(value: Any) -> str:
    s = normalize_text(value)
    if not s:
        return "0.00%，0人"

    parts = [p.strip() for p in s.split("，")]
    if len(parts) < 3:
        return s

    label = parts[0]
    ratio = parts[1]
    numer = _ratio_numerator_text(parts[2])
    return f"{label}，{ratio}，{numer}人"


def _set_percentile_title_rich(ws, cell_ref: str, prefix: str, red_num: Any, blue_num: Any) -> None:
    red_txt = _percentile_summary_text(red_num)
    blue_txt = _percentile_summary_text(blue_num)
    ws[cell_ref].value = f"{prefix}({red_txt})、({blue_txt})"


def _write_percentile_side(ws, row: int, start_col: int, rec: Optional[Dict[str, Any]]) -> None:
    cols_seq = [start_col + i for i in range(10)]
    name_cell = ws.cell(row, start_col)

    if not rec:
        for c in cols_seq:
            ws.cell(row, c).value = None
        name_cell.fill = copy(NO_FILL)
        return

    values = [
        rec.get("name"),
        rec.get("bday"),
        rec.get("id"),
        rec.get("score"),
        rec.get("note"),
        rec.get("value"),
        rec.get("last_dt"),
        rec.get("au"),
        rec.get("av"),
        rec.get("aw"),
    ]
    for c, v in zip(cols_seq, values):
        ws.cell(row, c).value = v

    if normalize_text(rec.get("name")):
        name_cell.fill = copy(PINK_FILL if rec.get("fill") == "pink" else BLUE_FILL)
    else:
        name_cell.fill = copy(NO_FILL)


def _apply_percentile_data_row_style(ws, row: int) -> None:
    """百分位資料列固定使用內建字型 14，避免沿用舊樣式造成字型跑掉。"""
    ws.row_dimensions[row].height = 66
    for start_col in (1, 14):
        for offset in range(10):
            cell = ws.cell(row, start_col + offset)
            cell.font = PERCENTILE_FONT


def populate_percentile_sheet(
    wb_tpl,
    hba_main_summary: str,
    hba_target_summary: str,
    ldl_main_summary: str,
    ldl_target_summary: str,
    cols: Dict[str, Optional[int]],
    data_start: int,
    last_row: int,
) -> None:
    if PERCENTILE_SHEET_NAME not in wb_tpl.sheetnames:
        ws = wb_tpl.create_sheet(PERCENTILE_SHEET_NAME)
    else:
        ws = wb_tpl[PERCENTILE_SHEET_NAME]

    ws_main = wb_tpl[Rules.SHEET_TARGET]
    ldl_records, hba_records = _collect_percentile_records(ws_main, cols, data_start, last_row)

    _set_percentile_title_rich(ws, "A1", "LDL百分位", ldl_main_summary, ldl_target_summary)
    _set_percentile_title_rich(ws, "N1", "HBA1C百分位", hba_main_summary, hba_target_summary)
    ws["A2"] = "紅色：達到標準、藍色：達到73.8%"
    ws["N2"] = "紅色：達到標準、藍色：達到73.8%"

    _clear_percentile_data_area(ws, start_row=5, end_col=24)

    data_rows = max(len(ldl_records), len(hba_records), 1)
    for idx in range(data_rows):
        row = 5 + idx
        _apply_percentile_data_row_style(ws, row)
        _write_percentile_side(ws, row, 1,  ldl_records[idx] if idx < len(ldl_records) else None)
        _write_percentile_side(ws, row, 14, hba_records[idx] if idx < len(hba_records) else None)

    for row in range(5, 5 + data_rows):
        for col in [2, 7, 15, 20]:
            cell = ws.cell(row, col)
            if isinstance(cell.value, (datetime.date, datetime.datetime)):
                cell.number_format = "yyyy-mm-dd"
    apply_full_grid(ws, ws.max_row, 24)


def _finalize_percentile_sheet_alignment(ws) -> None:
    end_row = ws.max_row
    while end_row >= 5:
        if any(ws.cell(end_row, col).value not in (None, "") for col in range(1, 25)):
            break
        end_row -= 1
    for row in range(5, end_row + 1):
        for start_col in (1, 14):
            for offset in range(10):
                cell = ws.cell(row, start_col + offset)
                if offset in PERCENTILE_LEFT_OFFSETS:
                    cell.alignment = _get_alignment("left", wrap_text=True)
                else:
                    cell.alignment = _get_alignment("center", wrap_text=True)


# ============================================================
# 全名單整合（v3 新增）
# ============================================================
def _id_from_sheet(sheet, header_aliases: List[str], search_rows: int = 10) -> Tuple[Optional[int], int]:
    """從 sheet 找 ID 欄（靠 header + 內容驗證）"""
    header_row = _find_header_row_contains_any(sheet, [header_aliases], search_rows=search_rows)
    if header_row is None:
        header_row = 1
    hmap = build_header_map(sheet, header_row)
    cand = find_column_exact(hmap, header_aliases)
    return find_id_col_by_content(sheet, header_row, cand), header_row


def _extract_id_name_map(sheet, id_aliases: List[str]) -> Dict[str, str]:
    """從任意 sheet 抽出 {id: name} 對應表"""
    if sheet is None:
        return {}
    result: Dict[str, str] = {}
    col_info = _id_from_sheet(sheet, id_aliases)
    id_col, hrow = col_info
    if id_col is None:
        return {}
    # 找姓名欄
    hmap = build_header_map(sheet, hrow)
    name_col = find_column_exact(hmap, ["姓名", "會員姓名", "名字"])
    for r in range(hrow + 1, sheet.max_row + 1):
        pid = normalize_id(sheet.cell(r, id_col).value)
        if pid and is_valid_tw_id(pid):
            name = sheet.cell(r, name_col).value if name_col else None
            result.setdefault(pid, normalize_text(name) if name else "")
    return result


def _empty_member() -> Dict[str, Any]:
    """回傳一筆空白會員記錄（所有欄位為 None / 空字串）。"""
    return {"name": "", "bday": None, "e_code": None, "abc": None,
            "phone": None, "mobile": None, "address": None,
            "dmk_raw": None, "cnt": None, "clinic": "", "ascvd": None,
            "case_category": None, "quality_roster": None,
            "multi_chronic_65": None, "high_visit": None,
            "chronic_mark": None, "non_chronic_mark": None,
            "same_clinic_previous_year": None, "three_highs": None,
            "hypertension": None, "hyperlipidemia": None,
            "hyperglycemia": None}


_MEMBER_ROSTER_EXTRA_COLUMNS: Dict[str, List[str]] = {
    "case_category": ["個案類別"],
    "quality_roster": ["論質名單"],
    "multi_chronic_65": ["65歲以上多重慢性病註記"],
    "high_visit": ["高診次註記"],
    "chronic_mark": ["慢性病註記"],
    "non_chronic_mark": ["非慢性病註記"],
    "same_clinic_previous_year": ["與前一年家醫收案診所相同", "與前一年家醫收案相同"],
    "three_highs": ["三高"],
    "hypertension": ["高血壓"],
    "hyperlipidemia": ["高血脂"],
    "hyperglycemia": ["高血糖"],
}


def _find_member_roster_extra_columns(hmap: Dict[str, int]) -> Dict[str, Optional[int]]:
    return {
        key: find_column_exact(hmap, aliases)
        for key, aliases in _MEMBER_ROSTER_EXTRA_COLUMNS.items()
    }


def _fill_member_roster_extra_fields(
    rec: Dict[str, Any],
    sheet: Any,
    row: int,
    extra_cols: Dict[str, Optional[int]],
) -> None:
    for key, col in extra_cols.items():
        if col:
            _fill_member_field(rec, key, sheet.cell(row, col).value)


def _fill_member_field(rec: Dict[str, Any], key: str, value: Any) -> None:
    """只在記錄中該欄位尚未有值時才填入（缺什麼補什麼）。"""
    if value is not None and value != "" and not rec.get(key):
        rec[key] = value


def _extract_member_partial_map(
    sheet,
    id_aliases: List[str],
    search_rows: int = 10,
) -> Dict[str, Dict[str, Any]]:
    """從任意 sheet 抽出可用的基本資料欄位，供會員主表缺漏回補。"""
    if sheet is None:
        return {}

    header_row = _find_header_row_contains_any(sheet, [id_aliases], search_rows=search_rows)
    if header_row is None:
        header_row = 1
    hmap = build_header_map(sheet, header_row)
    id_col = find_id_col_by_content(sheet, header_row, find_column_exact(hmap, id_aliases))
    if id_col is None:
        return {}
    data_start_row = _first_valid_id_row(sheet, id_col, header_row)

    name_col = find_column_exact(hmap, ["會員姓名", "姓名", "名字"])
    bday_col = find_bday_column(hmap)
    dmk_col = find_column_exact(hmap, ["疾病樣態"])
    ascvd_col = find_column_exact(hmap, ["ASCVD", "ascvd"])
    extra_cols = _find_member_roster_extra_columns(hmap)
    phone_cols = [
        c for hdr, c in hmap.items()
        if c != id_col and is_phone_header(hdr)
    ]
    addr_col = next(
        (c for hdr, c in hmap.items() if c != id_col and is_address_header(hdr)),
        None,
    )

    result: Dict[str, Dict[str, Any]] = {}
    for r in range(data_start_row, sheet.max_row + 1):
        pid = normalize_id(sheet.cell(r, id_col).value)
        if not pid or not is_valid_tw_id(pid):
            continue

        rec = result.setdefault(pid, _empty_member())
        if name_col:
            _fill_member_field(rec, "name", normalize_text(sheet.cell(r, name_col).value))
        if bday_col:
            _fill_member_field(rec, "bday", parse_date(sheet.cell(r, bday_col).value))
        if dmk_col:
            dmk_val = sheet.cell(r, dmk_col).value
            _fill_member_field(rec, "e_code", parse_disease_code(dmk_val))
            _fill_member_field(rec, "dmk_raw", dmk_val)
        if ascvd_col:
            ascvd_val = sheet.cell(r, ascvd_col).value
            if ascvd_val is not None and str(ascvd_val).strip() != "":
                _fill_member_field(rec, "ascvd", ascvd_val)
        _fill_member_roster_extra_fields(rec, sheet, r, extra_cols)
        if phone_cols:
            contact = pick_contact_from_values([sheet.cell(r, c).value for c in phone_cols])
            _fill_member_field(rec, "phone", contact.phone)
            _fill_member_field(rec, "mobile", contact.mobile)
        if addr_col:
            _fill_member_field(rec, "address", normalize_text(sheet.cell(r, addr_col).value) or None)

    return result


def _iter_member_union_keys(rec: Dict[str, Any]) -> List[Tuple[str, str]]:
    name = clean_spaces(rec.get("name")).upper()
    if not name:
        return []

    keys: List[Tuple[str, str]] = []
    bday = rec.get("bday")
    if isinstance(bday, datetime.date):
        keys.append(("name_bday", f"{name}|BDAY|{bday.strftime('%Y-%m-%d')}"))

    for raw in (rec.get("phone"), rec.get("mobile")):
        phone = normalize_phone_value(raw)
        if phone and len(phone) >= 7:
            keys.append(("name_phone", f"{name}|PHONE|{phone[-7:]}"))

    keys.append(("name_only", name))
    return list(dict.fromkeys(keys))


def _extract_member_noid_rows(
    sheet,
    id_aliases: List[str],
    search_rows: int = 10,
) -> List[Dict[str, Any]]:
    if sheet is None:
        return []

    id_col, header_row = _find_sheet_id_col(sheet, id_aliases, search_rows=search_rows)
    if id_col is not None:
        return []

    hmap = build_header_map(sheet, header_row)
    name_col = find_column_exact(hmap, ["會員姓名", "姓名", "名字"])
    if not name_col:
        return []

    bday_col = find_bday_column(hmap)
    ascvd_col = find_column_exact(hmap, ["ASCVD", "ascvd"])
    extra_cols = _find_member_roster_extra_columns(hmap)
    abc_col = find_column_exact(hmap, ["會員別"])
    dmk_col = find_column_exact(hmap, ["疾病樣態"])
    cnt_col = find_column_exact(hmap, ["就診次數"])
    phone_cols = [c for hdr, c in hmap.items() if c != name_col and is_phone_header(hdr)]
    addr_col = next((c for hdr, c in hmap.items() if c != name_col and is_address_header(hdr)), None)
    clinic_val = normalize_text(sheet.cell(1, 1).value)

    out: List[Dict[str, Any]] = []
    for r in range(header_row + 1, sheet.max_row + 1):
        name = normalize_text(sheet.cell(r, name_col).value) if name_col else ""
        if not name:
            continue
        rec = _empty_member()
        rec["name"] = name
        if bday_col:
            rec["bday"] = parse_date(sheet.cell(r, bday_col).value)
        if abc_col:
            rec["abc"] = sheet.cell(r, abc_col).value
        if dmk_col:
            rec["e_code"] = parse_disease_code(sheet.cell(r, dmk_col).value)
            rec["dmk_raw"] = sheet.cell(r, dmk_col).value
        if cnt_col:
            rec["cnt"] = sheet.cell(r, cnt_col).value
        if ascvd_col:
            ascvd_val = sheet.cell(r, ascvd_col).value
            if ascvd_val is not None and str(ascvd_val).strip() != "":
                rec["ascvd"] = ascvd_val
        _fill_member_roster_extra_fields(rec, sheet, r, extra_cols)
        if phone_cols:
            contact = pick_contact_from_values([sheet.cell(r, c).value for c in phone_cols])
            rec["phone"] = contact.phone
            rec["mobile"] = contact.mobile
        if addr_col:
            rec["address"] = normalize_text(sheet.cell(r, addr_col).value) or None
        if clinic_val:
            rec["clinic"] = clinic_val

        if any(rec.get(k) for k in ("name", "bday", "phone", "mobile", "address", "ascvd", "e_code", "cnt")):
            out.append(rec)
    return out


def collect_all_members(
    wb_src,
    partial_maps: Optional[Dict[str, Dict[str, Dict[str, Any]]]] = None,
) -> Dict[str, Dict[str, Any]]:
    """
    整合所有 sheet 的人員名單，以身份證號為鍵，去重複。
    各來源採「缺什麼補什麼」原則：已有值的欄位不會被空值覆蓋。
    回傳 {id: {name, bday, e_code, abc, phone, mobile, address, dmk_raw, cnt, clinic}}
    """
    members: Dict[str, Dict[str, Any]] = {}
    snames = wb_src.sheetnames
    id_aliases = ["身份證號", "身份證號碼", "身分證號", "身分證號碼", "身份証號", "身分証號",
                  "會員身份証", "會員身份證", "會員身分證", "ID", "家醫收案會員ID"]

    def _get_or_create(pid: str) -> Dict[str, Any]:
        if pid not in members:
            members[pid] = _empty_member()
        return members[pid]

    synthetic_counter = 0

    def _register_member_keys(pid: str, rec: Dict[str, Any]) -> None:
        for kind, key in _iter_member_union_keys(rec):
            if kind == "name_bday":
                name_bday_map.setdefault(key, set()).add(pid)
            elif kind == "name_phone":
                name_phone_map.setdefault(key, set()).add(pid)
            elif kind == "name_only":
                name_only_map.setdefault(key, set()).add(pid)

    def _merge_or_create_noid_member(rec: Dict[str, Any]) -> None:
        nonlocal synthetic_counter
        matched_ids: Optional[set[str]] = None
        for kind, key in _iter_member_union_keys(rec):
            candidate = (
                name_bday_map.get(key, set()) if kind == "name_bday"
                else name_phone_map.get(key, set()) if kind == "name_phone"
                else name_only_map.get(key, set())
            )
            if candidate:
                matched_ids = candidate
                break

        if matched_ids and len(matched_ids) == 1:
            pid = next(iter(matched_ids))
            target = _get_or_create(pid)
        else:
            synthetic_counter += 1
            pid = f"__NOID__{synthetic_counter:06d}"
            target = _get_or_create(pid)

        for key, value in rec.items():
            _fill_member_field(target, key, value)
        _register_member_keys(pid, target)

    name_bday_map: Dict[str, set[str]] = {}
    name_phone_map: Dict[str, set[str]] = {}
    name_only_map: Dict[str, set[str]] = {}

    # ── 1. 月份分頁（最低優先）──────────────────────────────────
    month_sheets = sorted(
        [s for s in snames if _sheet_year_bucket(s) in (114, 115) and _sheet_month(s)],
        key=lambda s: (int(s[:3]), _sheet_month(s))
    )
    processed_sheets = set(month_sheets)
    for sname in month_sheets:
        sh = wb_src[sname]
        partial_map = (partial_maps or {}).get(sname)
        if partial_map is None:
            partial_map = _extract_member_partial_map(sh, id_aliases, search_rows=5)
        for pid, partial in partial_map.items():
            rec = _get_or_create(pid)
            for key, value in partial.items():
                _fill_member_field(rec, key, value)

    # ── 2. ascvd / 115指定會員 / 自選名單 / 115X / P4P（補充）──
    # 115 指定會員與既有會員來源做聯集；既有值不被覆蓋，只補漏列會員與缺值。
    for sname in ["ascvd", "115指定會員", "自選名單", "115X", "P4P收案", "P4P追蹤"]:
        if sname not in snames:
            continue
        processed_sheets.add(sname)
        sh = wb_src[sname]
        partial_map = (partial_maps or {}).get(sname)
        if partial_map is None:
            partial_map = _extract_member_partial_map(sh, id_aliases)
        for pid, partial in partial_map.items():
            rec = _get_or_create(pid)
            for key, value in partial.items():
                _fill_member_field(rec, key, value)

    # ── 3. 其他工作表補齊：只要出現過有效 ID，就先納入主表 ──
    for sname in snames:
        if sname in processed_sheets or sname == "會員名單":
            continue
        sh = wb_src[sname]
        partial_map = (partial_maps or {}).get(sname)
        if partial_map is None:
            partial_map = _extract_member_partial_map(sh, id_aliases)
        for pid, partial in partial_map.items():
            rec = _get_or_create(pid)
            for key, value in partial.items():
                _fill_member_field(rec, key, value)

    # ── 4. 會員名單（最高優先）──────────────────────────────────
    if "會員名單" in snames:
        sh = wb_src["會員名單"]
        header_row = _find_header_row_contains_any(sh, [id_aliases], search_rows=10) or 5
        hmap = build_header_map(sh, header_row)
        id_col_cand = find_column_exact(hmap, id_aliases)
        id_col = find_id_col_by_content(sh, header_row, id_col_cand)
        name_col = find_column_exact(hmap, ["會員姓名", "姓名"])
        bday_col = find_bday_column(hmap)
        phone_cols = [c for hdr, c in hmap.items() if c != id_col and is_phone_header(hdr)]
        addr_col = next((c for hdr, c in hmap.items() if c != id_col and is_address_header(hdr)), None)
        abc_col = find_column_exact(hmap, ["會員別"])
        dmk_col = find_column_exact(hmap, ["疾病樣態"])
        cnt_col = find_column_exact(hmap, ["就診次數"])
        ascvd_col = find_column_exact(hmap, ["ASCVD", "ascvd"])
        extra_cols = _find_member_roster_extra_columns(hmap)
        clinic_val = normalize_text(sh.cell(1, 1).value)
        if id_col:
            data_start_row = _first_valid_id_row(sh, id_col, header_row)
            # 舊版 114 會員名單常沒有標準的 姓名/地址/電話 標題，
            # 但資料固定落在第 6/7/8 欄：姓名 / 地址 / 電話。
            if not name_col and sh.max_column >= 8 and data_start_row <= sh.max_row:
                sample_name = normalize_text(sh.cell(data_start_row, 6).value)
                sample_phone = normalize_phone_value(sh.cell(data_start_row, 8).value)
                if sample_name and sample_phone:
                    name_col = 6
                    if addr_col is None:
                        addr_col = 7
                    if not phone_cols:
                        phone_cols = [8]
            for r in range(data_start_row, sh.max_row + 1):
                pid = normalize_id(sh.cell(r, id_col).value)
                if not pid or not is_valid_tw_id(pid):
                    continue
                rec = _get_or_create(pid)
                contact = pick_contact_from_values([sh.cell(r, c).value for c in phone_cols])
                address = normalize_text(sh.cell(r, addr_col).value) if addr_col else ""
                if name_col:
                    _fill_member_field(rec, "name", normalize_text(sh.cell(r, name_col).value))
                if bday_col:
                    _fill_member_field(rec, "bday", parse_date(sh.cell(r, bday_col).value))
                if abc_col:
                    _fill_member_field(rec, "abc", sh.cell(r, abc_col).value)
                if dmk_col:
                    _fill_member_field(rec, "e_code", parse_disease_code(sh.cell(r, dmk_col).value))
                    _fill_member_field(rec, "dmk_raw", sh.cell(r, dmk_col).value)
                if cnt_col:
                    _fill_member_field(rec, "cnt", sh.cell(r, cnt_col).value)
                if ascvd_col:
                    ascvd_val = sh.cell(r, ascvd_col).value
                    if ascvd_val is not None and str(ascvd_val).strip() != "":
                        _fill_member_field(rec, "ascvd", ascvd_val)
                _fill_member_roster_extra_fields(rec, sh, r, extra_cols)
                _fill_member_field(rec, "phone", contact.phone)
                _fill_member_field(rec, "mobile", contact.mobile)
                _fill_member_field(rec, "address", address or None)
                _fill_member_field(rec, "clinic", clinic_val)

    for pid, rec in members.items():
        _register_member_keys(pid, rec)

    for sname in snames:
        sh = wb_src[sname]
        for rec in _extract_member_noid_rows(sh, id_aliases):
            _merge_or_create_noid_member(rec)

    return members


# ============================================================
# 聯絡資料整合（手機 / 電話 / 地址）
# ============================================================
def build_contact_map(sh_phone: Any) -> Dict[str, ContactInfo]:
    """從聯絡資料 sheet 建立 {id: ContactInfo} 對應。"""
    if sh_phone is None:
        return {}
    id_aliases = ["ID", "身份證號", "身分證號", "身份証號"]
    header_row = _find_header_row_contains_any(sh_phone, [id_aliases], search_rows=5)
    if header_row is None:
        header_row = 1
    hmap = build_header_map(sh_phone, header_row)
    id_col_cand = find_column_exact(hmap, id_aliases)
    id_col = find_id_col_by_content(sh_phone, header_row, id_col_cand)
    phone_cols = [
        c for hdr, c in hmap.items()
        if c != id_col and is_phone_header(hdr)
    ]
    addr_col = next(
        (c for hdr, c in hmap.items() if c != id_col and is_address_header(hdr)),
        None,
    )
    if id_col is None or (not phone_cols and addr_col is None):
        return {}
    result: Dict[str, ContactInfo] = {}
    for r in range(header_row + 1, sh_phone.max_row + 1):
        pid = normalize_id(sh_phone.cell(r, id_col).value)
        if not pid or not is_valid_tw_id(pid):
            continue
        contact = pick_contact_from_values([sh_phone.cell(r, c).value for c in phone_cols])
        address = normalize_text(sh_phone.cell(r, addr_col).value) if addr_col else ""
        if contact.phone or contact.mobile or address:
            incoming = ContactInfo(
                phone=contact.phone,
                mobile=contact.mobile,
                address=address or None,
            )
            result[pid] = merge_contact_info(result.get(pid), incoming)
    return result


def _log_duplicate_ids(sheet: Any, label: str) -> None:
    if sheet is None:
        return
    id_aliases = ["家醫收案會員ID", "ID", "身份證號", "身份證號碼", "身分證號", "身分證號碼"]
    hmap = build_header_map(sheet, 1)
    id_col = find_id_col_by_content(sheet, 1, find_column_exact(hmap, id_aliases))
    if id_col is None:
        return

    counts: Dict[str, int] = {}
    total = 0
    for r in range(2, sheet.max_row + 1):
        pid = normalize_id(sheet.cell(r, id_col).value)
        if not pid or not is_valid_tw_id(pid):
            continue
        total += 1
        counts[pid] = counts.get(pid, 0) + 1

    dup_ids = sorted(pid for pid, cnt in counts.items() if cnt > 1)
    if not dup_ids:
        return

    dup_rows = sum(counts[pid] - 1 for pid in dup_ids)
    sample = "、".join(dup_ids[:10])
    msg = f"{label} 有重複ID：原始{total}筆，唯一{len(counts)}筆，重複ID {len(dup_ids)}個，重複列 {dup_rows}筆"
    if sample:
        msg += f"；例如：{sample}"
    _log(msg)


# ============================================================
# P4P / 自選 / 114 / 115X 旗標
# ============================================================
def _p4p_status_priority(status: Any) -> int:
    text = normalize_text(status)
    if "本院收案" in text:
        return 0
    if "未收案" in text:
        return 1
    if "外院收案" in text or "外院收" in text:
        return 2
    return 3


def _sort_p4p_records(records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return sorted(
        records,
        key=lambda rec: (
            _p4p_status_priority(rec.get("status")),
            normalize_text(rec.get("plan")),
            normalize_text(rec.get("status")),
        ),
    )


def _upsert_p4p_record(
    records: List[Dict[str, Any]],
    *,
    plan: str = "",
    status: str = "",
    enroll_dt: Any = None,
    last_track_dt: Any = None,
    next_track_dt: Any = None,
    overdue: str = "",
) -> None:
    if not any((plan, status, enroll_dt, last_track_dt, next_track_dt, overdue)):
        return
    match = None
    if plan:
        for rec in records:
            if normalize_text(rec.get("plan")) == plan:
                match = rec
                break
    if match is None:
        match = {}
        records.append(match)
    if plan:
        match["plan"] = plan
    if status:
        match["status"] = status
    if enroll_dt is not None:
        match["enroll_dt"] = enroll_dt
    if last_track_dt is not None:
        match["last_track_dt"] = last_track_dt
    if next_track_dt is not None:
        match["next_track_dt"] = next_track_dt
    if overdue:
        match["overdue"] = overdue


def _finalize_p4p_member_record(member_p4p: Dict[str, Any]) -> Dict[str, Any]:
    records = _sort_p4p_records(member_p4p.get("records", []))
    member_p4p["records"] = records
    primary = records[0] if records else {}
    for key in ("plan", "status", "enroll_dt", "last_track_dt", "next_track_dt", "overdue"):
        if primary.get(key) not in (None, ""):
            member_p4p[key] = primary.get(key)
    return member_p4p


def format_p4p_record_display(record: Dict[str, Any]) -> str:
    status = normalize_text(record.get("status"))
    plan = normalize_text(record.get("plan"))
    if status and plan:
        return f"{status}({plan})"
    return status or plan


def build_p4p_map(sh_enroll: Any, sh_track: Any) -> Dict[str, Dict[str, Any]]:
    """
    合併 P4P收案 + P4P追蹤，回傳
    {id: {plan, status, enroll_dt, last_track_dt, next_track_dt, overdue, records}}
    records 保留同一病人的多筆 P4P 收案資料，並以本院收案優先排序。
    """
    result: Dict[str, Dict[str, Any]] = {}
    id_aliases = ["家醫收案會員ID", "ID", "身份證號", "身分證號"]

    # P4P收案：取收案計畫 / 收案狀態
    if sh_enroll is not None:
        hmap = build_header_map(sh_enroll, 1)
        id_col = find_id_col_by_content(sh_enroll, 1, find_column_exact(hmap, id_aliases))
        plan_col = find_column_exact(hmap, ["P4P收案計畫"])
        status_col = find_column_exact(hmap, ["收案狀態"])
        if id_col:
            for r in range(2, sh_enroll.max_row + 1):
                pid = normalize_id(sh_enroll.cell(r, id_col).value)
                if not pid or not is_valid_tw_id(pid):
                    continue
                d = result.setdefault(pid, {"records": []})
                plan = normalize_text(sh_enroll.cell(r, plan_col).value) if plan_col else ""
                status = normalize_text(sh_enroll.cell(r, status_col).value) if status_col else ""
                _upsert_p4p_record(d["records"], plan=plan, status=status)

    # P4P追蹤：取收案計畫 / 日期欄 / 逾期狀態
    if sh_track is not None:
        hmap = build_header_map(sh_track, 1)
        id_col = find_id_col_by_content(sh_track, 1, find_column_exact(hmap, id_aliases))
        plan_col       = find_column_exact(hmap, ["P4P收案計畫"])
        enroll_col     = find_column_exact(hmap, ["收案日期"])
        last_col       = find_column_exact(hmap, ["最後追蹤日"])
        next_col       = find_column_exact(hmap, ["下次應追蹤日"])
        status_col     = find_column_exact(hmap, ["收案狀態"])
        overdue_col    = find_column_exact(hmap, ["逾期未追蹤"])
        if id_col:
            for r in range(2, sh_track.max_row + 1):
                pid = normalize_id(sh_track.cell(r, id_col).value)
                if not pid or not is_valid_tw_id(pid):
                    continue
                d = result.setdefault(pid, {"records": []})
                plan = normalize_text(sh_track.cell(r, plan_col).value) if plan_col else ""
                status = normalize_text(sh_track.cell(r, status_col).value) if status_col else ""
                enroll_dt = parse_date(sh_track.cell(r, enroll_col).value) if enroll_col else None
                last_track_dt = parse_date(sh_track.cell(r, last_col).value) if last_col else None
                next_track_dt = parse_date(sh_track.cell(r, next_col).value) if next_col else None
                overdue = normalize_text(sh_track.cell(r, overdue_col).value) if overdue_col else ""
                _upsert_p4p_record(
                    d["records"],
                    plan=plan,
                    status=status,
                    enroll_dt=enroll_dt,
                    last_track_dt=last_track_dt,
                    next_track_dt=next_track_dt,
                    overdue=overdue,
                )

    for member_p4p in result.values():
        _finalize_p4p_member_record(member_p4p)
    return result


def build_id_set(sheet: Any, id_aliases: List[str]) -> set:
    """從 sheet 建立 ID set（自動偵測 header row）"""
    if sheet is None:
        return set()
    # 先嘗試找含有 ID 關鍵字的 header row
    header_row = _find_header_row_contains_any(sheet, [id_aliases], search_rows=10)
    if header_row is None:
        header_row = 1
    hmap = build_header_map(sheet, header_row)
    id_col = find_id_col_by_content(sheet, header_row, find_column_exact(hmap, id_aliases))
    if id_col is None:
        return set()
    result = set()
    for r in range(header_row + 1, sheet.max_row + 1):
        pid = normalize_id(sheet.cell(r, id_col).value)
        if pid and is_valid_tw_id(pid):
            result.add(pid)
    return result


def _find_sheet_id_col(sheet: Any, id_aliases: List[str], search_rows: int = 10) -> Tuple[Optional[int], int]:
    if sheet is None:
        return None, 1
    header_row = _find_header_row_contains_any(sheet, [id_aliases], search_rows=search_rows)
    if header_row is None:
        header_row = 1
    hmap = build_header_map(sheet, header_row)
    id_col = find_id_col_by_content(sheet, header_row, find_column_exact(hmap, id_aliases))
    return id_col, header_row


def _normalize_match_name(value: Any) -> str:
    return clean_spaces(value).upper()


def _bday_match_token(value: Any) -> Optional[str]:
    dt = parse_date(value)
    if not dt:
        return None
    return dt.strftime("%Y-%m-%d")


def _phone_match_tokens(phone: Optional[str]) -> List[str]:
    if not phone:
        return []
    if len(phone) >= 7:
        return [phone[-7:]]
    return []


def _compose_match_keys(name: Any, bday: Any, phones: List[Any]) -> set[str]:
    # 來源端規則：
    # 1. 有生日 -> 只用 姓名+生日
    # 2. 沒生日但有電話 -> 用 姓名+電話末7碼
    # 3. 都沒有 -> 才退回只用姓名
    normalized_name = _normalize_match_name(name)
    if not normalized_name:
        return set()

    bday_token = _bday_match_token(bday)
    if bday_token:
        return {f"{normalized_name}|BDAY|{bday_token}"}

    keys: set[str] = set()
    found_phone = False
    for raw in phones:
        phone = normalize_phone_value(raw)
        if phone:
            found_phone = True
        for token in _phone_match_tokens(phone):
            keys.add(f"{normalized_name}|PHONE|{token}")
    if not found_phone:
        keys.add(normalized_name)
    return keys


def _iter_source_match_rows(sheet: Any, id_aliases: List[str]) -> List[Tuple[Any, Any, List[Any]]]:
    if sheet is None:
        return []

    id_col, header_row = _find_sheet_id_col(sheet, id_aliases, search_rows=10)
    if id_col is not None:
        return []

    hmap = build_header_map(sheet, header_row)
    name_col = find_column_exact(hmap, ["姓名", "會員姓名", "名字"])
    if not name_col:
        return []

    bday_col = find_bday_column(hmap)
    phone_cols = [c for hdr, c in hmap.items() if c != name_col and is_phone_header(hdr)]
    result: List[Tuple[Any, Any, List[Any]]] = []
    for r in range(header_row + 1, sheet.max_row + 1):
        result.append(
            (
                sheet.cell(r, name_col).value,
                sheet.cell(r, bday_col).value if bday_col else None,
                [sheet.cell(r, c).value for c in phone_cols],
            )
        )
    return result


def _build_member_match_maps(all_members: Dict[str, Dict[str, Any]]) -> Tuple[Dict[str, set[str]], Dict[str, set[str]], Dict[str, set[str]]]:
    name_bday_map: Dict[str, set[str]] = {}
    name_phone_map: Dict[str, set[str]] = {}
    name_only_map: Dict[str, set[str]] = {}
    for pid, rec in all_members.items():
        for kind, key in _iter_member_union_keys(rec):
            if kind == "name_bday":
                name_bday_map.setdefault(key, set()).add(pid)
            elif kind == "name_phone":
                name_phone_map.setdefault(key, set()).add(pid)
            elif kind == "name_only":
                name_only_map.setdefault(key, set()).add(pid)
    return name_bday_map, name_phone_map, name_only_map


def _match_source_row_to_member_ids(
    name: Any,
    bday: Any,
    phones: List[Any],
    name_bday_map: Dict[str, set[str]],
    name_phone_map: Dict[str, set[str]],
    name_only_map: Dict[str, set[str]],
) -> Optional[set[str]]:
    matched_ids: Optional[set[str]] = None
    for key in _compose_match_keys(name, bday, phones):
        if "|BDAY|" in key:
            candidate = name_bday_map.get(key, set())
        elif "|PHONE|" in key:
            candidate = name_phone_map.get(key, set())
        else:
            candidate = name_only_map.get(key, set())
        if candidate:
            matched_ids = candidate
            break

    if (not matched_ids or len(matched_ids) != 1) and name:
        name_key = _normalize_match_name(name)
        if name_key:
            candidate = name_only_map.get(name_key, set())
            if len(candidate) == 1:
                matched_ids = candidate
    return matched_ids


def build_member_key_set_from_source(
    all_members: Dict[str, Dict[str, Any]],
    sheet: Any,
    id_aliases: List[str],
) -> set[str]:
    if sheet is None:
        return set()

    result: set[str] = set()
    direct_ids = build_id_set(sheet, id_aliases)
    for pid in direct_ids:
        if pid in all_members:
            result.add(pid)

    match_rows = _iter_source_match_rows(sheet, id_aliases)
    if not match_rows:
        return result

    name_bday_map, name_phone_map, name_only_map = _build_member_match_maps(all_members)
    for row_name, row_bday, row_phones in match_rows:
        matched_ids = _match_source_row_to_member_ids(
            row_name,
            row_bday,
            row_phones,
            name_bday_map,
            name_phone_map,
            name_only_map,
        )
        if matched_ids and len(matched_ids) == 1:
            result.add(next(iter(matched_ids)))
    return result


def build_member_key_set_from_sources(
    all_members: Dict[str, Dict[str, Any]],
    sheets: List[Any],
    id_aliases: List[str],
) -> set[str]:
    result: set[str] = set()
    for sheet in sheets:
        result.update(build_member_key_set_from_source(all_members, sheet, id_aliases))
    return result


def build_ascvd_member_id_set(sheet: Any, id_aliases: List[str]) -> set:
    """從會員名單抓 ID set，作為沒有獨立 ASCVD sheet 時的 114 名單來源。"""
    if sheet is None:
        return set()
    header_row = _find_header_row_contains_any(sheet, [id_aliases], search_rows=10)
    if header_row is None:
        header_row = 1
    hmap = build_header_map(sheet, header_row)
    id_col = find_id_col_by_content(sheet, header_row, find_column_exact(hmap, id_aliases))
    if id_col is None:
        return set()
    result = set()
    for r in range(header_row + 1, sheet.max_row + 1):
        pid = normalize_id(sheet.cell(r, id_col).value)
        if pid and is_valid_tw_id(pid):
            result.add(pid)
    return result


def _fill_extra_flags(
    ws,
    cols: Dict[str, Optional[int]],
    data_start: int,
    last_row: int,
    id_to_rows: Dict[str, List[int]],
    p4p_map: Dict[str, Dict[str, Any]],
    member_ids: set,
    self_select_ids: set,
    x115_ids: set,
) -> None:
    """填入後段旗標欄位。
    member_ids: 會員名單／較需要照護名單的 ID set，用於標記「是否為114會員名單」。
    """
    col_at = cols.get("p4p_status")
    col_plan = cols.get("p4p_plan")
    col_au = cols.get("p4p_enroll_dt")
    col_av = cols.get("p4p_last_dt")
    col_aw = cols.get("p4p_next_dt")
    col_overdue = cols.get("p4p_overdue")
    col_ax = cols.get("is_114")
    col_ay = cols.get("is_self_select")
    col_az = cols.get("is_115x")

    for pid, rows in id_to_rows.items():
        p4p = p4p_map.get(pid, {})
        is_114       = "✔" if pid in member_ids       else None
        is_self      = "✔" if pid in self_select_ids  else None
        is_115x      = "✔" if pid in x115_ids         else None
        plan         = p4p.get("plan") or None
        status       = p4p.get("status") or None
        enroll_dt    = p4p.get("enroll_dt")
        last_dt      = p4p.get("last_track_dt")
        next_dt      = p4p.get("next_track_dt")
        overdue      = p4p.get("overdue") or None

        for rr in rows:
            safe_set(ws, rr, col_plan, plan)
            safe_set(ws, rr, col_at, status)
            safe_set(ws, rr, col_au, enroll_dt)
            safe_set(ws, rr, col_av, last_dt)
            safe_set(ws, rr, col_aw, next_dt)
            safe_set(ws, rr, col_overdue, overdue)
            safe_set_check(ws, rr, col_ax, is_114)
            safe_set_check(ws, rr, col_ay, is_self)
            safe_set_check(ws, rr, col_az, is_115x)


def _recompute_member_category(
    ws,
    cols: Dict[str, Optional[int]],
    id_to_rows: Dict[str, List[int]],
    member_ids: set,
    self_select_ids: set,
    x115_ids: set,
) -> None:
    col_abc = cols.get("abc")
    col_dmk = cols.get("dmk_code")
    col_ascvd = cols.get("ascvd")
    if not col_abc:
        return

    for pid, rows in id_to_rows.items():
        in_member = pid in member_ids
        in_self_select = pid in self_select_ids
        in_115x = pid in x115_ids

        base = "D"
        if in_member:
            row0 = rows[0]
            e_code = parse_disease_code(ws.cell(row0, col_dmk).value) if col_dmk else None
            ascvd = parse_ascvd(ws.cell(row0, col_ascvd).value) if col_ascvd else AscvdCategory.NONE
            if e_code in (DiseaseCode.DM, DiseaseCode.CKD, DiseaseCode.DKD) or ascvd in (AscvdCategory.A, AscvdCategory.B):
                base = "A"
            else:
                base = "B"
        elif in_115x:
            base = "E2"
        elif in_self_select:
            base = "E1"

        if in_115x and base in ("A", "B"):
            value = f"{base}/E2"
        elif in_self_select and base in ("A", "B"):
            value = f"{base}/E1"
        else:
            value = base

        for rr in rows:
            safe_set(ws, rr, col_abc, value)

# ============================================================
# 資料填充子函數
# ============================================================
def _load_and_validate_source(
    source_path: str,
    profile: Optional[ProcessingProfile] = None,
):
    profile = _resolve_profile(profile)
    wb = profile.load_source_workbook(source_path)
    need = {
        "HealthCase": ("HealthCase", "健康管理列表", "個案健康管理列表"),
        "成人健檢": ("成人健檢", "成人預防保健"),
        "子宮抹片": ("子宮抹片", "子宮頸抹片"),
        "老人流感": ("老人流感", "65歲流感", "流感疫苗"),
        "糞便潛血": ("糞便潛血",),
        "肝炎篩檢": ("肝炎篩檢", "BC肝", "B肝C肝"),
    }
    missing = [label for label, aliases in need.items() if _first_sheet(wb, *aliases) is None]
    if missing:
        raise ValueError(f"原始檔缺少工作表：{'、'.join(missing)}")
    return wb


def _is_excel_source_file(filename: str) -> bool:
    ext = os.path.splitext(filename)[1].lower()
    base = os.path.basename(filename)
    if base.startswith("~$"):
        return False
    if re.match(r"^選會員\d{4}_\d{4}\.xlsx$", base):
        return False
    return ext in (".xlsx", ".xls", ".xlsm", ".xltx", ".xltm", ".csv", ".ods")


def _iter_source_files(folder_path: str) -> List[str]:
    files: List[str] = []
    for root, dirnames, filenames in os.walk(folder_path):
        dirnames.sort()
        for name in sorted(filenames):
            full = os.path.join(root, name)
            if os.path.isfile(full) and _is_excel_source_file(name):
                files.append(full)
    return sorted(
        files,
        key=lambda p: (
            HIS_MONTHLY_FEE_FOLDER_NAME in str(Path(p).parts),
            p,
        ),
    )


def _append_sheet_rows(src_ws, dst_ws, skip_header: bool = False) -> None:
    start_row = 2 if skip_header else 1
    for row in src_ws.iter_rows(min_row=start_row, values_only=True):
        dst_ws.append(list(row))


def detect_encoding(
    csv_path: str,
    encodings: Tuple[str, ...] = ("utf-16", "utf-16le", "utf-8-sig", "utf-8"),
) -> str:
    """
    優先判斷 Unicode 類 CSV。
    若無法判斷，呼叫端再退回 HIS 常見的 cp950 相容模式。
    """
    with open(csv_path, "rb") as f:
        raw = f.read(4)
    if raw.startswith(b"\xff\xfe") or raw.startswith(b"\xfe\xff"):
        return "utf-16"
    if raw.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"

    last_error: Optional[Exception] = None
    for encoding in encodings:
        try:
            with open(csv_path, "r", encoding=encoding, newline="") as f:
                for _ in f:
                    pass
            return encoding
        except UnicodeDecodeError as exc:
            last_error = exc
    raise RuntimeError(f"無法判斷 {os.path.basename(csv_path)} 的編碼: {last_error}") from last_error


def _load_csv_as_workbook(csv_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = os.path.splitext(os.path.basename(csv_path))[0][:31] or "Sheet1"

    try:
        try:
            enc = detect_encoding(csv_path)
            errors = "strict"
        except RuntimeError:
            # HIS 匯出的傳統中文 CSV 常混有少數不標準位元組，
            # 退回 cp950 相容模式可最大化保留內容。
            enc = "cp950"
            errors = "replace"

        with open(csv_path, "r", encoding=enc, errors=errors, newline="") as f:
            for row in csv.reader(f):
                ws.append(row)
        return wb
    except Exception as e:
        raise ValueError(f"CSV 檔案無法以支援編碼讀取：{os.path.basename(csv_path)}") from e


def _load_ods_as_workbook(ods_path: str):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    office_ns = "urn:oasis:names:tc:opendocument:xmlns:office:1.0"
    table_ns = "urn:oasis:names:tc:opendocument:xmlns:table:1.0"
    text_ns = "urn:oasis:names:tc:opendocument:xmlns:text:1.0"
    q_spreadsheet = f"{{{office_ns}}}spreadsheet"
    q_table = f"{{{table_ns}}}table"
    q_row = f"{{{table_ns}}}table-row"
    q_cell = f"{{{table_ns}}}table-cell"
    q_covered_cell = f"{{{table_ns}}}covered-table-cell"
    q_name = f"{{{table_ns}}}name"
    q_col_repeat = f"{{{table_ns}}}number-columns-repeated"
    q_row_repeat = f"{{{table_ns}}}number-rows-repeated"
    q_paragraph = f"{{{text_ns}}}p"

    with zipfile.ZipFile(ods_path, "r") as zf:
        root = ET.fromstring(zf.read("content.xml"))

    spreadsheet = root.find(f".//{q_spreadsheet}")
    if spreadsheet is None:
        ws = wb.create_sheet(title="Sheet1")
        return wb

    table_count = 0
    for idx, table in enumerate(spreadsheet, start=1):
        if table.tag != q_table:
            continue
        table_count += 1
        title = str(table.attrib.get(q_name) or f"Sheet{idx}")
        ws = wb.create_sheet(title=title[:31] or f"Sheet{idx}")
        for row in table:
            if row.tag != q_row:
                continue

            values: List[str] = []
            pending_blank_repeat = 0
            for cell in row:
                if cell.tag not in (q_cell, q_covered_cell):
                    continue
                repeat = int(cell.attrib.get(q_col_repeat, "1") or "1")
                text_parts = [
                    "".join(paragraph.itertext()).strip()
                    for paragraph in cell.iter(q_paragraph)
                ]
                text_parts = [part for part in text_parts if part]
                text = "\n".join(text_parts)
                if text:
                    if pending_blank_repeat:
                        values.extend([""] * pending_blank_repeat)
                        pending_blank_repeat = 0
                    values.extend([text] * repeat)
                else:
                    pending_blank_repeat += repeat

            if not values:
                continue

            row_repeat = int(row.attrib.get(q_row_repeat, "1") or "1")
            ws.append(values)
            if row_repeat > 1:
                values_copy = list(values)
                for _ in range(row_repeat - 1):
                    ws.append(values_copy)

        if table_count == 0:
            ws = wb.create_sheet(title="Sheet1")

    if table_count == 0:
        ws = wb.create_sheet(title="Sheet1")
    return wb


def _strip_worksheet_autofilters(xlsx_path: str) -> str:
    fd, repaired_path = tempfile.mkstemp(prefix="repair_", suffix=".xlsx")
    os.close(fd)
    main_ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"

    with zipfile.ZipFile(xlsx_path, "r") as zin, zipfile.ZipFile(repaired_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            data = zin.read(info.filename)
            if info.filename.startswith("xl/worksheets/") and info.filename.endswith(".xml"):
                try:
                    root = ET.fromstring(data)
                    changed = False
                    for parent in root.iter():
                        for child in list(parent):
                            if child.tag == f"{main_ns}autoFilter":
                                parent.remove(child)
                                changed = True
                    if changed:
                        data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                except Exception:
                    pass
            zout.writestr(info, data)

    return repaired_path


def _load_xlsx_as_workbook(xlsx_path: str):
    try:
        return openpyxl.load_workbook(xlsx_path, data_only=True)
    except ValueError as e:
        msg = str(e)
        if "could not read worksheets" not in msg and "wildcard" not in msg:
            raise
        repaired_path = _strip_worksheet_autofilters(xlsx_path)
        try:
            return openpyxl.load_workbook(repaired_path, data_only=True)
        finally:
            try:
                os.unlink(repaired_path)
            except OSError:
                pass


def _load_xls_as_workbook(xls_path: str):
    book = xlrd.open_workbook(xls_path)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for idx, sheet in enumerate(book.sheets(), start=1):
        title = str(sheet.name or f"Sheet{idx}")
        ws = wb.create_sheet(title=title[:31] or f"Sheet{idx}")
        for row_idx in range(sheet.nrows):
            values = [sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)]
            ws.append(values)
    return wb


def _normalize_sheet_lookup(text: Any) -> str:
    s = normalize_text(text)
    if not s:
        return ""
    s = s.lower()
    s = re.sub(r"[\s\-_()（）\[\]{}]+", "", s)
    return s


def _is_generic_sheet_title(title: str) -> bool:
    norm = _normalize_sheet_lookup(title)
    return bool(re.fullmatch(r"(sheet|工作表)\d*", norm))


_SOURCE_SHEET_ALIASES: Dict[str, Tuple[str, ...]] = {
    "會員名單": ("會員名單", "較需要照護名單", "家醫計畫較需要照護名單"),
    "114會員名單補充": ("114會員名單補充",),
    "ascvd": ("ascvd",),
    "HealthCase": ("healthcase", "健康管理列表", "個案健康管理列表"),
    "成人健檢": ("成人健檢", "成人預防保健"),
    "子宮抹片": ("子宮抹片", "子宮頸抹片", "抹片"),
    "老人流感": ("老人流感", "65歲流感", "流感疫苗"),
    "糞便潛血": ("糞便潛血", "潛血"),
    "肝炎篩檢": ("肝炎篩檢", "bc肝", "b肝c肝", "肝篩檢"),
    "自選名單": ("自選名單", "自選會員", "第一階段"),
    "P4P收案": ("p4p收案", "收案管理", "p4pcase"),
    "P4P追蹤": ("p4p追蹤", "追蹤管理", "p4ptrack"),
    "115X": ("115x",),
    "主次診斷": ("主次診斷", "主次診段"),
    "行動電話": ("行動電話", "基本資料檔列印", "基本資料列印"),
}


def _classify_r13210_file(file_path: str) -> Optional[str]:
    norm_file = _normalize_sheet_lookup(os.path.splitext(os.path.basename(file_path))[0])
    has_member_gate = ("會員" in norm_file) or ("選" in norm_file)
    has_self_select_token = any(token in norm_file for token in ("自選", "預選"))
    has_115_token = "115" in norm_file and "1150" not in norm_file and "1151" not in norm_file

    # 檔名分類規則：
    # 0. 先決條件：檔名需包含「會員」或「選」
    # 1. 檔名包含 [115x / 115X / 不選] 視為不選會員
    # 2. 其餘包含 [115 / 自選 / 預選] 視為自選會員
    #    但 1150 / 1151（例如 11501 月報）不視為自選判斷依據
    if not has_member_gate:
        return None

    if "115x" in norm_file or "不選" in norm_file or "不要" in norm_file:
        return "115X"

    if has_115_token or has_self_select_token:
        return "自選名單"

    if "r13210" not in norm_file:
        return None

    raise ValueError(
        f"檔名無法判定為預選或不選：{os.path.basename(file_path)}\n"
        "請更新檔名內容明確符合 115X 或 115 自選規則後，再重新進行處理。"
    )


def _classify_indexno_sheet(src_ws: Any) -> Optional[str]:
    try:
        header = normalize_text(src_ws.cell(1, 1).value)
        indicator = normalize_text(src_ws.cell(2, 1).value)
    except Exception:
        return None

    if header != "指標名稱" or not indicator:
        return None

    if "成人預防保健" in indicator or "成人健檢" in indicator:
        return "成人健檢"
    if "子宮頸抹片" in indicator or "子宮抹片" in indicator:
        return "子宮抹片"
    if "流感" in indicator:
        return "老人流感"
    if "糞便潛血" in indicator or "潛血" in indicator:
        return "糞便潛血"
    if "肝炎篩檢" in indicator or "bc肝" in indicator or "b、c肝" in indicator.lower():
        return "肝炎篩檢"
    return None


def _looks_like_member_roster_sheet(src_ws: Any) -> bool:
    id_aliases = ["ID", "身份證號", "身份證號碼", "身分證號", "身分證號碼"]
    header_row = _find_header_row_contains_any(src_ws, [id_aliases, ["疾病樣態"]], search_rows=10)
    return header_row is not None


def _looks_like_member_roster_with_ascvd_sheet(src_ws: Any) -> bool:
    id_aliases = ["ID", "身份證號", "身份證號碼", "身分證號", "身分證號碼"]
    # 單列 header：ID + 疾病樣態 + ASCVD 在同一列（常見格式）
    if _find_header_row_contains_any(src_ws, [id_aliases, ["疾病樣態"], ["ASCVD", "ascvd"]], search_rows=10) is not None:
        return True
    # 雙列 header：耀聖格式，ID/最後就診日在第1列，ASCVD 在第2列
    has_id = _find_header_row_contains_any(src_ws, [id_aliases], search_rows=5) is not None
    has_ascvd = _find_header_row_contains_any(src_ws, [["ASCVD", "ascvd"]], search_rows=5) is not None
    return has_id and has_ascvd


def _canonical_source_sheet_name(sheet_name: str, file_path: str, single_sheet: bool, src_ws: Any = None) -> str:
    norm_sheet = _normalize_sheet_lookup(sheet_name)
    norm_file = _normalize_sheet_lookup(os.path.splitext(os.path.basename(file_path))[0])
    base_name = os.path.basename(file_path)

    file_stem = os.path.splitext(os.path.basename(file_path))[0]
    if re.fullmatch(r"1(14|15)\d{2}", file_stem):
        return file_stem

    if re.fullmatch(r"1(14|15)\d{2}", str(sheet_name)):
        return str(sheet_name)

    if (
        any(token in norm_file for token in ("115指定會員", "115年指定會員"))
        and (single_sheet or _is_generic_sheet_title(sheet_name))
    ):
        return "115指定會員"

    if src_ws is not None and _looks_like_member_roster_with_ascvd_sheet(src_ws):
        return "ascvd"

    if "需照護名單" in os.path.basename(file_path) or "需要照護名單" in os.path.basename(file_path):
        if src_ws is not None and _looks_like_member_roster_with_ascvd_sheet(src_ws):
            return "ascvd"
        if src_ws is not None and _looks_like_member_roster_sheet(src_ws):
            return "會員名單"
        return "會員名單"

    if (
        "a115" in norm_file
        and "115x" not in norm_file
        and any(token in norm_file for token in ("預選會員", "自選會員", "自選名單", "預選"))
        and (single_sheet or _is_generic_sheet_title(sheet_name))
    ):
        return "自選名單115"

    if (
        "114" in norm_file
        and "會員" in norm_file
        and "115" not in norm_file
        and not any(token in norm_file for token in ("自選", "預選", "115x", "a115"))
        and (single_sheet or _is_generic_sheet_title(sheet_name))
    ):
        return "114會員名單補充"

    r13210_kind = _classify_r13210_file(file_path)
    if r13210_kind and (
        single_sheet
        or _is_generic_sheet_title(sheet_name)
        or "基本資料檔列印" in sheet_name
        or "特記分析報表" in sheet_name
    ):
        if r13210_kind == "自選名單" and ("a115" in norm_file or "預選會員" in base_name):
            return "自選名單115"
        return r13210_kind

    if any(token in norm_file for token in ("自選會員", "自選名單", "預選")):
        if single_sheet or _is_generic_sheet_title(sheet_name):
            return "自選名單"

    for canonical, aliases in _SOURCE_SHEET_ALIASES.items():
        if any(alias in norm_sheet for alias in aliases):
            return canonical

    if "家醫計畫" in os.path.basename(file_path):
        if src_ws is not None and _looks_like_member_roster_sheet(src_ws):
            return "會員名單"

    if src_ws is not None:
        indicator_canonical = _classify_indexno_sheet(src_ws)
        if indicator_canonical:
            return indicator_canonical

    if single_sheet or _is_generic_sheet_title(sheet_name):
        for canonical, aliases in _SOURCE_SHEET_ALIASES.items():
            if any(alias in norm_file for alias in aliases):
                return canonical

    return sheet_name


def _format_mapped_pairs(mapped_pairs: List[str], limit: int = 20) -> str:
    if not mapped_pairs:
        return ""

    grouped: Dict[str, List[str]] = {}
    for item in mapped_pairs[:limit]:
        left, dst_name = item.rsplit("->", 1)
        grouped.setdefault(dst_name, []).append(left)

    lines: List[str] = []
    for dst_name in sorted(grouped):
        sources = sorted(grouped[dst_name])
        preview = "、".join(sources[:3])
        extra = f" 等{len(sources)}筆" if len(sources) > 3 else ""
        lines.append(f"- {dst_name}：{preview}{extra}")
    return "\n".join(lines)


def _merge_source_folder(
    folder_path: str,
    profile: Optional[ProcessingProfile] = None,
):
    """
    將資料夾內所有 Excel 合併成一個暫時 workbook：
    - 同名 sheet：保留第一個檔案的表頭，後續檔案從第 2 列開始 append
    - 不同 sheet：直接新增
    """
    profile = _resolve_profile(profile)
    files = _iter_source_files(folder_path)
    if not files:
        raise ValueError("所選資料夾內找不到可用的 Excel 檔案（支援 .xlsx/.xlsm）")

    merged_wb = openpyxl.Workbook()
    merged_wb.remove(merged_wb.active)
    created = set()
    loaded_files: List[str] = []
    skipped_files: List[str] = []
    mapped_pairs: List[str] = []
    found_r13210 = False

    for path in files:
        base = os.path.basename(path)
        if _classify_r13210_file(path) is not None:
            found_r13210 = True
        try:
            wb = profile.load_source_workbook(path)
        except Exception:
            skipped_files.append(base)
            continue

        loaded_files.append(base)
        single_sheet = len(wb.sheetnames) == 1
        for sname in wb.sheetnames:
            src_ws = wb[sname]
            dst_name = profile.canonical_source_sheet_name(sname, path, single_sheet, src_ws)
            if dst_name != sname:
                mapped_pairs.append(f"{base}:{sname}->{dst_name}")
            if dst_name not in created:
                dst_ws = merged_wb.create_sheet(title=dst_name)
                _append_sheet_rows(src_ws, dst_ws, skip_header=False)
                created.add(dst_name)
            else:
                dst_ws = merged_wb[dst_name]
                _append_sheet_rows(src_ws, dst_ws, skip_header=True)

    if not loaded_files:
        raise ValueError("資料夾內的 Excel 檔案都無法讀取")

    need = ["HealthCase", "成人健檢", "子宮抹片", "老人流感", "糞便潛血", "肝炎篩檢"]
    missing = [s for s in need if s not in merged_wb.sheetnames]
    if missing:
        raise ValueError(
            "合併資料夾後仍缺少必要工作表：" + "、".join(missing) +
            "\n\n已讀取檔案：" + "、".join(loaded_files) +
            ("\n\n已自動對應：\n" + _format_mapped_pairs(mapped_pairs) if mapped_pairs else "") +
            ("\n略過檔案：" + "、".join(skipped_files) if skipped_files else "")
        )

    if mapped_pairs:
        _log("自動對應工作表：\n" + _format_mapped_pairs(mapped_pairs))
    if not found_r13210:
        _log("沒有預選/不選會員檔案")

    return merged_wb


def _fill_member_basic(
    ws,
    all_members: Dict[str, Dict[str, Any]],
    contact_map: Dict[str, ContactInfo],
    cols:        Dict[str, Optional[int]],
    data_start:  int,
    now:         datetime.date,
    clinic_val:  str = "",
) -> Tuple[Dict[str, List[int]], Dict[int, MemberMeta], int]:
    """
    v3：從 all_members（已整合所有 sheet）逐一寫入總表。
    聯絡資料以 contact_map 優先，其次用 all_members 內已整理好的電話/手機/地址。
    09 開頭寫入手機欄，其餘寫入電話欄；地址寫入隱藏輔助欄。
    """
    id_to_rows: Dict[str, List[int]] = {}
    meta: Dict[int, MemberMeta] = {}
    out_r = data_start
    wrote_any = False

    _max_col = ws.max_column

    # 欄號預先解出，避免每列重複 dict.get()
    _c_clinic  = cols.get("clinic")
    _c_name    = cols.get("name")
    _c_id      = cols.get("id")
    _c_bday    = cols.get("bday")
    _c_age     = cols.get("age")
    _c_tel     = cols.get("tel")
    _c_mobile  = cols.get("mobile")
    _c_abc     = cols.get("abc")
    _c_dmk     = cols.get("dmk_code")
    _c_ascvd   = cols.get("ascvd")
    _c_cnt     = cols.get("cnt")
    _c_sex     = cols.get("sex")
    _c_addr    = cols.get("address_hidden")

    for pid, info in all_members.items():
        name   = info.get("name") or ""
        bday   = info.get("bday")
        e_code = info.get("e_code")
        abc    = info.get("abc")
        ascvd  = info.get("ascvd")
        base_contact = ContactInfo(
            phone=info.get("phone"),
            mobile=info.get("mobile"),
            address=info.get("address"),
        )
        contact = merge_contact_info(base_contact, contact_map.get(pid))
        cnt    = info.get("cnt")
        clinic = info.get("clinic") or clinic_val
        age    = calc_age(bday, now) if isinstance(bday, datetime.date) else -1
        sex    = infer_gender_from_id(pid)
        output_pid = pid if is_valid_tw_id(pid) else None

        safe_set(ws, out_r, _c_clinic,  clinic)
        safe_set(ws, out_r, _c_name,    name or None)
        safe_set(ws, out_r, _c_id,      output_pid)
        safe_set(ws, out_r, _c_bday,    bday)
        safe_set(ws, out_r, _c_age,     age if age >= 0 else None)
        safe_set(ws, out_r, _c_tel,     contact.phone or None)
        safe_set(ws, out_r, _c_mobile,  contact.mobile or None)
        safe_set(ws, out_r, _c_abc,     abc)
        safe_set(ws, out_r, _c_dmk,     e_code.value if e_code else None)
        safe_set(ws, out_r, _c_ascvd,   ascvd)
        safe_set(ws, out_r, _c_cnt,     cnt)
        safe_set(ws, out_r, _c_sex,     sex)
        safe_set(ws, out_r, _c_addr,    contact.address or None)

        meta[out_r] = MemberMeta(
            row=out_r,
            pid=pid,
            bday=bday,
            age=age,
            e_code=e_code,
            ascvd=parse_ascvd(ascvd),
        )
        id_to_rows.setdefault(pid, []).append(out_r)

        out_r += 1
        wrote_any = True

    last_row = out_r - 1 if wrote_any else (data_start - 1)

    # 全部資料寫完後，一次性套用格式（比逐列套用少一半 cell 存取）
    for r in range(data_start, last_row + 1):
        _apply_member_row_style(ws, r, _max_col)

    return id_to_rows, meta, last_row


def _fill_ascvd(
    ws,
    sh_ascvd:  Any,
    cols:      Dict[str, Optional[int]],
    id_to_rows: Dict[str, List[int]],
    meta:      Dict[int, MemberMeta],
) -> None:
    # 耀聖名單的 ID / 最後就診日可能在第 1 列，ASCVD 在第 2 列。
    id_header_row = _find_header_row_contains_any(
        sh_ascvd, [["ID", "id"]], search_rows=10
    )
    ascvd_header_row = _find_header_row_contains_any(
        sh_ascvd, [["ASCVD", "ascvd"]], search_rows=10
    )
    if id_header_row is None or ascvd_header_row is None:
        raise ValueError("原始檔「ascvd」找不到 ID / ASCVD 欄位")

    header_rows = sorted({id_header_row, ascvd_header_row})
    amap_merged: Dict[str, int] = {}
    for header_row in header_rows:
        amap_merged.update(build_header_map(sh_ascvd, header_row))

    a_id  = find_column_exact(amap_merged, ["ID", "id"])
    a_asc = find_column_exact(amap_merged, ["ASCVD", "ascvd"])
    a_lv  = find_column_exact(amap_merged, ["最後就診日"])
    if a_id is None or a_asc is None:
        raise ValueError("原始檔「ascvd」找不到 ID / ASCVD 欄位")

    for r in range(max(header_rows) + 1, sh_ascvd.max_row + 1):
        pid = normalize_id(sh_ascvd.cell(r, a_id).value)
        val = sh_ascvd.cell(r, a_asc).value
        if not pid:
            continue
        if val is None or str(val).strip() in ("",):  # 0 也要保留寫入
            continue

        rows = id_to_rows.get(pid)
        if not rows:
            continue

        for tr in rows:
            safe_set(ws, tr, cols.get("ascvd"), val)
            if tr in meta:
                meta[tr].ascvd = parse_ascvd(val)
            # 最後就診日（ascvd T欄，優先填入）
            if a_lv:
                lv_dt = parse_date(sh_ascvd.cell(r, a_lv).value)
                if lv_dt:
                    existing = parse_date(ws.cell(tr, cols.get("last_visit") or 0).value) if cols.get("last_visit") else None
                    if existing is None or lv_dt > existing:
                        safe_set(ws, tr, cols.get("last_visit"), lv_dt)


def _fill_screening(
    ws,
    sheet,
    target_col: Optional[int],
    id_to_rows: Dict[str, List[int]],
) -> None:
    if not target_col:
        return
    header_row = _find_header_row_contains_any(
        sheet,
        [["ID", "身分證號", "身份證號"], ["最後篩檢日期"]],
        search_rows=10,
    ) or 1
    hmap    = build_header_map(sheet, header_row)
    sid_col = find_column_exact(hmap, ["ID", "身分證號", "身份證號"])
    dt_col  = find_column_exact(hmap, ["最後篩檢日期"])
    if sid_col is None or dt_col is None:
        missing = []
        if sid_col is None:
            missing.append("ID")
        if dt_col is None:
            missing.append("最後篩檢日期")
        raise ValueError(f"「{sheet.title}」找不到欄位：{' / '.join(missing)}")

    for rr in range(header_row + 1, sheet.max_row + 1):
        pid = normalize_id(sheet.cell(rr, sid_col).value)
        dt  = parse_date(sheet.cell(rr, dt_col).value)
        if not pid or dt is None:
            continue

        rows = id_to_rows.get(pid)
        if not rows:
            continue

        for tr in rows:
            existing = parse_date(ws.cell(tr, target_col).value)
            if existing is None or dt > existing:
                ws.cell(tr, target_col).value = dt


def _fill_health_case(
    ws,
    sh_health: Any,
    cols:      Dict[str, Optional[int]],
    id_to_rows: Dict[str, List[int]],
) -> None:
    field_aliases = {
        "hc_id":      ["家醫收案會員ID", "ID"],
        "hc_hba":     ["最近一次HbA1c檢查結果(%)"],
        "hc_hba_dt":  ["最近一次HbA1c檢查日期"],
        "hc_ldl":     ["最近一次LDL檢查結果(mg/dL)"],
        "hc_ldl_dt":  ["最近一次LDL檢查日期"],
        "hc_uacr":    ["最近一次UACR檢查結果(mg/gm)"],
        "hc_uacr_dt": ["最近一次UACR檢查日期"],
    }
    header_row = _find_header_row_contains_any(
        sh_health,
        [field_aliases["hc_id"], field_aliases["hc_hba"], field_aliases["hc_ldl"], field_aliases["hc_uacr"]],
        search_rows=10,
    ) or 1
    hmap = build_header_map(sh_health, header_row)
    fc = {k: find_column_exact(hmap, v) for k, v in field_aliases.items()}
    if any(v is None for v in fc.values()):
        missing_labels = {
            "hc_id": "家醫收案會員ID / ID",
            "hc_hba": "最近一次HbA1c檢查結果(%)",
            "hc_hba_dt": "最近一次HbA1c檢查日期",
            "hc_ldl": "最近一次LDL檢查結果(mg/dL)",
            "hc_ldl_dt": "最近一次LDL檢查日期",
            "hc_uacr": "最近一次UACR檢查結果(mg/gm)",
            "hc_uacr_dt": "最近一次UACR檢查日期",
        }
        missing = [missing_labels[k] for k, v in fc.items() if v is None]
        raise ValueError(
            f"原始檔「HealthCase」欄位不完整：{' / '.join(missing)}"
        )

    for r in range(header_row + 1, sh_health.max_row + 1):
        pid = normalize_id(sh_health.cell(r, fc["hc_id"]).value)
        if not pid:
            continue

        rows = id_to_rows.get(pid)
        if not rows:
            continue

        for tr in rows:
            for val_key, dt_key, col_val_key, col_dt_key in [
                ("hc_hba",  "hc_hba_dt",  "hba",  "hba_dt"),
                ("hc_ldl",  "hc_ldl_dt",  "ldl",  "ldl_dt"),
                ("hc_uacr", "hc_uacr_dt", "uacr", "uacr_dt"),
            ]:
                v  = parse_float(sh_health.cell(r, fc[val_key]).value)
                dt = parse_date(sh_health.cell(r, fc[dt_key]).value)
                # 以日期較新者優先：避免重複 ID 時舊資料蓋掉新資料
                existing_dt = parse_date(ws.cell(tr, cols[col_dt_key]).value) if cols.get(col_dt_key) else None
                if dt and (existing_dt is None or dt > existing_dt):
                    if v is not None and v != 0:
                        safe_set(ws, tr, cols.get(col_val_key), v)
                    safe_set(ws, tr, cols.get(col_dt_key), dt)


def _compute_all_derived(
    ws,
    cols:       Dict[str, Optional[int]],
    meta:       Dict[int, MemberMeta],
    data_start: int,
    last_row:   int,
    now:        datetime.date,
    claim_sums: Optional[Dict[str, Dict[str, float]]] = None,
    kpi_marks:  Optional[Dict[str, set]] = None,
    screening_member_ids: Optional[Dict[str, set[str]]] = None,
) -> None:
    kpi_marks = kpi_marks or {}

    # 預取欄號（loop 外一次取出，避免每列重複 dict lookup）
    _c_dmk       = cols["dmk_code"]
    _c_ascvd     = cols["ascvd"]
    _c_sex       = cols["sex"]
    _c_adult     = cols.get("adult")
    _c_pap       = cols.get("pap")
    _c_flu       = cols.get("flu")
    _c_fit       = cols.get("fit")
    _c_hep       = cols.get("hep")
    _c_hba       = cols["hba"]
    _c_hba_dt    = cols["hba_dt"]
    _c_ldl       = cols["ldl"]
    _c_ldl_dt    = cols["ldl_dt"]
    _c_uacr      = cols["uacr"]
    _c_id        = cols["id"]
    _c_m114      = cols.get("m_count_114")
    _c_n115      = cols.get("n_count_115")
    _c_r114      = cols.get("r_amount_114")
    _c_s115      = cols.get("s_amount_115")
    # 輸出欄號預解
    _c_disease   = cols.get("disease_text")
    _c_ay_mark   = cols.get("ay_mark")
    _c_az_mark   = cols.get("az_mark")
    _c_bb_mark   = cols.get("bb_mark")
    _c_bc_mark   = cols.get("bc_mark")
    _c_ak        = cols.get("ak")
    _c_ldl_pass  = cols.get("ldl_pass")
    _c_uacr_pass = cols.get("uacr_pass")
    _c_ax        = cols.get("ax")
    _c_note      = cols.get("note")
    _c_score     = cols.get("score")
    _c_breakdown = cols.get("breakdown")
    _c_metabolic = cols.get("metabolic_enroll")
    _c_au        = cols.get("au")
    _c_av        = cols.get("av")
    _c_aw        = cols.get("aw")
    # kpi_marks set 預先取出，避免每列重複 dict.get()
    _ay_set = kpi_marks.get("ay", set())
    _az_set = kpi_marks.get("az", set())
    _bb_set = kpi_marks.get("bb", set())
    _bc_set = kpi_marks.get("bc", set())

    def _get_dt_col(row: int, col: Optional[int]) -> Optional[datetime.date]:
        return parse_date(ws.cell(row, col).value) if col else None

    for rr in range(data_start, last_row + 1):
        e_code    = parse_disease_code(ws.cell(rr, _c_dmk).value)
        ascvd     = parse_ascvd(ws.cell(rr, _c_ascvd).value)
        ascvd_raw = ws.cell(rr, _c_ascvd).value

        m   = meta.get(rr, MemberMeta(row=rr))
        age = m.age if isinstance(m.age, int) else -1
        sex = normalize_text(ws.cell(rr, _c_sex).value)

        adult_dt = _get_dt_col(rr, _c_adult)
        pap_dt   = _get_dt_col(rr, _c_pap)
        flu_dt   = _get_dt_col(rr, _c_flu)
        fit_dt   = _get_dt_col(rr, _c_fit)
        hep_dt   = _get_dt_col(rr, _c_hep)
        hba_dt   = _get_dt_col(rr, _c_hba_dt)
        ldl_dt   = _get_dt_col(rr, _c_ldl_dt)

        hba_val = ws.cell(rr, _c_hba).value
        ldl_val = ws.cell(rr, _c_ldl).value
        uacr_val = ws.cell(rr, _c_uacr).value

        safe_set(ws, rr, _c_disease, disease_group_text(e_code, ascvd))

        safe_set_check(ws, rr, _c_ay_mark,   "✔" if rr in _ay_set else None)
        safe_set_check(ws, rr, _c_az_mark,   "✔" if rr in _az_set else None)
        safe_set_check(ws, rr, _c_bb_mark,   "✔" if rr in _bb_set else None)
        safe_set_check(ws, rr, _c_bc_mark,   "✔" if rr in _bc_set else None)

        # AK / LDL合格 / UACR合格：✔
        safe_set_check(ws, rr, _c_ak,
                       "✔" if should_check_ak(e_code=e_code, ascvd=ascvd, age=age, hba_val=hba_val) else None)
        safe_set_check(ws, rr, _c_ldl_pass,
                       "✔" if should_check_ldl_pass(e_code=e_code, ascvd_raw=ascvd_raw, ldl_val=ldl_val) else None)
        safe_set_check(ws, rr, _c_uacr_pass,
                       "✔" if should_check_uacr_pass(e_code=e_code, uacr_val=uacr_val) else None)

        # AX：漏檢項目
        ax_txt = build_ax_leak_item(
            d_code=e_code,
            ascvd_raw=ascvd_raw,
            hba_dt=hba_dt,
            ldl_dt=ldl_dt,
        )
        safe_set(ws, rr, _c_ax, ax_txt or None)

        # 年齡未知時不產生篩檢備註，避免以 0 歲誤判
        note = build_screening_note(
            age=age, e_code=e_code, sex=sex,
            hep_dt=hep_dt, fit_dt=fit_dt, pap_dt=pap_dt,
            adult_dt=adult_dt, flu_dt=flu_dt, today=now,
        ) if age >= 0 else ""
        safe_set(ws, rr, _c_note, note or None)

        hba_num = parse_float(hba_val)
        metabolic_flag = (
            "✔"
            if _c_metabolic
            and 20 <= age <= 64
            and hba_num is not None
            and 5.7 <= hba_num <= 6.4
            else None
        )
        safe_set_check(ws, rr, _c_metabolic, metabolic_flag)

        visit_count_114 = ws.cell(rr, _c_m114).value if _c_m114 else None
        visit_count_115 = ws.cell(rr, _c_n115).value if _c_n115 else None
        claim_amount_114 = ws.cell(rr, _c_r114).value if _c_r114 else None
        claim_amount_115 = ws.cell(rr, _c_s115).value if _c_s115 else None
        pid = normalize_id(ws.cell(rr, _c_id).value)
        adult_in_list = bool(pid and screening_member_ids and pid in screening_member_ids.get("adult", set()))
        pap_in_list = bool(pid and screening_member_ids and pid in screening_member_ids.get("pap", set()))
        flu_in_list = bool(pid and screening_member_ids and pid in screening_member_ids.get("flu", set()))
        fit_in_list = bool(pid and screening_member_ids and pid in screening_member_ids.get("fit", set()))
        hep_in_list = bool(pid and screening_member_ids and pid in screening_member_ids.get("hep", set()))
        fee_score_amount_114 = claim_amount_114
        fee_score_amount_115 = claim_amount_115
        if claim_sums is not None:
            pid = normalize_text(ws.cell(rr, _c_id).value).upper()
            data = claim_sums.get(pid)
            if data:
                v114a_total = data.get("114_amt_total", 0.0)
                v115a = data.get("115_amt", 0.0)
                v115_months = max(int(data.get("115_months", 0.0)), 1)
                fee_score_amount_114 = _to_excel_int(v114a_total / 12.0) if v114a_total != 0 else None
                fee_score_amount_115 = _to_excel_int(v115a / float(v115_months)) if v115a != 0 else None

        score, breakdown = calc_score(
            e_code=e_code, ascvd_raw=ascvd_raw,
            hba_val=hba_val, hba_dt=hba_dt,
            ldl_val=ldl_val, ldl_dt=ldl_dt,
            adult_dt=adult_dt, adult_in_list=adult_in_list,
            pap_dt=pap_dt, pap_in_list=pap_in_list,
            flu_dt=flu_dt, flu_in_list=flu_in_list,
            fit_dt=fit_dt, fit_in_list=fit_in_list,
            hep_dt=hep_dt, hep_in_list=hep_in_list,
            age=age, sex=sex,
            today=now,
            claim_amount_114=fee_score_amount_114,
            claim_amount_115=fee_score_amount_115,
            visit_count_114=visit_count_114,
            visit_count_115=visit_count_115,
        )
        safe_set(ws, rr, _c_score,     score)
        safe_set(ws, rr, _c_breakdown, breakdown)

        au_txt = build_au_note(
            e_code=e_code, ascvd=ascvd,
            hba_dt=hba_dt, ldl_dt=ldl_dt, today=now,
        )
        safe_set(ws, rr, _c_au, au_txt or None)

        if au_txt and not _should_skip_followup(au_txt):
            safe_set(ws, rr, _c_av, build_followup_note(au_txt, now, Rules.AV_OFFSET_DAYS))
            safe_set(ws, rr, _c_aw, build_followup_note(au_txt, now, Rules.AW_OFFSET_DAYS))
        else:
            safe_set(ws, rr, _c_av, None)
            safe_set(ws, rr, _c_aw, None)







# ============================================================
# 清空資料列
# ============================================================
def _clear_data_rows(
    ws, data_start: int, max_row: int, cols: Dict[str, Optional[int]]
) -> None:
    clear_keys = [
        "clinic", "name", "id", "bday", "age", "tel", "mobile", "cnt", "sex", "abc",
        "dmk_code", "ascvd", "main_sub_dx",
        "adult", "pap", "flu", "fit", "hep",
        "hba", "hba_dt", "ldl", "ldl_dt", "uacr", "uacr_dt",
        "disease_text", "ay_mark", "az_mark", "ak", "ldl_pass", "uacr_pass", "ax", "score", "breakdown", "note",
        "metabolic_enroll",
        "bb_mark", "bc_mark", "au", "av", "aw",
        "m_count_114", "m_count_114_q1", "n_count_115_q1", "n_count_115",
        "r_amount_114", "s_amount_115",
        "p4p_plan", "p4p_status", "p4p_enroll_dt", "p4p_last_dt", "p4p_next_dt", "p4p_overdue",
        "is_114", "is_self_select", "is_115x", "address_hidden",
        "last_visit", "dx_raw",
    ]
    col_ids = [cols[k] for k in clear_keys if cols.get(k)]
    for r in range(data_start, max_row + 1):
        for c in col_ids:
            ws.cell(r, c).value = None  # type: ignore[arg-type]


def _first_sheet(wb, *names: str):
    for name in names:
        if name in wb.sheetnames:
            return wb[name]
    return None


def _require_sheet(wb, *names: str):
    sheet = _first_sheet(wb, *names)
    if sheet is None:
        raise ValueError(f"原始檔缺少工作表：{' / '.join(names)}")
    return sheet


def _find_header_col_in_rows(ws, labels: List[str], max_scan_row: int = 3) -> Optional[int]:
    targets = {normalize_text(label) for label in labels if label}
    for row in range(1, min(max_scan_row, ws.max_row) + 1):
        for col in range(1, ws.max_column + 1):
            if normalize_text(ws.cell(row, col).value) in targets:
                return col
    return None


def _ensure_aux_header(
    ws,
    labels: List[str],
    preferred_letter: str,
    *,
    row1_value: str,
    row2_default: str = "",
) -> int:
    existing_col = _find_header_col_in_rows(ws, labels)
    if existing_col is not None:
        return existing_col

    col = column_index_from_string(preferred_letter)
    ws.cell(1, col).value = ws.cell(1, col).value or row1_value
    ws.cell(2, col).value = ws.cell(2, col).value or row2_default
    return col


def _main_alignment_end_col(ws) -> int:
    labels = [
        "預防保健提醒", "備註",
        "是否為114會員名單", "是否為自選會員", "是否為115X",
        "115年1-4月就診次數", "114年1-4月就診次數", "地址",
    ]
    found_cols = [
        col for col in (_find_header_col_in_rows(ws, [label]) for label in labels)
        if col is not None
    ]
    return max(found_cols) if found_cols else column_index_from_string(Rules.COL_114_COUNT_Q1_HIDDEN)


def prepare_template_layout(ws) -> None:
    """補齊 0325 樣板缺少但程式仍需使用的欄位/輔助欄，並先隱藏後段輔助欄。"""
    ws["M1"] = "114年實際申報總額"
    ws["O1"] = "115年實際申報總額"
    hidden_cols = {
        _ensure_aux_header(ws, ["預防保健提醒", "備註"], "AY", row1_value="預防保健提醒"),
        _ensure_aux_header(ws, ["是否為114會員名單"], "AZ", row1_value="是否為114會員名單"),
        _ensure_aux_header(ws, ["是否為自選會員"], "BA", row1_value="是否為自選會員"),
        _ensure_aux_header(ws, ["是否為115X"], "BB", row1_value="是否為115X"),
        _ensure_aux_header(ws, ["115年1-4月就診次數"], "BC", row1_value="115年1-4月就診次數"),
        _ensure_aux_header(ws, ["114年月平均"], "BD", row1_value="114年月平均"),
        _ensure_aux_header(ws, ["115年月平均"], "BE", row1_value="115年月平均"),
        _ensure_aux_header(ws, ["地址"], "BF", row1_value="地址"),
        _ensure_aux_header(ws, ["114年1-4月就診次數"], "BG", row1_value="114年1-4月就診次數"),
    }
    for col in range(min(hidden_cols), max(hidden_cols) + 1):
        ws.column_dimensions[get_column_letter(col)].hidden = True


# ============================================================
# 主流程
# ============================================================
# ============================================================
# 結構化流程（V10.3）
# ============================================================
def load_source(
    source_path: str,
    profile: Optional[ProcessingProfile] = None,
) -> SourceContext:
    profile = _resolve_profile(profile)
    if os.path.isdir(source_path):
        _log(f"掃描資料夾：{os.path.basename(os.path.abspath(source_path))}")
        wb_src = _merge_source_folder(source_path, profile=profile)
    else:
        wb_src = _load_and_validate_source(source_path, profile=profile)
    scan_cache = _scan_source_sheets(wb_src, profile=profile)
    all_members = collect_all_members(wb_src, partial_maps=scan_cache.partial_maps)
    claim_sums, claim_months_115 = profile.collect_monthly_claim_summaries(
        wb_src,
        monthly_scans=scan_cache.monthly_claim_sheets,
    )
    claim_sums, claim_months_115 = _supplement_claim_counts_from_hisb(
        wb_src,
        all_members,
        claim_sums,
        claim_months_115,
        hisb_scans=scan_cache.hisb_count_sheets,
        profile=profile,
    )
    sh_p4p_enroll = _first_sheet(wb_src, "P4P收案", "收案管理", "P4pCase")
    sh_p4p_track = _first_sheet(wb_src, "P4P追蹤", "追蹤管理", "P4pTrack")
    sh_115_designated = _first_sheet(wb_src, "115指定會員")
    designated_115_source_ids = build_id_set(
        sh_115_designated,
        ["ID", "身份證號", "身份證號碼", "身分證號", "身分證號碼"],
    ) if sh_115_designated is not None else set()
    designated_115_details = _extract_member_partial_map(
        sh_115_designated,
        ["ID", "身份證號", "身份證號碼", "身分證號", "身分證號碼"],
    ) if sh_115_designated is not None else {}
    designated_115_ids = {
        pid for pid in designated_115_source_ids
        if pid in all_members
    }
    if sh_115_designated is not None:
        _log(
            "115指定會員："
            f"有效 ID {len(designated_115_source_ids)} 筆，"
            f"會員總表命中 {len(designated_115_ids)} 筆"
        )
    p4p_map = build_p4p_map(sh_p4p_enroll, sh_p4p_track)
    return SourceContext(
        wb_src=wb_src,
        sh_member=_first_sheet(wb_src, "會員名單", "較需要照護名單"),
        sh_ascvd=_first_sheet(wb_src, "ascvd", "ASCVD"),
        sh_health=_require_sheet(wb_src, "HealthCase", "健康管理列表", "個案健康管理列表"),
        sh_main_sub_dx=_first_sheet(wb_src, "主次診斷", "主次診段"),
        sh_phone=_first_sheet(wb_src, "行動電話", "基本資料檔列印"),
        sh_self_select=_first_sheet(wb_src, "自選會員", "自選名單"),
        sh_115x=_first_sheet(wb_src, "115X"),
        sh_115_designated=sh_115_designated,
        sh_p4p_enroll=sh_p4p_enroll,
        sh_p4p_track=sh_p4p_track,
        screening_sheets={
            "adult": _require_sheet(wb_src, "成人健檢", "成人預防保健"),
            "pap":   _require_sheet(wb_src, "子宮抹片", "子宮頸抹片"),
            "flu":   _require_sheet(wb_src, "老人流感", "65歲流感", "65歲流感疫苗", "流感疫苗"),
            "fit":   _require_sheet(wb_src, "糞便潛血"),
            "hep":   _require_sheet(wb_src, "肝炎篩檢", "BC肝", "B肝C肝", "BC肝炎"),
        },
        claim_sums=claim_sums,
        claim_months_115=claim_months_115,
        all_members=all_members,
        p4p_map=p4p_map,
        designated_115_source_count=len(designated_115_source_ids),
        designated_115_ids=designated_115_ids,
        designated_115_details=designated_115_details,
    )


def load_template(template_path: str) -> TemplateContext:
    wb_tpl = openpyxl.load_workbook(template_path)
    if Rules.SHEET_TARGET not in wb_tpl.sheetnames:
        raise ValueError(f"模板檔缺少工作表：{Rules.SHEET_TARGET}")

    ws = wb_tpl[Rules.SHEET_TARGET]
    prepare_template_layout(ws)
    data_start = Rules.DATA_START_ROW
    cols = detect_template_columns(ws, data_start)
    _clear_data_rows(ws, data_start, ws.max_row, cols)

    return TemplateContext(
        wb_tpl=wb_tpl,
        ws=ws,
        cols=cols,
        data_start=data_start,
    )


def fill_basic_data(source_ctx: SourceContext, template_ctx: TemplateContext, now: datetime.date) -> RuntimeContext:
    contact_map = build_contact_map(source_ctx.sh_phone)
    clinic_val = ""
    if source_ctx.sh_member is not None:
        clinic_val = normalize_text(source_ctx.sh_member.cell(1, 1).value)

    id_to_rows, meta, last_row = _fill_member_basic(
        template_ctx.ws,
        source_ctx.all_members,
        contact_map,
        template_ctx.cols,
        template_ctx.data_start,
        now,
        clinic_val=clinic_val,
    )
    return RuntimeContext(
        id_to_rows=id_to_rows,
        meta=meta,
        last_row=last_row,
    )


def fill_external_data(source_ctx: SourceContext, template_ctx: TemplateContext, runtime_ctx: RuntimeContext) -> None:
    ws = template_ctx.ws
    cols = template_ctx.cols
    data_start = template_ctx.data_start
    last_row = runtime_ctx.last_row
    id_to_rows = runtime_ctx.id_to_rows
    meta = runtime_ctx.meta

    if source_ctx.sh_ascvd:
        _log("套用 ascvd 資料")
        _fill_ascvd(ws, source_ctx.sh_ascvd, cols, id_to_rows, meta)
    _log("套用主次診斷")
    _fill_main_sub_dx(ws, source_ctx.sh_main_sub_dx, source_ctx.wb_src, cols, id_to_rows)
    _sheet_label = {"adult": "成人健檢", "pap": "子宮抹片", "flu": "老人流感", "fit": "糞便潛血", "hep": "肝炎篩檢"}
    for key, sheet in source_ctx.screening_sheets.items():
        _log(f"套用篩檢資料：{_sheet_label.get(key, key)}")
        _fill_screening(ws, sheet, cols.get(key), id_to_rows)
    _log("套用 HealthCase 檢驗資料")
    _fill_health_case(ws, source_ctx.sh_health, cols, id_to_rows)
    _log("回填月份申報統計")
    fill_monthly_claim_summary_columns(
        ws,
        data_start,
        last_row,
        cols,
        source_ctx.claim_sums,
        meta=meta,
    )

    _log("回填旗標與 P4P 狀態")
    _log_duplicate_ids(source_ctx.sh_p4p_enroll, "P4P收案")
    _log_duplicate_ids(source_ctx.sh_p4p_track, "P4P追蹤")
    id_aliases = ["身份證號", "身份證號碼", "身分證號", "身分證號碼", "ID", "家醫收案會員ID"]
    member_sheets = [sheet for sheet in (
        source_ctx.sh_member,
        _first_sheet(source_ctx.wb_src, "114會員名單補充"),
        source_ctx.sh_ascvd,
    ) if sheet is not None]
    member_ids = build_member_key_set_from_sources(source_ctx.all_members, member_sheets, id_aliases)
    self_select_sheets = [sheet for sheet in (
        source_ctx.sh_self_select,
        _first_sheet(source_ctx.wb_src, "自選名單115"),
    ) if sheet is not None]
    self_select_ids = build_member_key_set_from_sources(source_ctx.all_members, self_select_sheets, id_aliases)
    x115_ids = build_member_key_set_from_source(source_ctx.all_members, source_ctx.sh_115x, id_aliases)
    _fill_extra_flags(
        ws, cols, data_start, last_row, id_to_rows,
        source_ctx.p4p_map, member_ids, self_select_ids, x115_ids,
    )
    _recompute_member_category(ws, cols, id_to_rows, member_ids, self_select_ids, x115_ids)


def compute_derived(
    template_ctx: TemplateContext,
    runtime_ctx: RuntimeContext,
    now: datetime.date,
    source_ctx: Optional[SourceContext] = None,
) -> Tuple[List[Tuple[int, float]], List[Tuple[int, float]]]:
    ws = template_ctx.ws
    cols = template_ctx.cols
    data_start = template_ctx.data_start
    last_row = runtime_ctx.last_row

    hba_candidates = _collect_hba_candidates(ws, cols, data_start, last_row)
    ldl_candidates = _collect_ldl_candidates(ws, cols, data_start, last_row)
    kpi_marks = collect_kpi_mark_sets(
        ws, cols, data_start, last_row,
        hba_candidates=hba_candidates,
        ldl_candidates=ldl_candidates,
    )

    screening_member_ids: Dict[str, set[str]] = {}
    if source_ctx is not None:
        screening_id_aliases = ["ID", "身份證號", "身份證號碼", "身分證號", "身分證號碼"]
        for key, sheet in source_ctx.screening_sheets.items():
            screening_member_ids[key] = build_member_key_set_from_source(
                source_ctx.all_members,
                sheet,
                screening_id_aliases,
            )

    _compute_all_derived(
        ws,
        cols,
        runtime_ctx.meta,
        data_start,
        last_row,
        now,
        claim_sums=source_ctx.claim_sums if source_ctx is not None else None,
        kpi_marks=kpi_marks,
        screening_member_ids=screening_member_ids,
    )
    apply_date_format(ws, cols, data_start, last_row)
    apply_amount_format(ws, cols, data_start, last_row)
    return hba_candidates, ldl_candidates


def compute_kpis(
    template_ctx: TemplateContext,
    runtime_ctx: RuntimeContext,
    now: datetime.date,
    source_ctx: Optional[SourceContext] = None,
    hba_candidates: Optional[List[Tuple[int, float]]] = None,
    ldl_candidates: Optional[List[Tuple[int, float]]] = None,
) -> None:
    ws = template_ctx.ws
    cols = template_ctx.cols
    data_start = template_ctx.data_start
    last_row = runtime_ctx.last_row

    _log("產生 KPI 摘要與附表")
    runtime_ctx.hba_main_summary, runtime_ctx.hba_target_summary = calc_hba_kpi_ay_az(
        ws, cols, data_start, last_row,
        hba_candidates=hba_candidates,
    )
    runtime_ctx.ldl_main_summary, runtime_ctx.ldl_target_summary = calc_ldl_percentiles(
        ws, cols, data_start, last_row,
        ldl_candidates=ldl_candidates,
    )
    _write_legacy_kpi_summary_cells(
        ws,
        hba_main_summary=runtime_ctx.hba_main_summary,
        hba_target_summary=runtime_ctx.hba_target_summary,
        ldl_main_summary=runtime_ctx.ldl_main_summary,
        ldl_target_summary=runtime_ctx.ldl_target_summary,
    )
    _log("產生百分位名單")
    populate_percentile_sheet(
        template_ctx.wb_tpl,
        runtime_ctx.hba_main_summary,
        runtime_ctx.hba_target_summary,
        runtime_ctx.ldl_main_summary,
        runtime_ctx.ldl_target_summary,
        cols,
        data_start,
        last_row,
    )
    screening_member_ids: Dict[str, set[str]] = {}
    if source_ctx is not None:
        screening_id_aliases = ["ID", "身份證號", "身份證號碼", "身分證號", "身分證號碼"]
        for key, sheet in source_ctx.screening_sheets.items():
            screening_member_ids[key] = build_member_key_set_from_source(
                source_ctx.all_members,
                sheet,
                screening_id_aliases,
            )
    _log("產生醫生看工作表")
    populate_doctor_sheet(
        template_ctx.wb_tpl,
        ws,
        cols,
        data_start,
        last_row,
        now,
        screening_member_ids=screening_member_ids,
        p4p_map=source_ctx.p4p_map if source_ctx is not None else None,
        claim_months_115=source_ctx.claim_months_115 if source_ctx is not None else None,
        all_members=source_ctx.all_members if source_ctx is not None else None,
        designated_115_ids=source_ctx.designated_115_ids if source_ctx is not None else None,
        designated_115_details=source_ctx.designated_115_details if source_ctx is not None else None,
    )
    if source_ctx:
        _log("產生自選名單工作表")
        populate_self_select_sheet(
            template_ctx.wb_tpl,
            ws,
            cols,
            data_start,
            last_row,
            source_ctx.sh_self_select,
        )


def _finalize_output_alignments(template_ctx: TemplateContext, runtime_ctx: Optional[RuntimeContext]) -> None:
    main_last_row = runtime_ctx.last_row if runtime_ctx else template_ctx.ws.max_row
    _log("整理會員總表對齊")
    _finalize_main_sheet_alignment(
        template_ctx.ws,
        template_ctx.data_start,
        main_last_row,
    )
    if PERCENTILE_SHEET_NAME in template_ctx.wb_tpl.sheetnames:
        _log("整理百分位名單對齊")
        _finalize_percentile_sheet_alignment(template_ctx.wb_tpl[PERCENTILE_SHEET_NAME])
    if DOCTOR_SHEET_NAME in template_ctx.wb_tpl.sheetnames:
        _log("整理醫生看對齊")
        _finalize_doctor_sheet_alignment(template_ctx.wb_tpl[DOCTOR_SHEET_NAME])
    if SELF_SELECT_SHEET_NAME in template_ctx.wb_tpl.sheetnames:
        _log("整理自選名單對齊")
        _finalize_self_select_sheet_alignment(template_ctx.wb_tpl[SELF_SELECT_SHEET_NAME])


def trim_worksheet(ws) -> None:
    """
    只刪除 worksheet 尾端完全空白的列與欄。
    判斷標準：
    - 儲存格值為 None 或 "" 視為空白
    - 中間的空白列/欄不動
    - 只刪除最後一個有值儲存格之後的尾端區塊
    """
    last_row = 0
    last_col = 0

    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                if cell.row > last_row:
                    last_row = cell.row
                if cell.column > last_col:
                    last_col = cell.column

    if last_row == 0:
        return

    if ws.max_row > last_row:
        ws.delete_rows(last_row + 1, ws.max_row - last_row)

    if ws.max_column > last_col:
        ws.delete_cols(last_col + 1, ws.max_column - last_col)


def trim_workbook(wb) -> None:
    """對活頁簿中的所有工作表執行尾端空白列/欄清理。"""
    for ws in wb.worksheets:
        trim_worksheet(ws)


def finalize_and_save(
    source_path: str,
    template_ctx: TemplateContext,
    now_dt: datetime.datetime,
    runtime_ctx: Optional[RuntimeContext] = None,
) -> str:
    wb_tpl = template_ctx.wb_tpl
    source_abs = os.path.abspath(source_path)
    source_dir = source_abs if os.path.isdir(source_abs) else os.path.dirname(source_abs)
    base_dir = os.path.dirname(source_dir)

    _finalize_output_alignments(template_ctx, runtime_ctx)
    trim_workbook(wb_tpl)

    for sht_name in (Rules.SHEET_TARGET, PERCENTILE_SHEET_NAME):
        if sht_name in wb_tpl.sheetnames:
            wb_tpl[sht_name].sheet_view.showGridLines = True

    clinic_name = ""
    clinic_code = _extract_clinic_code_from_source(source_abs)
    lookup = _load_clinic_name_lookup(source_dir)
    if clinic_code:
        clinic_name = lookup.get(clinic_code, "")
    clinic_name = _sanitize_filename_component(clinic_name)

    if clinic_name:
        filename = f"{clinic_name}選會員_{now_dt.strftime('%m%d_%H%M')}.xlsx"
    else:
        filename = f"選會員{now_dt.strftime('%m%d_%H%M')}.xlsx"

    out_path = os.path.join(base_dir, filename)
    _log("開始寫入 Excel")
    wb_tpl.save(out_path)
    return out_path


def _log_member_category_counts(template_ctx: TemplateContext, runtime_ctx: RuntimeContext) -> None:
    ws = template_ctx.ws
    cols = template_ctx.cols
    col_abc = cols.get("abc")
    if not col_abc:
        return

    counts: Dict[str, int] = {}
    for rr in range(template_ctx.data_start, runtime_ctx.last_row + 1):
        value = normalize_text(ws.cell(rr, col_abc).value)
        if not value:
            continue
        counts[value] = counts.get(value, 0) + 1

    order = ["A", "A/E1", "A/E2", "A/E1/E2", "B", "B/E1", "B/E2", "B/E1/E2", "E1", "E2", "E1/E2", "D"]
    summary = [f"會員{key} {counts.get(key, 0)}筆" for key in order if counts.get(key, 0)]
    if not summary:
        summary = ["會員分類 0筆"]
    _log(" | ".join(summary))


def _log(msg: str) -> None:
    """輸出帶台灣時間時間戳的進度訊息，並立即 flush 到 terminal。"""
    ts = datetime.datetime.now(_TZ_TW).strftime("%H:%M:%S")
    print(f"[{ts}] {msg}", flush=True)


def _sanitize_filename_component(name: str) -> str:
    name = normalize_text(name)
    if not name:
        return ""
    return re.sub(r'[\\/:*?"<>|]+', "_", name)


def _format_month_span(months: List[int]) -> str:
    months = sorted({month for month in months if 1 <= month <= 12})
    if not months:
        return ""
    if len(months) == 1:
        return f"{months[0]}月"
    if months == list(range(months[0], months[-1] + 1)):
        return f"{months[0]}-{months[-1]}月"
    return "、".join(f"{month}月" for month in months)


def _format_115_months_summary(months: List[int]) -> str:
    if not months:
        return "115年有效月份：未偵測到月份資料。"
    month_text = _format_month_span(months)
    return (
        f"115年有效月份：{month_text}，共{len(months)}個月；"
        f"醫生看 L 欄使用114年{month_text}同期次數，"
        f"M 欄使用115年{month_text}次數，"
        f"O 欄使用115年{month_text}總費用除以{len(months)}。"
    )


def _extract_clinic_code_from_source(source_path: str) -> Optional[str]:
    base = os.path.basename(os.path.abspath(source_path))
    m = re.match(r"^([A-Za-z0-9]{10})", base)
    if m:
        return m.group(1)
    return None


def _find_clinic_lookup_workbook(folder_path: str) -> Optional[str]:
    script_dir = os.path.dirname(os.path.abspath(__file__))
    preferred = os.path.join(script_dir, "醫療群_衛福部資料.xlsx")
    if os.path.isfile(preferred):
        return preferred

    for name in sorted(os.listdir(folder_path)):
        lower = name.lower()
        if not lower.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
            continue
        if "衛福部" in name and "資料" in name:
            return os.path.join(folder_path, name)
    return None


def _load_clinic_name_lookup(folder_path: str) -> Dict[str, str]:
    path = _find_clinic_lookup_workbook(folder_path)
    if not path:
        return {}

    try:
        wb = _load_xlsx_as_workbook(path)
    except Exception:
        return {}

    for sname in wb.sheetnames:
        ws = wb[sname]
        if ws.max_row < 2:
            continue
        header_row = find_header_row_contains(ws, ["醫事機構代碼", "醫事機構名稱"], 5) or 1
        hmap = build_header_map(ws, header_row)
        code_col = find_column_exact(hmap, ["醫事機構代碼"])
        name_col = find_column_exact(hmap, ["官方名稱", "醫事機構名稱"])
        if not code_col or not name_col:
            continue

        result: Dict[str, str] = {}
        for r in range(2, ws.max_row + 1):
            code = normalize_text(ws.cell(r, code_col).value)
            name = normalize_text(ws.cell(r, name_col).value)
            if code and name:
                result[code] = name
        if result:
            return result
    return {}


def process_excel(
    source_path: str,
    template_path: str,
    profile: Optional[ProcessingProfile] = None,
) -> str:
    profile = _resolve_profile(profile)
    now_dt = datetime.datetime.now(_TZ_TW)
    now = now_dt.date()

    _log("開始處理 Excel")

    _log(f"載入來源：{os.path.basename(os.path.abspath(source_path))}")
    source_ctx = load_source(source_path, profile=profile)

    _log(f"載入模板：{os.path.basename(template_path)}")
    template_ctx = load_template(template_path)
    _log("偵測模板欄位並清空舊資料")

    _log("整理聯絡資料與建立會員主表")
    runtime_ctx = fill_basic_data(source_ctx, template_ctx, now)

    if runtime_ctx.last_row < template_ctx.data_start:
        _log("無會員資料，直接輸出")
        return finalize_and_save(source_path, template_ctx, now_dt, runtime_ctx)

    _log("回填外部資料：ASCVD / 主次診斷 / 篩檢 / HealthCase / 月份統計")
    fill_external_data(source_ctx, template_ctx, runtime_ctx)

    _log("計算分數、追蹤提醒與衍生欄位")
    hba_candidates, ldl_candidates = compute_derived(template_ctx, runtime_ctx, now, source_ctx)

    _log("計算 KPI 標記名單")
    compute_kpis(
        template_ctx,
        runtime_ctx,
        now,
        source_ctx,
        hba_candidates=hba_candidates,
        ldl_candidates=ldl_candidates,
    )

    _log_member_category_counts(template_ctx, runtime_ctx)
    _log("寫入輸出檔案")
    out = finalize_and_save(source_path, template_ctx, now_dt, runtime_ctx)
    _log(_format_115_months_summary(source_ctx.claim_months_115))
    _log(f"完成輸出：{os.path.basename(out)}")
    return out


# ============================================================
# 進入點
# ============================================================
def main() -> None:
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()

    src = filedialog.askdirectory(
        title="選擇來源資料夾（程式會掃描裡面所有 Excel 檔案）"
    )
    if not src:
        return

    script_dir = os.path.dirname(os.path.abspath(__file__))
    try:
        template = _find_template(script_dir)
    except RuntimeError as exc:
        messagebox.showerror("錯誤", str(exc))
        return

    _log("已選擇來源資料夾，開始執行")
    try:
        out = process_excel(src, template)
        messagebox.showinfo("完成", f"已輸出：\n{out}")
        open_file_cross_platform(out)
    except (ValueError, KeyError, OSError, InvalidFileException) as e:
        messagebox.showerror("錯誤", str(e))
    except Exception as e:
        traceback.print_exc()
        messagebox.showerror("錯誤", f"未預期錯誤：{e}")


if __name__ == "__main__":
    main()
