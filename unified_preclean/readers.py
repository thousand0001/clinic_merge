from __future__ import annotations

import csv
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path
from typing import Iterable, List

from openpyxl import load_workbook

from .schema import SheetRows


def discover_files(source_dir: Path) -> List[Path]:
    from .schema import SUPPORTED_EXTENSIONS

    files: List[Path] = []
    for path in source_dir.rglob("*"):
        if not path.is_file() or path.name.startswith("~$"):
            continue
        if path.suffix.lower() in SUPPORTED_EXTENSIONS:
            files.append(path)
    return sorted(files)


def _trim_trailing_empty(row: Iterable[object]) -> List[object]:
    values = list(row)
    while values and values[-1] in (None, ""):
        values.pop()
    return values


def read_workbook(path: Path) -> List[SheetRows]:
    ext = path.suffix.lower()
    if ext in {".xlsx", ".xlsm"}:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            out: List[SheetRows] = []
            for ws in wb.worksheets:
                rows = [_trim_trailing_empty(row) for row in ws.iter_rows(values_only=True)]
                out.append(SheetRows(path, ws.title, rows))
            return out
        finally:
            wb.close()
    if ext == ".xls":
        import xlrd

        book = xlrd.open_workbook(str(path))
        out = []
        for sheet in book.sheets():
            rows = [_trim_trailing_empty(sheet.row_values(row_index)) for row_index in range(sheet.nrows)]
            out.append(SheetRows(path, sheet.name, rows))
        return out
    if ext == ".ods":
        return _read_ods(path)
    raise ValueError(f"不是工作簿格式：{path.name}")


def read_csv_like(path: Path) -> List[SheetRows]:
    encodings = ("utf-8-sig", "utf-16", "utf-16le", "cp950", "big5")
    last_error: Exception | None = None
    for encoding in encodings:
        try:
            with path.open("r", encoding=encoding, errors="strict", newline="") as file:
                rows = [_trim_trailing_empty(row) for row in csv.reader(file)]
            return [SheetRows(path, path.stem[:31] or "Sheet1", rows)]
        except UnicodeError as exc:
            last_error = exc
    with path.open("r", encoding="cp950", errors="replace", newline="") as file:
        rows = [_trim_trailing_empty(row) for row in csv.reader(file)]
    if last_error:
        rows.insert(0, [f"讀取警告：使用 cp950 replace；原錯誤：{last_error}"])
    return [SheetRows(path, path.stem[:31] or "Sheet1", rows)]


def read_txt(path: Path) -> List[SheetRows]:
    encodings = ("cp950", "big5", "utf-8-sig")
    text = ""
    for encoding in encodings:
        try:
            text = path.read_text(encoding=encoding)
            break
        except UnicodeError:
            continue
    if not text:
        text = path.read_text(encoding="cp950", errors="replace")
    rows = []
    for line in text.splitlines():
        if not line.strip():
            continue
        if "," in line:
            rows.append([part.strip() for part in line.split(",")])
        else:
            rows.append([line.strip()])
    return [SheetRows(path, path.stem[:31] or "Text", rows)]


def read_pdf_text(path: Path) -> List[SheetRows]:
    try:
        from pypdf import PdfReader
    except Exception:
        return [SheetRows(path, "PDF", [["PDF_READER_UNAVAILABLE"]])]
    try:
        reader = PdfReader(str(path))
        rows = []
        for page_index, page in enumerate(reader.pages, start=1):
            text = page.extract_text() or ""
            for line in text.splitlines():
                if line.strip():
                    rows.append([page_index, line.strip()])
        return [SheetRows(path, "PDF", rows)]
    except Exception as exc:
        return [SheetRows(path, "PDF", [[f"PDF_READ_ERROR: {exc}"]])]


def read_any(path: Path) -> List[SheetRows]:
    ext = path.suffix.lower()
    if ext in {".xlsx", ".xlsm", ".xls", ".ods"}:
        return read_workbook(path)
    if ext == ".csv":
        return read_csv_like(path)
    if ext == ".txt":
        return read_txt(path)
    if ext == ".pdf":
        return read_pdf_text(path)
    return [SheetRows(path, "", [])]


def _read_ods(path: Path) -> List[SheetRows]:
    office_ns = "urn:oasis:names:tc:opendocument:xmlns:office:1.0"
    table_ns = "urn:oasis:names:tc:opendocument:xmlns:table:1.0"
    text_ns = "urn:oasis:names:tc:opendocument:xmlns:text:1.0"
    q_spreadsheet = f"{{{office_ns}}}spreadsheet"
    q_table = f"{{{table_ns}}}table"
    q_row = f"{{{table_ns}}}table-row"
    q_cell = f"{{{table_ns}}}table-cell"
    q_covered_cell = f"{{{table_ns}}}covered-table-cell"
    q_name = f"{{{table_ns}}}name"
    q_col_repeat = f"{{{table_ns}}}number-columns-repeated"
    q_row_repeat = f"{{{table_ns}}}number-rows-repeated"
    q_paragraph = f"{{{text_ns}}}p"

    with zipfile.ZipFile(path, "r") as zf:
        root = ET.fromstring(zf.read("content.xml"))

    spreadsheet = root.find(f".//{q_spreadsheet}")
    if spreadsheet is None:
        return [SheetRows(path, "Sheet1", [])]

    sheets: List[SheetRows] = []
    for index, table in enumerate(spreadsheet, start=1):
        if table.tag != q_table:
            continue
        title = str(table.attrib.get(q_name) or f"Sheet{index}")[:31]
        rows: List[List[str]] = []
        for row in table:
            if row.tag != q_row:
                continue
            values: List[str] = []
            pending_blank_repeat = 0
            for cell in row:
                if cell.tag not in (q_cell, q_covered_cell):
                    continue
                repeat = int(cell.attrib.get(q_col_repeat, "1") or "1")
                text = "\n".join(
                    "".join(paragraph.itertext()).strip()
                    for paragraph in cell.iter(q_paragraph)
                    if "".join(paragraph.itertext()).strip()
                )
                if text:
                    if pending_blank_repeat:
                        values.extend([""] * pending_blank_repeat)
                        pending_blank_repeat = 0
                    values.extend([text] * repeat)
                else:
                    pending_blank_repeat += repeat
            if not values:
                continue
            row_repeat = int(row.attrib.get(q_row_repeat, "1") or "1")
            rows.append(values)
            for _ in range(row_repeat - 1):
                rows.append(list(values))
        sheets.append(SheetRows(path, title or f"Sheet{index}", rows))
    return sheets or [SheetRows(path, "Sheet1", [])]

