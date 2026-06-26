from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Iterable, List, Optional, Sequence, Set

from openpyxl import Workbook, load_workbook

from .schema import SourceFinding
from .utils import normalize_id


REPORT_HEADERS = [
    "來源檔",
    "副檔名",
    "sheet",
    "資料類型",
    "狀態",
    "原因",
    "表頭列",
    "總列數",
    "資料列數",
    "有效ID列數",
    "唯一ID數",
    "日期列數",
    "金額列數",
    "次數列數",
    "輸出命中ID數",
    "輸出未命中ID數",
    "欄位",
    "例子",
]


def load_output_ids(output_path: Optional[Path]) -> Optional[Set[str]]:
    if output_path is None:
        return None
    wb = load_workbook(output_path, read_only=True, data_only=True)
    try:
        if "會員總表" not in wb.sheetnames:
            return set()
        ws = wb["會員總表"]
        ids: Set[str] = set()
        for row in ws.iter_rows(min_row=3, values_only=True):
            pid = normalize_id(row[4] if len(row) > 4 else None)
            if pid:
                ids.add(pid)
        return ids
    finally:
        wb.close()


def finding_to_row(root: Path, finding: SourceFinding) -> List[object]:
    return [
        finding.relative_file(root),
        finding.extension,
        finding.sheet_name,
        finding.data_type,
        finding.status,
        finding.reason,
        finding.header_row or "",
        finding.row_count,
        finding.data_row_count,
        finding.valid_id_rows,
        finding.unique_id_count,
        finding.date_rows,
        finding.amount_rows,
        finding.count_rows,
        "" if finding.matched_output_ids is None else finding.matched_output_ids,
        "" if finding.missing_output_ids is None else finding.missing_output_ids,
        json.dumps(finding.columns, ensure_ascii=False, sort_keys=True),
        "；".join(finding.examples),
    ]


def write_csv_report(root: Path, findings: Sequence[SourceFinding], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(REPORT_HEADERS)
        for finding in findings:
            writer.writerow(finding_to_row(root, finding))
    return output_path


def write_xlsx_report(root: Path, findings: Sequence[SourceFinding], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "來源盤點"
    ws.append(REPORT_HEADERS)
    for finding in findings:
        ws.append(finding_to_row(root, finding))
    ws.freeze_panes = "A2"
    widths = {
        "A": 46,
        "B": 10,
        "C": 24,
        "D": 18,
        "E": 12,
        "F": 28,
        "Q": 52,
        "R": 36,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    wb.save(output_path)
    wb.close()
    return output_path


def summarize(findings: Iterable[SourceFinding]) -> dict:
    data = list(findings)
    by_type = {}
    status = {}
    for finding in data:
        by_type[finding.data_type] = by_type.get(finding.data_type, 0) + 1
        status[finding.status] = status.get(finding.status, 0) + 1
    return {
        "files_or_sheets": len(data),
        "by_type": dict(sorted(by_type.items())),
        "status": dict(sorted(status.items())),
        "unique_id_total_by_sheet_sum": sum(f.unique_id_count for f in data),
    }

