# -*- coding: utf-8 -*-
"""
展望系統解析器（source_system = "prospect"）

資料夾特徵：
- R11440/ 子資料夾：逐月就診明細 xlsx（每列=一次就診，含身分證號/總額）
- 照護名單 / 指定名單 xlsx：家醫名單與個案類別
- 自選會員 A115 / 不要自選 A115X xlsx
- 各類篩檢獨立 xlsx（成健/子抹/糞便/老流/BC肝）

月份代碼：從檔名擷取 5 位民國年月（例：11502）
"""
from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

from db_pipeline.config.models import ClinicConfig
from db_pipeline.datasets.models import (
    DatasetBundle,
    MemberRecord,
    MemberSelectionRecord,
    MonthlyClaimRecord,
    SourceTrace,
)
from db_pipeline.normalization import (
    normalize_id,
    normalize_name,
    normalize_phone,
    normalize_text,
    parse_date,
    parse_decimal,
    stable_row_hash,
)
from db_pipeline.parsers.contracts import ParseCoverage, ParseResult
from db_pipeline.validation.models import ValidationIssue

# ── 常數 ──────────────────────────────────────────────────────────────────────
ID_HEADERS = ("身分證號", "身份證號", "身分證號碼", "身份證號碼", "ID")
MEMBER_FIELDS = {
    "case_category":            ("個案類別",),
    "quality_roster":           ("論質名單",),
    "multi_chronic_65":         ("65歲以上多重慢性病註記",),
    "high_visit":               ("高診次註記",),
    "chronic_mark":             ("慢性病註記",),
    "non_chronic_mark":         ("非慢性病註記",),
    "same_clinic_previous_year":("與前一年家醫收案診所相同",),
    "disease_pattern":          ("疾病樣態",),
    "ascvd":                    ("ASCVD",),
    "three_highs":              ("三高",),
    "hypertension":             ("高血壓",),
    "hyperlipidemia":           ("高血脂",),
    "hyperglycemia":            ("高血糖",),
}
MONTH_CODE_RE = re.compile(r"(?<!\d)(1(?:14|15)\d{2})(?!\d)")
TW_ID_RE = re.compile(r"[A-Z][12]\d{8}")
XLSX_SUFFIXES = {".xlsx", ".xlsm"}
EXCLUDE_PREFIXES = ("~$", ".")


# ── 工具函式 ──────────────────────────────────────────────────────────────────
def _header_map(values: Sequence[Any]) -> Dict[str, int]:
    return {
        normalize_text(v): i
        for i, v in enumerate(values)
        if normalize_text(v)
    }


def _find_col(headers: Dict[str, int], aliases: Iterable[str]) -> Optional[int]:
    for alias in aliases:
        if alias in headers:
            return headers[alias]
    return None


def _find_header_row(ws: Any, required: Sequence[Sequence[str]]) -> Optional[int]:
    for row_no in range(1, min(ws.max_row, 12) + 1):
        vals = [ws.cell(row_no, c).value for c in range(1, ws.max_column + 1)]
        hmap = _header_map(vals)
        if all(_find_col(hmap, grp) is not None for grp in required):
            return row_no
    return None


def _month_code(text: str) -> Optional[str]:
    m = MONTH_CODE_RE.search(text)
    return m.group(1) if m else None


def _trace(
    config: ClinicConfig,
    batch_id: str,
    source_dir: Path,
    file_path: Path,
    sheet_name: str,
    row_no: int,
    values: Sequence[Any],
) -> SourceTrace:
    return SourceTrace(
        clinic_code=config.clinic_code,
        batch_id=batch_id,
        source_system=config.source_system,
        source_file=str(file_path.relative_to(source_dir)),
        source_sheet=sheet_name,
        source_row=row_no,
        raw_row_hash=stable_row_hash(values),
    )


# ── 主解析器 ──────────────────────────────────────────────────────────────────
class ProspectParser:
    source_system = "prospect"

    def parse(
        self,
        source_dir: Path,
        config: ClinicConfig,
        batch_id: str,
    ) -> ParseResult:
        bundle = DatasetBundle()
        coverage = ParseCoverage()
        issues: List[ValidationIssue] = []

        files = sorted(
            p for p in source_dir.rglob("*")
            if p.is_file() and not p.name.startswith(EXCLUDE_PREFIXES)
        )
        coverage.discovered_files = len(files)

        # 1. 照護名單 → members
        roster_files = [
            p for p in files
            if p.suffix.lower() in XLSX_SUFFIXES
            and any(kw in p.name for kw in ("照護名單", "指定名單", "指定會員", "家醫名單"))
        ]
        for path in roster_files:
            n = self._parse_member_workbook(source_dir, path, config, batch_id, bundle)
            if n:
                coverage.parsed_files += 1
                coverage.parsed_rows["members"] = coverage.parsed_rows.get("members", 0) + n

        member_indexes = self._build_indexes(bundle.members)

        # 2. R11440/ 逐月明細 → monthly_claims（聚合）
        r11440_files = [
            p for p in files
            if p.suffix.lower() in XLSX_SUFFIXES
            and "R11440" in p.name.upper()
            and _month_code(p.stem) is not None
        ]
        # 也掃子資料夾
        for subdir in ("R11440", "次數", "費用"):
            sd = source_dir / subdir
            if sd.is_dir():
                for p in sorted(sd.glob("*.xlsx")):
                    if not p.name.startswith(EXCLUDE_PREFIXES) and p not in r11440_files:
                        r11440_files.append(p)

        for path in r11440_files:
            parsed, unmatched = self._parse_claim_workbook(
                source_dir, path, config, batch_id, bundle)
            if parsed or unmatched:
                coverage.parsed_files += 1
                coverage.parsed_rows["monthly_claims"] = (
                    coverage.parsed_rows.get("monthly_claims", 0) + parsed)
                coverage.unmatched_rows["monthly_claims"] = (
                    coverage.unmatched_rows.get("monthly_claims", 0) + unmatched)

        # 3. 自選 / 不要 → member_selections
        select_files = [
            p for p in files
            if p.suffix.lower() in XLSX_SUFFIXES
            and any(kw in p.name for kw in ("自選", "A115", "115自選"))
            and not any(kw in p.name.upper() for kw in ("115X", "不要", "不選"))
        ]
        exclude_files = [
            p for p in files
            if p.suffix.lower() in XLSX_SUFFIXES
            and any(kw in p.name.upper() for kw in ("115X", "不要", "不選"))
        ]
        for path in select_files:
            n, u = self._parse_selection_workbook(
                source_dir, path, config, batch_id, member_indexes, bundle, "self_selected_115")
            if n or u:
                coverage.parsed_files += 1
                coverage.parsed_rows["member_selections"] = (
                    coverage.parsed_rows.get("member_selections", 0) + n)
                coverage.unmatched_rows["member_selections"] = (
                    coverage.unmatched_rows.get("member_selections", 0) + u)
        for path in exclude_files:
            n, u = self._parse_selection_workbook(
                source_dir, path, config, batch_id, member_indexes, bundle, "excluded_115x")
            if n or u:
                coverage.parsed_files += 1

        # 4. 未解析檔案
        parsed_paths = set(roster_files + r11440_files + select_files + exclude_files)
        for p in files:
            if p not in parsed_paths:
                coverage.skipped_files[str(p.relative_to(source_dir))] = (
                    "prospect v1 尚未實作此來源類型")

        if r11440_files and not bundle.monthly_claims:
            issues.append(ValidationIssue(
                severity="error", dataset="monthly_claims",
                code="claim_not_mapped",
                message="找到 R11440 來源檔，但沒有任何資料成功對應會員。"))
        for dataset, cnt in coverage.unmatched_rows.items():
            if cnt:
                issues.append(ValidationIssue(
                    severity="warning", dataset=dataset,
                    code="unmatched_source_rows",
                    message=f"有 {cnt} 筆來源資料無法唯一對應會員。"))

        return ParseResult(bundle=bundle, coverage=coverage, issues=issues)

    # ── 照護名單 ──────────────────────────────────────────────────────────────
    def _parse_member_workbook(
        self,
        source_dir: Path, path: Path,
        config: ClinicConfig, batch_id: str,
        bundle: DatasetBundle,
    ) -> int:
        wb = load_workbook(path, read_only=True, data_only=True)
        parsed = 0
        try:
            for ws in wb.worksheets:
                header_row = _find_header_row(ws, (ID_HEADERS, ("個案類別",)))
                if header_row is None:
                    continue
                hvals = [ws.cell(header_row, c).value
                         for c in range(1, ws.max_column + 1)]
                hmap = _header_map(hvals)
                id_col  = _find_col(hmap, ID_HEADERS)
                if id_col is None:
                    continue
                name_col   = _find_col(hmap, ("姓名",))
                birth_col  = _find_col(hmap, ("生日", "出生日期", "BIRTHDAY"))
                phone_col  = _find_col(hmap, ("電話",))
                mobile_col = _find_col(hmap, ("手機", "手機號碼", "行動電話"))
                addr_col   = _find_col(hmap, ("地址", "住址"))
                field_cols = {
                    fn: _find_col(hmap, aliases)
                    for fn, aliases in MEMBER_FIELDS.items()
                }
                for row_no, vals in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    pid = normalize_id(vals[id_col])
                    if not TW_ID_RE.fullmatch(pid):
                        continue
                    kwargs = {
                        fn: normalize_text(vals[col]) if col is not None else ""
                        for fn, col in field_cols.items()
                    }
                    trace = _trace(config, batch_id, source_dir, path, ws.title, row_no, vals)
                    bundle.members.append(MemberRecord(
                        trace=trace, person_id=pid,
                        name=normalize_name(vals[name_col]) if name_col is not None else "",
                        birth_date=parse_date(vals[birth_col]) if birth_col is not None else None,
                        phone=normalize_phone(vals[phone_col]) if phone_col is not None else "",
                        mobile=normalize_phone(vals[mobile_col]) if mobile_col is not None else "",
                        address=normalize_text(vals[addr_col]) if addr_col is not None else "",
                        **kwargs,
                    ))
                    bundle.member_selections.append(MemberSelectionRecord(
                        trace=trace, person_id=pid, selection_type="designated_114"))
                    parsed += 1
        finally:
            wb.close()
        return parsed

    # ── R11440 就診明細（聚合） ───────────────────────────────────────────────
    def _parse_claim_workbook(
        self,
        source_dir: Path, path: Path,
        config: ClinicConfig, batch_id: str,
        bundle: DatasetBundle,
    ) -> Tuple[int, int]:
        """
        展望 R11440 xlsx 每列 = 一次就診。
        按身分證號聚合成一筆 MonthlyClaimRecord（次數=出現次數，金額=總額欄加總）。
        """
        code = _month_code(path.stem) or _month_code(path.name)
        if code is None:
            return 0, 0
        roc_year = int(code[:3])
        month    = int(code[3:])

        wb = load_workbook(path, read_only=True, data_only=True)
        parsed = unmatched = 0
        try:
            for ws in wb.worksheets:
                header_row = _find_header_row(
                    ws, (ID_HEADERS,))
                if header_row is None:
                    continue
                hvals = [ws.cell(header_row, c).value
                         for c in range(1, ws.max_column + 1)]
                hmap = _header_map(hvals)
                id_col     = _find_col(hmap, ID_HEADERS)
                amount_col = _find_col(hmap, ("總額", "申請金額", "申報總額", "費用"))
                date_col   = _find_col(hmap, ("日期", "看診日期", "就醫日期"))
                if id_col is None:
                    continue

                # 聚合：{person_id: {count, amount, last_visit_date}}
                agg: Dict[str, dict] = {}
                for row_no, vals in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    pid = normalize_id(vals[id_col])
                    if not TW_ID_RE.fullmatch(pid):
                        continue
                    amt = parse_decimal(vals[amount_col]) if amount_col is not None else parse_decimal(0)
                    vdate = parse_date(vals[date_col]) if date_col is not None else None
                    if pid not in agg:
                        agg[pid] = {"count": 0, "amount": amt, "last_visit": vdate,
                                    "row_no": row_no, "vals": vals}
                    else:
                        agg[pid]["count"] += 1
                        agg[pid]["amount"] += amt
                        if vdate and (agg[pid]["last_visit"] is None
                                      or vdate > agg[pid]["last_visit"]):
                            agg[pid]["last_visit"] = vdate

                # 登錄
                existing_ids = {m.person_id for m in bundle.members}
                for pid, rec in agg.items():
                    if pid not in existing_ids:
                        unmatched += 1
                        continue
                    bundle.monthly_claims.append(MonthlyClaimRecord(
                        trace=_trace(config, batch_id, source_dir, path,
                                     ws.title, rec["row_no"], rec["vals"]),
                        person_id=pid,
                        roc_year=roc_year,
                        month=month,
                        visit_count=parse_decimal(rec["count"] + 1),
                        amount=rec["amount"],
                        last_visit_date=rec["last_visit"],
                    ))
                    parsed += 1
        finally:
            wb.close()
        return parsed, unmatched

    # ── 自選 / 不要 ──────────────────────────────────────────────────────────
    def _parse_selection_workbook(
        self,
        source_dir: Path, path: Path,
        config: ClinicConfig, batch_id: str,
        indexes: Dict[str, Any],
        bundle: DatasetBundle,
        selection_type: str,
    ) -> Tuple[int, int]:
        wb = load_workbook(path, read_only=True, data_only=True)
        parsed = unmatched = 0
        existing_ids = {m.person_id for m in bundle.members}
        try:
            for ws in wb.worksheets:
                header_row = _find_header_row(ws, (ID_HEADERS,))
                if header_row is None:
                    header_row = _find_header_row(ws, (("姓名",), ("電話", "手機")))
                if header_row is None:
                    continue
                hvals = [ws.cell(header_row, c).value
                         for c in range(1, ws.max_column + 1)]
                hmap = _header_map(hvals)
                id_col   = _find_col(hmap, ID_HEADERS)
                name_col = _find_col(hmap, ("姓名",))
                phone_col = _find_col(hmap, ("電話", "手機", "手機號碼"))

                for row_no, vals in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    pid = normalize_id(vals[id_col]) if id_col is not None else ""
                    if TW_ID_RE.fullmatch(pid):
                        pass  # 有直接 ID
                    else:
                        # 以姓名+電話比對
                        name = normalize_name(vals[name_col]) if name_col is not None else ""
                        phone = normalize_phone(vals[phone_col]) if phone_col is not None else ""
                        if not name or len(phone) < 7:
                            continue
                        ids = indexes.get("name_phone", {}).get((name, phone[-7:]), set())
                        if len(ids) != 1:
                            unmatched += 1
                            continue
                        pid = next(iter(ids))

                    if pid not in existing_ids:
                        unmatched += 1
                        continue
                    bundle.member_selections.append(MemberSelectionRecord(
                        trace=_trace(config, batch_id, source_dir, path,
                                     ws.title, row_no, vals),
                        person_id=pid,
                        selection_type=selection_type,
                    ))
                    parsed += 1
        finally:
            wb.close()
        return parsed, unmatched

    @staticmethod
    def _build_indexes(members: Sequence[MemberRecord]) -> Dict[str, Dict]:
        idxs: Dict[str, Dict] = {"name_phone": defaultdict(set)}
        for m in members:
            name = normalize_name(m.name)
            for phone in (m.phone, m.mobile):
                if phone:
                    idxs["name_phone"][(name, phone[-7:])].add(m.person_id)
        return idxs
