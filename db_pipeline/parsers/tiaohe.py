# -*- coding: utf-8 -*-
"""
調和系統解析器（source_system = "tiaohe"）

資料夾特徵：
- 自選會員 CSV（BIG5，可能含非法 XML 控制字元）
- 不要自選 CSV（BIG5）
- R11440次數/ 或 次數/ 子資料夾：逐月次數 xlsx（含身分證號+次數）
- sm_*_cliP4pCase / cliP4pTrack / cliAssay / cliScores xlsx

處理要點：
- CSV 先以多種編碼嘗試讀取，清除非法控制字元後儲存 UTF-8-SIG 暫存
- 次數 xlsx 直接含身分證號，不需姓名比對
- sm_* 部分與 SmParser 邏輯相同（P4P/篩檢/檢驗預留 v2）
"""
from __future__ import annotations

import csv
import io
import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

from db_pipeline.config.models import ClinicConfig
from db_pipeline.datasets.models import (
    DatasetBundle,
    MemberRecord,
    MemberSelectionRecord,
    MonthlyClaimRecord,
    SourceTrace,
)
from db_pipeline.normalization import (
    normalize_id,
    normalize_name,
    normalize_phone,
    normalize_text,
    parse_date,
    parse_decimal,
    stable_row_hash,
)
from db_pipeline.parsers.contracts import ParseCoverage, ParseResult
from db_pipeline.validation.models import ValidationIssue

# ── 常數 ──────────────────────────────────────────────────────────────────────
ID_HEADERS = ("身分證號", "身份證號", "身分證號碼", "身份證號碼", "ID", "家醫收案會員ID")
MEMBER_FIELDS = {
    "case_category":            ("個案類別",),
    "quality_roster":           ("論質名單",),
    "multi_chronic_65":         ("65歲以上多重慢性病註記",),
    "high_visit":               ("高診次註記",),
    "chronic_mark":             ("慢性病註記",),
    "non_chronic_mark":         ("非慢性病註記",),
    "same_clinic_previous_year":("與前一年家醫收案診所相同",),
    "disease_pattern":          ("疾病樣態",),
    "ascvd":                    ("ASCVD",),
    "three_highs":              ("三高",),
    "hypertension":             ("高血壓",),
    "hyperlipidemia":           ("高血脂",),
    "hyperglycemia":            ("高血糖",),
}
MONTH_CODE_RE     = re.compile(r"(?<!\d)(1(?:14|15)\d{2})(?!\d)")
TW_ID_RE          = re.compile(r"[A-Z][12]\d{8}")
ILLEGAL_CTRL_RE   = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")
XLSX_SUFFIXES     = {".xlsx", ".xlsm"}
CSV_ENCODINGS     = ["utf-8-sig", "utf-16", "cp950", "big5"]


# ── 工具函式 ──────────────────────────────────────────────────────────────────
def _header_map(values: Sequence[Any]) -> Dict[str, int]:
    return {normalize_text(v): i for i, v in enumerate(values) if normalize_text(v)}


def _find_col(headers: Dict[str, int], aliases: Iterable[str]) -> Optional[int]:
    for a in aliases:
        if a in headers:
            return headers[a]
    return None


def _find_header_row(ws: Any, required: Sequence[Sequence[str]]) -> Optional[int]:
    for row_no in range(1, min(ws.max_row, 12) + 1):
        vals = [ws.cell(row_no, c).value for c in range(1, ws.max_column + 1)]
        hmap = _header_map(vals)
        if all(_find_col(hmap, grp) is not None for grp in required):
            return row_no
    return None


def _month_code(text: str) -> Optional[str]:
    m = MONTH_CODE_RE.search(text)
    return m.group(1) if m else None


def _read_csv_rows(path: Path) -> List[List[str]]:
    """嘗試多種編碼讀取 CSV，並清除非法控制字元。"""
    raw: Optional[str] = None
    for enc in CSV_ENCODINGS:
        try:
            raw = path.read_text(encoding=enc, errors="strict")
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    if raw is None:
        raw = path.read_text(encoding="cp950", errors="replace")
    cleaned = ILLEGAL_CTRL_RE.sub("", raw)
    return list(csv.reader(io.StringIO(cleaned)))


def _trace(
    config: ClinicConfig, batch_id: str, source_dir: Path,
    file_path: Path, sheet_name: str, row_no: int, values: Sequence[Any],
) -> SourceTrace:
    return SourceTrace(
        clinic_code=config.clinic_code, batch_id=batch_id,
        source_system=config.source_system,
        source_file=str(file_path.relative_to(source_dir)),
        source_sheet=sheet_name, source_row=row_no,
        raw_row_hash=stable_row_hash(values),
    )


# ── 主解析器 ──────────────────────────────────────────────────────────────────
class TiaoheParser:
    source_system = "tiaohe"

    def parse(
        self,
        source_dir: Path,
        config: ClinicConfig,
        batch_id: str,
    ) -> ParseResult:
        bundle   = DatasetBundle()
        coverage = ParseCoverage()
        issues: List[ValidationIssue] = []

        files = sorted(
            p for p in source_dir.rglob("*")
            if p.is_file() and not p.name.startswith(("~$", "."))
        )
        coverage.discovered_files = len(files)

        # 1. 照護名單 xlsx → members
        roster_files = [
            p for p in files
            if p.suffix.lower() in XLSX_SUFFIXES
            and any(kw in p.name for kw in ("照護名單", "指定名單", "家醫名單", "指定會員"))
        ]
        for path in roster_files:
            n = self._parse_member_workbook(source_dir, path, config, batch_id, bundle)
            if n:
                coverage.parsed_files += 1
                coverage.parsed_rows["members"] = coverage.parsed_rows.get("members", 0) + n

        existing_ids = {m.person_id for m in bundle.members}

        # 2. 次數 xlsx（R11440次數/ 或 次數/ 或 主次代碼/）→ monthly_claims
        count_files: List[Path] = []
        for subdir_name in ("R11440次數", "次數", "主次代碼"):
            sd = source_dir / subdir_name
            if sd.is_dir():
                count_files.extend(sorted(
                    p for p in sd.glob("*.xlsx")
                    if not p.name.startswith(("~$", "."))
                    and _month_code(p.stem) is not None
                ))
        # 也掃根目錄下含月份代碼的 xlsx（R11440次數_轉檔OK.xlsx 等）
        for p in files:
            if (p.suffix.lower() in XLSX_SUFFIXES
                    and _month_code(p.stem) is not None
                    and p not in count_files
                    and p not in roster_files):
                count_files.append(p)

        for path in count_files:
            parsed, unmatched = self._parse_count_workbook(
                source_dir, path, config, batch_id, existing_ids, bundle)
            if parsed or unmatched:
                coverage.parsed_files += 1
                coverage.parsed_rows["monthly_claims"] = (
                    coverage.parsed_rows.get("monthly_claims", 0) + parsed)
                coverage.unmatched_rows["monthly_claims"] = (
                    coverage.unmatched_rows.get("monthly_claims", 0) + unmatched)

        # 3. 自選 / 不要 CSV（BIG5）
        select_csvs = [
            p for p in files
            if p.suffix.lower() == ".csv"
            and any(kw in p.name for kw in ("自選", "A115", "115自選"))
            and not any(kw in p.name.upper() for kw in ("115X", "不要"))
        ]
        exclude_csvs = [
            p for p in files
            if p.suffix.lower() == ".csv"
            and any(kw in p.name.upper() for kw in ("115X", "不要"))
        ]
        for path in select_csvs:
            n, u = self._parse_selection_csv(
                source_dir, path, config, batch_id, existing_ids, bundle, "self_selected_115")
            if n or u:
                coverage.parsed_files += 1
                coverage.parsed_rows["member_selections"] = (
                    coverage.parsed_rows.get("member_selections", 0) + n)
                coverage.unmatched_rows["member_selections"] = (
                    coverage.unmatched_rows.get("member_selections", 0) + u)
        for path in exclude_csvs:
            n, u = self._parse_selection_csv(
                source_dir, path, config, batch_id, existing_ids, bundle, "excluded_115x")
            if n or u:
                coverage.parsed_files += 1

        # 4. 跳過
        parsed_paths = set(roster_files + count_files + select_csvs + exclude_csvs)
        for p in files:
            if p not in parsed_paths:
                coverage.skipped_files[str(p.relative_to(source_dir))] = (
                    "tiaohe v1 尚未實作此來源類型（sm_* P4P/篩檢預留 v2）")

        if count_files and not bundle.monthly_claims:
            issues.append(ValidationIssue(
                severity="error", dataset="monthly_claims",
                code="claim_not_mapped",
                message="找到次數來源檔，但沒有任何資料成功對應會員。"))
        for dataset, cnt in coverage.unmatched_rows.items():
            if cnt:
                issues.append(ValidationIssue(
                    severity="warning", dataset=dataset,
                    code="unmatched_source_rows",
                    message=f"有 {cnt} 筆來源資料無法唯一對應會員。"))

        return ParseResult(bundle=bundle, coverage=coverage, issues=issues)

    # ── 照護名單 ──────────────────────────────────────────────────────────────
    def _parse_member_workbook(
        self,
        source_dir: Path, path: Path,
        config: ClinicConfig, batch_id: str,
        bundle: DatasetBundle,
    ) -> int:
        wb = load_workbook(path, read_only=True, data_only=True)
        parsed = 0
        try:
            for ws in wb.worksheets:
                header_row = _find_header_row(ws, (ID_HEADERS, ("個案類別",)))
                if header_row is None:
                    continue
                hvals = [ws.cell(header_row, c).value
                         for c in range(1, ws.max_column + 1)]
                hmap   = _header_map(hvals)
                id_col = _find_col(hmap, ID_HEADERS)
                if id_col is None:
                    continue
                name_col   = _find_col(hmap, ("姓名",))
                birth_col  = _find_col(hmap, ("生日", "出生日期", "BIRTHDAY"))
                phone_col  = _find_col(hmap, ("電話",))
                mobile_col = _find_col(hmap, ("手機", "手機號碼"))
                addr_col   = _find_col(hmap, ("地址", "住址"))
                field_cols = {fn: _find_col(hmap, aliases)
                              for fn, aliases in MEMBER_FIELDS.items()}
                for row_no, vals in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    pid = normalize_id(vals[id_col])
                    if not TW_ID_RE.fullmatch(pid):
                        continue
                    kwargs = {fn: normalize_text(vals[col]) if col is not None else ""
                              for fn, col in field_cols.items()}
                    trace = _trace(config, batch_id, source_dir, path, ws.title, row_no, vals)
                    bundle.members.append(MemberRecord(
                        trace=trace, person_id=pid,
                        name=normalize_name(vals[name_col]) if name_col is not None else "",
                        birth_date=parse_date(vals[birth_col]) if birth_col is not None else None,
                        phone=normalize_phone(vals[phone_col]) if phone_col is not None else "",
                        mobile=normalize_phone(vals[mobile_col]) if mobile_col is not None else "",
                        address=normalize_text(vals[addr_col]) if addr_col is not None else "",
                        **kwargs,
                    ))
                    bundle.member_selections.append(MemberSelectionRecord(
                        trace=trace, person_id=pid, selection_type="designated_114"))
                    parsed += 1
        finally:
            wb.close()
        return parsed

    # ── 次數 xlsx ─────────────────────────────────────────────────────────────
    def _parse_count_workbook(
        self,
        source_dir: Path, path: Path,
        config: ClinicConfig, batch_id: str,
        existing_ids: set,
        bundle: DatasetBundle,
    ) -> Tuple[int, int]:
        code = _month_code(path.stem) or _month_code(path.name)
        if code is None:
            return 0, 0
        roc_year = int(code[:3])
        month    = int(code[3:])

        wb = load_workbook(path, read_only=True, data_only=True)
        parsed = unmatched = 0
        try:
            for ws in wb.worksheets:
                header_row = _find_header_row(ws, (ID_HEADERS, ("次數", "件數")))
                if header_row is None:
                    continue
                hvals = [ws.cell(header_row, c).value
                         for c in range(1, ws.max_column + 1)]
                hmap      = _header_map(hvals)
                id_col    = _find_col(hmap, ID_HEADERS)
                count_col = _find_col(hmap, ("次數", "件數", "就診次數"))
                amount_col = _find_col(hmap, ("申請金額", "金額", "費用"))
                if id_col is None or count_col is None:
                    continue
                for row_no, vals in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    pid = normalize_id(vals[id_col])
                    if not TW_ID_RE.fullmatch(pid):
                        continue
                    if pid not in existing_ids:
                        unmatched += 1
                        continue
                    bundle.monthly_claims.append(MonthlyClaimRecord(
                        trace=_trace(config, batch_id, source_dir, path,
                                     ws.title, row_no, vals),
                        person_id=pid,
                        roc_year=roc_year,
                        month=month,
                        visit_count=parse_decimal(vals[count_col]),
                        amount=parse_decimal(vals[amount_col]) if amount_col is not None else parse_decimal(0),
                    ))
                    parsed += 1
        finally:
            wb.close()
        return parsed, unmatched

    # ── 自選 / 不要 CSV ───────────────────────────────────────────────────────
    def _parse_selection_csv(
        self,
        source_dir: Path, path: Path,
        config: ClinicConfig, batch_id: str,
        existing_ids: set,
        bundle: DatasetBundle,
        selection_type: str,
    ) -> Tuple[int, int]:
        rows = _read_csv_rows(path)
        if len(rows) < 2:
            return 0, 0

        # 找表頭（前 5 列）
        hdr_idx = 0
        for i, row in enumerate(rows[:5]):
            hmap = {normalize_text(v): j for j, v in enumerate(row)}
            if _find_col(hmap, ID_HEADERS) is not None:
                hdr_idx = i
                break
        hmap   = {normalize_text(v): j for j, v in enumerate(rows[hdr_idx])}
        id_col = _find_col(hmap, ID_HEADERS)
        if id_col is None:
            return 0, 0

        parsed = unmatched = 0
        for row_no, row in enumerate(rows[hdr_idx + 1:], start=hdr_idx + 2):
            if id_col >= len(row):
                continue
            pid = normalize_id(row[id_col])
            if not TW_ID_RE.fullmatch(pid):
                continue
            if pid not in existing_ids:
                unmatched += 1
                continue
            bundle.member_selections.append(MemberSelectionRecord(
                trace=SourceTrace(
                    clinic_code=config.clinic_code,
                    batch_id=batch_id,
                    source_system=config.source_system,
                    source_file=str(path.relative_to(source_dir)),
                    source_sheet="",
                    source_row=row_no,
                    raw_row_hash=stable_row_hash(row),
                ),
                person_id=pid,
                selection_type=selection_type,
            ))
            parsed += 1
        return parsed, unmatched
