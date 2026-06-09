# -*- coding: utf-8 -*-
"""
方鼎系統解析器（source_system = "fangding"）

資料夾特徵：
- 照護名單 xlsx（與其他系統格式相同）
- 次數/ 子資料夾：門診依就診次數統計表YYMM.csv（cp950，3 列標頭）
- 可選：P4P、個案管理、篩檢 xlsx（目前略過）
"""
from __future__ import annotations

import csv
import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

from db_pipeline.config.models import ClinicConfig
from db_pipeline.datasets.models import (
    DatasetBundle,
    MemberRecord,
    MemberSelectionRecord,
    MonthlyClaimRecord,
    SourceTrace,
)
from db_pipeline.normalization import (
    normalize_id,
    normalize_name,
    normalize_phone,
    normalize_text,
    parse_date,
    parse_decimal,
    stable_row_hash,
)
from db_pipeline.parsers.解析器介面 import ParseCoverage, ParseResult
from db_pipeline.validation.models import ValidationIssue

ID_HEADERS = ("ID", "身分證號", "身份證號", "身分證號碼", "身份證號碼")
MEMBER_FIELDS = {
    "case_category":             ("個案類別",),
    "quality_roster":            ("論質名單",),
    "multi_chronic_65":          ("65歲以上多重慢性病註記",),
    "high_visit":                ("高診次註記",),
    "chronic_mark":              ("慢性病註記",),
    "non_chronic_mark":          ("非慢性病註記",),
    "same_clinic_previous_year": ("與前一年家醫收案診所相同",),
    "disease_pattern":           ("疾病樣態",),
    "ascvd":                     ("ASCVD",),
    "three_highs":               ("三高",),
    "hypertension":              ("高血壓",),
    "hyperlipidemia":            ("高血脂",),
    "hyperglycemia":             ("高血糖",),
}
_MONTH_RE = re.compile(r"(1[01][0-9])(0[1-9]|1[0-2])")


def _header_map(values: Sequence[Any]) -> Dict[str, int]:
    return {normalize_text(v): i for i, v in enumerate(values) if normalize_text(v)}


def _find_col(headers: Dict[str, int], aliases: Iterable[str]) -> Optional[int]:
    for a in aliases:
        if a in headers:
            return headers[a]
    return None


def _find_header_row(ws: Any, required: Sequence[Sequence[str]]) -> Optional[int]:
    for r in range(1, min(ws.max_row, 12) + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        hdr = _header_map(vals)
        if all(_find_col(hdr, grp) is not None for grp in required):
            return r
    return None


def _trace(config, batch_id, source_dir, file_path, sheet, row_no, values):
    return SourceTrace(
        clinic_code=config.clinic_code,
        batch_id=batch_id,
        source_system=config.source_system,
        source_file=str(file_path.relative_to(source_dir)),
        source_sheet=sheet,
        source_row=row_no,
        raw_row_hash=stable_row_hash(values),
    )


class FangdingParser:
    source_system = "fangding"

    def parse(self, source_dir: Path, config: ClinicConfig, batch_id: str) -> ParseResult:
        bundle = DatasetBundle()
        coverage = ParseCoverage()
        issues: List[ValidationIssue] = []
        files = sorted(p for p in source_dir.rglob("*")
                       if p.is_file() and not p.name.startswith(("~$", ".")))
        coverage.discovered_files = len(files)

        roster_files = [
            p for p in files
            if p.suffix.lower() in (".xlsx", ".xlsm")
            and any(kw in p.name for kw in ("照護名單", "指定名單", "家醫名單", "指定會員"))
        ]
        for path in roster_files:
            n = self._parse_roster(source_dir, path, config, batch_id, bundle)
            if n:
                coverage.parsed_files += 1
                coverage.parsed_rows["members"] = coverage.parsed_rows.get("members", 0) + n

        existing_ids = {m.person_id for m in bundle.members}
        member_indexes = self._build_indexes(bundle.members)

        # claims: CSV files in 次數/ subfolder
        claim_csv_dir = source_dir / "次數"
        claim_files = sorted(claim_csv_dir.glob("*.csv")) if claim_csv_dir.is_dir() else []
        claim_files += [p for p in files if p.suffix.lower() == ".csv"
                        and "次數" in p.name and p.parent == source_dir]
        for path in claim_files:
            n = self._parse_claims_csv(source_dir, path, config, batch_id, bundle)
            if n:
                coverage.parsed_files += 1
                coverage.parsed_rows["monthly_claims"] = (
                    coverage.parsed_rows.get("monthly_claims", 0) + n)

        # self-select / exclusion
        for path in files:
            name = path.name
            if path.suffix.lower() not in (".xlsx", ".xlsm"):
                continue
            if "自選" in name and "不要" not in name and "115X" not in name.upper():
                n, _ = self._parse_flag_workbook(
                    source_dir, path, config, batch_id,
                    "self_selected_115", existing_ids, bundle)
                if n:
                    coverage.parsed_files += 1
            elif "不要" in name or "115X" in name.upper():
                n, _ = self._parse_flag_workbook(
                    source_dir, path, config, batch_id,
                    "excluded_115x", existing_ids, bundle)
                if n:
                    coverage.parsed_files += 1

        return ParseResult(bundle=bundle, coverage=coverage, issues=issues)

    # ── roster ──────────────────────────────────────────────────────────────
    def _parse_roster(self, source_dir, path, config, batch_id, bundle) -> int:
        wb = load_workbook(path, read_only=True, data_only=True)
        parsed = 0
        try:
            for ws in wb.worksheets:
                header_row = _find_header_row(ws, (ID_HEADERS, ("個案類別",)))
                if header_row is None:
                    continue
                hvals = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
                hdr = _header_map(hvals)
                id_col = _find_col(hdr, ID_HEADERS)
                if id_col is None:
                    continue
                birth_col  = _find_col(hdr, ("BIRTHDAY", "生日"))
                field_cols = {f: _find_col(hdr, aliases) for f, aliases in MEMBER_FIELDS.items()}
                for row_no, vals in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    pid = normalize_id(vals[id_col])
                    if not re.fullmatch(r"[A-Z][12][0-9]{8}", pid):
                        continue
                    kwargs = {
                        f: (normalize_text(vals[col]) if col is not None else "")
                        for f, col in field_cols.items()
                    }
                    bundle.members.append(MemberRecord(
                        trace=_trace(config, batch_id, source_dir, path, ws.title, row_no, vals),
                        person_id=pid,
                        birth_date=parse_date(vals[birth_col]) if birth_col is not None else None,
                        **kwargs,
                    ))
                    bundle.member_selections.append(MemberSelectionRecord(
                        trace=bundle.members[-1].trace,
                        person_id=pid,
                        selection_type="designated_114",
                    ))
                    parsed += 1
        finally:
            wb.close()
        return parsed

    # ── claims CSV ──────────────────────────────────────────────────────────
    def _parse_claims_csv(self, source_dir, path, config, batch_id, bundle) -> int:
        m = _MONTH_RE.search(path.stem)
        if not m:
            return 0
        roc_year, month = int(m.group(1)), int(m.group(2))
        parsed = 0
        for enc in ("cp950", "big5", "utf-8-sig"):
            try:
                with open(path, encoding=enc, errors="replace", newline="") as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                break
            except Exception:
                continue
        else:
            return 0

        # Find header row (contains 身份證字號)
        hdr_row_idx = None
        for i, row in enumerate(rows):
            if any("身份證" in str(c) for c in row):
                hdr_row_idx = i
                break
        if hdr_row_idx is None:
            return 0

        hdr = _header_map(rows[hdr_row_idx])
        id_col    = _find_col(hdr, ("身份證字號", "身分證號", "身份證號"))
        visit_col = _find_col(hdr, ("申報筆數", "就診次數", "次數"))
        amount_col= _find_col(hdr, ("合計金額", "申請金額", "金額"))
        if id_col is None or visit_col is None:
            return 0

        for row in rows[hdr_row_idx + 1:]:
            if not row or len(row) <= id_col:
                continue
            pid = normalize_id(row[id_col])
            if not re.fullmatch(r"[A-Z][12][0-9]{8}", pid):
                continue
            visit_count = parse_decimal(row[visit_col] if visit_col < len(row) else "")
            amount      = parse_decimal(row[amount_col] if amount_col and amount_col < len(row) else "")
            bundle.monthly_claims.append(MonthlyClaimRecord(
                trace=SourceTrace(
                    clinic_code=config.clinic_code, batch_id=batch_id,
                    source_system=config.source_system,
                    source_file=str(path.relative_to(source_dir)),
                    source_sheet="", source_row=0,
                    raw_row_hash=stable_row_hash(tuple(row)),
                ),
                person_id=pid,
                roc_year=roc_year, month=month,
                visit_count=visit_count,
                amount=amount,
            ))
            parsed += 1
        return parsed

    # ── flag workbook (self-select / exclusion) ──────────────────────────────
    def _parse_flag_workbook(
        self, source_dir, path, config, batch_id, flag_type, existing_ids, bundle
    ) -> Tuple[int, int]:
        wb = load_workbook(path, read_only=True, data_only=True)
        matched = unmatched = 0
        try:
            for ws in wb.worksheets:
                header_row = _find_header_row(ws, (ID_HEADERS,))
                if header_row is None:
                    continue
                hvals = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
                id_col = _find_col(_header_map(hvals), ID_HEADERS)
                if id_col is None:
                    continue
                for row_no, vals in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    pid = normalize_id(vals[id_col])
                    if not re.fullmatch(r"[A-Z][12][0-9]{8}", pid):
                        continue
                    bundle.member_selections.append(MemberSelectionRecord(
                        trace=_trace(config, batch_id, source_dir, path, ws.title, row_no, vals),
                        person_id=pid,
                        selection_type=flag_type,
                    ))
                    if pid in existing_ids:
                        matched += 1
                    else:
                        unmatched += 1
        finally:
            wb.close()
        return matched, unmatched

    def _build_indexes(self, members):
        idx = {"name_birth": defaultdict(set)}
        for m in members:
            if m.birth_date:
                idx["name_birth"][(normalize_name(m.name), m.birth_date.isoformat())].add(m.person_id)
        return idx
