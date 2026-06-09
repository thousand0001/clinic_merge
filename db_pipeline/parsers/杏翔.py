# -*- coding: utf-8 -*-
"""
杏翔系統解析器（source_system = "xingxiang"）

資料夾特徵：
- 照護會員收案追蹤報表 xlsx（格式同標準照護名單）
- 次數/ 子資料夾：YYMM.xlsx（無身分證，用姓名＋生日比對）
- 115自選會員(含成健).xlsx
- 115X不要會員.xlsx
- sm_* 系列（P4P/篩檢/檢驗）：略過（可加入 v2）
"""
from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

from db_pipeline.config.models import ClinicConfig
from db_pipeline.datasets.models import (
    DatasetBundle,
    MemberRecord,
    MemberSelectionRecord,
    MonthlyClaimRecord,
    SourceTrace,
)
from db_pipeline.normalization import (
    normalize_id,
    normalize_name,
    normalize_phone,
    normalize_text,
    parse_date,
    parse_decimal,
    stable_row_hash,
)
from db_pipeline.parsers.解析器介面 import ParseCoverage, ParseResult
from db_pipeline.validation.models import ValidationIssue

ID_HEADERS = ("ID", "身分證號", "身份證號", "身分證號碼", "身份證號碼", "家醫收案會員ID")
MEMBER_FIELDS = {
    "case_category":             ("個案類別",),
    "quality_roster":            ("論質名單",),
    "multi_chronic_65":          ("65歲以上多重慢性病註記",),
    "high_visit":                ("高診次註記",),
    "chronic_mark":              ("慢性病註記",),
    "non_chronic_mark":          ("非慢性病註記",),
    "same_clinic_previous_year": ("與前一年家醫收案診所相同",),
    "disease_pattern":           ("疾病樣態",),
    "ascvd":                     ("ASCVD",),
    "three_highs":               ("三高",),
    "hypertension":              ("高血壓",),
    "hyperlipidemia":            ("高血脂",),
    "hyperglycemia":             ("高血糖",),
}
_MONTH_RE   = re.compile(r"(1[01][0-9])(0[1-9]|1[0-2])")
_TW_ID_RE   = re.compile(r"[A-Z][12]\d{8}")
_XLSX_SUF   = {".xlsx", ".xlsm"}


def _header_map(values: Sequence[Any]) -> Dict[str, int]:
    return {normalize_text(v): i for i, v in enumerate(values) if normalize_text(v)}


def _find_col(headers: Dict[str, int], aliases: Iterable[str]) -> Optional[int]:
    for a in aliases:
        if a in headers:
            return headers[a]
    return None


def _find_header_row(ws: Any, required: Sequence[Sequence[str]]) -> Optional[int]:
    for r in range(1, min(ws.max_row, 12) + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        hdr = _header_map(vals)
        if all(_find_col(hdr, grp) is not None for grp in required):
            return r
    return None


def _trace(config, batch_id, source_dir, file_path, sheet, row_no, values):
    return SourceTrace(
        clinic_code=config.clinic_code, batch_id=batch_id,
        source_system=config.source_system,
        source_file=str(file_path.relative_to(source_dir)),
        source_sheet=sheet, source_row=row_no,
        raw_row_hash=stable_row_hash(values),
    )


class XingxiangParser:
    source_system = "xingxiang"

    def parse(self, source_dir: Path, config: ClinicConfig, batch_id: str) -> ParseResult:
        bundle   = DatasetBundle()
        coverage = ParseCoverage()
        issues: List[ValidationIssue] = []

        files = sorted(p for p in source_dir.rglob("*")
                       if p.is_file() and not p.name.startswith(("~$", ".")))
        coverage.discovered_files = len(files)

        # 1. 照護名單（照護會員收案追蹤報表 OR 照護名單 等標準名稱）
        roster_kws = ("照護會員收案追蹤報表", "照護名單", "指定名單", "家醫名單", "指定會員")
        roster_files = [
            p for p in files
            if p.suffix.lower() in _XLSX_SUF
            and any(kw in p.name for kw in roster_kws)
        ]
        for path in roster_files:
            n = self._parse_roster(source_dir, path, config, batch_id, bundle)
            if n:
                coverage.parsed_files += 1
                coverage.parsed_rows["members"] = coverage.parsed_rows.get("members", 0) + n

        existing_ids = {m.person_id for m in bundle.members}

        # Build name+birth → ID from sm_* files (roster has no name column)
        sm_files = [p for p in files if p.name.startswith("sm_") and p.suffix.lower() in _XLSX_SUF]
        name_birth_idx = self._build_name_birth_from_sm(sm_files, existing_ids)

        # 2. 次數/ XLSX → monthly_claims（姓名＋生日比對）
        claim_dir = source_dir / "次數"
        claim_files = sorted(claim_dir.glob("*.xlsx")) if claim_dir.is_dir() else []
        for path in claim_files:
            if path.name.startswith(("~$", ".")):
                continue
            n, u = self._parse_claims_xlsx(
                source_dir, path, config, batch_id, name_birth_idx, bundle)
            if n or u:
                coverage.parsed_files += 1
                coverage.parsed_rows["monthly_claims"] = (
                    coverage.parsed_rows.get("monthly_claims", 0) + n)
                coverage.unmatched_rows["monthly_claims"] = (
                    coverage.unmatched_rows.get("monthly_claims", 0) + u)

        # 3. 自選 / 不要
        for path in files:
            if path.suffix.lower() not in _XLSX_SUF:
                continue
            name = path.name
            is_excl = "115X" in name.upper() or "不要" in name
            is_self  = ("自選" in name or "115自選" in name) and not is_excl
            if is_self:
                n, u = self._parse_flag_workbook(
                    source_dir, path, config, batch_id,
                    "self_selected_115", existing_ids, name_birth_idx, bundle)
                if n or u:
                    coverage.parsed_files += 1
            elif is_excl:
                n, u = self._parse_flag_workbook(
                    source_dir, path, config, batch_id,
                    "excluded_115x", existing_ids, name_birth_idx, bundle)
                if n or u:
                    coverage.parsed_files += 1

        if coverage.unmatched_rows.get("monthly_claims", 0):
            issues.append(ValidationIssue(
                severity="warning", dataset="monthly_claims",
                code="unmatched_source_rows",
                message=(
                    f"次數 XLSX 有 {coverage.unmatched_rows['monthly_claims']} 筆"
                    "姓名＋生日無法比對照護名單，已略過。"
                )))
        return ParseResult(bundle=bundle, coverage=coverage, issues=issues)

    # ── 照護名單 ────────────────────────────────────────────────────────────────
    def _parse_roster(self, source_dir, path, config, batch_id, bundle) -> int:
        wb = load_workbook(path, read_only=True, data_only=True)
        parsed = 0
        try:
            for ws in wb.worksheets:
                header_row = _find_header_row(ws, (ID_HEADERS, ("個案類別",)))
                if header_row is None:
                    continue
                hvals = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
                hmap   = _header_map(hvals)
                id_col = _find_col(hmap, ID_HEADERS)
                if id_col is None:
                    continue
                name_col   = _find_col(hmap, ("姓名",))
                birth_col  = _find_col(hmap, ("生日", "出生日期", "BIRTHDAY"))
                phone_col  = _find_col(hmap, ("電話",))
                mobile_col = _find_col(hmap, ("手機", "手機號碼"))
                addr_col   = _find_col(hmap, ("地址", "住址"))
                field_cols = {fn: _find_col(hmap, aliases)
                              for fn, aliases in MEMBER_FIELDS.items()}
                for row_no, vals in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    pid = normalize_id(vals[id_col])
                    if not _TW_ID_RE.fullmatch(pid):
                        continue
                    kwargs = {fn: (normalize_text(vals[col]) if col is not None else "")
                              for fn, col in field_cols.items()}
                    trace = _trace(config, batch_id, source_dir, path, ws.title, row_no, vals)
                    bundle.members.append(MemberRecord(
                        trace=trace, person_id=pid,
                        name=normalize_name(vals[name_col]) if name_col is not None else "",
                        birth_date=parse_date(vals[birth_col]) if birth_col is not None else None,
                        phone=normalize_phone(vals[phone_col]) if phone_col is not None else "",
                        mobile=normalize_phone(vals[mobile_col]) if mobile_col is not None else "",
                        address=normalize_text(vals[addr_col]) if addr_col is not None else "",
                        **kwargs,
                    ))
                    bundle.member_selections.append(MemberSelectionRecord(
                        trace=trace, person_id=pid, selection_type="designated_114"))
                    parsed += 1
        finally:
            wb.close()
        return parsed

    # ── 次數 XLSX（姓名＋生日比對，無身分證） ─────────────────────────────────
    def _parse_claims_xlsx(
        self, source_dir, path, config, batch_id,
        name_birth_idx: Dict[Tuple[str, str], set], bundle
    ) -> Tuple[int, int]:
        m = _MONTH_RE.search(path.stem)
        if not m:
            return 0, 0
        roc_year, month = int(m.group(1)), int(m.group(2))

        wb = load_workbook(path, read_only=True, data_only=True)
        parsed = unmatched = 0
        try:
            for ws in wb.worksheets:
                header_row = _find_header_row(
                    ws, (("姓名", "患者姓名"), ("次數", "來診次數", "就診次數")))
                if header_row is None:
                    continue
                hvals = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
                hmap       = _header_map(hvals)
                name_col   = _find_col(hmap, ("姓名", "患者姓名"))
                birth_col  = _find_col(hmap, ("生日", "出生日期"))
                count_col  = _find_col(hmap, ("次數", "來診次數", "就診次數"))
                if name_col is None or count_col is None:
                    continue
                for row_no, vals in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    name = normalize_name(vals[name_col])
                    if not name:
                        continue
                    birth_str = ""
                    if birth_col is not None and birth_col < len(vals):
                        bd = parse_date(vals[birth_col])
                        birth_str = bd.isoformat() if bd else ""
                    pids = name_birth_idx.get((name, birth_str), set())
                    if len(pids) != 1:
                        unmatched += 1
                        continue
                    pid = next(iter(pids))
                    bundle.monthly_claims.append(MonthlyClaimRecord(
                        trace=_trace(config, batch_id, source_dir, path, ws.title, row_no, vals),
                        person_id=pid,
                        roc_year=roc_year, month=month,
                        visit_count=parse_decimal(vals[count_col]),
                        amount=parse_decimal(0),
                    ))
                    parsed += 1
        finally:
            wb.close()
        return parsed, unmatched

    # ── 自選 / 不要（優先 ID，退而求其次姓名＋生日） ───────────────────────────
    def _parse_flag_workbook(
        self, source_dir, path, config, batch_id, flag_type,
        existing_ids, name_birth_idx, bundle
    ) -> Tuple[int, int]:
        wb = load_workbook(path, read_only=True, data_only=True)
        matched = unmatched = 0
        try:
            for ws in wb.worksheets:
                # Try ID-based first
                header_row = _find_header_row(ws, (ID_HEADERS,))
                if header_row is None:
                    header_row = _find_header_row(ws, (("姓名", "患者姓名"),))
                if header_row is None:
                    continue
                hvals  = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
                hmap   = _header_map(hvals)
                id_col    = _find_col(hmap, ID_HEADERS)
                name_col  = _find_col(hmap, ("姓名", "患者姓名"))
                birth_col = _find_col(hmap, ("生日", "出生日期"))
                for row_no, vals in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    pid = normalize_id(vals[id_col]) if id_col is not None else ""
                    if not _TW_ID_RE.fullmatch(pid):
                        # Fall back to name+birth
                        name = normalize_name(vals[name_col]) if name_col is not None else ""
                        if not name:
                            continue
                        birth_str = ""
                        if birth_col is not None and birth_col < len(vals):
                            bd = parse_date(vals[birth_col])
                            birth_str = bd.isoformat() if bd else ""
                        pids = name_birth_idx.get((name, birth_str), set())
                        if len(pids) != 1:
                            unmatched += 1
                            continue
                        pid = next(iter(pids))

                    bundle.member_selections.append(MemberSelectionRecord(
                        trace=_trace(config, batch_id, source_dir, path, ws.title, row_no, vals),
                        person_id=pid,
                        selection_type=flag_type,
                    ))
                    if pid in existing_ids:
                        matched += 1
                    else:
                        unmatched += 1
        finally:
            wb.close()
        return matched, unmatched

    @staticmethod
    def _build_name_birth_index(members: List[MemberRecord]) -> Dict[Tuple[str, str], set]:
        idx: Dict[Tuple[str, str], set] = defaultdict(set)
        for m in members:
            name = normalize_name(m.name)
            birth = m.birth_date.isoformat() if m.birth_date else ""
            if name:
                idx[(name, birth)].add(m.person_id)
        return dict(idx)

    @staticmethod
    def _build_name_birth_from_sm(
        sm_files: List[Path], existing_ids: set
    ) -> Dict[Tuple[str, str], set]:
        """sm_ 系列檔都有 家醫收案會員ID / 姓名 / 生日，用來建立 name+birth→ID 索引。"""
        idx: Dict[Tuple[str, str], set] = defaultdict(set)
        id_aliases   = ("家醫收案會員ID", "ID", "身分證號", "身份證號")
        name_aliases = ("姓名",)
        birth_aliases= ("生日", "出生日期")
        for path in sm_files:
            try:
                wb = load_workbook(path, read_only=True, data_only=True)
                try:
                    for ws in wb.worksheets:
                        hvals = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
                        hmap  = _header_map(hvals)
                        id_col    = _find_col(hmap, id_aliases)
                        name_col  = _find_col(hmap, name_aliases)
                        birth_col = _find_col(hmap, birth_aliases)
                        if id_col is None or name_col is None:
                            continue
                        for vals in ws.iter_rows(min_row=2, values_only=True):
                            pid = normalize_id(vals[id_col])
                            if not _TW_ID_RE.fullmatch(pid) or pid not in existing_ids:
                                continue
                            name = normalize_name(vals[name_col])
                            if not name:
                                continue
                            birth_str = ""
                            if birth_col is not None and birth_col < len(vals):
                                bd = parse_date(vals[birth_col])
                                birth_str = bd.isoformat() if bd else ""
                            idx[(name, birth_str)].add(pid)
                finally:
                    wb.close()
            except Exception:
                continue
        return dict(idx)
