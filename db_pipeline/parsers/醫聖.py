# -*- coding: utf-8 -*-
"""
醫聖系統解析器（source_system = "medical_saint"）

資料夾特徵：
- R11440/ 子資料夾：逐月就診明細 BIG5 固定寬度 TXT（P11401.txt 等）
- 照護名單 / 指定名單 xlsx
- 自選會員 A115 / 不要自選 A115X xlsx
- 各類篩檢獨立 xlsx（成健/子抹/糞便/老流/BC肝）

TXT 格式：
- 編碼：CP950（BIG5），以逗號分隔（無引號保護）
- 主訴欄可能含逗號 → 解析時需容錯（_extract_row_fields）
- 每列 = 一次就診 → 按 ID + 月份聚合成 MonthlyClaimRecord
"""
from __future__ import annotations

import csv
import dataclasses
import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

from db_pipeline.config.models import ClinicConfig
from db_pipeline.datasets.models import (
    DatasetBundle,
    LabResultRecord,
    MemberRecord,
    MemberSelectionRecord,
    MonthlyClaimRecord,
    P4PCaseRecord,
    P4PTrackRecord,
    ScreeningRecord,
    SourceTrace,
)
from db_pipeline.normalization import (
    normalize_id,
    normalize_name,
    normalize_phone,
    normalize_text,
    parse_date,
    parse_decimal,
    stable_row_hash,
)
from db_pipeline.parsers.解析器介面 import ParseCoverage, ParseResult
from db_pipeline.validation.models import ValidationIssue

# ── 常數 ──────────────────────────────────────────────────────────────────────
CP950 = "cp950"
ID_HEADERS = ("身分證號", "身份證號", "身分證號碼", "身份證號碼", "ID")
MEMBER_FIELDS = {
    "case_category":            ("個案類別",),
    "quality_roster":           ("論質名單",),
    "multi_chronic_65":         ("65歲以上多重慢性病註記",),
    "high_visit":               ("高診次註記",),
    "chronic_mark":             ("慢性病註記",),
    "non_chronic_mark":         ("非慢性病註記",),
    "same_clinic_previous_year":("與前一年家醫收案診所相同",),
    "disease_pattern":          ("疾病樣態",),
    "ascvd":                    ("ASCVD",),
    "three_highs":              ("三高",),
    "hypertension":             ("高血壓",),
    "hyperlipidemia":           ("高血脂",),
    "hyperglycemia":            ("高血糖",),
}
MONTH_CODE_RE = re.compile(r"(?<!\d)(1(?:14|15)(?:0[1-9]|1[0-2]))(?!\d)")
TW_ID_RE      = re.compile(r"[A-Z][12]\d{8}")
DATE_RE       = re.compile(r"\d{2,3}[./-]\d{1,2}[./-]\d{1,2}|\d{4}[./-]\d{1,2}[./-]\d{1,2}")
AGE_RE        = re.compile(r"^\d{2,3}歲\d{1,2}月\d{1,2}天$")
MAX_AMOUNT    = 99_999
XLSX_SUFFIXES = {".xlsx", ".xlsm"}
P4P_ID_HEADERS = ("家醫收案會員ID",) + ID_HEADERS
SCREENING_FILENAME_MAP = {
    "成人": "成人健檢",
    "成健": "成人健檢",
    "子宮頸": "子宮抹片",
    "子宮抹片": "子宮抹片",
    "65歲以上老感": "老人流感",
    "老人流感": "老人流感",
    "老感": "老人流感",
    "老流": "老人流感",
    "糞便": "糞便潛血",
    "潛血": "糞便潛血",
    "b、c肝": "肝炎篩檢",
    "bc肝": "肝炎篩檢",
    "肝炎": "肝炎篩檢",
}


# ── 工具函式 ──────────────────────────────────────────────────────────────────
def _header_map(values: Sequence[Any]) -> Dict[str, int]:
    return {normalize_text(v): i for i, v in enumerate(values) if normalize_text(v)}


def _find_col(headers: Dict[str, int], aliases: Iterable[str]) -> Optional[int]:
    for a in aliases:
        if a in headers:
            return headers[a]
    return None


def _find_header_row(ws: Any, required: Sequence[Sequence[str]]) -> Optional[int]:
    for row_no in range(1, min(ws.max_row, 12) + 1):
        vals = [ws.cell(row_no, c).value for c in range(1, ws.max_column + 1)]
        hmap = _header_map(vals)
        if all(_find_col(hmap, grp) is not None for grp in required):
            return row_no
    return None


def _month_code(text: str) -> Optional[str]:
    m = MONTH_CODE_RE.search(text)
    return m.group(1) if m else None


def _screening_type(path: Path) -> Optional[str]:
    name = path.stem.lower()
    for token, screening_type in SCREENING_FILENAME_MAP.items():
        if token in name:
            return screening_type
    return None


def _lab_test_code(header: Any) -> Optional[str]:
    text = normalize_text(header).lower()
    if "hba1c" in text:
        return "HbA1c"
    if "ldl" in text:
        return "LDL"
    if "uacr" in text:
        return "UACR"
    return None


def _normalize_hdr(text: str) -> str:
    return re.sub(r"\s+", "", str(text).strip())


def _is_date(text: str) -> bool:
    return bool(DATE_RE.fullmatch(text.strip()))


def _is_tw_id(text: str) -> bool:
    return bool(TW_ID_RE.fullmatch(text.strip().upper()))


def _is_name(text: str) -> bool:
    s = text.strip()
    if not s or _is_tw_id(s) or _is_date(s) or re.fullmatch(r"\d+", s):
        return False
    if AGE_RE.fullmatch(s):
        return False
    return True


def _clean_amount(text: str) -> str:
    s = str(text).strip().replace(",", "")
    if not re.fullmatch(r"\d+(?:\.0+)?", s):
        return ""
    v = int(float(s))
    return str(v) if 0 <= v <= MAX_AMOUNT else ""


def _shifted_amount(row: List[str]) -> str:
    """主訴含逗號時，欄位右移；從列尾回推固定位置取金額。"""
    if len(row) >= 22:
        v = _clean_amount(row[-7])
        if v:
            return v
    for i, cell in enumerate(row):
        if AGE_RE.fullmatch(str(cell).strip()) and i + 4 < len(row):
            v = _clean_amount(row[i + 4])
            if v:
                return v
    return ""


def _extract_fields(row: List[str], mapping: Dict[str, int]) -> Tuple[str, str, str, str, str]:
    """從一列 TXT 資料萃取 (pid, name, birth, date, amount)。"""
    def pick(*keys: str) -> str:
        for k in keys:
            idx = mapping.get(k)
            if idx is not None and idx < len(row):
                return row[idx].strip()
        return ""

    pid    = pick("身分證", "身分證號", "身份證號", "ID")
    name   = pick("姓名", "病患姓名", "會員姓名")
    dt     = pick("看診日", "日期", "就醫日", "最後就診日")
    bday   = pick("生日", "出生日期")
    amount = pick("申請額", "申請金額", "申請額小計")
    amount = _clean_amount(amount) or _shifted_amount(row)

    if _is_tw_id(pid) and _is_name(name) and _is_date(dt):
        return pid.upper(), name, bday, dt, amount

    # 容錯推斷
    id_idx    = next((i for i, v in enumerate(row) if _is_tw_id(str(v))), None)
    date_idxs = [i for i, v in enumerate(row) if _is_date(str(v))]

    if id_idx is not None:
        pid = str(row[id_idx]).strip().upper()
    if (not dt or not _is_date(dt)) and date_idxs:
        dt = str(row[date_idxs[0]]).strip()
    if (not bday or not _is_date(bday) or bday == dt) and len(date_idxs) >= 2:
        bday = str(row[date_idxs[1]]).strip()
    if not _is_name(name):
        dt_i = next((i for i, v in enumerate(row) if str(v).strip() == dt), None)
        if dt_i and dt_i > 0 and _is_name(str(row[dt_i - 1])):
            name = str(row[dt_i - 1]).strip()
        elif id_idx is not None:
            for cand in (id_idx - 1, id_idx - 2, id_idx - 3, id_idx - 4):
                if 0 <= cand < len(row) and _is_name(str(row[cand])):
                    name = str(row[cand]).strip()
                    break
    if not amount:
        amount = _shifted_amount(row)

    return pid, name, bday, dt, amount


def _parse_txt(path: Path) -> List[Tuple[str, str, str, str]]:
    """回傳 [(pid, bday, visit_date, amount), ...]，每列 = 一次就診。"""
    with path.open("r", encoding=CP950, errors="replace", newline="") as f:
        raw = [line.rstrip("\r\n") for line in f if line.strip()]

    rows: List[List[str]] = []
    for line in raw:
        if "費用年月:" in line:
            continue
        parts = [p.strip() for p in line.split(",")]
        while parts and parts[-1] == "":
            parts.pop()
        if parts:
            rows.append(parts)

    if len(rows) < 2:
        return []

    # 找表頭
    hdr_idx = 0
    for i, row in enumerate(rows[:5]):
        norm = {_normalize_hdr(v) for v in row}
        if any(t in norm for t in ("身分證", "身分證號", "身份證號", "姓名", "看診日", "病歷號")):
            hdr_idx = i
            break

    mapping = {_normalize_hdr(v): idx for idx, v in enumerate(rows[hdr_idx])}
    result: List[Tuple[str, str, str, str]] = []
    for row in rows[hdr_idx + 1:]:
        pid, _name, bday, vdate, amount = _extract_fields(row, mapping)
        if pid and vdate:
            result.append((pid.upper(), bday, vdate, amount))
    return result


def _trace(
    config: ClinicConfig, batch_id: str, source_dir: Path,
    file_path: Path, sheet_name: str, row_no: int, values: Sequence[Any],
) -> SourceTrace:
    return SourceTrace(
        clinic_code=config.clinic_code, batch_id=batch_id,
        source_system=config.source_system,
        source_file=str(file_path.relative_to(source_dir)),
        source_sheet=sheet_name, source_row=row_no,
        raw_row_hash=stable_row_hash(values),
    )


# ── 主解析器 ──────────────────────────────────────────────────────────────────
class MedicalSaintParser:
    source_system = "medical_saint"

    def parse(
        self,
        source_dir: Path,
        config: ClinicConfig,
        batch_id: str,
    ) -> ParseResult:
        bundle   = DatasetBundle()
        coverage = ParseCoverage()
        issues: List[ValidationIssue] = []

        files = sorted(
            p for p in source_dir.rglob("*")
            if p.is_file() and not p.name.startswith(("~$", "."))
        )
        coverage.discovered_files = len(files)

        # 1. 照護名單
        roster_files = [
            p for p in files
            if p.suffix.lower() in XLSX_SUFFIXES
            and any(kw in p.name for kw in ("照護名單", "指定名單", "家醫名單", "指定會員"))
        ]
        for path in roster_files:
            n = self._parse_member_workbook(source_dir, path, config, batch_id, bundle)
            if n:
                coverage.parsed_files += 1
                coverage.parsed_rows["members"] = coverage.parsed_rows.get("members", 0) + n

        existing_ids = {m.person_id for m in bundle.members}

        # 2. BIG5 TXT 月份檔
        txt_files = [
            p for p in files
            if p.suffix.lower() == ".txt"
            and _month_code(p.stem) is not None
        ]
        for path in txt_files:
            parsed, unmatched = self._parse_txt_claims(
                source_dir, path, config, batch_id, existing_ids, bundle)
            if parsed or unmatched:
                coverage.parsed_files += 1
                coverage.parsed_rows["monthly_claims"] = (
                    coverage.parsed_rows.get("monthly_claims", 0) + parsed)
                coverage.unmatched_rows["monthly_claims"] = (
                    coverage.unmatched_rows.get("monthly_claims", 0) + unmatched)

        # 3. 自選 / 不要
        member_indexes = self._build_indexes(bundle.members)
        select_files = [
            p for p in files
            if p.suffix.lower() in XLSX_SUFFIXES
            and any(kw in p.name for kw in ("自選", "A115", "115自選"))
            and not any(kw in p.name.upper() for kw in ("115X", "不要"))
        ]
        exclude_files = [
            p for p in files
            if p.suffix.lower() in XLSX_SUFFIXES
            and any(kw in p.name.upper() for kw in ("115X", "不要"))
        ]
        p4p_case_files = [
            p for p in files
            if p.suffix.lower() in XLSX_SUFFIXES
            and re.search(r"P4P.*收案|P4p.*收案", p.name)
        ]
        p4p_track_files = [
            p for p in files
            if p.suffix.lower() in XLSX_SUFFIXES
            and re.search(r"P4P.*追蹤|P4p.*追蹤", p.name)
        ]
        health_files = [
            p for p in files
            if p.suffix.lower() in XLSX_SUFFIXES and "個案健康管理" in p.name
        ]
        screening_files = [
            p for p in files
            if p.suffix.lower() in XLSX_SUFFIXES and _screening_type(p) is not None
        ]
        for path in select_files:
            n, u = self._parse_selection_workbook(
                source_dir, path, config, batch_id, member_indexes, bundle, "self_selected_115")
            if n or u:
                coverage.parsed_files += 1
        for path in exclude_files:
            n, u = self._parse_selection_workbook(
                source_dir, path, config, batch_id, member_indexes, bundle, "excluded_115x")
            if n or u:
                coverage.parsed_files += 1

        # 4. P4P、篩檢、個案健康管理
        for path in p4p_case_files:
            n = self._parse_p4p_cases(source_dir, path, config, batch_id, bundle)
            coverage.parsed_files += 1
            coverage.parsed_rows["p4p_cases"] = coverage.parsed_rows.get("p4p_cases", 0) + n
        for path in p4p_track_files:
            n = self._parse_p4p_tracks(source_dir, path, config, batch_id, bundle)
            coverage.parsed_files += 1
            coverage.parsed_rows["p4p_tracks"] = coverage.parsed_rows.get("p4p_tracks", 0) + n
        for path in screening_files:
            n = self._parse_screenings(
                source_dir, path, config, batch_id, _screening_type(path) or "", bundle)
            coverage.parsed_files += 1
            coverage.parsed_rows["screenings"] = coverage.parsed_rows.get("screenings", 0) + n
        for path in health_files:
            n = self._parse_health_management(source_dir, path, config, batch_id, bundle)
            coverage.parsed_files += 1
            coverage.parsed_rows["lab_results"] = coverage.parsed_rows.get("lab_results", 0) + n

        # 5. 跳過
        parsed_paths = set(
            roster_files + txt_files + select_files + exclude_files
            + p4p_case_files + p4p_track_files + screening_files + health_files
        )
        for p in files:
            if p not in parsed_paths:
                coverage.skipped_files[str(p.relative_to(source_dir))] = (
                    "medical_saint v1 尚未實作此來源類型")

        if txt_files and not bundle.monthly_claims:
            issues.append(ValidationIssue(
                severity="error", dataset="monthly_claims",
                code="claim_not_mapped",
                message=(
                    "費用次數 TXT 月份檔已找到，但 0 筆可對應照護名單。"
                    "請確認照護名單與費用檔是否屬同一診所同一期別。"
                )))
        for dataset, cnt in coverage.unmatched_rows.items():
            if cnt:
                issues.append(ValidationIssue(
                    severity="warning", dataset=dataset,
                    code="unmatched_source_rows",
                    message=(
                        f"費用次數 TXT 月份檔有 {cnt} 筆身分證號不在照護名單"
                        "（一般門診病患，非家醫計畫會員），已以真實身分證號寫入。"
                    )))

        return ParseResult(bundle=bundle, coverage=coverage, issues=issues)

    # ── 照護名單 ──────────────────────────────────────────────────────────────
    def _parse_member_workbook(
        self,
        source_dir: Path, path: Path,
        config: ClinicConfig, batch_id: str,
        bundle: DatasetBundle,
    ) -> int:
        wb = load_workbook(path, read_only=True, data_only=True)
        parsed = 0
        # pid → index in bundle.members (for cross-sheet field merge)
        seen_pids: dict = {m.person_id: i for i, m in enumerate(bundle.members)}
        try:
            for ws in wb.worksheets:
                header_row = _find_header_row(ws, (ID_HEADERS, ("個案類別",)))
                if header_row is None:
                    continue
                hvals = [ws.cell(header_row, c).value
                         for c in range(1, ws.max_column + 1)]
                hmap   = _header_map(hvals)
                id_col = _find_col(hmap, ID_HEADERS)
                if id_col is None:
                    continue
                name_col   = _find_col(hmap, ("姓名",))
                birth_col  = _find_col(hmap, ("生日", "出生日期", "BIRTHDAY"))
                phone_col  = _find_col(hmap, ("電話",))
                mobile_col = _find_col(hmap, ("手機", "手機號碼"))
                addr_col   = _find_col(hmap, ("地址", "住址"))
                field_cols = {
                    fn: _find_col(hmap, aliases)
                    for fn, aliases in MEMBER_FIELDS.items()
                }
                for row_no, vals in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    pid = normalize_id(vals[id_col])
                    if not TW_ID_RE.fullmatch(pid):
                        continue
                    kwargs = {
                        fn: normalize_text(vals[col]) if col is not None else ""
                        for fn, col in field_cols.items()
                    }
                    trace = _trace(config, batch_id, source_dir, path, ws.title, row_no, vals)
                    new_rec = MemberRecord(
                        trace=trace, person_id=pid,
                        name=normalize_name(vals[name_col]) if name_col is not None else "",
                        birth_date=parse_date(vals[birth_col]) if birth_col is not None else None,
                        phone=normalize_phone(vals[phone_col]) if phone_col is not None else "",
                        mobile=normalize_phone(vals[mobile_col]) if mobile_col is not None else "",
                        address=normalize_text(vals[addr_col]) if addr_col is not None else "",
                        **kwargs,
                    )
                    if pid in seen_pids:
                        # Merge non-empty fields; always record every visited sheet in trace
                        idx = seen_pids[pid]
                        existing = bundle.members[idx]
                        updates = {
                            f.name: getattr(new_rec, f.name)
                            for f in dataclasses.fields(existing)
                            if f.name != "trace"
                            and (getattr(existing, f.name) is None or getattr(existing, f.name) == "")
                            and getattr(new_rec, f.name) not in (None, "")
                        }
                        existing_sheets = existing.trace.source_sheet
                        if ws.title not in existing_sheets:
                            updates["trace"] = dataclasses.replace(
                                existing.trace,
                                source_sheet=f"{existing_sheets}|{ws.title}",
                            )
                        if updates:
                            bundle.members[idx] = dataclasses.replace(existing, **updates)
                    else:
                        seen_pids[pid] = len(bundle.members)
                        bundle.members.append(new_rec)
                        bundle.member_selections.append(MemberSelectionRecord(
                            trace=trace, person_id=pid, selection_type="designated_114"))
                        parsed += 1
        finally:
            wb.close()
        return parsed

    # ── BIG5 TXT 月份就診（聚合） ─────────────────────────────────────────────
    def _parse_txt_claims(
        self,
        source_dir: Path, path: Path,
        config: ClinicConfig, batch_id: str,
        existing_ids: set,
        bundle: DatasetBundle,
    ) -> Tuple[int, int]:
        code = _month_code(path.stem) or _month_code(path.name)
        if code is None:
            return 0, 0
        roc_year = int(code[:3])
        month    = int(code[3:])

        visits = _parse_txt(path)
        if not visits:
            return 0, 0

        # 聚合
        agg: Dict[str, dict] = {}
        for pid, bday, vdate, amount in visits:
            if not TW_ID_RE.fullmatch(pid):
                continue
            rec = agg.setdefault(pid, {
                "count": 0, "amount": parse_decimal(0),
                "last_visit": None, "vdate_str": vdate, "bday": bday,
            })
            rec["count"] += 1
            rec["amount"] += parse_decimal(amount)
            dt = parse_date(vdate)
            if dt and (rec["last_visit"] is None or dt > rec["last_visit"]):
                rec["last_visit"] = dt

        parsed = unmatched = 0
        for pid, rec in agg.items():
            if pid not in existing_ids:
                unmatched += 1
            bundle.monthly_claims.append(MonthlyClaimRecord(
                trace=SourceTrace(
                    clinic_code=config.clinic_code,
                    batch_id=batch_id,
                    source_system=config.source_system,
                    source_file=str(path.relative_to(source_dir)),
                    source_sheet="",
                    source_row=0,
                    raw_row_hash=stable_row_hash([pid, code]),
                ),
                person_id=pid,
                roc_year=roc_year,
                month=month,
                visit_count=parse_decimal(rec["count"]),
                amount=rec["amount"],
                last_visit_date=rec["last_visit"],
            ))
            parsed += 1
        return parsed, unmatched

    # ── 自選 / 不要 ──────────────────────────────────────────────────────────
    def _parse_selection_workbook(
        self,
        source_dir: Path, path: Path,
        config: ClinicConfig, batch_id: str,
        indexes: Dict[str, Any],
        bundle: DatasetBundle,
        selection_type: str,
    ) -> Tuple[int, int]:
        wb = load_workbook(path, read_only=True, data_only=True)
        parsed = unmatched = 0
        existing_ids = {m.person_id for m in bundle.members}
        try:
            for ws in wb.worksheets:
                header_row = _find_header_row(ws, (ID_HEADERS,))
                if header_row is None:
                    continue
                hvals = [ws.cell(header_row, c).value
                         for c in range(1, ws.max_column + 1)]
                hmap   = _header_map(hvals)
                id_col = _find_col(hmap, ID_HEADERS)
                if id_col is None:
                    continue
                for row_no, vals in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    pid = normalize_id(vals[id_col])
                    if not TW_ID_RE.fullmatch(pid):
                        continue
                    if pid not in existing_ids:
                        unmatched += 1
                        continue
                    bundle.member_selections.append(MemberSelectionRecord(
                        trace=_trace(config, batch_id, source_dir, path,
                                     ws.title, row_no, vals),
                        person_id=pid,
                        selection_type=selection_type,
                    ))
                    parsed += 1
        finally:
            wb.close()
        return parsed, unmatched

    # ── P4P 收案 / 追蹤 ─────────────────────────────────────────────────────
    def _parse_p4p_cases(
        self,
        source_dir: Path, path: Path,
        config: ClinicConfig, batch_id: str,
        bundle: DatasetBundle,
    ) -> int:
        wb = load_workbook(path, read_only=True, data_only=True)
        parsed = 0
        try:
            for ws in wb.worksheets:
                header_row = _find_header_row(
                    ws, (P4P_ID_HEADERS, ("P4P收案計畫", "收案計畫")))
                if header_row is None:
                    continue
                values = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
                headers = _header_map(values)
                id_col = _find_col(headers, P4P_ID_HEADERS)
                plan_col = _find_col(headers, ("P4P收案計畫", "收案計畫"))
                status_col = _find_col(headers, ("收案狀態", "狀態"))
                enroll_col = _find_col(headers, ("收案日期",))
                if id_col is None:
                    continue
                for row_no, row in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    pid = normalize_id(row[id_col])
                    if not TW_ID_RE.fullmatch(pid):
                        continue
                    bundle.p4p_cases.append(P4PCaseRecord(
                        trace=_trace(config, batch_id, source_dir, path, ws.title, row_no, row),
                        person_id=pid,
                        plan=normalize_text(row[plan_col]) if plan_col is not None else "",
                        status=normalize_text(row[status_col]) if status_col is not None else "",
                        enrolled_at=parse_date(row[enroll_col]) if enroll_col is not None else None,
                    ))
                    parsed += 1
        finally:
            wb.close()
        return parsed

    def _parse_p4p_tracks(
        self,
        source_dir: Path, path: Path,
        config: ClinicConfig, batch_id: str,
        bundle: DatasetBundle,
    ) -> int:
        wb = load_workbook(path, read_only=True, data_only=True)
        parsed = 0
        try:
            for ws in wb.worksheets:
                header_row = _find_header_row(ws, (P4P_ID_HEADERS, ("最後追蹤日",)))
                if header_row is None:
                    continue
                values = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
                headers = _header_map(values)
                id_col = _find_col(headers, P4P_ID_HEADERS)
                plan_col = _find_col(headers, ("P4P收案計畫", "收案計畫"))
                last_col = _find_col(headers, ("最後追蹤日", "最近追蹤日"))
                next_col = _find_col(headers, ("下次應追蹤日", "下次追蹤日"))
                overdue_col = _find_col(headers, ("逾期未追蹤",))
                if id_col is None:
                    continue
                for row_no, row in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    pid = normalize_id(row[id_col])
                    if not TW_ID_RE.fullmatch(pid):
                        continue
                    bundle.p4p_tracks.append(P4PTrackRecord(
                        trace=_trace(config, batch_id, source_dir, path, ws.title, row_no, row),
                        person_id=pid,
                        plan=normalize_text(row[plan_col]) if plan_col is not None else "",
                        last_tracked_at=parse_date(row[last_col]) if last_col is not None else None,
                        next_track_at=parse_date(row[next_col]) if next_col is not None else None,
                        overdue=normalize_text(row[overdue_col]) if overdue_col is not None else "",
                    ))
                    parsed += 1
        finally:
            wb.close()
        return parsed

    # ── 篩檢 / 個案健康管理 ─────────────────────────────────────────────────
    def _parse_screenings(
        self,
        source_dir: Path, path: Path,
        config: ClinicConfig, batch_id: str,
        screening_type: str,
        bundle: DatasetBundle,
    ) -> int:
        wb = load_workbook(path, read_only=True, data_only=True)
        parsed = 0
        try:
            for ws in wb.worksheets:
                header_row = _find_header_row(
                    ws, (ID_HEADERS, ("最後篩檢日期", "篩檢日期", "最後檢查日期")))
                if header_row is None:
                    continue
                values = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
                headers = _header_map(values)
                id_col = _find_col(headers, ID_HEADERS)
                date_col = _find_col(
                    headers, ("最後篩檢日期", "篩檢日期", "最後檢查日期"))
                if id_col is None or date_col is None:
                    continue
                for row_no, row in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    pid = normalize_id(row[id_col])
                    screened_at = parse_date(row[date_col])
                    if not TW_ID_RE.fullmatch(pid) or screened_at is None:
                        continue
                    bundle.screenings.append(ScreeningRecord(
                        trace=_trace(config, batch_id, source_dir, path, ws.title, row_no, row),
                        person_id=pid,
                        screening_type=screening_type,
                        screened_at=screened_at,
                    ))
                    parsed += 1
        finally:
            wb.close()
        return parsed

    def _parse_health_management(
        self,
        source_dir: Path, path: Path,
        config: ClinicConfig, batch_id: str,
        bundle: DatasetBundle,
    ) -> int:
        wb = load_workbook(path, read_only=True, data_only=True)
        parsed = 0
        try:
            for ws in wb.worksheets:
                header_row = _find_header_row(ws, (P4P_ID_HEADERS,))
                if header_row is None:
                    continue
                header_values = [
                    ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)
                ]
                headers = _header_map(header_values)
                id_col = _find_col(headers, P4P_ID_HEADERS)
                if id_col is None:
                    continue
                pairs: List[Tuple[str, int, int]] = []
                for value_col, header in enumerate(header_values):
                    test_code = _lab_test_code(header)
                    if test_code is None or "結果" not in normalize_text(header):
                        continue
                    date_col = next((
                        idx for idx in range(value_col + 1, min(value_col + 4, len(header_values)))
                        if "日期" in normalize_text(header_values[idx])
                    ), None)
                    if date_col is not None:
                        pairs.append((test_code, value_col, date_col))
                for row_no, row in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    pid = normalize_id(row[id_col])
                    if not TW_ID_RE.fullmatch(pid):
                        continue
                    trace = _trace(config, batch_id, source_dir, path, ws.title, row_no, row)
                    for test_code, value_col, date_col in pairs:
                        result_value = normalize_text(row[value_col])
                        tested_at = parse_date(row[date_col])
                        if not result_value and tested_at is None:
                            continue
                        bundle.lab_results.append(LabResultRecord(
                            trace=trace,
                            person_id=pid,
                            test_code=test_code,
                            result_value=result_value,
                            tested_at=tested_at,
                        ))
                        parsed += 1
        finally:
            wb.close()
        return parsed

    @staticmethod
    def _build_indexes(members: Sequence[MemberRecord]) -> Dict[str, Dict]:
        idxs: Dict[str, Dict] = {"name_phone": defaultdict(set)}
        for m in members:
            name = normalize_name(m.name)
            for phone in (m.phone, m.mobile):
                if phone:
                    idxs["name_phone"][(name, phone[-7:])].add(m.person_id)
        return idxs
