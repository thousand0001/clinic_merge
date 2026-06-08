from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

from db_pipeline.config.models import ClinicConfig
from db_pipeline.datasets.models import (
    DatasetBundle,
    MemberRecord,
    MemberSelectionRecord,
    MonthlyClaimRecord,
    SourceTrace,
)
from db_pipeline.normalization import (
    normalize_id,
    normalize_name,
    normalize_phone,
    normalize_text,
    parse_date,
    parse_decimal,
    stable_row_hash,
)
from db_pipeline.parsers.解析器介面 import ParseCoverage, ParseResult
from db_pipeline.validation.models import ValidationIssue


ID_HEADERS = ("ID", "身分證號", "身份證號", "身分證號碼", "身份證號碼")
MEMBER_FIELDS = {
    "case_category": ("個案類別",),
    "quality_roster": ("論質名單",),
    "multi_chronic_65": ("65歲以上多重慢性病註記",),
    "high_visit": ("高診次註記",),
    "chronic_mark": ("慢性病註記",),
    "non_chronic_mark": ("非慢性病註記",),
    "same_clinic_previous_year": ("與前一年家醫收案診所相同",),
    "disease_pattern": ("疾病樣態",),
    "ascvd": ("ASCVD",),
    "three_highs": ("三高",),
    "hypertension": ("高血壓",),
    "hyperlipidemia": ("高血脂",),
    "hyperglycemia": ("高血糖",),
}


def _header_map(values: Sequence[Any]) -> Dict[str, int]:
    return {
        normalize_text(value): index
        for index, value in enumerate(values)
        if normalize_text(value)
    }


def _find_column(headers: Dict[str, int], aliases: Iterable[str]) -> Optional[int]:
    for alias in aliases:
        if alias in headers:
            return headers[alias]
    return None


def _find_header_row(ws: Any, required_groups: Sequence[Sequence[str]]) -> Optional[int]:
    for row_no in range(1, min(ws.max_row, 12) + 1):
        values = [ws.cell(row_no, col).value for col in range(1, ws.max_column + 1)]
        headers = _header_map(values)
        if all(_find_column(headers, aliases) is not None for aliases in required_groups):
            return row_no
    return None


def _infer_member_contact_columns(
    ws: Any,
    header_row: int,
    start_col: int,
    end_col: int,
) -> Dict[str, Optional[int]]:
    scores: Dict[str, Dict[int, int]] = {
        "name": {},
        "phone": {},
        "address": {},
    }
    address_tokens = "市縣區鄉鎮村里路街巷弄號樓"
    sample_end = min(ws.max_row, header_row + 30)

    for col in range(start_col, end_col):
        name_score = phone_score = address_score = 0
        for row_no in range(header_row + 1, sample_end + 1):
            text = normalize_text(ws.cell(row_no, col + 1).value)
            if not text:
                continue
            digits = re.sub(r"\D", "", text)
            has_cjk = bool(re.search(r"[\u4e00-\u9fff]", text))
            has_address_token = any(token in text for token in address_tokens)

            if 2 <= len(text) <= 10 and has_cjk and not has_address_token:
                if len(digits) < 3:
                    name_score += 1
            has_latin = bool(re.search(r"[A-Za-z]", text))
            if (
                8 <= len(digits) <= 10
                and not has_cjk
                and not has_latin
                and parse_date(text) is None
            ):
                phone_score += 1
            if has_address_token or len(text) > 12:
                address_score += 1

        scores["name"][col] = name_score
        scores["phone"][col] = phone_score
        scores["address"][col] = address_score

    selected: Dict[str, Optional[int]] = {
        "name": None,
        "phone": None,
        "address": None,
    }
    used = set()
    for field_name in ("phone", "address", "name"):
        candidates = sorted(
            scores[field_name].items(),
            key=lambda item: (-item[1], item[0]),
        )
        for col, score in candidates:
            if score > 0 and col not in used:
                selected[field_name] = col
                used.add(col)
                break
    return selected


def _trace(
    config: ClinicConfig,
    batch_id: str,
    source_dir: Path,
    file_path: Path,
    sheet_name: str,
    row_no: int,
    values: Sequence[Any],
) -> SourceTrace:
    return SourceTrace(
        clinic_code=config.clinic_code,
        batch_id=batch_id,
        source_system=config.source_system,
        source_file=str(file_path.relative_to(source_dir)),
        source_sheet=sheet_name,
        source_row=row_no,
        raw_row_hash=stable_row_hash(values),
    )


class SmParser:
    source_system = "sm"

    def parse(
        self,
        source_dir: Path,
        config: ClinicConfig,
        batch_id: str,
    ) -> ParseResult:
        bundle = DatasetBundle()
        coverage = ParseCoverage()
        issues: List[ValidationIssue] = []
        files = sorted(
            path
            for path in source_dir.rglob("*")
            if path.is_file() and not path.name.startswith(("~$", "."))
        )
        coverage.discovered_files = len(files)
        if not files:
            issues.append(
                ValidationIssue(
                    severity="error",
                    dataset="source",
                    code="empty_source_directory",
                    message="來源資料夾沒有可解析的檔案。",
                )
            )

        roster_files = [
            path
            for path in files
            if path.suffix.lower() in (".xlsx", ".xlsm")
            and ("照護名單" in path.name or "指定會員名單" in path.name)
        ]
        for path in roster_files:
            parsed = self._parse_member_workbook(
                source_dir,
                path,
                config,
                batch_id,
                bundle,
            )
            if parsed:
                coverage.parsed_files += 1
                coverage.parsed_rows["members"] = (
                    coverage.parsed_rows.get("members", 0) + parsed
                )

        member_indexes = self._build_member_indexes(bundle.members)
        claim_files = [
            path
            for path in files
            if path.suffix.lower() in (".xlsx", ".xlsm")
            and "R11440" in path.name.upper()
        ]
        for path in claim_files:
            parsed, unmatched = self._parse_monthly_claim_workbook(
                source_dir,
                path,
                config,
                batch_id,
                member_indexes,
                bundle,
            )
            if parsed or unmatched:
                coverage.parsed_files += 1
                coverage.parsed_rows["monthly_claims"] = (
                    coverage.parsed_rows.get("monthly_claims", 0) + parsed
                )
                coverage.unmatched_rows["monthly_claims"] = (
                    coverage.unmatched_rows.get("monthly_claims", 0) + unmatched
                )

        self_select_files = [
            path
            for path in files
            if path.suffix.lower() in (".xlsx", ".xlsm")
            and "自選" in path.name
            and "不要" not in path.name
            and "115X" not in path.name.upper()
        ]
        for path in self_select_files:
            parsed, unmatched = self._parse_self_selected_workbook(
                source_dir,
                path,
                config,
                batch_id,
                member_indexes,
                bundle,
            )
            if parsed or unmatched:
                coverage.parsed_files += 1
                coverage.parsed_rows["member_selections"] = (
                    coverage.parsed_rows.get("member_selections", 0) + parsed
                )
                coverage.unmatched_rows["member_selections"] = (
                    coverage.unmatched_rows.get("member_selections", 0) + unmatched
                )

        parsed_paths = set(roster_files + claim_files + self_select_files)
        for path in files:
            if path not in parsed_paths:
                coverage.skipped_files[str(path.relative_to(source_dir))] = (
                    "SM v1 尚未實作此來源類型"
                )

        if claim_files and not bundle.monthly_claims:
            issues.append(
                ValidationIssue(
                    severity="error",
                    dataset="monthly_claims",
                    code="claim_source_not_mapped",
                    message=(
                        "費用次數來源檔（R11440）已找到，但 0 筆可對應照護名單。"
                        "請確認照護名單與費用檔是否屬同一診所同一期別。"
                    ),
                )
            )
        for dataset, count in coverage.unmatched_rows.items():
            if count:
                issues.append(
                    ValidationIssue(
                        severity="warning",
                        dataset=dataset,
                        code="unmatched_source_rows",
                        message=(
                            f"費用次數檔有 {count} 筆姓名＋生日無法比對照護名單"
                            "（一般門診病患，非家醫計畫會員），已以病歷號作為暫代識別碼寫入，"
                            "身分證號欄位待後續補全。"
                        ),
                    )
                )

        return ParseResult(bundle=bundle, coverage=coverage, issues=issues)

    def _parse_member_workbook(
        self,
        source_dir: Path,
        path: Path,
        config: ClinicConfig,
        batch_id: str,
        bundle: DatasetBundle,
    ) -> int:
        wb = load_workbook(path, read_only=True, data_only=True)
        parsed = 0
        try:
            for ws in wb.worksheets:
                header_row = _find_header_row(ws, (ID_HEADERS, ("個案類別",)))
                if header_row is None:
                    continue
                header_values = [
                    ws.cell(header_row, col).value
                    for col in range(1, ws.max_column + 1)
                ]
                headers = _header_map(header_values)
                id_col = _find_column(headers, ID_HEADERS)
                if id_col is None:
                    continue
                name_col = _find_column(headers, ("姓名", "會員姓名"))
                birth_col = _find_column(headers, ("BIRTHDAY", "生日"))
                phone_col = _find_column(headers, ("電話", "聯絡電話"))
                mobile_col = _find_column(headers, ("手機", "手機號碼", "行動電話"))
                address_col = _find_column(headers, ("地址", "住址"))
                field_cols = {
                    field_name: _find_column(headers, aliases)
                    for field_name, aliases in MEMBER_FIELDS.items()
                }
                metadata_cols = [
                    col for col in field_cols.values() if col is not None
                ]
                first_metadata_col = (
                    min(metadata_cols) if metadata_cols else ws.max_column
                )
                if birth_col is not None:
                    inferred = _infer_member_contact_columns(
                        ws,
                        header_row,
                        birth_col + 1,
                        first_metadata_col,
                    )
                    if name_col is None:
                        name_col = inferred["name"]
                    if phone_col is None:
                        phone_col = inferred["phone"]
                    if address_col is None:
                        address_col = inferred["address"]

                for row_no, values in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    person_id = normalize_id(values[id_col])
                    if not re.fullmatch(r"[A-Z][12][0-9]{8}", person_id):
                        continue
                    kwargs = {
                        field_name: (
                            normalize_text(values[col]) if col is not None else ""
                        )
                        for field_name, col in field_cols.items()
                    }
                    bundle.members.append(
                        MemberRecord(
                            trace=_trace(
                                config,
                                batch_id,
                                source_dir,
                                path,
                                ws.title,
                                row_no,
                                values,
                            ),
                            person_id=person_id,
                            name=normalize_name(values[name_col]) if name_col is not None else "",
                            birth_date=parse_date(values[birth_col]) if birth_col is not None else None,
                            phone=normalize_phone(values[phone_col]) if phone_col is not None else "",
                            mobile=normalize_phone(values[mobile_col]) if mobile_col is not None else "",
                            address=normalize_text(values[address_col]) if address_col is not None else "",
                            **kwargs,
                        )
                    )
                    bundle.member_selections.append(
                        MemberSelectionRecord(
                            trace=bundle.members[-1].trace,
                            person_id=person_id,
                            selection_type="designated_114",
                        )
                    )
                    parsed += 1
        finally:
            wb.close()
        return parsed

    def _build_member_indexes(
        self,
        members: Sequence[MemberRecord],
    ) -> Dict[str, Dict[Tuple[str, str], set]]:
        indexes: Dict[str, Dict[Tuple[str, str], set]] = {
            "name_birth": defaultdict(set),
            "name_phone": defaultdict(set),
        }
        for member in members:
            name = normalize_name(member.name)
            if member.birth_date:
                indexes["name_birth"][(name, member.birth_date.isoformat())].add(
                    member.person_id
                )
            for phone in (member.phone, member.mobile):
                if phone:
                    indexes["name_phone"][(name, phone[-7:])].add(member.person_id)
        return indexes

    def _parse_monthly_claim_workbook(
        self,
        source_dir: Path,
        path: Path,
        config: ClinicConfig,
        batch_id: str,
        indexes: Dict[str, Dict[Tuple[str, str], set]],
        bundle: DatasetBundle,
    ) -> Tuple[int, int]:
        wb = load_workbook(path, read_only=True, data_only=True)
        parsed = unmatched = 0
        try:
            for ws in wb.worksheets:
                match = re.fullmatch(r"(114|115)(0[1-9]|1[0-2])", ws.title.strip())
                if not match:
                    continue
                header_row = _find_header_row(
                    ws,
                    (("姓名",), ("生日",), ("來診次數", "次數"), ("申報總金額", "總額")),
                )
                if header_row is None:
                    continue
                headers = _header_map(
                    [
                        ws.cell(header_row, col).value
                        for col in range(1, ws.max_column + 1)
                    ]
                )
                name_col = _find_column(headers, ("姓名",))
                birth_col = _find_column(headers, ("生日",))
                count_col = _find_column(headers, ("來診次數", "次數"))
                amount_col = _find_column(headers, ("申報總金額", "總額"))
                last_visit_col = _find_column(headers, ("最後回診日", "最後就診日"))
                chart_col = _find_column(headers, ("病歷號",))
                if None in (name_col, birth_col, count_col, amount_col):
                    continue

                for row_no, values in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    name = normalize_name(values[name_col])
                    birth_date = parse_date(values[birth_col])
                    if not name or birth_date is None:
                        continue
                    person_ids = indexes["name_birth"].get(
                        (name, birth_date.isoformat()),
                        set(),
                    )
                    if len(person_ids) == 1:
                        person_id = next(iter(person_ids))
                    else:
                        chart_raw = values[chart_col] if chart_col is not None else None
                        chart_no = str(int(chart_raw)) if chart_raw is not None else ""
                        person_id = f"chart:{chart_no}" if chart_no else ""
                        unmatched += 1
                    bundle.monthly_claims.append(
                        MonthlyClaimRecord(
                            trace=_trace(
                                config,
                                batch_id,
                                source_dir,
                                path,
                                ws.title,
                                row_no,
                                values,
                            ),
                            person_id=person_id,
                            roc_year=int(match.group(1)),
                            month=int(match.group(2)),
                            visit_count=parse_decimal(values[count_col]),
                            amount=parse_decimal(values[amount_col]),
                            last_visit_date=(
                                parse_date(values[last_visit_col])
                                if last_visit_col is not None
                                else None
                            ),
                        )
                    )
                    parsed += 1
        finally:
            wb.close()
        return parsed, unmatched

    def _parse_self_selected_workbook(
        self,
        source_dir: Path,
        path: Path,
        config: ClinicConfig,
        batch_id: str,
        indexes: Dict[str, Dict[Tuple[str, str], set]],
        bundle: DatasetBundle,
    ) -> Tuple[int, int]:
        wb = load_workbook(path, read_only=True, data_only=True)
        parsed = unmatched = 0
        try:
            for ws in wb.worksheets:
                header_row = _find_header_row(ws, (("姓名",), ("電話", "手機")))
                if header_row is None:
                    continue
                headers = _header_map(
                    [
                        ws.cell(header_row, col).value
                        for col in range(1, ws.max_column + 1)
                    ]
                )
                name_col = _find_column(headers, ("姓名",))
                phone_col = _find_column(headers, ("電話", "手機"))
                if name_col is None or phone_col is None:
                    continue
                for row_no, values in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    name = normalize_name(values[name_col])
                    phone = normalize_phone(values[phone_col])
                    if not name or len(phone) < 7:
                        continue
                    person_ids = indexes["name_phone"].get((name, phone[-7:]), set())
                    if len(person_ids) != 1:
                        unmatched += 1
                        continue
                    bundle.member_selections.append(
                        MemberSelectionRecord(
                            trace=_trace(
                                config,
                                batch_id,
                                source_dir,
                                path,
                                ws.title,
                                row_no,
                                values,
                            ),
                            person_id=next(iter(person_ids)),
                            selection_type="self_selected_115",
                        )
                    )
                    parsed += 1
        finally:
            wb.close()
        return parsed, unmatched
