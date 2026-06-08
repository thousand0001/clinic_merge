# -*- coding: utf-8 -*-
"""
宏誠系統解析器（source_system = "hongcheng"）

資料夾特徵：
- 費用/ 子資料夾：逐月就診全明細 xlsx（每列=一次就診，含身份證號/看診日期/申請金額）
- 次數/ 子資料夾：逐月就診次數 PDF（v1 僅記錄，不解析；待 v2 補上 pypdf）
- 照護名單 / 指定名單 xlsx 或 csv
- 各類篩檢獨立 xlsx

費用 xlsx 處理：
- 按 身份證號 + 月份 聚合：次數=該月出現列數，金額=申請金額加總
- 最後就診日取該月最大看診日期
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

from db_pipeline.config.models import ClinicConfig
from db_pipeline.datasets.models import (
    DatasetBundle,
    MemberRecord,
    MemberSelectionRecord,
    MonthlyClaimRecord,
    SourceTrace,
)
from db_pipeline.normalization import (
    normalize_id,
    normalize_name,
    normalize_phone,
    normalize_text,
    parse_date,
    parse_decimal,
    stable_row_hash,
)
from db_pipeline.parsers.contracts import ParseCoverage, ParseResult
from db_pipeline.validation.models import ValidationIssue

# ── 常數 ──────────────────────────────────────────────────────────────────────
ID_HEADERS = ("身份證號", "身分證號", "身份證號碼", "身分證號碼", "ID")
MEMBER_FIELDS = {
    "case_category":            ("個案類別",),
    "quality_roster":           ("論質名單",),
    "multi_chronic_65":         ("65歲以上多重慢性病註記",),
    "high_visit":               ("高診次註記",),
    "chronic_mark":             ("慢性病註記",),
    "non_chronic_mark":         ("非慢性病註記",),
    "same_clinic_previous_year":("與前一年家醫收案診所相同",),
    "disease_pattern":          ("疾病樣態",),
    "ascvd":                    ("ASCVD",),
    "three_highs":              ("三高",),
    "hypertension":             ("高血壓",),
    "hyperlipidemia":           ("高血脂",),
    "hyperglycemia":            ("高血糖",),
}
MONTH_CODE_RE = re.compile(r"(?<!\d)(1(?:14|15)\d{2})(?!\d)")
TW_ID_RE      = re.compile(r"[A-Z][12]\d{8}")
XLSX_SUFFIXES = {".xlsx", ".xlsm"}


# ── 工具函式 ──────────────────────────────────────────────────────────────────
def _header_map(values: Sequence[Any]) -> Dict[str, int]:
    return {normalize_text(v): i for i, v in enumerate(values) if normalize_text(v)}


def _find_col(headers: Dict[str, int], aliases: Iterable[str]) -> Optional[int]:
    for a in aliases:
        if a in headers:
            return headers[a]
    return None


def _find_header_row(ws: Any, required: Sequence[Sequence[str]]) -> Optional[int]:
    for row_no in range(1, min(ws.max_row, 12) + 1):
        vals = [ws.cell(row_no, c).value for c in range(1, ws.max_column + 1)]
        hmap = _header_map(vals)
        if all(_find_col(hmap, grp) is not None for grp in required):
            return row_no
    return None


def _month_code(text: str) -> Optional[str]:
    m = MONTH_CODE_RE.search(text)
    return m.group(1) if m else None


def _trace(
    config: ClinicConfig, batch_id: str, source_dir: Path,
    file_path: Path, sheet_name: str, row_no: int, values: Sequence[Any],
) -> SourceTrace:
    return SourceTrace(
        clinic_code=config.clinic_code, batch_id=batch_id,
        source_system=config.source_system,
        source_file=str(file_path.relative_to(source_dir)),
        source_sheet=sheet_name, source_row=row_no,
        raw_row_hash=stable_row_hash(values),
    )


# ── 主解析器 ──────────────────────────────────────────────────────────────────
class HongchengParser:
    source_system = "hongcheng"

    def parse(
        self,
        source_dir: Path,
        config: ClinicConfig,
        batch_id: str,
    ) -> ParseResult:
        bundle   = DatasetBundle()
        coverage = ParseCoverage()
        issues: List[ValidationIssue] = []

        files = sorted(
            p for p in source_dir.rglob("*")
            if p.is_file() and not p.name.startswith(("~$", "."))
        )
        coverage.discovered_files = len(files)

        # 1. 照護名單
        roster_files = [
            p for p in files
            if p.suffix.lower() in XLSX_SUFFIXES
            and any(kw in p.name for kw in ("照護名單", "指定名單", "家醫名單"))
        ]
        for path in roster_files:
            n = self._parse_member_workbook(source_dir, path, config, batch_id, bundle)
            if n:
                coverage.parsed_files += 1
                coverage.parsed_rows["members"] = coverage.parsed_rows.get("members", 0) + n

        existing_ids = {m.person_id for m in bundle.members}

        # 2. 費用/ xlsx → monthly_claims（聚合）
        fee_dir = source_dir / "費用"
        fee_files: List[Path] = []
        if fee_dir.is_dir():
            fee_files = sorted(
                p for p in fee_dir.glob("*.xlsx")
                if not p.name.startswith(("~$", "."))
                and _month_code(p.stem) is not None
            )
        for path in fee_files:
            parsed, unmatched = self._parse_fee_workbook(
                source_dir, path, config, batch_id, existing_ids, bundle)
            if parsed or unmatched:
                coverage.parsed_files += 1
                coverage.parsed_rows["monthly_claims"] = (
                    coverage.parsed_rows.get("monthly_claims", 0) + parsed)
                coverage.unmatched_rows["monthly_claims"] = (
                    coverage.unmatched_rows.get("monthly_claims", 0) + unmatched)

        # 3. 次數/ PDF → 記錄為 skipped（v2 補上）
        pdf_dir = source_dir / "次數"
        pdf_files: List[Path] = []
        if pdf_dir.is_dir():
            pdf_files = list(pdf_dir.glob("*.pdf"))
        for p in pdf_files:
            coverage.skipped_files[str(p.relative_to(source_dir))] = (
                "PDF 次數解析預留 v2（需 pypdf）")

        # 4. 自選 / 不要 xlsx
        select_files = [
            p for p in files
            if p.suffix.lower() in XLSX_SUFFIXES
            and any(kw in p.name for kw in ("自選", "A115", "115自選"))
            and not any(kw in p.name.upper() for kw in ("115X", "不要"))
        ]
        exclude_files = [
            p for p in files
            if p.suffix.lower() in XLSX_SUFFIXES
            and any(kw in p.name.upper() for kw in ("115X", "不要"))
        ]
        member_indexes = self._build_indexes(bundle.members)
        for path in select_files:
            n, u = self._parse_selection_workbook(
                source_dir, path, config, batch_id, existing_ids, member_indexes,
                bundle, "self_selected_115")
            if n or u:
                coverage.parsed_files += 1
        for path in exclude_files:
            n, u = self._parse_selection_workbook(
                source_dir, path, config, batch_id, existing_ids, member_indexes,
                bundle, "excluded_115x")
            if n or u:
                coverage.parsed_files += 1

        # 5. 跳過
        parsed_paths = set(roster_files + fee_files + select_files + exclude_files)
        for p in files:
            if p not in parsed_paths and p not in pdf_files:
                coverage.skipped_files[str(p.relative_to(source_dir))] = (
                    "hongcheng v1 尚未實作此來源類型")

        if fee_files and not bundle.monthly_claims:
            issues.append(ValidationIssue(
                severity="error", dataset="monthly_claims",
                code="claim_not_mapped",
                message="找到費用來源檔，但沒有任何資料成功對應會員。"))
        if pdf_files:
            issues.append(ValidationIssue(
                severity="warning", dataset="monthly_claims",
                code="pdf_claims_skipped",
                message=f"有 {len(pdf_files)} 個次數 PDF 尚未解析（v2 補上）。"))
        for dataset, cnt in coverage.unmatched_rows.items():
            if cnt:
                issues.append(ValidationIssue(
                    severity="warning", dataset=dataset,
                    code="unmatched_source_rows",
                    message=f"有 {cnt} 筆來源資料無法唯一對應會員。"))

        return ParseResult(bundle=bundle, coverage=coverage, issues=issues)

    # ── 照護名單 ──────────────────────────────────────────────────────────────
    def _parse_member_workbook(
        self,
        source_dir: Path, path: Path,
        config: ClinicConfig, batch_id: str,
        bundle: DatasetBundle,
    ) -> int:
        wb = load_workbook(path, read_only=True, data_only=True)
        parsed = 0
        try:
            for ws in wb.worksheets:
                header_row = _find_header_row(ws, (ID_HEADERS, ("個案類別",)))
                if header_row is None:
                    continue
                hvals = [ws.cell(header_row, c).value
                         for c in range(1, ws.max_column + 1)]
                hmap   = _header_map(hvals)
                id_col = _find_col(hmap, ID_HEADERS)
                if id_col is None:
                    continue
                name_col   = _find_col(hmap, ("姓名",))
                birth_col  = _find_col(hmap, ("生日", "出生日期"))
                phone_col  = _find_col(hmap, ("電話",))
                mobile_col = _find_col(hmap, ("手機", "行動電話"))
                addr_col   = _find_col(hmap, ("地址", "住址"))
                field_cols = {fn: _find_col(hmap, aliases)
                              for fn, aliases in MEMBER_FIELDS.items()}
                for row_no, vals in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    pid = normalize_id(vals[id_col])
                    if not TW_ID_RE.fullmatch(pid):
                        continue
                    kwargs = {fn: normalize_text(vals[col]) if col is not None else ""
                              for fn, col in field_cols.items()}
                    trace = _trace(config, batch_id, source_dir, path, ws.title, row_no, vals)
                    bundle.members.append(MemberRecord(
                        trace=trace, person_id=pid,
                        name=normalize_name(vals[name_col]) if name_col is not None else "",
                        birth_date=parse_date(vals[birth_col]) if birth_col is not None else None,
                        phone=normalize_phone(vals[phone_col]) if phone_col is not None else "",
                        mobile=normalize_phone(vals[mobile_col]) if mobile_col is not None else "",
                        address=normalize_text(vals[addr_col]) if addr_col is not None else "",
                        **kwargs,
                    ))
                    bundle.member_selections.append(MemberSelectionRecord(
                        trace=trace, person_id=pid, selection_type="designated_114"))
                    parsed += 1
        finally:
            wb.close()
        return parsed

    # ── 費用 xlsx（全明細，按月聚合） ────────────────────────────────────────
    def _parse_fee_workbook(
        self,
        source_dir: Path, path: Path,
        config: ClinicConfig, batch_id: str,
        existing_ids: set,
        bundle: DatasetBundle,
    ) -> Tuple[int, int]:
        code = _month_code(path.stem) or _month_code(path.name)
        if code is None:
            return 0, 0
        roc_year = int(code[:3])
        month    = int(code[3:])

        wb = load_workbook(path, read_only=True, data_only=True)
        parsed = unmatched = 0
        try:
            ws = wb.worksheets[0]
            header_row = _find_header_row(ws, (ID_HEADERS, ("申請金額", "看診日期")))
            if header_row is None:
                return 0, 0
            hvals = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
            hmap      = _header_map(hvals)
            id_col    = _find_col(hmap, ID_HEADERS)
            amount_col = _find_col(hmap, ("申請金額", "總額", "費用"))
            date_col   = _find_col(hmap, ("看診日期", "日期", "就醫日期"))
            if id_col is None:
                return 0, 0

            # 按 pid 聚合
            agg: Dict[str, dict] = {}
            for vals in ws.iter_rows(min_row=header_row + 1, values_only=True):
                pid = normalize_id(vals[id_col])
                if not TW_ID_RE.fullmatch(pid):
                    continue
                amt   = parse_decimal(vals[amount_col]) if amount_col is not None else parse_decimal(0)
                vdate = parse_date(vals[date_col]) if date_col is not None else None
                if pid not in agg:
                    agg[pid] = {"count": 0, "amount": parse_decimal(0),
                                "last_visit": None, "first_row": None, "first_vals": None}
                rec = agg[pid]
                rec["count"] += 1
                rec["amount"] += amt
                if vdate and (rec["last_visit"] is None or vdate > rec["last_visit"]):
                    rec["last_visit"] = vdate
                if rec["first_vals"] is None:
                    rec["first_vals"] = vals
                    rec["first_row"] = ws.max_row  # 近似列號

            for row_no, (pid, rec) in enumerate(agg.items(), start=header_row + 1):
                if pid not in existing_ids:
                    unmatched += 1
                    continue
                bundle.monthly_claims.append(MonthlyClaimRecord(
                    trace=_trace(config, batch_id, source_dir, path,
                                 ws.title, row_no, rec["first_vals"] or [pid]),
                    person_id=pid,
                    roc_year=roc_year,
                    month=month,
                    visit_count=parse_decimal(rec["count"]),
                    amount=rec["amount"],
                    last_visit_date=rec["last_visit"],
                ))
                parsed += 1
        finally:
            wb.close()
        return parsed, unmatched

    # ── 自選 / 不要 ──────────────────────────────────────────────────────────
    def _parse_selection_workbook(
        self,
        source_dir: Path, path: Path,
        config: ClinicConfig, batch_id: str,
        existing_ids: set,
        indexes: Dict[str, Any],
        bundle: DatasetBundle,
        selection_type: str,
    ) -> Tuple[int, int]:
        wb = load_workbook(path, read_only=True, data_only=True)
        parsed = unmatched = 0
        try:
            for ws in wb.worksheets:
                header_row = _find_header_row(ws, (ID_HEADERS,))
                if header_row is None:
                    continue
                hvals = [ws.cell(header_row, c).value
                         for c in range(1, ws.max_column + 1)]
                hmap   = _header_map(hvals)
                id_col = _find_col(hmap, ID_HEADERS)
                if id_col is None:
                    continue
                for row_no, vals in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    pid = normalize_id(vals[id_col])
                    if not TW_ID_RE.fullmatch(pid):
                        continue
                    if pid not in existing_ids:
                        unmatched += 1
                        continue
                    bundle.member_selections.append(MemberSelectionRecord(
                        trace=_trace(config, batch_id, source_dir, path,
                                     ws.title, row_no, vals),
                        person_id=pid,
                        selection_type=selection_type,
                    ))
                    parsed += 1
        finally:
            wb.close()
        return parsed, unmatched

    @staticmethod
    def _build_indexes(members: Sequence[MemberRecord]) -> Dict[str, Any]:
        from collections import defaultdict
        idxs: Dict[str, Any] = {"name_phone": defaultdict(set)}
        for m in members:
            name = normalize_name(m.name)
            for phone in (m.phone, m.mobile):
                if phone:
                    idxs["name_phone"][(name, phone[-7:])].add(m.person_id)
        return idxs
