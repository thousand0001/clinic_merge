# -*- coding: utf-8 -*-
"""
自行系統解析器（source_system = "custom"）

涵蓋：書田自行系統、耀聖（本一/周一珊/嘉齡/家齡/益民等）、仁康、延平等
無 sm_* 標準格式，但格式接近：

資料夾特徵：
- R11440.xlsx（多 sheet，sheet 名稱 = 月份代碼，如 "11502"）
  或 R11440/ 子資料夾（逐月 xlsx，檔名含月份代碼）
  或 單一月報合計 xlsx（含身分證 + 次數 + 金額）
- 照護名單 / 指定名單 xlsx
- 自選會員 A115 / 不要自選 A115X xlsx
- 各類篩檢獨立 xlsx（成健/子抹/糞便/老流/BC肝）
- sm_* 若有，P4P/篩檢部分預留 v2

月份代碼識別：
1. xlsx 的 sheet 名稱（fullmatch 1140x / 1150x）
2. 檔名中的 5 位數民國年月
"""
from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

from db_pipeline.config.models import ClinicConfig
from db_pipeline.datasets.models import (
    DatasetBundle,
    MemberRecord,
    MemberSelectionRecord,
    MonthlyClaimRecord,
    SourceTrace,
)
from db_pipeline.normalization import (
    normalize_id,
    normalize_name,
    normalize_phone,
    normalize_text,
    parse_date,
    parse_decimal,
    stable_row_hash,
)
from db_pipeline.parsers.contracts import ParseCoverage, ParseResult
from db_pipeline.validation.models import ValidationIssue

# ── 常數 ──────────────────────────────────────────────────────────────────────
ID_HEADERS = ("身分證號", "身份證號", "身分證號碼", "身份證號碼", "ID", "家醫收案會員ID")
MEMBER_FIELDS = {
    "case_category":            ("個案類別",),
    "quality_roster":           ("論質名單",),
    "multi_chronic_65":         ("65歲以上多重慢性病註記",),
    "high_visit":               ("高診次註記",),
    "chronic_mark":             ("慢性病註記",),
    "non_chronic_mark":         ("非慢性病註記",),
    "same_clinic_previous_year":("與前一年家醫收案診所相同",),
    "disease_pattern":          ("疾病樣態",),
    "ascvd":                    ("ASCVD",),
    "three_highs":              ("三高",),
    "hypertension":             ("高血壓",),
    "hyperlipidemia":           ("高血脂",),
    "hyperglycemia":            ("高血糖",),
}
MONTH_CODE_RE   = re.compile(r"(?<!\d)(1(?:14|15)(?:0[1-9]|1[0-2]))(?!\d)")
SHEET_CODE_RE   = re.compile(r"^(1(?:14|15)(?:0[1-9]|1[0-2]))$")
TW_ID_RE        = re.compile(r"[A-Z][12]\d{8}")
XLSX_SUFFIXES   = {".xlsx", ".xlsm"}


# ── 工具函式 ──────────────────────────────────────────────────────────────────
def _header_map(values: Sequence[Any]) -> Dict[str, int]:
    return {normalize_text(v): i for i, v in enumerate(values) if normalize_text(v)}


def _find_col(headers: Dict[str, int], aliases: Iterable[str]) -> Optional[int]:
    for a in aliases:
        if a in headers:
            return headers[a]
    return None


def _find_header_row(ws: Any, required: Sequence[Sequence[str]]) -> Optional[int]:
    for row_no in range(1, min(ws.max_row, 12) + 1):
        vals = [ws.cell(row_no, c).value for c in range(1, ws.max_column + 1)]
        hmap = _header_map(vals)
        if all(_find_col(hmap, grp) is not None for grp in required):
            return row_no
    return None


def _month_code_from_name(text: str) -> Optional[str]:
    m = MONTH_CODE_RE.search(text)
    return m.group(1) if m else None


def _trace(
    config: ClinicConfig, batch_id: str, source_dir: Path,
    file_path: Path, sheet_name: str, row_no: int, values: Sequence[Any],
) -> SourceTrace:
    return SourceTrace(
        clinic_code=config.clinic_code, batch_id=batch_id,
        source_system=config.source_system,
        source_file=str(file_path.relative_to(source_dir)),
        source_sheet=sheet_name, source_row=row_no,
        raw_row_hash=stable_row_hash(values),
    )


# ── 主解析器 ──────────────────────────────────────────────────────────────────
class CustomParser:
    source_system = "custom"

    def parse(
        self,
        source_dir: Path,
        config: ClinicConfig,
        batch_id: str,
    ) -> ParseResult:
        bundle   = DatasetBundle()
        coverage = ParseCoverage()
        issues: List[ValidationIssue] = []

        files = sorted(
            p for p in source_dir.rglob("*")
            if p.is_file() and not p.name.startswith(("~$", "."))
        )
        coverage.discovered_files = len(files)

        # 1. 照護名單
        roster_files = [
            p for p in files
            if p.suffix.lower() in XLSX_SUFFIXES
            and any(kw in p.name for kw in ("照護名單", "指定名單", "家醫名單", "指定會員"))
        ]
        for path in roster_files:
            n = self._parse_member_workbook(source_dir, path, config, batch_id, bundle)
            if n:
                coverage.parsed_files += 1
                coverage.parsed_rows["members"] = coverage.parsed_rows.get("members", 0) + n

        existing_ids = {m.person_id for m in bundle.members}

        # 2. R11440 系列 → monthly_claims
        #    2a. R11440.xlsx（多月份 sheet）
        r11440_single = [
            p for p in files
            if p.suffix.lower() in XLSX_SUFFIXES
            and "R11440" in p.name.upper()
            and _month_code_from_name(p.stem) is None  # 非月份單檔
        ]
        #    2b. 月份單檔（R11440/ 子資料夾或根目錄含月份代碼）
        r11440_monthly = [
            p for p in files
            if p.suffix.lower() in XLSX_SUFFIXES
            and "R11440" in p.name.upper()
            and _month_code_from_name(p.stem) is not None
        ]
        # 也掃 R11440/ 子資料夾
        r11440_dir = source_dir / "R11440"
        if r11440_dir.is_dir():
            for p in sorted(r11440_dir.glob("*.xlsx")):
                if not p.name.startswith(("~$", ".")) and p not in r11440_monthly:
                    r11440_monthly.append(p)

        for path in r11440_single:
            parsed, unmatched = self._parse_r11440_multisheet(
                source_dir, path, config, batch_id, existing_ids, bundle)
            if parsed or unmatched:
                coverage.parsed_files += 1
                coverage.parsed_rows["monthly_claims"] = (
                    coverage.parsed_rows.get("monthly_claims", 0) + parsed)
                coverage.unmatched_rows["monthly_claims"] = (
                    coverage.unmatched_rows.get("monthly_claims", 0) + unmatched)

        for path in r11440_monthly:
            parsed, unmatched = self._parse_r11440_single(
                source_dir, path, config, batch_id, existing_ids, bundle)
            if parsed or unmatched:
                coverage.parsed_files += 1
                coverage.parsed_rows["monthly_claims"] = (
                    coverage.parsed_rows.get("monthly_claims", 0) + parsed)
                coverage.unmatched_rows["monthly_claims"] = (
                    coverage.unmatched_rows.get("monthly_claims", 0) + unmatched)

        # 3. 自選 / 不要
        member_indexes = self._build_indexes(bundle.members)
        select_files = [
            p for p in files
            if p.suffix.lower() in XLSX_SUFFIXES
            and any(kw in p.name for kw in ("自選", "A115", "115自選"))
            and not any(kw in p.name.upper() for kw in ("115X", "不要"))
        ]
        exclude_files = [
            p for p in files
            if p.suffix.lower() in XLSX_SUFFIXES
            and any(kw in p.name.upper() for kw in ("115X", "不要"))
        ]
        for path in select_files:
            n, u = self._parse_selection_workbook(
                source_dir, path, config, batch_id, existing_ids, member_indexes,
                bundle, "self_selected_115")
            if n or u:
                coverage.parsed_files += 1
        for path in exclude_files:
            n, u = self._parse_selection_workbook(
                source_dir, path, config, batch_id, existing_ids, member_indexes,
                bundle, "excluded_115x")
            if n or u:
                coverage.parsed_files += 1

        # 4. 跳過
        parsed_paths = set(
            roster_files + r11440_single + r11440_monthly + select_files + exclude_files)
        for p in files:
            if p not in parsed_paths:
                coverage.skipped_files[str(p.relative_to(source_dir))] = (
                    "custom v1 尚未實作此來源類型")

        if (r11440_single or r11440_monthly) and not bundle.monthly_claims:
            issues.append(ValidationIssue(
                severity="error", dataset="monthly_claims",
                code="claim_not_mapped",
                message=(
                    "費用次數來源檔（R11440）已找到，但 0 筆可對應照護名單。"
                    "請確認照護名單與費用檔是否屬同一診所同一期別。"
                )))
        for dataset, cnt in coverage.unmatched_rows.items():
            if cnt:
                issues.append(ValidationIssue(
                    severity="warning", dataset=dataset,
                    code="unmatched_source_rows",
                    message=(
                        f"費用次數檔有 {cnt} 筆姓名＋生日無法比對照護名單"
                        "（可能為非家醫計畫病患或姓名格式差異），已略過。"
                    )))

        return ParseResult(bundle=bundle, coverage=coverage, issues=issues)

    # ── 照護名單 ──────────────────────────────────────────────────────────────
    def _parse_member_workbook(
        self,
        source_dir: Path, path: Path,
        config: ClinicConfig, batch_id: str,
        bundle: DatasetBundle,
    ) -> int:
        wb = load_workbook(path, read_only=True, data_only=True)
        parsed = 0
        try:
            for ws in wb.worksheets:
                header_row = _find_header_row(ws, (ID_HEADERS, ("個案類別",)))
                if header_row is None:
                    continue
                hvals = [ws.cell(header_row, c).value
                         for c in range(1, ws.max_column + 1)]
                hmap   = _header_map(hvals)
                id_col = _find_col(hmap, ID_HEADERS)
                if id_col is None:
                    continue
                name_col   = _find_col(hmap, ("姓名",))
                birth_col  = _find_col(hmap, ("生日", "出生日期", "BIRTHDAY"))
                phone_col  = _find_col(hmap, ("電話",))
                mobile_col = _find_col(hmap, ("手機", "手機號碼"))
                addr_col   = _find_col(hmap, ("地址", "住址"))
                field_cols = {fn: _find_col(hmap, aliases)
                              for fn, aliases in MEMBER_FIELDS.items()}
                for row_no, vals in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    pid = normalize_id(vals[id_col])
                    if not TW_ID_RE.fullmatch(pid):
                        continue
                    kwargs = {fn: normalize_text(vals[col]) if col is not None else ""
                              for fn, col in field_cols.items()}
                    trace = _trace(config, batch_id, source_dir, path, ws.title, row_no, vals)
                    bundle.members.append(MemberRecord(
                        trace=trace, person_id=pid,
                        name=normalize_name(vals[name_col]) if name_col is not None else "",
                        birth_date=parse_date(vals[birth_col]) if birth_col is not None else None,
                        phone=normalize_phone(vals[phone_col]) if phone_col is not None else "",
                        mobile=normalize_phone(vals[mobile_col]) if mobile_col is not None else "",
                        address=normalize_text(vals[addr_col]) if addr_col is not None else "",
                        **kwargs,
                    ))
                    bundle.member_selections.append(MemberSelectionRecord(
                        trace=trace, person_id=pid, selection_type="designated_114"))
                    parsed += 1
        finally:
            wb.close()
        return parsed

    # ── R11440.xlsx 多月份 sheet ──────────────────────────────────────────────
    def _parse_r11440_multisheet(
        self,
        source_dir: Path, path: Path,
        config: ClinicConfig, batch_id: str,
        existing_ids: set,
        bundle: DatasetBundle,
    ) -> Tuple[int, int]:
        wb = load_workbook(path, read_only=True, data_only=True)
        total_parsed = total_unmatched = 0
        try:
            for ws in wb.worksheets:
                code = ws.title.strip()
                if not SHEET_CODE_RE.fullmatch(code):
                    # 也試從 sheet 名稱擷取月份代碼
                    code = _month_code_from_name(code) or ""
                if not code:
                    continue
                p, u = self._parse_claim_sheet(
                    ws, code, source_dir, path, config, batch_id, existing_ids, bundle)
                total_parsed   += p
                total_unmatched += u
        finally:
            wb.close()
        return total_parsed, total_unmatched

    # ── 月份單檔 ─────────────────────────────────────────────────────────────
    def _parse_r11440_single(
        self,
        source_dir: Path, path: Path,
        config: ClinicConfig, batch_id: str,
        existing_ids: set,
        bundle: DatasetBundle,
    ) -> Tuple[int, int]:
        code = _month_code_from_name(path.stem) or _month_code_from_name(path.name)
        if not code:
            return 0, 0
        wb = load_workbook(path, read_only=True, data_only=True)
        total_parsed = total_unmatched = 0
        try:
            for ws in wb.worksheets:
                p, u = self._parse_claim_sheet(
                    ws, code, source_dir, path, config, batch_id, existing_ids, bundle)
                total_parsed   += p
                total_unmatched += u
        finally:
            wb.close()
        return total_parsed, total_unmatched

    def _parse_claim_sheet(
        self,
        ws: Any, code: str,
        source_dir: Path, path: Path,
        config: ClinicConfig, batch_id: str,
        existing_ids: set,
        bundle: DatasetBundle,
    ) -> Tuple[int, int]:
        roc_year = int(code[:3])
        month    = int(code[3:])

        # 優先找有 ID 欄的 header
        header_row = _find_header_row(ws, (ID_HEADERS,))
        if header_row is None:
            # 退而求其次：有姓名+次數
            header_row = _find_header_row(ws, (("姓名",), ("次數", "件數", "來診次數")))
        if header_row is None:
            return 0, 0

        hvals = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
        hmap      = _header_map(hvals)
        id_col    = _find_col(hmap, ID_HEADERS)
        name_col  = _find_col(hmap, ("姓名",))
        birth_col = _find_col(hmap, ("生日", "出生日期"))
        count_col = _find_col(hmap, ("次數", "件數", "來診次數", "就診次數"))
        amount_col = _find_col(hmap, ("申請金額", "總額", "費用", "申報總金額"))
        date_col  = _find_col(hmap, ("最後就診日", "最後回診日", "日期"))

        parsed = unmatched = 0
        for row_no, vals in enumerate(
            ws.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            if id_col is not None:
                pid = normalize_id(vals[id_col])
                if not TW_ID_RE.fullmatch(pid):
                    continue
            else:
                # 無 ID 欄：跳過（需姓名+生日比對，留 v2）
                continue

            if pid not in existing_ids:
                unmatched += 1
                continue

            bundle.monthly_claims.append(MonthlyClaimRecord(
                trace=_trace(config, batch_id, source_dir, path, ws.title, row_no, vals),
                person_id=pid,
                roc_year=roc_year,
                month=month,
                visit_count=parse_decimal(vals[count_col]) if count_col is not None else parse_decimal(0),
                amount=parse_decimal(vals[amount_col]) if amount_col is not None else parse_decimal(0),
                last_visit_date=parse_date(vals[date_col]) if date_col is not None else None,
            ))
            parsed += 1
        return parsed, unmatched

    # ── 自選 / 不要 ──────────────────────────────────────────────────────────
    def _parse_selection_workbook(
        self,
        source_dir: Path, path: Path,
        config: ClinicConfig, batch_id: str,
        existing_ids: set,
        indexes: Dict[str, Any],
        bundle: DatasetBundle,
        selection_type: str,
    ) -> Tuple[int, int]:
        wb = load_workbook(path, read_only=True, data_only=True)
        parsed = unmatched = 0
        try:
            for ws in wb.worksheets:
                header_row = _find_header_row(ws, (ID_HEADERS,))
                if header_row is None:
                    header_row = _find_header_row(ws, (("姓名",), ("電話", "手機")))
                if header_row is None:
                    continue
                hvals = [ws.cell(header_row, c).value
                         for c in range(1, ws.max_column + 1)]
                hmap    = _header_map(hvals)
                id_col  = _find_col(hmap, ID_HEADERS)
                name_col = _find_col(hmap, ("姓名",))
                phone_col = _find_col(hmap, ("電話", "手機", "手機號碼"))
                for row_no, vals in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    pid = normalize_id(vals[id_col]) if id_col is not None else ""
                    if not TW_ID_RE.fullmatch(pid):
                        # 姓名+電話比對
                        name  = normalize_name(vals[name_col]) if name_col is not None else ""
                        phone = normalize_phone(vals[phone_col]) if phone_col is not None else ""
                        if not name or len(phone) < 7:
                            continue
                        ids = indexes.get("name_phone", {}).get((name, phone[-7:]), set())
                        if len(ids) != 1:
                            unmatched += 1
                            continue
                        pid = next(iter(ids))

                    if pid not in existing_ids:
                        unmatched += 1
                        continue
                    bundle.member_selections.append(MemberSelectionRecord(
                        trace=_trace(config, batch_id, source_dir, path,
                                     ws.title, row_no, vals),
                        person_id=pid,
                        selection_type=selection_type,
                    ))
                    parsed += 1
        finally:
            wb.close()
        return parsed, unmatched

    @staticmethod
    def _build_indexes(members: Sequence[MemberRecord]) -> Dict[str, Any]:
        idxs: Dict[str, Any] = {"name_phone": defaultdict(set)}
        for m in members:
            name = normalize_name(m.name)
            for phone in (m.phone, m.mobile):
                if phone:
                    idxs["name_phone"][(name, phone[-7:])].add(m.person_id)
        return idxs
