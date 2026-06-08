# -*- coding: utf-8 -*-
"""
新耀聖系統解析器

與耀聖（SM）差異：
- 照護名單：姓名欄在 ID 與 BIRTHDAY 之間，無表頭
- 就診次數：有身份證號欄，直接對應，無需 name+birth 推導
- P4P 收案/追蹤：獨立 xlsx，直接用 ID
- 篩檢（BC肝/子宮頸/成健/糞便/老人流感）：獨立 xlsx，直接用 ID
- 個案健康管理：HbA1c、LDL 等最新檢驗值
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

from openpyxl import load_workbook

from db_pipeline.config.models import ClinicConfig
from db_pipeline.datasets.models import (
    DatasetBundle,
    LabResultRecord,
    MemberRecord,
    MemberSelectionRecord,
    MonthlyClaimRecord,
    P4PCaseRecord,
    P4PTrackRecord,
    ScreeningRecord,
)
from db_pipeline.normalization import (
    normalize_id,
    normalize_name,
    normalize_text,
    parse_date,
    parse_decimal,
    stable_row_hash,
)
from db_pipeline.parsers.contracts import ParseCoverage, ParseResult
from db_pipeline.validation.models import ValidationIssue
from .sm import (
    ID_HEADERS,
    MEMBER_FIELDS,
    _find_column,
    _find_header_row,
    _header_map,
    _trace,
)

_P4P_ID_HEADERS = ("家醫收案會員ID", "ID", "身分證號", "身份證號")
_CLAIM_ID_HEADERS = ("身份證號", "身份證號碼", "ID", "身分證號")

_SCREENING_FILENAME_MAP: Dict[str, str] = {
    "bc肝": "BC肝",
    "b型肝": "BC肝",
    "子宮頸": "子宮頸",
    "成健": "成健",
    "糞便": "糞便",
    "老人流感": "老人流感",
    "老流": "老人流感",
}


def _screening_type_from_filename(stem: str) -> Optional[str]:
    lower = stem.lower()
    for key, typ in _SCREENING_FILENAME_MAP.items():
        if key in lower:
            return typ
    return None


class NewSmParser:
    source_system = "new_sm"

    def parse(
        self,
        source_dir: Path,
        config: ClinicConfig,
        batch_id: str,
    ) -> ParseResult:
        bundle = DatasetBundle()
        coverage = ParseCoverage()
        issues: List[ValidationIssue] = []

        files = sorted(
            p for p in source_dir.rglob("*")
            if p.is_file()
            and not p.name.startswith(("~$", "."))
            and p.suffix.lower() in (".xlsx", ".xlsm")
        )
        coverage.discovered_files = len(files)
        if not files:
            issues.append(ValidationIssue(
                severity="error", dataset="source", code="empty_source_directory",
                message="來源資料夾沒有可解析的 xlsx 檔案。",
            ))

        roster_files    = [p for p in files if "照護名單" in p.name or "指定會員名單" in p.name]
        claim_files     = [p for p in files if "就診次數" in p.name or "會員就診" in p.name]
        p4p_case_files  = [p for p in files if re.search(r"P4P.*收案|P4p.*收案", p.name)]
        p4p_track_files = [p for p in files if re.search(r"P4P.*追蹤|P4p.*追蹤", p.name)]
        health_files    = [p for p in files if "個案健康管理" in p.name]
        screening_files = [p for p in files if _screening_type_from_filename(p.stem) is not None]

        parsed_paths: Set[Path] = set()

        for path in roster_files:
            n = self._parse_roster(source_dir, path, config, batch_id, bundle)
            if n:
                coverage.parsed_files += 1
                coverage.parsed_rows["members"] = coverage.parsed_rows.get("members", 0) + n
            parsed_paths.add(path)

        member_ids: Set[str] = {m.person_id for m in bundle.members}

        for path in claim_files:
            ok, ng = self._parse_claims(source_dir, path, config, batch_id, member_ids, bundle)
            if ok or ng:
                coverage.parsed_files += 1
                coverage.parsed_rows["monthly_claims"]   = coverage.parsed_rows.get("monthly_claims", 0) + ok
                coverage.unmatched_rows["monthly_claims"] = coverage.unmatched_rows.get("monthly_claims", 0) + ng
            parsed_paths.add(path)

        for path in p4p_case_files:
            n = self._parse_p4p_cases(source_dir, path, config, batch_id, bundle)
            if n:
                coverage.parsed_files += 1
                coverage.parsed_rows["p4p_cases"] = coverage.parsed_rows.get("p4p_cases", 0) + n
            parsed_paths.add(path)

        for path in p4p_track_files:
            n = self._parse_p4p_tracks(source_dir, path, config, batch_id, bundle)
            if n:
                coverage.parsed_files += 1
                coverage.parsed_rows["p4p_tracks"] = coverage.parsed_rows.get("p4p_tracks", 0) + n
            parsed_paths.add(path)

        for path in screening_files:
            stype = _screening_type_from_filename(path.stem)
            n = self._parse_screenings(source_dir, path, config, batch_id, stype, bundle)
            if n:
                coverage.parsed_files += 1
                coverage.parsed_rows["screenings"] = coverage.parsed_rows.get("screenings", 0) + n
            parsed_paths.add(path)

        for path in health_files:
            n = self._parse_health_mgmt(source_dir, path, config, batch_id, bundle)
            if n:
                coverage.parsed_files += 1
                coverage.parsed_rows["lab_results"] = coverage.parsed_rows.get("lab_results", 0) + n
            parsed_paths.add(path)

        for path in files:
            if path not in parsed_paths:
                coverage.skipped_files[str(path.relative_to(source_dir))] = "新耀聖 v1 尚未實作此來源類型"

        for dataset, count in coverage.unmatched_rows.items():
            if count:
                issues.append(ValidationIssue(
                    severity="warning", dataset=dataset,
                    code="non_roster_claims",
                    message=(
                        f"就診次數檔含 {count} 筆一般門診病患（非家醫計畫成員），"
                        "已全數寫入；這些記錄無對應照護名單資料，屬正常現象，毋需處理。"
                    ),
                ))

        return ParseResult(bundle=bundle, coverage=coverage, issues=issues)

    # ── 照護名單 ────────────────────────────────────────────────────────────────

    def _parse_roster(
        self,
        source_dir: Path,
        path: Path,
        config: ClinicConfig,
        batch_id: str,
        bundle: DatasetBundle,
    ) -> int:
        wb = load_workbook(path, read_only=True, data_only=True)
        parsed = 0
        try:
            for ws in wb.worksheets:
                header_row = _find_header_row(ws, (ID_HEADERS, ("個案類別",)))
                if header_row is None:
                    continue
                header_values = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
                headers = _header_map(header_values)
                id_col = _find_column(headers, ID_HEADERS)
                if id_col is None:
                    continue
                birth_col = _find_column(headers, ("BIRTHDAY", "生日"))
                name_col  = _find_column(headers, ("姓名", "會員姓名"))
                # 新耀聖格式：姓名欄在 ID 與 BIRTHDAY 之間，無表頭
                if name_col is None and birth_col is not None and birth_col > id_col + 1:
                    name_col = id_col + 1
                field_cols = {k: _find_column(headers, v) for k, v in MEMBER_FIELDS.items()}

                for row_no, values in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    person_id = normalize_id(values[id_col])
                    if not re.fullmatch(r"[A-Z][12][0-9]{8}", person_id):
                        continue
                    kwargs = {
                        k: (normalize_text(values[c]) if c is not None else "")
                        for k, c in field_cols.items()
                    }
                    bundle.members.append(MemberRecord(
                        trace=_trace(config, batch_id, source_dir, path, ws.title, row_no, values),
                        person_id=person_id,
                        name=normalize_name(values[name_col]) if name_col is not None else "",
                        birth_date=parse_date(values[birth_col]) if birth_col is not None else None,
                        **kwargs,
                    ))
                    bundle.member_selections.append(MemberSelectionRecord(
                        trace=bundle.members[-1].trace,
                        person_id=person_id,
                        selection_type="designated_114",
                    ))
                    parsed += 1
        finally:
            wb.close()
        return parsed

    # ── 就診次數 ────────────────────────────────────────────────────────────────

    def _parse_claims(
        self,
        source_dir: Path,
        path: Path,
        config: ClinicConfig,
        batch_id: str,
        member_ids: Set[str],
        bundle: DatasetBundle,
    ) -> Tuple[int, int]:
        """所有有效身分證號的就診記錄均寫入，不限於照護名單成員。
        非照護名單成員的 claims 不會有對應的 members 記錄，
        但就診次數與費用本身是完整資料。
        """
        wb = load_workbook(path, read_only=True, data_only=True)
        matched = unmatched = 0
        try:
            for ws in wb.worksheets:
                m = re.fullmatch(r"(114|115)(0[1-9]|1[0-2])", ws.title.strip())
                if not m:
                    continue
                header_row = _find_header_row(ws, (
                    _CLAIM_ID_HEADERS,
                    ("件數", "來診次數", "次數"),
                    ("申請金額", "申報總金額", "總額"),
                ))
                if header_row is None:
                    continue
                headers = _header_map(
                    [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
                )
                id_col     = _find_column(headers, _CLAIM_ID_HEADERS)
                count_col  = _find_column(headers, ("件數", "來診次數", "次數"))
                amount_col = _find_column(headers, ("申請金額", "申報總金額", "總額"))
                if None in (id_col, count_col, amount_col):
                    continue
                last_visit_col = _find_column(headers, ("最後回診日", "最後就診日"))

                for row_no, values in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    person_id = normalize_id(values[id_col])
                    if not re.fullmatch(r"[A-Z][12][0-9]{8}", person_id):
                        continue
                    in_roster = person_id in member_ids
                    bundle.monthly_claims.append(MonthlyClaimRecord(
                        trace=_trace(config, batch_id, source_dir, path, ws.title, row_no, values),
                        person_id=person_id,
                        roc_year=int(m.group(1)),
                        month=int(m.group(2)),
                        visit_count=parse_decimal(values[count_col]),
                        amount=parse_decimal(values[amount_col]),
                        last_visit_date=(
                            parse_date(values[last_visit_col]) if last_visit_col is not None else None
                        ),
                    ))
                    if in_roster:
                        matched += 1
                    else:
                        unmatched += 1
        finally:
            wb.close()
        return matched, unmatched

    # ── P4P 收案 ────────────────────────────────────────────────────────────────

    def _parse_p4p_cases(
        self,
        source_dir: Path,
        path: Path,
        config: ClinicConfig,
        batch_id: str,
        bundle: DatasetBundle,
    ) -> int:
        wb = load_workbook(path, read_only=True, data_only=True)
        parsed = 0
        try:
            for ws in wb.worksheets:
                header_row = _find_header_row(ws, (_P4P_ID_HEADERS, ("P4P收案計畫", "收案計畫")))
                if header_row is None:
                    continue
                headers = _header_map(
                    [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
                )
                id_col     = _find_column(headers, _P4P_ID_HEADERS)
                plan_col   = _find_column(headers, ("P4P收案計畫", "收案計畫"))
                status_col = _find_column(headers, ("收案狀態", "狀態"))
                if id_col is None:
                    continue
                for row_no, values in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    person_id = normalize_id(values[id_col])
                    if not re.fullmatch(r"[A-Z][12][0-9]{8}", person_id):
                        continue
                    bundle.p4p_cases.append(P4PCaseRecord(
                        trace=_trace(config, batch_id, source_dir, path, ws.title, row_no, values),
                        person_id=person_id,
                        plan=normalize_text(values[plan_col]) if plan_col is not None else "",
                        status=normalize_text(values[status_col]) if status_col is not None else "",
                    ))
                    parsed += 1
        finally:
            wb.close()
        return parsed

    # ── P4P 追蹤 ────────────────────────────────────────────────────────────────

    def _parse_p4p_tracks(
        self,
        source_dir: Path,
        path: Path,
        config: ClinicConfig,
        batch_id: str,
        bundle: DatasetBundle,
    ) -> int:
        wb = load_workbook(path, read_only=True, data_only=True)
        parsed = 0
        try:
            for ws in wb.worksheets:
                header_row = _find_header_row(ws, (_P4P_ID_HEADERS, ("最後追蹤日",)))
                if header_row is None:
                    continue
                headers = _header_map(
                    [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
                )
                id_col         = _find_column(headers, _P4P_ID_HEADERS)
                plan_col       = _find_column(headers, ("P4P收案計畫", "收案計畫"))
                last_track_col = _find_column(headers, ("最後追蹤日", "最近追蹤日"))
                next_track_col = _find_column(headers, ("下次應追蹤日", "下次追蹤日"))
                overdue_col    = _find_column(headers, ("逾期未追蹤",))
                if id_col is None:
                    continue
                for row_no, values in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    person_id = normalize_id(values[id_col])
                    if not re.fullmatch(r"[A-Z][12][0-9]{8}", person_id):
                        continue
                    bundle.p4p_tracks.append(P4PTrackRecord(
                        trace=_trace(config, batch_id, source_dir, path, ws.title, row_no, values),
                        person_id=person_id,
                        plan=normalize_text(values[plan_col]) if plan_col is not None else "",
                        last_tracked_at=(
                            parse_date(values[last_track_col]) if last_track_col is not None else None
                        ),
                        next_track_at=(
                            parse_date(values[next_track_col]) if next_track_col is not None else None
                        ),
                        overdue=normalize_text(values[overdue_col]) if overdue_col is not None else "",
                    ))
                    parsed += 1
        finally:
            wb.close()
        return parsed

    # ── 篩檢 ────────────────────────────────────────────────────────────────────

    def _parse_screenings(
        self,
        source_dir: Path,
        path: Path,
        config: ClinicConfig,
        batch_id: str,
        screening_type: str,
        bundle: DatasetBundle,
    ) -> int:
        wb = load_workbook(path, read_only=True, data_only=True)
        parsed = 0
        try:
            for ws in wb.worksheets:
                header_row = _find_header_row(ws, (ID_HEADERS, ("最後篩檢日期", "篩檢日期", "最後檢查日期")))
                if header_row is None:
                    continue
                headers = _header_map(
                    [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
                )
                id_col   = _find_column(headers, ID_HEADERS)
                date_col = _find_column(headers, ("最後篩檢日期", "篩檢日期", "最後檢查日期"))
                if id_col is None or date_col is None:
                    continue
                for row_no, values in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    person_id = normalize_id(values[id_col])
                    if not re.fullmatch(r"[A-Z][12][0-9]{8}", person_id):
                        continue
                    screened_at = parse_date(values[date_col])
                    if screened_at is None:
                        continue
                    bundle.screenings.append(ScreeningRecord(
                        trace=_trace(config, batch_id, source_dir, path, ws.title, row_no, values),
                        person_id=person_id,
                        screening_type=screening_type,
                        screened_at=screened_at,
                    ))
                    parsed += 1
        finally:
            wb.close()
        return parsed

    # ── 個案健康管理（lab results） ──────────────────────────────────────────────

    def _parse_health_mgmt(
        self,
        source_dir: Path,
        path: Path,
        config: ClinicConfig,
        batch_id: str,
        bundle: DatasetBundle,
    ) -> int:
        wb = load_workbook(path, read_only=True, data_only=True)
        parsed = 0
        try:
            for ws in wb.worksheets:
                header_row = _find_header_row(ws, (_P4P_ID_HEADERS,))
                if header_row is None:
                    continue
                header_values = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
                headers = _header_map(header_values)
                id_col = _find_column(headers, _P4P_ID_HEADERS)
                if id_col is None:
                    continue
                # 找「檢查結果」欄及其相鄰「日期」欄配對
                lab_pairs: List[Tuple[str, int, int]] = []
                for i, hdr in enumerate(header_values):
                    if hdr is None:
                        continue
                    h = normalize_text(str(hdr))
                    if "檢查結果" in h or "檢驗結果" in h:
                        test_code = re.sub(r"最近一次|最新一次", "", str(hdr))
                        test_code = test_code.split("(")[0].split("（")[0].strip()
                        date_idx: Optional[int] = None
                        for j in range(i + 1, min(i + 4, len(header_values))):
                            h2 = normalize_text(str(header_values[j])) if header_values[j] else ""
                            if "日期" in h2:
                                date_idx = j
                                break
                        if date_idx is not None:
                            lab_pairs.append((test_code, i, date_idx))

                if not lab_pairs:
                    continue

                for row_no, values in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    person_id = normalize_id(values[id_col])
                    if not re.fullmatch(r"[A-Z][12][0-9]{8}", person_id):
                        continue
                    trace = _trace(config, batch_id, source_dir, path, ws.title, row_no, values)
                    for test_code, val_idx, date_idx in lab_pairs:
                        result_val  = values[val_idx]
                        result_date = parse_date(values[date_idx])
                        if result_val is None and result_date is None:
                            continue
                        bundle.lab_results.append(LabResultRecord(
                            trace=trace,
                            person_id=person_id,
                            test_code=test_code,
                            result_value=str(result_val) if result_val is not None else "",
                            tested_at=result_date,
                        ))
                        parsed += 1
        finally:
            wb.close()
        return parsed
