from __future__ import annotations

import csv
import datetime as dt
import hashlib
import io
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Sequence

from openpyxl import load_workbook

from db_pipeline.datasets.models import RawSourceFile, RawSourceRow
from db_pipeline.normalization import normalize_text, stable_row_hash


CSV_ENCODINGS = ("utf-8-sig", "utf-16", "cp950", "big5")
EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
TEXT_SUFFIXES = {".csv", ".txt"}
ILLEGAL_CTRL_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")


@dataclass
class RawCollectionResult:
    source_files: List[RawSourceFile] = field(default_factory=list)
    rows: List[RawSourceRow] = field(default_factory=list)
    unreadable_files: Dict[str, str] = field(default_factory=dict)


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _row_payload(values: Sequence[Any]) -> Dict[str, object]:
    return {
        "values": [normalize_text(value) for value in values],
    }


def _append_row(
    result: RawCollectionResult,
    source_dir: Path,
    path: Path,
    sheet_name: str,
    row_no: int,
    values: Sequence[Any],
) -> None:
    if not any(normalize_text(value) for value in values):
        return
    result.rows.append(
        RawSourceRow(
            source_file=str(path.relative_to(source_dir)),
            file_name=path.name,
            sheet_name=sheet_name,
            row_no=row_no,
            row_data=_row_payload(values),
            row_hash=stable_row_hash(values),
        )
    )


def _read_text(path: Path) -> str:
    for encoding in CSV_ENCODINGS:
        try:
            return path.read_text(encoding=encoding, errors="strict")
        except UnicodeError:
            continue
    return path.read_text(encoding="cp950", errors="replace")


def collect_raw_sources(source_dir: Path) -> RawCollectionResult:
    result = RawCollectionResult()
    paths = sorted(
        path
        for path in source_dir.rglob("*")
        if path.is_file() and not path.name.startswith(("~$", "."))
    )

    for path in paths:
        stat = path.stat()
        suffix = path.suffix.lower()
        relative_path = str(path.relative_to(source_dir))
        result.source_files.append(
            RawSourceFile(
                relative_path=relative_path,
                file_name=path.name,
                file_size=stat.st_size,
                sha256=_sha256_file(path),
                file_mtime=dt.datetime.fromtimestamp(
                    stat.st_mtime,
                    tz=dt.timezone.utc,
                ),
                data_type=suffix.lstrip(".") or "unknown",
            )
        )

        try:
            if suffix in EXCEL_SUFFIXES:
                workbook = load_workbook(path, read_only=True, data_only=True)
                try:
                    for worksheet in workbook.worksheets:
                        for row_no, values in enumerate(
                            worksheet.iter_rows(values_only=True),
                            start=1,
                        ):
                            _append_row(
                                result,
                                source_dir,
                                path,
                                worksheet.title,
                                row_no,
                                values,
                            )
                finally:
                    workbook.close()
            elif suffix in TEXT_SUFFIXES:
                cleaned = ILLEGAL_CTRL_RE.sub("", _read_text(path))
                for row_no, values in enumerate(
                    csv.reader(io.StringIO(cleaned)),
                    start=1,
                ):
                    _append_row(
                        result,
                        source_dir,
                        path,
                        "",
                        row_no,
                        values,
                    )
            else:
                result.unreadable_files[relative_path] = (
                    "已保存檔案中繼資料與雜湊，尚未展開原始列"
                )
        except Exception as exc:
            result.unreadable_files[relative_path] = (
                f"已保存檔案中繼資料與雜湊，原始列讀取失敗：{exc}"
            )

    return result
