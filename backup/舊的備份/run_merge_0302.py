# -*- coding: utf-8 -*-
"""
選會員產檔工具（結構化單檔版）— 重構修正版（0301樣板）

功能概要
- 讀取原始檔（會員名單 / ascvd / HealthCase / 5 大篩檢）
- 套用模板（保留所有分頁；目前只寫入「會員指標」分頁），寫入資料列
- 計算：疾病樣態文字、備註、分數/明細、AU/AV/AW 文字提醒
- KPI
  - HbA1c：AY8（<=7 的比例）
  - HbA1c：AZ9（>=73.8% 所需人數），AZ8（該人數對應的 HbA1c 切點值）
  - LDL：BB8/BC8（<=100 / <=110）
- 全表格加網格線、輸出到原始檔同資料夾並自動開檔

注意
- Python 3.9 相容
- 不拆模組：同一支檔案即可維護
"""

from __future__ import annotations

import datetime
import math
import os
import re
import subprocess
import sys
from dataclasses import dataclass
from enum import Enum, auto
from typing import Dict, List, Optional, Any, Tuple

import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.utils.exceptions import InvalidFileException


# ============================================================
# 業務規則常數（每年調整這裡即可）
# ============================================================
class Rules:
    # 檢驗日期：哪些年份視為「有效」
    VALID_YEARS: Tuple[int, ...] = (2025, 2026)

    # 篩檢計分：哪一年的篩檢才給分
    SCREEN_YEAR: int = 2026

    # 2026 年：只要有檢查就給滿分（無條件給分）
    # 2025 年：需達標才給分
    YEAR_UNCONDITIONAL: int = 2026
    YEAR_VALUE_CHECK: int = 2025

    # 各項分數
    SCORE_HBA: int = 5
    SCORE_LDL: int = 5
    SCORE_ADULT: int = 6
    SCORE_PAP: int = 6
    SCORE_FLU: int = 4
    SCORE_FIT: int = 6
    SCORE_HEP: int = 6

    # 達標值（2025年）
    HBA_PASS_2025: float = 7.0
    LDL_PASS_2025: float = 120.0

    # KPI（HbA1c）
    HBA_CONTROL_THRESHOLD: float = 7.0  # AY8：HbA1c <= 7 的比例
    HBA_TARGET_PERCENT: float = 0.738   # AZ：要找 >=73.8% 的切點

    # KPI（LDL）
    LDL_KPI_THRESHOLDS: Tuple[float, float] = (100.0, 110.0)  # BB8/BC8

    # 回診追蹤
    AU_DAYS: int = 28
    AV_OFFSET_DAYS: int = 28
    AW_OFFSET_DAYS: int = 56

    # 模板設定（0301）
    TEMPLATE_NAME: str = "選會員樣板0301.xlsx"
    SHEET_TARGET: str = "會員指標"
    DATA_START_ROW: int = 10


# ============================================================
# Enum / Dataclass
# ============================================================
class DiseaseCode(Enum):
    DM = 1
    CKD = 2
    DKD = 3   # DM + CKD
    OTHER = 4 # 非 DM/CKD，可能有 ASCVD


class AscvdCategory(Enum):
    NONE = auto()  # 無 ASCVD
    A = auto()     # 極高風險
    B = auto()     # 非常高風險


@dataclass
class MemberMeta:
    """每位會員的衍生資訊，供後續計算用"""
    row: int
    bday: Optional[datetime.date] = None
    age: int = -1
    e_code: Optional[DiseaseCode] = None
    ascvd: AscvdCategory = AscvdCategory.NONE


# ============================================================
# 工具函數
# ============================================================
def open_file_cross_platform(path: str) -> None:
    """跨平台開啟檔案（盡量不拋錯）"""
    try:
        if sys.platform.startswith("darwin"):
            subprocess.call(("open", path))
        elif os.name == "nt":
            os.startfile(path)  # type: ignore[attr-defined]
        else:
            subprocess.call(("xdg-open", path))
    except Exception:
        pass


def normalize_text(v: Any) -> str:
    if v is None:
        return ""
    return str(v).replace("\t", "").replace("　", "").strip()


def normalize_header(v: Any) -> str:
    s = normalize_text(v)
    return s.replace(" ", "").replace("\n", "")


def clean_spaces(v: Any) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", "", str(v))


def safe_set(ws, row: int, col: Optional[int], value: Any) -> None:
    if col:
        ws.cell(row, col).value = value


def calc_age(bday: Optional[datetime.date], ref: datetime.date) -> int:
    if not isinstance(bday, datetime.date):
        return -1
    y = ref.year - bday.year
    if (ref.month, ref.day) < (bday.month, bday.day):
        y -= 1
    return y


def add_months(d: datetime.date, months: int) -> datetime.date:
    y = d.year + (d.month - 1 + months) // 12
    m = (d.month - 1 + months) % 12 + 1
    is_leap = (y % 4 == 0 and (y % 100 != 0 or y % 400 == 0))
    mdays = [31, 29 if is_leap else 28, 31, 30, 31, 30,
             31, 31, 30, 31, 30, 31][m - 1]
    return datetime.date(y, m, min(d.day, mdays))


def infer_gender_from_id(id_value: Any) -> str:
    """由台灣身分證/居留證第2碼推性別（1/8/A/C→男，2/9/B/D→女）"""
    if not id_value:
        return ""
    s = str(id_value).strip().upper()
    if len(s) < 2:
        return ""
    g = s[1]
    if g in ("1", "8", "A", "C"):
        return "男"
    if g in ("2", "9", "B", "D"):
        return "女"
    return ""


# ============================================================
# 解析函數
# ============================================================
def parse_date(value: Any) -> Optional[datetime.date]:
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value

    s = str(value).strip()
    if not s or s in ("-", "—", "–"):
        return None

    patterns = [
        (r"^(\d{4})[/-](\d{1,2})[/-](\d{1,2})$", False),
        (r"^(\d{2,3})(\d{2})(\d{2})$", True),
        (r"^(\d{2,3})[/-](\d{1,2})[/-](\d{1,2})$", True),
    ]
    for pat, is_roc in patterns:
        m = re.match(pat, s)
        if m:
            a, b, c = map(int, m.groups())
            year = a + 1911 if is_roc else a
            try:
                return datetime.date(year, b, c)
            except ValueError:
                return None
    return None


def parse_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    s = str(value).strip()
    if not s or s in ("-", "—", "–"):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def parse_disease_code(v: Any) -> Optional[DiseaseCode]:
    s = clean_spaces(v)
    if not s:
        return None
    if s.isdigit():
        mapping = {
            1: DiseaseCode.DM,
            2: DiseaseCode.CKD,
            3: DiseaseCode.DKD,
            4: DiseaseCode.OTHER,
        }
        return mapping.get(int(s))
    text_map = {
        "DM":     DiseaseCode.DM,
        "CKD":    DiseaseCode.CKD,
        "DKD":    DiseaseCode.DKD,
        "DM+CKD": DiseaseCode.DKD,
        "DMCKD":  DiseaseCode.DKD,
    }
    return text_map.get(s.upper())


def parse_ascvd(v: Any) -> AscvdCategory:
    if v is None:
        return AscvdCategory.NONE
    s = clean_spaces(v).lower()
    if not s or s == "0":
        return AscvdCategory.NONE
    if s in ("a", "1") or "極高" in s:
        return AscvdCategory.A
    if s in ("b", "2") or "非常高" in s:
        return AscvdCategory.B
    if "a" in s:
        return AscvdCategory.A
    if "b" in s:
        return AscvdCategory.B
    return AscvdCategory.NONE


def is_zero_like(v: Any) -> bool:
    """把 None / 空白 / '0' 視為零值"""
    if v is None:
        return True
    return str(v).strip() in ("", "0", "0.0")


# ============================================================
# 篩檢需求判斷
# ============================================================
def adult_check_interval_months(age: int) -> Optional[int]:
    if 30 <= age <= 39:
        return 60
    if 40 <= age <= 64:
        return 36
    if age >= 65:
        return 12
    return None


def pap_check_interval_months(age: int, sex: str) -> Optional[int]:
    if sex != "女":
        return None
    if 25 <= age <= 29:
        return 36
    if age >= 30:
        return 12
    return None


def need_flu(age: int) -> bool:
    return age >= 65


def need_fit(age: int) -> bool:
    return 45 <= age <= 75


def need_bc_hep(age: int) -> bool:
    return 45 <= age <= 80


# ============================================================
# 欄位偵測
# ============================================================
def build_header_map(sheet, header_row: int) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for c in range(1, sheet.max_column + 1):
        k = normalize_text(sheet.cell(header_row, c).value)
        if k:
            m[k] = c
    return m


def find_column_exact(hmap: Dict[str, int], aliases: List[str]) -> Optional[int]:
    for a in aliases:
        if a in hmap:
            return hmap[a]
    return None


def _get_merged_header_text(ws, header_row: int, c: int) -> str:
    upper = normalize_header(ws.cell(header_row - 1, c).value) if header_row > 1 else ""
    lower = normalize_header(ws.cell(header_row, c).value)
    return (upper + " " + lower).strip()


def find_col_by_keywords(ws, header_row: int, keywords: List[str]) -> Optional[int]:
    """模糊比對：合併上下兩列表頭，同時包含所有 keywords"""
    keys = [k.replace(" ", "").replace("\n", "") for k in keywords if k]
    for c in range(1, ws.max_column + 1):
        h = _get_merged_header_text(ws, header_row, c).replace(" ", "")
        if h and all(k in h for k in keys):
            return c
    return None


def find_col_by_keywords_any_row(ws, max_row: int, keywords: List[str]) -> Optional[int]:
    """跨多列找欄位：同欄所有儲存格合併後包含全部 keywords"""
    keys = [k.replace(" ", "").replace("\n", "") for k in keywords if k]
    for c in range(1, ws.max_column + 1):
        blob = "".join(
            normalize_header(ws.cell(r, c).value)
            for r in range(1, max_row + 1)
        )
        if blob and all(k in blob for k in keys):
            return c
    return None


def find_header_row_contains(sheet, must_have: List[str], search_rows: int = 250) -> Optional[int]:
    must = [m.replace(" ", "").replace("\n", "") for m in must_have]
    for r in range(1, min(search_rows, sheet.max_row) + 1):
        found = set()
        for c in range(1, sheet.max_column + 1):
            v = normalize_header(sheet.cell(r, c).value)
            if v in must:
                found.add(v)
        if all(m in found for m in must):
            return r
    return None


def _detect_basic_cols(ws, header_row: int) -> Dict[str, Optional[int]]:
    kw = find_col_by_keywords
    return {
        "clinic":   kw(ws, header_row, ["診所名稱或機構代碼"]),
        "name":     kw(ws, header_row, ["姓名"]),
        "id":       kw(ws, header_row, ["身份證號碼"]),
        "bday":     kw(ws, header_row, ["生日"]),
        "age":      kw(ws, header_row, ["年齡"]),
        "tel":      kw(ws, header_row, ["電話"]),
        "cnt":      kw(ws, header_row, ["次數"]),
        "sex":      kw(ws, header_row, ["性別"]),
        "abc":      kw(ws, header_row, ["A/B/C"]),
        "dmk_code": kw(ws, header_row, ["DM/CKD/DKD"]),
        "ascvd":    kw(ws, header_row, ["ASCVD"]),
    }


def _detect_screening_cols(ws, header_row: int) -> Dict[str, Optional[int]]:
    kw = find_col_by_keywords
    return {
        "adult": kw(ws, header_row, ["成人", "保健"]),
        "pap":   kw(ws, header_row, ["子宮", "抹片"]),
        "flu":   (kw(ws, header_row, ["老人", "流感"])
                  or kw(ws, header_row, ["流感"])
                  or kw(ws, header_row, ["流感"])),
        "fit":   kw(ws, header_row, ["糞便", "潛血"]),
        "hep":   kw(ws, header_row, ["肝炎"]),
    }


def _detect_lab_cols(ws, header_row: int) -> Dict[str, Optional[int]]:
    kw = find_col_by_keywords
    return {
        "hba":     kw(ws, header_row, ["HbA1c"]),
        "hba_dt":  kw(ws, header_row, ["HbA1c", "日期"]),
        "ldl":     kw(ws, header_row, ["LDL"]),
        "ldl_dt":  kw(ws, header_row, ["LDL", "日期"]),
        "uacr":    kw(ws, header_row, ["UACR"]),
        "uacr_dt": kw(ws, header_row, ["UACR", "日期"]),
    }


def _detect_output_cols(ws, header_row: int) -> Dict[str, Optional[int]]:
    kw = find_col_by_keywords
    return {
        "disease_text": (kw(ws, header_row, ["DM/CKD/DKD/ASCVD"])
                         or kw(ws, header_row, ["疾病樣態", "ASCVD"])),
        "score": kw(ws, header_row, ["分數"]),
        "note":  kw(ws, header_row, ["備註"]),
    }


def _detect_followup_cols(ws, max_scan_row: int) -> Dict[str, Optional[int]]:
    kw = find_col_by_keywords_any_row
    return {
        "au": (kw(ws, max_scan_row, ["第一次提醒", "28天"])
               or kw(ws, max_scan_row, ["回診追蹤", "28天"])
               or kw(ws, max_scan_row, ["28天"])),
        "av": (kw(ws, max_scan_row, ["第二次提醒", "56天"])
               or kw(ws, max_scan_row, ["回診追蹤", "56天"])),
        "aw": (kw(ws, max_scan_row, ["第三次提醒", "84天"])
               or kw(ws, max_scan_row, ["回診追蹤", "84天"])),
    }


def detect_template_columns(ws, data_start: int) -> Dict[str, Optional[int]]:
    header_row = (
        find_header_row_contains(ws, ["診所名稱或機構代碼", "姓名", "身份證號碼"], 250) or 5
    )
    max_scan_row = data_start - 1

    cols: Dict[str, Optional[int]] = {}
    cols.update(_detect_basic_cols(ws, header_row))
    cols.update(_detect_screening_cols(ws, header_row))
    cols.update(_detect_lab_cols(ws, header_row))
    cols.update(_detect_output_cols(ws, header_row))
    cols.update(_detect_followup_cols(ws, max_scan_row))

    cols["breakdown"] = None
    if cols.get("score"):
        cand = int(cols["score"]) + 13  # type: ignore[arg-type]
        if 1 <= cand <= ws.max_column:
            cols["breakdown"] = cand

    _require_cols(cols, [
        "clinic", "name", "id", "bday", "tel", "abc",
        "dmk_code", "cnt", "ascvd", "sex",
        "hba", "hba_dt", "ldl", "ldl_dt", "uacr", "uacr_dt",
        "disease_text", "score", "note", "au", "av", "aw",
    ])
    return cols


def _require_cols(cols: Dict[str, Optional[int]], required: List[str]) -> None:
    missing = [k for k in required if not cols.get(k)]
    if missing:
        raise ValueError(f"新模板欄位找不到：{'、'.join(missing)}")


# ============================================================
# 業務邏輯：疾病樣態文字
# ============================================================
def disease_group_text(
    e_code: Optional[DiseaseCode], ascvd: AscvdCategory
) -> Optional[str]:
    has_dm    = e_code in (DiseaseCode.DM, DiseaseCode.DKD)
    has_ckd   = e_code in (DiseaseCode.CKD, DiseaseCode.DKD)
    has_ascvd = ascvd != AscvdCategory.NONE

    table = {
        (True,  False, False): "DM",
        (False, True,  False): "CKD",
        (True,  True,  False): "DM+CKD（DKD）",
        (False, False, True):  "ASCVD",
        (False, True,  True):  "CKD+ASCVD",
        (True,  False, True):  "DM+ASCVD",
        (True,  True,  True):  "DM+CKD+ASCVD",
    }
    return table.get((has_dm, has_ckd, has_ascvd))


# ============================================================
# 業務邏輯：備註（需要篩檢提示）
# ============================================================
def build_screening_note(
    *,
    age:      int,
    sex:      str,
    hep_dt:   Optional[datetime.date],
    fit_dt:   Optional[datetime.date],
    pap_dt:   Optional[datetime.date],
    adult_dt: Optional[datetime.date],
    flu_dt:   Optional[datetime.date],
    today:    datetime.date,
) -> str:
    msgs: List[str] = []

    if need_bc_hep(age) and hep_dt is None:
        msgs.append("今年需檢測BC肝")

    if need_fit(age):
        if fit_dt is None or add_months(fit_dt, 24) <= today:
            msgs.append("今年需檢測糞便")

    pap_interval = pap_check_interval_months(age, sex)
    if pap_interval:
        if pap_dt is None or add_months(pap_dt, pap_interval) <= today:
            msgs.append("今年需檢測子抹")

    adult_interval = adult_check_interval_months(age)
    if adult_interval:
        if adult_dt is None or add_months(adult_dt, adult_interval) <= today:
            msgs.append("今年需檢測成健")

    if need_flu(age):
        if flu_dt is None or flu_dt.year < today.year:
            msgs.append("今年需檢測老流")

    return "\n".join(msgs)


# ============================================================
# 業務邏輯：分數計算
# ============================================================
def _score_hba(hba_val: Any, hba_dt: Optional[datetime.date]) -> int:
    if not isinstance(hba_dt, datetime.date):
        return 0
    if hba_dt.year == Rules.YEAR_UNCONDITIONAL:
        return Rules.SCORE_HBA
    if hba_dt.year == Rules.YEAR_VALUE_CHECK:
        v = parse_float(hba_val)
        if v is not None and v <= Rules.HBA_PASS_2025:
            return Rules.SCORE_HBA
    return 0


def _score_ldl(ldl_val: Any, ldl_dt: Optional[datetime.date]) -> int:
    if not isinstance(ldl_dt, datetime.date):
        return 0
    if ldl_dt.year == Rules.YEAR_UNCONDITIONAL:
        return Rules.SCORE_LDL
    if ldl_dt.year == Rules.YEAR_VALUE_CHECK:
        v = parse_float(ldl_val)
        if v is not None and v <= Rules.LDL_PASS_2025:
            return Rules.SCORE_LDL
    return 0


def _score_screening(dt: Optional[datetime.date], pts: int) -> int:
    if isinstance(dt, datetime.date) and dt.year == Rules.SCREEN_YEAR:
        return pts
    return 0


def calc_score(
    *,
    e_code:   Optional[DiseaseCode],
    ascvd:    AscvdCategory,
    hba_val:  Any,
    hba_dt:   Optional[datetime.date],
    ldl_val:  Any,
    ldl_dt:   Optional[datetime.date],
    adult_dt: Optional[datetime.date],
    pap_dt:   Optional[datetime.date],
    flu_dt:   Optional[datetime.date],
    fit_dt:   Optional[datetime.date],
    hep_dt:   Optional[datetime.date],
    age:      int,
    sex:      str,
) -> Tuple[int, Optional[str]]:
    has_dm    = e_code in (DiseaseCode.DM, DiseaseCode.DKD)
    has_ckd   = e_code in (DiseaseCode.CKD, DiseaseCode.DKD)
    has_ascvd = ascvd != AscvdCategory.NONE

    total = 0
    parts: List[str] = []

    def add(pts: int, label: str) -> None:
        nonlocal total
        if pts:
            total += pts
            parts.append(f"{label} {pts}分")

    if has_dm:
        add(_score_hba(hba_val, hba_dt), "HbA1c")

    if has_dm or has_ckd or has_ascvd:
        add(_score_ldl(ldl_val, ldl_dt), "LDL")

    if age >= 0:
        if adult_check_interval_months(age) is not None:
            add(_score_screening(adult_dt, Rules.SCORE_ADULT), "成健")
        if pap_check_interval_months(age, sex) is not None:
            add(_score_screening(pap_dt, Rules.SCORE_PAP), "子抹")
        if need_flu(age):
            add(_score_screening(flu_dt, Rules.SCORE_FLU), "老流")
        if need_fit(age):
            add(_score_screening(fit_dt, Rules.SCORE_FIT), "糞便")
        if need_bc_hep(age):
            add(_score_screening(hep_dt, Rules.SCORE_HEP), "BC肝")

    breakdown = "\n".join(parts) if parts else None
    return total, breakdown


# ============================================================
# 業務邏輯：AU 第一次提醒（28天）
# ============================================================
def _is_valid_year(dt: Optional[datetime.date]) -> bool:
    return isinstance(dt, datetime.date) and dt.year in Rules.VALID_YEARS


def _au_item_line(item: str, last_dt: Optional[datetime.date], today: datetime.date) -> str:
    if not _is_valid_year(last_dt):
        return f"{item}:超過2年未檢查"
    due = last_dt + datetime.timedelta(days=Rules.AU_DAYS)  # type: ignore[operator]
    if due <= today:
        return f"{item}:可立刻通知回診"
    return f"{item}:{due.strftime('%Y-%m-%d')}需回診"


def build_au_note(
    *,
    e_code: Optional[DiseaseCode],
    ascvd:  AscvdCategory,
    hba_dt: Optional[datetime.date],
    ldl_dt: Optional[datetime.date],
    today:  datetime.date,
) -> str:
    has_dm    = e_code in (DiseaseCode.DM, DiseaseCode.DKD)
    has_ckd   = e_code in (DiseaseCode.CKD, DiseaseCode.DKD)
    has_ascvd = ascvd != AscvdCategory.NONE

    need_hba = has_dm
    need_ldl = has_dm or has_ckd or has_ascvd

    if not need_hba and not need_ldl:
        return "非疾病不需回診"

    any_valid = (
        (need_hba and _is_valid_year(hba_dt))
        or (need_ldl and _is_valid_year(ldl_dt))
    )
    if not any_valid:
        lines: List[str] = []
        if need_hba:
            lines.append("HbA1c:超過2年未檢查")
        if need_ldl:
            lines.append("LDL:超過2年未檢查")
        return "\n".join(lines)

    lines2: List[str] = []
    if need_hba:
        lines2.append(_au_item_line("HbA1c", hba_dt, today))
    if need_ldl:
        lines2.append(_au_item_line("LDL", ldl_dt, today))
    return "\n".join(lines2)


# ============================================================
# 業務邏輯：AV/AW 從 AU 文字延後
# ============================================================
_DATE_RE = re.compile(r"(\d{4}-\d{2}-\d{2})")


def _should_skip_followup(au_txt: str) -> bool:
    s = (au_txt or "").replace(" ", "")
    return (not s) or ("非疾病" in s)


def _followup_item_line(
    item: str, base_dt: Optional[datetime.date], today: datetime.date, offset: int
) -> str:
    base = base_dt if isinstance(base_dt, datetime.date) else today
    due  = base + datetime.timedelta(days=offset)
    return f"{item}:{due.strftime('%Y-%m-%d')}需回診"


def build_followup_note(
    au_txt: str, today: datetime.date, offset_days: int
) -> Optional[str]:
    if not au_txt:
        return None

    out: List[str] = []
    for ln in (line.strip() for line in au_txt.splitlines() if line.strip()):
        if ln.startswith("HbA1c:"):
            item = "HbA1c"
        elif ln.startswith("LDL:"):
            item = "LDL"
        else:
            continue

        if "可立刻通知" in ln:
            out.append(_followup_item_line(item, None, today, offset_days))
            continue

        m = _DATE_RE.search(ln)
        if m:
            try:
                base = datetime.datetime.strptime(m.group(1), "%Y-%m-%d").date()
                out.append(_followup_item_line(item, base, today, offset_days))
            except ValueError:
                pass

    return "\n".join(out) if out else None


# ============================================================
# KPI：HbA1c（AY8 / AZ7 / AZ8 / AZ9）
# ============================================================
def _write_ratio(ws, addr: str, numer: int, denom: int) -> None:
    ws[addr].value = (numer / denom) if denom > 0 else 0
    ws[addr].number_format = "0.00%"


def calc_hba_kpi_ay_az(
    ws, cols: Dict[str, Optional[int]], data_start: int, last_row: int
) -> None:
    """
    HbA1c KPI
    - AY8：<=7 比例（百分比）
    - AY9：分子/分母（字串）
    - AZ7：cutoff 顯示（字串，例如 <=7.30）
    - AZ8：>=73.8% 的比例（k/denom，百分比）
    - AZ9：分子/分母（字串，k/denom）
    """
    hba_values: List[float] = []

    for r in range(data_start, last_row + 1):
        e = parse_disease_code(ws.cell(r, cols["dmk_code"]).value)  # type: ignore[index]
        if e not in (DiseaseCode.DM, DiseaseCode.DKD):
            continue
        if not is_zero_like(ws.cell(r, cols["ascvd"]).value):  # type: ignore[index]
            continue
        hba_dt = parse_date(ws.cell(r, cols["hba_dt"]).value)  # type: ignore[index]
        if not (isinstance(hba_dt, datetime.date) and hba_dt.year in Rules.VALID_YEARS):
            continue
        hba = parse_float(ws.cell(r, cols["hba"]).value)  # type: ignore[index]
        if hba is None:
            continue
        hba_values.append(hba)

    denom = len(hba_values)

    # AY：<=7
    numer_ay = sum(1 for v in hba_values if v <= Rules.HBA_CONTROL_THRESHOLD)
    _write_ratio(ws, "AY8", numer_ay, denom)
    ws["AY9"].value = f"{numer_ay}/{denom}" if denom > 0 else "0/0"

    # AZ：>=73.8%（排序找第 k 個）
    if denom <= 0:
        ws["AZ7"].value = None
        _write_ratio(ws, "AZ8", 0, 0)
        ws["AZ9"].value = "0/0"
        print("AZ 分母=0，分子=0，比例=0.00%，切點=None")
        return

    hba_values.sort()
    k = int(math.ceil(Rules.HBA_TARGET_PERCENT * denom))
    k = max(1, min(k, denom))
    cutoff = hba_values[k - 1]

    # ✅ 你要的格式：<=7.30（文字）
    ws["AZ7"].value = f"<={cutoff:.2f}"

    # AZ8：比例（百分比）
    _write_ratio(ws, "AZ8", k, denom)

    # AZ9：分子/分母（文字）
    ws["AZ9"].value = f"{k}/{denom}"

    ratio = k / denom
    print(f"AZ 分母={denom}，分子={k}，比例={ratio*100:.2f}%，切點={cutoff:.2f}")
# ============================================================
# KPI：LDL（BB8/BC8）
# ============================================================
def calc_ldl_percentiles(
    ws, cols: Dict[str, Optional[int]], data_start: int, last_row: int
) -> None:
    """
    LDL KPI（比照 HbA1c 版面）
    - BB8：<=100 比例（百分比）
    - BB9：分子/分母（字串）
    - BC7：cutoff 顯示（字串，例如 <=109）
    - BC8：>=73.8% 的比例（k/denom，百分比）
    - BC9：>=73.8% 的分子/分母（字串）
    """
    ldl_values: List[float] = []
    th_control = Rules.LDL_KPI_THRESHOLDS[0]  # 100

    for r in range(data_start, last_row + 1):
        e = parse_disease_code(ws.cell(r, cols["dmk_code"]).value)  # type: ignore[index]
        if e not in (DiseaseCode.DM, DiseaseCode.CKD, DiseaseCode.DKD, DiseaseCode.OTHER):
            continue

        ascvd = parse_ascvd(ws.cell(r, cols["ascvd"]).value)  # type: ignore[index]
        # 排除：E=4 且 F=0
        if e == DiseaseCode.OTHER and ascvd == AscvdCategory.NONE:
            continue

        ldl_dt = parse_date(ws.cell(r, cols["ldl_dt"]).value)  # type: ignore[index]
        if not (isinstance(ldl_dt, datetime.date) and ldl_dt.year in Rules.VALID_YEARS):
            continue

        ldl = parse_float(ws.cell(r, cols["ldl"]).value)  # type: ignore[index]
        if ldl is None:
            continue

        ldl_values.append(ldl)

    denom = len(ldl_values)

    # BB：<=100
    numer_bb = sum(1 for v in ldl_values if v <= th_control)
    _write_ratio(ws, "BB8", numer_bb, denom)
    ws["BB9"].value = f"{numer_bb}/{denom}" if denom > 0 else "0/0"

    # BC：>=73.8%（排序找第 k 個）
    if denom <= 0:
        ws["BC7"].value = None
        _write_ratio(ws, "BC8", 0, 0)
        ws["BC9"].value = "0/0"
        print("BC 分母=0，分子=0，比例=0.00%，切點=None")
        return

    ldl_values.sort()
    k = int(math.ceil(Rules.HBA_TARGET_PERCENT * denom))  # 73.8%
    k = max(1, min(k, denom))
    cutoff = ldl_values[k - 1]

    # ✅ 你要的格式：<=109（文字；一位小數會自動處理成 109.0 / 109.5）
    # 若你想永遠不要小數，把 .1f 改成 .0f
    ws["BC7"].value = f"<={cutoff:.1f}".replace(".0", "")

    # BC8：比例（百分比）
    _write_ratio(ws, "BC8", k, denom)

    # BC9：分子/分母（文字）
    ws["BC9"].value = f"{k}/{denom}"

    ratio = k / denom
    print(f"BC 分母={denom}，分子={k}，比例={ratio*100:.2f}%，切點={cutoff:.0f}")
# ============================================================
# 格式工具
# ============================================================
def apply_full_grid(ws, max_row: int, max_col: int) -> None:
    """僅對有資料的範圍加細邊框（避免對空白列浪費時間）"""
    thin = Side(style="thin")
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(r, c).border = grid


def apply_date_format(
    ws, cols: Dict[str, Optional[int]], data_start: int, last_row: int
) -> None:
    date_col_keys = ["bday", "adult", "pap", "flu", "fit", "hep",
                     "hba_dt", "ldl_dt", "uacr_dt"]
    date_cols = [cols.get(k) for k in date_col_keys if cols.get(k)]
    for r in range(data_start, last_row + 1):
        for c in date_cols:
            cell = ws.cell(r, c)  # type: ignore[arg-type]
            if isinstance(cell.value, (datetime.date, datetime.datetime)):
                cell.number_format = "yyyy-mm-dd"


# ============================================================
# 資料填充子函數
# ============================================================
def _load_and_validate_source(source_path: str):
    """讀取並驗證原始檔的工作表完整性，回傳 workbook"""
    wb = openpyxl.load_workbook(source_path, data_only=True)
    need = ["會員名單", "ascvd", "HealthCase",
            "成人健檢", "子宮抹片", "老人流感", "糞便潛血", "肝炎篩檢"]
    missing = [s for s in need if s not in wb.sheetnames]
    if missing:
        raise ValueError(f"原始檔缺少工作表：{'、'.join(missing)}")
    return wb


def _fill_member_basic(
    ws,
    sh_member,
    cols:       Dict[str, Optional[int]],
    data_start: int,
    now:        datetime.date,
) -> Tuple[Dict[str, List[int]], Dict[int, MemberMeta], int]:
    """
    填入會員基本資料。
    回傳 (id_to_rows, meta, last_row)。

    ✅ 修正：
    - id_to_rows：同一身分證可能多列 → 用 list 收集所有輸出列號
    - meta：每一列都寫，供後續衍生計算
    - last_row：以實際寫入列數（out_r - 1）為準
    """
    MEMBER_HEADER_ROW = 5
    mmap = build_header_map(sh_member, MEMBER_HEADER_ROW)

    m_name = find_column_exact(mmap, ["會員姓名"])
    m_id   = find_column_exact(mmap, ["會員身份証", "會員身份證", "會員身分證"])
    m_bday = find_column_exact(mmap, ["會員生日"])
    m_tel  = find_column_exact(mmap, ["電話"])
    m_abc  = find_column_exact(mmap, ["會員別"])
    m_dmk  = find_column_exact(mmap, ["疾病樣態"])
    m_cnt  = find_column_exact(mmap, ["就診次數"])

    if any(v is None for v in [m_name, m_id, m_bday, m_tel, m_abc, m_dmk, m_cnt]):
        raise ValueError(
            "原始檔「會員名單」欄位不完整"
            "（會員姓名/會員身分證/會員生日/電話/會員別/疾病樣態/就診次數）"
        )

    clinic_val = normalize_text(sh_member["A1"].value)
    id_to_rows: Dict[str, List[int]] = {}
    meta: Dict[int, MemberMeta] = {}
    out_r = data_start
    wrote_any = False

    for r in range(MEMBER_HEADER_ROW + 1, sh_member.max_row + 1):
        sid  = sh_member.cell(r, m_id).value    # type: ignore[arg-type]
        name = sh_member.cell(r, m_name).value  # type: ignore[arg-type]
        if (sid in (None, "")) and (name in (None, "")):
            continue

        wrote_any = True
        sid_s  = str(sid).strip() if sid is not None else ""
        bday   = parse_date(sh_member.cell(r, m_bday).value)   # type: ignore[arg-type]
        e_code = parse_disease_code(sh_member.cell(r, m_dmk).value)  # type: ignore[arg-type]
        age    = calc_age(bday, now) if bday else -1
        sex    = infer_gender_from_id(sid)

        safe_set(ws, out_r, cols.get("clinic"),   clinic_val)
        safe_set(ws, out_r, cols.get("name"),     name)
        safe_set(ws, out_r, cols.get("id"),       sid)
        safe_set(ws, out_r, cols.get("bday"),     bday)
        safe_set(ws, out_r, cols.get("age"),      age if age >= 0 else None)
        safe_set(ws, out_r, cols.get("tel"),      sh_member.cell(r, m_tel).value)  # type: ignore[arg-type]
        safe_set(ws, out_r, cols.get("abc"),      sh_member.cell(r, m_abc).value)  # type: ignore[arg-type]
        safe_set(ws, out_r, cols.get("dmk_code"), e_code.value if e_code else None)
        safe_set(ws, out_r, cols.get("cnt"),      sh_member.cell(r, m_cnt).value)  # type: ignore[arg-type]
        safe_set(ws, out_r, cols.get("sex"),      sex)

        meta[out_r] = MemberMeta(row=out_r, bday=bday, age=age, e_code=e_code)

        if sid_s:
            id_to_rows.setdefault(sid_s, []).append(out_r)

        out_r += 1

    last_row = out_r - 1 if wrote_any else (data_start - 1)
    return id_to_rows, meta, last_row


def _fill_ascvd(
    ws,
    sh_ascvd:  Any,
    cols:      Dict[str, Optional[int]],
    id_to_rows: Dict[str, List[int]],
    meta:      Dict[int, MemberMeta],
) -> None:
    ASCVD_HEADER_ROW = 5
    amap  = build_header_map(sh_ascvd, ASCVD_HEADER_ROW)
    a_id  = find_column_exact(amap, ["ID", "id"])
    a_asc = find_column_exact(amap, ["ASCVD", "ascvd"])
    if a_id is None or a_asc is None:
        raise ValueError("原始檔「ascvd」找不到 ID / ASCVD 欄位（表頭列預期第5列）")

    for r in range(ASCVD_HEADER_ROW + 1, sh_ascvd.max_row + 1):
        pid = sh_ascvd.cell(r, a_id).value
        val = sh_ascvd.cell(r, a_asc).value
        if pid is None or str(pid).strip() == "":
            continue
        if val is None or str(val).strip() in ("", "0"):
            continue

        rows = id_to_rows.get(str(pid).strip())
        if not rows:
            continue

        for tr in rows:
            safe_set(ws, tr, cols.get("ascvd"), val)
            if tr in meta:
                meta[tr].ascvd = parse_ascvd(val)


def _fill_screening(
    ws,
    sheet,
    target_col: Optional[int],
    id_to_rows: Dict[str, List[int]],
) -> None:
    """填入單一篩檢日期（保留最新日期）；同一 ID 多列全部回填"""
    if not target_col:
        return
    hmap    = build_header_map(sheet, 1)
    sid_col = find_column_exact(hmap, ["ID", "身分證號", "身份證號"])
    dt_col  = find_column_exact(hmap, ["最後篩檢日期"])
    if sid_col is None or dt_col is None:
        raise ValueError(f"「{sheet.title}」找不到 ID 或 最後篩檢日期 欄位")

    for rr in range(2, sheet.max_row + 1):
        pid = sheet.cell(rr, sid_col).value
        dt  = parse_date(sheet.cell(rr, dt_col).value)
        if pid is None or str(pid).strip() == "" or dt is None:
            continue

        rows = id_to_rows.get(str(pid).strip())
        if not rows:
            continue

        for tr in rows:
            existing = parse_date(ws.cell(tr, target_col).value)
            if existing is None or dt > existing:
                ws.cell(tr, target_col).value = dt


def _fill_health_case(
    ws,
    sh_health: Any,
    cols:      Dict[str, Optional[int]],
    id_to_rows: Dict[str, List[int]],
) -> None:
    """回填 HealthCase（同一 ID 多列全部回填）"""
    hmap = build_header_map(sh_health, 1)
    field_aliases = {
        "hc_id":      ["家醫收案會員ID", "ID"],
        "hc_hba":     ["最近一次HbA1c檢查結果(%)"],
        "hc_hba_dt":  ["最近一次HbA1c檢查日期"],
        "hc_ldl":     ["最近一次LDL檢查結果(mg/dL)"],
        "hc_ldl_dt":  ["最近一次LDL檢查日期"],
        "hc_uacr":    ["最近一次UACR檢查結果(mg/gm)"],
        "hc_uacr_dt": ["最近一次UACR檢查日期"],
    }
    fc = {k: find_column_exact(hmap, v) for k, v in field_aliases.items()}
    if any(v is None for v in fc.values()):
        raise ValueError(
            "原始檔「HealthCase」欄位不完整"
            "（家醫收案會員ID / HbA1c結果+日期 / LDL結果+日期 / UACR結果+日期）"
        )

    for r in range(2, sh_health.max_row + 1):
        pid = sh_health.cell(r, fc["hc_id"]).value
        if pid is None or str(pid).strip() == "":
            continue

        rows = id_to_rows.get(str(pid).strip())
        if not rows:
            continue

        for tr in rows:
            for val_key, dt_key, col_val_key, col_dt_key in [
                ("hc_hba",  "hc_hba_dt",  "hba",  "hba_dt"),
                ("hc_ldl",  "hc_ldl_dt",  "ldl",  "ldl_dt"),
                ("hc_uacr", "hc_uacr_dt", "uacr", "uacr_dt"),
            ]:
                v  = parse_float(sh_health.cell(r, fc[val_key]).value)
                dt = parse_date(sh_health.cell(r, fc[dt_key]).value)
                if v is not None and v != 0:
                    safe_set(ws, tr, cols.get(col_val_key), v)
                if dt:
                    safe_set(ws, tr, cols.get(col_dt_key), dt)


def _compute_all_derived(
    ws,
    cols:       Dict[str, Optional[int]],
    meta:       Dict[int, MemberMeta],
    data_start: int,
    last_row:   int,
    now:        datetime.date,
) -> None:
    """逐列計算疾病樣態文字、備註、分數、AU/AV/AW"""
    for rr in range(data_start, last_row + 1):
        e_code = parse_disease_code(ws.cell(rr, cols["dmk_code"]).value)  # type: ignore[index]
        ascvd  = parse_ascvd(ws.cell(rr, cols["ascvd"]).value)            # type: ignore[index]

        m   = meta.get(rr, MemberMeta(row=rr))
        age = m.age if isinstance(m.age, int) else -1
        sex = normalize_text(ws.cell(rr, cols["sex"]).value)              # type: ignore[index]

        def _get_dt(key: str) -> Optional[datetime.date]:
            c = cols.get(key)
            return parse_date(ws.cell(rr, c).value) if c else None  # type: ignore[arg-type]

        adult_dt = _get_dt("adult")
        pap_dt   = _get_dt("pap")
        flu_dt   = _get_dt("flu")
        fit_dt   = _get_dt("fit")
        hep_dt   = _get_dt("hep")
        hba_dt   = _get_dt("hba_dt")
        ldl_dt   = _get_dt("ldl_dt")

        hba_val = ws.cell(rr, cols["hba"]).value  # type: ignore[index]
        ldl_val = ws.cell(rr, cols["ldl"]).value  # type: ignore[index]

        safe_set(ws, rr, cols.get("disease_text"),
                 disease_group_text(e_code, ascvd))

        note = build_screening_note(
            age=max(age, 0), sex=sex,
            hep_dt=hep_dt, fit_dt=fit_dt, pap_dt=pap_dt,
            adult_dt=adult_dt, flu_dt=flu_dt, today=now,
        )
        safe_set(ws, rr, cols.get("note"), note or None)

        score, breakdown = calc_score(
            e_code=e_code, ascvd=ascvd,
            hba_val=hba_val, hba_dt=hba_dt,
            ldl_val=ldl_val, ldl_dt=ldl_dt,
            adult_dt=adult_dt, pap_dt=pap_dt,
            flu_dt=flu_dt, fit_dt=fit_dt, hep_dt=hep_dt,
            age=age, sex=sex,
        )
        safe_set(ws, rr, cols.get("score"),     score)
        safe_set(ws, rr, cols.get("breakdown"), breakdown)

        au_txt = build_au_note(
            e_code=e_code, ascvd=ascvd,
            hba_dt=hba_dt, ldl_dt=ldl_dt, today=now,
        )
        safe_set(ws, rr, cols.get("au"), au_txt or None)

        if au_txt and not _should_skip_followup(au_txt):
            safe_set(ws, rr, cols.get("av"),
                     build_followup_note(au_txt, now, Rules.AV_OFFSET_DAYS))
            safe_set(ws, rr, cols.get("aw"),
                     build_followup_note(au_txt, now, Rules.AW_OFFSET_DAYS))
        else:
            safe_set(ws, rr, cols.get("av"), None)
            safe_set(ws, rr, cols.get("aw"), None)


# ============================================================
# 清空資料列
# ============================================================
def _clear_data_rows(
    ws, data_start: int, max_row: int, cols: Dict[str, Optional[int]]
) -> None:
    clear_keys = [
        "clinic", "name", "id", "bday", "age", "tel", "cnt", "sex", "abc",
        "dmk_code", "ascvd",
        "adult", "pap", "flu", "fit", "hep",
        "hba", "hba_dt", "ldl", "ldl_dt", "uacr", "uacr_dt",
        "disease_text", "score", "breakdown", "note",
        "au", "av", "aw",
    ]
    col_ids = [cols[k] for k in clear_keys if cols.get(k)]
    for r in range(data_start, max_row + 1):
        for c in col_ids:
            ws.cell(r, c).value = None  # type: ignore[arg-type]


# ============================================================
# 主流程
# ============================================================
def process_excel(source_path: str, template_path: str) -> str:
    # 1. 讀取並驗證原始檔
    wb_src    = _load_and_validate_source(source_path)
    sh_member = wb_src["會員名單"]
    sh_ascvd  = wb_src["ascvd"]
    sh_health = wb_src["HealthCase"]
    screening_sheets = {
        "adult": wb_src["成人健檢"],
        "pap":   wb_src["子宮抹片"],
        "flu":   wb_src["老人流感"],
        "fit":   wb_src["糞便潛血"],
        "hep":   wb_src["肝炎篩檢"],
    }

    # 2. 讀模板（✅ 保留所有分頁；目前只寫入「會員指標」）
    wb_tpl = openpyxl.load_workbook(template_path)
    if Rules.SHEET_TARGET not in wb_tpl.sheetnames:
        raise ValueError(f"模板檔缺少工作表：{Rules.SHEET_TARGET}")
    ws = wb_tpl[Rules.SHEET_TARGET]

    DATA_START = Rules.DATA_START_ROW

    # 3. 偵測欄位
    cols = detect_template_columns(ws, DATA_START)

    # 4. 清空舊資料列
    _clear_data_rows(ws, DATA_START, ws.max_row, cols)

    # 5. 取得台灣時間（今天）
    tz_tw = datetime.timezone(datetime.timedelta(hours=8))
    now   = datetime.datetime.now(tz_tw).date()

    # 6. 填入基本資料
    id_to_rows, meta, last_row = _fill_member_basic(
        ws, sh_member, cols, DATA_START, now
    )

    # 若完全沒有資料列，仍可輸出空模板
    if last_row < DATA_START:
        base_dir = os.path.dirname(os.path.abspath(source_path))
        ts = datetime.datetime.now(tz_tw).strftime("%m%d_%H%M%S")
        out_path = os.path.join(base_dir, f"選會員{ts}.xlsx")
        wb_tpl.save(out_path)
        return out_path

    # 7. 回填 ASCVD（同一 ID 多列全部回填）
    _fill_ascvd(ws, sh_ascvd, cols, id_to_rows, meta)

    # 8. 回填 5 大篩檢日期（同一 ID 多列全部回填）
    for key, sheet in screening_sheets.items():
        _fill_screening(ws, sheet, cols.get(key), id_to_rows)

    # 9. 回填 HealthCase 檢驗值（同一 ID 多列全部回填）
    _fill_health_case(ws, sh_health, cols, id_to_rows)

    # 10. 計算衍生欄位（疾病樣態/備註/分數/AU/AV/AW）
    _compute_all_derived(ws, cols, meta, DATA_START, last_row, now)

    # 11. 日期格式
    apply_date_format(ws, cols, DATA_START, last_row)

    # 12. KPI（HbA1c / LDL）
    calc_hba_kpi_ay_az(ws, cols, DATA_START, last_row)
    calc_ldl_percentiles(ws, cols, DATA_START, last_row)

    # 13. 全表邊框（只對有資料的範圍）
    apply_full_grid(ws, max_row=last_row, max_col=ws.max_column)

    # 14. 輸出
    base_dir = os.path.dirname(os.path.abspath(source_path))
    ts       = datetime.datetime.now(tz_tw).strftime("%m%d_%H%M%S")
    out_path = os.path.join(base_dir, f"選會員{ts}.xlsx")
    wb_tpl.save(out_path)
    return out_path


# ============================================================
# 進入點
# ============================================================
def main() -> None:
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()

    src = filedialog.askopenfilename(
        title="選擇原始 Excel 檔案（沒有會員指標的那份）",
        filetypes=[("Excel files", "*.xlsx")],
    )
    if not src:
        return

    script_dir = os.path.dirname(os.path.abspath(__file__))
    template   = os.path.join(script_dir, Rules.TEMPLATE_NAME)

    if not os.path.exists(template):
        messagebox.showerror(
            "錯誤",
            f"找不到模板檔：\n{template}\n\n"
            f"請把模板放到此 .py 同資料夾，檔名需為：\n{Rules.TEMPLATE_NAME}",
        )
        return

    try:
        out = process_excel(src, template)
        messagebox.showinfo("完成", f"已輸出：\n{out}")
        open_file_cross_platform(out)
    except (ValueError, KeyError, OSError, InvalidFileException) as e:
        messagebox.showerror("錯誤", str(e))
    except Exception as e:
        messagebox.showerror("錯誤", f"未預期錯誤：{e}")


if __name__ == "__main__":
    main()