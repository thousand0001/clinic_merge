# -*- coding: utf-8 -*-
"""
選會員產檔工具（結構化單檔版）— V5.10（0307樣板）

功能概要
- 讀取原始檔（會員名單 / ascvd / HealthCase / 5 大篩檢）
- 套用模板（保留所有分頁；目前只寫入「會員指標」分頁），寫入資料列
- 計算：疾病樣態文字、備註、分數/分數說明、AU/AV/AW 文字提醒
- 新增：
  - AK：HbA1c 控制打勾（✔）
  - AX：漏檢項目（HbA1c / LDL 漏檢）
- KPI
  - HbA1c：AY8（<=7 的比例）
  - HbA1c：AZ9（>=73.8% 所需人數），AZ7（該人數對應的 HbA1c 切點值）
  - LDL：BB8/BC8（<=100 / <=110）
- 全表格加網格線、輸出到原始檔同資料夾並自動開檔

注意
- Python 3.9 相容
- 不拆模組：同一支檔案即可維護
"""

from __future__ import annotations

import datetime
import math
import os
import re
import subprocess
import sys
from copy import copy
from dataclasses import dataclass
from enum import Enum, auto
from typing import Dict, List, Optional, Any, Tuple

import openpyxl
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import column_index_from_string
from openpyxl.utils.exceptions import InvalidFileException


# ============================================================
# 業務規則常數（每年調整這裡即可）
# ============================================================
class Rules:
    # 檢驗日期：哪些年份視為「有效」
    VALID_YEARS: Tuple[int, ...] = (2025, 2026)

    # 篩檢計分：哪一年的篩檢才給分
    SCREEN_YEAR: int = 2026

    # 2026 年：只要有檢查就給滿分（無條件給分）
    # 2025 年：需達標才給分
    YEAR_UNCONDITIONAL: int = 2026
    YEAR_VALUE_CHECK: int = 2025

    # 各項分數
    SCORE_HBA: int = 5
    SCORE_LDL: int = 5
    SCORE_ADULT: int = 6
    SCORE_PAP: int = 6
    SCORE_FLU: int = 4
    SCORE_FIT: int = 6
    SCORE_HEP: int = 6

    # 達標值（2025年）
    HBA_PASS_2025: float = 7.0
    LDL_PASS_2025: float = 120.0

    # KPI（HbA1c）
    HBA_CONTROL_THRESHOLD: float = 7.0  # AY8：HbA1c <= 7 的比例
    HBA_TARGET_PERCENT: float = 0.738   # AZ：要找 >=73.8% 的切點

    # KPI（LDL）
    LDL_KPI_THRESHOLDS: Tuple[float, float] = (100.0, 110.0)  # BB8/BC8
    LDL_TARGET_PERCENT: float = 0.738   # BC：要找 >=73.8% 的切點（與 HbA1c 目前相同，但獨立維護）

    # 回診追蹤
    AU_DAYS: int = 28
    AV_OFFSET_DAYS: int = 28
    AW_OFFSET_DAYS: int = 56

    # 模板設定（0307）
    TEMPLATE_NAME: str = "選會員樣板0307.xlsx"
    SHEET_TARGET: str = "會員指標"
    DATA_START_ROW: int = 10

    # 月份申請統計輸出欄位
    COL_114_COUNT: str = "M"
    COL_115_COUNT: str = "N"
    COL_114_AMOUNT: str = "R"
    COL_115_AMOUNT: str = "S"

    # KPI 分子打勾欄位（明細列）
    COL_AY_MARK: str = "AY"
    COL_AZ_MARK: str = "AZ"
    COL_BB_MARK: str = "BB"
    COL_BC_MARK: str = "BC"


# ============================================================
# Enum / Dataclass
# ============================================================
class DiseaseCode(Enum):
    DM = 1
    CKD = 2
    DKD = 3   # DM + CKD
    OTHER = 4 # 非 DM/CKD，可能有 ASCVD


class AscvdCategory(Enum):
    NONE = auto()  # 無 ASCVD
    A = auto()     # 極高風險
    B = auto()     # 非常高風險


@dataclass
class MemberMeta:
    """每位會員的衍生資訊，供後續計算用"""
    row: int
    bday: Optional[datetime.date] = None
    age: int = -1
    e_code: Optional[DiseaseCode] = None
    ascvd: AscvdCategory = AscvdCategory.NONE


# ============================================================
# 工具函數
# ============================================================
def open_file_cross_platform(path: str) -> None:
    """跨平台開啟檔案（盡量不拋錯）"""
    try:
        if sys.platform.startswith("darwin"):
            subprocess.call(("open", path))
        elif os.name == "nt":
            os.startfile(path)  # type: ignore[attr-defined]
        else:
            subprocess.call(("xdg-open", path))
    except Exception:
        pass


def normalize_text(v: Any) -> str:
    if v is None:
        return ""
    return str(v).replace("\t", "").replace("　", "").strip().lstrip("'")


def normalize_header(v: Any) -> str:
    s = normalize_text(v)
    return s.replace(" ", "").replace("\n", "")


def clean_spaces(v: Any) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", "", str(v))


def safe_set(ws, row: int, col: Optional[int], value: Any) -> None:
    if col:
        ws.cell(row, col).value = value


def calc_age(bday: Optional[datetime.date], ref: datetime.date) -> int:
    if not isinstance(bday, datetime.date):
        return -1
    y = ref.year - bday.year
    if (ref.month, ref.day) < (bday.month, bday.day):
        y -= 1
    return y


def add_months(d: datetime.date, months: int) -> datetime.date:
    y = d.year + (d.month - 1 + months) // 12
    m = (d.month - 1 + months) % 12 + 1
    is_leap = (y % 4 == 0 and (y % 100 != 0 or y % 400 == 0))
    mdays = [31, 29 if is_leap else 28, 31, 30, 31, 30,
             31, 31, 30, 31, 30, 31][m - 1]
    return datetime.date(y, m, min(d.day, mdays))


def infer_gender_from_id(id_value: Any) -> str:
    """由台灣身分證/居留證第2碼推性別（1/8/A/C→男，2/9/B/D→女）"""
    if not id_value:
        return ""
    s = str(id_value).strip().upper()
    if len(s) < 2:
        return ""
    g = s[1]
    if g in ("1", "8", "A", "C"):
        return "男"
    if g in ("2", "9", "B", "D"):
        return "女"
    return ""


# ============================================================
# 解析函數
# ============================================================
def parse_date(value: Any) -> Optional[datetime.date]:
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value

    s = str(value).strip()
    if not s or s in ("-", "—", "–"):
        return None

    patterns = [
        (r"^(\d{4})[/-](\d{1,2})[/-](\d{1,2})$", False),
        (r"^(\d{2,3})(\d{2})(\d{2})$", True),
        (r"^(\d{2,3})[/-](\d{1,2})[/-](\d{1,2})$", True),
    ]
    for pat, is_roc in patterns:
        m = re.match(pat, s)
        if m:
            a, b, c = map(int, m.groups())
            year = a + 1911 if is_roc else a
            try:
                return datetime.date(year, b, c)
            except ValueError:
                return None
    return None


def parse_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    s = str(value).strip().lstrip("'").replace(",", "")
    if not s or s in ("-", "—", "–"):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def parse_disease_code(v: Any) -> Optional[DiseaseCode]:
    s = clean_spaces(v)
    if not s:
        return None
    if s.isdigit():
        mapping = {
            1: DiseaseCode.DM,
            2: DiseaseCode.CKD,
            3: DiseaseCode.DKD,
            4: DiseaseCode.OTHER,
        }
        return mapping.get(int(s))
    text_map = {
        "DM":     DiseaseCode.DM,
        "CKD":    DiseaseCode.CKD,
        "DKD":    DiseaseCode.DKD,
        "DM+CKD": DiseaseCode.DKD,
        "DMCKD":  DiseaseCode.DKD,
    }
    return text_map.get(s.upper())


def parse_ascvd(v: Any) -> AscvdCategory:
    if v is None:
        return AscvdCategory.NONE
    s = clean_spaces(v).lower()
    if not s or s == "0":
        return AscvdCategory.NONE
    if s in ("a", "1") or "極高" in s:
        return AscvdCategory.A
    if s in ("b", "2") or "非常高" in s:
        return AscvdCategory.B
    if "a" in s:
        return AscvdCategory.A
    if "b" in s:
        return AscvdCategory.B
    return AscvdCategory.NONE


def is_zero_like(v: Any) -> bool:
    """把 None / 空白 / '0' 視為零值"""
    if v is None:
        return True
    return str(v).strip() in ("", "0", "0.0")


# ============================================================
# 月份申請統計（114xx / 115xx）
# ============================================================
def _find_header_row_contains_any(sheet, alias_groups: List[List[str]], search_rows: int = 30) -> Optional[int]:
    # 正規化 alias 一次，避免每列每組重複計算
    norm_groups = [[a.replace(" ", "").replace("\n", "") for a in grp] for grp in alias_groups]
    for r in range(1, min(search_rows, sheet.max_row) + 1):
        row_headers = {normalize_header(sheet.cell(r, c).value) for c in range(1, sheet.max_column + 1)}
        if all(any(a in row_headers for a in grp) for grp in norm_groups):
            return r
    return None


def _sheet_year_bucket(title: str) -> Optional[int]:
    s = str(title).strip()
    if not re.fullmatch(r"\d{5}", s):
        return None
    if s.startswith("114"):
        return 114
    if s.startswith("115"):
        return 115
    return None


def collect_monthly_claim_summaries(wb_src) -> Dict[str, Dict[str, float]]:
    """
    掃描 11401~11412、11501~11512 這類月份分頁，
    依 ID 彙總：
      - 114 D欄件數 -> M
      - 115 D欄件數 -> N
      - 114 E欄金額 -> R
      - 115 E欄金額 -> S
    若某 ID 完全沒有資料，後續保持空白，不填 0。
    """
    out: Dict[str, Dict[str, float]] = {}
    id_aliases = ["ID", "身分證號", "身分證號碼", "身份證號", "身份證號碼"]

    for sheet_name in wb_src.sheetnames:
        year_bucket = _sheet_year_bucket(sheet_name)
        if year_bucket not in (114, 115):
            continue

        sh = wb_src[sheet_name]
        header_row = _find_header_row_contains_any(
            sh,
            [id_aliases, ["件數"], ["申請金額"]],
            search_rows=30,
        )
        if header_row is None:
            continue

        hmap = build_header_map(sh, header_row)
        id_col = find_column_exact(hmap, id_aliases)
        if id_col is None:
            continue

        count_col = 4  # D 欄：件數
        amount_col = 5 # E 欄：申請金額

        for r in range(header_row + 1, sh.max_row + 1):
            pid_raw = sh.cell(r, id_col).value
            pid = normalize_text(pid_raw).upper()
            if not pid:
                continue

            cnt = parse_float(sh.cell(r, count_col).value)
            amt = parse_float(sh.cell(r, amount_col).value)
            if cnt is None and amt is None:
                continue

            bucket = out.setdefault(pid, {
                "114_cnt": 0.0,
                "115_cnt": 0.0,
                "114_amt": 0.0,
                "115_amt": 0.0,
            })

            prefix = str(year_bucket)  # "114" 或 "115"
            if cnt is not None:
                bucket[f"{prefix}_cnt"] += cnt
            if amt is not None:
                bucket[f"{prefix}_amt"] += amt

    return out


def _to_excel_number(v: Optional[float]) -> Optional[float]:
    if v is None:
        return None
    if abs(v - round(v)) < 1e-9:
        return int(round(v))
    return round(v, 2)


def fill_monthly_claim_summary_columns(
    ws,
    data_start: int,
    last_row: int,
    cols: Dict[str, Optional[int]],
    claim_sums: Dict[str, Dict[str, float]],
) -> None:
    col_m = cols.get("m_count_114")
    col_n = cols.get("n_count_115")
    col_r = cols.get("r_amount_114")
    col_s = cols.get("s_amount_115")

    if not all([col_m, col_n, col_r, col_s, cols.get("id")]):
        raise ValueError("模板找不到 M/N/R/S 或 ID 欄位，無法填入月份申請統計")

    for rr in range(data_start, last_row + 1):
        pid = normalize_text(ws.cell(rr, cols["id"]).value).upper()  # type: ignore[index]
        data = claim_sums.get(pid)

        if not data:
            ws.cell(rr, col_m).value = None
            ws.cell(rr, col_n).value = None
            ws.cell(rr, col_r).value = None
            ws.cell(rr, col_s).value = None
            continue

        v114c = data.get("114_cnt", 0.0)
        v115c = data.get("115_cnt", 0.0)
        v114a = data.get("114_amt", 0.0)
        v115a = data.get("115_amt", 0.0)

        ws.cell(rr, col_m).value = _to_excel_number(v114c) if v114c != 0 else None
        ws.cell(rr, col_n).value = _to_excel_number(v115c) if v115c != 0 else None
        ws.cell(rr, col_r).value = _to_excel_number(v114a) if v114a != 0 else None
        ws.cell(rr, col_s).value = _to_excel_number(v115a) if v115a != 0 else None


# ============================================================
# 篩檢需求判斷
# ============================================================
def adult_check_interval_months(age: int) -> Optional[int]:
    if 30 <= age <= 39:
        return 60
    if 40 <= age <= 64:
        return 36
    if age >= 65:
        return 12
    return None


def pap_check_interval_months(age: int, sex: str) -> Optional[int]:
    if sex != "女":
        return None
    if 25 <= age <= 29:
        return 36
    if age >= 30:
        return 12
    return None


def need_flu(age: int) -> bool:
    return age >= 65


def need_fit(age: int) -> bool:
    return 45 <= age <= 75


def need_bc_hep(age: int) -> bool:
    return 45 <= age <= 80


# ============================================================
# 欄位偵測
# ============================================================
def build_header_map(sheet, header_row: int) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for c in range(1, sheet.max_column + 1):
        k = normalize_header(sheet.cell(header_row, c).value)
        if k:
            m[k] = c
    return m


def find_column_exact(hmap: Dict[str, int], aliases: List[str]) -> Optional[int]:
    for a in aliases:
        if a in hmap:
            return hmap[a]
    return None


def _get_merged_header_text(ws, header_row: int, c: int) -> str:
    upper = normalize_header(ws.cell(header_row - 1, c).value) if header_row > 1 else ""
    lower = normalize_header(ws.cell(header_row, c).value)
    return (upper + " " + lower).strip()


def find_col_by_keywords(ws, header_row: int, keywords: List[str]) -> Optional[int]:
    """模糊比對：合併上下兩列表頭，同時包含所有 keywords"""
    keys = [k.replace(" ", "").replace("\n", "") for k in keywords if k]
    for c in range(1, ws.max_column + 1):
        h = _get_merged_header_text(ws, header_row, c).replace(" ", "")
        if h and all(k in h for k in keys):
            return c
    return None


def find_col_by_keywords_any_row(ws, max_row: int, keywords: List[str]) -> Optional[int]:
    """跨多列找欄位：同欄所有儲存格合併後包含全部 keywords"""
    keys = [k.replace(" ", "").replace("\n", "") for k in keywords if k]
    for c in range(1, ws.max_column + 1):
        blob = "".join(
            normalize_header(ws.cell(r, c).value)
            for r in range(1, max_row + 1)
        )
        if blob and all(k in blob for k in keys):
            return c
    return None


def find_header_row_contains(sheet, must_have: List[str], search_rows: int = 250) -> Optional[int]:
    must = [m.replace(" ", "").replace("\n", "") for m in must_have]
    for r in range(1, min(search_rows, sheet.max_row) + 1):
        found = set()
        for c in range(1, sheet.max_column + 1):
            v = normalize_header(sheet.cell(r, c).value)
            if v in must:
                found.add(v)
        if all(m in found for m in must):
            return r
    return None


def _detect_basic_cols(ws, header_row: int) -> Dict[str, Optional[int]]:
    kw = find_col_by_keywords
    return {
        "clinic":   kw(ws, header_row, ["診所名稱或機構代碼"]),
        "name":     kw(ws, header_row, ["姓名"]),
        "id":       kw(ws, header_row, ["身份證號碼"]),
        "bday":     kw(ws, header_row, ["生日"]),
        "age":      kw(ws, header_row, ["年齡"]),
        "tel":      kw(ws, header_row, ["電話"]),
        "cnt":      kw(ws, header_row, ["次數"]),
        "sex":      kw(ws, header_row, ["性別"]),
        "abc":      kw(ws, header_row, ["A/B/C"]),
        "dmk_code": kw(ws, header_row, ["DM/CKD/DKD"]),
        "ascvd":    kw(ws, header_row, ["ASCVD"]),
    }


def _detect_screening_cols(ws, header_row: int) -> Dict[str, Optional[int]]:
    kw = find_col_by_keywords
    return {
        "adult": kw(ws, header_row, ["成人", "保健"]),
        "pap":   kw(ws, header_row, ["子宮", "抹片"]),
        "flu":   (kw(ws, header_row, ["老人", "流感"])
                  or kw(ws, header_row, ["流感"])),
        "fit":   kw(ws, header_row, ["糞便", "潛血"]),
        "hep":   kw(ws, header_row, ["肝炎"]),
    }


def _detect_lab_cols(ws, header_row: int) -> Dict[str, Optional[int]]:
    kw = find_col_by_keywords
    return {
        "hba":     kw(ws, header_row, ["HbA1c"]),
        "hba_dt":  kw(ws, header_row, ["HbA1c", "日期"]),
        "ldl":     kw(ws, header_row, ["LDL"]),
        "ldl_dt":  kw(ws, header_row, ["LDL", "日期"]),
        "uacr":    kw(ws, header_row, ["UACR"]),
        "uacr_dt": kw(ws, header_row, ["UACR", "日期"]),
    }


def _detect_output_cols(ws, header_row: int, max_scan_row: int) -> Dict[str, Optional[int]]:
    kw = find_col_by_keywords
    kw_any = find_col_by_keywords_any_row
    return {
        "disease_text": (kw(ws, header_row, ["DM/CKD/DKD/ASCVD"])
                         or kw(ws, header_row, ["疾病樣態", "ASCVD"])),
        "score": kw(ws, header_row, ["分數"]),
        "breakdown": (kw(ws, header_row, ["分數說明"])
                        or kw_any(ws, max_scan_row, ["分數說明"])
                        or column_index_from_string("BD")),
        "note":  kw(ws, header_row, ["備註"]),
        # AK：合併大區塊，用 1~data_start-1 掃描
        "ak":    (kw_any(ws, max_scan_row, ["打勾"])
                  or kw_any(ws, max_scan_row, ["HbA1c", "打勾"])
                  or kw(ws, header_row, ["打勾"])),
        # AX：漏檢項目
        "ax":    (kw_any(ws, max_scan_row, ["漏檢項目"])
                  or kw_any(ws, max_scan_row, ["漏檢"])
                  or kw(ws, header_row, ["漏檢項目"])
                  or kw(ws, header_row, ["漏檢"])),
    }


def _detect_followup_cols(ws, max_scan_row: int) -> Dict[str, Optional[int]]:
    kw = find_col_by_keywords_any_row
    return {
        "au": (kw(ws, max_scan_row, ["第一次提醒", "28天"])
               or kw(ws, max_scan_row, ["回診追蹤", "28天"])
               or kw(ws, max_scan_row, ["28天"])),
        "av": (kw(ws, max_scan_row, ["第二次提醒", "56天"])
               or kw(ws, max_scan_row, ["回診追蹤", "56天"])),
        "aw": (kw(ws, max_scan_row, ["第三次提醒", "84天"])
               or kw(ws, max_scan_row, ["回診追蹤", "84天"])),
    }


def detect_template_columns(ws, data_start: int) -> Dict[str, Optional[int]]:
    header_row = (
        find_header_row_contains(ws, ["診所名稱或機構代碼", "姓名", "身份證號碼"], 250) or 5
    )
    max_scan_row = data_start - 1

    cols: Dict[str, Optional[int]] = {}
    cols.update(_detect_basic_cols(ws, header_row))
    cols.update(_detect_screening_cols(ws, header_row))
    cols.update(_detect_lab_cols(ws, header_row))
    cols.update(_detect_output_cols(ws, header_row, max_scan_row))
    cols.update(_detect_followup_cols(ws, max_scan_row))

    cols["m_count_114"] = column_index_from_string(Rules.COL_114_COUNT)
    cols["n_count_115"] = column_index_from_string(Rules.COL_115_COUNT)
    cols["r_amount_114"] = column_index_from_string(Rules.COL_114_AMOUNT)
    cols["s_amount_115"] = column_index_from_string(Rules.COL_115_AMOUNT)
    cols["ay_mark"] = column_index_from_string(Rules.COL_AY_MARK)
    cols["az_mark"] = column_index_from_string(Rules.COL_AZ_MARK)
    cols["bb_mark"] = column_index_from_string(Rules.COL_BB_MARK)
    cols["bc_mark"] = column_index_from_string(Rules.COL_BC_MARK)

    _require_cols(cols, [
        "clinic", "name", "id", "bday", "tel", "abc",
        "dmk_code", "cnt", "ascvd", "sex",
        "hba", "hba_dt", "ldl", "ldl_dt", "uacr", "uacr_dt",
        "disease_text", "score", "breakdown", "note",
        "au", "av", "aw",
        "ak", "ax",
        "m_count_114", "n_count_115", "r_amount_114", "s_amount_115",
        "ay_mark", "az_mark", "bb_mark", "bc_mark",
    ])
    return cols


def _require_cols(cols: Dict[str, Optional[int]], required: List[str]) -> None:
    missing = [k for k in required if not cols.get(k)]
    if missing:
        raise ValueError(f"新模板欄位找不到：{'、'.join(missing)}")


# ============================================================
# 業務邏輯：疾病樣態文字（✅ DM+CKD 後面不加 (DKD)）
# ============================================================
def disease_group_text(
    e_code: Optional[DiseaseCode], ascvd: AscvdCategory
) -> Optional[str]:
    has_dm    = e_code in (DiseaseCode.DM, DiseaseCode.DKD)
    has_ckd   = e_code in (DiseaseCode.CKD, DiseaseCode.DKD)
    has_ascvd = ascvd != AscvdCategory.NONE

    table = {
        (True,  False, False): "DM",
        (False, True,  False): "CKD",
        (True,  True,  False): "DM+CKD",
        (False, False, True):  "ASCVD",
        (False, True,  True):  "CKD+ASCVD",
        (True,  False, True):  "DM+ASCVD",
        (True,  True,  True):  "DM+CKD+ASCVD",
    }
    return table.get((has_dm, has_ckd, has_ascvd))


# ============================================================
# 業務邏輯：AK（HbA1c 打勾 ✔）
# ============================================================
def should_check_ak(
    *,
    e_code: Optional[DiseaseCode],
    ascvd: AscvdCategory,
    age: int,
    hba_val: Any,
) -> bool:
    """
    AK 打勾規則（✔）：
    - DM：HbA1c < 8
    - DM+CKD / DM+ASCVD / DM+CKD+ASCVD：
        - 年齡 >= 80：HbA1c < 8
        - 其餘：HbA1c < 7
    其他疾病：不勾
    """
    has_dm  = e_code in (DiseaseCode.DM, DiseaseCode.DKD)
    has_ckd = e_code in (DiseaseCode.CKD, DiseaseCode.DKD)
    has_asc = ascvd != AscvdCategory.NONE

    if not has_dm:
        return False

    v = parse_float(hba_val)
    if v is None:
        return False

    # DM only
    if (not has_ckd) and (not has_asc):
        return v < 8.0

    # DM+CKD / DM+ASCVD / DM+CKD+ASCVD
    if age < 0:
        return False

    if age >= 80:
        return v < 8.0
    return v < 7.0


# ============================================================
# 業務邏輯：AX（漏檢項目）
# ============================================================
def _ascvd_token(v: Any) -> str:
    """AX 用：ASCVD 欄只接受 0/a/b（0 也算有值）"""
    s = clean_spaces(v).lower()
    if s in ("0", "a", "b"):
        return s
    return ""


def build_ax_leak_item(
    *,
    d_code: Optional[DiseaseCode],  # D欄：1/2/3/4
    ascvd_raw: Any,                 # E欄：0/a/b
    hba_dt: Optional[datetime.date],  # AD：HbA1c 檢查日期（程式內用 hba_dt）
    ldl_dt: Optional[datetime.date],  # AF：LDL 檢查日期（程式內用 ldl_dt）
) -> Optional[str]:
    """
    條件（同時成立）：
    - D(疾病樣態)=1/2/3/4
    - E(ascvd)=0/a/b（0 也算有值）
    - 排除：D=4 且 E=0
    - 只有 D=1 或 D=3 判斷漏檢
    - HbA1c日期 與 LDL日期 皆有值且不同天

    顯示：
    - HbA1c日 > LDL日 → LDL漏檢
    - LDL日 > HbA1c日 → HbA1c漏檢
    - 相同日 → 不顯示
    """
    if d_code not in (DiseaseCode.DM, DiseaseCode.CKD, DiseaseCode.DKD, DiseaseCode.OTHER):
        return None

    token = _ascvd_token(ascvd_raw)
    if not token:  # 必須是 0/a/b
        return None

    # 排除：D=4 且 E=0
    if d_code == DiseaseCode.OTHER and token == "0":
        return None

    # 只有 D=1 或 D=3 判斷漏檢
    if d_code not in (DiseaseCode.DM, DiseaseCode.DKD):
        return None

    if not isinstance(hba_dt, datetime.date) or not isinstance(ldl_dt, datetime.date):
        return None

    if hba_dt == ldl_dt:
        return None

    if hba_dt > ldl_dt:
        return "LDL漏檢"
    if ldl_dt > hba_dt:
        return "HbA1c漏檢"
    return None


# ============================================================
# 業務邏輯：備註（需要篩檢提示）
# ============================================================
def build_screening_note(
    *,
    age:      int,
    sex:      str,
    hep_dt:   Optional[datetime.date],
    fit_dt:   Optional[datetime.date],
    pap_dt:   Optional[datetime.date],
    adult_dt: Optional[datetime.date],
    flu_dt:   Optional[datetime.date],
    today:    datetime.date,
) -> str:
    msgs: List[str] = []

    if need_bc_hep(age) and hep_dt is None:
        msgs.append("今年需檢測BC肝")

    if need_fit(age):
        if fit_dt is None or add_months(fit_dt, 24) <= today:
            msgs.append("今年需檢測糞便")

    pap_interval = pap_check_interval_months(age, sex)
    if pap_interval:
        if pap_dt is None or add_months(pap_dt, pap_interval) <= today:
            msgs.append("今年需檢測子抹")

    adult_interval = adult_check_interval_months(age)
    if adult_interval:
        if adult_dt is None or add_months(adult_dt, adult_interval) <= today:
            msgs.append("今年需檢測成健")

    if need_flu(age):
        if flu_dt is None or flu_dt.year < today.year:
            msgs.append("今年需檢測老流")

    return "\n".join(msgs)


# ============================================================
# 業務邏輯：分數計算
# ============================================================
def _score_hba(hba_val: Any, hba_dt: Optional[datetime.date]) -> int:
    if not isinstance(hba_dt, datetime.date):
        return 0
    if hba_dt.year == Rules.YEAR_UNCONDITIONAL:
        return Rules.SCORE_HBA
    if hba_dt.year == Rules.YEAR_VALUE_CHECK:
        v = parse_float(hba_val)
        if v is not None and v <= Rules.HBA_PASS_2025:
            return Rules.SCORE_HBA
    return 0


def _score_ldl(ldl_val: Any, ldl_dt: Optional[datetime.date]) -> int:
    if not isinstance(ldl_dt, datetime.date):
        return 0
    if ldl_dt.year == Rules.YEAR_UNCONDITIONAL:
        return Rules.SCORE_LDL
    if ldl_dt.year == Rules.YEAR_VALUE_CHECK:
        v = parse_float(ldl_val)
        if v is not None and v <= Rules.LDL_PASS_2025:
            return Rules.SCORE_LDL
    return 0


def _score_screening(dt: Optional[datetime.date], pts: int) -> int:
    if isinstance(dt, datetime.date) and dt.year == Rules.SCREEN_YEAR:
        return pts
    return 0


def _score_age(age: int, sex: str) -> int:
    sex = normalize_text(sex)
    if age < 18:  # 年齡未知（-1）或不合理（<18）均不給分
        return 0
    if sex == "女" and age <= 25:
        return 28
    if sex == "男" and age <= 30:
        return 28
    return 0


def _safe_int(value: Any) -> int:
    if value is None:
        return 0
    try:
        v = int(float(str(value).strip().lstrip("'").replace(",", "")))
        return max(v, 0)
    except Exception:
        return 0


def _score_visit_count_115(count_115: Any) -> int:
    n = _safe_int(count_115)
    return 0 if n <= 0 else n + 1


def calc_score(
    *,
    hba_val:  Any,
    hba_dt:   Optional[datetime.date],
    ldl_val:  Any,
    ldl_dt:   Optional[datetime.date],
    adult_dt: Optional[datetime.date],
    pap_dt:   Optional[datetime.date],
    flu_dt:   Optional[datetime.date],
    fit_dt:   Optional[datetime.date],
    hep_dt:   Optional[datetime.date],
    age:      int,
    sex:      str,
    visit_count_115: Any,
) -> Tuple[int, str]:
    age_score = _score_age(age, sex)

    prevention_score = (
        _score_screening(adult_dt, Rules.SCORE_ADULT)
        + _score_screening(pap_dt, Rules.SCORE_PAP)
        + _score_screening(flu_dt, Rules.SCORE_FLU)
        + _score_screening(fit_dt, Rules.SCORE_FIT)
        + _score_screening(hep_dt, Rules.SCORE_HEP)
    )

    exam_score = _score_hba(hba_val, hba_dt) + _score_ldl(ldl_val, ldl_dt)
    visit_score = _score_visit_count_115(visit_count_115)

    total = age_score + prevention_score + exam_score + visit_score
    parts: List[str] = []
    if age_score > 0:
        parts.append(f"1.年齡{age_score}分")
    if prevention_score > 0:
        parts.append(f"2.預防保健{prevention_score}分")
    if exam_score > 0:
        parts.append(f"3.檢查{exam_score}分")
    if visit_score > 0:
        parts.append(f"4.固定就診次數{visit_score}分")
    breakdown = "，\n".join(parts) if parts else "0分"
    return total, breakdown


# ============================================================
# 業務邏輯：AU 第一次提醒（28天）
# ============================================================
def _is_valid_year(dt: Optional[datetime.date]) -> bool:
    return isinstance(dt, datetime.date) and dt.year in Rules.VALID_YEARS


def _au_item_line(item: str, last_dt: Optional[datetime.date], today: datetime.date) -> str:
    if not _is_valid_year(last_dt):
        return f"{item}:超過2年未檢查"
    due = last_dt + datetime.timedelta(days=Rules.AU_DAYS)  # type: ignore[operator]
    if due <= today:
        return f"{item}:可立刻通知回診"
    return f"{item}:{due.strftime('%Y-%m-%d')}需回診"


def build_au_note(
    *,
    e_code: Optional[DiseaseCode],
    ascvd:  AscvdCategory,
    hba_dt: Optional[datetime.date],
    ldl_dt: Optional[datetime.date],
    today:  datetime.date,
) -> str:
    has_dm    = e_code in (DiseaseCode.DM, DiseaseCode.DKD)
    has_ckd   = e_code in (DiseaseCode.CKD, DiseaseCode.DKD)
    has_ascvd = ascvd != AscvdCategory.NONE

    need_hba = has_dm
    need_ldl = has_dm or has_ckd or has_ascvd

    if not need_hba and not need_ldl:
        return "非疾病不需回診"

    any_valid = (
        (need_hba and _is_valid_year(hba_dt))
        or (need_ldl and _is_valid_year(ldl_dt))
    )
    lines: List[str] = []
    if not any_valid:
        if need_hba:
            lines.append("HbA1c:超過2年未檢查")
        if need_ldl:
            lines.append("LDL:超過2年未檢查")
        return "\n".join(lines)

    if need_hba:
        lines.append(_au_item_line("HbA1c", hba_dt, today))
    if need_ldl:
        lines.append(_au_item_line("LDL", ldl_dt, today))
    return "\n".join(lines)


# ============================================================
# 業務邏輯：AV/AW 從 AU 文字延後
# ============================================================
_DATE_RE = re.compile(r"(\d{4}-\d{2}-\d{2})")


def _should_skip_followup(au_txt: str) -> bool:
    s = (au_txt or "").replace(" ", "")
    return (not s) or ("非疾病" in s)


def _followup_item_line(
    item: str, base_dt: Optional[datetime.date], today: datetime.date, offset: int
) -> str:
    base = base_dt if isinstance(base_dt, datetime.date) else today
    due  = base + datetime.timedelta(days=offset)
    return f"{item}:{due.strftime('%Y-%m-%d')}需回診"


def build_followup_note(
    au_txt: str, today: datetime.date, offset_days: int
) -> Optional[str]:
    if not au_txt:
        return None

    out: List[str] = []
    for ln in (line.strip() for line in au_txt.splitlines() if line.strip()):
        if ln.startswith("HbA1c:"):
            item = "HbA1c"
        elif ln.startswith("LDL:"):
            item = "LDL"
        else:
            continue

        if "可立刻通知" in ln:
            out.append(_followup_item_line(item, None, today, offset_days))
            continue

        m = _DATE_RE.search(ln)
        if m:
            try:
                base = datetime.datetime.strptime(m.group(1), "%Y-%m-%d").date()
                out.append(_followup_item_line(item, base, today, offset_days))
            except ValueError:
                pass

    return "\n".join(out) if out else None


# ============================================================
# KPI：HbA1c（AY8 / AZ7 / AZ8 / AZ9）
# ============================================================
def _write_ratio(ws, addr: str, numer: int, denom: int) -> None:
    ws[addr].value = (numer / denom) if denom > 0 else 0
    ws[addr].number_format = "0.00%"


def _collect_hba_candidates(
    ws, cols: Dict[str, Optional[int]], data_start: int, last_row: int
) -> List[Tuple[int, float]]:
    candidates: List[Tuple[int, float]] = []
    for r in range(data_start, last_row + 1):
        e = parse_disease_code(ws.cell(r, cols["dmk_code"]).value)  # type: ignore[index]
        if e not in (DiseaseCode.DM, DiseaseCode.DKD):
            continue
        if not is_zero_like(ws.cell(r, cols["ascvd"]).value):  # type: ignore[index]
            continue
        hba_dt = parse_date(ws.cell(r, cols["hba_dt"]).value)  # type: ignore[index]
        if not (isinstance(hba_dt, datetime.date) and hba_dt.year in Rules.VALID_YEARS):
            continue
        hba = parse_float(ws.cell(r, cols["hba"]).value)  # type: ignore[index]
        if hba is None:
            continue
        candidates.append((r, hba))
    return candidates


def _collect_ldl_candidates(
    ws, cols: Dict[str, Optional[int]], data_start: int, last_row: int
) -> List[Tuple[int, float]]:
    candidates: List[Tuple[int, float]] = []
    for r in range(data_start, last_row + 1):
        e = parse_disease_code(ws.cell(r, cols["dmk_code"]).value)  # type: ignore[index]
        if e not in (DiseaseCode.DM, DiseaseCode.CKD, DiseaseCode.DKD, DiseaseCode.OTHER):
            continue

        ascvd = parse_ascvd(ws.cell(r, cols["ascvd"]).value)  # type: ignore[index]
        if e == DiseaseCode.OTHER and ascvd == AscvdCategory.NONE:
            continue

        ldl_dt = parse_date(ws.cell(r, cols["ldl_dt"]).value)  # type: ignore[index]
        if not (isinstance(ldl_dt, datetime.date) and ldl_dt.year in Rules.VALID_YEARS):
            continue

        ldl = parse_float(ws.cell(r, cols["ldl"]).value)  # type: ignore[index]
        if ldl is None:
            continue

        candidates.append((r, ldl))
    return candidates


def collect_kpi_mark_sets(
    ws, cols: Dict[str, Optional[int]], data_start: int, last_row: int,
    hba_candidates: Optional[List[Tuple[int, float]]] = None,
    ldl_candidates: Optional[List[Tuple[int, float]]] = None,
) -> Dict[str, set]:
    if hba_candidates is None:
        hba_candidates = _collect_hba_candidates(ws, cols, data_start, last_row)
    if ldl_candidates is None:
        ldl_candidates = _collect_ldl_candidates(ws, cols, data_start, last_row)

    ay_rows = {r for r, v in hba_candidates if v <= Rules.HBA_CONTROL_THRESHOLD}

    hba_sorted = sorted(hba_candidates, key=lambda x: (x[1], x[0]))
    hba_k = int(math.ceil(Rules.HBA_TARGET_PERCENT * len(hba_sorted))) if hba_sorted else 0
    az_rows = {r for r, _ in hba_sorted[:hba_k]}

    ldl_threshold = Rules.LDL_KPI_THRESHOLDS[0]
    bb_rows = {r for r, v in ldl_candidates if v <= ldl_threshold}

    ldl_sorted = sorted(ldl_candidates, key=lambda x: (x[1], x[0]))
    ldl_k = int(math.ceil(Rules.LDL_TARGET_PERCENT * len(ldl_sorted))) if ldl_sorted else 0
    bc_rows = {r for r, _ in ldl_sorted[:ldl_k]}

    return {
        "ay": ay_rows,
        "az": az_rows,
        "bb": bb_rows,
        "bc": bc_rows,
    }


def calc_hba_kpi_ay_az(
    ws, cols: Dict[str, Optional[int]], data_start: int, last_row: int,
    hba_candidates: Optional[List[Tuple[int, float]]] = None,
) -> None:
    if hba_candidates is None:
        hba_candidates = _collect_hba_candidates(ws, cols, data_start, last_row)
    hba_values = [v for _, v in hba_candidates]
    denom = len(hba_values)

    numer_ay = sum(1 for v in hba_values if v <= Rules.HBA_CONTROL_THRESHOLD)
    _write_ratio(ws, "AY8", numer_ay, denom)
    ws["AY9"].value = f"{numer_ay}/{denom}" if denom > 0 else "0/0"

    if denom <= 0:
        ws["AZ7"].value = None
        _write_ratio(ws, "AZ8", 0, 0)
        ws["AZ9"].value = "0/0"
        print("AZ 分母=0，分子=0，比例=0.00%，切點=None")
        return

    hba_values.sort()
    k = int(math.ceil(Rules.HBA_TARGET_PERCENT * denom))
    k = max(1, min(k, denom))
    cutoff = hba_values[k - 1]

    ws["AZ7"].value = f"<={cutoff:.2f}"
    _write_ratio(ws, "AZ8", k, denom)
    ws["AZ9"].value = f"{k}/{denom}"

    ratio = k / denom
    print(f"AZ 分母={denom}，分子={k}，比例={ratio*100:.2f}%，切點={cutoff:.2f}")


# ============================================================
# KPI：LDL（BB8/BC8）
# ============================================================
def calc_ldl_percentiles(
    ws, cols: Dict[str, Optional[int]], data_start: int, last_row: int,
    ldl_candidates: Optional[List[Tuple[int, float]]] = None,
) -> None:
    if ldl_candidates is None:
        ldl_candidates = _collect_ldl_candidates(ws, cols, data_start, last_row)
    ldl_values = [v for _, v in ldl_candidates]
    th_control = Rules.LDL_KPI_THRESHOLDS[0]  # 100
    denom = len(ldl_values)

    numer_bb = sum(1 for v in ldl_values if v <= th_control)
    _write_ratio(ws, "BB8", numer_bb, denom)
    ws["BB9"].value = f"{numer_bb}/{denom}" if denom > 0 else "0/0"

    if denom <= 0:
        ws["BC7"].value = None
        _write_ratio(ws, "BC8", 0, 0)
        ws["BC9"].value = "0/0"
        print("BC 分母=0，分子=0，比例=0.00%，切點=None")
        return

    ldl_values.sort()
    k = int(math.ceil(Rules.LDL_TARGET_PERCENT * denom))
    k = max(1, min(k, denom))
    cutoff = ldl_values[k - 1]

    ws["BC7"].value = f"<={cutoff:.1f}".replace(".0", "")
    _write_ratio(ws, "BC8", k, denom)
    ws["BC9"].value = f"{k}/{denom}"

    ratio = k / denom
    print(f"BC 分母={denom}，分子={k}，比例={ratio*100:.2f}%，切點={cutoff:.0f}")


# ============================================================
# 格式工具
# ============================================================
def apply_full_grid(ws, max_row: int, max_col: int) -> None:
    thin = Side(style="thin")
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    # 只對實際有資料的儲存格設定邊框，跳過空值以提升效能
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            if cell.value is not None or r == 1:
                cell.border = grid


def apply_date_format(
    ws, cols: Dict[str, Optional[int]], data_start: int, last_row: int
) -> None:
    date_col_keys = ["bday", "adult", "pap", "flu", "fit", "hep",
                     "hba_dt", "ldl_dt", "uacr_dt"]
    date_cols = [cols.get(k) for k in date_col_keys if cols.get(k)]
    for r in range(data_start, last_row + 1):
        for c in date_cols:
            cell = ws.cell(r, c)  # type: ignore[arg-type]
            if isinstance(cell.value, (datetime.date, datetime.datetime)):
                cell.number_format = "yyyy-mm-dd"


# ============================================================
# 百分位名單 sheet（sheet2）
# ============================================================
PERCENTILE_SHEET_NAME = "百分位名單"
PINK_FILL = openpyxl.styles.PatternFill(fill_type="solid", fgColor="EAC0C0")
BLUE_FILL = openpyxl.styles.PatternFill(fill_type="solid", fgColor="A9C2D9")
NO_FILL = openpyxl.styles.PatternFill(fill_type=None)


def _copy_row_style(ws, src_row: int, dst_row: int, start_col: int = 1, end_col: int = 24) -> None:
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for c in range(start_col, end_col + 1):
        src = ws.cell(src_row, c)
        dst = ws.cell(dst_row, c)
        dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)


def _clear_percentile_data_area(ws, start_row: int = 5, end_col: int = 24) -> None:
    for r in range(start_row, ws.max_row + 1):
        for c in range(1, end_col + 1):
            ws.cell(r, c).value = None


def _make_percentile_record(ws_main, row: int, cols: Dict[str, Optional[int]], metric: str, name_fill: str) -> Dict[str, Any]:
    value_col = cols["ldl"] if metric == "ldl" else cols["hba"]
    date_col = cols["ldl_dt"] if metric == "ldl" else cols["hba_dt"]
    return {
        "row": row,
        "name": ws_main.cell(row, cols["name"]).value,  # type: ignore[index]
        "bday": ws_main.cell(row, cols["bday"]).value,  # type: ignore[index]
        "id": ws_main.cell(row, cols["id"]).value,      # type: ignore[index]
        "score": ws_main.cell(row, cols["score"]).value,  # type: ignore[index]
        "note": ws_main.cell(row, cols["note"]).value,    # type: ignore[index]
        "value": ws_main.cell(row, value_col).value if value_col else None,
        "last_dt": ws_main.cell(row, date_col).value if date_col else None,
        "au": ws_main.cell(row, cols["au"]).value,      # type: ignore[index]
        "av": ws_main.cell(row, cols["av"]).value,      # type: ignore[index]
        "aw": ws_main.cell(row, cols["aw"]).value,      # type: ignore[index]
        "fill": name_fill,
    }


def _collect_side_records_sorted(
    ws_main,
    cols: Dict[str, Optional[int]],
    data_start: int,
    last_row: int,
    primary_mark_key: str,
    secondary_mark_key: str,
    metric: str,
) -> List[Dict[str, Any]]:
    primary_rows: List[int] = []
    secondary_rows: List[int] = []

    for row in range(data_start, last_row + 1):
        primary = normalize_text(ws_main.cell(row, cols[primary_mark_key]).value) if cols.get(primary_mark_key) else ""
        secondary = normalize_text(ws_main.cell(row, cols[secondary_mark_key]).value) if cols.get(secondary_mark_key) else ""

        if primary == "✔":
            primary_rows.append(row)
        elif secondary == "✔":
            secondary_rows.append(row)

    out: List[Dict[str, Any]] = []
    for row in primary_rows:
        out.append(_make_percentile_record(ws_main, row, cols, metric, "pink"))
    for row in secondary_rows:
        out.append(_make_percentile_record(ws_main, row, cols, metric, "blue"))
    return out


def _collect_percentile_records(ws_main, cols: Dict[str, Optional[int]], data_start: int, last_row: int) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    ldl_records = _collect_side_records_sorted(
        ws_main, cols, data_start, last_row,
        primary_mark_key="bb_mark", secondary_mark_key="bc_mark", metric="ldl"
    )
    hba_records = _collect_side_records_sorted(
        ws_main, cols, data_start, last_row,
        primary_mark_key="ay_mark", secondary_mark_key="az_mark", metric="hba"
    )
    return ldl_records, hba_records




def _ratio_numerator_text(value: Any) -> str:
    s = normalize_text(value)
    if not s:
        return "0"
    return s.split("/", 1)[0].strip() or "0"


def _set_percentile_title_rich(ws, cell_ref: str, prefix: str, red_num: Any, blue_num: Any) -> None:
    ws[cell_ref].value = f"{prefix}({_ratio_numerator_text(red_num)}/{_ratio_numerator_text(blue_num)}人)"


def _write_percentile_side(ws, row: int, start_col: int, rec: Optional[Dict[str, Any]]) -> None:
    cols_seq = [start_col + i for i in range(10)]
    name_cell = ws.cell(row, start_col)

    if not rec:
        for c in cols_seq:
            ws.cell(row, c).value = None
        name_cell.fill = copy(NO_FILL)
        return

    values = [
        rec.get("name"),
        rec.get("bday"),
        rec.get("id"),
        rec.get("score"),
        rec.get("note"),
        rec.get("value"),
        rec.get("last_dt"),
        rec.get("au"),
        rec.get("av"),
        rec.get("aw"),
    ]
    for c, v in zip(cols_seq, values):
        ws.cell(row, c).value = v

    if normalize_text(rec.get("name")):
        name_cell.fill = copy(PINK_FILL if rec.get("fill") == "pink" else BLUE_FILL)
    else:
        name_cell.fill = copy(NO_FILL)


def populate_percentile_sheet(
    wb_tpl,
    ws_main,
    cols: Dict[str, Optional[int]],
    data_start: int,
    last_row: int,
) -> None:
    if PERCENTILE_SHEET_NAME not in wb_tpl.sheetnames:
        ws = wb_tpl.create_sheet(PERCENTILE_SHEET_NAME)
    else:
        ws = wb_tpl[PERCENTILE_SHEET_NAME]

    ldl_records, hba_records = _collect_percentile_records(ws_main, cols, data_start, last_row)

    _set_percentile_title_rich(ws, "A1", "LDL百分位", ws_main["BB9"].value, ws_main["BC9"].value)
    _set_percentile_title_rich(ws, "N1", "HBA1C百分位", ws_main["AY9"].value, ws_main["AZ9"].value)
    ws["A2"] = "紅色：達到標準、藍色：達到73.8%"
    ws["N2"] = "紅色：達到標準、藍色：達到73.8%"

    _clear_percentile_data_area(ws, start_row=5, end_col=24)

    data_rows = max(len(ldl_records), len(hba_records), 1)
    style_src_row = 5
    for idx in range(data_rows):
        row = 5 + idx
        if row != style_src_row:
            _copy_row_style(ws, style_src_row, row, 1, 24)
        ws.row_dimensions[row].height = ws.row_dimensions[style_src_row].height or 66
        _write_percentile_side(ws, row, 1, ldl_records[idx] if idx < len(ldl_records) else None)
        _write_percentile_side(ws, row, 14, hba_records[idx] if idx < len(hba_records) else None)

    for row in range(5, 5 + data_rows):
        for col in [2, 7, 15, 20]:
            cell = ws.cell(row, col)
            if isinstance(cell.value, (datetime.date, datetime.datetime)):
                cell.number_format = "yyyy-mm-dd"
    apply_full_grid(ws, ws.max_row, 24)

# ============================================================
# 資料填充子函數
# ============================================================
def _load_and_validate_source(source_path: str):
    wb = openpyxl.load_workbook(source_path, data_only=True)
    need = ["會員名單", "ascvd", "HealthCase",
            "成人健檢", "子宮抹片", "老人流感", "糞便潛血", "肝炎篩檢"]
    missing = [s for s in need if s not in wb.sheetnames]
    if missing:
        raise ValueError(f"原始檔缺少工作表：{'、'.join(missing)}")
    return wb


def _fill_member_basic(
    ws,
    sh_member,
    cols:       Dict[str, Optional[int]],
    data_start: int,
    now:        datetime.date,
) -> Tuple[Dict[str, List[int]], Dict[int, MemberMeta], int]:
    MEMBER_HEADER_ROW = 5
    mmap = build_header_map(sh_member, MEMBER_HEADER_ROW)

    m_name = find_column_exact(mmap, ["會員姓名"])
    m_id   = find_column_exact(mmap, ["會員身份証", "會員身份證", "會員身分證"])
    m_bday = find_column_exact(mmap, ["會員生日"])
    m_tel  = find_column_exact(mmap, ["電話"])
    m_abc  = find_column_exact(mmap, ["會員別"])
    m_dmk  = find_column_exact(mmap, ["疾病樣態"])
    m_cnt  = find_column_exact(mmap, ["就診次數"])

    if any(v is None for v in [m_name, m_id, m_bday, m_tel, m_abc, m_dmk, m_cnt]):
        raise ValueError(
            "原始檔「會員名單」欄位不完整"
            "（會員姓名/會員身分證/會員生日/電話/會員別/疾病樣態/就診次數）"
        )

    clinic_val = normalize_text(sh_member["A1"].value)
    id_to_rows: Dict[str, List[int]] = {}
    meta: Dict[int, MemberMeta] = {}
    out_r = data_start
    wrote_any = False

    for r in range(MEMBER_HEADER_ROW + 1, sh_member.max_row + 1):
        sid  = sh_member.cell(r, m_id).value    # type: ignore[arg-type]
        name = sh_member.cell(r, m_name).value  # type: ignore[arg-type]
        if (sid in (None, "")) and (name in (None, "")):
            continue

        wrote_any = True
        sid_s  = str(sid).strip() if sid is not None else ""
        bday   = parse_date(sh_member.cell(r, m_bday).value)   # type: ignore[arg-type]
        e_code = parse_disease_code(sh_member.cell(r, m_dmk).value)  # type: ignore[arg-type]
        age    = calc_age(bday, now) if bday else -1
        sex    = infer_gender_from_id(sid)

        safe_set(ws, out_r, cols.get("clinic"),   clinic_val)
        safe_set(ws, out_r, cols.get("name"),     name)
        safe_set(ws, out_r, cols.get("id"),       sid)
        safe_set(ws, out_r, cols.get("bday"),     bday)
        safe_set(ws, out_r, cols.get("age"),      age if age >= 0 else None)
        safe_set(ws, out_r, cols.get("tel"),      sh_member.cell(r, m_tel).value)  # type: ignore[arg-type]
        safe_set(ws, out_r, cols.get("abc"),      sh_member.cell(r, m_abc).value)  # type: ignore[arg-type]
        safe_set(ws, out_r, cols.get("dmk_code"), e_code.value if e_code else None)
        safe_set(ws, out_r, cols.get("cnt"),      sh_member.cell(r, m_cnt).value)  # type: ignore[arg-type]
        safe_set(ws, out_r, cols.get("sex"),      sex)

        meta[out_r] = MemberMeta(row=out_r, bday=bday, age=age, e_code=e_code)

        if sid_s:
            id_to_rows.setdefault(sid_s, []).append(out_r)

        out_r += 1

    last_row = out_r - 1 if wrote_any else (data_start - 1)
    return id_to_rows, meta, last_row


def _fill_ascvd(
    ws,
    sh_ascvd:  Any,
    cols:      Dict[str, Optional[int]],
    id_to_rows: Dict[str, List[int]],
    meta:      Dict[int, MemberMeta],
) -> None:
    ASCVD_HEADER_ROW = 5
    amap  = build_header_map(sh_ascvd, ASCVD_HEADER_ROW)
    a_id  = find_column_exact(amap, ["ID", "id"])
    a_asc = find_column_exact(amap, ["ASCVD", "ascvd"])
    if a_id is None or a_asc is None:
        raise ValueError("原始檔「ascvd」找不到 ID / ASCVD 欄位（表頭列預期第5列）")

    for r in range(ASCVD_HEADER_ROW + 1, sh_ascvd.max_row + 1):
        pid = sh_ascvd.cell(r, a_id).value
        val = sh_ascvd.cell(r, a_asc).value
        if pid is None or str(pid).strip() == "":
            continue
        if val is None or str(val).strip() in ("",):  # 0 也要保留寫入
            continue

        rows = id_to_rows.get(str(pid).strip())
        if not rows:
            continue

        for tr in rows:
            safe_set(ws, tr, cols.get("ascvd"), val)
            if tr in meta:
                meta[tr].ascvd = parse_ascvd(val)


def _fill_screening(
    ws,
    sheet,
    target_col: Optional[int],
    id_to_rows: Dict[str, List[int]],
) -> None:
    if not target_col:
        return
    hmap    = build_header_map(sheet, 1)
    sid_col = find_column_exact(hmap, ["ID", "身分證號", "身份證號"])
    dt_col  = find_column_exact(hmap, ["最後篩檢日期"])
    if sid_col is None or dt_col is None:
        raise ValueError(f"「{sheet.title}」找不到 ID 或 最後篩檢日期 欄位")

    for rr in range(2, sheet.max_row + 1):
        pid = sheet.cell(rr, sid_col).value
        dt  = parse_date(sheet.cell(rr, dt_col).value)
        if pid is None or str(pid).strip() == "" or dt is None:
            continue

        rows = id_to_rows.get(str(pid).strip())
        if not rows:
            continue

        for tr in rows:
            existing = parse_date(ws.cell(tr, target_col).value)
            if existing is None or dt > existing:
                ws.cell(tr, target_col).value = dt


def _fill_health_case(
    ws,
    sh_health: Any,
    cols:      Dict[str, Optional[int]],
    id_to_rows: Dict[str, List[int]],
) -> None:
    hmap = build_header_map(sh_health, 1)
    field_aliases = {
        "hc_id":      ["家醫收案會員ID", "ID"],
        "hc_hba":     ["最近一次HbA1c檢查結果(%)"],
        "hc_hba_dt":  ["最近一次HbA1c檢查日期"],
        "hc_ldl":     ["最近一次LDL檢查結果(mg/dL)"],
        "hc_ldl_dt":  ["最近一次LDL檢查日期"],
        "hc_uacr":    ["最近一次UACR檢查結果(mg/gm)"],
        "hc_uacr_dt": ["最近一次UACR檢查日期"],
    }
    fc = {k: find_column_exact(hmap, v) for k, v in field_aliases.items()}
    if any(v is None for v in fc.values()):
        raise ValueError(
            "原始檔「HealthCase」欄位不完整"
            "（家醫收案會員ID / HbA1c結果+日期 / LDL結果+日期 / UACR結果+日期）"
        )

    for r in range(2, sh_health.max_row + 1):
        pid = sh_health.cell(r, fc["hc_id"]).value
        if pid is None or str(pid).strip() == "":
            continue

        rows = id_to_rows.get(str(pid).strip())
        if not rows:
            continue

        for tr in rows:
            for val_key, dt_key, col_val_key, col_dt_key in [
                ("hc_hba",  "hc_hba_dt",  "hba",  "hba_dt"),
                ("hc_ldl",  "hc_ldl_dt",  "ldl",  "ldl_dt"),
                ("hc_uacr", "hc_uacr_dt", "uacr", "uacr_dt"),
            ]:
                v  = parse_float(sh_health.cell(r, fc[val_key]).value)
                dt = parse_date(sh_health.cell(r, fc[dt_key]).value)
                # 以日期較新者優先：避免重複 ID 時舊資料蓋掉新資料
                existing_dt = parse_date(ws.cell(tr, cols[col_dt_key]).value) if cols.get(col_dt_key) else None
                if dt and (existing_dt is None or dt >= existing_dt):
                    if v is not None and v != 0:
                        safe_set(ws, tr, cols.get(col_val_key), v)
                    safe_set(ws, tr, cols.get(col_dt_key), dt)


def _compute_all_derived(
    ws,
    cols:       Dict[str, Optional[int]],
    meta:       Dict[int, MemberMeta],
    data_start: int,
    last_row:   int,
    now:        datetime.date,
    kpi_marks:  Optional[Dict[str, set]] = None,
) -> None:
    kpi_marks = kpi_marks or {}

    def _get_dt(row: int, key: str) -> Optional[datetime.date]:
        """從模板工作表取得指定列、指定欄的日期值。"""  
        c = cols.get(key)
        return parse_date(ws.cell(row, c).value) if c else None  # type: ignore[arg-type]

    for rr in range(data_start, last_row + 1):
        e_code = parse_disease_code(ws.cell(rr, cols["dmk_code"]).value)  # type: ignore[index]
        ascvd  = parse_ascvd(ws.cell(rr, cols["ascvd"]).value)            # type: ignore[index]
        ascvd_raw = ws.cell(rr, cols["ascvd"]).value                      # type: ignore[index]

        m   = meta.get(rr, MemberMeta(row=rr))
        age = m.age if isinstance(m.age, int) else -1
        sex = normalize_text(ws.cell(rr, cols["sex"]).value)              # type: ignore[index]

        adult_dt = _get_dt(rr, "adult")
        pap_dt   = _get_dt(rr, "pap")
        flu_dt   = _get_dt(rr, "flu")
        fit_dt   = _get_dt(rr, "fit")
        hep_dt   = _get_dt(rr, "hep")
        hba_dt   = _get_dt(rr, "hba_dt")   # 你的模板是 AD
        ldl_dt   = _get_dt(rr, "ldl_dt")   # 你的模板是 AF

        hba_val = ws.cell(rr, cols["hba"]).value  # type: ignore[index]
        ldl_val = ws.cell(rr, cols["ldl"]).value  # type: ignore[index]

        safe_set(ws, rr, cols.get("disease_text"),
                 disease_group_text(e_code, ascvd))

        for _mark_key in ("ay", "az", "bb", "bc"):
            safe_set(ws, rr, cols.get(f"{_mark_key}_mark"),
                     "✔" if rr in kpi_marks.get(_mark_key, set()) else None)

        # AK：✔
        safe_set(ws, rr, cols.get("ak"),
                 "✔" if should_check_ak(e_code=e_code, ascvd=ascvd, age=age, hba_val=hba_val) else None)

        # AX：漏檢項目
        ax_txt = build_ax_leak_item(
            d_code=e_code,
            ascvd_raw=ascvd_raw,
            hba_dt=hba_dt,
            ldl_dt=ldl_dt,
        )
        safe_set(ws, rr, cols.get("ax"), ax_txt or None)

        # 年齡未知時不產生篩檢備註，避免以 0 歲誤判
        note = build_screening_note(
            age=age, sex=sex,
            hep_dt=hep_dt, fit_dt=fit_dt, pap_dt=pap_dt,
            adult_dt=adult_dt, flu_dt=flu_dt, today=now,
        ) if age >= 0 else ""
        safe_set(ws, rr, cols.get("note"), note or None)

        visit_count_115 = ws.cell(rr, cols["n_count_115"]).value if cols.get("n_count_115") else None

        score, breakdown = calc_score(
            hba_val=hba_val, hba_dt=hba_dt,
            ldl_val=ldl_val, ldl_dt=ldl_dt,
            adult_dt=adult_dt, pap_dt=pap_dt,
            flu_dt=flu_dt, fit_dt=fit_dt, hep_dt=hep_dt,
            age=age, sex=sex,
            visit_count_115=visit_count_115,
        )
        safe_set(ws, rr, cols.get("score"),     score)
        safe_set(ws, rr, cols.get("breakdown"), breakdown)

        au_txt = build_au_note(
            e_code=e_code, ascvd=ascvd,
            hba_dt=hba_dt, ldl_dt=ldl_dt, today=now,
        )
        safe_set(ws, rr, cols.get("au"), au_txt or None)

        if au_txt and not _should_skip_followup(au_txt):
            safe_set(ws, rr, cols.get("av"),
                     build_followup_note(au_txt, now, Rules.AV_OFFSET_DAYS))
            safe_set(ws, rr, cols.get("aw"),
                     build_followup_note(au_txt, now, Rules.AW_OFFSET_DAYS))
        else:
            safe_set(ws, rr, cols.get("av"), None)
            safe_set(ws, rr, cols.get("aw"), None)




# NOTE: 此函數目前未被呼叫，保留備用（如需啟用請在 process_excel 中加入呼叫）
def _format_breakdown_column(ws, cols: Dict[str, Optional[int]], data_start: int, last_row: int) -> None:
    c = cols.get("breakdown")
    if not c:
        return

    col_letter = ws.cell(1, c).column_letter
    try:
        if (ws.column_dimensions[col_letter].width or 0) < 32:
            ws.column_dimensions[col_letter].width = 32
    except Exception:
        ws.column_dimensions[col_letter].width = 32

    for r in range(1, last_row + 1):
        cell = ws.cell(r, c)
        base = copy(cell.alignment) if cell.alignment else Alignment()
        cell.alignment = Alignment(
            horizontal=base.horizontal,
            vertical=base.vertical or "center",
            text_rotation=base.text_rotation,
            wrap_text=True,
            shrink_to_fit=base.shrink_to_fit,
            indent=base.indent,
        )


# ============================================================
# 清空資料列
# ============================================================
def _clear_data_rows(
    ws, data_start: int, max_row: int, cols: Dict[str, Optional[int]]
) -> None:
    clear_keys = [
        "clinic", "name", "id", "bday", "age", "tel", "cnt", "sex", "abc",
        "dmk_code", "ascvd",
        "adult", "pap", "flu", "fit", "hep",
        "hba", "hba_dt", "ldl", "ldl_dt", "uacr", "uacr_dt",
        "disease_text", "ay_mark", "az_mark", "ak", "ax", "score", "breakdown", "note",
        "bb_mark", "bc_mark", "au", "av", "aw",
        "m_count_114", "n_count_115", "r_amount_114", "s_amount_115",
    ]
    col_ids = [cols[k] for k in clear_keys if cols.get(k)]
    for r in range(data_start, max_row + 1):
        for c in col_ids:
            ws.cell(r, c).value = None  # type: ignore[arg-type]


# ============================================================
# 主流程
# ============================================================
# ============================================================
# 結構化流程（V10.3）
# ============================================================
def load_source(source_path: str) -> Dict[str, Any]:
    wb_src = _load_and_validate_source(source_path)
    return {
        "wb_src": wb_src,
        "sh_member": wb_src["會員名單"],
        "sh_ascvd": wb_src["ascvd"],
        "sh_health": wb_src["HealthCase"],
        "screening_sheets": {
            "adult": wb_src["成人健檢"],
            "pap": wb_src["子宮抹片"],
            "flu": wb_src["老人流感"],
            "fit": wb_src["糞便潛血"],
            "hep": wb_src["肝炎篩檢"],
        },
        "claim_sums": collect_monthly_claim_summaries(wb_src),
    }


def load_template(template_path: str) -> Dict[str, Any]:
    wb_tpl = openpyxl.load_workbook(template_path)
    if Rules.SHEET_TARGET not in wb_tpl.sheetnames:
        raise ValueError(f"模板檔缺少工作表：{Rules.SHEET_TARGET}")

    ws = wb_tpl[Rules.SHEET_TARGET]
    data_start = Rules.DATA_START_ROW
    cols = detect_template_columns(ws, data_start)
    _clear_data_rows(ws, data_start, ws.max_row, cols)

    return {
        "wb_tpl": wb_tpl,
        "ws": ws,
        "cols": cols,
        "data_start": data_start,
    }


def fill_basic_data(source_ctx: Dict[str, Any], template_ctx: Dict[str, Any], now: datetime.date) -> Dict[str, Any]:
    id_to_rows, meta, last_row = _fill_member_basic(
        template_ctx["ws"],
        source_ctx["sh_member"],
        template_ctx["cols"],
        template_ctx["data_start"],
        now,
    )
    return {
        "id_to_rows": id_to_rows,
        "meta": meta,
        "last_row": last_row,
    }


def fill_external_data(source_ctx: Dict[str, Any], template_ctx: Dict[str, Any], runtime_ctx: Dict[str, Any]) -> None:
    ws = template_ctx["ws"]
    cols = template_ctx["cols"]
    data_start = template_ctx["data_start"]
    last_row = runtime_ctx["last_row"]
    id_to_rows = runtime_ctx["id_to_rows"]
    meta = runtime_ctx["meta"]

    _fill_ascvd(ws, source_ctx["sh_ascvd"], cols, id_to_rows, meta)
    for key, sheet in source_ctx["screening_sheets"].items():
        _fill_screening(ws, sheet, cols.get(key), id_to_rows)
    _fill_health_case(ws, source_ctx["sh_health"], cols, id_to_rows)
    fill_monthly_claim_summary_columns(ws, data_start, last_row, cols, source_ctx["claim_sums"])


def compute_derived(template_ctx: Dict[str, Any], runtime_ctx: Dict[str, Any], now: datetime.date) -> None:
    ws = template_ctx["ws"]
    cols = template_ctx["cols"]
    data_start = template_ctx["data_start"]
    last_row = runtime_ctx["last_row"]

    hba_candidates = _collect_hba_candidates(ws, cols, data_start, last_row)
    ldl_candidates = _collect_ldl_candidates(ws, cols, data_start, last_row)
    kpi_marks = collect_kpi_mark_sets(
        ws, cols, data_start, last_row,
        hba_candidates=hba_candidates,
        ldl_candidates=ldl_candidates,
    )

    _compute_all_derived(ws, cols, runtime_ctx["meta"], data_start, last_row, now, kpi_marks)
    apply_date_format(ws, cols, data_start, last_row)

    runtime_ctx["hba_candidates"] = hba_candidates
    runtime_ctx["ldl_candidates"] = ldl_candidates


def compute_kpis(template_ctx: Dict[str, Any], runtime_ctx: Dict[str, Any]) -> None:
    ws = template_ctx["ws"]
    cols = template_ctx["cols"]
    data_start = template_ctx["data_start"]
    last_row = runtime_ctx["last_row"]

    calc_hba_kpi_ay_az(
        ws, cols, data_start, last_row,
        hba_candidates=runtime_ctx.get("hba_candidates"),
    )
    calc_ldl_percentiles(
        ws, cols, data_start, last_row,
        ldl_candidates=runtime_ctx.get("ldl_candidates"),
    )
    populate_percentile_sheet(template_ctx["wb_tpl"], ws, cols, data_start, last_row)


def finalize_and_save(source_path: str, template_ctx: Dict[str, Any], now_dt: datetime.datetime) -> str:
    wb_tpl = template_ctx["wb_tpl"]
    base_dir = os.path.dirname(os.path.abspath(source_path))

    for sht_name in (Rules.SHEET_TARGET, PERCENTILE_SHEET_NAME):
        if sht_name in wb_tpl.sheetnames:
            wb_tpl[sht_name].sheet_view.showGridLines = True

    out_path = os.path.join(base_dir, f"選會員{now_dt.strftime('%m%d_%H%M%S')}.xlsx")
    wb_tpl.save(out_path)
    return out_path


def process_excel(source_path: str, template_path: str) -> str:
    tz_tw = datetime.timezone(datetime.timedelta(hours=8))
    now_dt = datetime.datetime.now(tz_tw)
    now = now_dt.date()

    source_ctx = load_source(source_path)
    template_ctx = load_template(template_path)
    runtime_ctx = fill_basic_data(source_ctx, template_ctx, now)

    if runtime_ctx["last_row"] < template_ctx["data_start"]:
        return finalize_and_save(source_path, template_ctx, now_dt)

    fill_external_data(source_ctx, template_ctx, runtime_ctx)
    compute_derived(template_ctx, runtime_ctx, now)
    compute_kpis(template_ctx, runtime_ctx)
    return finalize_and_save(source_path, template_ctx, now_dt)


# ============================================================
# 進入點
# ============================================================
def main() -> None:
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()

    src = filedialog.askopenfilename(
        title="選擇原始 Excel 檔案（沒有會員指標的那份）",
        filetypes=[("Excel files", "*.xlsx")],
    )
    if not src:
        return

    script_dir = os.path.dirname(os.path.abspath(__file__))
    template   = os.path.join(script_dir, Rules.TEMPLATE_NAME)

    if not os.path.exists(template):
        messagebox.showerror(
            "錯誤",
            f"找不到模板檔：\n{template}\n\n"
            f"請把模板放到此 .py 同資料夾，檔名需為：\n{Rules.TEMPLATE_NAME}",
        )
        return

    try:
        out = process_excel(src, template)
        messagebox.showinfo("完成", f"已輸出：\n{out}")
        open_file_cross_platform(out)
    except (ValueError, KeyError, OSError, InvalidFileException) as e:
        messagebox.showerror("錯誤", str(e))
    except Exception as e:
        messagebox.showerror("錯誤", f"未預期錯誤：{e}")


if __name__ == "__main__":
    main()