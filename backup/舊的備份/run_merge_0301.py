# -*- coding: utf-8 -*-
"""
選會員產檔工具（結構化單檔版）— 0301 樣板版

功能概要
- 讀取原始檔（會員名單 / ascvd / HealthCase / 5 大篩檢）
- 套用模板（只保留「會員指標」分頁），寫入資料列
- 計算：疾病樣態文字、備註、分數/明細、AU/AV/AW 文字提醒
- KPI 百分位（0301 樣板）
  - HbA1c：AY8/AZ8（<=7 / <=7.3） 〔BA8 已刪除，不再填〕
  - LDL：BB8/BC8（<=100 / <=110）〔<=105 與 BE/BF 舊欄已刪除，不再填〕
- 全表格加網格線、輸出到原始檔同資料夾並自動開檔（macOS/Windows/Linux）

注意
- Python 3.9 相容
- 不拆模組：同一支檔案即可維護
"""

import os
import re
import sys
import datetime
import subprocess
from typing import Optional, Dict, Any, Tuple, List

import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.utils.exceptions import InvalidFileException


# =========================
# 固定設定
# =========================
TEMPLATE_NAME = "選會員樣板0301.xlsx"   # 新模板檔名：放在本 .py 同資料夾
SHEET_TARGET = "會員指標"


# =========================
# 小工具
# =========================
def open_file_cross_platform(path: str) -> None:
    """跨平台打開檔案（盡量不拋錯）"""
    try:
        if sys.platform.startswith("darwin"):
            subprocess.call(("open", path))
        elif os.name == "nt":
            os.startfile(path)  # type: ignore[attr-defined]
        else:
            subprocess.call(("xdg-open", path))
    except Exception:
        pass


def normalize_text(v: Any) -> str:
    if v is None:
        return ""
    return str(v).replace("\t", "").replace("　", "").strip()


def normalize_header(v: Any) -> str:
    """表頭用：移除空白+換行，方便模糊比對"""
    s = normalize_text(v)
    s = s.replace(" ", "").replace("\n", "")
    return s


def build_header_map(sheet, header_row: int) -> Dict[str, int]:
    """完全匹配用：key = 原樣表頭文字（不做 normalize_header 破壞）"""
    m: Dict[str, int] = {}
    for c in range(1, sheet.max_column + 1):
        k = normalize_text(sheet.cell(header_row, c).value)
        if k:
            m[k] = c
    return m


def find_column_exact(hmap: Dict[str, int], aliases: List[str]) -> Optional[int]:
    for a in aliases:
        if a in hmap:
            return hmap[a]
    return None


def get_merged_header_text(ws, header_row: int, c: int) -> str:
    """
    多層表頭支援：
    - header_row-1：大類（可能合併）
    - header_row：欄名/說明
    合併兩列做模糊比對
    """
    upper = normalize_header(ws.cell(header_row - 1, c).value) if header_row > 1 else ""
    lower = normalize_header(ws.cell(header_row, c).value)
    return (upper + " " + lower).strip()


def find_col_by_keywords(ws, header_row: int, keywords: List[str]) -> Optional[int]:
    """
    模糊比對：表頭(去空白/去換行) 必須同時包含 keywords 的所有字
    支援多層表頭：用 (header_row-1 + header_row) 合併字串來找
    """
    keys = [k.replace(" ", "").replace("\n", "") for k in keywords if k]
    for c in range(1, ws.max_column + 1):
        h = get_merged_header_text(ws, header_row, c).replace(" ", "")
        if not h:
            continue
        if all(k in h for k in keys):
            return c
    return None


def find_col_by_keywords_any_row(ws, max_row: int, keywords: List[str]) -> Optional[int]:
    """
    跨多列找欄位：在 1..max_row 的同一欄中，只要有任一儲存格文字
    同時包含 keywords 全部關鍵字，就回傳該欄。
    """
    keys = [k.replace(" ", "").replace("\n", "") for k in keywords if k]
    for c in range(1, ws.max_column + 1):
        blob = ""
        for r in range(1, max_row + 1):
            v = normalize_header(ws.cell(r, c).value)
            if v:
                blob += v
        if not blob:
            continue
        if all(k in blob for k in keys):
            return c
    return None


def find_header_row_contains(sheet, must_have: List[str], search_rows: int = 250) -> Optional[int]:
    """
    找到同一列同時包含 must_have 幾個表頭（模糊等於：去空白/去換行後比對）
    """
    must = [m.replace(" ", "").replace("\n", "") for m in must_have]
    for r in range(1, min(search_rows, sheet.max_row) + 1):
        found = set()
        for c in range(1, sheet.max_column + 1):
            v = normalize_header(sheet.cell(r, c).value)
            if not v:
                continue
            for m in must:
                if v == m:
                    found.add(m)
        if all(m in found for m in must):
            return r
    return None


def parse_date(value: Any) -> Optional[datetime.date]:
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value

    s = str(value).strip()
    if s == "" or s in ("-", "—", "–"):
        return None

    # yyyy/mm/dd or yyyy-mm-dd
    m = re.match(r"^(\d{4})[/-](\d{1,2})[/-](\d{1,2})$", s)
    if m:
        y, mo, d = map(int, m.groups())
        return datetime.date(y, mo, d)

    # ROC yyyMMdd
    m = re.match(r"^(\d{2,3})(\d{2})(\d{2})$", s)
    if m:
        roc, mo, d = map(int, m.groups())
        return datetime.date(roc + 1911, mo, d)

    # ROC yyy/mm/dd
    m = re.match(r"^(\d{2,3})[/-](\d{1,2})[/-](\d{1,2})$", s)
    if m:
        roc, mo, d = map(int, m.groups())
        return datetime.date(roc + 1911, mo, d)

    return None


def parse_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    s = str(value).strip()
    if s == "" or s in ("-", "—", "–"):
        return None
    try:
        return float(s)
    except Exception:
        return None


def infer_gender_from_id(id_value: Any) -> str:
    if not id_value:
        return ""
    s = str(id_value).strip().upper()
    if len(s) < 2:
        return ""
    g = s[1]
    if g in ("1", "8"):
        return "男"
    if g in ("2", "9"):
        return "女"
    return ""


def clean_spaces(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value))


def parse_dmk_to_code(v: Any) -> Optional[int]:
    """
    疾病樣態碼：
    1:DM, 2:CKD, 3:DM+CKD(DKD), 其他：沒有DM或CKD（可能是4或空）
    """
    s = clean_spaces(v)
    if s == "":
        return None
    if s.isdigit():
        return int(s)
    su = s.upper()
    if su == "DM":
        return 1
    if su == "CKD":
        return 2
    if su in ("DKD", "DM+CKD", "DMCKD"):
        return 3
    return None


def classify_ascvd(v: Any) -> Optional[str]:
    """
    ASCVD 欄：
    - a/b 代表有 ASCVD
    - 0/空白/None 代表沒有 ASCVD
    """
    if v is None:
        return None
    s = clean_spaces(v).lower()
    if s == "" or s == "0":
        return None
    if s in ("a", "1") or "極高" in s:
        return "a"
    if s in ("b", "2") or "非常高" in s:
        return "b"
    if "a" in s:
        return "a"
    if "b" in s:
        return "b"
    return None


def calc_age(bday: Optional[datetime.date], ref: datetime.date) -> int:
    if not isinstance(bday, datetime.date):
        return -1
    y = ref.year - bday.year
    if (ref.month, ref.day) < (bday.month, bday.day):
        y -= 1
    return y


def add_full_grid(ws) -> None:
    """全表加細邊框（網格線）"""
    thin = Side(style="thin")
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).border = grid


# =========================
# 欄位集中管理（cols dict）
# =========================
def require_cols(cols: Dict[str, Optional[int]], required_keys: List[str], title: str = "模板欄位") -> None:
    missing = [k for k in required_keys if not cols.get(k)]
    if missing:
        raise ValueError(f"{title}找不到：{ '、'.join(missing) }")


def clear_data_rows(ws, data_start: int, max_row: int, cols: Dict[str, Optional[int]], keys_to_clear: List[str]) -> None:
    """清空資料列的指定欄位（不動表頭/樣式）"""
    col_ids = [cols.get(k) for k in keys_to_clear if cols.get(k)]
    for r in range(data_start, max_row + 1):
        for c in col_ids:
            ws.cell(r, c).value = None


def detect_template_columns(ws, data_start: int) -> Dict[str, Optional[int]]:
    """
    從模板「會員指標」偵測欄位，回傳 cols dict（key -> column index）
    """
    header_row = find_header_row_contains(ws, ["診所名稱或機構代碼", "姓名", "身份證號碼"], 250) or 5
    max_scan_row = data_start - 1

    cols: Dict[str, Optional[int]] = {}

    # 基本欄位
    cols["clinic"] = find_col_by_keywords(ws, header_row, ["診所名稱或機構代碼"])
    cols["name"]   = find_col_by_keywords(ws, header_row, ["姓名"])
    cols["id"]     = find_col_by_keywords(ws, header_row, ["身份證號碼"])
    cols["bday"]   = find_col_by_keywords(ws, header_row, ["生日"])
    cols["age"]    = find_col_by_keywords(ws, header_row, ["年齡"])
    cols["tel"]    = find_col_by_keywords(ws, header_row, ["電話"])
    cols["cnt"]    = find_col_by_keywords(ws, header_row, ["次數"])
    cols["sex"]    = find_col_by_keywords(ws, header_row, ["性別"])
    cols["abc"]    = find_col_by_keywords(ws, header_row, ["A/B/C"])

    # 疾病/ASCVD
    cols["dmk_code"] = find_col_by_keywords(ws, header_row, ["DM/CKD/DKD"])
    cols["ascvd"]    = find_col_by_keywords(ws, header_row, ["ASCVD"])

    # 五大篩檢日期
    cols["adult"] = find_col_by_keywords(ws, header_row, ["成人", "保健"])
    cols["pap"]   = find_col_by_keywords(ws, header_row, ["子宮", "抹片"])
    cols["flu"]   = find_col_by_keywords(ws, header_row, ["流感"]) or find_col_by_keywords(ws, header_row, ["流感"])
    cols["fit"]   = find_col_by_keywords(ws, header_row, ["糞便", "潛血"])
    cols["hep"]   = find_col_by_keywords(ws, header_row, ["肝炎"])

    # 檢驗值/日期
    cols["hba"]     = find_col_by_keywords(ws, header_row, ["HbA1c"])
    cols["hba_dt"]  = find_col_by_keywords(ws, header_row, ["HbA1c", "日期"])
    cols["ldl"]     = find_col_by_keywords(ws, header_row, ["LDL"])
    cols["ldl_dt"]  = find_col_by_keywords(ws, header_row, ["LDL", "日期"])
    cols["uacr"]    = find_col_by_keywords(ws, header_row, ["UACR"])
    cols["uacr_dt"] = find_col_by_keywords(ws, header_row, ["UACR", "日期"])

    # 文字/分數/備註
    cols["disease_text"] = (
        find_col_by_keywords(ws, header_row, ["DM/CKD/DKD/ASCVD"])
        or find_col_by_keywords(ws, header_row, ["疾病樣態", "ASCVD"])
    )
    cols["score"] = find_col_by_keywords(ws, header_row, ["分數"])
    cols["note"]  = find_col_by_keywords(ws, header_row, ["備註"])

    # AU/AV/AW：跨列掃描 1..(data_start-1)
    cols["au"] = (
        find_col_by_keywords_any_row(ws, max_scan_row, ["第一次提醒", "28天"])
        or find_col_by_keywords_any_row(ws, max_scan_row, ["回診追蹤", "28天"])
        or find_col_by_keywords_any_row(ws, max_scan_row, ["28天"])
    )
    cols["av"] = (
        find_col_by_keywords_any_row(ws, max_scan_row, ["第二次提醒", "56天"])
        or find_col_by_keywords_any_row(ws, max_scan_row, ["回診追蹤", "56天"])
    )
    cols["aw"] = (
        find_col_by_keywords_any_row(ws, max_scan_row, ["第三次提醒", "84天"])
        or find_col_by_keywords_any_row(ws, max_scan_row, ["回診追蹤", "84天"])
    )

    # breakdown：分數欄往右 13 欄（保留你原本的規則）
    cols["breakdown"] = None
    if cols.get("score"):
        cand = int(cols["score"]) + 13  # type: ignore[arg-type]
        if 1 <= cand <= ws.max_column:
            cols["breakdown"] = cand

    # 必填檢查
    require_cols(cols, [
        "clinic","name","id","bday","tel","abc","dmk_code","cnt","ascvd","sex",
        "hba","hba_dt","ldl","ldl_dt","uacr","uacr_dt",
        "disease_text","score","note","au","av","aw"
    ], title="新模板欄位/欄名")

    return cols


# =========================
# AV 備註（不靠公式）與「需要篩檢」判斷
# =========================
def add_months(d: datetime.date, months: int) -> datetime.date:
    """date + N months（EDATE 等效）"""
    y = d.year + (d.month - 1 + months) // 12
    m = (d.month - 1 + months) % 12 + 1
    is_leap = (y % 4 == 0 and (y % 100 != 0 or y % 400 == 0))
    mdays = [31, 29 if is_leap else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][m - 1]
    day = min(d.day, mdays)
    return datetime.date(y, m, day)


def need_bc_hep(age: int) -> bool:
    return 45 <= age <= 80


def need_fit(age: int) -> bool:
    return 45 <= age <= 75


def need_pap(age: int, sex: str) -> bool:
    if sex != "女":
        return False
    return age >= 25


def need_adult(age: int) -> bool:
    return age >= 30


def need_flu(age: int) -> bool:
    return age >= 65


def build_av_note(
    age: int,
    sex: str,
    hep_dt: Optional[datetime.date],
    fit_dt: Optional[datetime.date],
    pap_dt: Optional[datetime.date],
    adult_dt: Optional[datetime.date],
    flu_dt: Optional[datetime.date],
    today: datetime.date
) -> str:
    msgs: List[str] = []

    if need_bc_hep(age) and hep_dt is None:
        msgs.append("今年需檢測BC肝")

    if need_fit(age):
        if fit_dt is None or add_months(fit_dt, 24) <= today:
            msgs.append("今年需檢測糞便")

    if sex == "女" and need_pap(age, sex):
        if 25 <= age <= 29:
            if pap_dt is None or add_months(pap_dt, 36) <= today:
                msgs.append("今年需檢測子抹")
        elif age >= 30:
            if pap_dt is None or add_months(pap_dt, 12) <= today:
                msgs.append("今年需檢測子抹")

    if need_adult(age):
        if 30 <= age <= 39:
            if adult_dt is None or add_months(adult_dt, 60) <= today:
                msgs.append("今年需檢測成健")
        elif 40 <= age <= 64:
            if adult_dt is None or add_months(adult_dt, 36) <= today:
                msgs.append("今年需檢測成健")
        elif age >= 65:
            if adult_dt is None or add_months(adult_dt, 12) <= today:
                msgs.append("今年需檢測成健")

    if need_flu(age):
        if flu_dt is None or flu_dt.year < today.year:
            msgs.append("今年需檢測老流")

    return "\n".join(msgs).strip()


# =========================
# 疾病樣態文字
# =========================
def disease_group_text(e_code: Optional[int], f_ascvd: Optional[str]) -> Optional[str]:
    has_dm = (e_code == 1) or (e_code == 3)
    has_ckd = (e_code == 2) or (e_code == 3)
    has_ascvd = f_ascvd in ("a", "b")

    if has_dm and (not has_ckd) and (not has_ascvd):
        return "DM"
    if has_ckd and (not has_dm) and (not has_ascvd):
        return "CKD"
    if has_dm and has_ckd and (not has_ascvd):
        return "DM+CKD（DKD）"
    if (not has_dm) and (not has_ckd) and has_ascvd:
        return "ASCVD"
    if has_ckd and (not has_dm) and has_ascvd:
        return "CKD+ASCVD"
    if has_dm and (not has_ckd) and has_ascvd:
        return "DM+ASCVD"
    if has_dm and has_ckd and has_ascvd:
        return "DM+CKD+ASCVD"
    return None


# =========================
# AU 分數 / 明細
# =========================
SCORE_YEAR_SCREEN_2026 = 2026


def score_hba(hba_val: Any, hba_dt: Optional[datetime.date]) -> int:
    if not isinstance(hba_dt, datetime.date):
        return 0
    if hba_dt.year == 2026:
        return 5
    if hba_dt.year == 2025:
        if isinstance(hba_val, (int, float)) and float(hba_val) <= 7.0:
            return 5
    return 0


def score_ldl(ldl_val: Any, ldl_dt: Optional[datetime.date]) -> int:
    if not isinstance(ldl_dt, datetime.date):
        return 0
    if ldl_dt.year == 2026:
        return 5
    if ldl_dt.year == 2025:
        if isinstance(ldl_val, (int, float)) and float(ldl_val) <= 120.0:
            return 5
    return 0


def score_screening_year_only(dt: Optional[datetime.date], year: int, pts: int) -> int:
    if isinstance(dt, datetime.date) and dt.year == year:
        return pts
    return 0


def au_score_with_breakdown(
    *,
    e_code: Optional[int],
    f_ascvd: Optional[str],
    hba_val: Any,
    hba_dt: Optional[datetime.date],
    ldl_val: Any,
    ldl_dt: Optional[datetime.date],
    adult_dt: Optional[datetime.date],
    pap_dt: Optional[datetime.date],
    flu_dt: Optional[datetime.date],
    fit_dt: Optional[datetime.date],
    hep_dt: Optional[datetime.date],
    age: int,
    sex: str
) -> Tuple[int, Optional[str]]:
    has_dm = (e_code == 1) or (e_code == 3)
    has_ckd = (e_code == 2) or (e_code == 3)
    has_ascvd = f_ascvd in ("a", "b")

    total = 0
    parts: List[str] = []

    if has_dm:
        pts = score_hba(hba_val, hba_dt)
        if pts:
            total += pts
            parts.append(f"HbA1c {pts}分")

    if has_dm or has_ckd or has_ascvd:
        pts = score_ldl(ldl_val, ldl_dt)
        if pts:
            total += pts
            parts.append(f"LDL {pts}分")

    if age >= 0:
        if need_adult(age):
            pts = score_screening_year_only(adult_dt, SCORE_YEAR_SCREEN_2026, 6)
            if pts:
                total += pts
                parts.append(f"成健 {pts}分")

        if need_pap(age, sex):
            pts = score_screening_year_only(pap_dt, SCORE_YEAR_SCREEN_2026, 6)
            if pts:
                total += pts
                parts.append(f"子抹 {pts}分")

        if need_flu(age):
            pts = score_screening_year_only(flu_dt, SCORE_YEAR_SCREEN_2026, 4)
            if pts:
                total += pts
                parts.append(f"老流 {pts}分")

        if need_fit(age):
            pts = score_screening_year_only(fit_dt, SCORE_YEAR_SCREEN_2026, 6)
            if pts:
                total += pts
                parts.append(f"糞便 {pts}分")

        if need_bc_hep(age):
            pts = score_screening_year_only(hep_dt, SCORE_YEAR_SCREEN_2026, 6)
            if pts:
                total += pts
                parts.append(f"BC肝 {pts}分")

    breakdown = "\n".join(parts).strip() if parts else None
    return total, breakdown


# =========================
# AU：第一次提醒（28天）— HbA1c/LDL
# =========================
def build_au_28_note(
    *,
    e_code: Optional[int],
    f_ascvd: Optional[str],
    hba_dt: Optional[datetime.date],
    ldl_dt: Optional[datetime.date],
    today: datetime.date
) -> str:
    """
    AU 欄位（第一次提醒 28 天）規則：
    - 若不需檢查 -> 非疾病不需回診
    - 若 <2025 或 無日期 -> 超過2年未檢查
    - 若 2025/2026：
      dt+28 <= today -> 可立刻通知回診
      else           -> yyyy-mm-dd需回診（顯示 dt+28）
    """
    e = e_code if e_code in (1, 2, 3, 4) else 4
    has_ascvd = f_ascvd in ("a", "b")

    need_hba = e in (1, 3)
    need_ldl = (e in (1, 2, 3)) or (e == 4 and has_ascvd)

    if not need_hba and not need_ldl:
        return "非疾病不需回診"

    def is_2025_or_2026(dt: Optional[datetime.date]) -> bool:
        return isinstance(dt, datetime.date) and dt.year in (2025, 2026)

    def build_au_item_line(item_name: str, last_check_date: Optional[datetime.date], today_date: datetime.date) -> str:
        if not is_2025_or_2026(last_check_date):
            return f"{item_name}:超過2年未檢查"
        due = last_check_date + datetime.timedelta(days=28)  # type: ignore[operator]
        if due <= today_date:
            return f"{item_name}:可立刻通知回診"
        return f"{item_name}:{due.strftime('%Y-%m-%d')}需回診"

    eligible = is_2025_or_2026(hba_dt) or is_2025_or_2026(ldl_dt)
    if not eligible:
        lines: List[str] = []
        if need_hba:
            lines.append("HbA1c:超過2年未檢查")
        if need_ldl:
            lines.append("LDL:超過2年未檢查")
        return "\n".join(lines).strip()

    lines2: List[str] = []
    if need_hba:
        lines2.append(build_au_item_line("HbA1c", hba_dt, today))
    if need_ldl:
        lines2.append(build_au_item_line("LDL", ldl_dt, today))

    return "\n".join(lines2).strip()


# =========================
# AV/AW：從 AU 文字生成追蹤文字
# =========================
DATE_RE = re.compile(r"(\d{4}-\d{2}-\d{2})")


def au_should_skip_followup(au_txt: str) -> bool:
    """AU 不須通知 or 超過兩年未檢查 -> AV/AW 不填"""
    if not au_txt:
        return True
    s = au_txt.replace(" ", "")
    if "超過2年未檢查" in s or "超過兩年未檢查" in s:
        return True
    if "非疾病" in s:
        return True
    return False


def build_followup_text_from_au(au_txt: str, today: datetime.date, offset_days: int) -> Optional[str]:
    """
    將 AU 內容轉成追蹤欄位文字（AV/AW）：
    - 每一行獨立處理
    - 行內若有 yyyy-mm-dd -> 該日期 + offset_days
    - 行內若是「可立刻通知回診」 -> today + offset_days
    - 輸出仍為「HbA1c:YYYY-MM-DD需回診」「LDL:YYYY-MM-DD需回診」
    """
    if not au_txt:
        return None

    lines = [ln.strip() for ln in str(au_txt).splitlines() if ln.strip()]
    out: List[str] = []

    def build_followup_item_line(item_name: str, base_date: Optional[datetime.date], today_date: datetime.date) -> str:
        base = base_date if isinstance(base_date, datetime.date) else today_date
        due = base + datetime.timedelta(days=offset_days)
        return f"{item_name}:{due.strftime('%Y-%m-%d')}需回診"

    for ln in lines:
        if not (ln.startswith("HbA1c:") or ln.startswith("LDL:")):
            continue

        item_name = "HbA1c" if ln.startswith("HbA1c:") else "LDL"

        if "可立刻通知" in ln:
            out.append(build_followup_item_line(item_name, None, today))
            continue

        m = DATE_RE.search(ln)
        if m:
            try:
                base = datetime.datetime.strptime(m.group(1), "%Y-%m-%d").date()
                out.append(build_followup_item_line(item_name, base, today))
            except Exception:
                pass

    txt = "\n".join(out).strip()
    return txt if txt else None


# =========================
# KPI 百分位計算（統一函數）
# =========================
def _is_zero_like(v: Any) -> bool:
    """把 None / 空白 / '0' / '0.0' 視為 0"""
    if v is None:
        return True
    s = str(v).strip()
    return s == "" or s == "0" or s == "0.0"


def write_ratio_cell(ws, addr: str, numer: int, denom: int) -> None:
    """把 numer/denom 寫入指定儲存格（百分比格式），denom=0 則寫 0"""
    ws[addr].value = (numer / denom) if denom > 0 else 0
    ws[addr].number_format = "0.00%"


def calc_hba_percentiles_and_write(ws, cols: Dict[str, Optional[int]], data_start: int, last_row: int) -> None:
    """
    0301 樣板：
    AY8/AZ8：HbA1c 百分位（<=7 / <=7.3）
    （原本 BA8 <=7.5 已刪除，不再填）
    條件：
    - 疾病樣態碼：1 或 3
    - ASCVD：0 / 空白
    - HbA1c 日期：2025 或 2026
    - HbA1c 數值：可轉成數字才算分母
    """
    denom = 0
    n7 = 0
    n73 = 0

    for r in range(data_start, last_row + 1):
        e_raw = ws.cell(r, cols["dmk_code"]).value  # type: ignore[index]
        try:
            e = int(e_raw) if e_raw is not None and str(e_raw).strip() != "" else None
        except Exception:
            e = parse_dmk_to_code(e_raw)

        if e not in (1, 3):
            continue

        if not _is_zero_like(ws.cell(r, cols["ascvd"]).value):  # type: ignore[index]
            continue

        hba_dt = parse_date(ws.cell(r, cols["hba_dt"]).value)  # type: ignore[index]
        if not (isinstance(hba_dt, datetime.date) and hba_dt.year in (2025, 2026)):
            continue

        hba_val = parse_float(ws.cell(r, cols["hba"]).value)  # type: ignore[index]
        if hba_val is None:
            continue

        denom += 1
        if hba_val <= 7:
            n7 += 1
        if hba_val <= 7.3:
            n73 += 1

    write_ratio_cell(ws, "AY8", n7, denom)
    write_ratio_cell(ws, "AZ8", n73, denom)


def calc_ldl_percentiles_and_write(ws, cols: Dict[str, Optional[int]], data_start: int, last_row: int) -> None:
    """
    0301 樣板：
    BB8/BC8：LDL 百分位（<=100 / <=110）
    （原本 BD8 <=105、以及舊版 BE/BF 欄位已刪除，不再填）
    條件：
    - 疾病樣態碼：1/2/3/4
    - ASCVD：0 / a / b（都可），但「排除 E=4 且 F=0」
    - LDL 日期：2025 或 2026
    - LDL 數值：可轉成數字才算分母
    """
    denom = 0
    n100 = 0
    n110 = 0

    for r in range(data_start, last_row + 1):
        e_raw = ws.cell(r, cols["dmk_code"]).value  # type: ignore[index]
        try:
            e = int(e_raw) if e_raw is not None and str(e_raw).strip() != "" else None
        except Exception:
            e = parse_dmk_to_code(e_raw)

        if e not in (1, 2, 3, 4):
            continue

        f_raw = ws.cell(r, cols["ascvd"]).value  # type: ignore[index]
        f_cat = classify_ascvd(f_raw)
        ascvd_is_zero = (f_cat is None) and _is_zero_like(f_raw)

        # 排除：疾病樣態=4 且 ASCVD=0
        if e == 4 and ascvd_is_zero:
            continue

        ldl_dt = parse_date(ws.cell(r, cols["ldl_dt"]).value)  # type: ignore[index]
        if not (isinstance(ldl_dt, datetime.date) and ldl_dt.year in (2025, 2026)):
            continue

        ldl_val = parse_float(ws.cell(r, cols["ldl"]).value)  # type: ignore[index]
        if ldl_val is None:
            continue

        denom += 1
        if ldl_val <= 100:
            n100 += 1
        if ldl_val <= 110:
            n110 += 1

    write_ratio_cell(ws, "BB8", n100, denom)
    write_ratio_cell(ws, "BC8", n110, denom)


# =========================
# 主流程
# =========================
def process_excel(source_xlsx_path: str, template_xlsx_path: str) -> str:
    # ---------- 讀原始檔 ----------
    wb_src = openpyxl.load_workbook(source_xlsx_path, data_only=True)
    need_sheets = ["會員名單", "ascvd", "HealthCase", "成人健檢", "子宮抹片", "老人流感", "糞便潛血", "肝炎篩檢"]
    for sn in need_sheets:
        if sn not in wb_src.sheetnames:
            raise ValueError(f"原始檔缺少工作表：{sn}")

    sh_member = wb_src["會員名單"]
    sh_ascvd = wb_src["ascvd"]
    sh_health = wb_src["HealthCase"]
    sh_adult = wb_src["成人健檢"]
    sh_pap = wb_src["子宮抹片"]
    sh_flu = wb_src["老人流感"]
    sh_fit = wb_src["糞便潛血"]
    sh_hep = wb_src["肝炎篩檢"]

    # ---------- 讀模板（只保留會員指標） ----------
    wb_tpl = openpyxl.load_workbook(template_xlsx_path)
    if SHEET_TARGET not in wb_tpl.sheetnames:
        raise ValueError("模板檔缺少工作表：會員指標")

    for sn in list(wb_tpl.sheetnames):
        if sn != SHEET_TARGET:
            del wb_tpl[sn]
    ws = wb_tpl[SHEET_TARGET]

    # 0301 模板：第 10 列開始是資料列（你樣板內公式也用 D10:D...）
    DATA_START = 10

    # ---------- 偵測模板欄位（集中成 cols） ----------
    cols = detect_template_columns(ws, DATA_START)

    # ---------- 清資料列（不動表頭/樣式） ----------
    clear_data_rows(ws, DATA_START, ws.max_row, cols, keys_to_clear=[
        "clinic","name","id","bday","age","tel","cnt","sex","abc",
        "dmk_code","ascvd",
        "adult","pap","flu","fit","hep",
        "hba","hba_dt","ldl","ldl_dt","uacr","uacr_dt",
        "disease_text","score","breakdown","note",
        "au","av","aw",
    ])

    # ---------- 會員名單欄位 ----------
    member_header_row = 5
    mmap = build_header_map(sh_member, member_header_row)
    m_name = find_column_exact(mmap, ["會員姓名"])
    m_id = find_column_exact(mmap, ["會員身份証", "會員身份證", "會員身分證"])
    m_bday = find_column_exact(mmap, ["會員生日"])
    m_tel = find_column_exact(mmap, ["電話"])
    m_abc = find_column_exact(mmap, ["會員別"])
    m_dmk = find_column_exact(mmap, ["疾病樣態"])
    m_cnt = find_column_exact(mmap, ["就診次數"])

    if any(v is None for v in [m_name, m_id, m_bday, m_tel, m_abc, m_dmk, m_cnt]):
        raise ValueError("原始檔「會員名單」欄位不完整（會員姓名/會員身分證/會員生日/電話/會員別/疾病樣態/就診次數）")

    clinic_val = normalize_text(sh_member["A1"].value)

    # ---------- 寫入基本資料：建立 id_to_row ----------
    id_to_row: Dict[str, int] = {}
    meta: Dict[int, Dict[str, Any]] = {}
    now_tw = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=8))).date()

    out_r = DATA_START
    for r in range(member_header_row + 1, sh_member.max_row + 1):
        sid = sh_member.cell(r, m_id).value  # type: ignore[arg-type]
        name = sh_member.cell(r, m_name).value  # type: ignore[arg-type]
        if (sid in (None, "")) and (name in (None, "")):
            continue

        sid_s = str(sid).strip()
        bday = parse_date(sh_member.cell(r, m_bday).value)  # type: ignore[arg-type]
        e_code = parse_dmk_to_code(sh_member.cell(r, m_dmk).value)  # type: ignore[arg-type]
        age = calc_age(bday, now_tw) if bday else -1

        ws.cell(out_r, cols["clinic"]).value = clinic_val  # type: ignore[arg-type]
        ws.cell(out_r, cols["name"]).value = name  # type: ignore[arg-type]
        ws.cell(out_r, cols["id"]).value = sid  # type: ignore[arg-type]
        ws.cell(out_r, cols["bday"]).value = bday  # type: ignore[arg-type]
        if cols.get("age"):
            ws.cell(out_r, cols["age"]).value = (age if age >= 0 else None)  # type: ignore[arg-type]
        ws.cell(out_r, cols["tel"]).value = sh_member.cell(r, m_tel).value  # type: ignore[arg-type]
        ws.cell(out_r, cols["abc"]).value = sh_member.cell(r, m_abc).value  # type: ignore[arg-type]
        ws.cell(out_r, cols["dmk_code"]).value = e_code  # type: ignore[arg-type]
        ws.cell(out_r, cols["cnt"]).value = sh_member.cell(r, m_cnt).value  # type: ignore[arg-type]
        ws.cell(out_r, cols["sex"]).value = infer_gender_from_id(sid)  # type: ignore[arg-type]

        if sid_s and sid_s not in id_to_row:
            id_to_row[sid_s] = out_r
            meta[out_r] = {"bday": bday, "age": age, "e_code": e_code, "ascvd": None}

        out_r += 1

    last_row = out_r - 1

    # ---------- ASCVD 回填 ----------
    ascvd_header_row = 5
    amap = build_header_map(sh_ascvd, ascvd_header_row)
    a_id = find_column_exact(amap, ["ID", "id"])
    a_asc = find_column_exact(amap, ["ASCVD", "ascvd"])
    if a_id is None or a_asc is None:
        raise ValueError("原始檔「ascvd」找不到 ID / ASCVD 欄位（表頭列預期第5列）")

    for r in range(ascvd_header_row + 1, sh_ascvd.max_row + 1):
        pid = sh_ascvd.cell(r, a_id).value
        val = sh_ascvd.cell(r, a_asc).value
        if pid is None or str(pid).strip() == "":
            continue
        if val is None or str(val).strip() == "" or str(val).strip() == "0":
            continue
        tr = id_to_row.get(str(pid).strip())
        if tr:
            ws.cell(tr, cols["ascvd"]).value = val  # type: ignore[arg-type]
            meta.setdefault(tr, {})["ascvd"] = val

    # ---------- 5 大篩檢日期回填 ----------
    def fill_screening(sheet, target_col: Optional[int]) -> None:
        if not target_col:
            return
        hmap = build_header_map(sheet, 1)
        sid_col = find_column_exact(hmap, ["ID", "身分證號", "身份證號"])
        dt_col = find_column_exact(hmap, ["最後篩檢日期"])
        if sid_col is None or dt_col is None:
            raise ValueError(f"「{sheet.title}」找不到 ID 或 最後篩檢日期 欄位")

        for rr in range(2, sheet.max_row + 1):
            pid = sheet.cell(rr, sid_col).value
            dt = parse_date(sheet.cell(rr, dt_col).value)
            if pid is None or str(pid).strip() == "" or dt is None:
                continue
            tr = id_to_row.get(str(pid).strip())
            if tr:
                ws.cell(tr, target_col).value = dt

    fill_screening(sh_adult, cols.get("adult"))
    fill_screening(sh_pap, cols.get("pap"))
    fill_screening(sh_flu, cols.get("flu"))
    fill_screening(sh_fit, cols.get("fit"))
    fill_screening(sh_hep, cols.get("hep"))

    # ---------- HealthCase 檢驗回填 ----------
    hc_map = build_header_map(sh_health, 1)
    hc_id = find_column_exact(hc_map, ["家醫收案會員ID", "ID"])
    hc_hba = find_column_exact(hc_map, ["最近一次HbA1c檢查結果(%)"])
    hc_hba_dt = find_column_exact(hc_map, ["最近一次HbA1c檢查日期"])
    hc_ldl = find_column_exact(hc_map, ["最近一次LDL檢查結果(mg/dL)"])
    hc_ldl_dt = find_column_exact(hc_map, ["最近一次LDL檢查日期"])
    hc_uacr = find_column_exact(hc_map, ["最近一次UACR檢查結果(mg/gm)"])
    hc_uacr_dt = find_column_exact(hc_map, ["最近一次UACR檢查日期"])

    if any(v is None for v in [hc_id, hc_hba, hc_hba_dt, hc_ldl, hc_ldl_dt, hc_uacr, hc_uacr_dt]):
        raise ValueError("原始檔「HealthCase」欄位不完整（家醫收案會員ID / HbA1c結果+日期 / LDL結果+日期 / UACR結果+日期）")

    for r in range(2, sh_health.max_row + 1):
        pid = sh_health.cell(r, hc_id).value  # type: ignore[arg-type]
        if pid is None or str(pid).strip() == "":
            continue
        tr = id_to_row.get(str(pid).strip())
        if not tr:
            continue

        hba = parse_float(sh_health.cell(r, hc_hba).value)  # type: ignore[arg-type]
        if hba is not None and hba != 0:
            ws.cell(tr, cols["hba"]).value = hba  # type: ignore[arg-type]
        hba_dt = parse_date(sh_health.cell(r, hc_hba_dt).value)  # type: ignore[arg-type]
        if hba_dt:
            ws.cell(tr, cols["hba_dt"]).value = hba_dt  # type: ignore[arg-type]

        ldl = parse_float(sh_health.cell(r, hc_ldl).value)  # type: ignore[arg-type]
        if ldl is not None and ldl != 0:
            ws.cell(tr, cols["ldl"]).value = ldl  # type: ignore[arg-type]
        ldl_dt = parse_date(sh_health.cell(r, hc_ldl_dt).value)  # type: ignore[arg-type]
        if ldl_dt:
            ws.cell(tr, cols["ldl_dt"]).value = ldl_dt  # type: ignore[arg-type]

        uacr = parse_float(sh_health.cell(r, hc_uacr).value)  # type: ignore[arg-type]
        if uacr is not None and uacr != 0:
            ws.cell(tr, cols["uacr"]).value = uacr  # type: ignore[arg-type]
        uacr_dt = parse_date(sh_health.cell(r, hc_uacr_dt).value)  # type: ignore[arg-type]
        if uacr_dt:
            ws.cell(tr, cols["uacr_dt"]).value = uacr_dt  # type: ignore[arg-type]

    # ---------- 疾病樣態文字 + 備註 + 分數/明細 + AU/AV/AW ----------
    for rr in range(DATA_START, last_row + 1):
        e_raw = ws.cell(rr, cols["dmk_code"]).value  # type: ignore[index]
        e_code = int(e_raw) if isinstance(e_raw, (int, float)) else parse_dmk_to_code(e_raw)

        f_val = ws.cell(rr, cols["ascvd"]).value  # type: ignore[index]
        f_asc = classify_ascvd(f_val)

        ws.cell(rr, cols["disease_text"]).value = disease_group_text(e_code, f_asc)  # type: ignore[arg-type]

        age = meta.get(rr, {}).get("age", -1)
        if not isinstance(age, int):
            try:
                age = int(age)
            except Exception:
                age = -1
        sex = normalize_text(ws.cell(rr, cols["sex"]).value)  # type: ignore[arg-type]

        adult_dt = parse_date(ws.cell(rr, cols["adult"]).value) if cols.get("adult") else None  # type: ignore[index]
        pap_dt = parse_date(ws.cell(rr, cols["pap"]).value) if cols.get("pap") else None  # type: ignore[index]
        flu_dt = parse_date(ws.cell(rr, cols["flu"]).value) if cols.get("flu") else None  # type: ignore[index]
        fit_dt = parse_date(ws.cell(rr, cols["fit"]).value) if cols.get("fit") else None  # type: ignore[index]
        hep_dt = parse_date(ws.cell(rr, cols["hep"]).value) if cols.get("hep") else None  # type: ignore[index]

        note = build_av_note(
            age=age if age >= 0 else 0,
            sex=sex,
            hep_dt=hep_dt,
            fit_dt=fit_dt,
            pap_dt=pap_dt,
            adult_dt=adult_dt,
            flu_dt=flu_dt,
            today=now_tw,
        )
        ws.cell(rr, cols["note"]).value = note if note else None  # type: ignore[arg-type]

        hba_val = ws.cell(rr, cols["hba"]).value  # type: ignore[index]
        hba_dt = parse_date(ws.cell(rr, cols["hba_dt"]).value) if cols.get("hba_dt") else None  # type: ignore[index]
        ldl_val = ws.cell(rr, cols["ldl"]).value  # type: ignore[index]
        ldl_dt = parse_date(ws.cell(rr, cols["ldl_dt"]).value) if cols.get("ldl_dt") else None  # type: ignore[index]

        score, breakdown = au_score_with_breakdown(
            e_code=e_code,
            f_ascvd=f_asc,
            hba_val=hba_val,
            hba_dt=hba_dt,
            ldl_val=ldl_val,
            ldl_dt=ldl_dt,
            adult_dt=adult_dt,
            pap_dt=pap_dt,
            flu_dt=flu_dt,
            fit_dt=fit_dt,
            hep_dt=hep_dt,
            age=age,
            sex=sex,
        )
        ws.cell(rr, cols["score"]).value = score  # type: ignore[arg-type]
        if cols.get("breakdown"):
            ws.cell(rr, cols["breakdown"]).value = breakdown  # type: ignore[arg-type]

        # AU：第一次提醒（28天）
        au_txt = build_au_28_note(
            e_code=e_code,
            f_ascvd=f_asc,
            hba_dt=hba_dt,
            ldl_dt=ldl_dt,
            today=now_tw,
        )
        ws.cell(rr, cols["au"]).value = au_txt if au_txt else None  # type: ignore[arg-type]

        # AV/AW：從 AU 文字延後日期
        if au_txt and (not au_should_skip_followup(au_txt)):
            av_txt = build_followup_text_from_au(au_txt, today=now_tw, offset_days=28)
            aw_txt = build_followup_text_from_au(au_txt, today=now_tw, offset_days=56)
            ws.cell(rr, cols["av"]).value = av_txt if av_txt else None  # type: ignore[arg-type]
            ws.cell(rr, cols["aw"]).value = aw_txt if aw_txt else None  # type: ignore[arg-type]
        else:
            ws.cell(rr, cols["av"]).value = None  # type: ignore[arg-type]
            ws.cell(rr, cols["aw"]).value = None  # type: ignore[arg-type]

    # ---------- 日期欄格式 ----------
    date_cols = [
        cols.get("bday"),
        cols.get("adult"),
        cols.get("pap"),
        cols.get("flu"),
        cols.get("fit"),
        cols.get("hep"),
        cols.get("hba_dt"),
        cols.get("ldl_dt"),
        cols.get("uacr_dt"),
    ]
    for r in range(DATA_START, last_row + 1):
        for c in date_cols:
            if not c:
                continue
            cell = ws.cell(r, c)
            if isinstance(cell.value, (datetime.date, datetime.datetime)):
                cell.number_format = "yyyy-mm-dd"

    # ---------- KPI 百分位（寫到固定儲存格） ----------
    calc_hba_percentiles_and_write(ws, cols, DATA_START, last_row)
    calc_ldl_percentiles_and_write(ws, cols, DATA_START, last_row)

    # ---------- 邊框 ----------
    add_full_grid(ws)

    # ---------- 輸出 ----------
    base_dir = os.path.dirname(os.path.abspath(source_xlsx_path))
    ts = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=8))).strftime("%m%d_%H%M")
    out_path = os.path.join(base_dir, f"選會員{ts}.xlsx")
    wb_tpl.save(out_path)
    return out_path


def main() -> None:
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()

    src = filedialog.askopenfilename(
        title="選擇原始 Excel 檔案（沒有會員指標的那份）",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not src:
        return

    script_dir = os.path.dirname(os.path.abspath(__file__))
    template = os.path.join(script_dir, TEMPLATE_NAME)

    if not os.path.exists(template):
        messagebox.showerror(
            "錯誤",
            f"找不到模板檔：\n{template}\n\n"
            f"請把模板放到此 .py 同資料夾，檔名需為：\n{TEMPLATE_NAME}"
        )
        return

    try:
        out = process_excel(src, template)
        messagebox.showinfo("完成", f"已輸出：\n{out}")
        open_file_cross_platform(out)
    except (ValueError, KeyError, OSError, InvalidFileException) as e:
        messagebox.showerror("錯誤", str(e))
    except Exception as e:
        messagebox.showerror("錯誤", f"未預期錯誤：{e}")


if __name__ == "__main__":
    main()