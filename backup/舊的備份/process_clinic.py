#!/usr/bin/env python3
"""
從「門診診療次數月報表」產生「撈出資料」格式的 Excel 檔
輸出檔案名稱：輸入檔前6個字 + "_撈出.xlsx"
"""

import pandas as pd
import os
import sys
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ── 欄位索引（row 3 為標題列，0-based）
COL_NAME  = 6   # 姓名
COL_ID    = 7   # 身分證號
COL_DATE  = 3   # 日期
COL_DX    = 14  # 病1（主診斷）
COL_COUNT = 17  # 次數
COL_TOTAL = 43  # 總額

# ──────────────────────────────────────────────
def load_input(path: str) -> dict:
    """讀取所有工作表，回傳 {sheet_name: DataFrame}"""
    print(f"[讀取] {path}")
    all_sheets = pd.read_excel(path, sheet_name=None, header=None)
    print(f"  → 共 {len(all_sheets)} 個工作表：{list(all_sheets.keys())}")
    return all_sheets

# ──────────────────────────────────────────────
def process_monthly(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    """
    將單月工作表整理為：身份證號 | 姓名 | 件數(次數) | 申請金額(總額)
    同一人可能多筆，加總次數與總額
    """
    # 取 row 4 以後為資料（0=標題頁, 1=列印日期, 2=空, 3=欄名, 4+=資料）
    data = df.iloc[4:].copy()

    # 過濾有效身份證（長度10，非空）
    data = data[data[COL_ID].notna()]
    data = data[data[COL_ID].astype(str).str.len() == 10]

    if data.empty:
        print(f"  [{sheet_name}] 無有效資料，略過")
        return pd.DataFrame(columns=['身份證號', '姓名', '件數(次數)', '申請金額(總額)'])

    # 選出需要的欄位並轉型
    subset = data[[COL_ID, COL_NAME, COL_COUNT, COL_TOTAL]].copy()
    subset.columns = ['身份證號', '姓名', '件數(次數)', '申請金額(總額)']
    subset['件數(次數)']    = pd.to_numeric(subset['件數(次數)'],    errors='coerce').fillna(0)
    subset['申請金額(總額)'] = pd.to_numeric(subset['申請金額(總額)'], errors='coerce').fillna(0)

    # 以身份證號 + 姓名 group by，加總
    grouped = (
        subset.groupby(['身份證號', '姓名'], as_index=False, sort=False)
              .agg({'件數(次數)': 'sum', '申請金額(總額)': 'sum'})
    )
    grouped['件數(次數)']    = grouped['件數(次數)'].astype(int)
    grouped['申請金額(總額)'] = grouped['申請金額(總額)'].astype(int)

    print(f"  [{sheet_name}] {len(grouped)} 人，次數合計 {grouped['件數(次數)'].sum()}，總額合計 {grouped['申請金額(總額)'].sum():,}")
    return grouped

# ──────────────────────────────────────────────
def build_main_diag(all_sheets: dict) -> pd.DataFrame:
    """
    從所有月份資料找每位病患的最後就診日及主診斷
    輸出：姓名 | 身份證號 | 最後就診日 | 診斷代碼(病1)
    """
    print("\n[處理] 建立「主次診斷」工作表...")
    all_records = []
    for sheet_name, df in all_sheets.items():
        data = df.iloc[4:].copy()
        data = data[data[COL_ID].notna()]
        data = data[data[COL_ID].astype(str).str.len() == 10]
        if data.empty:
            continue
        records = data[[COL_NAME, COL_ID, COL_DATE, COL_DX]].copy()
        records.columns = ['姓名', '身份證號', '日期', '病1']
        all_records.append(records)

    if not all_records:
        return pd.DataFrame(columns=['姓名', '身份證號', '最後就診日(日期)\n(以最新的日期為主)', '診斷代碼(病1)'])

    all_df = pd.concat(all_records, ignore_index=True)
    all_df['日期'] = pd.to_numeric(all_df['日期'], errors='coerce')
    all_df = all_df.dropna(subset=['日期'])
    # 取每人最新就診那一筆
    latest = (
        all_df.sort_values('日期', ascending=False)
              .drop_duplicates('身份證號')
              .reset_index(drop=True)
    )
    latest = latest.rename(columns={
        '日期': '最後就診日(日期)\n(以最新的日期為主)',
        '病1':  '診斷代碼(病1)'
    })[['姓名', '身份證號', '最後就診日(日期)\n(以最新的日期為主)', '診斷代碼(病1)']]
    latest['最後就診日(日期)\n(以最新的日期為主)'] = latest['最後就診日(日期)\n(以最新的日期為主)'].astype(int)

    # 依身份證號倒序排列（符合原始輸出格式）
    latest = latest.sort_values('身份證號', ascending=False).reset_index(drop=True)
    print(f"  → 主次診斷共 {len(latest)} 位病患")
    return latest

# ──────────────────────────────────────────────
def apply_header_style(ws, header_fill_hex='4472C4'):
    """套用標題列樣式"""
    fill = PatternFill('solid', start_color=header_fill_hex, end_color=header_fill_hex)
    font = Font(bold=True, color='FFFFFF', name='Arial')
    for cell in ws[1]:
        cell.fill  = fill
        cell.font  = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def write_sheet(ws, df: pd.DataFrame, header: list):
    """寫入標題與資料，並自動調整欄寬"""
    ws.append(header)
    for row in df.itertuples(index=False):
        ws.append(list(row))
    apply_header_style(ws)
    # 自動欄寬
    for col_idx, _ in enumerate(header, 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(header[col_idx - 1])),
            *(len(str(ws.cell(r, col_idx).value or '')) for r in range(2, ws.max_row + 1))
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

# ──────────────────────────────────────────────
def pick_input_file() -> str:
    """彈出檔案選擇視窗，回傳選擇的路徑；取消則結束程式"""
    root = tk.Tk()
    root.withdraw()          # 隱藏主視窗
    root.attributes('-topmost', True)
    path = filedialog.askopenfilename(
        title='請選擇門診診療次數月報表',
        filetypes=[('Excel 檔案', '*.xlsx *.xls *.XLSX *.XLS'), ('所有檔案', '*.*')]
    )
    root.destroy()
    if not path:
        print("❌ 未選擇檔案，程式結束。")
        sys.exit(0)
    return path

# ──────────────────────────────────────────────
def main():
    # 彈出檔案選擇視窗
    input_path = pick_input_file()

    # 輸出檔名：輸入檔名前6個字 + "_撈出.xlsx"
    base_name   = os.path.basename(input_path)
    name_no_ext = os.path.splitext(base_name)[0]
    prefix6     = name_no_ext[:6]
    output_filename = f"{prefix6}_撈出.xlsx"
    output_dir  = os.path.dirname(input_path)   # 輸出至與輸入相同目錄
    output_path = os.path.join(output_dir, output_filename)

    print("=" * 55)
    print(f"輸入檔案：{base_name}")
    print(f"輸出目錄：{output_dir}")
    print(f"輸出檔案：{output_filename}")
    print("=" * 55)

    # 讀取輸入
    all_sheets = load_input(input_path)

    # 建立輸出 Workbook
    wb = Workbook()
    wb.remove(wb.active)  # 移除預設空工作表

    monthly_header = ['身份證號', '姓名', '件數(次數)', '申請金額(總額)']
    diag_header    = ['姓名', '身份證號', '最後就診日(日期)\n(以最新的日期為主)', '診斷代碼(病1)']

    print("\n[處理] 各月份工作表...")
    for sheet_name, df in all_sheets.items():
        result = process_monthly(df, sheet_name)
        ws = wb.create_sheet(sheet_name)
        write_sheet(ws, result, monthly_header)

    # 空工作表（11503, 11504），放在月份工作表之後、主次診斷之前
    existing = set(all_sheets.keys())
    for extra in ['11503', '11504']:
        if extra not in existing:
            ws_extra = wb.create_sheet(extra)
            ws_extra.append(monthly_header)
            apply_header_style(ws_extra)
            print(f"  [建立] 空工作表 {extra}")

    # 主次診斷放在最後一個
    diag_df = build_main_diag(all_sheets)
    ws_diag = wb.create_sheet('主次診斷')
    write_sheet(ws_diag, diag_df, diag_header)

    wb.save(output_path)
    print("\n" + "=" * 55)
    print(f"✅ 輸出完成：{output_path}")
    print(f"   工作表數量：{len(wb.sheetnames)}")
    print(f"   工作表清單：{wb.sheetnames}")
    print("=" * 55)

if __name__ == '__main__':
    main()
