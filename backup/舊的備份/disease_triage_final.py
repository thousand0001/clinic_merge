# -*- coding: utf-8 -*-
"""
disease_triage_gui_v5_1.py

✅ V5.1 版本重點（在不改你原本「分流/備註邏輯」前提下做格式容錯）
- ✅ 忽略 Excel 前導單引號 `'`（ID / 疾病樣態 / ASCVD 都會處理）
- ✅ 欄名換行/全形空白/多空白容錯（疾病\n樣態 也抓得到）
- ✅ 避免 used range：只處理實際資料列（用 ID 欄找最後資料列）
- ✅ 7 分流 + 分析表（邏輯不動）
- ✅ 分流頁樣式：備註左對齊可換行，其餘置中；不加粗/不放大姓名
- ✅ 輸出：未達控制mmdd_hhmm.xlsx（放 B 檔同資料夾），並自動開啟
- ✅ 分析表：刪除原本第2列；原本第4,8,12,...備註文字列列高=1.2倍
- ✅ 寫出前：日期/文字欄位值為 '0' 或 '0.0' 轉成 '-'
- ✅ 備註欄寬固定 185pt（openpyxl 欄寬近似：30）
- ✅ 依「人工作業範例」：不是整列上色，是「指定欄位」上色；黑色隔線(border)不動

支援：
- .xlsx / .xls / .ods（.ods 需 pip install odfpy）
"""

import re
import os
import sys
import subprocess
import traceback
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE_YEAR = 2026

SHEET_ORDER = ["DM", "CKD", "ASCVD", "DKD", "CKD+ASCVD", "DM+ASCVD", "DKD+ASCVD"]

SCREENING_EXPORT_COLS = [
    "成人預防保健(6分)",
    "子宮頸抹片檢查率(6分)",
    "65歲以上老人流感注射率(4分)",
    "糞便潛血檢查率(6分)",
    "BC肝炎篩檢(6分)",
]

EXPORT_COLS = [
    "ID", "疾病樣態", "ASCVD", "姓名", "生日",
    "最近一次HbA1c檢查結果(%)", "最近一次HbA1c檢查日期",
    "最近一次LDL檢查結果(mg/dL)", "最近一次LDL檢查日期",
    "最近一次UACR檢查結果(mg/gm)", "最近一次UACR檢查日期",
    "備註",
] + SCREENING_EXPORT_COLS

NO_DISEASE_SHEET = "無疾病(4,0)"
NO_DISEASE_EXPORT_COLS = [
    "ID", "疾病樣態", "ASCVD", "姓名", "生日",
] + SCREENING_EXPORT_COLS


# ------------------------- utils -------------------------
def open_file(path: Path):
    p = str(path)
    if sys.platform.startswith("darwin"):
        subprocess.run(["open", p], check=False)
    elif os.name == "nt":
        os.startfile(p)  # type: ignore[attr-defined]
    else:
        subprocess.run(["xdg-open", p], check=False)


def _strip_excel_apostrophe(x: str) -> str:
    # Excel 常用 ' 強制文字格式：在儲存格看不到，但值會包含它
    if x is None:
        return ""
    s = str(x)
    return s.lstrip("'")


def norm_id(s: pd.Series) -> pd.Series:
    s = s.astype(str).str.strip()

    # ✅ 忽略 Excel 前導單引號（ID 常見）
    s = s.str.lstrip("'")

    # 去掉尾端 .0
    s = s.str.replace(r"\.0$", "", regex=True)

    # 去掉各種空白（含全形）
    s = s.str.replace("\u3000", "", regex=False)
    s = s.str.replace(r"\s+", "", regex=True)

    s = s.replace({"nan": pd.NA, "None": pd.NA, "": pd.NA})
    return s


def _norm_colname(x: str) -> str:
    return (str(x).strip().lower()
            .replace("\n", "")
            .replace("\r", "")
            .replace("\\n", "")
            .replace("\\r", "")
            .replace(" ", "")
            .replace("\u3000", ""))


def pick_col_exact(cols: List[str], candidates: List[str]) -> Optional[str]:
    norm_map = {_norm_colname(c): c for c in cols}
    for cand in candidates:
        k = _norm_colname(cand)
        if k in norm_map:
            return norm_map[k]
    return None


def pick_col_contains(cols: List[str], keywords: List[str]) -> Optional[str]:
    norm_cols = [(_norm_colname(c), c) for c in cols]
    for kw in keywords:
        nkw = _norm_colname(kw)
        for nc, orig in norm_cols:
            if nkw and (nkw in nc):
                return orig
    return None


def to_dt(series: pd.Series) -> pd.Series:
    """
    穩定日期解析：支援
    - YYYY-MM-DD / YYYY/MM/DD
    - 民國 114/01/30 或 1140130
    """
    s = series.astype(str).str.strip()
    s = s.replace({"nan": pd.NA, "None": pd.NA, "": pd.NA})

    def normalize_one(x: str) -> Optional[str]:
        if x is None or pd.isna(x):
            return None
        x = str(x).strip()
        x = _strip_excel_apostrophe(x)
        if not x or x.lower() in ("nan", "none"):
            return None

        x2 = x.replace(".", "/").replace("-", "/")

        m = re.match(r"^(\d{2,3})/(\d{1,2})/(\d{1,2})$", x2)  # 民國 114/01/30
        if m:
            y, mo, d = map(int, m.groups())
            y += 1911
            return f"{y:04d}-{mo:02d}-{d:02d}"

        m = re.match(r"^(\d{2,3})(\d{2})(\d{2})$", x)  # 民國 1140130
        if m:
            y, mo, d = map(int, m.groups())
            y += 1911
            return f"{y:04d}-{mo:02d}-{d:02d}"

        m = re.match(r"^(\d{4})/(\d{1,2})/(\d{1,2})$", x2)  # 西元 2026/01/30
        if m:
            y, mo, d = map(int, m.groups())
            return f"{y:04d}-{mo:02d}-{d:02d}"

        return x

    s2 = s.map(normalize_one)
    return pd.to_datetime(s2, format="%Y-%m-%d", errors="coerce")


def parse_mixed_date(v):
    """
    穩定版單值日期解析：
    - datetime/date/Timestamp
    - Excel serial date
    - YYYY-MM-DD / YYYY/MM/DD / YYYYMMDD
    - 民國 115/03/06、115-03-06、1150306、115年03月06日
    失敗回傳 pd.NaT
    """
    import datetime as dt

    if v is None or pd.isna(v):
        return pd.NaT

    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime()
    if isinstance(v, dt.datetime):
        return v
    if isinstance(v, dt.date):
        return dt.datetime(v.year, v.month, v.day)

    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            fv = float(v)
            if 20000 <= fv <= 60000:
                base = dt.datetime(1899, 12, 30)
                return base + dt.timedelta(days=fv)
        except Exception:
            pass

    s = str(v).strip()
    s = _strip_excel_apostrophe(s)
    if not s or s.lower() in ("nan", "none", "nat") or s in ("-", "0", "0.0"):
        return pd.NaT

    s2 = (s.replace('年', '/')
            .replace('月', '/')
            .replace('日', '')
            .replace('.', '/')
            .replace('\\', '/')
            .replace('-', '/'))
    s2 = re.sub(r'/+', '/', s2).strip('/')

    for fmt in ("%Y/%m/%d", "%Y%m%d"):
        try:
            return dt.datetime.strptime(s2 if fmt == "%Y/%m/%d" else s2.replace('/', ''), fmt)
        except Exception:
            pass

    m = re.match(r'^(\d{2,3})/(\d{1,2})/(\d{1,2})$', s2)
    if m:
        y, mo, d = map(int, m.groups())
        return dt.datetime(y + 1911, mo, d)

    m = re.match(r'^(\d{2,3})(\d{2})(\d{2})$', s2.replace('/', ''))
    if m:
        y, mo, d = map(int, m.groups())
        return dt.datetime(y + 1911, mo, d)

    return pd.NaT


def age_on(bday: pd.Timestamp, ref_date: pd.Timestamp):
    if pd.isna(bday):
        return None
    return ref_date.year - bday.year - ((ref_date.month, ref_date.day) < (bday.month, bday.day))


def ascvd_norm(x) -> str:
    v = str(x).strip().lower()
    v = _strip_excel_apostrophe(v)  # ✅ 忽略 Excel 前導單引號
    v = v.strip()
    return v if v in ("a", "b") else ""


# ✅ ODS 支援：依副檔名選 engine
def _excel_engine_for(path: Path) -> Optional[str]:
    ext = path.suffix.lower()
    if ext == ".ods":
        return "odf"  # 需要：pip install odfpy
    return None


def read_table(path: Path, **kwargs) -> pd.DataFrame:
    eng = _excel_engine_for(path)
    if eng:
        kwargs.setdefault("engine", eng)
    return pd.read_excel(path, **kwargs)


def open_workbook(path: Path) -> pd.ExcelFile:
    eng = _excel_engine_for(path)
    if eng:
        return pd.ExcelFile(path, engine=eng)
    return pd.ExcelFile(path)


def find_sheet_and_header_for_b(path_b: Path) -> Tuple[str, int]:
    xl = open_workbook(path_b)

    def row_has_required(tokens):
        tokens = [_norm_colname(t) for t in tokens if str(t).strip()]
        has_id = any(t == "id" for t in tokens)
        has_dis = any("疾病樣態" in t for t in tokens)
        has_ascvd = any("ascvd" in t for t in tokens)
        return has_id and has_dis and has_ascvd

    for sh in xl.sheet_names:
        probe = read_table(path_b, sheet_name=sh, header=None, nrows=25, dtype=str)
        for r in range(len(probe)):
            if row_has_required(probe.iloc[r].tolist()):
                return sh, r

    # 找不到就回落（維持你原本策略）
    return xl.sheet_names[0], 5


# ------------------------- rules (LOGIC NOT CHANGED) -------------------------
def required_items(sheet_name: str) -> List[str]:
    if sheet_name in ("DM", "DM+ASCVD"):
        return ["HbA1c", "LDL"]
    if sheet_name in ("CKD", "CKD+ASCVD"):
        return ["LDL", "UACR"]
    if sheet_name in ("DKD", "DKD+ASCVD"):
        return ["HbA1c", "LDL", "UACR"]
    if sheet_name == "ASCVD":
        return ["LDL"]
    raise ValueError(sheet_name)


def ldl_target(row: pd.Series, sheet_name: str) -> float:
    a = ascvd_norm(row.get("ASCVD", ""))
    if a in ("a", "b") and sheet_name in ("ASCVD", "DM+ASCVD", "CKD+ASCVD", "DKD+ASCVD"):
        return 55.0 if a == "a" else 70.0
    if sheet_name == "CKD":
        return 130.0
    return 100.0


def remark_code(row: pd.Series, sheet_name: str, base_year: int) -> int:
    items = required_items(sheet_name)

    ref = pd.Timestamp(f"{base_year}-06-30")
    age = age_on(row["_生日_dt"], ref)
    h_target = 8.0 if (age is not None and age >= 80) else 7.0
    l_target = ldl_target(row, sheet_name)
    u_target = 30.0

    tested = {}
    bad = {}
    years = []

    for it in items:
        dt = row.get(f"_{it}_dt", pd.NaT)
        val = row.get(f"_{it}_num", pd.NA)

        is_y = (not pd.isna(dt)) and (dt.year == base_year)
        tested[it] = is_y

        if not pd.isna(dt):
            years.append(int(dt.year))

        if is_y and not pd.isna(val):
            if it == "HbA1c":
                bad[it] = val >= h_target
            elif it == "LDL":
                bad[it] = val >= l_target
            elif it == "UACR":
                bad[it] = val >= u_target
            else:
                bad[it] = False
        else:
            bad[it] = False

    all_tested = all(tested.values())
    all_missing = not any(tested.values())
    any_bad = any(bad.values())
    last_year = max(years) if years else None

    if all_tested and (not any_bad):
        return 0
    if (not all_tested) and any_bad:
        return 3
    if (not all_tested) and (not all_missing):
        return 2
    if all_tested and any_bad:
        return 1

    if all_missing:
        any_prev = False
        for it in items:
            dt = row.get(f"_{it}_dt", pd.NaT)
            if not pd.isna(dt) and dt.year == (base_year - 1):
                any_prev = True
                break
        if any_prev:
            return 4

        if last_year is None:
            return 5
        if (base_year - last_year) >= 2:
            return 5

        return 2

    return 2


def remark_text(code: int, x_items: int, base_year: int) -> str:
    if code == 0:
        return f"0.{base_year}年{x_items}項前測控制良好，符合收案標準,建議收為會員。"
    if code == 1:
        return f"1.{base_year}年已受檢未達控制良好"
    if code == 2:
        return f"2.{base_year}年漏檢,請安排回診"
    if code == 3:
        return f"3.{base_year}年漏檢+未達控制良好,回診複查"
    if code == 4:
        return f"4.{base_year-1}已檢：提醒{base_year}年度回診"
    if code == 5:
        return "5.逾兩年未檢：暫緩收案"
    return ""


# ------------------------- build df (A name, B dis/ascvd) -------------------------
def build_master_df(path_a: Path, path_b: Path) -> Tuple[pd.DataFrame, pd.DataFrame]:
    A0 = read_table(path_a, dtype=str)
    A0.columns = A0.columns.astype(str).str.strip()
    cols_a = list(A0.columns)

    col_id_a = pick_col_exact(cols_a, ["ID", "Id", "id"]) or pick_col_contains(cols_a, ["id"])
    if not col_id_a:
        raise ValueError("A 檔找不到欄位：ID")

    col_name_a = pick_col_exact(cols_a, ["姓名", "欄1", "name"]) or pick_col_contains(cols_a, ["姓名", "欄1"])
    if not col_name_a:
        raise ValueError("A 檔找不到欄位：姓名（或 欄1）")

    col_bday_a = pick_col_exact(cols_a, ["生日", "BIRTHDAY", "出生日期"]) or pick_col_contains(cols_a, ["生日", "birthday", "出生"])

    col_h_val = pick_col_exact(cols_a, ["最近一次HbA1c檢查結果(%)", "最近一次HbA1c檢查結果 (%)"]) or pick_col_contains(cols_a, ["最近一次hba1c檢查結果"])
    col_h_dt = pick_col_exact(cols_a, ["最近一次HbA1c檢查日期"]) or pick_col_contains(cols_a, ["最近一次hba1c檢查日期"])
    col_l_val = pick_col_exact(cols_a, ["最近一次LDL檢查結果(mg/dL)", "最近一次LDL檢查結果 (mg/dL)"]) or pick_col_contains(cols_a, ["最近一次ldl檢查結果"])
    col_l_dt = pick_col_exact(cols_a, ["最近一次LDL檢查日期"]) or pick_col_contains(cols_a, ["最近一次ldl檢查日期"])
    col_u_val = pick_col_exact(cols_a, ["最近一次UACR檢查結果(mg/gm)", "最近一次UACR檢查結果 (mg/gm)"]) or pick_col_contains(cols_a, ["最近一次uacr檢查結果"])
    col_u_dt = pick_col_exact(cols_a, ["最近一次UACR檢查日期"]) or pick_col_contains(cols_a, ["最近一次uacr檢查日期"])

    keep_a = [col_id_a, col_name_a, col_bday_a, col_h_val, col_h_dt, col_l_val, col_l_dt, col_u_val, col_u_dt]
    keep_a = [c for c in keep_a if c]
    A = A0[keep_a].copy().rename(columns={col_id_a: "ID", col_name_a: "姓名"})
    A["生日"] = A0[col_bday_a] if col_bday_a else pd.NA
    A["ID"] = norm_id(A["ID"])

    b_sheet, b_header = find_sheet_and_header_for_b(path_b)
    B0 = read_table(path_b, sheet_name=b_sheet, header=b_header, dtype=str)
    B0.columns = B0.columns.astype(str).str.strip()
    cols_b = list(B0.columns)

    col_id_b = pick_col_exact(cols_b, ["ID", "Id", "id"]) or pick_col_contains(cols_b, ["id"])
    col_dis_b = pick_col_exact(cols_b, ["疾病樣態"]) or pick_col_contains(cols_b, ["疾病樣態"])
    col_ascvd_b = pick_col_exact(cols_b, ["ASCVD", "ascvd"]) or pick_col_contains(cols_b, ["ascvd"])
    if not col_id_b or not col_dis_b or not col_ascvd_b:
        raise ValueError("B 檔欄位找不到（需要 ID/疾病樣態/ASCVD）")

    B = B0[[col_id_b, col_dis_b, col_ascvd_b]].copy().rename(
        columns={col_id_b: "ID", col_dis_b: "疾病樣態", col_ascvd_b: "ASCVD"}
    )
    B["ID"] = norm_id(B["ID"])

    # ✅ merge
    M = A.merge(B, on="ID", how="left", indicator=True)

    # ✅ 疾病樣態：忽略 Excel 前導單引號 ' + 清空白/換行/全形空白
    dis_clean = (M["疾病樣態"].astype(str)
                 .str.replace("\u3000", "", regex=False)
                 .str.replace(r"\s+", "", regex=True)
                 .str.replace("\n", "", regex=False)
                 .str.replace("\r", "", regex=False)
                 .str.lstrip("'")
                 .str.strip())

    df = pd.DataFrame({
        "ID": M["ID"],
        "疾病樣態": pd.to_numeric(dis_clean, errors="coerce"),
        "ASCVD": M["ASCVD"],
        "姓名": M["姓名"],
        "生日": M["生日"],
        "最近一次HbA1c檢查結果(%)": M[col_h_val] if col_h_val else pd.NA,
        "最近一次HbA1c檢查日期": M[col_h_dt] if col_h_dt else pd.NA,
        "最近一次LDL檢查結果(mg/dL)": M[col_l_val] if col_l_val else pd.NA,
        "最近一次LDL檢查日期": M[col_l_dt] if col_l_dt else pd.NA,
        "最近一次UACR檢查結果(mg/gm)": M[col_u_val] if col_u_val else pd.NA,
        "最近一次UACR檢查日期": M[col_u_dt] if col_u_dt else pd.NA,
    })
    df["_merge_flag"] = M["_merge"]

    # ✅ 日期解析（也忽略前導 '）
    df["_生日_dt"] = to_dt(df["生日"])
    df["_HbA1c_dt"] = to_dt(df["最近一次HbA1c檢查日期"])
    df["_LDL_dt"] = to_dt(df["最近一次LDL檢查日期"])
    df["_UACR_dt"] = to_dt(df["最近一次UACR檢查日期"])

    df["_HbA1c_num"] = pd.to_numeric(df["最近一次HbA1c檢查結果(%)"].astype(str).str.lstrip("'"), errors="coerce")
    df["_LDL_num"] = pd.to_numeric(df["最近一次LDL檢查結果(mg/dL)"].astype(str).str.lstrip("'"), errors="coerce")
    df["_UACR_num"] = pd.to_numeric(df["最近一次UACR檢查結果(mg/gm)"].astype(str).str.lstrip("'"), errors="coerce")

    df["生日"] = df["_生日_dt"].dt.strftime("%Y-%m-%d")
    df["最近一次HbA1c檢查日期"] = df["_HbA1c_dt"].dt.strftime("%Y-%m-%d")
    df["最近一次LDL檢查日期"] = df["_LDL_dt"].dt.strftime("%Y-%m-%d")
    df["最近一次UACR檢查日期"] = df["_UACR_dt"].dt.strftime("%Y-%m-%d")

    df["ASCVD"] = df["ASCVD"].apply(ascvd_norm)

    issues = []
    miss_b = df[df["疾病樣態"].isna()].copy()
    if not miss_b.empty:
        miss_b["問題"] = "B檔找不到此ID的疾病樣態/ASCVD（merge不到或格式無法轉數字）"
        issues.append(miss_b)
    issues_df = pd.concat(issues, ignore_index=True) if issues else pd.DataFrame()

    return df, issues_df


SCREENING_SHEET_ALIASES: Dict[str, List[str]] = {
    "成人健檢": ["成人健檢", "成人健檢名單", "成人預防保健", "成人預防保健名單"],
    "子宮抹片": ["子宮抹片", "子宮頸抹片", "子宮抹片名單", "子宮頸抹片名單"],
    "老人流感": ["老人流感", "老人流感名單", "老人流感注射", "65歲以上老人流感", "65歲以上老人流感注射率"],
    "糞便潛血": ["糞便潛血", "糞便潛血名單", "糞便潛血檢查", "糞便潛血檢查名單"],
    "肝炎篩檢": ["肝炎篩檢", "BC肝炎篩檢", "B肝炎篩檢", "C肝炎篩檢", "BC肝炎篩檢名單", "肝炎篩檢名單"],
}


def resolve_screening_sheet_name(path_a: Path, target_name: str) -> Optional[str]:
    xl = open_workbook(path_a)
    norm_to_real = {_norm_colname(sh): sh for sh in xl.sheet_names}
    aliases = SCREENING_SHEET_ALIASES.get(target_name, [target_name])

    for alias in aliases:
        n_alias = _norm_colname(alias)
        if n_alias in norm_to_real:
            return norm_to_real[n_alias]

    for alias in aliases:
        n_alias = _norm_colname(alias)
        for n_sheet, real in norm_to_real.items():
            if n_alias and (n_alias in n_sheet or n_sheet in n_alias):
                return real

    return None


def find_header_row_for_screening_sheet(path_a: Path, sheet_name: str) -> int:
    probe = read_table(path_a, sheet_name=sheet_name, header=None, nrows=15, dtype=str)
    date_keywords = [
        "最後篩檢日期", "最後檢查日期", "篩檢日期", "檢查日期",
        "受檢日期", "最後一次檢查日期", "最後一次篩檢日期", "日期"
    ]
    for r in range(len(probe)):
        tokens = [_norm_colname(v) for v in probe.iloc[r].tolist() if str(v).strip()]
        has_id = any(t == "id" for t in tokens)
        has_date = any(any(_norm_colname(k) in t for k in date_keywords) for t in tokens)
        if has_id and has_date:
            return r
    return 0


def build_screening_lookup(path_a: Path, sheet_name: str) -> pd.DataFrame:
    """
    從 A 檔指定篩檢 sheet 建立 lookup：
    回傳欄位固定為 ["ID", "篩檢日期"]
    - 自動容錯 sheet 名稱
    - 自動容錯表頭列/欄名
    - 同一 ID 多筆時取最新日期
    - 使用單值穩定日期解析，避免 pandas 混合格式 warning
    """
    resolved_name = resolve_screening_sheet_name(path_a, sheet_name)
    if not resolved_name:
        print(f"[WARN] 找不到 sheet：{sheet_name} → 已跳過")
        return pd.DataFrame(columns=["ID", "篩檢日期"])

    raw = read_table(path_a, sheet_name=resolved_name, header=None, dtype=object)
    if raw.empty:
        print(f"[WARN] sheet「{resolved_name}」是空的 → 已跳過")
        return pd.DataFrame(columns=["ID", "篩檢日期"])

    date_keywords = [
        "最後篩檢日期", "最後檢查日期", "篩檢日期", "檢查日期",
        "受檢日期", "最後一次檢查日期", "最後一次篩檢日期", "日期"
    ]
    id_keywords = ["ID", "id", "Id", "身份證", "身分證", "身份證號", "身分證號"]

    header_row = None
    scan_rows = min(len(raw), 15)
    for r in range(scan_rows):
        vals = [_norm_colname(v) for v in raw.iloc[r].tolist()]
        has_id = any(any(_norm_colname(k) == v or _norm_colname(k) in v for k in id_keywords) for v in vals if v)
        has_date = any(any(_norm_colname(k) in v for k in date_keywords) for v in vals if v)
        if has_id and has_date:
            header_row = r
            break
    if header_row is None:
        header_row = 0

    header_vals = raw.iloc[header_row].tolist()
    data = raw.iloc[header_row + 1:].copy().reset_index(drop=True)
    data.columns = [str(c).strip() if c is not None and not pd.isna(c) else f"col_{i}" for i, c in enumerate(header_vals)]

    cols = list(data.columns)

    def find_col(candidates_exact=None, candidates_kw=None):
        candidates_exact = candidates_exact or []
        candidates_kw = candidates_kw or []

        for c in cols:
            nc = _norm_colname(c)
            for name in candidates_exact:
                if nc == _norm_colname(name):
                    return c

        for c in cols:
            nc = _norm_colname(c)
            for kw in candidates_kw:
                nkw = _norm_colname(kw)
                if nkw and nkw in nc:
                    return c
        return None

    col_id = find_col(
        candidates_exact=["ID", "id", "Id"],
        candidates_kw=["id", "身份證", "身分證", "身份證號", "身分證號"]
    )
    col_dt = find_col(
        candidates_exact=["最後篩檢日期", "最後檢查日期", "篩檢日期", "檢查日期", "受檢日期", "最後一次檢查日期", "最後一次篩檢日期", "日期"],
        candidates_kw=["最後篩檢日期", "最後檢查日期", "篩檢日期", "檢查日期", "受檢日期", "最後一次檢查日期", "最後一次篩檢日期", "日期"]
    )

    if not col_id:
        print(f"[WARN] sheet「{resolved_name}」找不到 ID 欄位 → 已跳過")
        return pd.DataFrame(columns=["ID", "篩檢日期"])

    out = data[[col_id]].copy()
    out = out.rename(columns={col_id: "ID"})
    out["ID"] = norm_id(out["ID"])
    out = out[out["ID"].notna()].copy()

    if col_dt:
        out["篩檢日期"] = data[col_dt].apply(parse_mixed_date)
        latest = out.groupby("ID", as_index=False)["篩檢日期"].max()
        latest["篩檢日期"] = latest["篩檢日期"].apply(
            lambda x: x.strftime("%Y-%m-%d") if pd.notna(x) else pd.NA
        )
        s = latest["篩檢日期"].astype(str).str.strip()
        latest.loc[s.isin(["NaT", "nan", "None", "<NA>", "-", ""]), "篩檢日期"] = pd.NA
        return latest[["ID", "篩檢日期"]]

    out = out.drop_duplicates(subset=["ID"], keep="first")
    out["篩檢日期"] = pd.NA
    return out[["ID", "篩檢日期"]]



def build_all_screening_lookup(path_a: Path) -> pd.DataFrame:
    """
    建立 5 項篩檢的共用 lookup，回傳欄位：
    ID + SCREENING_EXPORT_COLS
    """
    result = pd.DataFrame(columns=["ID"])
    lookup_specs = [
        ("成人健檢", "成人預防保健(6分)"),
        ("子宮抹片", "子宮頸抹片檢查率(6分)"),
        ("老人流感", "65歲以上老人流感注射率(4分)"),
        ("糞便潛血", "糞便潛血檢查率(6分)"),
        ("肝炎篩檢", "BC肝炎篩檢(6分)"),
    ]

    for logical_sheet_name, out_col in lookup_specs:
        lookup = build_screening_lookup(path_a, logical_sheet_name)
        if lookup.empty:
            print(f"[INFO] {logical_sheet_name} 無資料或已跳過")
            result[out_col] = pd.NA
            continue

        lookup = lookup.rename(columns={"篩檢日期": out_col})
        if result.empty:
            result = lookup[["ID", out_col]].copy()
        else:
            result = result.merge(lookup[["ID", out_col]], on="ID", how="outer")

    if "ID" not in result.columns:
        result = pd.DataFrame(columns=["ID"] + SCREENING_EXPORT_COLS)

    for col in SCREENING_EXPORT_COLS:
        if col not in result.columns:
            result[col] = pd.NA

    result["ID"] = norm_id(result["ID"])
    result = result[result["ID"].notna()].copy()
    return result[["ID"] + SCREENING_EXPORT_COLS].drop_duplicates(subset=["ID"], keep="first")


def attach_screening_columns(df: pd.DataFrame, path_a: Path) -> pd.DataFrame:
    """
    把 5 項篩檢日期加到主 df，供 7 個分流 sheet 與無疾病 sheet 共用
    """
    lookup = build_all_screening_lookup(path_a)
    out = df.merge(lookup, on="ID", how="left")

    for col in SCREENING_EXPORT_COLS:
        if col not in out.columns:
            out[col] = pd.NA
        s = out[col].astype(str).str.strip()
        out.loc[s.isin(["NaT", "nan", "None", "<NA>", "-", ""]), col] = pd.NA

    return out

def build_no_disease_df(df: pd.DataFrame, path_a: Path) -> pd.DataFrame:
    base = df[
        (~df["疾病樣態"].isin([1, 2, 3])) &
        (~df["ASCVD"].isin(["a", "b"]))
    ].copy()

    if base.empty:
        return pd.DataFrame(columns=NO_DISEASE_EXPORT_COLS)

    for col in NO_DISEASE_EXPORT_COLS:
        if col not in base.columns:
            base[col] = pd.NA

    for col in SCREENING_EXPORT_COLS:
        s = base[col].astype(str).str.strip()
        base.loc[s.isin(["NaT", "nan", "None", "<NA>", "-"]), col] = pd.NA

    return base[NO_DISEASE_EXPORT_COLS].sort_values("ID", ascending=True, kind="mergesort").reset_index(drop=True)


# ------------------------- split -------------------------
def split_into_sheets(df: pd.DataFrame, base_year: int) -> Tuple[Dict[str, pd.DataFrame], Dict[str, Dict[str, int]]]:
    is_dm = df["疾病樣態"] == 1
    is_ckd = df["疾病樣態"] == 2
    is_dkd = df["疾病樣態"] == 3
    has_ascvd = df["ASCVD"].isin(["a", "b"])

    sheets = {
        "DM": df[is_dm & (~has_ascvd)].copy(),
        "CKD": df[is_ckd & (~has_ascvd)].copy(),
        "DKD": df[is_dkd & (~has_ascvd)].copy(),
        "DM+ASCVD": df[is_dm & has_ascvd].copy(),
        "CKD+ASCVD": df[is_ckd & has_ascvd].copy(),
        "DKD+ASCVD": df[is_dkd & has_ascvd].copy(),
        "ASCVD": df[has_ascvd & (~(is_dm | is_ckd | is_dkd))].copy(),
    }

    analysis_counts: Dict[str, Dict[str, int]] = {}

    for name, d in sheets.items():
        if len(d) > 0:
            x = len(required_items(name))
            codes = d.apply(lambda r: remark_code(r, name, base_year), axis=1)
            d["備註"] = [remark_text(int(c), x, base_year) for c in codes.tolist()]
            d = d.sort_values("ID", ascending=True, kind="mergesort")
        else:
            d = pd.DataFrame(columns=EXPORT_COLS)

        if len(d) > 0:
            vc = d["備註"].astype(str).str.extract(r"^(\d)\.")[0].value_counts()
        else:
            vc = pd.Series(dtype="int64")
        analysis_counts[name] = {str(i): int(vc.get(str(i), 0)) for i in range(6)}

        sheets[name] = d[EXPORT_COLS].copy()

    return sheets, analysis_counts


# ------------------------- formatting helpers -------------------------
def autosize_columns(ws, max_width: int = 45):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _parse_year(v) -> Optional[int]:
    if v is None:
        return None
    s = str(v).strip()
    s = _strip_excel_apostrophe(s)
    if not s or s in ("-", "0", "0.0"):
        return None
    m = re.search(r"(\d{4})", s)
    return int(m.group(1)) if m else None


def _parse_float(v) -> Optional[float]:
    if v is None:
        return None
    s = str(v).strip()
    s = _strip_excel_apostrophe(s)
    if not s or s in ("-", "0", "0.0"):
        return None
    try:
        return float(s)
    except Exception:
        return None


def style_triage_sheet(ws, sheet_name: str):
    """
    分頁上色（依人工範例）：不是整列，只標註指定欄位
    - 不改 border → 黑色隔線保留
    - 對齊：備註左對齊 wrap；其餘置中
    - 不加粗/不放大姓名
    - 只處理實際資料列，避免 used range
    """
    if ws.max_row < 2:
        return

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    def col_idx(name: str):
        try:
            return headers.index(name) + 1
        except ValueError:
            return None

    c_id = col_idx("ID") or 1
    c_dis = col_idx("疾病樣態")
    c_asc = col_idx("ASCVD")
    c_name = col_idx("姓名")
    c_bday = col_idx("生日")
    c_remark = col_idx("備註")

    c_h_val = col_idx("最近一次HbA1c檢查結果(%)")
    c_h_dt = col_idx("最近一次HbA1c檢查日期")
    c_l_val = col_idx("最近一次LDL檢查結果(mg/dL)")
    c_l_dt = col_idx("最近一次LDL檢查日期")
    c_u_val = col_idx("最近一次UACR檢查結果(mg/gm)")
    c_u_dt = col_idx("最近一次UACR檢查日期")

    screening_date_cols = [
        col_idx("成人預防保健(6分)"),
        col_idx("子宮頸抹片檢查率(6分)"),
        col_idx("65歲以上老人流感注射率(4分)"),
        col_idx("糞便潛血檢查率(6分)"),
        col_idx("BC肝炎篩檢(6分)"),
    ]
    screening_date_cols = [c for c in screening_date_cols if c]

    member_cols = [c for c in [c_id, c_dis, c_asc, c_name, c_bday] if c]

    FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")   # 今年日期
    FILL_PINK = PatternFill("solid", fgColor="FF6699")    # 去年日期
    FILL_YELLOW = PatternFill("solid", fgColor="FFFF00")  # 漏檢/空白/其他年
    FILL_ORANGE = PatternFill("solid", fgColor="FFC000")  # 異常值
    FILL_PURPLE = PatternFill("solid", fgColor="DDD9EE")  # code 4 會員欄

    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    normal_font = Font(size=12, bold=False)

    # ✅ 修正：統一黑色細框，所有儲存格都必須明確設定
    thin_black = Side(style="thin", color="000000")
    cell_border = Border(
        left=thin_black, right=thin_black,
        top=thin_black, bottom=thin_black,
    )

    # 找最後一筆資料列（用 ID 欄判斷）
    last_data_row = 1
    for r in range(ws.max_row, 1, -1):
        v = ws.cell(r, c_id).value
        if v is not None and str(v).strip() != "":
            last_data_row = r
            break
    if last_data_row == 1:
        return

    def paint_date(row_r: int, col_dt: Optional[int]):
        if not col_dt:
            return
        cell = ws.cell(row_r, col_dt)
        y = _parse_year(cell.value)

        # 無有效日期：不要上色
        if y is None:
            cell.border = cell_border
            return

        if y == BASE_YEAR:
            cell.fill = FILL_GREEN
        elif y == BASE_YEAR - 1:
            cell.fill = FILL_PINK
        else:
            cell.fill = FILL_YELLOW
        cell.border = cell_border  # ✅ 填色後補回邊框

    def paint_screening_date(row_r: int, col_dt: Optional[int]):
        if not col_dt:
            return
        cell = ws.cell(row_r, col_dt)
        y = _parse_year(cell.value)
        if y is None:
            cell.border = cell_border
            return
        if y == BASE_YEAR:
            cell.fill = FILL_GREEN
        elif y == BASE_YEAR - 1:
            cell.fill = FILL_PINK
        else:
            cell.fill = FILL_YELLOW
        cell.border = cell_border

    def paint_value_if_bad(row_r: int, col_val: Optional[int], col_dt: Optional[int], threshold: float):
        if (not col_val) or (not col_dt):
            return
        y = _parse_year(ws.cell(row_r, col_dt).value)
        v = _parse_float(ws.cell(row_r, col_val).value)
        if y == BASE_YEAR and (v is not None) and v >= threshold:
            cell = ws.cell(row_r, col_val)
            cell.fill = FILL_ORANGE
            cell.border = cell_border  # ✅ 填色後補回邊框

    items = required_items(sheet_name)

    for r in range(2, last_data_row + 1):
        ws.row_dimensions[r].height = 20

        # ✅ 修正：對齊/字型/邊框一併設定
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.font = normal_font
            cell.alignment = left_wrap if (c_remark and c == c_remark) else center
            cell.border = cell_border  # ← 補上黑色邊框

        # 取備註代碼 0~5
        code = ""
        if c_remark:
            rv = ws.cell(r, c_remark).value
            if rv:
                s = str(rv).strip()
                if s and s[0].isdigit():
                    code = s[0]

        # code 4 / 5：會員欄底色
        if code == "4":
            for c in member_cols:
                cell = ws.cell(r, c)
                cell.fill = FILL_PURPLE
                cell.border = cell_border  # ✅ 填色後補回邊框
        elif code == "5":
            for c in member_cols:
                cell = ws.cell(r, c)
                cell.fill = FILL_YELLOW
                cell.border = cell_border  # ✅ 填色後補回邊框

        # 門檻計算（沿用你原邏輯）
        row_ascvd = ascvd_norm(ws.cell(r, c_asc).value if c_asc else "")
        bday_str = ws.cell(r, c_bday).value if c_bday else None
        bday_dt = pd.to_datetime(_strip_excel_apostrophe(str(bday_str)), errors="coerce") if bday_str not in (None, "-", "0", "0.0") else pd.NaT

        ref = pd.Timestamp(f"{BASE_YEAR}-06-30")
        age = age_on(bday_dt, ref)
        h_target = 8.0 if (age is not None and age >= 80) else 7.0
        l_target = ldl_target(pd.Series({"ASCVD": row_ascvd}), sheet_name)
        u_target = 30.0

        if "HbA1c" in items:
            paint_date(r, c_h_dt)
            paint_value_if_bad(r, c_h_val, c_h_dt, h_target)
        if "LDL" in items:
            paint_date(r, c_l_dt)
            paint_value_if_bad(r, c_l_val, c_l_dt, l_target)
        if "UACR" in items:
            paint_date(r, c_u_dt)
            paint_value_if_bad(r, c_u_val, c_u_dt, u_target)

        for c in screening_date_cols:
            paint_screening_date(r, c)


# ------------------------- analysis sheet (layout tweaks) -------------------------
def build_analysis_sheet(ws, analysis_counts: Dict[str, Dict[str, int]], base_year: int):
    order = ["DM", "CKD", "DKD", "ASCVD", "CKD+ASCVD", "DM+ASCVD", "DKD+ASCVD"]

    group_fills = {
        "DM": PatternFill("solid", fgColor="D9E8EE"),
        "CKD": PatternFill("solid", fgColor="F3DED3"),
        "DKD": PatternFill("solid", fgColor="DFF0D8"),
        "ASCVD": PatternFill("solid", fgColor="D9E8EE"),
        "CKD+ASCVD": PatternFill("solid", fgColor="E6E6E6"),
        "DM+ASCVD": PatternFill("solid", fgColor="E5D8EE"),
        "DKD+ASCVD": PatternFill("solid", fgColor="D8E6EE"),
    }

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    title_font = Font(name="標楷體", size=20, bold=True)
    group_font = Font(name="Times New Roman", size=14, bold=False)
    header_font = Font(name="標楷體", size=14, bold=True)
    text_font = Font(name="標楷體", size=12, bold=False)
    red_font = Font(name="標楷體", size=12, color="FF0000", bold=True)
    num_font = Font(name="Times New Roman", size=14, bold=False)

    ws.column_dimensions["A"].width = 12
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 22

    ws.merge_cells("A1:G1")
    ws["A1"].value = "115年收案會員選案建議評估"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 40

    r = 2
    base_h = 33
    remark_h = base_h * 1.2

    def remark_texts_for_group(g: str) -> List[str]:
        x = len(required_items(g))
        return [
            f"0.{base_year}年{x}項前測控制良好，符合收案標準,建議收為會員",
            f"1.{base_year}年已受檢未達控制良好",
            f"2.{base_year}年漏檢,請安排回診",
            f"3.{base_year}年漏檢+未達控制良好,回診複查",
            f"4.{base_year-1}已檢：提醒{base_year}年度回診",
            "5.逾兩年未檢：暫緩收案",
        ]

    for g in order:
        fill = group_fills.get(g, PatternFill("solid", fgColor="EEEEEE"))

        ws.row_dimensions[r].height = base_h
        ws.row_dimensions[r + 1].height = remark_h
        ws.row_dimensions[r + 2].height = base_h

        ws.merge_cells(f"A{r}:A{r + 2}")
        cA = ws[f"A{r}"]
        cA.value = g
        cA.font = group_font
        cA.alignment = center
        cA.fill = fill
        cA.border = border

        ws.merge_cells(f"B{r}:G{r}")
        cTot = ws[f"B{r}"]
        cTot.value = "總計"
        cTot.font = header_font
        cTot.alignment = center
        cTot.fill = fill
        cTot.border = border

        texts = remark_texts_for_group(g)
        for i, col in enumerate("BCDEFG"):
            cell = ws[f"{col}{r + 1}"]
            cell.value = texts[i]
            cell.alignment = left_wrap
            cell.fill = fill
            cell.border = border
            cell.font = red_font if any(k in texts[i] for k in ["建議收", "未達控制", "漏檢", "提醒"]) else text_font

        counts = analysis_counts.get(g, {str(i): 0 for i in range(6)})
        for i, col in enumerate("BCDEFG"):
            cell = ws[f"{col}{r + 2}"]
            cell.value = int(counts.get(str(i), 0))
            cell.font = num_font
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.fill = PatternFill("solid", fgColor="FFFFFF")
            cell.border = border

        for rr in (r + 1, r + 2):
            ws[f"A{rr}"].fill = fill
            ws[f"A{rr}"].border = border

        for col in "BCDEFG":
            ws[f"{col}{r}"].border = border
            ws[f"{col}{r}"].fill = fill

        sep_row = r + 3
        ws.row_dimensions[sep_row].height = 15
        r = sep_row + 1

    ws.sheet_view.showGridLines = False



def style_no_disease_sheet(ws):
    """無疾病(4,0) 分頁樣式。"""
    if ws.max_row < 1:
        return

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    def col_idx(name: str):
        try:
            return headers.index(name) + 1
        except ValueError:
            return None

    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    header_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    normal_font = Font(size=12, bold=False)
    header_font = Font(size=12, bold=True)
    thin_black = Side(style="thin", color="000000")
    border = Border(left=thin_black, right=thin_black, top=thin_black, bottom=thin_black)
    fill_green  = PatternFill("solid", fgColor="C6EFCE")  # 今年
    fill_pink   = PatternFill("solid", fgColor="FF6699")  # 去年
    fill_yellow = PatternFill("solid", fgColor="FFFF00")  # 其他年

    check_cols = [
        col_idx("成人預防保健(6分)"),
        col_idx("子宮頸抹片檢查率(6分)"),
        col_idx("65歲以上老人流感注射率(4分)"),
        col_idx("糞便潛血檢查率(6分)"),
        col_idx("BC肝炎篩檢(6分)"),
    ]
    check_cols = [c for c in check_cols if c]

    for c in range(1, ws.max_column + 1):
        cell = ws.cell(1, c)
        cell.alignment = header_center
        cell.font = header_font
        cell.border = border

    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 20
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.alignment = center
            cell.font = normal_font
            cell.border = border

        for c in check_cols:
            cell = ws.cell(r, c)
            y = _parse_year(cell.value)
            if y is None:
                pass  # 空值不填色
            elif y == BASE_YEAR:
                cell.fill = fill_green
            elif y == BASE_YEAR - 1:
                cell.fill = fill_pink
            else:
                cell.fill = fill_yellow

    autosize_columns(ws)
    c_name = col_idx("姓名")
    if c_name:
        ws.column_dimensions[get_column_letter(c_name)].width = 10
    c_birth = col_idx("生日")
    if c_birth:
        birth_col = get_column_letter(c_birth)
        if ws.column_dimensions[birth_col].width < 12:
            ws.column_dimensions[birth_col].width = 12

    ws.sheet_view.showGridLines = False


# ------------------------- write output -------------------------
def write_output(out_path: Path, sheets: Dict[str, pd.DataFrame], analysis_counts: Dict[str, Dict[str, int]], no_disease_df: Optional[pd.DataFrame] = None):
    wb = Workbook()
    wb.remove(wb.active)

    ws_analysis = wb.create_sheet("分析表", 0)
    build_analysis_sheet(ws_analysis, analysis_counts, BASE_YEAR)

    def _excel_safe(v):
        """openpyxl 可接受的值：None / str / int / float / datetime"""
        if v is None:
            return None
        try:
            if pd.isna(v):
                return None
        except Exception:
            pass
        if str(v) == "<NA>":
            return None
        return v

    for name in SHEET_ORDER:
        ws = wb.create_sheet(title=name)
        df = sheets.get(name, pd.DataFrame(columns=EXPORT_COLS)).copy()

        # 1) '0'/'0.0' -> '-'
        text_cols = ["ID", "疾病樣態", "ASCVD", "姓名", "生日", "備註"]
        date_cols = ["最近一次HbA1c檢查日期", "最近一次LDL檢查日期", "最近一次UACR檢查日期"] + SCREENING_EXPORT_COLS
        val_cols = ["最近一次HbA1c檢查結果(%)", "最近一次LDL檢查結果(mg/dL)", "最近一次UACR檢查結果(mg/gm)"]
        fix_cols = [c for c in (text_cols + date_cols + val_cols) if c in df.columns]

        for col in fix_cols:
            s = df[col].astype(str).str.strip().str.lstrip("'")
            mask = df[col].notna() & s.isin(["0", "0.0"])
            if mask.any():
                df[col] = df[col].astype("object")
                df.loc[mask, col] = "-"

        # 2) 寫入
        for c, h in enumerate(df.columns, start=1):
            ws.cell(row=1, column=c, value=_excel_safe(h))

        for r_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
            for c_idx, v in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=_excel_safe(v))

        # 3) 欄寬：先 autosize，再固定姓名/備註
        autosize_columns(ws)

        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

        if "姓名" in headers:
            name_col = get_column_letter(headers.index("姓名") + 1)
            ws.column_dimensions[name_col].width = 10

        if "備註" in headers:
            remark_col = get_column_letter(headers.index("備註") + 1)
            ws.column_dimensions[remark_col].width = 30  # 185pt 近似

        # 4) 分頁樣式
        style_triage_sheet(ws, sheet_name=name)

    ws_no = wb.create_sheet(title=NO_DISEASE_SHEET)
    no_df = no_disease_df.copy() if no_disease_df is not None else pd.DataFrame(columns=NO_DISEASE_EXPORT_COLS)

    for col in [c for c in NO_DISEASE_EXPORT_COLS if c in no_df.columns]:
        s = no_df[col].astype(str).str.strip().str.lstrip("'")
        mask = no_df[col].notna() & s.isin(["0", "0.0"])
        if mask.any():
            no_df[col] = no_df[col].astype("object")
            no_df.loc[mask, col] = "-"

    for c, h in enumerate(NO_DISEASE_EXPORT_COLS, start=1):
        ws_no.cell(row=1, column=c, value=_excel_safe(h))

    for r_idx, row in enumerate(no_df.reindex(columns=NO_DISEASE_EXPORT_COLS).itertuples(index=False, name=None), start=2):
        for c_idx, v in enumerate(row, start=1):
            ws_no.cell(row=r_idx, column=c_idx, value=_excel_safe(v))
    style_no_disease_sheet(ws_no)

    wb.save(out_path)


# ------------------------- GUI main -------------------------
def main():
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()

    try:
        messagebox.showinfo("步驟1", "請選擇 A 檔（個案健康管理）")
        path_a = filedialog.askopenfilename(
            title="選擇 A 檔（姓名 + 檢驗結果）",
            filetypes=[("Spreadsheet files", "*.xlsx *.xls *.ods")]
        )
        if not path_a:
            return

        messagebox.showinfo("步驟2", "請選擇 B 檔（照護名單，含ASCVD）")
        path_b = filedialog.askopenfilename(
            title="選擇 B 檔（疾病樣態 + ASCVD）",
            filetypes=[("Spreadsheet files", "*.xlsx *.xls *.ods")]
        )
        if not path_b:
            return

        out_dir = Path(path_b).parent
        ts = datetime.now().strftime("%m%d_%H%M")
        out_path = out_dir / f"未達控制{ts}.xlsx"

        print("▶ 讀取並合併 A/B 檔...")
        df, _ = build_master_df(Path(path_a), Path(path_b))
        print(f"  ✓ 合併完成，共 {len(df)} 筆")

        print("▶ 附加篩檢欄位...")
        df = attach_screening_columns(df, Path(path_a))
        print("  ✓ 篩檢欄位附加完成")

        print("▶ 分流中...")
        sheets, analysis_counts = split_into_sheets(df, BASE_YEAR)
        for sname, sdf in sheets.items():
            print(f"  {sname}: {len(sdf)} 筆")

        print("▶ 建立無疾病分頁...")
        no_disease_df = build_no_disease_df(df, Path(path_a))
        print(f"  ✓ 無疾病共 {len(no_disease_df)} 筆")

        print("▶ 寫出 Excel...")
        write_output(out_path, sheets, analysis_counts, no_disease_df)
        print(f"  ✓ 輸出完成：{out_path}")

        msg = f"報表完成（年度固定 {BASE_YEAR}）：\n{out_path}"
        messagebox.showinfo("完成", msg)
        open_file(out_path)

    except Exception:
        tb = traceback.format_exc()
        print(f"\n❌ 發生錯誤：\n{tb}")
        messagebox.showerror("錯誤(Traceback)", tb)


if __name__ == "__main__":
    main()