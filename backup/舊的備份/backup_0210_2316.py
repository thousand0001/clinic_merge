import os
import re
import datetime
from typing import Optional, Dict, Tuple, Any

import openpyxl
from openpyxl.styles import Font, Border, Side
from openpyxl.utils.exceptions import InvalidFileException


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", "").replace("\t", "").replace("　", "").strip()


def build_header_map(sheet, header_row: int) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    for col in range(1, sheet.max_column + 1):
        key = normalize_text(sheet.cell(header_row, col).value)
        if key:
            header_map[key] = col
    return header_map


def find_column(header_map: Dict[str, int], aliases) -> Optional[int]:
    for name in aliases:
        if name in header_map:
            return header_map[name]
    return None


def find_column_contains(sheet, header_row: int, keywords) -> Tuple[Optional[int], Optional[str]]:
    keys = [k.replace(" ", "") for k in keywords]
    for col in range(1, sheet.max_column + 1):
        header = normalize_text(sheet.cell(header_row, col).value)
        header_no_space = header.replace(" ", "")
        for k in keys:
            if k and k in header_no_space:
                return col, header
    return None, None


def find_header_row(sheet, required_headers, search_rows: int = 80) -> Optional[int]:
    required = [h.lower() for h in required_headers]
    for row in range(1, min(search_rows, sheet.max_row) + 1):
        found = set()
        for col in range(1, sheet.max_column + 1):
            v = sheet.cell(row, col).value
            if v is None:
                continue
            s = str(v).strip().lower()
            if s in required:
                found.add(s)
        if all(h in found for h in required):
            return row
    return None


def parse_date(value: Any) -> Optional[datetime.date]:
    if value is None:
        return None

    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value

    s = str(value).strip()
    if s == "" or s in ("-", "—", "–"):
        return None

    # yyyy/mm/dd or yyyy-mm-dd
    m = re.match(r"^(\d{4})[/-](\d{1,2})[/-](\d{1,2})$", s)
    if m:
        y, mo, d = map(int, m.groups())
        return datetime.date(y, mo, d)

    # ROC yyyMMdd
    m = re.match(r"^(\d{2,3})(\d{2})(\d{2})$", s)
    if m:
        roc, mo, d = map(int, m.groups())
        return datetime.date(roc + 1911, mo, d)

    # ROC yyy/mm/dd
    m = re.match(r"^(\d{2,3})[/-](\d{1,2})[/-](\d{1,2})$", s)
    if m:
        roc, mo, d = map(int, m.groups())
        return datetime.date(roc + 1911, mo, d)

    return None


def parse_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    s = str(value).strip()
    if s == "" or s in ("-", "—", "–"):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def infer_gender_from_id(id_value: Any) -> str:
    if not id_value:
        return ""
    s = str(id_value).strip().upper()
    if len(s) < 2:
        return ""
    g = s[1]
    if g in ("1", "8"):
        return "男"
    if g in ("2", "9"):
        return "女"
    return ""


def get_sheet_ci(workbook, name: str):
    target = name.strip().lower()
    for s in workbook.sheetnames:
        if s.strip().lower() == target:
            return workbook[s]
    return None


def rename_header(sheet, header_row: int, old_name: str, new_name: str) -> None:
    for col in range(1, sheet.max_column + 1):
        if normalize_text(sheet.cell(header_row, col).value) == old_name:
            sheet.cell(header_row, col).value = new_name


def ensure_column_after(sheet, header_row: int, after_header: str, new_header: str) -> int:
    for col in range(1, sheet.max_column + 1):
        if normalize_text(sheet.cell(header_row, col).value) == new_header:
            return col

    after_col = None
    for col in range(1, sheet.max_column + 1):
        if normalize_text(sheet.cell(header_row, col).value) == after_header:
            after_col = col
            break
    if after_col is None:
        raise ValueError(f"找不到欄位「{after_header}」，無法插入「{new_header}」")

    sheet.insert_cols(after_col + 1)
    sheet.cell(header_row, after_col + 1).value = new_header
    return after_col + 1


def style_sheet(sheet, font_name: str = "標楷體", font_size: int = 19) -> None:
    font_style = Font(name=font_name, size=font_size)
    thin = Side(style="thin")
    border_style = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            cell = sheet.cell(r, c)
            cell.font = font_style
            cell.border = border_style


def set_date_format(sheet, col_indices, start_row: int, fmt: str = "yyyy-mm-dd") -> None:
    for r in range(start_row, sheet.max_row + 1):
        for c in col_indices:
            if c is None:
                continue
            cell = sheet.cell(r, c)
            if isinstance(cell.value, (datetime.date, datetime.datetime)):
                cell.number_format = fmt


def process_excel(xlsx_path: str) -> str:
    workbook = openpyxl.load_workbook(xlsx_path)

    sheet_target = get_sheet_ci(workbook, "會員指標")
    sheet_member = get_sheet_ci(workbook, "會員名單")
    sheet_ascvd = get_sheet_ci(workbook, "ascvd")
    sheet_health = get_sheet_ci(workbook, "HealthCase")

    sheet_adult = get_sheet_ci(workbook, "成人健檢")
    sheet_pap = get_sheet_ci(workbook, "子宮抹片")
    sheet_flu = get_sheet_ci(workbook, "老人流感")
    sheet_fit = get_sheet_ci(workbook, "糞便潛血")
    sheet_hep = get_sheet_ci(workbook, "肝炎篩檢")

    missing_sheets = []
    for name, sh in [
        ("會員指標", sheet_target),
        ("會員名單", sheet_member),
        ("ascvd", sheet_ascvd),
        ("HealthCase", sheet_health),
        ("成人健檢", sheet_adult),
        ("子宮抹片", sheet_pap),
        ("老人流感", sheet_flu),
        ("糞便潛血", sheet_fit),
        ("肝炎篩檢", sheet_hep),
    ]:
        if sh is None:
            missing_sheets.append(name)
    if missing_sheets:
        raise ValueError("缺少工作表：" + "、".join(missing_sheets))

    header_row_target = 5
    header_row_member = 5

    # /ASCVD -> ASCVD
    rename_header(sheet_target, header_row_target, "/ASCVD", "ASCVD")

    # Insert HealthCase date columns (after results)
    ensure_column_after(sheet_target, header_row_target, "HbA1c", "HbA1c檢查日期")
    ensure_column_after(sheet_target, header_row_target, "LDL", "LDL檢查日期")
    ensure_column_after(sheet_target, header_row_target, "UACR", "UACR檢查日期")

    # Target columns
    tmap = build_header_map(sheet_target, header_row_target)
    col_name = find_column(tmap, ["姓名"])
    col_id = find_column(tmap, ["身份證號碼", "身分證號碼", "ID"])
    col_bday = find_column(tmap, ["生日"])
    col_tel = find_column(tmap, ["電話"])
    col_abc = find_column(tmap, ["A/B/CC=06B=空白其他為A", "會員a/b/c", "A/B/C"])
    col_dmk = find_column(tmap, ["DM/CKD/DKD"])
    col_cnt = find_column(tmap, ["次數"])
    col_ascvd = find_column(tmap, ["ASCVD"])
    col_sex = find_column(tmap, ["性別"])

    col_hba1c = find_column(tmap, ["HbA1c"])
    col_hba1c_date = find_column(tmap, ["HbA1c檢查日期"])
    col_ldl = find_column(tmap, ["LDL"])
    col_ldl_date = find_column(tmap, ["LDL檢查日期"])
    col_uacr = find_column(tmap, ["UACR"])
    col_uacr_date = find_column(tmap, ["UACR檢查日期"])

    # Screening target columns by keyword
    col_adult, _ = find_column_contains(sheet_target, header_row_target, ["成人預防保健"])
    col_pap, _ = find_column_contains(sheet_target, header_row_target, ["子宮頸抹片"])
    col_flu, _ = find_column_contains(sheet_target, header_row_target, ["老人流感", "老人流感", "流感注射", "流感"])
    col_fit, _ = find_column_contains(sheet_target, header_row_target, ["糞便潛血"])
    col_hep, _ = find_column_contains(sheet_target, header_row_target, ["肝炎篩檢", "B、C肝炎", "BC肝炎"])

    if col_sex is None:
        col_sex = sheet_target.max_column + 1
        sheet_target.cell(header_row_target, col_sex).value = "性別"

    required_cols = {
        "姓名": col_name,
        "身份證號碼": col_id,
        "生日": col_bday,
        "電話": col_tel,
        "A/B/C": col_abc,
        "DM/CKD/DKD": col_dmk,
        "次數": col_cnt,
        "ASCVD": col_ascvd,
        "HbA1c": col_hba1c,
        "HbA1c檢查日期": col_hba1c_date,
        "LDL": col_ldl,
        "LDL檢查日期": col_ldl_date,
        "UACR": col_uacr,
        "UACR檢查日期": col_uacr_date,
        "性別": col_sex,
        "成人預防保健": col_adult,
        "子宮頸抹片": col_pap,
        "老人流感": col_flu,
        "糞便潛血": col_fit,
        "肝炎篩檢": col_hep,
    }
    missing_cols = [k for k, v in required_cols.items() if v is None]
    if missing_cols:
        raise ValueError("會員指標缺少欄位/欄名不符：" + "、".join(missing_cols))

    # Member columns
    mmap = build_header_map(sheet_member, header_row_member)
    m_name = find_column(mmap, ["會員姓名"])
    m_id = find_column(mmap, ["會員身份証", "會員身份證", "會員身分證"])
    m_bday = find_column(mmap, ["會員生日"])
    m_tel = find_column(mmap, ["電話"])
    m_abc = find_column(mmap, ["會員別"])
    m_dmk = find_column(mmap, ["疾病樣態"])
    m_cnt = find_column(mmap, ["就診次數"])

    missing_member_cols = [k for k, v in {
        "會員姓名": m_name,
        "會員身份證": m_id,
        "會員生日": m_bday,
        "電話": m_tel,
        "會員別": m_abc,
        "疾病樣態": m_dmk,
        "就診次數": m_cnt,
    }.items() if v is None]
    if missing_member_cols:
        raise ValueError("會員名單缺少欄位：" + "、".join(missing_member_cols))

    # 追加名單
    append_row = 9
    while sheet_target.cell(append_row, col_id).value not in (None, ""):
        append_row += 1

    added_count = 0
    for member_row in range(6, sheet_member.max_row + 1):
        sid = sheet_member.cell(member_row, m_id).value
        name = sheet_member.cell(member_row, m_name).value
        if (sid in (None, "")) and (name in (None, "")):
            continue

        sheet_target.cell(append_row, col_name).value = name
        sheet_target.cell(append_row, col_id).value = sid
        sheet_target.cell(append_row, col_bday).value = parse_date(sheet_member.cell(member_row, m_bday).value)
        sheet_target.cell(append_row, col_tel).value = sheet_member.cell(member_row, m_tel).value
        sheet_target.cell(append_row, col_abc).value = sheet_member.cell(member_row, m_abc).value
        sheet_target.cell(append_row, col_dmk).value = sheet_member.cell(member_row, m_dmk).value
        sheet_target.cell(append_row, col_cnt).value = sheet_member.cell(member_row, m_cnt).value
        sheet_target.cell(append_row, col_sex).value = infer_gender_from_id(sid)

        added_count += 1
        append_row += 1

    # Build ID -> row
    id_to_row: Dict[str, int] = {}
    for r in range(header_row_target + 1, sheet_target.max_row + 1):
        vid = sheet_target.cell(r, col_id).value
        if vid is None:
            continue
        vid_s = str(vid).strip()
        if vid_s and vid_s not in id_to_row:
            id_to_row[vid_s] = r

    # ASCVD fill (ASCVD != 0)
    ascvd_header_row = find_header_row(sheet_ascvd, ["ID", "ASCVD"], search_rows=80)
    if ascvd_header_row is None:
        raise ValueError("ascvd 表找不到同時包含「ID」與「ASCVD」的標題列")

    amap = build_header_map(sheet_ascvd, ascvd_header_row)
    a_id = find_column(amap, ["ID", "id"])
    a_asc = find_column(amap, ["ASCVD", "ascvd"])
    if a_id is None or a_asc is None:
        raise ValueError("ascvd 表找不到「ID」或「ASCVD」欄位")

    ascvd_filled = 0
    for row_idx in range(ascvd_header_row + 1, sheet_ascvd.max_row + 1):
        pid_val = sheet_ascvd.cell(row_idx, a_id).value
        ascvd_val = sheet_ascvd.cell(row_idx, a_asc).value
        if pid_val is None or str(pid_val).strip() == "":
            continue
        if ascvd_val is None or str(ascvd_val).strip() == "":
            continue
        if str(ascvd_val).strip() == "0":
            continue
        target_row = id_to_row.get(str(pid_val).strip())
        if target_row is not None:
            sheet_target.cell(target_row, col_ascvd).value = ascvd_val
            ascvd_filled += 1

    # Screening fill (ID + 最後篩檢日期)
    def fill_screening(source_sheet, target_col: int) -> int:
        src_map = build_header_map(source_sheet, 1)
        src_id = find_column(src_map, ["ID", "身分證號", "身份證號"])
        src_dt = find_column(src_map, ["最後篩檢日期"])
        if src_id is None or src_dt is None:
            raise ValueError(f"「{source_sheet.title}」找不到 ID 或 最後篩檢日期 欄位")

        filled = 0
        for row_idx in range(2, source_sheet.max_row + 1):
            pid_val = source_sheet.cell(row_idx, src_id).value
            dt_val = parse_date(source_sheet.cell(row_idx, src_dt).value)
            if pid_val is None or str(pid_val).strip() == "":
                continue
            if dt_val is None:
                continue
            target_row = id_to_row.get(str(pid_val).strip())
            if target_row is not None:
                sheet_target.cell(target_row, target_col).value = dt_val
                filled += 1
        return filled

    adult_filled = fill_screening(sheet_adult, col_adult)
    pap_filled = fill_screening(sheet_pap, col_pap)
    flu_filled = fill_screening(sheet_flu, col_flu)
    fit_filled = fill_screening(sheet_fit, col_fit)
    hep_filled = fill_screening(sheet_hep, col_hep)

    # HealthCase fill
    hc_map = build_header_map(sheet_health, 1)
    hc_id = find_column(hc_map, ["家醫收案會員ID", "ID"])
    hc_hba = find_column(hc_map, ["最近一次HbA1c檢查結果(%)"])
    hc_hba_dt = find_column(hc_map, ["最近一次HbA1c檢查日期"])
    hc_ldl = find_column(hc_map, ["最近一次LDL檢查結果(mg/dL)"])
    hc_ldl_dt = find_column(hc_map, ["最近一次LDL檢查日期"])
    hc_uacr = find_column(hc_map, ["最近一次UACR檢查結果(mg/gm)"])
    hc_uacr_dt = find_column(hc_map, ["最近一次UACR檢查日期"])

    hc_missing = [k for k, v in {
        "家醫收案會員ID": hc_id,
        "HbA1c結果": hc_hba,
        "HbA1c日期": hc_hba_dt,
        "LDL結果": hc_ldl,
        "LDL日期": hc_ldl_dt,
        "UACR結果": hc_uacr,
        "UACR日期": hc_uacr_dt,
    }.items() if v is None]
    if hc_missing:
        raise ValueError("HealthCase 缺少欄位：" + "、".join(hc_missing))

    filled_hba = filled_ldl = filled_uacr = 0
    for row_idx in range(2, sheet_health.max_row + 1):
        pid_val = sheet_health.cell(row_idx, hc_id).value
        if pid_val is None or str(pid_val).strip() == "":
            continue
        target_row = id_to_row.get(str(pid_val).strip())
        if target_row is None:
            continue

        hba_val = parse_float(sheet_health.cell(row_idx, hc_hba).value)
        if hba_val is not None and hba_val != 0:
            sheet_target.cell(target_row, col_hba1c).value = hba_val
            filled_hba += 1
        hba_dt_val = parse_date(sheet_health.cell(row_idx, hc_hba_dt).value)
        if hba_dt_val is not None:
            sheet_target.cell(target_row, col_hba1c_date).value = hba_dt_val

        ldl_val = parse_float(sheet_health.cell(row_idx, hc_ldl).value)
        if ldl_val is not None and ldl_val != 0:
            sheet_target.cell(target_row, col_ldl).value = ldl_val
            filled_ldl += 1
        ldl_dt_val = parse_date(sheet_health.cell(row_idx, hc_ldl_dt).value)
        if ldl_dt_val is not None:
            sheet_target.cell(target_row, col_ldl_date).value = ldl_dt_val

        uacr_val = parse_float(sheet_health.cell(row_idx, hc_uacr).value)
        if uacr_val is not None and uacr_val != 0:
            sheet_target.cell(target_row, col_uacr).value = uacr_val
            filled_uacr += 1
        uacr_dt_val = parse_date(sheet_health.cell(row_idx, hc_uacr_dt).value)
        if uacr_dt_val is not None:
            sheet_target.cell(target_row, col_uacr_date).value = uacr_dt_val

    # Style + date format
    style_sheet(sheet_target, font_name="標楷體", font_size=19)
    set_date_format(
        sheet_target,
        [col_bday, col_adult, col_pap, col_flu, col_fit, col_hep, col_hba1c_date, col_ldl_date, col_uacr_date],
        start_row=header_row_target + 1,
        fmt="yyyy-mm-dd",
    )

    # Output
    base_dir = os.path.dirname(os.path.abspath(xlsx_path))
    base_name = os.path.splitext(os.path.basename(xlsx_path))[0]
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    out_path = os.path.join(base_dir, "{}_處理完成_{}.xlsx".format(base_name, ts))
    workbook.save(out_path)

    print("完成！輸出檔案：", out_path)
    print("新增名單：", added_count)
    print("ASCVD 回填：", ascvd_filled)
    print("篩檢日期回填：成人健檢{}、子宮抹片{}、老人流感{}、糞便潛血{}、肝炎篩檢{}".format(
        adult_filled, pap_filled, flu_filled, fit_filled, hep_filled
    ))
    print("HealthCase 回填：HbA1c{}、LDL{}、UACR{}".format(filled_hba, filled_ldl, filled_uacr))

    return out_path


def main() -> None:
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()

    xlsx_path = filedialog.askopenfilename(
        title="選擇 Excel 檔案",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not xlsx_path:
        return

    try:
        out_path = process_excel(xlsx_path)
        messagebox.showinfo("完成", "已輸出：\n{}".format(out_path))
    except (ValueError, KeyError, OSError, InvalidFileException) as e:
        messagebox.showerror("錯誤", str(e))


if __name__ == "__main__":
    main()
