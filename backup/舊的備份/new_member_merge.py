# -*- coding: utf-8 -*-
"""
整合同一個 Excel 檔案內所有 sheets 的基本資料，
以身分證為 key 去重複、補空白欄位，
最後「用新的整合結果取代原本的會員名單」分頁。

規則：
1. 以「身分證」為唯一 key。
2. 身分證先清洗：去空白、去特殊字元、轉大寫。
3. 依工作表順序整合，前面的 sheet 優先；後面的 sheet 只補空白，不覆蓋已存在值。
4. 若後面的 sheet 出現新身分證，則新增一列。
5. 表頭列不一定在第 1 列，會自動偵測。
6. 原本的「會員名單」會先納入整合來源，再被刪除。
7. 最後建立一張新的「會員名單」在最前面，等於取代舊的「會員名單」。
8. 完成後自動打開輸出檔所在資料夾。

輸出欄位：
- 會員姓名
- 會員身份証
- 會員生日
- 電話
- 會員別
- 疾病樣態
- ascvd

使用方式：
python new_member_merge.py
或
python new_member_merge.py /path/to/input.xlsx
"""

from __future__ import annotations

import os
import re
import sys
import subprocess
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

try:
    import tkinter as tk
    from tkinter import filedialog
except Exception:
    tk = None
    filedialog = None


TARGET_SHEET_NAME = "會員名單"
OUTPUT_HEADERS = [
    "會員姓名",
    "會員身份証",
    "會員生日",
    "電話",
    "會員別",
    "疾病樣態",
    "ascvd",
]

CANONICAL_FIELDS = {
    "name": "會員姓名",
    "idno": "會員身份証",
    "birthday": "會員生日",
    "phone": "電話",
    "member_type": "會員別",
    "disease": "疾病樣態",
    "ascvd": "ascvd",
}

ALIASES: Dict[str, List[str]] = {
    "name": [
        "會員姓名", "姓名", "name", "個案姓名", "病患姓名", "患者姓名",
    ],
    "idno": [
        "會員身份証", "會員身份證", "會員身分證", "會員身分証",
        "身份證號", "身分證號", "身份証號", "身分証號",
        "身份證字號", "身分證字號", "證號", "id", "idno",
        "家醫收案會員id", "會員id", "個案id", "病歷號+id",
        "身份證", "身分證",
    ],
    "birthday": [
        "會員生日", "會員生曰", "生日", "出生年月日", "出生日期", "birth", "birthday", "bday",
    ],
    "phone": [
        "電話", "電話號碼", "聯絡電話", "手機", "手機號碼", "tel", "phone", "mobile",
    ],
    "member_type": [
        "會員別", "個案類別", "會員類別", "收案類別", "家醫會員別",
    ],
    "disease": [
        "疾病樣態", "疾病型態", "疾病", "疾病註記", "疾病樣態註記",
    ],
    "ascvd": [
        "ascvd", "ASCVD", "ascvd分類", "ascvd類別", "風險等級", "心血管風險",
    ],
}


@dataclass
class MemberRecord:
    order: int
    values: Dict[str, Any] = field(default_factory=dict)
    source_sheet: str = ""


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\u3000", " ").replace("\t", " ").strip().lstrip("'")


def normalize_header(value: Any) -> str:
    s = normalize_text(value)
    s = s.replace("\n", "").replace(" ", "")
    s = s.replace("（", "(").replace("）", ")")
    s = s.replace("証", "證")
    s = s.replace("身份", "身分")
    s = s.replace("生曰", "生日")
    return s.lower()


def normalize_id(value: Any) -> str:
    s = normalize_text(value).upper()
    if not s:
        return ""
    return re.sub(r"[^A-Z0-9]", "", s)


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return normalize_text(value) == ""
    return False


def choose_input_file() -> str:
    if len(sys.argv) >= 2:
        return sys.argv[1]
    if tk is None or filedialog is None:
        raise SystemExit("請在命令列提供 Excel 路徑，例如：python new_member_merge.py input.xlsx")

    root = tk.Tk()
    root.withdraw()
    path = filedialog.askopenfilename(
        title="選擇要整合的 Excel 檔案",
        filetypes=[("Excel files", "*.xlsx *.xlsm *.xltx *.xltm")],
    )
    root.destroy()
    if not path:
        raise SystemExit("未選擇檔案，已取消。")
    return path


def open_folder_cross_platform(path: str) -> None:
    folder = os.path.dirname(os.path.abspath(path))
    try:
        if sys.platform.startswith("darwin"):
            subprocess.Popen(["open", folder])
        elif os.name == "nt":
            os.startfile(folder)  # type: ignore[attr-defined]
        else:
            subprocess.Popen(["xdg-open", folder])
    except Exception:
        pass


def build_alias_lookup() -> Dict[str, str]:
    lookup: Dict[str, str] = {}
    for field, names in ALIASES.items():
        for name in names:
            lookup[normalize_header(name)] = field
    return lookup


ALIAS_LOOKUP = build_alias_lookup()


def find_header_row(ws: Worksheet, max_scan_rows: int = 30) -> Tuple[Optional[int], Dict[str, int]]:
    best_row: Optional[int] = None
    best_map: Dict[str, int] = {}
    best_score = -1

    max_row = min(ws.max_row, max_scan_rows)
    max_col = ws.max_column

    for r in range(1, max_row + 1):
        field_to_col: Dict[str, int] = {}
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            hv = normalize_header(v)
            if not hv:
                continue

            matched_field = ALIAS_LOOKUP.get(hv)
            if matched_field and matched_field not in field_to_col:
                field_to_col[matched_field] = c
                continue

            if "ascvd" in hv and "ascvd" not in field_to_col:
                field_to_col["ascvd"] = c
            elif "姓名" in hv and "name" not in field_to_col:
                field_to_col["name"] = c
            elif (("生日" in hv) or ("出生" in hv)) and "birthday" not in field_to_col:
                field_to_col["birthday"] = c
            elif (("電話" in hv) or ("手機" in hv) or ("tel" in hv) or ("phone" in hv)) and "phone" not in field_to_col:
                field_to_col["phone"] = c
            elif (("會員別" in hv) or ("個案類別" in hv) or ("會員類別" in hv)) and "member_type" not in field_to_col:
                field_to_col["member_type"] = c
            elif (("疾病樣態" in hv) or ("疾病型態" in hv) or (hv == "疾病")) and "disease" not in field_to_col:
                field_to_col["disease"] = c
            elif (
                hv == "id"
                or "身分證" in hv or "身份證" in hv
                or "字號" in hv or "證號" in hv
                or "會員id" in hv or "家醫收案會員id" in hv
            ) and "idno" not in field_to_col:
                field_to_col["idno"] = c

        score = len(field_to_col)
        if "idno" not in field_to_col:
            continue
        if score > best_score:
            best_score = score
            best_row = r
            best_map = field_to_col

    return best_row, best_map


def get_cell_value(ws: Worksheet, row_idx: int, col_idx: Optional[int]) -> Any:
    if not col_idx:
        return None
    return ws.cell(row_idx, col_idx).value


def record_from_row(ws: Worksheet, row_idx: int, field_map: Dict[str, int]) -> Tuple[str, Dict[str, Any]]:
    raw_id = get_cell_value(ws, row_idx, field_map.get("idno"))
    clean_id = normalize_id(raw_id)
    values = {
        "name": get_cell_value(ws, row_idx, field_map.get("name")),
        "idno": clean_id,
        "birthday": get_cell_value(ws, row_idx, field_map.get("birthday")),
        "phone": get_cell_value(ws, row_idx, field_map.get("phone")),
        "member_type": get_cell_value(ws, row_idx, field_map.get("member_type")),
        "disease": get_cell_value(ws, row_idx, field_map.get("disease")),
        "ascvd": get_cell_value(ws, row_idx, field_map.get("ascvd")),
    }
    return clean_id, values


def get_source_sheets(wb: Workbook) -> List[Worksheet]:
    sheets = list(wb.worksheets)

    member_sheet = wb[TARGET_SHEET_NAME] if TARGET_SHEET_NAME in wb.sheetnames else None
    others = [ws for ws in sheets if ws.title != TARGET_SHEET_NAME]

    ordered: List[Worksheet] = []
    if member_sheet is not None:
        ordered.append(member_sheet)
    ordered.extend(others)
    return ordered


def merge_workbook(wb: Workbook) -> List[MemberRecord]:
    merged: Dict[str, MemberRecord] = {}
    order_counter = 0

    for ws in get_source_sheets(wb):
        header_row, field_map = find_header_row(ws)
        if not header_row or "idno" not in field_map:
            continue

        for r in range(header_row + 1, ws.max_row + 1):
            clean_id, row_values = record_from_row(ws, r, field_map)
            if not clean_id:
                continue

            if clean_id not in merged:
                order_counter += 1
                merged[clean_id] = MemberRecord(
                    order=order_counter,
                    values=dict(row_values),
                    source_sheet=ws.title,
                )
                continue

            rec = merged[clean_id]
            for field in CANONICAL_FIELDS.keys():
                if field == "idno":
                    continue
                existing = rec.values.get(field)
                incoming = row_values.get(field)
                if is_blank(existing) and not is_blank(incoming):
                    rec.values[field] = incoming

    return sorted(merged.values(), key=lambda x: x.order)


def remove_target_sheet_if_exists(wb: Workbook) -> None:
    if TARGET_SHEET_NAME in wb.sheetnames:
        wb.remove(wb[TARGET_SHEET_NAME])


def create_output_sheet(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet(title=TARGET_SHEET_NAME, index=0)
    for c, title in enumerate(OUTPUT_HEADERS, start=1):
        ws.cell(1, c).value = title
    ws.freeze_panes = "A2"
    return ws


def auto_width(ws: Worksheet) -> None:
    widths = {
        1: 18,
        2: 16,
        3: 14,
        4: 18,
        5: 12,
        6: 14,
        7: 10,
    }
    for col_idx, width in widths.items():
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width


def write_output(ws: Worksheet, records: List[MemberRecord]) -> None:
    for row_idx, rec in enumerate(records, start=2):
        ws.cell(row_idx, 1).value = rec.values.get("name")
        ws.cell(row_idx, 2).value = rec.values.get("idno")
        ws.cell(row_idx, 3).value = rec.values.get("birthday")
        ws.cell(row_idx, 4).value = rec.values.get("phone")
        ws.cell(row_idx, 5).value = rec.values.get("member_type")
        ws.cell(row_idx, 6).value = rec.values.get("disease")
        ws.cell(row_idx, 7).value = rec.values.get("ascvd")

    auto_width(ws)


def build_output_path(input_path: str) -> str:
    base, ext = os.path.splitext(input_path)
    return f"{base}_會員名單整合{ext}"


def process_excel(input_path: str) -> str:
    wb = openpyxl.load_workbook(input_path)

    records = merge_workbook(wb)

    # 舊的「會員名單」先刪除，再由新的整合結果取代
    remove_target_sheet_if_exists(wb)
    ws_out = create_output_sheet(wb)
    write_output(ws_out, records)

    output_path = build_output_path(input_path)
    wb.save(output_path)
    return output_path


def main() -> None:
    input_path = choose_input_file()
    if not os.path.exists(input_path):
        raise SystemExit(f"找不到檔案：{input_path}")

    try:
        out = process_excel(input_path)
    except Exception as e:
        raise SystemExit(f"處理失敗：{e}")

    print("完成")
    print(f"輸出檔案：{out}")
    open_folder_cross_platform(out)


if __name__ == "__main__":
    main()
