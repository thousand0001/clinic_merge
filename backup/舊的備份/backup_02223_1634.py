# -*- coding: utf-8 -*-
import os
import re
import sys
import datetime
import subprocess
from typing import Optional, Dict, Any, Tuple

import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.utils import column_index_from_string
from openpyxl.utils.exceptions import InvalidFileException


# =========================
# 固定設定
# =========================
TEMPLATE_NAME = "會員指標模板.xlsx"   # 模板檔名：同資料夾
SHEET_TARGET = "會員指標"

# 你指定：打勾欄位位置
COL_CHK_HBA = column_index_from_string("AK")
COL_CHK_LDL = column_index_from_string("AL")
COL_CHK_UACR = column_index_from_string("AM")
CHECK = "✓"


# =========================
# 小工具
# =========================
def open_file_cross_platform(path: str) -> None:
    try:
        if sys.platform.startswith("darwin"):
            subprocess.call(("open", path))
        elif os.name == "nt":
            os.startfile(path)  # type: ignore[attr-defined]
        else:
            subprocess.call(("xdg-open", path))
    except Exception:
        pass


def normalize_text(v: Any) -> str:
    if v is None:
        return ""
    return str(v).replace("\t", "").replace("　", "").strip()


def normalize_header(v: Any) -> str:
    """表頭用：移除空白+換行，方便模糊比對"""
    s = normalize_text(v)
    s = s.replace(" ", "").replace("\n", "")
    return s


def build_header_map(sheet, header_row: int) -> Dict[str, int]:
    """完全匹配用（保留原功能）"""
    m: Dict[str, int] = {}
    for c in range(1, sheet.max_column + 1):
        k = normalize_text(sheet.cell(header_row, c).value)
        if k:
            m[k] = c
    return m


def find_column_exact(hmap: Dict[str, int], aliases) -> Optional[int]:
    for a in aliases:
        if a in hmap:
            return hmap[a]
    return None


def find_col_by_keywords(sheet, header_row: int, keywords) -> Optional[int]:
    """
    模糊比對：表頭(去空白/去換行) 必須同時包含 keywords 的所有字
    例：keywords=["HbA1c","日期"] 可匹配 "最近一次HbA1c檢查日期"
    """
    keys = [k.replace(" ", "").replace("\n", "") for k in keywords]
    for c in range(1, sheet.max_column + 1):
        h = normalize_header(sheet.cell(header_row, c).value)
        if not h:
            continue
        ok = True
        for k in keys:
            if k and k not in h:
                ok = False
                break
        if ok:
            return c
    return None


def find_header_row_contains(sheet, must_have, search_rows: int = 250) -> Optional[int]:
    """
    找到同一列同時包含 must_have 幾個表頭（模糊等於：去空白/去換行後比對）
    """
    must = [m.replace(" ", "").replace("\n", "") for m in must_have]
    for r in range(1, min(search_rows, sheet.max_row) + 1):
        found = set()
        for c in range(1, sheet.max_column + 1):
            v = normalize_header(sheet.cell(r, c).value)
            if not v:
                continue
            for m in must:
                if v == m:
                    found.add(m)
        if all(m in found for m in must):
            return r
    return None


def parse_date(value: Any) -> Optional[datetime.date]:
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value

    s = str(value).strip()
    if s == "" or s in ("-", "—", "–"):
        return None

    # yyyy/mm/dd or yyyy-mm-dd
    m = re.match(r"^(\d{4})[/-](\d{1,2})[/-](\d{1,2})$", s)
    if m:
        y, mo, d = map(int, m.groups())
        return datetime.date(y, mo, d)

    # ROC yyyMMdd
    m = re.match(r"^(\d{2,3})(\d{2})(\d{2})$", s)
    if m:
        roc, mo, d = map(int, m.groups())
        return datetime.date(roc + 1911, mo, d)

    # ROC yyy/mm/dd
    m = re.match(r"^(\d{2,3})[/-](\d{1,2})[/-](\d{1,2})$", s)
    if m:
        roc, mo, d = map(int, m.groups())
        return datetime.date(roc + 1911, mo, d)

    return None


def parse_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    s = str(value).strip()
    if s == "" or s in ("-", "—", "–"):
        return None
    try:
        return float(s)
    except Exception:
        return None


def infer_gender_from_id(id_value: Any) -> str:
    if not id_value:
        return ""
    s = str(id_value).strip().upper()
    if len(s) < 2:
        return ""
    g = s[1]
    if g in ("1", "8"):
        return "男"
    if g in ("2", "9"):
        return "女"
    return ""


def clean_spaces(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value))


def parse_dmk_to_code(v: Any) -> Optional[int]:
    """
    疾病樣態內容 1:DM,2:CKD,3:DKD(DM+CKD),4:none
    """
    s = clean_spaces(v)
    if s == "":
        return None
    if s.isdigit():
        n = int(s)
        if n in (1, 2, 3, 4):
            return n
    su = s.upper()
    if su == "DM":
        return 1
    if su == "CKD":
        return 2
    if su in ("DKD", "DM+CKD", "DMCKD"):
        return 3
    if su in ("NONE", "NA", "N/A", "0"):
        return 4
    # 兜底
    has_dm = "DM" in su
    has_ckd = "CKD" in su or "DKD" in su
    if has_dm and has_ckd:
        return 3
    if has_dm:
        return 1
    if has_ckd:
        return 2
    return 4


def classify_ascvd(v: Any) -> Optional[str]:
    """
    ASCVD 分 a/b（a 極高、b 非常高）
    可能來源：a/b/中文/1/2
    """
    if v is None:
        return None
    s = clean_spaces(v).lower()
    if s == "" or s == "0":
        return None
    if s in ("a", "1") or "極高" in s:
        return "a"
    if s in ("b", "2") or "非常高" in s:
        return "b"
    # 兜底：含字母
    if "a" in s:
        return "a"
    if "b" in s:
        return "b"
    return None


def calc_age(bday: Optional[datetime.date], ref: datetime.date) -> int:
    if not isinstance(bday, datetime.date):
        return -1
    y = ref.year - bday.year
    if (ref.month, ref.day) < (bday.month, bday.day):
        y -= 1
    return y


def add_full_grid(ws) -> None:
    thin = Side(style="thin")
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).border = grid


# =========================
# 主流程
# =========================
def process_excel(source_xlsx_path: str, template_xlsx_path: str) -> str:
    # --- 讀原始檔（未含會員指標） ---
    wb_src = openpyxl.load_workbook(source_xlsx_path, data_only=True)
    need_sheets = ["會員名單", "ascvd", "HealthCase", "成人健檢", "子宮抹片", "老人流感", "糞便潛血", "肝炎篩檢"]
    for sn in need_sheets:
        if sn not in wb_src.sheetnames:
            raise ValueError(f"原始檔缺少工作表：{sn}")

    sh_member = wb_src["會員名單"]
    sh_ascvd = wb_src["ascvd"]
    sh_health = wb_src["HealthCase"]
    sh_adult = wb_src["成人健檢"]
    sh_pap = wb_src["子宮抹片"]
    sh_flu = wb_src["老人流感"]
    sh_fit = wb_src["糞便潛血"]
    sh_hep = wb_src["肝炎篩檢"]

    # --- 讀模板（只保留會員指標） ---
    wb_tpl = openpyxl.load_workbook(template_xlsx_path)
    if SHEET_TARGET not in wb_tpl.sheetnames:
        raise ValueError("模板檔缺少工作表：會員指標")

    for sn in list(wb_tpl.sheetnames):
        if sn != SHEET_TARGET:
            del wb_tpl[sn]
    ws = wb_tpl[SHEET_TARGET]

    # --- 找模板表頭列 ---
    header_row = find_header_row_contains(ws, ["診所名稱或機構代碼", "姓名", "身份證號碼"], 250) or 5

    # --- 用「模糊比對」找欄位（避免換行/說明文字） ---
    col_clinic = find_col_by_keywords(ws, header_row, ["診所名稱或機構代碼"])
    col_name = find_col_by_keywords(ws, header_row, ["姓名"])
    col_id = find_col_by_keywords(ws, header_row, ["身份證號碼"])
    col_bday = find_col_by_keywords(ws, header_row, ["生日"])
    col_tel = find_col_by_keywords(ws, header_row, ["電話"])
    col_cnt = find_col_by_keywords(ws, header_row, ["次數"])
    col_sex = find_col_by_keywords(ws, header_row, ["性別"])
    col_abc = find_col_by_keywords(ws, header_row, ["A/B/C"])  # ★ 修正：模糊抓 A/B/C
    col_dmk = find_col_by_keywords(ws, header_row, ["DM/CKD/DKD"])
    col_ascvd = find_col_by_keywords(ws, header_row, ["ASCVD"])

    col_adult = find_col_by_keywords(ws, header_row, ["成人預防保健"])
    col_pap = find_col_by_keywords(ws, header_row, ["子宮", "抹片"])
    col_flu = find_col_by_keywords(ws, header_row, ["流感"])
    col_fit = find_col_by_keywords(ws, header_row, ["糞便", "潛血"])
    col_hep = find_col_by_keywords(ws, header_row, ["肝炎"])

    col_hba = find_col_by_keywords(ws, header_row, ["HbA1c"])
    col_ldl = find_col_by_keywords(ws, header_row, ["LDL"])
    col_uacr = find_col_by_keywords(ws, header_row, ["UACR"])

    # ★ 你確認：日期欄位就是這三個（用模糊比對）
    col_hba_dt = find_col_by_keywords(ws, header_row, ["HbA1c", "日期"])
    col_ldl_dt = find_col_by_keywords(ws, header_row, ["LDL", "日期"])
    col_uacr_dt = find_col_by_keywords(ws, header_row, ["UACR", "日期"])

    must = [
        ("診所名稱或機構代碼", col_clinic),
        ("姓名", col_name),
        ("身份證號碼", col_id),
        ("生日", col_bday),
        ("電話", col_tel),
        ("A/B/C", col_abc),
        ("DM/CKD/DKD", col_dmk),
        ("次數", col_cnt),
        ("ASCVD", col_ascvd),
        ("性別", col_sex),
        ("HbA1c", col_hba),
        ("HbA1c日期", col_hba_dt),
        ("LDL", col_ldl),
        ("LDL日期", col_ldl_dt),
        ("UACR", col_uacr),
        ("UACR日期", col_uacr_dt),
    ]
    missing = [k for k, v in must if v is None]
    if missing:
        raise ValueError("模板會員指標缺少欄位/欄名找不到：" + "、".join(missing))

    # --- 清資料列（不動表頭/樣式）---
    DATA_START = 9
    cols_to_clear = [
        col_clinic, col_name, col_id, col_bday, col_tel, col_cnt, col_sex, col_abc, col_dmk, col_ascvd,
        col_adult, col_pap, col_flu, col_fit, col_hep,
        col_hba, col_hba_dt, col_ldl, col_ldl_dt, col_uacr, col_uacr_dt,
        COL_CHK_HBA, COL_CHK_LDL, COL_CHK_UACR
    ]
    cols_to_clear = [c for c in cols_to_clear if c]
    for r in range(DATA_START, ws.max_row + 1):
        for c in cols_to_clear:
            ws.cell(r, c).value = None

    # --- 會員名單欄位 ---
    member_header_row = 5
    mmap = build_header_map(sh_member, member_header_row)
    m_name = find_column_exact(mmap, ["會員姓名"])
    m_id = find_column_exact(mmap, ["會員身份証", "會員身份證", "會員身分證"])
    m_bday = find_column_exact(mmap, ["會員生日"])
    m_tel = find_column_exact(mmap, ["電話"])
    m_abc = find_column_exact(mmap, ["會員別"])
    m_dmk = find_column_exact(mmap, ["疾病樣態"])
    m_cnt = find_column_exact(mmap, ["就診次數"])
    if any(v is None for v in [m_name, m_id, m_bday, m_tel, m_abc, m_dmk, m_cnt]):
        raise ValueError("原始檔「會員名單」欄位不完整（會員姓名/會員身分證/會員生日/電話/會員別/疾病樣態/就診次數）")

    clinic_val = normalize_text(sh_member["A1"].value)

    # --- 寫入基本資料，建 id_to_row + meta ---
    id_to_row: Dict[str, int] = {}
    meta: Dict[int, Dict[str, Any]] = {}

    out_r = DATA_START
    for r in range(member_header_row + 1, sh_member.max_row + 1):
        sid = sh_member.cell(r, m_id).value
        name = sh_member.cell(r, m_name).value
        if (sid in (None, "")) and (name in (None, "")):
            continue

        sid_s = str(sid).strip()
        bday = parse_date(sh_member.cell(r, m_bday).value)
        dmk_code = parse_dmk_to_code(sh_member.cell(r, m_dmk).value)

        ws.cell(out_r, col_clinic).value = clinic_val
        ws.cell(out_r, col_name).value = name
        ws.cell(out_r, col_id).value = sid
        ws.cell(out_r, col_bday).value = bday
        ws.cell(out_r, col_tel).value = sh_member.cell(r, m_tel).value
        ws.cell(out_r, col_abc).value = sh_member.cell(r, m_abc).value
        ws.cell(out_r, col_dmk).value = dmk_code
        ws.cell(out_r, col_cnt).value = sh_member.cell(r, m_cnt).value
        ws.cell(out_r, col_sex).value = infer_gender_from_id(sid)

        if sid_s and sid_s not in id_to_row:
            id_to_row[sid_s] = out_r
            meta[out_r] = {"bday": bday, "dmk": dmk_code, "ascvd": None}

        out_r += 1

    last_row = out_r - 1

    # --- ASCVD 回填 ---
    ascvd_header_row = 5
    amap = build_header_map(sh_ascvd, ascvd_header_row)
    a_id = find_column_exact(amap, ["ID", "id"])
    a_asc = find_column_exact(amap, ["ASCVD", "ascvd"])
    if a_id is None or a_asc is None:
        raise ValueError("原始檔「ascvd」找不到 ID / ASCVD 欄位（表頭列預期第5列）")

    for r in range(ascvd_header_row + 1, sh_ascvd.max_row + 1):
        pid = sh_ascvd.cell(r, a_id).value
        val = sh_ascvd.cell(r, a_asc).value
        if pid is None or str(pid).strip() == "":
            continue
        if val is None or str(val).strip() == "" or str(val).strip() == "0":
            continue
        tr = id_to_row.get(str(pid).strip())
        if tr:
            ws.cell(tr, col_ascvd).value = val
            meta.setdefault(tr, {})["ascvd"] = val

    # --- 5 大篩檢日期回填 ---
    def fill_screening(sheet, target_col: Optional[int]) -> None:
        if not target_col:
            return
        hmap = build_header_map(sheet, 1)
        sid_col = find_column_exact(hmap, ["ID", "身分證號", "身份證號"])
        dt_col = find_column_exact(hmap, ["最後篩檢日期"])
        if sid_col is None or dt_col is None:
            raise ValueError(f"「{sheet.title}」找不到 ID 或 最後篩檢日期 欄位")

        for rr in range(2, sheet.max_row + 1):
            pid = sheet.cell(rr, sid_col).value
            dt = parse_date(sheet.cell(rr, dt_col).value)
            if pid is None or str(pid).strip() == "" or dt is None:
                continue
            tr = id_to_row.get(str(pid).strip())
            if tr:
                ws.cell(tr, target_col).value = dt

    fill_screening(sh_adult, col_adult)
    fill_screening(sh_pap, col_pap)
    fill_screening(sh_flu, col_flu)
    fill_screening(sh_fit, col_fit)
    fill_screening(sh_hep, col_hep)

    # --- HealthCase 檢驗回填 ---
    hc_map = build_header_map(sh_health, 1)
    hc_id = find_column_exact(hc_map, ["家醫收案會員ID", "ID"])
    hc_hba = find_column_exact(hc_map, ["最近一次HbA1c檢查結果(%)"])
    hc_hba_dt = find_column_exact(hc_map, ["最近一次HbA1c檢查日期"])
    hc_ldl = find_column_exact(hc_map, ["最近一次LDL檢查結果(mg/dL)"])
    hc_ldl_dt = find_column_exact(hc_map, ["最近一次LDL檢查日期"])
    hc_uacr = find_column_exact(hc_map, ["最近一次UACR檢查結果(mg/gm)"])
    hc_uacr_dt = find_column_exact(hc_map, ["最近一次UACR檢查日期"])

    if any(v is None for v in [hc_id, hc_hba, hc_hba_dt, hc_ldl, hc_ldl_dt, hc_uacr, hc_uacr_dt]):
        raise ValueError("原始檔「HealthCase」欄位不完整（家醫收案會員ID / HbA1c結果+日期 / LDL結果+日期 / UACR結果+日期）")

    for r in range(2, sh_health.max_row + 1):
        pid = sh_health.cell(r, hc_id).value
        if pid is None or str(pid).strip() == "":
            continue
        tr = id_to_row.get(str(pid).strip())
        if not tr:
            continue

        hba = parse_float(sh_health.cell(r, hc_hba).value)
        if hba is not None and hba != 0:
            ws.cell(tr, col_hba).value = hba
        hba_dt = parse_date(sh_health.cell(r, hc_hba_dt).value)
        if hba_dt:
            ws.cell(tr, col_hba_dt).value = hba_dt

        ldl = parse_float(sh_health.cell(r, hc_ldl).value)
        if ldl is not None and ldl != 0:
            ws.cell(tr, col_ldl).value = ldl
        ldl_dt = parse_date(sh_health.cell(r, hc_ldl_dt).value)
        if ldl_dt:
            ws.cell(tr, col_ldl_dt).value = ldl_dt

        uacr = parse_float(sh_health.cell(r, hc_uacr).value)
        if uacr is not None and uacr != 0:
            ws.cell(tr, col_uacr).value = uacr
        uacr_dt = parse_date(sh_health.cell(r, hc_uacr_dt).value)
        if uacr_dt:
            ws.cell(tr, col_uacr_dt).value = uacr_dt

    # --- 打勾邏輯（寫入 AH/AI/AJ） ---
    now_tw = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=8))).date()

    def has_flags(dmk_code: Optional[int], ascvd_val: Any) -> Tuple[bool, bool, bool, Optional[str]]:
        has_dm = dmk_code in (1, 3)
        has_ckd = dmk_code in (2, 3)
        ascat = classify_ascvd(ascvd_val)
        has_ascvd = ascat is not None
        return has_dm, has_ckd, has_ascvd, ascat

    for r in range(DATA_START, last_row + 1):
        dmk_code = meta.get(r, {}).get("dmk")
        bday = meta.get(r, {}).get("bday")
        ascvd_val = meta.get(r, {}).get("ascvd")
        if ascvd_val is None:
            ascvd_val = ws.cell(r, col_ascvd).value

        age = calc_age(bday, now_tw)
        has_dm, has_ckd, has_ascvd, ascvd_cat = has_flags(dmk_code, ascvd_val)

        hba = ws.cell(r, col_hba).value
        ldl = ws.cell(r, col_ldl).value
        uacr = ws.cell(r, col_uacr).value

        # HbA1c
        # 1) DM：<8
        # 2) DM+CKD / DM+ASCVD / DM+CKD+ASCVD：>=80 <8；<80 <7
        if isinstance(hba, (int, float)) and hba != 0 and has_dm:
            if has_dm and (not has_ckd) and (not has_ascvd):  # DM
                if hba < 8.0:
                    ws.cell(r, COL_CHK_HBA).value = CHECK
            else:
                if age >= 80 and age != -1:
                    if hba < 8.0:
                        ws.cell(r, COL_CHK_HBA).value = CHECK
                else:
                    if hba < 7.0:
                        ws.cell(r, COL_CHK_HBA).value = CHECK

        # LDL
        # DM:<110
        # CKD:<140
        # ASCVDa:<80
        # ASCVDb:<100
        # DM+CKD:<110
        if isinstance(ldl, (int, float)) and ldl != 0:
            ok = False
            if has_dm and (not has_ckd) and (not has_ascvd):
                ok = ldl < 110.0
            elif has_ckd and (not has_dm) and (not has_ascvd):
                ok = ldl < 140.0
            elif has_ascvd and (not has_dm) and (not has_ckd):
                if ascvd_cat == "a":
                    ok = ldl < 80.0
                elif ascvd_cat == "b":
                    ok = ldl < 100.0
            elif has_dm and has_ckd and (not has_ascvd):
                ok = ldl < 110.0
            if ok:
                ws.cell(r, COL_CHK_LDL).value = CHECK

        # UACR
        # CKD:<35
        # DM+CKD<30
        # CKD+ASCVD<30
        # DM+CKD+ASCVD<30
        if isinstance(uacr, (int, float)) and uacr != 0:
            ok = False
            if has_ckd and (not has_dm) and (not has_ascvd):
                ok = uacr < 35.0
            elif has_dm and has_ckd and (not has_ascvd):
                ok = uacr < 30.0
            elif has_ckd and has_ascvd and (not has_dm):
                ok = uacr < 30.0
            elif has_dm and has_ckd and has_ascvd:
                ok = uacr < 30.0
            if ok:
                ws.cell(r, COL_CHK_UACR).value = CHECK

    # 日期欄格式（只格式實際 date）
    date_cols = [col_bday, col_adult, col_pap, col_flu, col_fit, col_hep, col_hba_dt, col_ldl_dt, col_uacr_dt]
    for r in range(DATA_START, last_row + 1):
        for c in date_cols:
            if not c:
                continue
            cell = ws.cell(r, c)
            if isinstance(cell.value, (datetime.date, datetime.datetime)):
                cell.number_format = "yyyy-mm-dd"

    # 全表格線
    add_full_grid(ws)

    # 輸出（台灣時間）
    base_dir = os.path.dirname(os.path.abspath(source_xlsx_path))
    ts = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=8))).strftime("%m%d_%H%M")
    out_path = os.path.join(base_dir, f"選會員{ts}.xlsx")
    wb_tpl.save(out_path)

    return out_path


def main() -> None:
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()

    src = filedialog.askopenfilename(
        title="選擇原始 Excel 檔案（沒有會員指標的那份）",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not src:
        return

    # 取得「目前這個 .py 檔」所在資料夾
    script_dir = os.path.dirname(os.path.abspath(__file__))
    template = os.path.join(script_dir, TEMPLATE_NAME)

    if not os.path.exists(template):
        messagebox.showerror("錯誤", f"找不到模板檔：\n{template}\n\n請把模板放到原始檔同資料夾，檔名需為：\n{TEMPLATE_NAME}")
        return

    try:
        out = process_excel(src, template)

        # 先顯示完成訊息
        messagebox.showinfo("完成", f"已輸出：\n{out}")

        # 使用者按 OK 後才開檔
        open_file_cross_platform(out)
    except (ValueError, KeyError, OSError, InvalidFileException) as e:
        messagebox.showerror("錯誤", str(e))
    except Exception as e:
        messagebox.showerror("錯誤", f"未預期錯誤：{e}")


if __name__ == "__main__":
    main()
