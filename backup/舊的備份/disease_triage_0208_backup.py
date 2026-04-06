# -*- coding: utf-8 -*-
"""
disease_triage_gui_clean.py  （乾淨版：GUI選檔 / 無參數 / 無template / 年度固定）

✅ 特點
- 不用任何命令列參數：PyCharm 直接 Run
- 依序跳出選檔：A檔、B檔、輸出檔
- 不需要 template：分析表由程式產生
- 年度固定：BASE_YEAR=2026（不再詢問輸入）
- B 檔自動偵測正確 sheet + header 列（常見第6列 header）
- 欄名採「模糊包含」匹配（避免『疾病樣態(說明...)』抓不到）
- 若仍找不到必要欄位，會彈窗顯示：偵測到的 sheet/header + 欄位列表（好除錯）
- 若有 merge 找不到 ID / 疾病樣態缺失，會自動輸出 _issues.csv（不影響 8 分頁）

輸入：
  A：個案健康管理報表（ODS 轉 xlsx；含最近一次 HbA1c/LDL/UACR 結果與日期）
  B：收案追蹤報表（某一列是欄名；含 ID/姓名/BIRTHDAY/疾病樣態/ASCVD）

輸出：
  out.xlsx：8 個分頁（分析表 + 7 分流表）
"""

import re
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

BASE_YEAR = 2026  # ← 固定，不再輸入


# ------------------------- helpers -------------------------
def norm_id(s: pd.Series) -> pd.Series:
    s = s.astype(str).str.strip()
    s = s.str.replace(r"\.0$", "", regex=True)
    s = s.str.replace(r"\s+", "", regex=True)
    s = s.replace({"nan": pd.NA, "None": pd.NA, "": pd.NA})
    return s


def _norm_colname(x: str) -> str:
    return str(x).strip().lower().replace("\n", "").replace(" ", "").replace("\u3000", "")


def pick_col_exact(cols: List[str], candidates: List[str]) -> Optional[str]:
    norm_map = {_norm_colname(c): c for c in cols}
    for cand in candidates:
        k = _norm_colname(cand)
        if k in norm_map:
            return norm_map[k]
    return None


def pick_col_contains(cols: List[str], keywords: List[str]) -> Optional[str]:
    norm_cols = [(_norm_colname(c), c) for c in cols]
    for kw in keywords:
        nkw = _norm_colname(kw)
        for nc, orig in norm_cols:
            if nkw and (nkw in nc):
                return orig
    return None


def to_dt(series: pd.Series) -> pd.Series:
    """日期解析：先 to_datetime，失敗再補民國格式（114/01/30 或 1140130）。"""
    s = series.astype(str).str.strip()
    s = s.replace({"nan": pd.NA, "None": pd.NA, "": pd.NA})

    dt = pd.to_datetime(s, errors="coerce")

    mask = dt.isna() & s.notna()
    if mask.any():

        def parse_minguo(x: str):
            x = str(x).strip()
            if not x or x.lower() in ("nan", "none"):
                return pd.NaT
            x = x.replace("-", "/")

            m = re.match(r"^(\d{2,3})/(\d{1,2})/(\d{1,2})$", x)  # 114/01/30
            if m:
                y, mo, d = map(int, m.groups())
                try:
                    return pd.Timestamp(year=y + 1911, month=mo, day=d)
                except Exception:
                    return pd.NaT

            m = re.match(r"^(\d{2,3})(\d{2})(\d{2})$", x)  # 1140130
            if m:
                y, mo, d = map(int, m.groups())
                try:
                    return pd.Timestamp(year=y + 1911, month=mo, day=d)
                except Exception:
                    return pd.NaT

            return pd.NaT

        dt2 = s[mask].map(parse_minguo)
        dt.loc[mask] = dt2

    return dt


def age_on(bday: pd.Timestamp, ref_date: pd.Timestamp):
    if pd.isna(bday):
        return None
    return ref_date.year - bday.year - ((ref_date.month, ref_date.day) < (bday.month, bday.day))


def ascvd_norm(x) -> str:
    v = str(x).strip().lower()
    return v if v in ("a", "b") else ""


def find_sheet_and_header_for_b(path_b: Path, probe_rows: int = 40) -> Tuple[str, int]:
    """
    自動找 B 檔 sheet + header 列（0-based）。
    規則：同一列同時包含 ID / BIRTHDAY / 疾病樣態 / ASCVD（允許欄名包含額外描述）。
    找不到時保險回傳：(第一張分頁, header=5) 也就是「第6列」。
    """
    xl = pd.ExcelFile(path_b)
    for sh in xl.sheet_names:
        probe = pd.read_excel(path_b, sheet_name=sh, header=None, nrows=probe_rows, dtype=str)
        for r in range(len(probe)):
            row_vals = probe.iloc[r].tolist()
            tokens = [_norm_colname(v) for v in row_vals if v is not None and str(v).strip() != ""]

            has_id = any(t == "id" for t in tokens)
            has_bday = any(t == "birthday" for t in tokens)
            has_dis = any("疾病樣態" in t for t in tokens)
            has_ascvd = any("ascvd" in t for t in tokens)

            if has_id and has_bday and has_dis and has_ascvd:
                return sh, r

    # fallback：你說 B 標題欄常在第6列
    return xl.sheet_names[0], 5


# ------------------------- rules -------------------------
def required_items(sheet_name: str) -> List[str]:
    if sheet_name in ("DM", "DM+ASCVD"):
        return ["HbA1c", "LDL"]
    if sheet_name in ("CKD", "CKD+ASCVD"):
        return ["LDL", "UACR"]
    if sheet_name in ("DKD", "DKD+ASCVD"):
        return ["HbA1c", "LDL", "UACR"]
    if sheet_name == "ASCVD":
        return ["LDL"]
    raise ValueError(sheet_name)


def ldl_target(row: pd.Series, sheet_name: str) -> float:
    a = ascvd_norm(row.get("ASCVD", ""))
    if a in ("a", "b") and sheet_name in ("ASCVD", "DM+ASCVD", "CKD+ASCVD", "DKD+ASCVD"):
        return 55.0 if a == "a" else 70.0
    if sheet_name == "CKD":
        return 130.0
    return 100.0


def remark_code(row: pd.Series, sheet_name: str, base_year: int) -> int:
    """
    回傳 0..5
    優先序：0 -> 3 -> 2 -> 1 -> 4 -> 5
    5：只看年份差 (base_year - last_year) >= 2（不看月日）
    """
    items = required_items(sheet_name)

    ref = pd.Timestamp(f"{base_year}-06-30")
    age = age_on(row["_生日_dt"], ref)
    h_target = 8.0 if (age is not None and age >= 80) else 7.0
    l_target = ldl_target(row, sheet_name)
    u_target = 30.0

    tested = {}
    bad = {}
    years = []

    for it in items:
        dt = row.get(f"_{it}_dt", pd.NaT)
        val = row.get(f"_{it}_num", pd.NA)

        is_y = (not pd.isna(dt)) and (dt.year == base_year)
        tested[it] = is_y

        if not pd.isna(dt):
            years.append(int(dt.year))

        if is_y and not pd.isna(val):
            if it == "HbA1c":
                bad[it] = val >= h_target
            elif it == "LDL":
                bad[it] = val >= l_target
            elif it == "UACR":
                bad[it] = val >= u_target
            else:
                bad[it] = False
        else:
            bad[it] = False

    all_tested = all(tested.values())
    any_missing = not all_tested
    all_missing = not any(tested.values())
    any_bad = any(bad.values())
    last_year = max(years) if years else None

    if all_tested and (not any_bad):
        return 0
    if any_missing and any_bad:
        return 3
    if any_missing and (not all_missing):
        return 2
    if all_tested and any_bad:
        return 1

    if all_missing:
        any_prev = False
        for it in items:
            dt = row.get(f"_{it}_dt", pd.NaT)
            if not pd.isna(dt) and dt.year == (base_year - 1):
                any_prev = True
                break
        if any_prev:
            return 4

        if last_year is None:
            return 5
        if (base_year - last_year) >= 2:
            return 5

        return 2

    return 2


def remark_text(code: int, x_items: int, base_year: int) -> str:
    if code == 0:
        return f"0.{base_year}年{x_items}項前測控制良好，符合收案標準,建議收為會員。（{x_items}是檢查的數量）"
    if code == 1:
        return f"1.{base_year}年已受檢未達控制良好"
    if code == 2:
        return f"2.{base_year}年漏檢,請安排回診"
    if code == 3:
        return f"3.{base_year}年漏檢+未達控制良好,回診複查"
    if code == 4:
        return f"4.{base_year-1}已檢：提醒{base_year}年度回診"
    if code == 5:
        return "5.逾兩年未檢：暫緩收案"
    return ""


# ------------------------- build df from A+B -------------------------
def build_master_df(path_a: Path, path_b: Path) -> Tuple[pd.DataFrame, pd.DataFrame]:
    # A
    A = pd.read_excel(path_a, dtype=str)
    A.columns = A.columns.astype(str).str.strip()
    col_id_a = pick_col_exact(list(A.columns), ["ID", "Id", "id"]) or pick_col_contains(list(A.columns), ["id"])
    if not col_id_a:
        raise ValueError("A 檔找不到欄位：ID")
    A = A.rename(columns={col_id_a: "ID"})
    A["ID"] = norm_id(A["ID"])

    # B
    b_sheet, b_header = find_sheet_and_header_for_b(path_b)
    B = pd.read_excel(path_b, sheet_name=b_sheet, header=b_header, dtype=str)
    B.columns = B.columns.astype(str).str.strip()
    cols = list(B.columns)

    col_id_b = pick_col_exact(cols, ["ID", "Id", "id"]) or pick_col_contains(cols, ["id"])
    col_name = pick_col_exact(cols, ["欄1", "姓名", "name"]) or pick_col_contains(cols, ["欄1", "姓名"])
    col_bday = pick_col_exact(cols, ["BIRTHDAY", "生日", "出生日期"]) or pick_col_contains(cols, ["birthday", "生日", "出生"])
    col_dis = pick_col_exact(cols, ["疾病樣態"]) or pick_col_contains(cols, ["疾病樣態"])
    col_ascvd = pick_col_exact(cols, ["ASCVD", "ascvd"]) or pick_col_contains(cols, ["ascvd"])

    missing = []
    if not col_id_b: missing.append("ID")
    if not col_name: missing.append("姓名(欄1)")
    if not col_bday: missing.append("BIRTHDAY(生日)")
    if not col_dis: missing.append("疾病樣態")
    if not col_ascvd: missing.append("ASCVD")

    if missing:
        raise ValueError(
            "B 檔找不到欄位：" + ", ".join(missing) +
            f"\n偵測到：分頁={b_sheet}，header 第{b_header+1}列\n\n"
            "B 欄位列表：\n" + "\n".join([str(c) for c in cols])
        )

    B = B.rename(columns={col_id_b: "ID", col_name: "姓名", col_bday: "生日", col_dis: "疾病樣態", col_ascvd: "ASCVD"})
    B["ID"] = norm_id(B["ID"])
    B_sub = B[["ID", "姓名", "生日", "疾病樣態", "ASCVD"]].copy()

    # merge
    M = A.merge(B_sub, on="ID", how="left", indicator=True)

    # A columns (fuzzy)
    mcols = list(M.columns)
    col_h_val = pick_col_exact(mcols, ["最近一次HbA1c檢查結果(%)", "最近一次HbA1c檢查結果 (%)"]) or pick_col_contains(mcols, ["最近一次hba1c檢查結果"])
    col_h_dt  = pick_col_exact(mcols, ["最近一次HbA1c檢查日期"]) or pick_col_contains(mcols, ["最近一次hba1c檢查日期"])

    col_l_val = pick_col_exact(mcols, ["最近一次LDL檢查結果(mg/dL)", "最近一次LDL檢查結果 (mg/dL)"]) or pick_col_contains(mcols, ["最近一次ldl檢查結果"])
    col_l_dt  = pick_col_exact(mcols, ["最近一次LDL檢查日期"]) or pick_col_contains(mcols, ["最近一次ldl檢查日期"])

    col_u_val = pick_col_exact(mcols, ["最近一次UACR檢查結果(mg/gm)", "最近一次UACR檢查結果 (mg/gm)"]) or pick_col_contains(mcols, ["最近一次uacr檢查結果"])
    col_u_dt  = pick_col_exact(mcols, ["最近一次UACR檢查日期"]) or pick_col_contains(mcols, ["最近一次uacr檢查日期"])

    df = pd.DataFrame({
        "ID": M["ID"],
        "疾病樣態": pd.to_numeric(M["疾病樣態"], errors="coerce"),
        "ASCVD": M["ASCVD"],
        "姓名": M["姓名"],
        "生日": M["生日"],
        "最近一次HbA1c檢查結果(%)": M[col_h_val] if col_h_val else pd.NA,
        "最近一次HbA1c檢查日期": M[col_h_dt] if col_h_dt else pd.NA,
        "最近一次LDL檢查結果(mg/dL)": M[col_l_val] if col_l_val else pd.NA,
        "最近一次LDL檢查日期": M[col_l_dt] if col_l_dt else pd.NA,
        "最近一次UACR檢查結果(mg/gm)": M[col_u_val] if col_u_val else pd.NA,
        "最近一次UACR檢查日期": M[col_u_dt] if col_u_dt else pd.NA,
    })

    df["_merge_flag"] = M["_merge"]  # both / left_only

    # helpers
    df["_生日_dt"] = to_dt(df["生日"])
    df["_HbA1c_dt"] = to_dt(df["最近一次HbA1c檢查日期"])
    df["_LDL_dt"] = to_dt(df["最近一次LDL檢查日期"])
    df["_UACR_dt"] = to_dt(df["最近一次UACR檢查日期"])

    df["_HbA1c_num"] = pd.to_numeric(df["最近一次HbA1c檢查結果(%)"], errors="coerce")
    df["_LDL_num"] = pd.to_numeric(df["最近一次LDL檢查結果(mg/dL)"], errors="coerce")
    df["_UACR_num"] = pd.to_numeric(df["最近一次UACR檢查結果(mg/gm)"], errors="coerce")

    # output date format
    df["生日"] = df["_生日_dt"].dt.strftime("%Y-%m-%d")
    df["最近一次HbA1c檢查日期"] = df["_HbA1c_dt"].dt.strftime("%Y-%m-%d")
    df["最近一次LDL檢查日期"] = df["_LDL_dt"].dt.strftime("%Y-%m-%d")
    df["最近一次UACR檢查日期"] = df["_UACR_dt"].dt.strftime("%Y-%m-%d")

    df["ASCVD"] = df["ASCVD"].apply(ascvd_norm)

    # issues
    issues = []
    issues_merge = df[df["_merge_flag"] != "both"].copy()
    if not issues_merge.empty:
        issues_merge["問題"] = "B檔找不到ID(merge left_only)"
        issues.append(issues_merge)

    issues_dis = df[df["疾病樣態"].isna()].copy()
    if not issues_dis.empty:
        issues_dis["問題"] = "疾病樣態缺失/非數字"
        issues.append(issues_dis)

    issues_df = pd.concat(issues, ignore_index=True) if issues else pd.DataFrame()
    return df, issues_df


# ------------------------- split to 7 sheets + remarks + sort -------------------------
def split_into_sheets(df: pd.DataFrame, base_year: int) -> Tuple[Dict[str, pd.DataFrame], Dict[str, Dict[str, int]]]:
    is_dm = df["疾病樣態"] == 1
    is_ckd = df["疾病樣態"] == 2
    is_dkd = df["疾病樣態"] == 3
    has_ascvd = df["ASCVD"].isin(["a", "b"])

    sheets = {
        "DM": df[is_dm & (~has_ascvd)].copy(),
        "CKD": df[is_ckd & (~has_ascvd)].copy(),
        "DKD": df[is_dkd & (~has_ascvd)].copy(),
        "DM+ASCVD": df[is_dm & has_ascvd].copy(),
        "CKD+ASCVD": df[is_ckd & has_ascvd].copy(),
        "DKD+ASCVD": df[is_dkd & has_ascvd].copy(),
        "ASCVD": df[has_ascvd & (~(is_dm | is_ckd | is_dkd))].copy(),
    }

    export_cols = [
        "ID", "疾病樣態", "ASCVD", "姓名", "生日",
        "最近一次HbA1c檢查結果(%)", "最近一次HbA1c檢查日期",
        "最近一次LDL檢查結果(mg/dL)", "最近一次LDL檢查日期",
        "最近一次UACR檢查結果(mg/gm)", "最近一次UACR檢查日期",
        "備註",
    ]

    analysis_counts: Dict[str, Dict[str, int]] = {}

    for name, d in sheets.items():
        if len(d) > 0:
            x = len(required_items(name))
            codes = d.apply(lambda r: remark_code(r, name, base_year), axis=1)
            d["備註"] = [remark_text(int(c), x, base_year) for c in codes.tolist()]
            d = d.sort_values("ID", ascending=True, kind="mergesort")
        else:
            d = pd.DataFrame(columns=export_cols)

        if len(d) > 0:
            vc = d["備註"].astype(str).str.extract(r"^(\d)\.")[0].value_counts()
        else:
            vc = pd.Series(dtype="int64")
        analysis_counts[name] = {str(i): int(vc.get(str(i), 0)) for i in range(6)}

        sheets[name] = d[export_cols].copy()

    return sheets, analysis_counts


# ------------------------- analysis sheet (no template) -------------------------
def build_analysis_table_rows(analysis_counts: Dict[str, Dict[str, int]]) -> List[List]:
    order = ["DM", "CKD", "ASCVD", "DKD", "CKD+ASCVD", "DM+ASCVD", "DKD+ASCVD"]
    rows = []
    for g in order:
        c = analysis_counts.get(g, {str(i): 0 for i in range(6)})
        total = sum(int(c.get(str(i), 0)) for i in range(6))
        rows.append([g, int(c["0"]), int(c["1"]), int(c["2"]), int(c["3"]), int(c["4"]), int(c["5"]), total])
    return rows


def autosize_columns(ws, max_width: int = 45):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def write_output_no_template(out_path: Path, sheets: Dict[str, pd.DataFrame], analysis_counts: Dict[str, Dict[str, int]]):
    wb = Workbook()
    wb.remove(wb.active)

    ws_a = wb.create_sheet("分析表", 0)
    ws_a.append(["群組", "0", "1", "2", "3", "4", "5", "總計"])
    for row in build_analysis_table_rows(analysis_counts):
        ws_a.append(row)
    autosize_columns(ws_a)

    order = ["DM", "CKD", "ASCVD", "DKD", "CKD+ASCVD", "DM+ASCVD", "DKD+ASCVD"]
    for name in order:
        ws = wb.create_sheet(title=name)
        df = sheets.get(name)

        if df is None or df.empty:
            cols = sheets[order[0]].columns.tolist() if order[0] in sheets else ["ID"]
            ws.append(cols)
        else:
            ws.append(list(df.columns))
            for row in df.itertuples(index=False, name=None):
                ws.append(list(row))

        autosize_columns(ws)

    wb.save(out_path)


# ------------------------- GUI main (ONLY ONE) -------------------------
def main():
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()

    try:
        messagebox.showinfo("步驟1", "請選擇 A 檔：個案健康管理報表")
        path_a = filedialog.askopenfilename(
            title="選擇 A 檔（個案健康管理報表）",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if not path_a:
            return

        messagebox.showinfo("步驟2", "請選擇 B 檔：收案追蹤報表")
        path_b = filedialog.askopenfilename(
            title="選擇 B 檔（收案追蹤報表）",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if not path_b:
            return

        messagebox.showinfo("步驟3", "請選擇輸出檔案位置")
        path_out = filedialog.asksaveasfilename(
            title="儲存分流報表",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not path_out:
            return

        df, issues_df = build_master_df(Path(path_a), Path(path_b))
        sheets, analysis_counts = split_into_sheets(df, BASE_YEAR)
        write_output_no_template(Path(path_out), sheets, analysis_counts)

        msg = f"報表完成（年度固定 {BASE_YEAR}）：\n{path_out}"
        if issues_df is not None and not issues_df.empty:
            issues_path = Path(path_out).with_suffix("").as_posix() + "_issues.csv"
            issues_df.to_csv(issues_path, index=False, encoding="utf-8-sig")
            msg += f"\n\n另產生 issues：\n{issues_path}"

        messagebox.showinfo("完成", msg)

    except Exception as e:
        messagebox.showerror("錯誤", str(e))


if __name__ == "__main__":
    main()
