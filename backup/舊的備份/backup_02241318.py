# -*- coding: utf-8 -*-
"""
選會員產檔工具（單檔版、已結構化）

功能摘要
- 讀取原始 Excel（包含：會員名單、ascvd、HealthCase、五大篩檢等工作表）
- 以「選會員樣板0223.xlsx」為模板，只保留「會員指標」分頁並寫入資料
- 產出：選會員MMDD_HHMM.xlsx（儲存在原始檔同資料夾），並自動開啟

本檔案結構
1) 小工具：文字/表頭/欄位偵測、日期/數字解析
2) 規則工具：疾病樣態/ASCVD 分類、備註 AV 文字、AU/AV/AW 回診追蹤文字
3) KPI/統計：HbA1c/LDL 百分位（AY/AZ/BA、BC/BD/BE）
4) 主流程 process_excel：欄位定位 → 清資料 → 寫入/回填 → 計算欄位 → KPI → 儲存
"""

import os
import re
import sys
import datetime
import subprocess
from typing import Optional, Dict, Any, Tuple, List, Callable

import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.utils.exceptions import InvalidFileException


# =========================
# 固定設定
# =========================
TEMPLATE_NAME = "選會員樣板0223.xlsx"   # 模板檔名：放在本 .py 同資料夾
SHEET_TARGET = "會員指標"
CHECK = "✓"


# =========================
# 小工具：跨平台開檔
# =========================
def open_file_cross_platform(path: str) -> None:
    """嘗試用系統預設程式開啟檔案。失敗則忽略。"""
    try:
        if sys.platform.startswith("darwin"):
            subprocess.call(("open", path))
        elif os.name == "nt":
            os.startfile(path)  # type: ignore[attr-defined]
        else:
            subprocess.call(("xdg-open", path))
    except Exception:
        pass


# =========================
# 小工具：文字/表頭/欄位偵測
# =========================
def normalize_text(v: Any) -> str:
    """一般文字正規化：去除 tab、全形空白、前後空白。"""
    if v is None:
        return ""
    return str(v).replace("\t", "").replace("　", "").strip()


def normalize_header(v: Any) -> str:
    """表頭正規化：移除空白+換行，方便模糊比對。"""
    s = normalize_text(v)
    return s.replace(" ", "").replace("\n", "")


def build_header_map(sheet, header_row: int) -> Dict[str, int]:
    """完全匹配用：回傳 {表頭文字: 欄號}。"""
    m: Dict[str, int] = {}
    for c in range(1, sheet.max_column + 1):
        k = normalize_text(sheet.cell(header_row, c).value)
        if k:
            m[k] = c
    return m


def find_column_exact(hmap: Dict[str, int], aliases) -> Optional[int]:
    """完全匹配：別名任一存在即回傳該欄。"""
    for a in aliases:
        if a in hmap:
            return hmap[a]
    return None


def get_merged_header_text(ws, header_row: int, c: int) -> str:
    """
    支援多層表頭：
    - header_row-1：大類（可能合併）
    - header_row：欄名/說明
    合併兩列文字後回傳，用於模糊比對。
    """
    upper = normalize_header(ws.cell(header_row - 1, c).value) if header_row > 1 else ""
    lower = normalize_header(ws.cell(header_row, c).value)
    return (upper + " " + lower).strip()


def find_col_by_keywords(ws, header_row: int, keywords: List[str]) -> Optional[int]:
    """
    模糊比對：表頭(去空白/去換行) 必須同時包含 keywords 的所有字。
    支援多層表頭：用 (header_row-1 + header_row) 合併字串來找。
    """
    keys = [k.replace(" ", "").replace("\n", "") for k in keywords if k]
    for c in range(1, ws.max_column + 1):
        h = get_merged_header_text(ws, header_row, c).replace(" ", "")
        if not h:
            continue
        if all(k in h for k in keys):
            return c
    return None


def find_col_by_keywords_any_row(ws, max_row: int, keywords: List[str]) -> Optional[int]:
    """
    跨多列找欄位：在 1..max_row 的同一欄中，
    只要任一儲存格文字整體拼接後同時包含 keywords 全部關鍵字，就回傳該欄。
    （用於 AU/AV/AW 這種「表頭可能在多列」的模板）
    """
    keys = [k.replace(" ", "").replace("\n", "") for k in keywords if k]
    for c in range(1, ws.max_column + 1):
        blob = ""
        for r in range(1, max_row + 1):
            v = normalize_header(ws.cell(r, c).value)
            if v:
                blob += v
        if not blob:
            continue
        if all(k in blob for k in keys):
            return c
    return None


def find_header_row_contains(sheet, must_have, search_rows: int = 250) -> Optional[int]:
    """
    找到同一列同時包含 must_have 幾個表頭（模糊等於：去空白/去換行後比對）。
    """
    must = [m.replace(" ", "").replace("\n", "") for m in must_have]
    for r in range(1, min(search_rows, sheet.max_row) + 1):
        found = set()
        for c in range(1, sheet.max_column + 1):
            v = normalize_header(sheet.cell(r, c).value)
            if not v:
                continue
            for m in must:
                if v == m:
                    found.add(m)
        if all(m in found for m in must):
            return r
    return None


# =========================
# 小工具：日期/數字解析
# =========================
def parse_date(value: Any) -> Optional[datetime.date]:
    """支援 datetime/date、yyyy-mm-dd、yyyy/mm/dd、民國 yyyMMdd、民國 yyy/mm/dd。"""
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value

    s = str(value).strip()
    if s == "" or s in ("-", "—", "–"):
        return None

    # yyyy/mm/dd or yyyy-mm-dd
    m = re.match(r"^(\d{4})[/-](\d{1,2})[/-](\d{1,2})$", s)
    if m:
        y, mo, d = map(int, m.groups())
        return datetime.date(y, mo, d)

    # ROC yyyMMdd
    m = re.match(r"^(\d{2,3})(\d{2})(\d{2})$", s)
    if m:
        roc, mo, d = map(int, m.groups())
        return datetime.date(roc + 1911, mo, d)

    # ROC yyy/mm/dd
    m = re.match(r"^(\d{2,3})[/-](\d{1,2})[/-](\d{1,2})$", s)
    if m:
        roc, mo, d = map(int, m.groups())
        return datetime.date(roc + 1911, mo, d)

    return None


def parse_float(value: Any) -> Optional[float]:
    """將儲存格內容轉 float（允許字串數字）。空白/破折號回 None。"""
    if value is None:
        return None
    s = str(value).strip()
    if s == "" or s in ("-", "—", "–"):
        return None
    try:
        return float(s)
    except Exception:
        return None


# =========================
# 小工具：格式/基本欄位推導
# =========================
def infer_gender_from_id(id_value: Any) -> str:
    """從身分證第 2 碼推斷性別（1/8 男，2/9 女），無法判斷回空字串。"""
    if not id_value:
        return ""
    s = str(id_value).strip().upper()
    if len(s) < 2:
        return ""
    g = s[1]
    if g in ("1", "8"):
        return "男"
    if g in ("2", "9"):
        return "女"
    return ""


def clean_spaces(value: Any) -> str:
    """移除所有空白（含多個空白/換行）。"""
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value))


def calc_age(bday: Optional[datetime.date], ref: datetime.date) -> int:
    """計算整歲；無生日回 -1。"""
    if not isinstance(bday, datetime.date):
        return -1
    y = ref.year - bday.year
    if (ref.month, ref.day) < (bday.month, bday.day):
        y -= 1
    return y


def add_full_grid(ws) -> None:
    """整張工作表套細框線（用於模板輸出一致）。"""
    thin = Side(style="thin")
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).border = grid


# =========================
# 規則：疾病樣態/ASCVD 分類
# =========================
def parse_dmk_to_code(v: Any) -> Optional[int]:
    """
    疾病樣態碼（模板欄位 DM/CKD/DKD）：
    1:DM, 2:CKD, 3:DM+CKD(DKD), 其他：視為無（可能是 4 或空）
    """
    s = clean_spaces(v)
    if s == "":
        return None
    if s.isdigit():
        return int(s)
    su = s.upper()
    if su == "DM":
        return 1
    if su == "CKD":
        return 2
    if su in ("DKD", "DM+CKD", "DMCKD"):
        return 3
    return None


def classify_ascvd(v: Any) -> Optional[str]:
    """
    ASCVD：
    - a/b 代表有 ASCVD（含中文：極高/非常高）
    - 空白或 0 視為無（None）
    """
    if v is None:
        return None
    s = clean_spaces(v).lower()
    if s == "" or s == "0":
        return None
    if s in ("a", "1") or "極高" in s:
        return "a"
    if s in ("b", "2") or "非常高" in s:
        return "b"
    if "a" in s:
        return "a"
    if "b" in s:
        return "b"
    return None


def disease_group_text(e_code: Optional[int], f_ascvd: Optional[str]) -> Optional[str]:
    """把疾病樣態碼 + ASCVD 分類組合成顯示文字。"""
    has_dm = (e_code == 1) or (e_code == 3)
    has_ckd = (e_code == 2) or (e_code == 3)
    has_ascvd = f_ascvd in ("a", "b")

    if has_dm and (not has_ckd) and (not has_ascvd):
        return "DM"
    if has_ckd and (not has_dm) and (not has_ascvd):
        return "CKD"
    if has_dm and has_ckd and (not has_ascvd):
        return "DM+CKD（DKD）"
    if (not has_dm) and (not has_ckd) and has_ascvd:
        return "ASCVD"
    if has_ckd and (not has_dm) and has_ascvd:
        return "CKD+ASCVD"
    if has_dm and (not has_ckd) and has_ascvd:
        return "DM+ASCVD"
    if has_dm and has_ckd and has_ascvd:
        return "DM+CKD+ASCVD"
    return None


# =========================
# 規則：AV 備註（不靠公式）
# =========================
def add_months(d: datetime.date, months: int) -> datetime.date:
    """date + N months（EDATE 等效）。"""
    y = d.year + (d.month - 1 + months) // 12
    m = (d.month - 1 + months) % 12 + 1
    is_leap = (y % 4 == 0 and (y % 100 != 0 or y % 400 == 0))
    mdays = [31, 29 if is_leap else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][m - 1]
    day = min(d.day, mdays)
    return datetime.date(y, m, day)


def need_bc_hep(age: int) -> bool:
    return 45 <= age <= 80


def need_fit(age: int) -> bool:
    return 45 <= age <= 75


def need_pap(age: int, sex: str) -> bool:
    if sex != "女":
        return False
    return age >= 25


def need_adult(age: int) -> bool:
    return age >= 30


def need_flu(age: int) -> bool:
    return age >= 65


def build_av_note(
    age: int,
    sex: str,
    hep_dt: Optional[datetime.date],
    fit_dt: Optional[datetime.date],
    pap_dt: Optional[datetime.date],
    adult_dt: Optional[datetime.date],
    flu_dt: Optional[datetime.date],
    today: datetime.date
) -> str:
    """依規則輸出備註（每行一項）。"""
    msgs: List[str] = []

    if need_bc_hep(age) and hep_dt is None:
        msgs.append("今年需檢測BC肝")

    if need_fit(age):
        if fit_dt is None or add_months(fit_dt, 24) <= today:
            msgs.append("今年需檢測糞便")

    if sex == "女" and need_pap(age, sex):
        if 25 <= age <= 29:
            if pap_dt is None or add_months(pap_dt, 36) <= today:
                msgs.append("今年需檢測子抹")
        elif age >= 30:
            if pap_dt is None or add_months(pap_dt, 12) <= today:
                msgs.append("今年需檢測子抹")

    if need_adult(age):
        if 30 <= age <= 39:
            if adult_dt is None or add_months(adult_dt, 60) <= today:
                msgs.append("今年需檢測成健")
        elif 40 <= age <= 64:
            if adult_dt is None or add_months(adult_dt, 36) <= today:
                msgs.append("今年需檢測成健")
        elif age >= 65:
            if adult_dt is None or add_months(adult_dt, 12) <= today:
                msgs.append("今年需檢測成健")

    if need_flu(age):
        if flu_dt is None or flu_dt.year < today.year:
            msgs.append("今年需檢測老流")

    return "\n".join(msgs).strip()


# =========================
# 規則：AU 分數/明細（保留原邏輯）
# =========================
SCORE_YEAR_SCREEN_2026 = 2026


def score_hba(hba_val: Any, hba_dt: Optional[datetime.date]) -> int:
    if not isinstance(hba_dt, datetime.date):
        return 0
    if hba_dt.year == 2026:
        return 5
    if hba_dt.year == 2025:
        if isinstance(hba_val, (int, float)) and float(hba_val) <= 7.0:
            return 5
    return 0


def score_ldl(ldl_val: Any, ldl_dt: Optional[datetime.date]) -> int:
    if not isinstance(ldl_dt, datetime.date):
        return 0
    if ldl_dt.year == 2026:
        return 5
    if ldl_dt.year == 2025:
        if isinstance(ldl_val, (int, float)) and float(ldl_val) <= 120.0:
            return 5
    return 0


def score_screening_year_only(dt: Optional[datetime.date], year: int, pts: int) -> int:
    if isinstance(dt, datetime.date) and dt.year == year:
        return pts
    return 0


def au_score_with_breakdown(
    *,
    e_code: Optional[int],
    f_ascvd: Optional[str],
    hba_val: Any,
    hba_dt: Optional[datetime.date],
    ldl_val: Any,
    ldl_dt: Optional[datetime.date],
    adult_dt: Optional[datetime.date],
    pap_dt: Optional[datetime.date],
    flu_dt: Optional[datetime.date],
    fit_dt: Optional[datetime.date],
    hep_dt: Optional[datetime.date],
    age: int,
    sex: str
) -> Tuple[int, Optional[str]]:
    has_dm = (e_code == 1) or (e_code == 3)
    has_ckd = (e_code == 2) or (e_code == 3)
    has_ascvd = f_ascvd in ("a", "b")

    total = 0
    parts: List[str] = []

    if has_dm:
        pts = score_hba(hba_val, hba_dt)
        if pts:
            total += pts
            parts.append(f"HbA1c {pts}分")

    if has_dm or has_ckd or has_ascvd:
        pts = score_ldl(ldl_val, ldl_dt)
        if pts:
            total += pts
            parts.append(f"LDL {pts}分")

    if age >= 0:
        if need_adult(age):
            pts = score_screening_year_only(adult_dt, SCORE_YEAR_SCREEN_2026, 6)
            if pts:
                total += pts
                parts.append(f"成健 {pts}分")

        if need_pap(age, sex):
            pts = score_screening_year_only(pap_dt, SCORE_YEAR_SCREEN_2026, 6)
            if pts:
                total += pts
                parts.append(f"子抹 {pts}分")

        if need_flu(age):
            pts = score_screening_year_only(flu_dt, SCORE_YEAR_SCREEN_2026, 4)
            if pts:
                total += pts
                parts.append(f"老流 {pts}分")

        if need_fit(age):
            pts = score_screening_year_only(fit_dt, SCORE_YEAR_SCREEN_2026, 6)
            if pts:
                total += pts
                parts.append(f"糞便 {pts}分")

        if need_bc_hep(age):
            pts = score_screening_year_only(hep_dt, SCORE_YEAR_SCREEN_2026, 6)
            if pts:
                total += pts
                parts.append(f"BC肝 {pts}分")

    breakdown = "\n".join(parts).strip() if parts else None
    return total, breakdown


# =========================
# 規則：AU（第一次提醒 28 天）— HbA1c/LDL（保留你定義）
# =========================
def build_au_28_note(
    *,
    e_code: Optional[int],
    f_ascvd: Optional[str],
    hba_dt: Optional[datetime.date],
    ldl_dt: Optional[datetime.date],
    today: datetime.date
) -> str:
    """
    AU 欄位（第一次提醒 28 天）規則：
    - 若不需檢查 -> 非疾病不需回診
    - 若 <2025 或 無日期 -> 超過2年未檢查
    - 若 2025/2026：
      dt+28 <= today -> 可立刻通知回診
      else           -> yyyy-mm-dd需回診（顯示 dt+28）
    """
    e = e_code if e_code in (1, 2, 3, 4) else 4
    has_ascvd = f_ascvd in ("a", "b")

    need_hba = e in (1, 3)
    need_ldl = (e in (1, 2, 3)) or (e == 4 and has_ascvd)

    if not need_hba and not need_ldl:
        return "非疾病不需回診"

    def is_2025_or_2026(dt: Optional[datetime.date]) -> bool:
        return isinstance(dt, datetime.date) and dt.year in (2025, 2026)

    eligible = is_2025_or_2026(hba_dt) or is_2025_or_2026(ldl_dt)
    if not eligible:
        lines: List[str] = []
        if need_hba:
            lines.append("HbA1c:超過2年未檢查")
        if need_ldl:
            lines.append("LDL:超過2年未檢查")
        return "\n".join(lines).strip()

    lines2: List[str] = []

    if need_hba:
        if not is_2025_or_2026(hba_dt):
            lines2.append("HbA1c:超過2年未檢查")
        else:
            due = hba_dt + datetime.timedelta(days=28)  # type: ignore[operator]
            if due <= today:
                lines2.append("HbA1c:可立刻通知回診")
            else:
                lines2.append(f"HbA1c:{due.strftime('%Y-%m-%d')}需回診")

    if need_ldl:
        if not is_2025_or_2026(ldl_dt):
            lines2.append("LDL:超過2年未檢查")
        else:
            due = ldl_dt + datetime.timedelta(days=28)  # type: ignore[operator]
            if due <= today:
                lines2.append("LDL:可立刻通知回診")
            else:
                lines2.append(f"LDL:{due.strftime('%Y-%m-%d')}需回診")

    return "\n".join(lines2).strip()


# =========================
# 規則：AV/AW 追蹤（由 AU 文字延後）
# =========================
DATE_RE = re.compile(r"(\d{4}-\d{2}-\d{2})")


def au_should_skip_followup(au_txt: str) -> bool:
    """AU 不須通知 or 超過兩年未檢查 -> AV/AW 不填。"""
    if not au_txt:
        return True
    s = au_txt.replace(" ", "")
    if "超過2年未檢查" in s or "超過兩年未檢查" in s:
        return True
    if "非疾病" in s:
        return True
    return False


def build_followup_text_from_au(au_txt: str, today: datetime.date, offset_days: int) -> Optional[str]:
    """
    將 AU 內容轉成追蹤欄位文字（AV/AW）：
    - 每一行獨立處理
    - 行內若有 yyyy-mm-dd -> 該日期 + offset_days
    - 行內若是「可立刻通知回診」 -> today + offset_days
    - 只處理 HbA1c/LDL 行
    """
    if not au_txt:
        return None

    lines = [ln.strip() for ln in str(au_txt).splitlines() if ln.strip()]
    out: List[str] = []

    for ln in lines:
        if not (ln.startswith("HbA1c:") or ln.startswith("LDL:")):
            continue

        prefix = "HbA1c:" if ln.startswith("HbA1c:") else "LDL:"

        if "可立刻通知" in ln:
            due = today + datetime.timedelta(days=offset_days)
            out.append(f"{prefix}{due.strftime('%Y-%m-%d')}需回診")
            continue

        m = DATE_RE.search(ln)
        if m:
            try:
                base = datetime.datetime.strptime(m.group(1), "%Y-%m-%d").date()
                due = base + datetime.timedelta(days=offset_days)
                out.append(f"{prefix}{due.strftime('%Y-%m-%d')}需回診")
            except Exception:
                pass

    txt = "\n".join(out).strip()
    return txt if txt else None


# =========================
# KPI/統計：通用百分位計算器
# =========================
def is_ascvd_zero_like(v: Any) -> bool:
    """ASCVD 視為 0：None / 空白 / '0' / '0.0'。"""
    if v is None:
        return True
    s = str(v).strip()
    return s == "" or s == "0" or s == "0.0"


def compute_percentiles(
    *,
    ws,
    data_start: int,
    last_row: int,
    col_dmk_code: int,
    col_ascvd: int,
    col_value: int,
    col_date: int,
    valid_dmk: Tuple[int, ...],
    valid_years: Tuple[int, ...],
    thresholds: List[float],
    exclude_rule: Optional[Callable[[int, int, Any], bool]] = None,
) -> Tuple[int, Dict[float, float]]:
    """
    通用「分子/分母百分位」計算器。

    分母條件（通用）：
    - 疾病樣態碼在 valid_dmk
    - 日期欄為有效日期，且年份在 valid_years
    - 數值欄可轉成數字

    分子條件：
    - 在符合分母前提下：數值 <= threshold

    exclude_rule（可選）：
    - callable(row, e_code, ascvd_val) -> bool
    - True 表示排除此列（不進分母）

    回傳：
    - denom: 分母
    - ratios: {threshold: ratio(0~1)}
    """
    denom = 0
    numer_map: Dict[float, int] = {t: 0 for t in thresholds}

    for r in range(data_start, last_row + 1):
        # 疾病樣態碼
        e_raw = ws.cell(r, col_dmk_code).value
        try:
            e = int(e_raw) if e_raw is not None and str(e_raw).strip() != "" else None
        except Exception:
            e = parse_dmk_to_code(e_raw)

        if e not in valid_dmk:
            continue

        ascvd_val = ws.cell(r, col_ascvd).value

        if exclude_rule and exclude_rule(r, int(e), ascvd_val):  # type: ignore[arg-type]
            continue

        dt = parse_date(ws.cell(r, col_date).value) if col_date else None
        if not (isinstance(dt, datetime.date) and dt.year in valid_years):
            continue

        val = parse_float(ws.cell(r, col_value).value) if col_value else None
        if val is None:
            continue

        denom += 1
        for t in thresholds:
            if val <= t:
                numer_map[t] += 1

    ratios = {t: (numer_map[t] / denom if denom > 0 else 0) for t in thresholds}
    return denom, ratios


def write_percent_cells(ws, mapping: Dict[str, float]) -> None:
    """將 {儲存格: ratio(0~1)} 寫入，並設定為 0.00% 格式。"""
    for addr, ratio in mapping.items():
        ws[addr].value = ratio
        ws[addr].number_format = "0.00%"


# =========================
# 主流程
# =========================
def process_excel(source_xlsx_path: str, template_xlsx_path: str) -> str:
    wb_src = openpyxl.load_workbook(source_xlsx_path, data_only=True)

    need_sheets = ["會員名單", "ascvd", "HealthCase", "成人健檢", "子宮抹片", "老人流感", "糞便潛血", "肝炎篩檢"]
    for sn in need_sheets:
        if sn not in wb_src.sheetnames:
            raise ValueError(f"原始檔缺少工作表：{sn}")

    sh_member = wb_src["會員名單"]
    sh_ascvd = wb_src["ascvd"]
    sh_health = wb_src["HealthCase"]
    sh_adult = wb_src["成人健檢"]
    sh_pap = wb_src["子宮抹片"]
    sh_flu = wb_src["老人流感"]
    sh_fit = wb_src["糞便潛血"]
    sh_hep = wb_src["肝炎篩檢"]

    wb_tpl = openpyxl.load_workbook(template_xlsx_path)
    if SHEET_TARGET not in wb_tpl.sheetnames:
        raise ValueError("模板檔缺少工作表：會員指標")

    # 僅保留目標分頁
    for sn in list(wb_tpl.sheetnames):
        if sn != SHEET_TARGET:
            del wb_tpl[sn]
    ws = wb_tpl[SHEET_TARGET]

    # 你的模板：第9列開始是資料列
    DATA_START = 9

    # --- 找模板表頭列（主要表頭在第5列） ---
    header_row = find_header_row_contains(ws, ["診所名稱或機構代碼", "姓名", "身份證號碼"], 250) or 5

    # --- 用「多層表頭」模糊比對找欄位 ---
    col_clinic = find_col_by_keywords(ws, header_row, ["診所名稱或機構代碼"])
    col_name = find_col_by_keywords(ws, header_row, ["姓名"])
    col_id = find_col_by_keywords(ws, header_row, ["身份證號碼"])
    col_bday = find_col_by_keywords(ws, header_row, ["生日"])
    col_age = find_col_by_keywords(ws, header_row, ["年齡"])
    col_tel = find_col_by_keywords(ws, header_row, ["電話"])
    col_cnt = find_col_by_keywords(ws, header_row, ["次數"])
    col_sex = find_col_by_keywords(ws, header_row, ["性別"])
    col_abc = find_col_by_keywords(ws, header_row, ["A/B/C"])

    col_dmk_code = find_col_by_keywords(ws, header_row, ["DM/CKD/DKD"])
    col_ascvd = find_col_by_keywords(ws, header_row, ["ASCVD"])

    col_adult = find_col_by_keywords(ws, header_row, ["成人", "保健"])
    col_pap = find_col_by_keywords(ws, header_row, ["子宮", "抹片"])
    col_flu = find_col_by_keywords(ws, header_row, ["流感"]) or find_col_by_keywords(ws, header_row, ["流感"])
    col_fit = find_col_by_keywords(ws, header_row, ["糞便", "潛血"])
    col_hep = find_col_by_keywords(ws, header_row, ["肝炎"])

    col_hba = find_col_by_keywords(ws, header_row, ["HbA1c"])
    col_ldl = find_col_by_keywords(ws, header_row, ["LDL"])
    col_uacr = find_col_by_keywords(ws, header_row, ["UACR"])
    col_hba_dt = find_col_by_keywords(ws, header_row, ["HbA1c", "日期"])
    col_ldl_dt = find_col_by_keywords(ws, header_row, ["LDL", "日期"])
    col_uacr_dt = find_col_by_keywords(ws, header_row, ["UACR", "日期"])

    col_disease_text = (
        find_col_by_keywords(ws, header_row, ["DM/CKD/DKD/ASCVD"])
        or find_col_by_keywords(ws, header_row, ["疾病樣態", "ASCVD"])
    )
    col_score = find_col_by_keywords(ws, header_row, ["分數"])
    col_note = find_col_by_keywords(ws, header_row, ["備註"])

    # AU/AV/AW：跨列掃描 1..8
    max_scan_row = DATA_START - 1
    col_au_remind = (
        find_col_by_keywords_any_row(ws, max_scan_row, ["第一次提醒", "28天"])
        or find_col_by_keywords_any_row(ws, max_scan_row, ["回診追蹤", "28天"])
        or find_col_by_keywords_any_row(ws, max_scan_row, ["28天"])
    )
    col_av_remind = (
        find_col_by_keywords_any_row(ws, max_scan_row, ["第二次提醒", "56天"])
        or find_col_by_keywords_any_row(ws, max_scan_row, ["回診追蹤", "56天"])
    )
    col_aw_remind = (
        find_col_by_keywords_any_row(ws, max_scan_row, ["第三次提醒", "84天"])
        or find_col_by_keywords_any_row(ws, max_scan_row, ["回診追蹤", "84天"])
    )

    # BH：分數明細（保留你原本邏輯：分數欄往右 13 欄）
    col_breakdown = None
    if col_score:
        cand = col_score + 13
        if 1 <= cand <= ws.max_column:
            col_breakdown = cand

    must = [
        ("診所名稱或機構代碼", col_clinic),
        ("姓名", col_name),
        ("身份證號碼", col_id),
        ("生日", col_bday),
        ("電話", col_tel),
        ("A/B/C", col_abc),
        ("DM/CKD/DKD", col_dmk_code),
        ("次數", col_cnt),
        ("ASCVD", col_ascvd),
        ("性別", col_sex),
        ("HbA1c", col_hba),
        ("HbA1c日期", col_hba_dt),
        ("LDL", col_ldl),
        ("LDL日期", col_ldl_dt),
        ("UACR", col_uacr),
        ("UACR日期", col_uacr_dt),
        ("疾病樣態文字欄", col_disease_text),
        ("分數欄", col_score),
        ("備註欄", col_note),
        ("AU(第一次提醒28天)", col_au_remind),
        ("AV(第二次提醒56天)", col_av_remind),
        ("AW(第三次提醒84天)", col_aw_remind),
    ]
    missing = [k for k, v in must if v is None]
    if missing:
        raise ValueError("新模板欄位/欄名找不到：" + "、".join(missing))

    # --- 清資料列（不動表頭/樣式）---
    cols_to_clear = [
        col_clinic, col_name, col_id, col_bday, col_age, col_tel, col_cnt, col_sex, col_abc,
        col_dmk_code, col_ascvd,
        col_adult, col_pap, col_flu, col_fit, col_hep,
        col_hba, col_hba_dt, col_ldl, col_ldl_dt, col_uacr, col_uacr_dt,
        col_disease_text, col_score, col_breakdown, col_note,
        col_au_remind, col_av_remind, col_aw_remind,
    ]
    cols_to_clear = [c for c in cols_to_clear if c]

    for r in range(DATA_START, ws.max_row + 1):
        for c in cols_to_clear:
            ws.cell(r, c).value = None

    # --- 會員名單欄位 ---
    member_header_row = 5
    mmap = build_header_map(sh_member, member_header_row)
    m_name = find_column_exact(mmap, ["會員姓名"])
    m_id = find_column_exact(mmap, ["會員身份証", "會員身份證", "會員身分證"])
    m_bday = find_column_exact(mmap, ["會員生日"])
    m_tel = find_column_exact(mmap, ["電話"])
    m_abc = find_column_exact(mmap, ["會員別"])
    m_dmk = find_column_exact(mmap, ["疾病樣態"])
    m_cnt = find_column_exact(mmap, ["就診次數"])

    if any(v is None for v in [m_name, m_id, m_bday, m_tel, m_abc, m_dmk, m_cnt]):
        raise ValueError("原始檔「會員名單」欄位不完整（會員姓名/會員身分證/會員生日/電話/會員別/疾病樣態/就診次數）")

    clinic_val = normalize_text(sh_member["A1"].value)

    # --- 寫入基本資料，建立 id_to_row / meta ---
    id_to_row: Dict[str, int] = {}
    meta: Dict[int, Dict[str, Any]] = {}
    now_tw = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=8))).date()

    out_r = DATA_START
    for r in range(member_header_row + 1, sh_member.max_row + 1):
        sid = sh_member.cell(r, m_id).value
        name = sh_member.cell(r, m_name).value
        if (sid in (None, "")) and (name in (None, "")):
            continue

        sid_s = str(sid).strip()
        bday = parse_date(sh_member.cell(r, m_bday).value)
        e_code = parse_dmk_to_code(sh_member.cell(r, m_dmk).value)
        age = calc_age(bday, now_tw) if bday else -1

        ws.cell(out_r, col_clinic).value = clinic_val
        ws.cell(out_r, col_name).value = name
        ws.cell(out_r, col_id).value = sid
        ws.cell(out_r, col_bday).value = bday
        if col_age:
            ws.cell(out_r, col_age).value = (age if age >= 0 else None)
        ws.cell(out_r, col_tel).value = sh_member.cell(r, m_tel).value
        ws.cell(out_r, col_abc).value = sh_member.cell(r, m_abc).value
        ws.cell(out_r, col_dmk_code).value = e_code
        ws.cell(out_r, col_cnt).value = sh_member.cell(r, m_cnt).value
        ws.cell(out_r, col_sex).value = infer_gender_from_id(sid)

        if sid_s and sid_s not in id_to_row:
            id_to_row[sid_s] = out_r
            meta[out_r] = {"bday": bday, "age": age, "e_code": e_code, "ascvd": None}

        out_r += 1

    last_row = out_r - 1

    # --- ASCVD 回填 ---
    ascvd_header_row = 5
    amap = build_header_map(sh_ascvd, ascvd_header_row)
    a_id = find_column_exact(amap, ["ID", "id"])
    a_asc = find_column_exact(amap, ["ASCVD", "ascvd"])
    if a_id is None or a_asc is None:
        raise ValueError("原始檔「ascvd」找不到 ID / ASCVD 欄位（表頭列預期第5列）")

    for r in range(ascvd_header_row + 1, sh_ascvd.max_row + 1):
        pid = sh_ascvd.cell(r, a_id).value
        val = sh_ascvd.cell(r, a_asc).value
        if pid is None or str(pid).strip() == "":
            continue
        if val is None or str(val).strip() == "" or str(val).strip() == "0":
            continue
        tr = id_to_row.get(str(pid).strip())
        if tr:
            ws.cell(tr, col_ascvd).value = val
            meta.setdefault(tr, {})["ascvd"] = val

    # --- 5 大篩檢日期回填 ---
    def fill_screening(sheet, target_col: Optional[int]) -> None:
        if not target_col:
            return
        hmap = build_header_map(sheet, 1)
        sid_col = find_column_exact(hmap, ["ID", "身分證號", "身份證號"])
        dt_col = find_column_exact(hmap, ["最後篩檢日期"])
        if sid_col is None or dt_col is None:
            raise ValueError(f"「{sheet.title}」找不到 ID 或 最後篩檢日期 欄位")

        for rr in range(2, sheet.max_row + 1):
            pid = sheet.cell(rr, sid_col).value
            dt = parse_date(sheet.cell(rr, dt_col).value)
            if pid is None or str(pid).strip() == "" or dt is None:
                continue
            tr = id_to_row.get(str(pid).strip())
            if tr:
                ws.cell(tr, target_col).value = dt

    fill_screening(sh_adult, col_adult)
    fill_screening(sh_pap, col_pap)
    fill_screening(sh_flu, col_flu)
    fill_screening(sh_fit, col_fit)
    fill_screening(sh_hep, col_hep)

    # --- HealthCase 檢驗回填 ---
    hc_map = build_header_map(sh_health, 1)
    hc_id = find_column_exact(hc_map, ["家醫收案會員ID", "ID"])
    hc_hba = find_column_exact(hc_map, ["最近一次HbA1c檢查結果(%)"])
    hc_hba_dt = find_column_exact(hc_map, ["最近一次HbA1c檢查日期"])
    hc_ldl = find_column_exact(hc_map, ["最近一次LDL檢查結果(mg/dL)"])
    hc_ldl_dt = find_column_exact(hc_map, ["最近一次LDL檢查日期"])
    hc_uacr = find_column_exact(hc_map, ["最近一次UACR檢查結果(mg/gm)"])
    hc_uacr_dt = find_column_exact(hc_map, ["最近一次UACR檢查日期"])

    if any(v is None for v in [hc_id, hc_hba, hc_hba_dt, hc_ldl, hc_ldl_dt, hc_uacr, hc_uacr_dt]):
        raise ValueError("原始檔「HealthCase」欄位不完整（家醫收案會員ID / HbA1c結果+日期 / LDL結果+日期 / UACR結果+日期）")

    for r in range(2, sh_health.max_row + 1):
        pid = sh_health.cell(r, hc_id).value
        if pid is None or str(pid).strip() == "":
            continue
        tr = id_to_row.get(str(pid).strip())
        if not tr:
            continue

        hba = parse_float(sh_health.cell(r, hc_hba).value)
        if hba is not None and hba != 0:
            ws.cell(tr, col_hba).value = hba
        hba_dt = parse_date(sh_health.cell(r, hc_hba_dt).value)
        if hba_dt:
            ws.cell(tr, col_hba_dt).value = hba_dt

        ldl = parse_float(sh_health.cell(r, hc_ldl).value)
        if ldl is not None and ldl != 0:
            ws.cell(tr, col_ldl).value = ldl
        ldl_dt = parse_date(sh_health.cell(r, hc_ldl_dt).value)
        if ldl_dt:
            ws.cell(tr, col_ldl_dt).value = ldl_dt

        uacr = parse_float(sh_health.cell(r, hc_uacr).value)
        if uacr is not None and uacr != 0:
            ws.cell(tr, col_uacr).value = uacr
        uacr_dt = parse_date(sh_health.cell(r, hc_uacr_dt).value)
        if uacr_dt:
            ws.cell(tr, col_uacr_dt).value = uacr_dt

    # --- 疾病樣態文字 + 備註 + 分數/明細 + AU/AV/AW(文字型) ---
    for rr in range(DATA_START, last_row + 1):
        e_raw = ws.cell(rr, col_dmk_code).value
        e_code = int(e_raw) if isinstance(e_raw, (int, float)) else parse_dmk_to_code(e_raw)

        f_val = ws.cell(rr, col_ascvd).value
        f_asc = classify_ascvd(f_val)

        ws.cell(rr, col_disease_text).value = disease_group_text(e_code, f_asc)

        age = meta.get(rr, {}).get("age", -1)
        if not isinstance(age, int):
            try:
                age = int(age)
            except Exception:
                age = -1
        sex = normalize_text(ws.cell(rr, col_sex).value)

        adult_dt = parse_date(ws.cell(rr, col_adult).value) if col_adult else None
        pap_dt = parse_date(ws.cell(rr, col_pap).value) if col_pap else None
        flu_dt = parse_date(ws.cell(rr, col_flu).value) if col_flu else None
        fit_dt = parse_date(ws.cell(rr, col_fit).value) if col_fit else None
        hep_dt = parse_date(ws.cell(rr, col_hep).value) if col_hep else None

        note = build_av_note(
            age=age if age >= 0 else 0,
            sex=sex,
            hep_dt=hep_dt,
            fit_dt=fit_dt,
            pap_dt=pap_dt,
            adult_dt=adult_dt,
            flu_dt=flu_dt,
            today=now_tw,
        )
        ws.cell(rr, col_note).value = note if note else None

        hba_val = ws.cell(rr, col_hba).value
        hba_dt = parse_date(ws.cell(rr, col_hba_dt).value) if col_hba_dt else None
        ldl_val = ws.cell(rr, col_ldl).value
        ldl_dt = parse_date(ws.cell(rr, col_ldl_dt).value) if col_ldl_dt else None

        score, breakdown = au_score_with_breakdown(
            e_code=e_code,
            f_ascvd=f_asc,
            hba_val=hba_val,
            hba_dt=hba_dt,
            ldl_val=ldl_val,
            ldl_dt=ldl_dt,
            adult_dt=adult_dt,
            pap_dt=pap_dt,
            flu_dt=flu_dt,
            fit_dt=fit_dt,
            hep_dt=hep_dt,
            age=age,
            sex=sex,
        )
        ws.cell(rr, col_score).value = score
        if col_breakdown:
            ws.cell(rr, col_breakdown).value = breakdown

        # AU：第一次提醒（28天）
        au_txt = build_au_28_note(
            e_code=e_code,
            f_ascvd=f_asc,
            hba_dt=hba_dt,
            ldl_dt=ldl_dt,
            today=now_tw,
        )
        ws.cell(rr, col_au_remind).value = au_txt if au_txt else None

        # AV/AW：由 AU 推導（日期延後）
        if au_txt and (not au_should_skip_followup(au_txt)):
            av_txt = build_followup_text_from_au(au_txt, today=now_tw, offset_days=28)
            aw_txt = build_followup_text_from_au(au_txt, today=now_tw, offset_days=56)
            ws.cell(rr, col_av_remind).value = av_txt if av_txt else None
            ws.cell(rr, col_aw_remind).value = aw_txt if aw_txt else None
        else:
            ws.cell(rr, col_av_remind).value = None
            ws.cell(rr, col_aw_remind).value = None

    # --- 日期欄格式（只格式實際 date）---
    date_cols = [col_bday, col_adult, col_pap, col_flu, col_fit, col_hep, col_hba_dt, col_ldl_dt, col_uacr_dt]
    for r in range(DATA_START, last_row + 1):
        for c in date_cols:
            if not c:
                continue
            cell = ws.cell(r, c)
            if isinstance(cell.value, (datetime.date, datetime.datetime)):
                cell.number_format = "yyyy-mm-dd"

    # --- 整張框線 ---
    add_full_grid(ws)

    # =========================
    # KPI：HbA1c 百分位（AY8/AZ8/BA8）
    # 條件：
    # - 疾病樣態：1 或 3
    # - ASCVD：必須為 0/空白（視為無 ASCVD）
    # - HbA1c 日期：2025 或 2026
    # =========================
    def hba_exclude(_r: int, _e: int, ascvd_val: Any) -> bool:
        # 排除 ASCVD 非 0（也就是有 a/b 或任何非零內容）
        return not is_ascvd_zero_like(ascvd_val)

    _, hba_ratios = compute_percentiles(
        ws=ws,
        data_start=DATA_START,
        last_row=last_row,
        col_dmk_code=col_dmk_code,
        col_ascvd=col_ascvd,
        col_value=col_hba,
        col_date=col_hba_dt,
        valid_dmk=(1, 3),
        valid_years=(2025, 2026),
        thresholds=[7.0, 7.3, 7.5],
        exclude_rule=hba_exclude,
    )
    write_percent_cells(ws, {
        "AY8": hba_ratios[7.0],
        "AZ8": hba_ratios[7.3],
        "BA8": hba_ratios[7.5],
    })

    # =========================
    # KPI：LDL 百分位（BC8/BD8/BE8）
    # 條件（依你截圖）：
    # - 疾病樣態：1/2/3/4
    # - ASCVD：0/a/b 都可
    # - 排除：疾病樣態=4 且 ASCVD=0
    # - LDL 日期：2025 或 2026
    # =========================
    def ldl_exclude(_r: int, e: int, ascvd_val: Any) -> bool:
        return (e == 4) and is_ascvd_zero_like(ascvd_val)

    _, ldl_ratios = compute_percentiles(
        ws=ws,
        data_start=DATA_START,
        last_row=last_row,
        col_dmk_code=col_dmk_code,
        col_ascvd=col_ascvd,
        col_value=col_ldl,
        col_date=col_ldl_dt,
        valid_dmk=(1, 2, 3, 4),
        valid_years=(2025, 2026),
        thresholds=[100.0, 105.0, 110.0],
        exclude_rule=ldl_exclude,
    )
    write_percent_cells(ws, {
        "BC8": ldl_ratios[100.0],
        "BD8": ldl_ratios[105.0],
        "BE8": ldl_ratios[110.0],
    })

    # --- 輸出 ---
    base_dir = os.path.dirname(os.path.abspath(source_xlsx_path))
    ts = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=8))).strftime("%m%d_%H%M")
    out_path = os.path.join(base_dir, f"選會員{ts}.xlsx")
    wb_tpl.save(out_path)
    return out_path


def main() -> None:
    """GUI：選檔 → 執行 → 顯示結果並自動開啟。"""
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()

    src = filedialog.askopenfilename(
        title="選擇原始 Excel 檔案（沒有會員指標的那份）",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not src:
        return

    script_dir = os.path.dirname(os.path.abspath(__file__))
    template = os.path.join(script_dir, TEMPLATE_NAME)

    if not os.path.exists(template):
        messagebox.showerror(
            "錯誤",
            f"找不到模板檔：\n{template}\n\n"
            f"請把模板放到此 .py 同資料夾，檔名需為：\n{TEMPLATE_NAME}"
        )
        return

    try:
        out = process_excel(src, template)
        messagebox.showinfo("完成", f"已輸出：\n{out}")
        open_file_cross_platform(out)
    except (ValueError, KeyError, OSError, InvalidFileException) as e:
        messagebox.showerror("錯誤", str(e))
    except Exception as e:
        messagebox.showerror("錯誤", f"未預期錯誤：{e}")


if __name__ == "__main__":
    main()