# -*- coding: utf-8 -*-
"""
選會員產檔工具 — v5（0322 樣板）

功能概要
- 整合所有來源 sheet 人員名單（會員名單 / ascvd / 自選名單 / 115X / P4P / 月份分頁）
- 套用 0322 樣板，寫入「會員指標」分頁，同步輸出：
  - 百分位名單（LDL / HbA1c 百分位）
  - 醫生看
  - 自選名單
- 填入欄位：疾病樣態 / 主次診斷 / 分數 / 備註 / 追蹤提醒 / 電話分流 / 隱藏輔助欄
- KPI：HbA1c（BG/BH）、LDL（BJ/BK）百分位

注意
- Python 3.9 相容；不拆模組，單一檔案維護
- 模板：選會員樣板0322.xlsx
"""

from __future__ import annotations

import datetime
import math
import os
import re
import subprocess
import sys
import traceback
from copy import copy
from dataclasses import dataclass
from enum import Enum, auto
from typing import Dict, List, Optional, Any, Tuple

import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import column_index_from_string
from openpyxl.utils.exceptions import InvalidFileException


# ============================================================
# KPI 儲存格地址（沿用既有樣板固定位置，直接寫死）
# BG = HbA1c 主, BH = HbA1c 73.8% 目標
# BJ = LDL 主,   BK = LDL 73.8% 目標
# ============================================================
# ============================================================
# 會員總表資料列固定樣式
# AM(39)~AO(41), AV(48), AW(49): 靠左、字型14
# 其餘欄位: 置中、字型24
# ============================================================
_LEFT_COLS  = frozenset([39, 40, 41, 48, 49])   # AM, AN, AO, AV, AW

_FONT_14    = Font(size=14)
_FONT_24    = Font(size=24)
_ALIGN_LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
_ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _apply_member_row_style(ws, row: int, max_col: int) -> None:
    for c in range(1, max_col + 1):
        cell = ws.cell(row, c)
        if c in _LEFT_COLS:
            cell.font      = _FONT_14
            cell.alignment = _ALIGN_LEFT
        else:
            cell.font      = _FONT_24
            cell.alignment = _ALIGN_CENTER


SUMMARY_CELLS = {
    "hba_main_summary":    "AQ2",
    "hba_target_summary":  "AR2",
    "ldl_main_summary":    "AS2",
    "ldl_target_summary":  "AT2",
    "breakdown_backup":    "BL1",
}


def sc(name: str) -> str:
    """回傳 KPI 儲存格地址字串"""
    return SUMMARY_CELLS[name]

# ============================================================
# 模板欄位 alias（標題偵測優先，不再依賴單一名稱）
# ============================================================
TEMPLATE_ALIASES: Dict[str, List[List[str]]] = {
    "clinic": [
        ["診所名稱或機構代碼"],
        ["診所名稱"],
        ["機構代碼"],
        ["院所名稱"],
    ],
    "name": [
        ["姓名"],
        ["個案姓名"],
        ["會員姓名"],
    ],
    "id": [
        ["身份證號碼"],
        ["身分證號碼"],
        ["身份證號"],
        ["身分證號"],
        ["身份証號"],
        ["身分証號"],
        ["ID"],
        ["會員身份證"],
        ["會員身分證"],
    ],
    "bday": [
        ["生日"],
        ["出生日期"],
        ["出生年月日"],
    ],
    "age": [
        ["年齡"],
        ["歲"],
    ],
    "tel": [
        ["電話"],
        ["聯絡電話"],
    ],
    "mobile": [
        ["手機"],
        ["行動電話"],
        ["手機號碼"],
    ],
    "cnt": [
        ["次數"],
        ["件數"],
        ["就診次數"],
    ],
    "sex": [
        ["性別"],
        ["男女"],
    ],
    "abc": [
        ["A/B/C"],
        ["ABC"],
        ["會員"],
    ],
    "dmk_code": [
        ["DM/CKD/DKD"],
        ["疾病樣態編號"],
        ["DM", "CKD", "DKD"],
        ["疾病樣態"],
    ],
    "ascvd": [
        ["ASCVD"],
    ],
    "last_visit": [
        ["最後就診日"],
        ["最近就診日"],
        ["最後看診日"],
    ],
}



# ============================================================
# 業務規則常數（每年調整這裡即可）
# ============================================================
class Rules:
    # 檢驗日期：哪些年份視為「有效」
    VALID_YEARS: Tuple[int, ...] = (2025, 2026)

    # 篩檢計分：哪一年的篩檢才給分
    SCREEN_YEAR: int = 2026

    # 2026 年：只要有檢查就給滿分（無條件給分）
    # 2025 年：需達標才給分
    YEAR_UNCONDITIONAL: int = 2026
    YEAR_VALUE_CHECK: int = 2025

    # 各項分數
    SCORE_HBA: int = 5
    SCORE_LDL: int = 5
    SCORE_ADULT: int = 6
    SCORE_PAP: int = 6
    SCORE_FLU: int = 4
    SCORE_FIT: int = 6
    SCORE_HEP: int = 6

    # 達標值（2025年）
    HBA_PASS_2025: float = 7.0
    LDL_PASS_2025: float = 120.0

    # KPI（HbA1c）
    HBA_CONTROL_THRESHOLD: float = 7.0  # AY8：HbA1c <= 7 的比例
    HBA_TARGET_PERCENT: float = 0.738   # AZ：要找 >=73.8% 的切點

    # KPI（LDL）
    LDL_KPI_THRESHOLDS: Tuple[float, float] = (100.0, 110.0)  # BB8/BC8
    LDL_TARGET_PERCENT: float = 0.738   # BC：要找 >=73.8% 的切點（與 HbA1c 目前相同，但獨立維護）

    # 回診追蹤
    AU_DAYS: int = 28
    AV_OFFSET_DAYS: int = 28
    AW_OFFSET_DAYS: int = 56

    # 模板設定（0322）
    TEMPLATE_NAME: str = "選會員樣板0322.xlsx"
    SHEET_TARGET: str = "會員總表"
    DATA_START_ROW: int = 3

    # 月份申請統計輸出欄位（0322）
    # L=114件數, M=114每月平均金額, N=115件數, O=115每月平均金額
    COL_114_COUNT: str = "L"
    COL_114_COUNT_FULL: str = "BB"   # 隱藏輔助欄：114全年件數
    COL_115_COUNT: str = "N"
    COL_114_AMOUNT: str = "M"
    COL_115_AMOUNT: str = "O"
    COL_114_AMOUNT_TOTAL: str = "BC"  # 隱藏輔助欄：114總金額
    COL_115_AMOUNT_TOTAL: str = "BD"  # 隱藏輔助欄：115總金額
    COL_ADDRESS_HIDDEN: str = "BE"    # 隱藏輔助欄：地址
    COL_IS_114: str = "AX"
    COL_IS_SELF_SELECT: str = "AY"
    COL_IS_115X: str = "AZ"
    COL_NOTE: str = "AW"





# ============================================================
# Enum / Dataclass
# ============================================================
class DiseaseCode(Enum):
    DM = 1
    CKD = 2
    DKD = 3   # DM + CKD
    OTHER = 4 # 非 DM/CKD，可能有 ASCVD


class AscvdCategory(Enum):
    NONE = auto()  # 無 ASCVD
    A = auto()     # 極高風險
    B = auto()     # 非常高風險


@dataclass
class MemberMeta:
    """每位會員的衍生資訊，供後續計算用"""
    row: int
    bday: Optional[datetime.date] = None
    age: int = -1
    e_code: Optional[DiseaseCode] = None
    ascvd: AscvdCategory = AscvdCategory.NONE


@dataclass
class SourceContext:
    wb_src: Any
    sh_member: Any
    sh_ascvd: Any
    sh_health: Any
    sh_main_sub_dx: Any
    sh_phone: Any
    sh_self_select: Any
    sh_115x: Any
    sh_p4p_enroll: Any
    sh_p4p_track: Any
    screening_sheets: Dict[str, Any]
    claim_sums: Dict[str, Dict[str, float]]
    all_members: Dict[str, Dict[str, Any]]


@dataclass
class TemplateContext:
    wb_tpl: Any
    ws: Any
    cols: Dict[str, Optional[int]]
    data_start: int


@dataclass
class RuntimeContext:
    id_to_rows: Dict[str, List[int]]
    meta: Dict[int, MemberMeta]
    last_row: int
    hba_candidates: Optional[List[Tuple[int, float]]] = None
    ldl_candidates: Optional[List[Tuple[int, float]]] = None


@dataclass
class ContactInfo:
    phone: Optional[str] = None
    mobile: Optional[str] = None
    address: Optional[str] = None


# ============================================================
# 工具函數
# ============================================================
def open_file_cross_platform(path: str) -> None:
    """跨平台開啟檔案（盡量不拋錯）"""
    try:
        if sys.platform.startswith("darwin"):
            subprocess.call(("open", path))
        elif os.name == "nt":
            os.startfile(path)  # type: ignore[attr-defined]
        else:
            subprocess.call(("xdg-open", path))
    except Exception:
        pass


def normalize_text(v: Any) -> str:
    if v is None:
        return ""
    return str(v).replace("\t", "").replace("　", "").strip().lstrip("'")


def normalize_header(v: Any) -> str:
    s = normalize_text(v)
    return s.replace(" ", "").replace("\n", "")


def clean_spaces(v: Any) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", "", str(v))


def is_phone_header(text: str) -> bool:
    s = normalize_header(text)
    return ("電話" in s) or ("手機" in s) or ("行動電話" in s)


def is_address_header(text: str) -> bool:
    s = normalize_header(text)
    return any(k in s for k in ("地址", "住址", "聯絡地址", "會員地址", "戶籍地址"))


def normalize_phone_value(v: Any) -> Optional[str]:
    if v is None:
        return None
    digits = re.sub(r"\D+", "", str(v))
    if len(digits) < 8:
        return None
    return digits


def pick_contact_from_values(values: List[Any]) -> ContactInfo:
    info = ContactInfo()
    for raw in values:
        phone = normalize_phone_value(raw)
        if not phone:
            continue
        if phone.startswith("09"):
            if not info.mobile:
                info.mobile = phone
        elif not info.phone:
            info.phone = phone
    return info


def merge_contact_info(primary: Optional[ContactInfo], secondary: Optional[ContactInfo]) -> ContactInfo:
    p = primary or ContactInfo()
    s = secondary or ContactInfo()
    return ContactInfo(
        phone=p.phone or s.phone,
        mobile=p.mobile or s.mobile,
        address=p.address or s.address,
    )


def safe_set(ws, row: int, col: Optional[int], value: Any) -> None:
    if col:
        ws.cell(row, col).value = value


def safe_set_check(ws, row: int, col: Optional[int], value: Any) -> None:
    """寫入打勾符號並置中對齊"""
    if col:
        cell = ws.cell(row, col)
        cell.value = value
        if value is not None:
            cell.alignment = Alignment(horizontal="center", vertical="center")


def calc_age(bday: Optional[datetime.date], ref: datetime.date) -> int:
    if not isinstance(bday, datetime.date):
        return -1
    y = ref.year - bday.year
    if (ref.month, ref.day) < (bday.month, bday.day):
        y -= 1
    return y


def add_months(d: datetime.date, months: int) -> datetime.date:
    y = d.year + (d.month - 1 + months) // 12
    m = (d.month - 1 + months) % 12 + 1
    is_leap = (y % 4 == 0 and (y % 100 != 0 or y % 400 == 0))
    mdays = [31, 29 if is_leap else 28, 31, 30, 31, 30,
             31, 31, 30, 31, 30, 31][m - 1]
    return datetime.date(y, m, min(d.day, mdays))


def infer_gender_from_id(id_value: Any) -> str:
    """由台灣身分證/居留證第2碼推性別（1/8/A/C→男，2/9/B/D→女）"""
    if not id_value:
        return ""
    s = str(id_value).strip().upper()
    if len(s) < 2:
        return ""
    g = s[1]
    if g in ("1", "8", "A", "C"):
        return "男"
    if g in ("2", "9", "B", "D"):
        return "女"
    return ""


# ============================================================
# 解析函數
# ============================================================
def parse_date(value: Any) -> Optional[datetime.date]:
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value

    s = str(value).strip()
    if not s or s in ("-", "—", "–"):
        return None

    patterns = [
        (r"^(\d{4})[/-](\d{1,2})[/-](\d{1,2})$", False),
        (r"^(\d{2,3})(\d{2})(\d{2})$", True),
        (r"^(\d{2,3})[/-](\d{1,2})[/-](\d{1,2})$", True),
    ]
    for pat, is_roc in patterns:
        m = re.match(pat, s)
        if m:
            a, b, c = map(int, m.groups())
            year = a + 1911 if is_roc else a
            try:
                return datetime.date(year, b, c)
            except ValueError:
                return None
    return None


def parse_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    s = str(value).strip().lstrip("'").replace(",", "")
    if not s or s in ("-", "—", "–"):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def parse_disease_code(v: Any) -> Optional[DiseaseCode]:
    s = clean_spaces(v)
    if not s:
        return None
    if s.isdigit():
        mapping = {
            1: DiseaseCode.DM,
            2: DiseaseCode.CKD,
            3: DiseaseCode.DKD,
            4: DiseaseCode.OTHER,
        }
        return mapping.get(int(s))
    text_map = {
        "DM":     DiseaseCode.DM,
        "CKD":    DiseaseCode.CKD,
        "DKD":    DiseaseCode.DKD,
        "DM+CKD": DiseaseCode.DKD,
        "DMCKD":  DiseaseCode.DKD,
    }
    return text_map.get(s.upper())




# ============================================================
# 主次診斷代碼（DM / CKD）
# ============================================================
DM_CODE_PREFIXES: Tuple[str, ...] = ("E08", "E09", "E10", "E11", "E12", "E13")

# 依使用者提供之「主次診斷代碼.docx」整理
CKD_CODES = {
    'A18.11',
    'A52.75',
    'C64.1',
    'C64.2',
    'C64.9',
    'C7A.093',
    'D59.3',
    'E08.21',
    'E08.22',
    'E08.29',
    'E08.65',
    'E09.21',
    'E09.22',
    'E09.29',
    'E09.65',
    'E10.21',
    'E10.22',
    'E10.29',
    'E10.65',
    'E12.21',
    'E12.22',
    'E12.29',
    'E12.65',
    'E11.21',
    'E11.22',
    'E11.29',
    'E11.65',
    'E13.21',
    'E13.22',
    'E13.29',
    'E13.65',
    'E74.8',
    'I70.1',
    'I72.2',
    'I75.81',
    'I77.3',
    'I77.73',
    'K76.7',
    'M10.30',
    'M10.311',
    'M10.312',
    'M10.319',
    'M10.321',
    'M10.322',
    'M10.329',
    'M10.331',
    'M10.332',
    'M10.339',
    'M10.341',
    'M10.342',
    'M10.349',
    'M10.351',
    'M10.352',
    'M10.359',
    'M10.361',
    'M10.362',
    'M10.369',
    'M10.371',
    'M10.372',
    'M10.379',
    'M10.38',
    'M10.39',
    'N00.0',
    'N00.1',
    'N00.2',
    'N00.3',
    'N00.4',
    'N00.5',
    'N00.6',
    'N00.7',
    'N00.8',
    'N00.9',
    'N01.0',
    'N01.1',
    'N01.2',
    'N01.3',
    'N01.4',
    'N01.5',
    'N01.6',
    'N01.7',
    'N01.8',
    'N01.9',
    'N02.0',
    'N02.1',
    'N02.2',
    'N02.3',
    'N02.4',
    'N02.5',
    'N02.6',
    'N02.7',
    'N02.8',
    'N02.9',
    'N03.0',
    'N03.1',
    'N03.2',
    'N03.3',
    'N03.4',
    'N03.5',
    'N03.6',
    'N03.7',
    'N03.8',
    'N03.9',
    'N04.0',
    'N04.1',
    'N04.2',
    'N04.3',
    'N04.4',
    'N04.5',
    'N04.6',
    'N04.7',
    'N04.8',
    'N04.9',
    'N05.0',
    'N05.1',
    'N05.2',
    'N05.3',
    'N05.4',
    'N05.5',
    'N05.6',
    'N05.7',
    'N05.8',
    'N05.9',
    'N06.0',
    'N06.1',
    'N06.2',
    'N06.3',
    'N06.4',
    'N06.5',
    'N06.6',
    'N06.7',
    'N06.8',
    'N06.9',
    'N07.0',
    'N07.1',
    'N07.2',
    'N07.3',
    'N07.4',
    'N07.5',
    'N07.6',
    'N07.7',
    'N07.8',
    'N07.9',
    'N08',
    'N14.0',
    'N14.1',
    'N14.2',
    'N14.3',
    'N14.4',
    'N15.0',
    'N15.8',
    'N15.9',
    'N16',
    'N17.0',
    'N17.1',
    'N17.2',
    'N17.8',
    'N17.9',
    'N18.1',
    'N18.2',
    'N18.3',
    'N20.0',
    'N25.0',
    'N25.1',
    'N25.81',
    'N25.89',
    'N25.9',
    'N26.9',
    'O10.211',
    'O10.212',
    'O10.213',
    'O10.219',
    'O10.22',
    'O10.23',
    'O10.311',
    'O10.312',
    'O10.313',
    'O10.319',
    'O10.32',
    'O10.33',
    'O10.411',
    'O10.412',
    'O10.413',
    'O10.419',
    'O10.42',
    'O10.43',
    'Q61.01',
    'Q61.02',
    'Q61.11',
    'Q61.19',
    'Q61.2',
    'Q61.3',
    'Q61.4',
    'Q61.5',
    'Q61.8',
    'Q62.0',
    'Q62.10',
    'Q62.11',
    'Q62.12',
    'Q62.2',
    'Q62.31',
    'Q62.32',
    'Q62.39',
    'R94.4'
}


def normalize_id(v: Any) -> str:
    return normalize_text(v).upper()


_ID_RE = re.compile(r'^[A-Z]{1,2}\d{8,9}$')

def is_valid_tw_id(v: Any) -> bool:
    """台灣身分證/居留證格式：1英+9數 或 2英+8數"""
    s = normalize_text(v).upper().replace(" ", "")
    return bool(_ID_RE.match(s))


def find_id_col_by_content(sheet, header_row: int, id_col_candidate: Optional[int]) -> Optional[int]:
    """
    先靠 header 找到候選 ID 欄，再往下掃內容驗證格式。
    若 header 找不到則逐欄掃所有欄位內容。
    """
    def _col_has_valid_id(col: int) -> bool:
        for r in range(header_row + 1, min(header_row + 30, sheet.max_row + 1)):
            v = sheet.cell(r, col).value
            if v and is_valid_tw_id(v):
                return True
        return False

    if id_col_candidate and _col_has_valid_id(id_col_candidate):
        return id_col_candidate

    # fallback：掃所有欄
    for c in range(1, sheet.max_column + 1):
        hdr = normalize_header(sheet.cell(header_row, c).value)
        if any(kw in hdr for kw in ["身分", "身份", "ID", "id"]):
            if _col_has_valid_id(c):
                return c
    return None


def normalize_icd_code(v: Any) -> str:
    s = normalize_text(v).upper().replace(" ", "")
    if not s:
        return ""
    # 容忍 E8.9 / E9 / N8 這種少一個 0 的寫法
    s = re.sub(r"^([A-Z])(\d)(?=\.|$)", r"\g<1>0\g<2>", s)
    return s


def extract_icd_codes_from_cell(v: Any) -> List[str]:
    s = normalize_text(v).upper()
    if not s:
        return []
    # 正確使用 word boundary，擷取像 E11.21 / N18.3 / N08 這類 ICD 代碼
    # ICD-10 格式：標準(E11.21/N18.3/J069) + 特殊(C7A.093)
    found = re.findall(r"\b[A-Z]\d{2,3}(?:\.\d+)?\b|\b[A-Z]\d[A-Z](?:\.\d+)?\b", s)
    out: List[str] = []
    seen = set()
    for raw in found:
        code = normalize_icd_code(raw)
        if code and code not in seen:
            seen.add(code)
            out.append(code)
    return out

def is_dm_icd_code(code: str) -> bool:
    c = normalize_icd_code(code)
    return any(c.startswith(p) for p in DM_CODE_PREFIXES)


_DM_CKD_RENAL_RE = re.compile(r'^E(08|09|10|11|12|13)\.2')

def is_ckd_icd_code(code: str) -> bool:
    c = normalize_icd_code(code)
    # E08~E13 小數點後第一位為 2 → 腎臟併發症，直接算 CKD
    if _DM_CKD_RENAL_RE.match(c):
        return True
    return c in CKD_CODES


def classify_main_sub_dx(codes: List[str]) -> str:
    has_dm = any(is_dm_icd_code(c) for c in codes)
    has_ckd = any(is_ckd_icd_code(c) for c in codes)
    if has_dm and has_ckd:
        return "DM+CKD"
    if has_dm:
        return "DM"
    if has_ckd:
        return "CKD"
    return ""


def _build_main_sub_dx_map(sh_dx: Any) -> Dict[str, str]:
    id_aliases = [
        "ID", "id", "身份證號", "身份證號碼", "身分證號", "身分證號碼",
        "會員身份証", "會員身份證", "會員身分證",
    ]
    # 只用 id_aliases 找 header row
    header_row = _find_header_row_contains_any(sh_dx, [id_aliases], search_rows=20)
    if header_row is None:
        raise ValueError("原始檔「主次診斷」找不到 ID 欄位")

    hmap = build_header_map(sh_dx, header_row)
    id_col = find_id_col_by_content(sh_dx, header_row, find_column_exact(hmap, id_aliases))
    if id_col is None:
        raise ValueError("原始檔「主次診斷」找不到 ID 欄位")

    # 包含以下任一關鍵字即視為診斷碼欄
    _DX_KW = ["主診斷", "次診斷", "診斷碼", "診斷代碼", "診斷", "代碼", "病1", "病23", "icd", "疾病碼"]
    dx_cols: List[int] = []
    for hdr, c in hmap.items():
        if c == id_col:
            continue
        if any(kw.lower() in hdr.lower() for kw in _DX_KW):
            dx_cols.append(c)
    dx_cols = sorted(set(dx_cols))
    if not dx_cols:
        raise ValueError("原始檔「主次診斷」找不到診斷碼相關欄位")

    # 找最後就診日欄（可選）—— 模糊比對「就診日」
    last_visit_col = next(
        (c for hdr, c in hmap.items() if "就診日" in hdr and c != id_col),
        None
    )

    id_to_codes: Dict[str, List[str]] = {}
    id_to_last_visit: Dict[str, Optional[datetime.date]] = {}
    for r in range(header_row + 1, sh_dx.max_row + 1):
        pid = normalize_id(sh_dx.cell(r, id_col).value)
        if not pid:
            continue
        codes_here: List[str] = []
        for c in dx_cols:
            codes_here.extend(extract_icd_codes_from_cell(sh_dx.cell(r, c).value))
        if codes_here:
            bucket = id_to_codes.setdefault(pid, [])
            for code in codes_here:
                if code not in bucket:
                    bucket.append(code)
        # 最後就診日（取最新）
        if last_visit_col:
            dt = parse_date(sh_dx.cell(r, last_visit_col).value)
            if dt:
                existing = id_to_last_visit.get(pid)
                if existing is None or dt > existing:
                    id_to_last_visit[pid] = dt

    result: Dict[str, Dict[str, Any]] = {}
    all_pids = set(id_to_codes.keys()) | set(id_to_last_visit.keys())
    for pid in all_pids:
        codes   = id_to_codes.get(pid, [])
        label   = classify_main_sub_dx(codes)
        raw_str = ",".join(codes)
        lv      = id_to_last_visit.get(pid)
        if label or raw_str or lv:
            result[pid] = {"label": label, "raw": raw_str, "last_visit": lv}
    return result


def _fill_main_sub_dx(
    ws,
    sh_dx: Any,
    cols: Dict[str, Optional[int]],
    id_to_rows: Dict[str, List[int]],
) -> None:
    if sh_dx is None:
        return

    dx_map = _build_main_sub_dx_map(sh_dx)
    col_ad = cols.get("main_sub_dx")   # AD：分類結果
    col_q  = cols.get("dx_raw")        # Q：原始診斷碼
    col_l  = cols.get("last_visit")    # L：最後就診日

    for pid, rows in id_to_rows.items():
        info = dx_map.get(normalize_id(pid))
        if not info:
            continue
        for rr in rows:
            if col_ad and info.get("label"):
                safe_set(ws, rr, col_ad, info["label"])
            if col_q and info.get("raw"):
                safe_set(ws, rr, col_q, info["raw"])
            # 最後就診日：只在目前空白時才填（ascvd 已填過的優先）
            if col_l and info.get("last_visit"):
                existing = parse_date(ws.cell(rr, col_l).value)
                new_dt   = info["last_visit"]
                if existing is None or new_dt > existing:
                    safe_set(ws, rr, col_l, new_dt)

def parse_ascvd(v: Any) -> AscvdCategory:
    if v is None:
        return AscvdCategory.NONE
    s = clean_spaces(v).lower()
    if not s or s == "0":
        return AscvdCategory.NONE
    if s in ("a", "1") or "極高" in s:
        return AscvdCategory.A
    if s in ("b", "2") or "非常高" in s:
        return AscvdCategory.B
    if "a" in s:
        return AscvdCategory.A
    if "b" in s:
        return AscvdCategory.B
    return AscvdCategory.NONE


def is_zero_like(v: Any) -> bool:
    """把 None / 空白 / '0' 視為零值"""
    if v is None:
        return True
    return str(v).strip() in ("", "0", "0.0")


# ============================================================
# 月份申請統計（114xx / 115xx）
# ============================================================
def _find_header_row_contains_any(sheet, alias_groups: List[List[str]], search_rows: int = 30) -> Optional[int]:
    # 正規化 alias 一次，避免每列每組重複計算
    norm_groups = [[a.replace(" ", "").replace("\n", "") for a in grp] for grp in alias_groups]
    for r in range(1, min(search_rows, sheet.max_row) + 1):
        row_headers = {normalize_header(sheet.cell(r, c).value) for c in range(1, sheet.max_column + 1)}
        if all(any(a in row_headers for a in grp) for grp in norm_groups):
            return r
    return None


def _sheet_year_bucket(title: str) -> Optional[int]:
    s = str(title).strip()
    if not re.fullmatch(r"\d{5}", s):
        return None
    if s.startswith("114"):
        return 114
    if s.startswith("115"):
        return 115
    return None


def _sheet_month(title: str) -> Optional[int]:
    """從分頁名稱取月份數字（11401→1, 11412→12），不符合則 None"""
    s = str(title).strip()
    if not re.fullmatch(r"\d{5}", s):
        return None
    return int(s[3:5])


def collect_monthly_claim_summaries(wb_src) -> Dict[str, Dict[str, float]]:
    """
    掃描 11401~11412、11501~11512 這類月份分頁，
    依 ID 彙總（目前月份分頁格式）：
      - 114 Q1(11401-11404) D欄件數 -> 114_cnt
      - 114 全年(11401-11412) D欄件數 -> 114_cnt_full（新）
      - 115 Q1(11501-11504) D欄件數 -> 115_cnt
      - 114 E欄金額 -> 114_amt
      - 115 E欄金額 -> 115_amt
    若某 ID 完全沒有資料，後續保持空白，不填 0。
    """
    out: Dict[str, Dict[str, float]] = {}
    id_aliases = ["ID", "身分證號", "身分證號碼", "身份證號", "身份證號碼"]
    seen_115_months: set = set()

    for sheet_name in wb_src.sheetnames:
        year_bucket = _sheet_year_bucket(sheet_name)
        if year_bucket not in (114, 115):
            continue

        month = _sheet_month(sheet_name)
        if month is None:
            continue

        sh = wb_src[sheet_name]
        header_row = _find_header_row_contains_any(
            sh,
            [id_aliases, ["件數"], ["申請金額"]],
            search_rows=30,
        )
        if header_row is None:
            continue

        hmap = build_header_map(sh, header_row)
        id_col_cand = find_column_exact(hmap, id_aliases)
        id_col = find_id_col_by_content(sh, header_row, id_col_cand)
        if id_col is None:
            continue

        # 靠 header 找件數、金額欄（相容不同診所格式）
        count_col  = find_column_exact(hmap, ["件數"])
        amount_col = find_column_exact(hmap, ["申請金額"])
        if count_col is None or amount_col is None:
            continue

        # 是否屬於 Q1（1~4月）
        is_q1 = (month <= 4)
        if year_bucket == 115 and is_q1:
            seen_115_months.add(month)

        for r in range(header_row + 1, sh.max_row + 1):
            pid_raw = sh.cell(r, id_col).value
            pid = normalize_text(pid_raw).upper()
            if not pid or not is_valid_tw_id(pid):
                continue

            cnt = parse_float(sh.cell(r, count_col).value)
            amt = parse_float(sh.cell(r, amount_col).value)
            if cnt is None and amt is None:
                continue

            bucket = out.setdefault(pid, {
                "114_cnt": 0.0,
                "114_cnt_full": 0.0,
                "115_cnt": 0.0,
                "114_amt": 0.0,
                "115_amt": 0.0,
                "115_months": 0.0,
            })

            prefix = str(year_bucket)  # "114" 或 "115"
            if cnt is not None:
                if is_q1:
                    bucket[f"{prefix}_cnt"] += cnt
                if year_bucket == 114:
                    bucket["114_cnt_full"] += cnt
            # 金額只取 Q1
            if amt is not None and is_q1:
                bucket[f"{prefix}_amt"] += amt
    month_count_115 = float(len(seen_115_months) or 1)
    for bucket in out.values():
        bucket["115_months"] = month_count_115

    return out


def _to_excel_number(v: Optional[float]) -> Optional[float]:
    if v is None:
        return None
    if abs(v - round(v)) < 1e-9:
        return int(round(v))
    return round(v, 2)


def _to_excel_int(v: Optional[float]) -> Optional[int]:
    if v is None:
        return None
    return int(round(v))


def _fmt_percent(v: float, denom: int) -> str:
    if denom <= 0:
        return "0.00%"
    return f"{v * 100:.2f}%"


def fill_monthly_claim_summary_columns(
    ws,
    data_start: int,
    last_row: int,
    cols: Dict[str, Optional[int]],
    claim_sums: Dict[str, Dict[str, float]],
) -> None:
    col_m    = cols.get("m_count_114")            # L：114件數
    col_n_fy = cols.get("m_count_114_full")       # BB：114全年件數（輔助）
    col_o    = cols.get("n_count_115")            # N：115件數
    col_s    = cols.get("r_amount_114")           # M：114每月平均金額
    col_t    = cols.get("s_amount_115")           # O：115每月平均金額
    col_s_total = cols.get("r_amount_114_total")  # BC：114總金額（輔助）
    col_t_total = cols.get("s_amount_115_total")  # BD：115總金額（輔助）

    if not all([col_m, col_o, col_s, col_t, cols.get("id")]):
        raise ValueError("模板找不到 L/M/N/O 或 ID 欄位，無法填入月份申請統計")

    for rr in range(data_start, last_row + 1):
        pid = normalize_text(ws.cell(rr, cols["id"]).value).upper()  # type: ignore[index]
        data = claim_sums.get(pid)

        if not data:
            ws.cell(rr, col_m).value = None
            if col_n_fy:
                ws.cell(rr, col_n_fy).value = None
            ws.cell(rr, col_o).value = None
            ws.cell(rr, col_s).value = None
            ws.cell(rr, col_t).value = None
            if col_s_total:
                ws.cell(rr, col_s_total).value = None
            if col_t_total:
                ws.cell(rr, col_t_total).value = None
            continue

        v114c    = data.get("114_cnt", 0.0)
        v114c_fy = data.get("114_cnt_full", 0.0)
        v115c    = data.get("115_cnt", 0.0)
        v114a    = data.get("114_amt", 0.0)
        v115a    = data.get("115_amt", 0.0)
        v115_months = max(int(data.get("115_months", 0.0)), 1)
        v114a_avg = (v114a / 12.0) if v114a != 0 else 0.0
        v115a_avg = (v115a / float(v115_months)) if v115a != 0 else 0.0

        ws.cell(rr, col_m).value = _to_excel_number(v114c) if v114c != 0 else None
        if col_n_fy:
            ws.cell(rr, col_n_fy).value = _to_excel_number(v114c_fy) if v114c_fy != 0 else None
        ws.cell(rr, col_o).value = _to_excel_number(v115c) if v115c != 0 else None
        ws.cell(rr, col_s).value = _to_excel_int(v114a_avg) if v114a != 0 else None
        ws.cell(rr, col_t).value = _to_excel_int(v115a_avg) if v115a != 0 else None
        if col_s_total:
            ws.cell(rr, col_s_total).value = _to_excel_int(v114a) if v114a != 0 else None
        if col_t_total:
            ws.cell(rr, col_t_total).value = _to_excel_int(v115a) if v115a != 0 else None


# ============================================================
# 篩檢需求判斷
# ============================================================
def adult_check_interval_months(age: int) -> Optional[int]:
    if 30 <= age <= 39:
        return 60
    if 40 <= age <= 64:
        return 36
    if age >= 65:
        return 12
    return None


def pap_check_interval_months(age: int, sex: str) -> Optional[int]:
    if sex != "女":
        return None
    if 25 <= age <= 29:
        return 36
    if age >= 30:
        return 12
    return None


def need_flu(age: int) -> bool:
    return age >= 65


def need_fit(age: int) -> bool:
    return 45 <= age <= 75


def need_bc_hep(age: int) -> bool:
    return 45 <= age <= 80


# ============================================================
# 欄位偵測
# ============================================================
def build_header_map(sheet, header_row: int) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for c in range(1, sheet.max_column + 1):
        k = normalize_header(sheet.cell(header_row, c).value)
        if k:
            m[k] = c
    return m


def find_column_exact(hmap: Dict[str, int], aliases: List[str]) -> Optional[int]:
    for a in aliases:
        if a in hmap:
            return hmap[a]
    return None


def _get_merged_header_text(ws, header_row: int, c: int) -> str:
    upper = normalize_header(ws.cell(header_row - 1, c).value) if header_row > 1 else ""
    lower = normalize_header(ws.cell(header_row, c).value)
    return (upper + " " + lower).strip()


def find_col_by_keywords(ws, header_row: int, keywords: List[str]) -> Optional[int]:
    """模糊比對：合併上下兩列表頭，同時包含所有 keywords"""
    keys = [k.replace(" ", "").replace("\n", "") for k in keywords if k]
    for c in range(1, ws.max_column + 1):
        h = _get_merged_header_text(ws, header_row, c).replace(" ", "")
        if h and all(k in h for k in keys):
            return c
    return None


def find_col_by_keywords_any_row(ws, max_row: int, keywords: List[str]) -> Optional[int]:
    """跨多列找欄位：同欄所有儲存格合併後包含全部 keywords"""
    keys = [k.replace(" ", "").replace("\n", "") for k in keywords if k]
    for c in range(1, ws.max_column + 1):
        blob = "".join(
            normalize_header(ws.cell(r, c).value)
            for r in range(1, max_row + 1)
        )
        if blob and all(k in blob for k in keys):
            return c
    return None


def find_header_row_contains(sheet, must_have: List[str], search_rows: int = 250) -> Optional[int]:
    must = [m.replace(" ", "").replace("\n", "") for m in must_have]
    for r in range(1, min(search_rows, sheet.max_row) + 1):
        found = set()
        for c in range(1, sheet.max_column + 1):
            v = normalize_header(sheet.cell(r, c).value)
            if v in must:
                found.add(v)
        if all(m in found for m in must):
            return r
    return None


def find_col_by_alias_groups(ws, header_row: int, alias_groups: List[List[str]]) -> Optional[int]:
    """
    依序嘗試多組 alias。
    每組 alias 內的字串需同時出現在同一欄標題中；先命中的先用。
    """
    for group in alias_groups:
        col = find_col_by_keywords(ws, header_row, group)
        if col:
            return col
    return None


def find_header_row_contains_alias_groups(
    ws,
    required_fields: List[str],
    search_rows: int = 250,
) -> Optional[int]:
    """
    以 alias 規則找最像模板主表頭的列。
    至少要命中所有 required_fields 對應欄位。
    """
    max_r = min(search_rows, ws.max_row)
    for r in range(1, max_r + 1):
        ok = True
        for field in required_fields:
            alias_groups = TEMPLATE_ALIASES.get(field, [])
            if not alias_groups:
                ok = False
                break
            if not find_col_by_alias_groups(ws, r, alias_groups):
                ok = False
                break
        if ok:
            return r
    return None


def _detect_basic_cols(ws, header_row: int) -> Dict[str, Optional[int]]:
    alias = find_col_by_alias_groups
    return {
        "clinic":     alias(ws, header_row, TEMPLATE_ALIASES["clinic"]),
        "name":       alias(ws, header_row, TEMPLATE_ALIASES["name"]),
        "id":         alias(ws, header_row, TEMPLATE_ALIASES["id"]),
        "bday":       alias(ws, header_row, TEMPLATE_ALIASES["bday"]),
        "age":        alias(ws, header_row, TEMPLATE_ALIASES["age"]),
        "tel":        alias(ws, header_row, TEMPLATE_ALIASES["tel"]),
        "mobile":     alias(ws, header_row, TEMPLATE_ALIASES["mobile"]),
        "cnt":        alias(ws, header_row, TEMPLATE_ALIASES["cnt"]),
        "sex":        alias(ws, header_row, TEMPLATE_ALIASES["sex"]),
        "abc":        alias(ws, header_row, TEMPLATE_ALIASES["abc"]),
        "dmk_code":   alias(ws, header_row, TEMPLATE_ALIASES["dmk_code"]),
        "ascvd":      alias(ws, header_row, TEMPLATE_ALIASES["ascvd"]),
        "last_visit": alias(ws, header_row, TEMPLATE_ALIASES["last_visit"]),
    }


def _detect_screening_cols(ws, header_row: int) -> Dict[str, Optional[int]]:
    """
    5大篩檢欄：只依表頭關鍵字偵測，不依賴固定欄位位置。
    """
    kw     = find_col_by_keywords
    kw_any = find_col_by_keywords_any_row
    max_scan = header_row + 3  # 多掃幾列以防標題跨列

    def _find_screening(keywords: List[str]) -> Optional[int]:
        return (
            kw(ws, header_row, keywords)
            or kw_any(ws, max_scan, keywords)
        )

    return {
        "adult": _find_screening(["成人", "健檢"]),
        "pap":   _find_screening(["子宮", "抹片"]),
        "flu":   _find_screening(["老人", "流感"]),
        "fit":   _find_screening(["糞便", "潛血"]),
        "hep":   _find_screening(["BC肝炎"]),
    }


def _detect_lab_cols(ws, header_row: int) -> Dict[str, Optional[int]]:
    kw = find_col_by_keywords
    return {
        "hba":     kw(ws, header_row, ["HbA1c"]),
        "hba_dt":  kw(ws, header_row, ["HbA1c", "日期"]),
        "ldl":     kw(ws, header_row, ["LDL"]),
        "ldl_dt":  kw(ws, header_row, ["LDL", "日期"]),
        "uacr":    kw(ws, header_row, ["UACR"]),
        "uacr_dt": kw(ws, header_row, ["UACR", "日期"]),
    }


def _detect_output_cols(ws, header_row: int, max_scan_row: int) -> Dict[str, Optional[int]]:
    kw = find_col_by_keywords
    kw_any = find_col_by_keywords_any_row
    return {
        "disease_text": (kw(ws, header_row, ["DM/CKD/DKD/ASCVD"])
                         or kw(ws, header_row, ["疾病樣態", "ASCVD"])
                         or kw(ws, header_row, ["疾病樣態", "分類"])),
        "score": (kw(ws, header_row, ["分數"])),
        "breakdown": (kw(ws, header_row, ["分數說明"])
                        or kw_any(ws, max_scan_row, ["分數說明"])),
        "note":  (kw(ws, header_row, ["備註"])),
        # AK：合併大區塊，用 1~data_start-1 掃描
        "ak":    (kw_any(ws, max_scan_row, ["打勾"])
                  or kw_any(ws, max_scan_row, ["HbA1c", "打勾"])
                  or kw(ws, header_row, ["打勾"])
                  or kw(ws, header_row, ["HbA1c", "合格"])),
        # AX：漏檢項目
        "ax":    (kw_any(ws, max_scan_row, ["漏檢項目"])
                  or kw_any(ws, max_scan_row, ["漏檢"])
                  or kw(ws, header_row, ["漏檢項目"])
                  or kw(ws, header_row, ["漏檢"])),
        "main_sub_dx": (kw(ws, header_row, ["主次診斷", "分類"])
                        or kw(ws, header_row, ["主次診斷"])
                        or kw(ws, header_row, ["DM/CKD"])),
        "dx_raw":      (kw(ws, header_row, ["病123"])
                        or kw(ws, header_row, ["英文字照抄"])),
        # 後段輔助欄（旗標 / 統計 / 隱藏地址）
        "p4p_status":     kw(ws, header_row, ["P4P收案狀態"]),
        "p4p_enroll_dt":  kw(ws, header_row, ["收案日期"]),
        "p4p_last_dt":    kw(ws, header_row, ["最後追蹤日"]),
        "p4p_next_dt":    (kw(ws, header_row, ["下次應追蹤日"])
                           or kw(ws, header_row, ["下次追蹤日"])),
        "is_114":         kw(ws, header_row, ["是否為114會員名單"]),
        "is_self_select": kw(ws, header_row, ["是否為自選會員"]),
        "is_115x":        kw(ws, header_row, ["是否為115X"]),
        "m_count_114":    kw(ws, header_row, ["114年", "就診次數"]),
        "n_count_115":    kw(ws, header_row, ["115年", "就診次數"]),
        "r_amount_114":   kw(ws, header_row, ["114年", "申報金額", "月"]),
        "s_amount_115":   kw(ws, header_row, ["115年", "申報金額", "月"]),
        "m_count_114_full": kw(ws, header_row, ["114年全年就診次數"]),
        "r_amount_114_total": kw(ws, header_row, ["114年申報總金額"]),
        "s_amount_115_total": kw(ws, header_row, ["115年申報總金額"]),
        "address_hidden": kw(ws, header_row, ["地址"]),
    }


def _detect_kpi_col(ws, max_scan_row: int, keywords: List[str]) -> Optional[int]:
    """找 KPI 百分位標記欄，只依表頭文字判斷。"""
    keys = [normalize_header(k).lower() for k in keywords if k]
    for c in range(1, ws.max_column + 1):
        blob = "".join(
            normalize_header(ws.cell(r, c).value)
            for r in range(1, max_scan_row + 1)
        ).lower()
        if blob and all(k in blob for k in keys):
            return c
    return None


def _detect_followup_cols(ws, max_scan_row: int) -> Dict[str, Optional[int]]:
    kw = find_col_by_keywords_any_row
    return {
        "au": (kw(ws, max_scan_row, ["第一次提醒", "28天"])
               or kw(ws, max_scan_row, ["回診追蹤", "28天"])
               or kw(ws, max_scan_row, ["28天"])),
        "av": (kw(ws, max_scan_row, ["第二次提醒", "56天"])
               or kw(ws, max_scan_row, ["回診追蹤", "56天"])),
        "aw": (kw(ws, max_scan_row, ["第三次提醒", "84天"])
               or kw(ws, max_scan_row, ["回診追蹤", "84天"])),
    }


def detect_template_columns(ws, data_start: int) -> Dict[str, Optional[int]]:
    header_row = (
        find_header_row_contains_alias_groups(ws, ["name", "id"], 250)
        or find_header_row_contains(ws, ["姓名", "身份證號碼"], 250)
        or 1
    )
    max_scan_row = data_start - 1

    cols: Dict[str, Optional[int]] = {}
    cols.update(_detect_basic_cols(ws, header_row))
    cols.update(_detect_screening_cols(ws, header_row))
    cols.update(_detect_lab_cols(ws, header_row))

    followup_cols = _detect_followup_cols(ws, max_scan_row)
    cols.update(followup_cols)
    cols.update(_detect_output_cols(ws, header_row, max_scan_row))

    # KPI 標記欄（依表頭文字偵測）
    cols["ay_mark"] = (
        _detect_kpi_col(ws, max_scan_row, ["HbA1c<=7"])
        or _detect_kpi_col(ws, max_scan_row, ["HAb1c<=7"])
    )
    cols["az_mark"] = (
        _detect_kpi_col(ws, max_scan_row, ["HbA1c百分位", "73.8"])
        or _detect_kpi_col(ws, max_scan_row, ["HAb1c百分位", "73.8"])
    )
    cols["bb_mark"] = _detect_kpi_col(ws, max_scan_row, ["LDL百分位<=100"])
    cols["bc_mark"] = _detect_kpi_col(ws, max_scan_row, ["LDL百分位", "73.8"])

    # 必填欄位驗證：集中在最後檢查，避免偵測流程夾雜太多錯誤處理
    _require_cols(cols, [
        "name", "id", "bday", "tel", "abc",
        "mobile",
        "dmk_code", "ascvd", "sex",
        "hba", "hba_dt", "ldl", "ldl_dt", "uacr", "uacr_dt",
        "disease_text", "score", "breakdown", "note",
        "au", "av", "aw",
        "ak", "ax",
        "m_count_114", "m_count_114_full", "n_count_115", "r_amount_114", "s_amount_115",
        "ay_mark", "az_mark", "bb_mark", "bc_mark",
    ])
    return cols


def _require_cols(cols: Dict[str, Optional[int]], required: List[str]) -> None:
    missing = [k for k in required if not cols.get(k)]
    if missing:
        raise ValueError(f"新模板欄位找不到：{'、'.join(missing)}")


# ============================================================
# 業務邏輯：疾病樣態文字（✅ DM+CKD 後面不加 (DKD)）
# ============================================================
def disease_group_text(
    e_code: Optional[DiseaseCode], ascvd: AscvdCategory
) -> Optional[str]:
    has_dm    = e_code in (DiseaseCode.DM, DiseaseCode.DKD)
    has_ckd   = e_code in (DiseaseCode.CKD, DiseaseCode.DKD)
    has_ascvd = ascvd != AscvdCategory.NONE

    table = {
        (True,  False, False): "DM",
        (False, True,  False): "CKD",
        (True,  True,  False): "DM+CKD",
        (False, False, True):  "ASCVD",
        (False, True,  True):  "CKD+ASCVD",
        (True,  False, True):  "DM+ASCVD",
        (True,  True,  True):  "DM+CKD+ASCVD",
    }
    return table.get((has_dm, has_ckd, has_ascvd))


# ============================================================
# 業務邏輯：AK（HbA1c 打勾 ✔）
# ============================================================
def should_check_ak(
    *,
    e_code: Optional[DiseaseCode],
    ascvd: AscvdCategory,
    age: int,
    hba_val: Any,
) -> bool:
    """
    AK 打勾規則（✔）：
    - DM：HbA1c < 8
    - DM+CKD / DM+ASCVD / DM+CKD+ASCVD：
        - 年齡 >= 80：HbA1c < 8
        - 其餘：HbA1c < 7
    其他疾病：不勾
    """
    has_dm  = e_code in (DiseaseCode.DM, DiseaseCode.DKD)
    has_ckd = e_code in (DiseaseCode.CKD, DiseaseCode.DKD)
    has_asc = ascvd != AscvdCategory.NONE

    if not has_dm:
        return False

    v = parse_float(hba_val)
    if v is None:
        return False

    # DM only
    if (not has_ckd) and (not has_asc):
        return v < 8.0

    # DM+CKD / DM+ASCVD / DM+CKD+ASCVD
    if age < 0:
        return False

    if age >= 80:
        return v < 8.0
    return v < 7.0


# ============================================================
# 業務邏輯：AX（漏檢項目）
# ============================================================
def _ascvd_token(v: Any) -> str:
    """AX 用：ASCVD 欄只接受 0/a/b（0 也算有值）"""
    s = clean_spaces(v).lower()
    if s in ("0", "a", "b"):
        return s
    return ""


def build_ax_leak_item(
    *,
    d_code: Optional[DiseaseCode],  # D欄：1/2/3/4
    ascvd_raw: Any,                 # E欄：0/a/b
    hba_dt: Optional[datetime.date],  # AD：HbA1c 檢查日期（程式內用 hba_dt）
    ldl_dt: Optional[datetime.date],  # AF：LDL 檢查日期（程式內用 ldl_dt）
) -> Optional[str]:
    """
    條件（同時成立）：
    - D(疾病樣態)=1/2/3/4
    - E(ascvd)=0/a/b（0 也算有值）
    - 排除：D=4 且 E=0
    - 只有 D=1 或 D=3 判斷漏檢
    - HbA1c日期 與 LDL日期 皆有值且不同天

    顯示：
    - HbA1c日 > LDL日 → LDL漏檢
    - LDL日 > HbA1c日 → HbA1c漏檢
    - 相同日 → 不顯示
    """
    if d_code not in (DiseaseCode.DM, DiseaseCode.CKD, DiseaseCode.DKD, DiseaseCode.OTHER):
        return None

    token = _ascvd_token(ascvd_raw)
    if not token:  # 必須是 0/a/b
        return None

    # 排除：D=4 且 E=0
    if d_code == DiseaseCode.OTHER and token == "0":
        return None

    # 只有 D=1 或 D=3 判斷漏檢
    if d_code not in (DiseaseCode.DM, DiseaseCode.DKD):
        return None

    if not isinstance(hba_dt, datetime.date) or not isinstance(ldl_dt, datetime.date):
        return None

    if hba_dt == ldl_dt:
        return None

    if hba_dt > ldl_dt:
        return "LDL漏檢"
    if ldl_dt > hba_dt:
        return "HbA1c漏檢"
    return None


# ============================================================
# 業務邏輯：備註（需要篩檢提示）
# ============================================================
def build_screening_note(
    *,
    age:      int,
    sex:      str,
    hep_dt:   Optional[datetime.date],
    fit_dt:   Optional[datetime.date],
    pap_dt:   Optional[datetime.date],
    adult_dt: Optional[datetime.date],
    flu_dt:   Optional[datetime.date],
    today:    datetime.date,
) -> str:
    msgs: List[str] = []

    if need_bc_hep(age) and hep_dt is None:
        msgs.append("今年需檢測BC肝")

    if need_fit(age):
        if fit_dt is None or add_months(fit_dt, 24) <= today:
            msgs.append("今年需檢測糞便")

    pap_interval = pap_check_interval_months(age, sex)
    if pap_interval:
        if pap_dt is None or add_months(pap_dt, pap_interval) <= today:
            msgs.append("今年需檢測子抹")

    adult_interval = adult_check_interval_months(age)
    if adult_interval:
        if adult_dt is None or add_months(adult_dt, adult_interval) <= today:
            msgs.append("今年需檢測成健")

    if need_flu(age):
        if flu_dt is None or flu_dt.year < today.year:
            msgs.append("今年需檢測老流")

    return "\n".join(msgs)


# ============================================================
# 業務邏輯：分數計算
# ============================================================
def _score_hba(hba_val: Any, hba_dt: Optional[datetime.date]) -> int:
    if not isinstance(hba_dt, datetime.date):
        return 0
    if hba_dt.year == Rules.YEAR_UNCONDITIONAL:
        return Rules.SCORE_HBA
    if hba_dt.year == Rules.YEAR_VALUE_CHECK:
        v = parse_float(hba_val)
        if v is not None and v <= Rules.HBA_PASS_2025:
            return Rules.SCORE_HBA
    return 0


def _score_ldl(ldl_val: Any, ldl_dt: Optional[datetime.date]) -> int:
    if not isinstance(ldl_dt, datetime.date):
        return 0
    if ldl_dt.year == Rules.YEAR_UNCONDITIONAL:
        return Rules.SCORE_LDL
    if ldl_dt.year == Rules.YEAR_VALUE_CHECK:
        v = parse_float(ldl_val)
        if v is not None and v <= Rules.LDL_PASS_2025:
            return Rules.SCORE_LDL
    return 0


def _score_screening(dt: Optional[datetime.date], pts: int) -> int:
    if isinstance(dt, datetime.date) and dt.year == Rules.SCREEN_YEAR:
        return pts
    return 0


def _score_age(age: int, sex: str) -> int:
    sex = normalize_text(sex)
    if age < 18:  # 年齡未知（-1）或不合理（<18）均不給分
        return 0
    # 業務規則：年輕族群額外加分，女性 25 歲以下 / 男性 30 歲以下均給 28 分。
    if sex == "女" and age <= 25:
        return 28
    if sex == "男" and age <= 30:
        return 28
    return 0


def _safe_int(value: Any) -> int:
    if value is None:
        return 0
    try:
        v = int(float(str(value).strip().lstrip("'").replace(",", "")))
        return max(v, 0)
    except Exception:
        return 0


def _score_visit_count_115(count_115: Any) -> int:
    n = _safe_int(count_115)
    if n <= 0:
        return 0
    return min(n + 1, 10)

def calc_score(
    *,
    hba_val:  Any,
    hba_dt:   Optional[datetime.date],
    ldl_val:  Any,
    ldl_dt:   Optional[datetime.date],
    adult_dt: Optional[datetime.date],
    pap_dt:   Optional[datetime.date],
    flu_dt:   Optional[datetime.date],
    fit_dt:   Optional[datetime.date],
    hep_dt:   Optional[datetime.date],
    age:      int,
    sex:      str,
    visit_count_115: Any,
) -> Tuple[int, str]:
    age_score = _score_age(age, sex)

    prevention_score = (
        _score_screening(adult_dt, Rules.SCORE_ADULT)
        + _score_screening(pap_dt, Rules.SCORE_PAP)
        + _score_screening(flu_dt, Rules.SCORE_FLU)
        + _score_screening(fit_dt, Rules.SCORE_FIT)
        + _score_screening(hep_dt, Rules.SCORE_HEP)
    )

    exam_score = _score_hba(hba_val, hba_dt) + _score_ldl(ldl_val, ldl_dt)
    visit_score = _score_visit_count_115(visit_count_115)

    total = age_score + prevention_score + exam_score + visit_score
    parts: List[str] = []
    if age_score > 0:
        parts.append(f"1.年齡{age_score}分")
    if prevention_score > 0:
        parts.append(f"2.預防保健{prevention_score}分")
    if exam_score > 0:
        parts.append(f"3.檢查{exam_score}分")
    if visit_score > 0:
        parts.append(f"4.固定就診次數{visit_score}分")
    breakdown = "，\n".join(parts) if parts else "0分"
    return total, breakdown


# ============================================================
# 業務邏輯：AU 第一次提醒（28天）
# ============================================================
def _is_valid_year(dt: Optional[datetime.date]) -> bool:
    return isinstance(dt, datetime.date) and dt.year in Rules.VALID_YEARS


def _au_item_line(item: str, last_dt: Optional[datetime.date], today: datetime.date) -> str:
    if not _is_valid_year(last_dt):
        return f"{item}:超過2年未檢查"
    due = last_dt + datetime.timedelta(days=Rules.AU_DAYS)  # type: ignore[operator]
    if due <= today:
        return f"{item}:可立刻通知回診"
    return f"{item}:{due.strftime('%Y-%m-%d')}需回診"


def build_au_note(
    *,
    e_code: Optional[DiseaseCode],
    ascvd:  AscvdCategory,
    hba_dt: Optional[datetime.date],
    ldl_dt: Optional[datetime.date],
    today:  datetime.date,
) -> str:
    has_dm    = e_code in (DiseaseCode.DM, DiseaseCode.DKD)
    has_ckd   = e_code in (DiseaseCode.CKD, DiseaseCode.DKD)
    has_ascvd = ascvd != AscvdCategory.NONE

    need_hba = has_dm
    need_ldl = has_dm or has_ckd or has_ascvd

    if not need_hba and not need_ldl:
        return "非疾病不需回診"

    any_valid = (
        (need_hba and _is_valid_year(hba_dt))
        or (need_ldl and _is_valid_year(ldl_dt))
    )
    lines: List[str] = []
    if not any_valid:
        if need_hba:
            lines.append("HbA1c:超過2年未檢查")
        if need_ldl:
            lines.append("LDL:超過2年未檢查")
        return "\n".join(lines)

    if need_hba:
        lines.append(_au_item_line("HbA1c", hba_dt, today))
    if need_ldl:
        lines.append(_au_item_line("LDL", ldl_dt, today))
    return "\n".join(lines)


# ============================================================
# 業務邏輯：AV/AW 從 AU 文字延後
# ============================================================
_DATE_RE = re.compile(r"(\d{4}-\d{2}-\d{2})")


def _should_skip_followup(au_txt: str) -> bool:
    s = (au_txt or "").replace(" ", "")
    return (not s) or ("非疾病" in s)


def _followup_item_line(
    item: str, base_dt: Optional[datetime.date], today: datetime.date, offset: int
) -> str:
    base = base_dt if isinstance(base_dt, datetime.date) else today
    due  = base + datetime.timedelta(days=offset)
    return f"{item}:{due.strftime('%Y-%m-%d')}需回診"


def build_followup_note(
    au_txt: str, today: datetime.date, offset_days: int
) -> Optional[str]:
    if not au_txt:
        return None

    out: List[str] = []
    for ln in (line.strip() for line in au_txt.splitlines() if line.strip()):
        if ln.startswith("HbA1c:"):
            item = "HbA1c"
        elif ln.startswith("LDL:"):
            item = "LDL"
        else:
            continue

        if "可立刻通知" in ln:
            out.append(_followup_item_line(item, None, today, offset_days))
            continue

        m = _DATE_RE.search(ln)
        if m:
            try:
                base = datetime.datetime.strptime(m.group(1), "%Y-%m-%d").date()
                out.append(_followup_item_line(item, base, today, offset_days))
            except ValueError:
                pass

    return "\n".join(out) if out else None


# ============================================================
# KPI：HbA1c（AY8 / AZ7 / AZ8 / AZ9）
# ============================================================
def _write_ratio(ws, addr: str, numer: int, denom: int) -> None:
    ws[addr].value = (numer / denom) if denom > 0 else 0
    ws[addr].number_format = "0.00%"


def _write_summary_text(ws, addr: str, label: Optional[str], numer: int, denom: int) -> None:
    label_txt = label or ""
    ratio_txt = _fmt_percent((numer / denom) if denom > 0 else 0.0, denom)
    fraction_txt = f"{numer}/{denom}" if denom > 0 else "0/0"
    ws[addr].value = f"{label_txt}，{ratio_txt}，{fraction_txt}"


def _collect_hba_candidates(
    ws, cols: Dict[str, Optional[int]], data_start: int, last_row: int
) -> List[Tuple[int, float]]:
    candidates: List[Tuple[int, float]] = []
    for r in range(data_start, last_row + 1):
        e = parse_disease_code(ws.cell(r, cols["dmk_code"]).value)  # type: ignore[index]
        if e not in (DiseaseCode.DM, DiseaseCode.DKD):
            continue
        if not is_zero_like(ws.cell(r, cols["ascvd"]).value):  # type: ignore[index]
            continue
        hba_dt = parse_date(ws.cell(r, cols["hba_dt"]).value)  # type: ignore[index]
        if not (isinstance(hba_dt, datetime.date) and hba_dt.year in Rules.VALID_YEARS):
            continue
        hba = parse_float(ws.cell(r, cols["hba"]).value)  # type: ignore[index]
        if hba is None:
            continue
        candidates.append((r, hba))
    return candidates


def _collect_ldl_candidates(
    ws, cols: Dict[str, Optional[int]], data_start: int, last_row: int
) -> List[Tuple[int, float]]:
    candidates: List[Tuple[int, float]] = []
    for r in range(data_start, last_row + 1):
        e = parse_disease_code(ws.cell(r, cols["dmk_code"]).value)  # type: ignore[index]
        if e not in (DiseaseCode.DM, DiseaseCode.CKD, DiseaseCode.DKD, DiseaseCode.OTHER):
            continue

        ascvd = parse_ascvd(ws.cell(r, cols["ascvd"]).value)  # type: ignore[index]
        if e == DiseaseCode.OTHER and ascvd == AscvdCategory.NONE:
            continue

        ldl_dt = parse_date(ws.cell(r, cols["ldl_dt"]).value)  # type: ignore[index]
        if not (isinstance(ldl_dt, datetime.date) and ldl_dt.year in Rules.VALID_YEARS):
            continue

        ldl = parse_float(ws.cell(r, cols["ldl"]).value)  # type: ignore[index]
        if ldl is None:
            continue

        candidates.append((r, ldl))
    return candidates


def collect_kpi_mark_sets(
    ws, cols: Dict[str, Optional[int]], data_start: int, last_row: int,
    hba_candidates: Optional[List[Tuple[int, float]]] = None,
    ldl_candidates: Optional[List[Tuple[int, float]]] = None,
) -> Dict[str, set]:
    if hba_candidates is None:
        hba_candidates = _collect_hba_candidates(ws, cols, data_start, last_row)
    if ldl_candidates is None:
        ldl_candidates = _collect_ldl_candidates(ws, cols, data_start, last_row)

    ay_rows = {r for r, v in hba_candidates if v <= Rules.HBA_CONTROL_THRESHOLD}

    hba_sorted = sorted(hba_candidates, key=lambda x: (x[1], x[0]))
    hba_k = int(math.ceil(Rules.HBA_TARGET_PERCENT * len(hba_sorted))) if hba_sorted else 0
    az_rows = {r for r, _ in hba_sorted[:hba_k]}

    ldl_threshold = Rules.LDL_KPI_THRESHOLDS[0]
    bb_rows = {r for r, v in ldl_candidates if v <= ldl_threshold}

    ldl_sorted = sorted(ldl_candidates, key=lambda x: (x[1], x[0]))
    ldl_k = int(math.ceil(Rules.LDL_TARGET_PERCENT * len(ldl_sorted))) if ldl_sorted else 0
    bc_rows = {r for r, _ in ldl_sorted[:ldl_k]}

    return {
        "ay": ay_rows,
        "az": az_rows,
        "bb": bb_rows,
        "bc": bc_rows,
    }


def calc_hba_kpi_ay_az(
    ws, cols: Dict[str, Optional[int]], data_start: int, last_row: int,
    hba_candidates: Optional[List[Tuple[int, float]]] = None,
) -> None:
    if hba_candidates is None:
        hba_candidates = _collect_hba_candidates(ws, cols, data_start, last_row)
    hba_values = [v for _, v in hba_candidates]
    denom = len(hba_values)

    numer_ay = sum(1 for v in hba_values if v <= Rules.HBA_CONTROL_THRESHOLD)
    _write_summary_text(ws, sc("hba_main_summary"), "<=7", numer_ay, denom)

    if denom <= 0:
        _write_summary_text(ws, sc("hba_target_summary"), "", 0, 0)
        print("AZ 分母=0，分子=0，比例=0.00%，切點=None")
        return

    hba_values.sort()
    k = int(math.ceil(Rules.HBA_TARGET_PERCENT * denom))
    k = max(1, min(k, denom))
    cutoff = hba_values[k - 1]

    _write_summary_text(ws, sc("hba_target_summary"), f"<={cutoff:.2f}", k, denom)

    ratio = k / denom
    print(f"AZ 分母={denom}，分子={k}，比例={ratio*100:.2f}%，切點={cutoff:.2f}")


# ============================================================
# KPI：LDL（BB8/BC8）
# ============================================================
def calc_ldl_percentiles(
    ws, cols: Dict[str, Optional[int]], data_start: int, last_row: int,
    ldl_candidates: Optional[List[Tuple[int, float]]] = None,
) -> None:
    if ldl_candidates is None:
        ldl_candidates = _collect_ldl_candidates(ws, cols, data_start, last_row)
    ldl_values = [v for _, v in ldl_candidates]
    th_control = Rules.LDL_KPI_THRESHOLDS[0]  # 100
    denom = len(ldl_values)

    numer_bb = sum(1 for v in ldl_values if v <= th_control)
    _write_summary_text(ws, sc("ldl_main_summary"), "<=100", numer_bb, denom)

    if denom <= 0:
        _write_summary_text(ws, sc("ldl_target_summary"), "", 0, 0)
        print("BC 分母=0，分子=0，比例=0.00%，切點=None")
        return

    ldl_values.sort()
    k = int(math.ceil(Rules.LDL_TARGET_PERCENT * denom))
    k = max(1, min(k, denom))
    cutoff = ldl_values[k - 1]

    _write_summary_text(ws, sc("ldl_target_summary"), f"<={cutoff:.1f}".replace(".0", ""), k, denom)

    ratio = k / denom
    print(f"BC 分母={denom}，分子={k}，比例={ratio*100:.2f}%，切點={cutoff:.0f}")


# ============================================================
# 格式工具
# ============================================================
def apply_full_grid(ws, max_row: int, max_col: int) -> None:
    thin = Side(style="thin")
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    # 只對實際有資料的儲存格設定邊框，跳過空值以提升效能
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            if cell.value is not None or r == 1:
                cell.border = grid


def apply_date_format(
    ws, cols: Dict[str, Optional[int]], data_start: int, last_row: int
) -> None:
    date_col_keys = ["bday", "adult", "pap", "flu", "fit", "hep",
                     "hba_dt", "ldl_dt", "uacr_dt",
                     "last_visit", "p4p_enroll_dt", "p4p_last_dt", "p4p_next_dt"]
    date_cols = [cols.get(k) for k in date_col_keys if cols.get(k)]
    for r in range(data_start, last_row + 1):
        for c in date_cols:
            cell = ws.cell(r, c)  # type: ignore[arg-type]
            if isinstance(cell.value, (datetime.date, datetime.datetime)):
                cell.number_format = "yyyy-mm-dd"


def apply_amount_format(
    ws, cols: Dict[str, Optional[int]], data_start: int, last_row: int
) -> None:
    amount_col_keys = [
        "r_amount_114", "s_amount_115",
        "r_amount_114_total", "s_amount_115_total",
    ]
    amount_cols = [cols.get(k) for k in amount_col_keys if cols.get(k)]
    for r in range(data_start, last_row + 1):
        for c in amount_cols:
            cell = ws.cell(r, c)  # type: ignore[arg-type]
            if cell.value is not None:
                cell.number_format = "#,##0"



# ============================================================
# 醫生看 sheet（從會員指標 Key 過來）
# ============================================================

# ============================================================
def _build_src_col_map(
    col_map: List[Tuple[str, str]],
    cols: Dict[str, Optional[int]],
) -> Dict[str, Optional[int]]:
    """
    根據 col_map（來源欄位 key, 目標欄字母）和 cols dict，
    建立來源欄位 key → 實際欄號的對照表。
    """
    result: Dict[str, Optional[int]] = {}
    for src_key, _ in col_map:
        result[src_key] = cols.get(src_key)
    return result


def _copy_sheet_rows(
    ws_main,
    ws_out,
    col_map: List[Tuple[str, str]],
    src_col_map: Dict[str, Optional[int]],
    data_start: int,
    last_row: int,
    dst_data_start: int = 3,
    filter_ids: Optional[set] = None,
    id_main_col: Optional[int] = None,
) -> int:
    """
    通用：把 ws_main 的資料按 col_map 逐列 copy 到 ws_out。
    filter_ids: 若指定，只 copy ID 在此 set 內的列。
    回傳實際寫入列數。
    """
    # 清空目標舊資料
    for r in range(dst_data_start, ws_out.max_row + 1):
        for c in range(1, ws_out.max_column + 1):
            ws_out.cell(r, c).value = None

    # 預先把 (src欄號, dst欄號) 轉換好，避免每列重複呼叫 column_index_from_string
    _col_pairs: List[Tuple[Optional[int], int]] = [
        (src_col_map.get(src_key), column_index_from_string(dst_letter))
        for src_key, dst_letter in col_map
    ]
    _center_align = Alignment(horizontal="center", vertical="center")

    dst_row = dst_data_start
    for src_row in range(data_start, last_row + 1):
        # 過濾 ID
        if filter_ids is not None:
            pid = normalize_id(ws_main.cell(src_row, id_main_col).value) if id_main_col else ""
            if pid not in filter_ids:
                continue

        for src_c, dst_c in _col_pairs:
            if src_c:
                val = ws_main.cell(src_row, src_c).value
                cell = ws_out.cell(dst_row, dst_c)
                cell.value = val
                if val == "v":
                    cell.alignment = _center_align
        dst_row += 1

    ws_out.sheet_view.showGridLines = True
    return dst_row - dst_data_start


SELF_SELECT_SHEET_NAME = "自選名單(從會員指標內容Key過來)"

# ============================================================
# 醫生看 sheet
# ============================================================
DOCTOR_SHEET_NAME = "醫生看(從會員指標內容Key過來)"

_DOCTOR_COL_MAP: List[Tuple[str, str]] = [
    ("id", "A"),
    ("name", "B"),
    ("bday", "C"),
    ("mobile", "D"),
    ("dx_raw", "E"),
    ("dmk_code", "F"),
    ("ascvd", "G"),
    ("last_visit", "H"),
    ("m_count_114_full", "I"),
    ("m_count_114", "J"),
    ("n_count_115", "K"),
    ("r_amount_114_total", "L"),
    ("s_amount_115_total", "M"),
    ("adult", "N"),
    ("pap", "O"),
    ("flu", "P"),
    ("fit", "Q"),
    ("hep", "R"),
    ("hba", "S"),
    ("hba_dt", "T"),
    ("ldl", "U"),
    ("ldl_dt", "V"),
    ("p4p_status", "W"),
    ("p4p_enroll_dt", "X"),
    ("p4p_last_dt", "Y"),
    ("p4p_next_dt", "Z"),
    ("is_115x", "AA"),
    ("is_self_select", "AB"),
    ("is_114", "AC"),
    ("score", "AD"),
]


def populate_doctor_sheet(
    wb_tpl,
    ws_main,
    cols: Dict[str, Optional[int]],
    data_start: int,
    last_row: int,
) -> None:
    if DOCTOR_SHEET_NAME not in wb_tpl.sheetnames:
        return
    src_col_map = _build_src_col_map(_DOCTOR_COL_MAP, cols)
    ws_doc = wb_tpl[DOCTOR_SHEET_NAME]
    n = _copy_sheet_rows(ws_main, ws_doc, _DOCTOR_COL_MAP, src_col_map, data_start, last_row)
    print(f"醫生看 sheet 已寫入 {n} 列")


# ============================================================
# 自選名單 sheet
# ============================================================
_SELF_SELECT_COL_MAP: List[Tuple[str, str]] = [
    ("name", "A"),
    ("id", "B"),
    ("m_count_114_full", "C"),
    ("m_count_114", "D"),
    ("n_count_115", "E"),
    ("r_amount_114_total", "F"),
    ("s_amount_115_total", "G"),
    ("adult", "H"),
    ("pap", "I"),
    ("flu", "J"),
    ("fit", "K"),
    ("hep", "L"),
    ("hba", "M"),
    ("hba_dt", "N"),
    ("ldl", "O"),
    ("ldl_dt", "P"),
    ("p4p_status", "Q"),
    ("p4p_enroll_dt", "R"),
    ("p4p_last_dt", "S"),
    ("p4p_next_dt", "T"),
    ("is_114", "U"),
]


def populate_self_select_sheet(
    wb_tpl,
    ws_main,
    cols: Dict[str, Optional[int]],
    data_start: int,
    last_row: int,
    sh_self_select: Any,
) -> None:
    if SELF_SELECT_SHEET_NAME not in wb_tpl.sheetnames:
        return
    if sh_self_select is None:
        return

    # 建立自選名單的 ID set
    id_aliases = ["身份證號", "身分證號", "ID", "身份証號"]
    hmap = build_header_map(sh_self_select, 1)
    id_col = find_id_col_by_content(sh_self_select, 1, find_column_exact(hmap, id_aliases))
    if id_col is None:
        print("自選名單 sheet 找不到 ID 欄，略過")
        return

    self_select_ids: set = {
        normalize_id(sh_self_select.cell(r, id_col).value)
        for r in range(2, sh_self_select.max_row + 1)
        if is_valid_tw_id(sh_self_select.cell(r, id_col).value)
    }

    src_col_map = _build_src_col_map(_SELF_SELECT_COL_MAP, cols)
    ws_out = wb_tpl[SELF_SELECT_SHEET_NAME]
    n = _copy_sheet_rows(
        ws_main, ws_out, _SELF_SELECT_COL_MAP, src_col_map,
        data_start, last_row,
        filter_ids=self_select_ids,
        id_main_col=cols.get("id"),
    )
    print(f"自選名單 sheet 已寫入 {n} 列")

# ============================================================
# 百分位名單 sheet（sheet2）
# ============================================================
PERCENTILE_SHEET_NAME = "百分位名單"
PINK_FILL = openpyxl.styles.PatternFill(fill_type="solid", fgColor="EAC0C0")
BLUE_FILL = openpyxl.styles.PatternFill(fill_type="solid", fgColor="A9C2D9")
NO_FILL = openpyxl.styles.PatternFill(fill_type=None)


def _copy_row_style(ws, src_row: int, dst_row: int, start_col: int = 1, end_col: int = 24) -> None:
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for c in range(start_col, end_col + 1):
        src = ws.cell(src_row, c)
        dst = ws.cell(dst_row, c)
        if src.number_format:
            dst.number_format = src.number_format
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)


def _cache_row_styles(ws, src_row: int, start_col: int, end_col: int) -> dict:
    """
    快取模板列每欄的 _style 索引與列高。
    openpyxl 內部以整數索引參照共享樣式表，直接複製索引比 copy() 物件快約 20 倍，
    且框線、底色、字型、格式均完整保留。
    """
    cache: dict = {
        "_height": ws.row_dimensions[src_row].height,
        "_styles": [ws.cell(src_row, c)._style for c in range(start_col, end_col + 1)],
        "_start":  start_col,
    }
    return cache


def _apply_cached_style(ws, dst_row: int, style_cache: dict, start_col: int, end_col: int) -> None:
    """直接寫入 _style 索引，不建立任何新物件，速度遠快於逐屬性 copy()。"""
    ws.row_dimensions[dst_row].height = style_cache["_height"]
    styles = style_cache["_styles"]
    base   = style_cache["_start"]
    for c in range(start_col, end_col + 1):
        ws.cell(dst_row, c)._style = styles[c - base]


def _clear_percentile_data_area(ws, start_row: int = 5, end_col: int = 24) -> None:
    for r in range(start_row, ws.max_row + 1):
        for c in range(1, end_col + 1):
            ws.cell(r, c).value = None


def _filter_followup_by_metric(text: Any, metric: str) -> Optional[str]:
    """
    AU/AV/AW 欄位可能同時包含 HbA1c 與 LDL 兩行（用 \\n 分隔）。
    百分位名單第二頁只顯示對應指標那一行，避免 LDL 側出現 HbA1c 文字或反之。
    """
    if text is None:
        return None
    prefix = "LDL:" if metric == "ldl" else "HbA1c:"
    lines = [ln for ln in str(text).splitlines() if ln.strip().startswith(prefix)]
    return "\n".join(lines) if lines else None


def _make_percentile_record(ws_main, row: int, cols: Dict[str, Optional[int]], metric: str, name_fill: str) -> Dict[str, Any]:
    value_col = cols["ldl"] if metric == "ldl" else cols["hba"]
    date_col = cols["ldl_dt"] if metric == "ldl" else cols["hba_dt"]

    raw_au = ws_main.cell(row, cols["au"]).value  # type: ignore[index]
    raw_av = ws_main.cell(row, cols["av"]).value  # type: ignore[index]
    raw_aw = ws_main.cell(row, cols["aw"]).value  # type: ignore[index]

    return {
        "row": row,
        "name": ws_main.cell(row, cols["name"]).value,  # type: ignore[index]
        "bday": ws_main.cell(row, cols["bday"]).value,  # type: ignore[index]
        "id": ws_main.cell(row, cols["id"]).value,      # type: ignore[index]
        "score": ws_main.cell(row, cols["score"]).value,  # type: ignore[index]
        "note": ws_main.cell(row, cols["note"]).value,    # type: ignore[index]
        "value": ws_main.cell(row, value_col).value if value_col else None,
        "last_dt": ws_main.cell(row, date_col).value if date_col else None,
        "au": _filter_followup_by_metric(raw_au, metric),
        "av": _filter_followup_by_metric(raw_av, metric),
        "aw": _filter_followup_by_metric(raw_aw, metric),
        "fill": name_fill,
    }


def _collect_side_records_sorted(
    ws_main,
    cols: Dict[str, Optional[int]],
    data_start: int,
    last_row: int,
    primary_mark_key: str,
    secondary_mark_key: str,
    metric: str,
) -> List[Dict[str, Any]]:
    primary_rows: List[int] = []
    secondary_rows: List[int] = []

    for row in range(data_start, last_row + 1):
        primary = normalize_text(ws_main.cell(row, cols[primary_mark_key]).value) if cols.get(primary_mark_key) else ""
        secondary = normalize_text(ws_main.cell(row, cols[secondary_mark_key]).value) if cols.get(secondary_mark_key) else ""

        if primary == "✔":
            primary_rows.append(row)
        elif secondary == "✔":
            secondary_rows.append(row)

    out: List[Dict[str, Any]] = []
    for row in primary_rows:
        out.append(_make_percentile_record(ws_main, row, cols, metric, "pink"))
    for row in secondary_rows:
        out.append(_make_percentile_record(ws_main, row, cols, metric, "blue"))
    return out


def _collect_percentile_records(ws_main, cols: Dict[str, Optional[int]], data_start: int, last_row: int) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    ldl_records = _collect_side_records_sorted(
        ws_main, cols, data_start, last_row,
        primary_mark_key="bb_mark", secondary_mark_key="bc_mark", metric="ldl"
    )
    hba_records = _collect_side_records_sorted(
        ws_main, cols, data_start, last_row,
        primary_mark_key="ay_mark", secondary_mark_key="az_mark", metric="hba"
    )
    return ldl_records, hba_records




def _ratio_numerator_text(value: Any) -> str:
    s = normalize_text(value)
    if not s:
        return "0"
    return s.split("/", 1)[0].strip() or "0"


def _set_percentile_title_rich(ws, cell_ref: str, prefix: str, red_num: Any, blue_num: Any) -> None:
    ws[cell_ref].value = f"{prefix}({_ratio_numerator_text(red_num)}/{_ratio_numerator_text(blue_num)}人)"


def _write_percentile_side(ws, row: int, start_col: int, rec: Optional[Dict[str, Any]]) -> None:
    cols_seq = [start_col + i for i in range(10)]
    name_cell = ws.cell(row, start_col)

    if not rec:
        for c in cols_seq:
            ws.cell(row, c).value = None
        name_cell.fill = copy(NO_FILL)
        return

    values = [
        rec.get("name"),
        rec.get("bday"),
        rec.get("id"),
        rec.get("score"),
        rec.get("note"),
        rec.get("value"),
        rec.get("last_dt"),
        rec.get("au"),
        rec.get("av"),
        rec.get("aw"),
    ]
    for c, v in zip(cols_seq, values):
        ws.cell(row, c).value = v

    if normalize_text(rec.get("name")):
        name_cell.fill = copy(PINK_FILL if rec.get("fill") == "pink" else BLUE_FILL)
    else:
        name_cell.fill = copy(NO_FILL)


def populate_percentile_sheet(
    wb_tpl,
    ws_main,
    cols: Dict[str, Optional[int]],
    data_start: int,
    last_row: int,
) -> None:
    if PERCENTILE_SHEET_NAME not in wb_tpl.sheetnames:
        ws = wb_tpl.create_sheet(PERCENTILE_SHEET_NAME)
    else:
        ws = wb_tpl[PERCENTILE_SHEET_NAME]

    ldl_records, hba_records = _collect_percentile_records(ws_main, cols, data_start, last_row)

    _set_percentile_title_rich(ws, "A1", "LDL百分位", ws_main[sc("ldl_main_summary")].value, ws_main[sc("ldl_target_summary")].value)
    _set_percentile_title_rich(ws, "N1", "HBA1C百分位", ws_main[sc("hba_main_summary")].value, ws_main[sc("hba_target_summary")].value)
    ws["A2"] = "紅色：達到標準、藍色：達到73.8%"
    ws["N2"] = "紅色：達到標準、藍色：達到73.8%"

    _clear_percentile_data_area(ws, start_row=5, end_col=24)

    data_rows = max(len(ldl_records), len(hba_records), 1)
    style_src_row = 5
    for idx in range(data_rows):
        row = 5 + idx
        if row != style_src_row:
            _copy_row_style(ws, style_src_row, row, 1, 24)
        ws.row_dimensions[row].height = ws.row_dimensions[style_src_row].height or 66
        _write_percentile_side(ws, row, 1, ldl_records[idx] if idx < len(ldl_records) else None)
        _write_percentile_side(ws, row, 14, hba_records[idx] if idx < len(hba_records) else None)

    for row in range(5, 5 + data_rows):
        for col in [2, 7, 15, 20]:
            cell = ws.cell(row, col)
            if isinstance(cell.value, (datetime.date, datetime.datetime)):
                cell.number_format = "yyyy-mm-dd"
    apply_full_grid(ws, ws.max_row, 24)


# ============================================================
# 全名單整合（v3 新增）
# ============================================================
def _id_from_sheet(sheet, header_aliases: List[str], search_rows: int = 10) -> Optional[int]:
    """從 sheet 找 ID 欄（靠 header + 內容驗證）"""
    header_row = _find_header_row_contains_any(sheet, [header_aliases], search_rows=search_rows)
    if header_row is None:
        header_row = 1
    hmap = build_header_map(sheet, header_row)
    cand = find_column_exact(hmap, header_aliases)
    return find_id_col_by_content(sheet, header_row, cand), header_row


def _extract_id_name_map(sheet, id_aliases: List[str]) -> Dict[str, str]:
    """從任意 sheet 抽出 {id: name} 對應表"""
    if sheet is None:
        return {}
    result: Dict[str, str] = {}
    col_info = _id_from_sheet(sheet, id_aliases)
    id_col, hrow = col_info
    if id_col is None:
        return {}
    # 找姓名欄
    hmap = build_header_map(sheet, hrow)
    name_col = find_column_exact(hmap, ["姓名", "會員姓名", "名字"])
    for r in range(hrow + 1, sheet.max_row + 1):
        pid = normalize_id(sheet.cell(r, id_col).value)
        if pid and is_valid_tw_id(pid):
            name = sheet.cell(r, name_col).value if name_col else None
            result.setdefault(pid, normalize_text(name) if name else "")
    return result


def _empty_member() -> Dict[str, Any]:
    """回傳一筆空白會員記錄（所有欄位為 None / 空字串）。"""
    return {"name": "", "bday": None, "e_code": None, "abc": None,
            "phone": None, "mobile": None, "address": None,
            "dmk_raw": None, "cnt": None, "clinic": ""}


def _fill_member_field(rec: Dict[str, Any], key: str, value: Any) -> None:
    """只在記錄中該欄位尚未有值時才填入（缺什麼補什麼）。"""
    if value is not None and value != "" and not rec.get(key):
        rec[key] = value


def collect_all_members(wb_src) -> Dict[str, Dict[str, Any]]:
    """
    整合所有 sheet 的人員名單，以身份證號為鍵，去重複。
    優先順序（高蓋低）：會員名單 > 其他名單 > 月份分頁
    各來源採「缺什麼補什麼」原則：已有值的欄位不會被空值覆蓋。
    回傳 {id: {name, bday, e_code, abc, phone, mobile, address, dmk_raw, cnt, clinic}}
    """
    members: Dict[str, Dict[str, Any]] = {}
    snames = wb_src.sheetnames
    id_aliases = ["身份證號", "身分證號", "身份証號", "身分証號",
                  "會員身份証", "會員身份證", "會員身分證", "ID", "家醫收案會員ID"]

    def _get_or_create(pid: str) -> Dict[str, Any]:
        if pid not in members:
            members[pid] = _empty_member()
        return members[pid]

    # ── 1. 月份分頁（最低優先）──────────────────────────────────
    month_sheets = sorted(
        [s for s in snames if _sheet_year_bucket(s) in (114, 115) and _sheet_month(s)],
        key=lambda s: (int(s[:3]), _sheet_month(s))
    )
    _month_header_row: Optional[int] = None
    _month_id_col:     Optional[int] = None
    _month_cols:       Dict[str, Optional[int]] = {}
    for sname in month_sheets:
        sh = wb_src[sname]
        if _month_id_col is None:
            _month_header_row = _find_header_row_contains_any(sh, [id_aliases], search_rows=5) or 1
            hmap = build_header_map(sh, _month_header_row)
            _month_id_col = find_id_col_by_content(sh, _month_header_row, find_column_exact(hmap, id_aliases))
            _month_cols = {
                "name":    find_column_exact(hmap, ["姓名", "會員姓名"]),
                "bday":    find_column_exact(hmap, ["生日", "出生日期", "出生年月日"]),
                "phone":   next((c for hdr, c in hmap.items() if is_phone_header(hdr)), None),
                "address": next((c for hdr, c in hmap.items() if is_address_header(hdr)), None),
            }
        if _month_id_col is None:
            continue
        for r in range(_month_header_row + 1, sh.max_row + 1):
            pid = normalize_id(sh.cell(r, _month_id_col).value)
            if not pid or not is_valid_tw_id(pid):
                continue
            rec = _get_or_create(pid)
            if _month_cols.get("name"):
                _fill_member_field(rec, "name", normalize_text(sh.cell(r, _month_cols["name"]).value))
            if _month_cols.get("bday"):
                _fill_member_field(rec, "bday", parse_date(sh.cell(r, _month_cols["bday"]).value))
            if _month_cols.get("phone"):
                contact = pick_contact_from_values([sh.cell(r, _month_cols["phone"]).value])
                _fill_member_field(rec, "phone",  contact.phone)
                _fill_member_field(rec, "mobile", contact.mobile)
            if _month_cols.get("address"):
                _fill_member_field(rec, "address", normalize_text(sh.cell(r, _month_cols["address"]).value) or None)

    # ── 2. ascvd / 自選名單 / 115X / P4P收案 / P4P追蹤（補充）──
    for sname in ["ascvd", "自選名單", "115X", "P4P收案", "P4P追蹤"]:
        if sname not in snames:
            continue
        sh = wb_src[sname]
        m = _extract_id_name_map(sh, id_aliases)
        for pid, name in m.items():
            rec = _get_or_create(pid)
            _fill_member_field(rec, "name", name)

    # ── 3. 會員名單（最高優先）──────────────────────────────────
    if "會員名單" in snames:
        sh = wb_src["會員名單"]
        header_row = _find_header_row_contains_any(sh, [id_aliases], search_rows=10) or 5
        hmap = build_header_map(sh, header_row)
        id_col_cand = find_column_exact(hmap, id_aliases)
        id_col   = find_id_col_by_content(sh, header_row, id_col_cand)
        name_col = find_column_exact(hmap, ["會員姓名", "姓名"])
        bday_col = find_column_exact(hmap, ["會員生日", "生日"])
        phone_cols = [c for hdr, c in hmap.items() if c != id_col and is_phone_header(hdr)]
        addr_col = next((c for hdr, c in hmap.items() if c != id_col and is_address_header(hdr)), None)
        abc_col  = find_column_exact(hmap, ["會員別"])
        dmk_col  = find_column_exact(hmap, ["疾病樣態"])
        cnt_col  = find_column_exact(hmap, ["就診次數"])
        clinic_val = normalize_text(sh.cell(1, 1).value)
        if id_col:
            for r in range(header_row + 1, sh.max_row + 1):
                pid = normalize_id(sh.cell(r, id_col).value)
                if not pid or not is_valid_tw_id(pid):
                    continue
                rec = _get_or_create(pid)
                contact = pick_contact_from_values([sh.cell(r, c).value for c in phone_cols])
                address = normalize_text(sh.cell(r, addr_col).value) if addr_col else ""
                # 會員名單的值直接覆蓋（最高優先）；空值則不蓋掉已有資料
                if name_col:
                    _fill_member_field(rec, "name",    normalize_text(sh.cell(r, name_col).value))
                if bday_col:
                    _fill_member_field(rec, "bday",    parse_date(sh.cell(r, bday_col).value))
                if abc_col:
                    _fill_member_field(rec, "abc",     sh.cell(r, abc_col).value)
                if dmk_col:
                    _fill_member_field(rec, "e_code",  parse_disease_code(sh.cell(r, dmk_col).value))
                    _fill_member_field(rec, "dmk_raw", sh.cell(r, dmk_col).value)
                if cnt_col:
                    _fill_member_field(rec, "cnt",     sh.cell(r, cnt_col).value)
                _fill_member_field(rec, "phone",   contact.phone)
                _fill_member_field(rec, "mobile",  contact.mobile)
                _fill_member_field(rec, "address", address or None)
                _fill_member_field(rec, "clinic",  clinic_val)

    return members


# ============================================================
# 聯絡資料整合（手機 / 電話 / 地址）
# ============================================================
def build_contact_map(sh_phone: Any) -> Dict[str, ContactInfo]:
    """從聯絡資料 sheet 建立 {id: ContactInfo} 對應。"""
    if sh_phone is None:
        return {}
    id_aliases = ["ID", "身份證號", "身分證號", "身份証號"]
    header_row = _find_header_row_contains_any(sh_phone, [id_aliases], search_rows=5)
    if header_row is None:
        header_row = 1
    hmap = build_header_map(sh_phone, header_row)
    id_col_cand = find_column_exact(hmap, id_aliases)
    id_col = find_id_col_by_content(sh_phone, header_row, id_col_cand)
    phone_cols = [
        c for hdr, c in hmap.items()
        if c != id_col and is_phone_header(hdr)
    ]
    addr_col = next(
        (c for hdr, c in hmap.items() if c != id_col and is_address_header(hdr)),
        None,
    )
    if id_col is None or (not phone_cols and addr_col is None):
        return {}
    result: Dict[str, ContactInfo] = {}
    for r in range(header_row + 1, sh_phone.max_row + 1):
        pid = normalize_id(sh_phone.cell(r, id_col).value)
        if not pid or not is_valid_tw_id(pid):
            continue
        contact = pick_contact_from_values([sh_phone.cell(r, c).value for c in phone_cols])
        address = normalize_text(sh_phone.cell(r, addr_col).value) if addr_col else ""
        if contact.phone or contact.mobile or address:
            result[pid] = ContactInfo(
                phone=contact.phone,
                mobile=contact.mobile,
                address=address or None,
            )
    return result


# ============================================================
# P4P / 自選 / 114 / 115X 旗標
# ============================================================
def build_p4p_map(sh_enroll: Any, sh_track: Any) -> Dict[str, Dict[str, Any]]:
    """
    合併 P4P收案 + P4P追蹤，回傳 {id: {status, enroll_dt, last_track_dt, next_track_dt}}
    """
    result: Dict[str, Dict[str, Any]] = {}
    id_aliases = ["家醫收案會員ID", "ID", "身份證號", "身分證號"]

    # P4P收案：取收案狀態
    if sh_enroll is not None:
        hmap = build_header_map(sh_enroll, 1)
        id_col = find_id_col_by_content(sh_enroll, 1, find_column_exact(hmap, id_aliases))
        status_col = find_column_exact(hmap, ["收案狀態"])
        if id_col:
            for r in range(2, sh_enroll.max_row + 1):
                pid = normalize_id(sh_enroll.cell(r, id_col).value)
                if not pid or not is_valid_tw_id(pid):
                    continue
                status = normalize_text(sh_enroll.cell(r, status_col).value) if status_col else ""
                result.setdefault(pid, {})["status"] = status or None

    # P4P追蹤：取日期欄
    if sh_track is not None:
        hmap = build_header_map(sh_track, 1)
        id_col = find_id_col_by_content(sh_track, 1, find_column_exact(hmap, id_aliases))
        enroll_col     = find_column_exact(hmap, ["收案日期"])
        last_col       = find_column_exact(hmap, ["最後追蹤日"])
        next_col       = find_column_exact(hmap, ["下次應追蹤日"])
        status_col     = find_column_exact(hmap, ["收案狀態"])
        if id_col:
            for r in range(2, sh_track.max_row + 1):
                pid = normalize_id(sh_track.cell(r, id_col).value)
                if not pid or not is_valid_tw_id(pid):
                    continue
                d = result.setdefault(pid, {})
                if status_col:
                    s = normalize_text(sh_track.cell(r, status_col).value)
                    if s:
                        d["status"] = s
                if enroll_col:
                    d["enroll_dt"]     = parse_date(sh_track.cell(r, enroll_col).value)
                if last_col:
                    d["last_track_dt"] = parse_date(sh_track.cell(r, last_col).value)
                if next_col:
                    d["next_track_dt"] = parse_date(sh_track.cell(r, next_col).value)

    return result


def build_id_set(sheet: Any, id_aliases: List[str]) -> set:
    """從 sheet 建立 ID set（自動偵測 header row）"""
    if sheet is None:
        return set()
    # 先嘗試找含有 ID 關鍵字的 header row
    header_row = _find_header_row_contains_any(sheet, [id_aliases], search_rows=10)
    if header_row is None:
        header_row = 1
    hmap = build_header_map(sheet, header_row)
    id_col = find_id_col_by_content(sheet, header_row, find_column_exact(hmap, id_aliases))
    if id_col is None:
        return set()
    result = set()
    for r in range(header_row + 1, sheet.max_row + 1):
        pid = normalize_id(sheet.cell(r, id_col).value)
        if pid and is_valid_tw_id(pid):
            result.add(pid)
    return result


def _fill_extra_flags(
    ws,
    cols: Dict[str, Optional[int]],
    data_start: int,
    last_row: int,
    id_to_rows: Dict[str, List[int]],
    p4p_map: Dict[str, Dict[str, Any]],
    ascvd_ids: set,
    self_select_ids: set,
    x115_ids: set,
) -> None:
    """填入後段旗標欄位。"""
    col_at = cols.get("p4p_status")
    col_au = cols.get("p4p_enroll_dt")
    col_av = cols.get("p4p_last_dt")
    col_aw = cols.get("p4p_next_dt")
    col_ax = cols.get("is_114")
    col_ay = cols.get("is_self_select")
    col_az = cols.get("is_115x")

    for pid, rows in id_to_rows.items():
        p4p = p4p_map.get(pid, {})
        is_114       = "v" if pid in ascvd_ids       else None
        is_self      = "v" if pid in self_select_ids  else None
        is_115x      = "v" if pid in x115_ids         else None
        status       = p4p.get("status") or None
        enroll_dt    = p4p.get("enroll_dt")
        last_dt      = p4p.get("last_track_dt")
        next_dt      = p4p.get("next_track_dt")

        for rr in rows:
            safe_set(ws, rr, col_at, status)
            safe_set(ws, rr, col_au, enroll_dt)
            safe_set(ws, rr, col_av, last_dt)
            safe_set(ws, rr, col_aw, next_dt)
            safe_set_check(ws, rr, col_ax, is_114)
            safe_set_check(ws, rr, col_ay, is_self)
            safe_set_check(ws, rr, col_az, is_115x)

# ============================================================
# 資料填充子函數
# ============================================================
def _load_and_validate_source(source_path: str):
    wb = openpyxl.load_workbook(source_path, data_only=True)
    need = ["HealthCase", "成人健檢", "子宮抹片", "老人流感", "糞便潛血", "肝炎篩檢"]
    missing = [s for s in need if s not in wb.sheetnames]
    if missing:
        raise ValueError(f"原始檔缺少工作表：{'、'.join(missing)}")
    return wb


def _fill_member_basic(
    ws,
    all_members: Dict[str, Dict[str, Any]],
    contact_map: Dict[str, ContactInfo],
    cols:        Dict[str, Optional[int]],
    data_start:  int,
    now:         datetime.date,
    clinic_val:  str = "",
) -> Tuple[Dict[str, List[int]], Dict[int, MemberMeta], int]:
    """
    v3：從 all_members（已整合所有 sheet）逐一寫入總表。
    聯絡資料以 contact_map 優先，其次用 all_members 內已整理好的電話/手機/地址。
    09 開頭寫入手機欄，其餘寫入電話欄；地址寫入隱藏輔助欄。
    """
    id_to_rows: Dict[str, List[int]] = {}
    meta: Dict[int, MemberMeta] = {}
    out_r = data_start
    wrote_any = False

    _max_col = ws.max_column

    # 欄號預先解出，避免每列重複 dict.get()
    _c_clinic  = cols.get("clinic")
    _c_name    = cols.get("name")
    _c_id      = cols.get("id")
    _c_bday    = cols.get("bday")
    _c_age     = cols.get("age")
    _c_tel     = cols.get("tel")
    _c_mobile  = cols.get("mobile")
    _c_abc     = cols.get("abc")
    _c_dmk     = cols.get("dmk_code")
    _c_cnt     = cols.get("cnt")
    _c_sex     = cols.get("sex")
    _c_addr    = cols.get("address_hidden")

    for pid, info in all_members.items():
        name   = info.get("name") or ""
        bday   = info.get("bday")
        e_code = info.get("e_code")
        abc    = info.get("abc")
        base_contact = ContactInfo(
            phone=info.get("phone"),
            mobile=info.get("mobile"),
            address=info.get("address"),
        )
        contact = merge_contact_info(contact_map.get(pid), base_contact)
        cnt    = info.get("cnt")
        clinic = info.get("clinic") or clinic_val
        age    = calc_age(bday, now) if isinstance(bday, datetime.date) else -1
        sex    = infer_gender_from_id(pid)

        safe_set(ws, out_r, _c_clinic,  clinic)
        safe_set(ws, out_r, _c_name,    name or None)
        safe_set(ws, out_r, _c_id,      pid)
        safe_set(ws, out_r, _c_bday,    bday)
        safe_set(ws, out_r, _c_age,     age if age >= 0 else None)
        safe_set(ws, out_r, _c_tel,     contact.phone or None)
        safe_set(ws, out_r, _c_mobile,  contact.mobile or None)
        safe_set(ws, out_r, _c_abc,     abc)
        safe_set(ws, out_r, _c_dmk,     e_code.value if e_code else None)
        safe_set(ws, out_r, _c_cnt,     cnt)
        safe_set(ws, out_r, _c_sex,     sex)
        safe_set(ws, out_r, _c_addr,    contact.address or None)

        meta[out_r] = MemberMeta(row=out_r, bday=bday, age=age, e_code=e_code)
        id_to_rows.setdefault(pid, []).append(out_r)

        out_r += 1
        wrote_any = True

    last_row = out_r - 1 if wrote_any else (data_start - 1)

    # 全部資料寫完後，一次性套用格式（比逐列套用少一半 cell 存取）
    for r in range(data_start, last_row + 1):
        _apply_member_row_style(ws, r, _max_col)

    return id_to_rows, meta, last_row


def _fill_ascvd(
    ws,
    sh_ascvd:  Any,
    cols:      Dict[str, Optional[int]],
    id_to_rows: Dict[str, List[int]],
    meta:      Dict[int, MemberMeta],
) -> None:
    # 動態找 header row（含 ID 和 ASCVD 的列）
    ASCVD_HEADER_ROW = _find_header_row_contains_any(
        sh_ascvd, [["ID", "id"], ["ASCVD", "ascvd"]], search_rows=10
    ) or 5
    amap  = build_header_map(sh_ascvd, ASCVD_HEADER_ROW)
    # 最後就診日可能在 header_row-1（上一列）
    amap_upper = build_header_map(sh_ascvd, ASCVD_HEADER_ROW - 1) if ASCVD_HEADER_ROW > 1 else {}
    amap_merged = {**amap_upper, **amap}
    a_id  = find_column_exact(amap, ["ID", "id"])
    a_asc = find_column_exact(amap, ["ASCVD", "ascvd"])
    a_lv  = find_column_exact(amap_merged, ["最後就診日"])
    if a_id is None or a_asc is None:
        raise ValueError("原始檔「ascvd」找不到 ID / ASCVD 欄位")

    for r in range(ASCVD_HEADER_ROW + 1, sh_ascvd.max_row + 1):
        pid = normalize_id(sh_ascvd.cell(r, a_id).value)
        val = sh_ascvd.cell(r, a_asc).value
        if not pid:
            continue
        if val is None or str(val).strip() in ("",):  # 0 也要保留寫入
            continue

        rows = id_to_rows.get(pid)
        if not rows:
            continue

        for tr in rows:
            safe_set(ws, tr, cols.get("ascvd"), val)
            if tr in meta:
                meta[tr].ascvd = parse_ascvd(val)
            # 最後就診日（ascvd T欄，優先填入）
            if a_lv:
                lv_dt = parse_date(sh_ascvd.cell(r, a_lv).value)
                if lv_dt:
                    existing = parse_date(ws.cell(tr, cols.get("last_visit") or 0).value) if cols.get("last_visit") else None
                    if existing is None or lv_dt > existing:
                        safe_set(ws, tr, cols.get("last_visit"), lv_dt)


def _fill_screening(
    ws,
    sheet,
    target_col: Optional[int],
    id_to_rows: Dict[str, List[int]],
) -> None:
    if not target_col:
        return
    hmap    = build_header_map(sheet, 1)
    sid_col = find_column_exact(hmap, ["ID", "身分證號", "身份證號"])
    dt_col  = find_column_exact(hmap, ["最後篩檢日期"])
    if sid_col is None or dt_col is None:
        raise ValueError(f"「{sheet.title}」找不到 ID 或 最後篩檢日期 欄位")

    for rr in range(2, sheet.max_row + 1):
        pid = normalize_id(sheet.cell(rr, sid_col).value)
        dt  = parse_date(sheet.cell(rr, dt_col).value)
        if not pid or dt is None:
            continue

        rows = id_to_rows.get(pid)
        if not rows:
            continue

        for tr in rows:
            existing = parse_date(ws.cell(tr, target_col).value)
            if existing is None or dt > existing:
                ws.cell(tr, target_col).value = dt


def _fill_health_case(
    ws,
    sh_health: Any,
    cols:      Dict[str, Optional[int]],
    id_to_rows: Dict[str, List[int]],
) -> None:
    hmap = build_header_map(sh_health, 1)
    field_aliases = {
        "hc_id":      ["家醫收案會員ID", "ID"],
        "hc_hba":     ["最近一次HbA1c檢查結果(%)"],
        "hc_hba_dt":  ["最近一次HbA1c檢查日期"],
        "hc_ldl":     ["最近一次LDL檢查結果(mg/dL)"],
        "hc_ldl_dt":  ["最近一次LDL檢查日期"],
        "hc_uacr":    ["最近一次UACR檢查結果(mg/gm)"],
        "hc_uacr_dt": ["最近一次UACR檢查日期"],
    }
    fc = {k: find_column_exact(hmap, v) for k, v in field_aliases.items()}
    if any(v is None for v in fc.values()):
        raise ValueError(
            "原始檔「HealthCase」欄位不完整"
            "（家醫收案會員ID / HbA1c結果+日期 / LDL結果+日期 / UACR結果+日期）"
        )

    for r in range(2, sh_health.max_row + 1):
        pid = normalize_id(sh_health.cell(r, fc["hc_id"]).value)
        if not pid:
            continue

        rows = id_to_rows.get(pid)
        if not rows:
            continue

        for tr in rows:
            for val_key, dt_key, col_val_key, col_dt_key in [
                ("hc_hba",  "hc_hba_dt",  "hba",  "hba_dt"),
                ("hc_ldl",  "hc_ldl_dt",  "ldl",  "ldl_dt"),
                ("hc_uacr", "hc_uacr_dt", "uacr", "uacr_dt"),
            ]:
                v  = parse_float(sh_health.cell(r, fc[val_key]).value)
                dt = parse_date(sh_health.cell(r, fc[dt_key]).value)
                # 以日期較新者優先：避免重複 ID 時舊資料蓋掉新資料
                existing_dt = parse_date(ws.cell(tr, cols[col_dt_key]).value) if cols.get(col_dt_key) else None
                if dt and (existing_dt is None or dt > existing_dt):
                    if v is not None and v != 0:
                        safe_set(ws, tr, cols.get(col_val_key), v)
                    safe_set(ws, tr, cols.get(col_dt_key), dt)


def _compute_all_derived(
    ws,
    cols:       Dict[str, Optional[int]],
    meta:       Dict[int, MemberMeta],
    data_start: int,
    last_row:   int,
    now:        datetime.date,
    kpi_marks:  Optional[Dict[str, set]] = None,
) -> None:
    kpi_marks = kpi_marks or {}

    # 預取欄號（loop 外一次取出，避免每列重複 dict lookup）
    _c_dmk    = cols["dmk_code"]
    _c_ascvd  = cols["ascvd"]
    _c_sex    = cols["sex"]
    _c_adult  = cols.get("adult")
    _c_pap    = cols.get("pap")
    _c_flu    = cols.get("flu")
    _c_fit    = cols.get("fit")
    _c_hep    = cols.get("hep")
    _c_hba    = cols["hba"]
    _c_hba_dt = cols["hba_dt"]
    _c_ldl    = cols["ldl"]
    _c_ldl_dt = cols["ldl_dt"]
    _c_n115   = cols.get("n_count_115")

    def _get_dt_col(row: int, col: Optional[int]) -> Optional[datetime.date]:
        return parse_date(ws.cell(row, col).value) if col else None

    for rr in range(data_start, last_row + 1):
        e_code    = parse_disease_code(ws.cell(rr, _c_dmk).value)
        ascvd     = parse_ascvd(ws.cell(rr, _c_ascvd).value)
        ascvd_raw = ws.cell(rr, _c_ascvd).value

        m   = meta.get(rr, MemberMeta(row=rr))
        age = m.age if isinstance(m.age, int) else -1
        sex = normalize_text(ws.cell(rr, _c_sex).value)

        adult_dt = _get_dt_col(rr, _c_adult)
        pap_dt   = _get_dt_col(rr, _c_pap)
        flu_dt   = _get_dt_col(rr, _c_flu)
        fit_dt   = _get_dt_col(rr, _c_fit)
        hep_dt   = _get_dt_col(rr, _c_hep)
        hba_dt   = _get_dt_col(rr, _c_hba_dt)
        ldl_dt   = _get_dt_col(rr, _c_ldl_dt)

        hba_val = ws.cell(rr, _c_hba).value
        ldl_val = ws.cell(rr, _c_ldl).value

        safe_set(ws, rr, cols.get("disease_text"),
                 disease_group_text(e_code, ascvd))

        for _mark_key in ("ay", "az", "bb", "bc"):
            safe_set_check(ws, rr, cols.get(f"{_mark_key}_mark"),
                           "✔" if rr in kpi_marks.get(_mark_key, set()) else None)

        # AK：✔
        safe_set_check(ws, rr, cols.get("ak"),
                       "✔" if should_check_ak(e_code=e_code, ascvd=ascvd, age=age, hba_val=hba_val) else None)

        # AX：漏檢項目
        ax_txt = build_ax_leak_item(
            d_code=e_code,
            ascvd_raw=ascvd_raw,
            hba_dt=hba_dt,
            ldl_dt=ldl_dt,
        )
        safe_set(ws, rr, cols.get("ax"), ax_txt or None)

        # 年齡未知時不產生篩檢備註，避免以 0 歲誤判
        note = build_screening_note(
            age=age, sex=sex,
            hep_dt=hep_dt, fit_dt=fit_dt, pap_dt=pap_dt,
            adult_dt=adult_dt, flu_dt=flu_dt, today=now,
        ) if age >= 0 else ""
        safe_set(ws, rr, cols.get("note"), note or None)

        visit_count_115 = ws.cell(rr, _c_n115).value if _c_n115 else None

        score, breakdown = calc_score(
            hba_val=hba_val, hba_dt=hba_dt,
            ldl_val=ldl_val, ldl_dt=ldl_dt,
            adult_dt=adult_dt, pap_dt=pap_dt,
            flu_dt=flu_dt, fit_dt=fit_dt, hep_dt=hep_dt,
            age=age, sex=sex,
            visit_count_115=visit_count_115,
        )
        safe_set(ws, rr, cols.get("score"),     score)
        safe_set(ws, rr, cols.get("breakdown"), breakdown)

        au_txt = build_au_note(
            e_code=e_code, ascvd=ascvd,
            hba_dt=hba_dt, ldl_dt=ldl_dt, today=now,
        )
        safe_set(ws, rr, cols.get("au"), au_txt or None)

        if au_txt and not _should_skip_followup(au_txt):
            safe_set(ws, rr, cols.get("av"),
                     build_followup_note(au_txt, now, Rules.AV_OFFSET_DAYS))
            safe_set(ws, rr, cols.get("aw"),
                     build_followup_note(au_txt, now, Rules.AW_OFFSET_DAYS))
        else:
            safe_set(ws, rr, cols.get("av"), None)
            safe_set(ws, rr, cols.get("aw"), None)







# ============================================================
# 清空資料列
# ============================================================
def _clear_data_rows(
    ws, data_start: int, max_row: int, cols: Dict[str, Optional[int]]
) -> None:
    clear_keys = [
        "clinic", "name", "id", "bday", "age", "tel", "mobile", "cnt", "sex", "abc",
        "dmk_code", "ascvd", "main_sub_dx",
        "adult", "pap", "flu", "fit", "hep",
        "hba", "hba_dt", "ldl", "ldl_dt", "uacr", "uacr_dt",
        "disease_text", "ay_mark", "az_mark", "ak", "ax", "score", "breakdown", "note",
        "bb_mark", "bc_mark", "au", "av", "aw",
        "m_count_114", "m_count_114_full", "n_count_115",
        "r_amount_114", "s_amount_115", "r_amount_114_total", "s_amount_115_total",
        "p4p_status", "p4p_enroll_dt", "p4p_last_dt", "p4p_next_dt",
        "is_114", "is_self_select", "is_115x", "address_hidden",
        "last_visit", "dx_raw",
    ]
    col_ids = [cols[k] for k in clear_keys if cols.get(k)]
    for r in range(data_start, max_row + 1):
        for c in col_ids:
            ws.cell(r, c).value = None  # type: ignore[arg-type]


def _first_sheet(wb, *names: str):
    for name in names:
        if name in wb.sheetnames:
            return wb[name]
    return None


def _require_sheet(wb, *names: str):
    sheet = _first_sheet(wb, *names)
    if sheet is None:
        raise ValueError(f"原始檔缺少工作表：{' / '.join(names)}")
    return sheet


def prepare_template_layout(ws) -> None:
    """補齊 0322 樣板缺少但程式仍需使用的欄位/輔助欄，並先隱藏後段輔助欄。"""
    ws["AW1"] = ws["AW1"].value or "備註"
    ws["AW2"] = ws["AW2"].value or ""
    ws["AX1"] = ws["AX1"].value or "是否為114會員名單"
    ws["AY1"] = ws["AY1"].value or "是否為自選會員"
    ws["AZ1"] = ws["AZ1"].value or "是否為115X"
    ws["BB1"] = ws["BB1"].value or "114年全年就診次數"
    ws["BC1"] = ws["BC1"].value or "114年申報總金額"
    ws["BD1"] = ws["BD1"].value or "115年申報總金額"
    ws["BE1"] = ws["BE1"].value or "地址"
    for addr in ("AX2", "AY2", "AZ2", "BB2", "BC2", "BD2", "BE2"):
        ws[addr] = ws[addr].value or ""
    for col in range(column_index_from_string("AX"), column_index_from_string("BE") + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].hidden = True


# ============================================================
# 主流程
# ============================================================
# ============================================================
# 結構化流程（V10.3）
# ============================================================
def load_source(source_path: str) -> SourceContext:
    wb_src = _load_and_validate_source(source_path)
    return SourceContext(
        wb_src=wb_src,
        sh_member=_first_sheet(wb_src, "會員名單"),
        sh_ascvd=_first_sheet(wb_src, "ascvd"),
        sh_health=_require_sheet(wb_src, "HealthCase"),
        sh_main_sub_dx=_first_sheet(wb_src, "主次診斷"),
        sh_phone=_first_sheet(wb_src, "行動電話"),
        sh_self_select=_first_sheet(wb_src, "自選會員", "自選名單"),
        sh_115x=_first_sheet(wb_src, "115X"),
        sh_p4p_enroll=_first_sheet(wb_src, "P4P收案"),
        sh_p4p_track=_first_sheet(wb_src, "P4P追蹤"),
        screening_sheets={
            "adult": _require_sheet(wb_src, "成人健檢"),
            "pap":   _require_sheet(wb_src, "子宮抹片"),
            "flu":   _require_sheet(wb_src, "老人流感"),
            "fit":   _require_sheet(wb_src, "糞便潛血"),
            "hep":   _require_sheet(wb_src, "肝炎篩檢"),
        },
        claim_sums=collect_monthly_claim_summaries(wb_src),
        all_members=collect_all_members(wb_src),
    )


def load_template(template_path: str) -> TemplateContext:
    wb_tpl = openpyxl.load_workbook(template_path)
    if Rules.SHEET_TARGET not in wb_tpl.sheetnames:
        raise ValueError(f"模板檔缺少工作表：{Rules.SHEET_TARGET}")

    ws = wb_tpl[Rules.SHEET_TARGET]
    prepare_template_layout(ws)
    data_start = Rules.DATA_START_ROW
    cols = detect_template_columns(ws, data_start)
    _clear_data_rows(ws, data_start, ws.max_row, cols)

    return TemplateContext(
        wb_tpl=wb_tpl,
        ws=ws,
        cols=cols,
        data_start=data_start,
    )


def fill_basic_data(source_ctx: SourceContext, template_ctx: TemplateContext, now: datetime.date) -> RuntimeContext:
    contact_map = build_contact_map(source_ctx.sh_phone)
    clinic_val = ""
    if source_ctx.sh_member is not None:
        clinic_val = normalize_text(source_ctx.sh_member.cell(1, 1).value)

    id_to_rows, meta, last_row = _fill_member_basic(
        template_ctx.ws,
        source_ctx.all_members,
        contact_map,
        template_ctx.cols,
        template_ctx.data_start,
        now,
        clinic_val=clinic_val,
    )
    return RuntimeContext(
        id_to_rows=id_to_rows,
        meta=meta,
        last_row=last_row,
    )


def fill_external_data(source_ctx: SourceContext, template_ctx: TemplateContext, runtime_ctx: RuntimeContext) -> None:
    ws = template_ctx.ws
    cols = template_ctx.cols
    data_start = template_ctx.data_start
    last_row = runtime_ctx.last_row
    id_to_rows = runtime_ctx.id_to_rows
    meta = runtime_ctx.meta

    if source_ctx.sh_ascvd:
        _log("套用 ascvd 資料")
        _fill_ascvd(ws, source_ctx.sh_ascvd, cols, id_to_rows, meta)
    _log("套用主次診斷")
    _fill_main_sub_dx(ws, source_ctx.sh_main_sub_dx, cols, id_to_rows)
    _sheet_label = {"adult": "成人健檢", "pap": "子宮抹片", "flu": "老人流感", "fit": "糞便潛血", "hep": "肝炎篩檢"}
    for key, sheet in source_ctx.screening_sheets.items():
        _log(f"套用篩檢資料：{_sheet_label.get(key, key)}")
        _fill_screening(ws, sheet, cols.get(key), id_to_rows)
    _log("套用 HealthCase 檢驗資料")
    _fill_health_case(ws, source_ctx.sh_health, cols, id_to_rows)
    _log("回填月份申報統計")
    fill_monthly_claim_summary_columns(ws, data_start, last_row, cols, source_ctx.claim_sums)

    _log("回填旗標與 P4P 狀態")
    id_aliases = ["身份證號", "身分證號", "ID", "家醫收案會員ID"]
    ascvd_ids = build_id_set(source_ctx.sh_ascvd, id_aliases)
    self_select_ids = build_id_set(source_ctx.sh_self_select, id_aliases)
    x115_ids = build_id_set(source_ctx.sh_115x, id_aliases)
    p4p_map = build_p4p_map(source_ctx.sh_p4p_enroll, source_ctx.sh_p4p_track)
    _fill_extra_flags(
        ws, cols, data_start, last_row, id_to_rows,
        p4p_map, ascvd_ids, self_select_ids, x115_ids,
    )


def compute_derived(template_ctx: TemplateContext, runtime_ctx: RuntimeContext, now: datetime.date) -> None:
    ws = template_ctx.ws
    cols = template_ctx.cols
    data_start = template_ctx.data_start
    last_row = runtime_ctx.last_row

    hba_candidates = _collect_hba_candidates(ws, cols, data_start, last_row)
    ldl_candidates = _collect_ldl_candidates(ws, cols, data_start, last_row)
    kpi_marks = collect_kpi_mark_sets(
        ws, cols, data_start, last_row,
        hba_candidates=hba_candidates,
        ldl_candidates=ldl_candidates,
    )

    _compute_all_derived(ws, cols, runtime_ctx.meta, data_start, last_row, now, kpi_marks)
    apply_date_format(ws, cols, data_start, last_row)
    apply_amount_format(ws, cols, data_start, last_row)

    runtime_ctx.hba_candidates = hba_candidates
    runtime_ctx.ldl_candidates = ldl_candidates


def compute_kpis(
    template_ctx: TemplateContext,
    runtime_ctx: RuntimeContext,
    source_ctx: Optional[SourceContext] = None,
) -> None:
    ws = template_ctx.ws
    cols = template_ctx.cols
    data_start = template_ctx.data_start
    last_row = runtime_ctx.last_row

    _log("產生 KPI 摘要與附表")
    calc_hba_kpi_ay_az(
        ws, cols, data_start, last_row,
        hba_candidates=runtime_ctx.hba_candidates,
    )
    calc_ldl_percentiles(
        ws, cols, data_start, last_row,
        ldl_candidates=runtime_ctx.ldl_candidates,
    )
    _log("產生百分位名單")
    populate_percentile_sheet(template_ctx.wb_tpl, ws, cols, data_start, last_row)
    _log("產生醫生看工作表")
    populate_doctor_sheet(template_ctx.wb_tpl, ws, cols, data_start, last_row)
    if source_ctx:
        _log("產生自選名單工作表")
        populate_self_select_sheet(
            template_ctx.wb_tpl,
            ws,
            cols,
            data_start,
            last_row,
            source_ctx.sh_self_select,
        )


def finalize_and_save(source_path: str, template_ctx: TemplateContext, now_dt: datetime.datetime) -> str:
    wb_tpl = template_ctx.wb_tpl
    base_dir = os.path.dirname(os.path.abspath(source_path))

    for sht_name in (Rules.SHEET_TARGET, PERCENTILE_SHEET_NAME):
        if sht_name in wb_tpl.sheetnames:
            wb_tpl[sht_name].sheet_view.showGridLines = True

    out_path = os.path.join(base_dir, f"選會員{now_dt.strftime('%m%d_%H%M%S')}.xlsx")
    wb_tpl.save(out_path)
    return out_path


def _log(msg: str) -> None:
    """輸出帶台灣時間時間戳的進度訊息，並立即 flush 到 terminal。"""
    tz_tw = datetime.timezone(datetime.timedelta(hours=8))
    ts = datetime.datetime.now(tz_tw).strftime("%H:%M:%S")
    print(f"[{ts}] {msg}", flush=True)


def process_excel(source_path: str, template_path: str) -> str:
    tz_tw = datetime.timezone(datetime.timedelta(hours=8))
    now_dt = datetime.datetime.now(tz_tw)
    now = now_dt.date()

    _log("開始處理 Excel")

    _log(f"載入原始檔：{os.path.basename(source_path)}")
    source_ctx = load_source(source_path)

    _log(f"載入模板：{os.path.basename(template_path)}")
    template_ctx = load_template(template_path)
    _log("偵測模板欄位並清空舊資料")

    _log("整理聯絡資料與建立會員主表")
    runtime_ctx = fill_basic_data(source_ctx, template_ctx, now)

    if runtime_ctx.last_row < template_ctx.data_start:
        _log("無會員資料，直接輸出")
        return finalize_and_save(source_path, template_ctx, now_dt)

    _log("回填外部資料：ASCVD / 主次診斷 / 篩檢 / HealthCase / 月份統計")
    fill_external_data(source_ctx, template_ctx, runtime_ctx)

    _log("計算分數、追蹤提醒與衍生欄位")
    compute_derived(template_ctx, runtime_ctx, now)

    _log("計算 KPI 標記名單")
    compute_kpis(template_ctx, runtime_ctx, source_ctx)

    _log("寫入輸出檔案")
    out = finalize_and_save(source_path, template_ctx, now_dt)
    _log(f"完成輸出：{os.path.basename(out)}")
    return out


# ============================================================
# 進入點
# ============================================================
def main() -> None:
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()

    src = filedialog.askopenfilename(
        title="選擇原始 Excel 檔案（沒有會員指標的那份）",
        filetypes=[("Excel files", "*.xlsx")],
    )
    if not src:
        return

    script_dir = os.path.dirname(os.path.abspath(__file__))
    template   = os.path.join(script_dir, Rules.TEMPLATE_NAME)

    if not os.path.exists(template):
        messagebox.showerror(
            "錯誤",
            f"找不到模板檔：\n{template}\n\n"
            f"請把模板放到此 .py 同資料夾，檔名需為：\n{Rules.TEMPLATE_NAME}",
        )
        return

    _log("已選擇原始檔，開始執行")
    try:
        out = process_excel(src, template)
        messagebox.showinfo("完成", f"已輸出：\n{out}")
        open_file_cross_platform(out)
    except (ValueError, KeyError, OSError, InvalidFileException) as e:
        messagebox.showerror("錯誤", str(e))
    except Exception as e:
        traceback.print_exc()
        messagebox.showerror("錯誤", f"未預期錯誤：{e}")


if __name__ == "__main__":
    main()
