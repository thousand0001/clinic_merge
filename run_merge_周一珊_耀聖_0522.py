#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
周一珊耀聖專用總表產檔。

資料來源很少：
- 家115.xls：姓名 / 電話 / 地址
- 次數/*.xls：病歷號 / 姓名 / 生日 / 電話 / 次數 / 地址

輸出只保留原會員總表中有資料可填的相關欄位，刪除篩檢、費用、
P4P、疾病樣態、分數等本資料夾無來源的欄位。
"""

from __future__ import annotations

import argparse
import datetime as dt
import os
import re
import subprocess
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, Optional, Sequence

import openpyxl
import xlrd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TARGET_HEADERS = [
    "姓名",
    "身份證號碼",
    "生日",
    "年齡",
    "電話",
    "手機號碼",
    "最後就診日",
    "114年就診次數",
    "115年就診次數",
    "病歷號",
    "地址",
    "備註",
]


@dataclass
class MemberRecord:
    name: str = ""
    pid: str = "-"
    bday: Optional[dt.date] = None
    chart_no: str = ""
    phone: str = ""
    mobile: str = ""
    address: str = ""
    last_visit: Optional[dt.date] = None
    count_114: float = 0.0
    count_115: float = 0.0
    sources: set[str] = field(default_factory=set)
    notes: set[str] = field(default_factory=set)


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    if text.endswith(".0") and re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return re.sub(r"\s+", "", text)


def display_text(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    if text.endswith(".0") and re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return text


def normalize_chart(value: Any) -> str:
    text = display_text(value)
    if not text:
        return ""
    return text.zfill(6) if text.isdigit() else text


def normalize_phone(value: Any) -> str:
    text = display_text(value)
    if not text:
        return ""
    digits = re.sub(r"\D+", "", text)
    if len(digits) == 9 and digits.startswith("9"):
        digits = "0" + digits
    return digits


def split_phone(value: Any) -> tuple[str, str]:
    phone = normalize_phone(value)
    if phone.startswith("09"):
        return "", phone
    return phone, ""


def parse_float(value: Any) -> float:
    try:
        return float(display_text(value).replace(",", "") or 0)
    except Exception:
        return 0.0


def parse_roc_birth(value: Any) -> Optional[dt.date]:
    text = display_text(value)
    if not text:
        return None
    digits = re.sub(r"\D+", "", text)
    if len(digits) < 5:
        return None
    if len(digits) <= 6:
        year = int(digits[:-4])
    else:
        year = int(digits[:-4])
    month = int(digits[-4:-2])
    day = int(digits[-2:])
    try:
        return dt.date(year + 1911, month, day)
    except ValueError:
        return None


def calc_age(bday: Optional[dt.date], today: dt.date) -> Optional[int]:
    if not bday:
        return None
    age = today.year - bday.year - ((today.month, today.day) < (bday.month, bday.day))
    return age if age >= 0 else None


def month_to_visit_date(month_code: str) -> Optional[dt.date]:
    if not re.fullmatch(r"1(?:14|15)\d{2}", month_code):
        return None
    roc_year = int(month_code[:3])
    month = int(month_code[3:5])
    try:
        return dt.date(roc_year + 1911, month, 1)
    except ValueError:
        return None


def read_xls_rows(path: Path) -> Iterable[list[Any]]:
    book = xlrd.open_workbook(str(path))
    sheet = book.sheet_by_index(0)
    for row_idx in range(sheet.nrows):
        yield [sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)]


def get_or_create(records: Dict[str, MemberRecord], key: str) -> MemberRecord:
    if key not in records:
        records[key] = MemberRecord()
    return records[key]


def record_key(chart_no: str, name: str, phone: str = "") -> str:
    if chart_no:
        return f"CHART:{chart_no}"
    compact_name = normalize_text(name)
    compact_phone = normalize_phone(phone)
    if compact_name and compact_phone:
        return f"NAMEPHONE:{compact_name}:{compact_phone}"
    return f"NAME:{compact_name}"


def merge_basic(rec: MemberRecord, *, name: Any = "", phone: Any = "", address: Any = "", source: str = "") -> None:
    if display_text(name) and not rec.name:
        rec.name = display_text(name)
    tel, mobile = split_phone(phone)
    if tel and not rec.phone:
        rec.phone = tel
    if mobile and not rec.mobile:
        rec.mobile = mobile
    if display_text(address) and not rec.address:
        rec.address = display_text(address)
    if source:
        rec.sources.add(source)


def collect_data(source_dir: Path) -> Dict[str, MemberRecord]:
    records: Dict[str, MemberRecord] = {}

    home_path = source_dir / "家115.xls"
    if home_path.exists():
        for row_idx, row in enumerate(read_xls_rows(home_path)):
            if row_idx == 0 or len(row) < 3:
                continue
            name, phone, address = row[0], row[1], row[2]
            if not display_text(name):
                continue
            key = record_key("", display_text(name), display_text(phone))
            rec = get_or_create(records, key)
            merge_basic(rec, name=name, phone=phone, address=address, source=home_path.name)

    count_dir = source_dir / "次數"
    if count_dir.is_dir():
        for path in sorted(count_dir.glob("*.xls")):
            visit_date = month_to_visit_date(path.stem)
            if visit_date is None:
                continue
            year_bucket = visit_date.year - 1911
            for row_idx, row in enumerate(read_xls_rows(path)):
                if row_idx == 0 or len(row) < 6:
                    continue
                chart_no = normalize_chart(row[0])
                name = display_text(row[1])
                if not chart_no and not name:
                    continue
                bday = parse_roc_birth(row[2])
                phone = row[3]
                count = parse_float(row[4])
                address = row[5]
                key = record_key(chart_no, name, phone)
                rec = get_or_create(records, key)
                merge_basic(rec, name=name, phone=phone, address=address, source=f"次數/{path.name}")
                if chart_no:
                    rec.chart_no = chart_no
                if bday and not rec.bday:
                    rec.bday = bday
                if year_bucket == 114:
                    rec.count_114 += count
                elif year_bucket == 115:
                    rec.count_115 += count
                if rec.last_visit is None or visit_date > rec.last_visit:
                    rec.last_visit = visit_date

    # 用姓名+電話把家115資料合併到已有病歷號資料，避免同一人拆成兩列。
    keyed_by_name_phone: Dict[str, MemberRecord] = {}
    for rec in records.values():
        phone = rec.mobile or rec.phone
        if rec.name and phone and rec.chart_no:
            keyed_by_name_phone[f"{normalize_text(rec.name)}:{phone}"] = rec

    for key in list(records.keys()):
        rec = records[key]
        phone = rec.mobile or rec.phone
        merge_target = keyed_by_name_phone.get(f"{normalize_text(rec.name)}:{phone}") if phone else None
        if merge_target is None or merge_target is rec:
            continue
        merge_target.sources.update(rec.sources)
        merge_target.notes.update(rec.notes)
        if rec.address and not merge_target.address:
            merge_target.address = rec.address
        del records[key]

    for rec in records.values():
        if not rec.chart_no:
            rec.notes.add("僅在「家115」檔案")
        elif not rec.bday:
            rec.notes.add("缺生日")
    return records


def sanitize_filename(text: str) -> str:
    return re.sub(r'[\\/:*?"<>|]+', "_", text).strip() or "耀聖"


def extract_clinic_name(source_dir: Path) -> str:
    match = re.search(r"\d{10}(.+?)耀聖$", source_dir.name)
    return match.group(1) if match else source_dir.name


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if cell.row == 1:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = {
        "A": 14, "B": 14, "C": 12, "D": 8, "E": 14, "F": 14,
        "G": 12, "H": 14, "I": 14, "J": 12, "K": 36, "L": 28,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def create_output(source_dir: Path, output_dir: Optional[Path] = None) -> Path:
    records = collect_data(source_dir)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "會員總表"
    ws.append(TARGET_HEADERS)

    today = dt.date.today()

    def sort_key(rec: MemberRecord) -> tuple[int, str, str]:
        return (0 if rec.chart_no else 1, normalize_text(rec.name), rec.chart_no)

    for rec in sorted(records.values(), key=sort_key):
        ws.append([
            rec.name or "-",
            rec.pid or "-",
            rec.bday,
            calc_age(rec.bday, today),
            rec.phone or "-",
            rec.mobile or "-",
            rec.last_visit,
            int(rec.count_114) if rec.count_114 else "-",
            int(rec.count_115) if rec.count_115 else "-",
            rec.chart_no or "-",
            rec.address or "-",
            "；".join(sorted(rec.notes)) if rec.notes else "-",
        ])

    for row in ws.iter_rows(min_row=2):
        for col in (3, 7):
            cell = row[col - 1]
            if isinstance(cell.value, dt.date):
                cell.number_format = "yyyy-mm-dd"

    style_sheet(ws)

    if output_dir is None:
        output_dir = source_dir.parent
    output_dir.mkdir(parents=True, exist_ok=True)
    clinic_name = sanitize_filename(extract_clinic_name(source_dir))
    timestamp = dt.datetime.now().strftime("%m%d_%H%M")
    out_path = output_dir / f"{clinic_name}耀聖總表_{timestamp}.xlsx"
    wb.save(out_path)
    return out_path


def open_file_cross_platform(path: Path) -> None:
    if sys.platform == "darwin":
        subprocess.call(("open", str(path)))
    elif os.name == "nt":
        os.startfile(str(path))  # type: ignore[attr-defined]
    else:
        subprocess.call(("xdg-open", str(path)))


def gui_main() -> None:
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()
    source = filedialog.askdirectory(title="選擇周一珊耀聖來源資料夾")
    if not source:
        return
    try:
        out_path = create_output(Path(source).expanduser().resolve())
        messagebox.showinfo("完成", f"已輸出：\n{out_path}")
        open_file_cross_platform(out_path)
    except Exception as exc:
        messagebox.showerror("錯誤", str(exc))


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="周一珊耀聖專用總表產檔")
    parser.add_argument("source_dir", nargs="?", help="周一珊耀聖來源資料夾")
    parser.add_argument("-o", "--output-dir", help="輸出資料夾；預設為來源資料夾上一層")
    args = parser.parse_args(argv)
    if not args.source_dir:
        gui_main()
        return 0
    source_dir = Path(args.source_dir).expanduser().resolve()
    if not source_dir.is_dir():
        print(f"找不到資料夾：{source_dir}")
        return 1
    output_dir = Path(args.output_dir).expanduser().resolve() if args.output_dir else None
    out_path = create_output(source_dir, output_dir)
    print(f"已輸出：{out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
