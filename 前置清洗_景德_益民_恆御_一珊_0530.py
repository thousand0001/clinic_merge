# -*- coding: utf-8 -*-
"""
景德 / 益民 / 恆御 / 一珊 前置清洗 + 通用主程式包裝（0530）

支援：
- 景德展望：舊版 R11440 .XLS 月報
- 益民耀聖：門診人次統計明細，姓名補 ID/生日後回填次數
- 恆御復健科醫聖：R11440 TXT 月報，補齊通用必要空白表
- 周一珊耀聖：次數 .XLS 月報，清成通用 HISB 次數格式

原則：
- 不修改 run_merge_通用_0430_1.py
- 清洗完呼叫通用程式
- 輸出後核對會員總表 L/M/N/O 與清洗階段累計 totals
"""

from __future__ import annotations

import csv
import datetime as _dt
import hashlib
import importlib.util
import os
import re
import shutil
import sys
import tempfile
from collections import defaultdict
from pathlib import Path
from typing import Any, DefaultDict, Dict, Iterable, List, Optional, Tuple

import xlrd
from openpyxl import Workbook, load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
TW_ID_RE = re.compile(r"^[A-Z][12]\d{8}$")
AGE_TEXT_RE = re.compile(r"^\d{2,3}歲\d{1,2}月\d{1,2}天$")
MAX_REASONABLE_CLAIM_AMOUNT = 50000
ILLEGAL_XLSX_CHARS_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")


def _find_generic_script(script_dir: Path) -> Path:
    candidates = sorted(script_dir.glob("run_merge_通用_*.py"), reverse=True)
    if not candidates:
        raise RuntimeError("找不到通用程式 run_merge_通用_*.py")
    return candidates[0]


GENERIC_SCRIPT = _find_generic_script(SCRIPT_DIR)


def _load_generic_module():
    spec = importlib.util.spec_from_file_location("run_merge_generic", GENERIC_SCRIPT)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"無法載入通用程式：{GENERIC_SCRIPT}")
    mod = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = mod
    spec.loader.exec_module(mod)
    return mod


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return re.sub(r"\s+", "", s)


def _normalize_id(value: Any) -> str:
    s = _normalize_text(value).upper()
    if s.endswith(".0"):
        s = s[:-2]
    return s.replace(" ", "")


def _is_valid_tw_id(value: Any) -> bool:
    return bool(TW_ID_RE.fullmatch(_normalize_id(value)))


def _clean_number(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and abs(value - round(value)) < 1e-9:
        return str(int(round(value)))
    s = str(value).strip().replace(",", "")
    if s.endswith(".0") and s[:-2].lstrip("-").isdigit():
        s = s[:-2]
    return s


def _clean_amount(value: Any) -> str:
    s = _clean_number(value)
    if not re.fullmatch(r"-?\d+(?:\.\d+)?", s or ""):
        return ""
    amount = int(round(float(s)))
    if abs(amount) > MAX_REASONABLE_CLAIM_AMOUNT:
        return ""
    return str(amount)


def _safe_excel_value(value: Any) -> Any:
    if isinstance(value, str):
        return ILLEGAL_XLSX_CHARS_RE.sub("", value)
    return value


def _extract_month_code(text: Any) -> Optional[str]:
    matches = re.findall(r"(1(?:14|15)\d{2})", str(text))
    return matches[-1] if matches else None


def _sheet_date_from_month(month_code: str) -> str:
    return f"{month_code[:3]}.{month_code[3:]}.01"


def _to_float(value: Any) -> float:
    s = _clean_number(value)
    if not s or not re.fullmatch(r"-?\d+(?:\.\d+)?", s):
        return 0.0
    return float(s)


def _decode_latin1_cp950(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    try:
        return value.encode("latin1").decode("cp950")
    except Exception:
        return value


def _read_xls_rows(path: Path) -> List[List[Any]]:
    try:
        book = xlrd.open_workbook(str(path))
        decoder = lambda x: x
    except UnicodeDecodeError:
        book = xlrd.open_workbook(str(path), encoding_override="latin1")
        decoder = _decode_latin1_cp950

    out: List[List[Any]] = []
    for sheet in book.sheets():
        for r in range(sheet.nrows):
            out.append([decoder(sheet.cell_value(r, c)) for c in range(sheet.ncols)])
    return out


def _header_map(headers: Iterable[Any]) -> Dict[str, int]:
    return {_normalize_text(v): idx for idx, v in enumerate(headers) if _normalize_text(v)}


def _pick_idx(hmap: Dict[str, int], *names: str) -> Optional[int]:
    for name in names:
        key = _normalize_text(name)
        if key in hmap:
            return hmap[key]
    return None


def _cell(row: List[Any], idx: Optional[int]) -> Any:
    if idx is None or idx >= len(row):
        return ""
    return row[idx]


def _append_month_row(
    months: DefaultDict[str, List[List[Any]]],
    totals: Dict[str, float],
    month_code: str,
    row: List[Any],
    count: Any,
    amount: Any,
) -> None:
    cnt = _to_float(count)
    amt = _to_float(amount)
    if cnt == 0 and amt == 0:
        return
    months[month_code].append(row)
    if month_code.startswith("114"):
        totals["L"] += cnt
        totals["M"] += amt
    elif month_code.startswith("115"):
        totals["N"] += cnt
        totals["O"] += amt


def _pseudo_id(seed: str) -> str:
    # 通用程式需要像身分證格式的 key 才能回填月份統計；輸出後會清空 ZZ 開頭內部 ID。
    digest = hashlib.sha1(seed.encode("utf-8")).hexdigest()
    number = int(digest[:12], 16) % 100000000
    return f"ZZ{number:08d}"


def _save_month_workbook(source_dir: Path, months: Dict[str, List[List[Any]]]) -> Optional[Path]:
    if not months:
        return None
    wb = Workbook()
    wb.remove(wb.active)
    for month_code in sorted(months):
        ws = wb.create_sheet(month_code)
        for row in months[month_code]:
            ws.append([_safe_excel_value(v) for v in row])
    out_path = source_dir / "000_清洗_月份統計.xlsx"
    wb.save(out_path)
    return out_path


def _find_unique_name_roster(source_dir: Path) -> Dict[str, Tuple[str, str]]:
    candidates = list(source_dir.glob("*自選*.xlsx")) + list(source_dir.glob("*A115*.xlsx"))
    by_name: DefaultDict[str, List[Tuple[str, str]]] = defaultdict(list)

    for path in sorted(set(candidates)):
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        try:
            for ws in wb.worksheets:
                header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                hmap = _header_map(header)
                id_idx = _pick_idx(hmap, "身分證", "身分證號", "身份證號", "ID")
                name_idx = _pick_idx(hmap, "姓名", "病患姓名", "會員姓名")
                bday_idx = _pick_idx(hmap, "生日", "出生日期")
                if id_idx is None:
                    # 益民 A115 舊表第一欄含 A171+ID+西元生日。
                    id_idx = 0
                if name_idx is None:
                    continue
                for row in ws.iter_rows(min_row=2, values_only=True):
                    raw_id = _cell(list(row), id_idx)
                    pid = _normalize_id(raw_id)
                    bday = _cell(list(row), bday_idx)
                    if not _is_valid_tw_id(pid):
                        m = re.search(r"([A-Z][12]\d{8})(\d{8})?", _normalize_text(raw_id).upper())
                        if not m:
                            continue
                        pid = m.group(1)
                        if not bday and m.group(2):
                            bday = m.group(2)
                    name = _normalize_text(_cell(list(row), name_idx))
                    if name:
                        by_name[name].append((pid, _clean_number(bday)))
        finally:
            wb.close()

    unique: Dict[str, Tuple[str, str]] = {}
    for name, rows in by_name.items():
        ids = {(pid, bday) for pid, bday in rows if _is_valid_tw_id(pid)}
        if len(ids) == 1:
            unique[name] = next(iter(ids))
    return unique


def _clean_jingde(source_dir: Path, totals: Dict[str, float]) -> int:
    month_rows: DefaultDict[str, List[List[Any]]] = defaultdict(
        lambda: [["ID", "姓名", "生日", "日期", "次數", "申請金額", "地址", "電話"]]
    )
    written = 0
    month_files = [
        path for path in sorted((source_dir / "R11440").iterdir())
        if path.is_file() and path.suffix.lower() in {".xls", ".xlsx"}
    ]
    for path in month_files:
        month_code = _extract_month_code(path.name)
        if not month_code:
            continue
        rows = _read_xls_rows(path)
        if not rows:
            continue
        hmap = _header_map(rows[0])
        id_idx = _pick_idx(hmap, "身分證號", "身份證號", "ID")
        name_idx = _pick_idx(hmap, "姓名")
        bday_idx = _pick_idx(hmap, "生日")
        date_idx = _pick_idx(hmap, "日期")
        count_idx = _pick_idx(hmap, "次數", "件數")
        amount_idx = _pick_idx(hmap, "總額", "總金額", "申請金額")
        addr_idx = _pick_idx(hmap, "地址")
        tel_idx = _pick_idx(hmap, "電話")
        if id_idx is None or count_idx is None or amount_idx is None:
            continue
        for row in rows[1:]:
            pid = _normalize_id(_cell(row, id_idx))
            if not _is_valid_tw_id(pid):
                continue
            out_row = [
                pid,
                _normalize_text(_cell(row, name_idx)),
                _clean_number(_cell(row, bday_idx)),
                _clean_number(_cell(row, date_idx)) or _sheet_date_from_month(month_code),
                _clean_number(_cell(row, count_idx)),
                _clean_amount(_cell(row, amount_idx)),
                _cell(row, addr_idx),
                _cell(row, tel_idx),
            ]
            _append_month_row(month_rows, totals, month_code, out_row, out_row[4], out_row[5])
            written += 1
    if written:
        _save_month_workbook(source_dir, month_rows)
        for path in month_files:
            path.unlink(missing_ok=True)
    return written


def _clean_zhou_yishan(source_dir: Path, totals: Dict[str, float]) -> int:
    month_dir = source_dir / "次數"
    if not month_dir.is_dir():
        return 0
    month_rows: DefaultDict[str, List[List[Any]]] = defaultdict(
        lambda: [["ID", "病歷號", "姓名", "生日", "日期", "次數", "申請金額", "地址", "電話"]]
    )
    written = 0
    month_files = [
        path for path in sorted(month_dir.iterdir())
        if path.is_file() and path.suffix.lower() in {".xls", ".xlsx"}
    ]
    for path in month_files:
        month_code = _extract_month_code(path.name)
        if not month_code:
            continue
        rows = _read_xls_rows(path)
        if not rows:
            continue
        hmap = _header_map(rows[0])
        chart_idx = _pick_idx(hmap, "病歷號", "病歷號碼")
        name_idx = _pick_idx(hmap, "姓名", "姓    名")
        bday_idx = _pick_idx(hmap, "生日", "生   日")
        tel_idx = _pick_idx(hmap, "電話", "電      話")
        count_idx = _pick_idx(hmap, "次數")
        addr_idx = _pick_idx(hmap, "地址", "地                                    址")
        if name_idx is None or bday_idx is None or count_idx is None:
            continue
        for row in rows[1:]:
            name = _normalize_text(_cell(row, name_idx))
            bday = _clean_number(_cell(row, bday_idx))
            cnt = _clean_number(_cell(row, count_idx))
            if not name or not bday or not cnt:
                continue
            chart_no = _clean_number(_cell(row, chart_idx))
            pid = _pseudo_id(f"周一珊|{chart_no}|{name}|{bday}")
            out_row = [
                pid,
                chart_no,
                name,
                bday,
                _sheet_date_from_month(month_code),
                cnt,
                "",
                _cell(row, addr_idx),
                _cell(row, tel_idx),
            ]
            _append_month_row(month_rows, totals, month_code, out_row, cnt, 0)
            written += 1
    if written:
        _save_month_workbook(source_dir, month_rows)
        for path in month_files:
            path.unlink(missing_ok=True)
    return written


def _clean_yimin(source_dir: Path, totals: Dict[str, float]) -> int:
    r11440 = source_dir / "R11440"
    if not r11440.is_dir():
        return 0
    name_map = _find_unique_name_roster(source_dir)
    month_rows: DefaultDict[str, List[List[Any]]] = defaultdict(
        lambda: [["ID", "姓名", "生日", "日期", "次數", "申請金額", "病歷號"]]
    )
    written = 0
    month_files = [path for path in sorted(r11440.glob("*.xlsx")) if path.is_file()]
    for path in month_files:
        month_code = _extract_month_code(path.name)
        if not month_code:
            continue
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            if "門診人次統計-資料明細" not in wb.sheetnames:
                continue
            ws = wb["門診人次統計-資料明細"]
            rows = list(ws.iter_rows(values_only=True))
        finally:
            wb.close()
        if len(rows) < 5:
            continue
        hmap = _header_map(rows[3])
        date_idx = _pick_idx(hmap, "看診日期", "日期")
        chart_idx = _pick_idx(hmap, "病歷號", "病歷號碼")
        name_idx = _pick_idx(hmap, "病患姓名", "姓名")
        if date_idx is None or name_idx is None:
            continue
        for row in rows[4:]:
            vals = list(row)
            name = _normalize_text(_cell(vals, name_idx))
            if not name:
                continue
            mapped = name_map.get(name)
            chart_no = _clean_number(_cell(vals, chart_idx))
            if mapped:
                pid, bday = mapped
            else:
                pid = _pseudo_id(f"益民|{chart_no}|{name}")
                bday = ""
            out_row = [
                pid,
                name,
                bday,
                _cell(vals, date_idx) or _sheet_date_from_month(month_code),
                1,
                "",
                chart_no,
            ]
            _append_month_row(month_rows, totals, month_code, out_row, 1, 0)
            written += 1
    if written:
        _save_month_workbook(source_dir, month_rows)
        for path in month_files:
            path.unlink(missing_ok=True)
    return written


def _parse_txt_lines(txt_path: Path) -> List[List[str]]:
    with txt_path.open("r", encoding="cp950", errors="replace", newline="") as f:
        lines = [line.rstrip("\r\n") for line in f if line.strip()]
    rows: List[List[str]] = []
    for line in lines:
        if "費用年月:" in line:
            continue
        parts = [part.strip() for part in line.split(",")]
        while parts and parts[-1] == "":
            parts.pop()
        if parts:
            rows.append(parts)
    return rows


def _looks_like_date(text: Any) -> bool:
    s = str(text).strip()
    return bool(re.fullmatch(r"\d{2,3}[./-]\d{1,2}[./-]\d{1,2}", s) or re.fullmatch(r"\d{4}[./-]\d{1,2}[./-]\d{1,2}", s))


def _extract_shifted_claim_amount(row: List[str]) -> str:
    if len(row) >= 22:
        amount = _clean_amount(row[-7])
        if amount:
            return amount
    for idx, value in enumerate(row):
        if AGE_TEXT_RE.fullmatch(str(value).strip()) and idx + 4 < len(row):
            amount = _clean_amount(row[idx + 4])
            if amount:
                return amount
    return ""


def _convert_yisheng_txt_monthlies(source_dir: Path, totals: Dict[str, float]) -> int:
    month_rows: DefaultDict[str, List[List[Any]]] = defaultdict(
        lambda: [["ID", "姓名", "生日", "日期", "次數", "申請金額"]]
    )
    written = 0
    for txt_path in sorted(source_dir.rglob("*.txt")):
        month_code = _extract_month_code(txt_path.name)
        if not month_code:
            continue
        rows = _parse_txt_lines(txt_path)
        if len(rows) < 2:
            continue
        hmap = _header_map(rows[0])
        id_idx = _pick_idx(hmap, "身分證", "身分證號", "身份證號", "ID")
        name_idx = _pick_idx(hmap, "姓名", "病患姓名")
        bday_idx = _pick_idx(hmap, "生日")
        date_idx = _pick_idx(hmap, "看診日", "日期", "就醫日")
        amount_idx = _pick_idx(hmap, "申請額", "申請金額")
        for row in rows[1:]:
            pid = _normalize_id(_cell(row, id_idx))
            name = _normalize_text(_cell(row, name_idx))
            dt = _cell(row, date_idx)
            bday = _cell(row, bday_idx)
            amount = _clean_amount(_cell(row, amount_idx)) or _extract_shifted_claim_amount(row)
            if not _is_valid_tw_id(pid):
                id_pos = next((i for i, v in enumerate(row) if _is_valid_tw_id(v)), None)
                if id_pos is not None:
                    pid = _normalize_id(row[id_pos])
            if not _looks_like_date(dt):
                date_pos = next((i for i, v in enumerate(row) if _looks_like_date(v)), None)
                if date_pos is not None:
                    dt = row[date_pos]
                    if date_pos + 3 < len(row) and not name:
                        name = _normalize_text(row[date_pos + 1])
            if not _is_valid_tw_id(pid) or not dt:
                continue
            out_row = [pid, name, bday, dt, 1, amount]
            _append_month_row(month_rows, totals, month_code, out_row, 1, amount)
            written += 1
    if written:
        _save_month_workbook(source_dir, month_rows)
    return written


def _existing_canonical_sheets(source_dir: Path) -> set[str]:
    generic = _load_generic_module()
    try:
        wb = generic._merge_source_folder(str(source_dir))
    except Exception:
        return set()
    return set(wb.sheetnames)


def _add_missing_required_workbooks(source_dir: Path) -> Optional[Path]:
    required = {
        "HealthCase": ["家醫收案會員ID", "姓名", "生日", "最近一次HbA1c檢查結果(%)", "最近一次HbA1c檢查日期", "最近一次LDL檢查結果(mg/dL)", "最近一次LDL檢查日期", "最近一次UACR檢查結果(mg/gm)", "最近一次UACR檢查日期"],
        "成人健檢": ["ID", "最後篩檢日期"],
        "子宮抹片": ["ID", "最後篩檢日期"],
        "老人流感": ["ID", "最後篩檢日期"],
        "糞便潛血": ["ID", "最後篩檢日期"],
        "肝炎篩檢": ["ID", "最後篩檢日期"],
    }
    existing = _existing_canonical_sheets(source_dir)
    missing = [name for name in required if name not in existing]
    if not missing:
        return None
    wb = Workbook()
    wb.remove(wb.active)
    for name in missing:
        ws = wb.create_sheet(name)
        ws.append(required[name])
    out_path = source_dir / "__清洗_必要工作表補齊.xlsx"
    wb.save(out_path)
    return out_path


def _sum_output_member_totals(output_path: Path) -> Dict[str, float]:
    wb = load_workbook(output_path, read_only=True, data_only=True)
    try:
        if "會員總表" not in wb.sheetnames:
            raise ValueError("輸出檔缺少會員總表，無法核對 L/M/N/O")
        ws = wb["會員總表"]
        totals = {"L": 0.0, "M": 0.0, "N": 0.0, "O": 0.0}
        for row in ws.iter_rows(min_row=3, min_col=12, max_col=15, values_only=True):
            for key, value in zip(("L", "M", "N", "O"), row):
                totals[key] += _to_float(value)
        return totals
    finally:
        wb.close()


def _validate_output_totals(output_path: Path, expected: Dict[str, float]) -> None:
    actual = _sum_output_member_totals(output_path)
    labels = {"L": "114年次數", "M": "114年費用", "N": "115年次數", "O": "115年費用"}
    mismatches = []
    for key in ("L", "M", "N", "O"):
        if round(actual.get(key, 0.0), 2) != round(expected.get(key, 0.0), 2):
            mismatches.append(f"{key}欄{labels[key]}：輸入 {expected.get(key, 0.0):,.0f} / 輸出 {actual.get(key, 0.0):,.0f}")
    if mismatches:
        raise ValueError("會員總表 L/M/N/O 核對不一致：\n" + "\n".join(mismatches))
    print(
        "會員總表 L/M/N/O 核對相符："
        f"L={actual['L']:,.0f}, M={actual['M']:,.0f}, N={actual['N']:,.0f}, O={actual['O']:,.0f}",
        flush=True,
    )


def _blank_internal_ids(output_path: Path) -> None:
    generic = _load_generic_module()
    wb = load_workbook(output_path)
    try:
        if "會員總表" not in wb.sheetnames:
            return
        ws = wb["會員總表"]
        cols = generic.detect_template_columns(ws, 3)
        id_col = cols.get("id")
        if not id_col:
            return
        changed = False
        for row in range(3, ws.max_row + 1):
            value = _normalize_text(ws.cell(row, id_col).value).upper()
            if value.startswith("ZZ") and re.fullmatch(r"ZZ\d{8}", value):
                ws.cell(row, id_col).value = None
                changed = True
        if changed:
            wb.save(output_path)
    finally:
        wb.close()


def _preclean(source_dir: Path) -> Dict[str, float]:
    totals = {"L": 0.0, "M": 0.0, "N": 0.0, "O": 0.0}
    name = source_dir.name

    written = 0
    if "景德" in name:
        written = _clean_jingde(source_dir, totals)
    elif "益民" in name:
        written = _clean_yimin(source_dir, totals)
    elif "周一珊" in name:
        written = _clean_zhou_yishan(source_dir, totals)
    elif "恆御" in name or "醫聖" in name:
        written = _convert_yisheng_txt_monthlies(source_dir, totals)
    else:
        written = (
            _clean_jingde(source_dir, totals)
            or _clean_yimin(source_dir, totals)
            or _clean_zhou_yishan(source_dir, totals)
            or _convert_yisheng_txt_monthlies(source_dir, totals)
        )

    _add_missing_required_workbooks(source_dir)
    print(f"已產生清洗後月報明細 {written} 筆", flush=True)
    return totals


def process_excel(source_path: str, template_path: Optional[str] = None) -> str:
    generic = _load_generic_module()
    source_dir = Path(source_path).resolve()
    if not source_dir.is_dir():
        raise ValueError("請選擇資料夾。")

    template = template_path or generic._find_template(str(SCRIPT_DIR))
    if not os.path.exists(template):
        raise ValueError(f"找不到模板檔：{template}")

    temp_root = Path(tempfile.mkdtemp(prefix="newclinic_clean_"))
    temp_source = temp_root / source_dir.name
    try:
        shutil.copytree(source_dir, temp_source)
        expected = _preclean(temp_source)
        temp_output = Path(generic.process_excel(str(temp_source), template))
        _blank_internal_ids(temp_output)
        _validate_output_totals(temp_output, expected)
        final_output = source_dir.parent / temp_output.name
        if final_output.exists():
            final_output.unlink()
        shutil.move(str(temp_output), str(final_output))
        return str(final_output)
    finally:
        shutil.rmtree(temp_root, ignore_errors=True)


def main() -> None:
    generic = _load_generic_module()
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()

    src = filedialog.askdirectory(title="選擇新診所來源資料夾")
    if not src:
        return

    template = generic._find_template(str(SCRIPT_DIR))
    try:
        out = process_excel(src, template)
        messagebox.showinfo("完成", f"已輸出：\n{out}")
        generic.open_file_cross_platform(out)
    except Exception as e:
        messagebox.showerror("錯誤", str(e))


if __name__ == "__main__":
    main()
