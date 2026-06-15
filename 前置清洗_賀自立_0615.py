#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""賀自立耳鼻喉科診所前置清洗 + 最新共用核心包裝（0615）。

自選／不選來源依檔名分類：
- 含 115X、不選、不要會員：不選會員（E2），優先於自選判斷。
- 其餘含自選、要選、預選或獨立的 115：自選會員（E1）。
- 同一 ID 同時出現時保留兩個旗標，由共用核心顯示 E1/E2，
  並從最終自選名單分頁排除。

新版來源可直接提供身分證字號；舊版若沒有 ID，仍可從同資料夾其他
Excel 或「賀自立會員ID對照*.xlsx」建立姓名 → ID 唯一對照。

指定欄位：
- F 看診數       → 會員總表 L
- G 最後看診日   → 會員總表 K
- H 去年費用     → 會員總表 M
- I 今年費用     → 會員總表 O
- P 成健執行日   → 會員總表 Q
- N 子抹執行日   → 會員總表 R
- O 潛血執行日   → 會員總表 T

正式執行時若有無效或無法唯一取得的 ID，仍會納入會員輸出、產生待補
ID 報表，並在終端機及完成提示視窗顯示警告。
"""

from __future__ import annotations

import argparse
import datetime as dt
import importlib.util
import os
import re
import shutil
import subprocess
import sys
import tempfile
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl


PROJECT_DIR = Path(__file__).resolve().parent
CLINIC_CODE = "3531030942"
CLINIC_NAME = "賀自立耳鼻喉科診所"
VERSION_TAG = "0615"


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def normalize_name(value: Any) -> str:
    return re.sub(r"[\s　]+", "", normalize_text(value))


def normalize_id(value: Any) -> str:
    return re.sub(r"\s+", "", normalize_text(value).upper().strip("'"))


def normalize_phone(value: Any) -> str:
    digits = re.sub(r"\D", "", normalize_text(value))
    if len(digits) == 9 and digits.startswith(("9", "2")):
        return f"0{digits}"
    if len(digits) == 8:
        return f"02{digits}"
    return digits


def is_valid_id(value: Any) -> bool:
    return bool(re.fullmatch(r"[A-Z][1289]\d{8}", normalize_id(value)))


def parse_number(value: Any) -> float:
    text = normalize_text(value).replace(",", "")
    if text in {"", "-", "—", "–"}:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def parse_date(value: Any) -> Optional[dt.date]:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = normalize_text(value)
    if text in {"", "-", "—", "–"}:
        return None
    digits = re.sub(r"\D", "", text.split()[0])
    try:
        if len(digits) == 7:
            return dt.date(int(digits[:3]) + 1911, int(digits[3:5]), int(digits[5:7]))
        if len(digits) == 8:
            year = int(digits[:4])
            if year < 1911:
                year += 1911
            return dt.date(year, int(digits[4:6]), int(digits[6:8]))
    except ValueError:
        return None
    return None


def compact_header(value: Any) -> str:
    return re.sub(r"[\s　]+", "", normalize_text(value)).upper()


def find_header(ws: Any, required: Sequence[Sequence[str]], search_rows: int = 12) -> Optional[int]:
    groups = [{compact_header(alias) for alias in aliases} for aliases in required]
    for row_no in range(1, min(ws.max_row, search_rows) + 1):
        values = {
            compact_header(ws.cell(row_no, col).value)
            for col in range(1, ws.max_column + 1)
        }
        if all(group & values for group in groups):
            return row_no
    return None


def header_map(ws: Any, row_no: int) -> Dict[str, int]:
    return {
        compact_header(ws.cell(row_no, col).value): col
        for col in range(1, ws.max_column + 1)
        if compact_header(ws.cell(row_no, col).value)
    }


def find_col(columns: Dict[str, int], aliases: Iterable[str]) -> Optional[int]:
    for alias in aliases:
        column = columns.get(compact_header(alias))
        if column is not None:
            return column
    return None


def latest_core() -> Path:
    candidates = sorted(
        PROJECT_DIR.glob("選會員_共用核心_*.py"),
        key=lambda path: path.name,
        reverse=True,
    )
    if not candidates:
        raise RuntimeError("找不到共用核心：選會員_共用核心_*.py")
    return candidates[0]


def latest_template() -> Path:
    candidates = sorted(
        PROJECT_DIR.glob("選會員模板*.xlsx"),
        key=lambda path: path.name,
        reverse=True,
    )
    if not candidates:
        raise RuntimeError("找不到模板：選會員模板*.xlsx")
    return candidates[0]


def classify_selection_file(path: Path) -> Optional[str]:
    name = re.sub(r"[\s\-_()（）\[\]{}]+", "", path.stem.lower())
    if "會員" not in name and "選" not in name:
        return None
    if "115x" in name or "不選" in name or "不要會員" in name or "不要" in name:
        return "exclude"
    has_115 = "115" in name and "1150" not in name and "1151" not in name
    if has_115 or any(token in name for token in ("自選", "要選", "預選")):
        return "select"
    return None


def discover_selection_files(source_dir: Path) -> List[Tuple[Path, str]]:
    files: List[Tuple[Path, str]] = []
    for path in sorted(source_dir.glob("*.xlsx")):
        if path.name.startswith("~$") or path.name.startswith("賀自立會員ID對照_"):
            continue
        kind = classify_selection_file(path)
        if kind:
            files.append((path, kind))
    if not files:
        raise FileNotFoundError(
            "找不到自選／不選會員檔案；檔名需包含自選、要選、預選、115、"
            "115X、不選或不要會員。"
        )
    return files


def load_selection_file(path: Path, kind: str) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows: List[Dict[str, Any]] = []
    try:
        for ws in wb.worksheets:
            header_row = find_header(
                ws,
                (
                    ("姓名", "會員姓名"),
                    ("身分證字號", "身分證號", "身份證字號", "身份證號", "ID", "病歷號碼"),
                ),
            )
            if header_row is None:
                continue
            columns = header_map(ws, header_row)
            aliases = {
                "chart": ("病歷號碼", "病歷號"),
                "pid": ("身分證字號", "身分證號", "身份證字號", "身份證號", "ID"),
                "name": ("姓名",),
                "sex": ("性別",),
                "age": ("年齡",),
                "count": ("看診數",),
                "last_visit": ("最後看診日",),
                "amount_114": ("去年費用", "去年"),
                "amount_115": ("今年費用", "今年"),
                "pap": ("子抹執行日", "子抹"),
                "fit": ("潛血執行日", "潛血"),
                "adult": ("成健執行日", "成健"),
                "birth": ("出生日期", "生日"),
                "phone": ("電話",),
                "address": ("住址", "地址"),
            }
            indexes = {key: find_col(columns, names) for key, names in aliases.items()}
            if indexes["name"] is None:
                continue

            def value(row_no: int, key: str) -> Any:
                column = indexes[key]
                return ws.cell(row_no, column).value if column else None

            for row_no in range(header_row + 1, ws.max_row + 1):
                name = normalize_name(value(row_no, "name"))
                if not name:
                    continue
                rows.append({
                    "source_file": path.name,
                    "source_kind": kind,
                    "source_row": row_no,
                    "chart": normalize_text(value(row_no, "chart")),
                    "pid": normalize_id(value(row_no, "pid")),
                    "name": name,
                    "birth": parse_date(value(row_no, "birth")),
                    "sex": normalize_text(value(row_no, "sex")),
                    "age": normalize_text(value(row_no, "age")),
                    "count": parse_number(value(row_no, "count")),
                    "last_visit": parse_date(value(row_no, "last_visit")),
                    "amount_114": parse_number(value(row_no, "amount_114")),
                    "amount_115": parse_number(value(row_no, "amount_115")),
                    "pap": parse_date(value(row_no, "pap")),
                    "fit": parse_date(value(row_no, "fit")),
                    "adult": parse_date(value(row_no, "adult")),
                    "phone": normalize_phone(value(row_no, "phone")),
                    "address": normalize_text(value(row_no, "address")),
                })
    finally:
        wb.close()
    if not rows:
        raise ValueError(f"{path.name} 找不到有效會員資料")
    return rows


def load_selection_rows(
    selection_files: Sequence[Tuple[Path, str]],
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for path, kind in selection_files:
        rows.extend(load_selection_file(path, kind))
    return rows


def build_identity_index(
    source_dir: Path,
    selection_paths: set[Path],
) -> Dict[str, List[Dict[str, Any]]]:
    identities: Dict[str, Dict[str, Dict[str, Any]]] = defaultdict(dict)
    for path in sorted(source_dir.glob("*.xlsx")):
        if path.name.startswith("~$") or path in selection_paths:
            continue
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                header_row = find_header(
                    ws,
                    (("姓名", "會員姓名"), ("ID", "家醫收案會員ID", "身分證號", "身份證號")),
                )
                if header_row is None:
                    continue
                columns = header_map(ws, header_row)
                name_col = find_col(columns, ("姓名", "會員姓名"))
                id_col = find_col(columns, ("ID", "家醫收案會員ID", "身分證號", "身份證號"))
                birth_col = find_col(columns, ("生日", "出生日期", "BIRTHDAY"))
                if name_col is None or id_col is None:
                    continue
                for row_no in range(header_row + 1, ws.max_row + 1):
                    name = normalize_name(ws.cell(row_no, name_col).value)
                    pid = normalize_id(ws.cell(row_no, id_col).value)
                    if not name or not is_valid_id(pid):
                        continue
                    birth = parse_date(ws.cell(row_no, birth_col).value) if birth_col else None
                    current = identities[name].setdefault(pid, {
                        "pid": pid,
                        "birth": birth,
                        "source": path.name,
                    })
                    if current["birth"] is None and birth is not None:
                        current["birth"] = birth
        finally:
            wb.close()
    return {name: list(by_id.values()) for name, by_id in identities.items()}


def match_identities(
    rows: List[Dict[str, Any]],
    identity_index: Dict[str, List[Dict[str, Any]]],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    matched: List[Dict[str, Any]] = []
    unresolved: List[Dict[str, Any]] = []
    for row in rows:
        if is_valid_id(row.get("pid")):
            matched.append(row)
            continue
        candidates = identity_index.get(row["name"], [])
        if len(candidates) == 1:
            matched.append({**row, **candidates[0]})
            continue
        reason = "找不到姓名對應的身分證號" if not candidates else "同名對應到多個身分證號"
        unresolved.append({
            **row,
            "reason": reason,
            "candidate_ids": "|".join(item["pid"] for item in candidates),
        })
    return matched, unresolved


def write_unresolved_report(source_dir: Path, unresolved: List[Dict[str, Any]]) -> Path:
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    path = source_dir / f"賀自立會員ID對照_{VERSION_TAG}_{stamp}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "待補ID"
    headers = ["來源列", "病歷號碼", "姓名", "電話", "住址", "身分證號", "原因", "候選ID"]
    ws.append(headers)
    for row in unresolved:
        ws.append([
            row["source_row"], row["chart"], row["name"], row["phone"], row["address"],
            row.get("pid", ""), row["reason"], row["candidate_ids"],
        ])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{ws.max_row}"
    widths = [10, 14, 14, 16, 36, 18, 24, 24]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    wb.save(path)
    wb.close()
    return path


def write_table(path: Path, sheet_name: str, headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(list(headers))
    for row in rows:
        ws.append(list(row))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()


def write_claims(path: Path, rows: List[Dict[str, Any]]) -> None:
    headers_114 = ["身分證號", "姓名", "次數", "總額"]
    headers_115 = ["身分證號", "姓名", "日期", "次數", "總額"]
    wb = openpyxl.Workbook()
    ws114 = wb.active
    ws114.title = "11401"
    ws114.append(headers_114)
    ws115 = wb.create_sheet("11501")
    ws115.append(headers_115)
    for row in rows:
        ws114.append([row["pid"], row["name"], row["count"], row["amount_114"]])
        ws115.append([row["pid"], row["name"], row["last_visit"], 0, row["amount_115"]])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()


def copy_supporting_files(
    source_dir: Path,
    clean_dir: Path,
    selection_paths: set[Path],
) -> None:
    for path in sorted(source_dir.glob("*.xlsx")):
        if path.name.startswith("~$") or path in selection_paths:
            continue
        lower = path.name.lower()
        if "cliscores" in lower:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            try:
                indicator = normalize_text(wb.active.cell(2, 1).value)
            finally:
                wb.close()
            if any(token in indicator for token in ("成人預防", "子宮", "糞便潛血")):
                continue
        shutil.copy2(path, clean_dir / path.name)


def merge_rows_by_id(
    matched: Sequence[Dict[str, Any]],
) -> Tuple[List[Dict[str, Any]], set[str], set[str]]:
    by_id: Dict[str, Dict[str, Any]] = {}
    self_ids: set[str] = set()
    exclude_ids: set[str] = set()
    fill_keys = (
        "name", "birth", "chart", "sex", "age", "count", "last_visit",
        "amount_114", "amount_115", "pap", "fit", "adult", "phone", "address",
    )
    for row in matched:
        pid = row["pid"]
        if row["source_kind"] == "exclude":
            exclude_ids.add(pid)
        else:
            self_ids.add(pid)
        current = by_id.setdefault(pid, dict(row))
        for key in fill_keys:
            current_value = current.get(key)
            new_value = row.get(key)
            if (
                current_value in (None, "")
                or (key in {"count", "amount_114", "amount_115"} and current_value == 0)
            ) and new_value not in (None, "", 0):
                current[key] = row[key]
    return list(by_id.values()), self_ids, exclude_ids


def write_clean_data(
    clean_dir: Path,
    rows: List[Dict[str, Any]],
    self_ids: set[str],
    exclude_ids: set[str],
) -> Dict[str, float]:
    clean_dir.mkdir(parents=True, exist_ok=True)
    write_table(
        clean_dir / "自選會員_標準.xlsx",
        "自選會員",
        ["ID", "姓名", "生日", "電話", "地址"],
        [
            [row["pid"], row["name"], row.get("birth"), row["phone"], row["address"]]
            for row in rows
            if row["pid"] in self_ids
        ],
    )
    write_table(
        clean_dir / "115X不選會員_標準.xlsx",
        "115X",
        ["ID", "姓名", "生日", "電話", "地址"],
        [
            [row["pid"], row["name"], row.get("birth"), row["phone"], row["address"]]
            for row in rows
            if row["pid"] in exclude_ids
        ],
    )
    write_claims(clean_dir / "門診次數費用_標準.xlsx", rows)
    screening_specs = [
        ("成人健檢_整合檔.xlsx", "成人健檢", "成人預防保健檢查", "adult"),
        ("子宮抹片_整合檔.xlsx", "子宮抹片", "子宮頸抹片檢查", "pap"),
        ("糞便潛血_整合檔.xlsx", "糞便潛血", "糞便潛血檢查", "fit"),
    ]
    for filename, sheet_name, indicator, key in screening_specs:
        write_table(
            clean_dir / filename,
            sheet_name,
            ["指標名稱", "ID", "生日", "姓名", "最後篩檢日期"],
            [
                [indicator, row["pid"], row.get("birth"), row["name"], row[key]]
                for row in rows
                if row[key] is not None
            ],
        )
    overlap_ids = self_ids & exclude_ids
    return {
        "members": len(rows),
        "count_l": sum(row["count"] for row in rows),
        "amount_m": sum(row["amount_114"] for row in rows),
        "count_n": 0,
        "amount_o": sum(row["amount_115"] for row in rows),
        "last_visit": sum(row["last_visit"] is not None for row in rows),
        "adult": sum(row["adult"] is not None for row in rows),
        "pap": sum(row["pap"] is not None for row in rows),
        "fit": sum(row["fit"] is not None for row in rows),
        "pure_e1": len(self_ids - exclude_ids),
        "pure_e2": len(exclude_ids - self_ids),
        "overlap_e1_e2": len(overlap_ids),
        "self_output": len(self_ids - exclude_ids),
    }


def prepare_clean_dir(
    source_dir: Path,
    tmp_root: Path,
) -> Tuple[Path, Dict[str, Any], Optional[Path]]:
    selection_files = discover_selection_files(source_dir)
    selection_paths = {path for path, _kind in selection_files}
    rows = load_selection_rows(selection_files)
    matched, unresolved = match_identities(
        rows,
        build_identity_index(source_dir, selection_paths),
    )
    report = write_unresolved_report(source_dir, unresolved) if unresolved else None
    if unresolved:
        print(
            f"警告：預選名單共 {len(rows)} 筆，其中 {len(unresolved)} 筆因 ID 無效"
            "或無法唯一取得，仍納入會員輸出但 ID 待補："
        )
        for row in unresolved:
            print(
                f"  - {row['source_file']} 第 {row['source_row']} 列，"
                f"{row['name']}，ID={row.get('pid') or '空白'}，原因：{row['reason']}"
            )
        print(f"待補 ID 報表：{report}")
    run_tag = dt.datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    clean_dir = tmp_root / f"{CLINIC_CODE}{CLINIC_NAME}_{run_tag}"
    clean_dir.mkdir(parents=True, exist_ok=True)
    merged, self_ids, exclude_ids = merge_rows_by_id(matched)
    copy_supporting_files(source_dir, clean_dir, selection_paths)
    counters = write_clean_data(clean_dir, merged, self_ids, exclude_ids)
    counters["source_members"] = len(rows)
    counters["unresolved"] = len(unresolved)
    counters["unresolved_rows"] = unresolved
    counters["selection_files"] = len(selection_files)
    return clean_dir, counters, report


def run_common_core(clean_dir: Path, output_dir: Path) -> Path:
    core_path = latest_core()
    spec = importlib.util.spec_from_file_location("hezili_common_core", core_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"無法載入共用核心：{core_path}")
    core = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = core
    spec.loader.exec_module(core)
    generated = Path(core.process_excel(str(clean_dir), str(latest_template())))
    output_dir.mkdir(parents=True, exist_ok=True)
    target = output_dir / generated.name
    if target.exists():
        stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        target = output_dir / f"{generated.stem}_{stamp}{generated.suffix}"
    shutil.move(str(generated), target)
    return target


def choose_source_dir() -> Optional[Path]:
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    try:
        selected = filedialog.askdirectory(title=f"選擇{CLINIC_NAME}來源資料夾")
    finally:
        root.destroy()
    return Path(selected) if selected else None


def open_file(path: Path) -> None:
    try:
        if sys.platform.startswith("darwin"):
            subprocess.call(("open", str(path)))
        elif os.name == "nt":
            os.startfile(str(path))  # type: ignore[attr-defined]
        else:
            subprocess.call(("xdg-open", str(path)))
    except Exception:
        pass


def show_completion_message(
    output: Path,
    counters: Dict[str, Any],
    report: Optional[Path],
) -> None:
    try:
        import tkinter as tk
        from tkinter import messagebox

        root = tk.Tk()
        root.withdraw()
        try:
            if counters["unresolved"]:
                messagebox.showwarning(
                    "完成，但有 ID 待補",
                    f"正式 Excel 已產生：\n{output}\n\n"
                    f"警告：{int(counters['unresolved'])} 筆因 ID 無效或無法唯一取得，"
                    "已納入會員輸出，但 ID 仍待補正。\n\n"
                    + "\n".join(
                        f"{row['source_file']} 第 {row['source_row']} 列："
                        f"{row['name']}，ID={row.get('pid') or '空白'}"
                        for row in counters["unresolved_rows"]
                    )
                    + "\n\n"
                    f"待補 ID 報表：\n{report}",
                    parent=root,
                )
            else:
                messagebox.showinfo(
                    "完成",
                    f"正式 Excel 已產生：\n{output}",
                    parent=root,
                )
        finally:
            root.destroy()
    except Exception as exc:
        print(f"提示視窗無法顯示：{exc}")


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=f"{CLINIC_NAME}前置清洗與選會員輸出")
    parser.add_argument("source_dir", nargs="?", help="來源資料夾；未指定時開啟選擇視窗")
    parser.add_argument("--output-dir", help="輸出資料夾，預設為來源資料夾上一層")
    parser.add_argument("--no-open", action="store_true", help="完成後不自動開啟 Excel")
    parser.add_argument("--no-dialog", action="store_true", help="完成後不顯示提示視窗")
    args = parser.parse_args(argv)

    source_dir = Path(args.source_dir).expanduser().resolve() if args.source_dir else choose_source_dir()
    if source_dir is None:
        print("未選擇來源資料夾，程式已取消。")
        return 0
    if not source_dir.is_dir():
        raise FileNotFoundError(f"找不到來源資料夾：{source_dir}")

    output_dir = (
        Path(args.output_dir).expanduser().resolve()
        if args.output_dir
        else source_dir.parent
    )
    with tempfile.TemporaryDirectory(prefix="clinic_merge_hezili_") as tmp:
        clean_dir, counters, report = prepare_clean_dir(source_dir, Path(tmp))
        output = run_common_core(clean_dir, output_dir)
    print(
        f"完成：{output}\n"
        f"會員 {int(counters['members'])} 筆；"
        f"L={counters['count_l']:.0f}、M={counters['amount_m']:.0f}、"
        f"N={counters['count_n']:.0f}、O={counters['amount_o']:.0f}；"
        f"純E1={int(counters['pure_e1'])}、純E2={int(counters['pure_e2'])}、"
        f"E1/E2={int(counters['overlap_e1_e2'])}、"
        f"自選名單={int(counters['self_output'])}；"
        f"K={int(counters['last_visit'])}、Q={int(counters['adult'])}、"
        f"R={int(counters['pap'])}、T={int(counters['fit'])}。"
    )
    if counters["unresolved"]:
        print(
            f"警告：{int(counters['unresolved'])} 筆 ID 異常資料已納入會員輸出，"
            "但 ID 仍待補正；"
            f"待補 ID 報表：{report}"
        )
        print("待補 ID 資料：")
        for row in counters["unresolved_rows"]:
            print(
                f"  - {row['source_file']} 第 {row['source_row']} 列｜"
                f"姓名：{row['name']}｜原 ID：{row.get('pid') or '空白'}｜"
                f"原因：{row['reason']}"
            )
    if not args.no_dialog:
        show_completion_message(output, counters, report)
    if not args.no_open:
        open_file(output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
