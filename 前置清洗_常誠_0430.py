# -*- coding: utf-8 -*-
"""
常誠前置清洗 + 通用主程式包裝

用途：
- 將常誠主名單的雙工作表來源拆成乾淨的 ascvd / 會員名單
- 移除會誤判成會員名單的百分位計算工作表干擾
- 再呼叫 run_merge_通用（自動偵測最新版）完成輸出
"""

from __future__ import annotations

import importlib.util
import os
import shutil
import sys
import tempfile
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent


def _find_generic_script(script_dir: Path) -> Path:
    candidates = sorted(script_dir.glob("run_merge_通用_*.py"), reverse=True)
    if not candidates:
        raise RuntimeError("找不到通用程式 run_merge_通用_*.py")
    return candidates[0]

GENERIC_SCRIPT = _find_generic_script(SCRIPT_DIR)


def _load_generic_module():
    spec = importlib.util.spec_from_file_location("run_merge_generic", GENERIC_SCRIPT)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"無法載入通用程式：{GENERIC_SCRIPT}")
    mod = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = mod
    spec.loader.exec_module(mod)
    return mod


def _copy_sheet_values(src_path: Path, sheet_name: str, out_path: Path, target_title: str) -> None:
    wb_src = load_workbook(src_path, data_only=True, read_only=True)
    try:
        ws_src = wb_src[sheet_name]
        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = target_title
        for row in ws_src.iter_rows(values_only=True):
            ws_out.append(list(row))
        wb_out.save(out_path)
    finally:
        wb_src.close()


def _sanitize_member_roster(source_dir: Path) -> bool:
    candidates = sorted(source_dir.glob("*需照護名單*.xlsx"))
    if not candidates:
        candidates = sorted(source_dir.glob("*需照護名單*.xls"))
    if not candidates:
        return False

    roster_path = candidates[0]
    wb = load_workbook(roster_path, data_only=True, read_only=True)
    try:
        first_sheet = wb.sheetnames[0]
    finally:
        wb.close()

    _copy_sheet_values(roster_path, first_sheet, source_dir / "常誠_ascvd_補正.xlsx", "ascvd")
    _copy_sheet_values(roster_path, first_sheet, source_dir / "常誠_會員名單_補正.xlsx", "會員名單")
    roster_path.unlink(missing_ok=True)
    return True


def process_excel(source_path: str, template_path: Optional[str] = None) -> str:
    generic = _load_generic_module()
    source_dir = Path(source_path).resolve()
    if not source_dir.is_dir():
        raise ValueError("請選擇資料夾。")

    template = template_path or generic._find_template(str(SCRIPT_DIR))
    if not os.path.exists(template):
        raise ValueError(f"找不到模板檔：{template}")

    temp_root = Path(tempfile.mkdtemp(prefix="chang_clean_"))
    temp_source = temp_root / source_dir.name

    try:
        shutil.copytree(source_dir, temp_source)
        _sanitize_member_roster(temp_source)
        temp_output = Path(generic.process_excel(str(temp_source), template))
        final_output = source_dir.parent / temp_output.name
        if final_output.exists():
            final_output.unlink()
        shutil.move(str(temp_output), str(final_output))
        return str(final_output)
    finally:
        shutil.rmtree(temp_root, ignore_errors=True)


def main() -> None:
    generic = _load_generic_module()
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()

    src = filedialog.askdirectory(title="選擇常誠來源資料夾")
    if not src:
        return

    template = generic._find_template(str(SCRIPT_DIR))

    try:
        out = process_excel(src, template)
        messagebox.showinfo("完成", f"已輸出：\n{out}")
        generic.open_file_cross_platform(out)
    except Exception as e:
        messagebox.showerror("錯誤", str(e))


if __name__ == "__main__":
    main()
