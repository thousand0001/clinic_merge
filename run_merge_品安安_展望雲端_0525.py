#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
品安安診所展望雲端專用選會員產檔。

輸出比照「選會員模板」的會員總表，並額外加上：
- 病歷號
- 備註

品安安展望資料夾目前包含：
1. A115自選會員.XLSX：自選會員基本資料
2. 115X不要會員.XLSX：不要會員名單，讀取時排除
3. R11440/*.XLSX：門診診療次數月報表，直接用身分證號歸戶並累計次數與總額
"""

from __future__ import annotations

import argparse
import copy
import datetime as dt
import os
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter

import run_merge_通用_0430_1 as generic


TEMPLATE_GLOB = "選會員模板*.xlsx"
TARGET_SHEET = "會員總表"
PERCENTILE_SHEET_NAME = "百分位名單"
DOCTOR_SHEET_NAME = "醫生看(從會員指標內容Key過來)"
SELF_SELECT_SHEET_NAME = "自選名單(從會員指標內容Key過來)"
DOCTOR_FORMAT_REFERENCE = Path(
    "/Users/thousand0001/CloudStation/醫療群/梓寧給的資料/run_merge跑資料/方鼎系統/芝妍皮膚專科診所選會員_0525_1900.xlsx"
)
SCREENING_COLUMNS = {
    "成人健檢": "成人健檢\n最後篩檢日",
    "子宮抹片": "子宮抹片最後篩檢日",
    "老人流感": "老人流感最後注射日",
    "糞便潛血": "糞便潛血最後篩檢日",
    "BC肝炎": "BC肝炎最篩檢日期",
}


class TableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.rows: List[List[str]] = []
        self._row: Optional[List[str]] = None
        self._cell: Optional[List[str]] = None
        self._in_cell = False

    def handle_starttag(self, tag: str, attrs: Sequence[Tuple[str, Optional[str]]]) -> None:
        tag = tag.lower()
        if tag == "tr":
            self._row = []
        elif tag in ("td", "th") and self._row is not None:
            self._cell = []
            self._in_cell = True

    def handle_data(self, data: str) -> None:
        if self._in_cell and self._cell is not None:
            self._cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag in ("td", "th") and self._in_cell:
            text = " ".join("".join(self._cell or []).split())
            if self._row is not None:
                self._row.append(text)
            self._cell = None
            self._in_cell = False
        elif tag == "tr" and self._row is not None:
            if any(str(value).strip() for value in self._row):
                self.rows.append(self._row)
            self._row = None


@dataclass
class MemberRecord:
    name: str = ""
    pid: str = ""
    bday: Optional[dt.date] = None
    charts: Set[str] = field(default_factory=set)
    sex: str = ""
    clinic: str = ""
    ascvd: Any = None
    dmk_raw: Any = None
    health: Dict[str, Any] = field(default_factory=dict)
    p4p: Dict[str, Any] = field(default_factory=dict)
    screenings: Dict[str, Any] = field(default_factory=dict)
    count_114: float = 0.0
    count_114_q1: float = 0.0
    count_115: float = 0.0
    count_115_q1: float = 0.0
    amount_114: float = 0.0
    amount_114_q1: float = 0.0
    amount_115: float = 0.0
    amount_115_q1: float = 0.0
    last_visit: Optional[dt.date] = None
    notes: Set[str] = field(default_factory=set)
    sources: Set[str] = field(default_factory=set)
    is_self_select: bool = False


def normalize_name(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).upper()


def display_text(value: Any) -> str:
    return str(value or "").strip()


def normalize_id(value: Any) -> str:
    return str(value or "").strip().upper().replace(" ", "")


def is_valid_id(value: Any) -> bool:
    text = normalize_id(value)
    return bool(
        re.fullmatch(r"[A-Z][12]\d{8}", text)
        or re.fullmatch(r"[A-Z][A-Z0-9]\d{8}", text)
        or re.fullmatch(r"[A-Z]{1,2}\d{8,10}", text)
    )


def normalize_chart(value: Any) -> str:
    text = str(value or "").strip()
    return text.zfill(7) if text.isdigit() else text


def parse_number(value: Any) -> float:
    try:
        return float(str(value or "0").replace(",", "").strip() or 0)
    except Exception:
        return 0.0


def parse_roc_or_ad_date(value: Any) -> Optional[dt.date]:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = str(value).strip()
    if not text or text == "-":
        return None
    compact = re.sub(r"\D", "", text)
    if re.fullmatch(r"\d{6,8}", compact):
        parsed = parse_compact_roc_or_ad_date(compact)
        if parsed:
            return parsed
    text = text.split(" ")[0].replace("-", "/")
    match = re.fullmatch(r"(\d{2,4})/(\d{1,2})/(\d{1,2})", text)
    if not match:
        return None
    year = int(match.group(1))
    if year < 1911:
        year += 1911
    try:
        return dt.date(year, int(match.group(2)), int(match.group(3)))
    except ValueError:
        return None


def parse_compact_roc_or_ad_date(value: Any) -> Optional[dt.date]:
    text = str(value or "").strip()
    if not re.fullmatch(r"\d{6,8}", text):
        return None
    if len(text) == 7:
        year = int(text[:3]) + 1911
        month = int(text[3:5])
        day = int(text[5:7])
    elif len(text) == 8:
        year = int(text[:4])
        month = int(text[4:6])
        day = int(text[6:8])
        if year < 1911:
            year += 1911
    else:
        year = int(text[:2]) + 1911
        month = int(text[2:4])
        day = int(text[4:6])
    try:
        return dt.date(year, month, day)
    except ValueError:
        return None


def parse_roc_yyyymmdd(value: Any) -> Optional[dt.date]:
    text = str(value or "").strip()
    return parse_compact_roc_or_ad_date(text)


def format_date(value: Any) -> Any:
    parsed = parse_roc_or_ad_date(value)
    return parsed if parsed else None


def read_html_table(path: Path) -> List[List[str]]:
    parser = TableParser()
    parser.feed(path.read_text(encoding="utf-8-sig", errors="replace"))
    return parser.rows


def iter_workbook_rows(path: Path) -> Iterable[Tuple[str, List[List[Any]]]]:
    if path.suffix.lower() == ".ods":
        wb = generic._load_ods_as_workbook(str(path))
    else:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    for ws in wb.worksheets:
        rows: List[List[Any]] = []
        for row in ws.iter_rows(values_only=True):
            values = list(row)
            if any(value is not None and str(value).strip() != "" for value in values):
                rows.append(values)
        yield ws.title, rows


def compact_header(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).upper()


def find_col(header: Sequence[Any], aliases: Sequence[str]) -> Optional[int]:
    targets = {compact_header(alias) for alias in aliases}
    for index, value in enumerate(header):
        if compact_header(value) in targets:
            return index
    return None


def find_header(rows: Sequence[Sequence[Any]], required_groups: Sequence[Sequence[str]], scan_rows: int = 30) -> Optional[int]:
    for index, row in enumerate(rows[:scan_rows]):
        if all(find_col(row, group) is not None for group in required_groups):
            return index
    return None


def find_template() -> Path:
    candidates = sorted(Path(__file__).resolve().parent.glob(TEMPLATE_GLOB))
    candidates = [path for path in candidates if path.suffix.lower() == ".xlsx"]
    if not candidates:
        raise FileNotFoundError(f"找不到模板：{TEMPLATE_GLOB}")
    return candidates[-1]


def infer_gender(pid: str) -> str:
    text = normalize_id(pid)
    if len(text) >= 2:
        if text[1] == "1":
            return "男"
        if text[1] == "2":
            return "女"
    return ""


def age_at_today(bday: Optional[dt.date], today: dt.date) -> Optional[int]:
    if not bday:
        return None
    age = today.year - bday.year - ((today.month, today.day) < (bday.month, bday.day))
    return age if age >= 0 else None


def sheet_header_map(ws) -> Dict[str, int]:
    return {compact_header(ws.cell(1, col).value): col for col in range(1, ws.max_column + 1)}


def get_or_create_by_id(records: Dict[str, MemberRecord], pid: str) -> Optional[MemberRecord]:
    pid = normalize_id(pid)
    if not is_valid_id(pid):
        return None
    key = f"ID:{pid}"
    rec = records.setdefault(key, MemberRecord(pid=pid))
    rec.pid = pid
    return rec


def get_or_create_by_name(records: Dict[str, MemberRecord], name: str) -> Optional[MemberRecord]:
    norm = normalize_name(name)
    if not norm:
        return None
    key = f"NAME:{norm}"
    rec = records.setdefault(key, MemberRecord(name=display_text(name)))
    if not rec.name:
        rec.name = display_text(name)
    return rec


def add_identity(
    records: Dict[str, MemberRecord],
    name_to_ids: Dict[str, Set[str]],
    chart_to_ids: Dict[str, Set[str]],
    *,
    name: Any = "",
    pid: Any = "",
    chart: Any = "",
    bday: Any = "",
    source: str = "",
) -> Optional[MemberRecord]:
    text_pid = normalize_id(pid)
    rec = get_or_create_by_id(records, text_pid) if is_valid_id(text_pid) else get_or_create_by_name(records, display_text(name))
    if rec is None:
        return None

    if display_text(name) and not rec.name:
        rec.name = display_text(name)
    if is_valid_id(text_pid):
        rec.pid = text_pid
        if normalize_name(name):
            name_to_ids[normalize_name(name)].add(text_pid)
    parsed_bday = parse_roc_or_ad_date(bday)
    if parsed_bday and not rec.bday:
        rec.bday = parsed_bday
    chart_text = normalize_chart(chart)
    if chart_text:
        rec.charts.add(chart_text)
        if is_valid_id(text_pid):
            chart_to_ids[chart_text].add(text_pid)
    if source:
        rec.sources.add(source)
    return rec


def merge_name_records(records: Dict[str, MemberRecord], name_to_ids: Dict[str, Set[str]]) -> None:
    for key in [key for key in records if key.startswith("NAME:")]:
        rec = records[key]
        ids = name_to_ids.get(normalize_name(rec.name), set())
        if len(ids) != 1:
            continue
        target = get_or_create_by_id(records, next(iter(ids)))
        if target is None or target is rec:
            continue
        if not target.name:
            target.name = rec.name
        if not target.bday:
            target.bday = rec.bday
        target.charts.update(rec.charts)
        target.screenings.update({k: v for k, v in rec.screenings.items() if k not in target.screenings})
        target.health.update({k: v for k, v in rec.health.items() if k not in target.health})
        target.p4p.update({k: v for k, v in rec.p4p.items() if k not in target.p4p})
        target.count_114 += rec.count_114
        target.count_114_q1 += rec.count_114_q1
        target.count_115 += rec.count_115
        target.count_115_q1 += rec.count_115_q1
        target.amount_114 += rec.amount_114
        target.amount_114_q1 += rec.amount_114_q1
        target.amount_115 += rec.amount_115
        target.amount_115_q1 += rec.amount_115_q1
        if rec.last_visit and (target.last_visit is None or rec.last_visit > target.last_visit):
            target.last_visit = rec.last_visit
        target.notes.update(rec.notes)
        target.sources.update(rec.sources)
        del records[key]


def scan_root_files(
    source_dir: Path,
    records: Dict[str, MemberRecord],
    name_to_ids: Dict[str, Set[str]],
    chart_to_ids: Dict[str, Set[str]],
) -> None:
    for path in sorted(source_dir.iterdir()):
        if not path.is_file() or path.name.startswith("~$") or path.suffix.lower() not in (".xlsx", ".ods"):
            continue
        if is_excluded_source(path.stem):
            continue
        source_is_self_select = is_self_select_source(path.stem)
        for sheet_name, rows in iter_workbook_rows(path):
            if is_excluded_source(sheet_name):
                continue
            sheet_is_self_select = source_is_self_select or is_self_select_source(sheet_name)
            if not rows:
                continue
            header_index = find_header(rows, [["姓名", "會員姓名", "病患姓名"], ["ID", "身分證", "身分證號", "身分證號碼", "身份證號", "身份證號碼", "家醫收案會員ID"]])
            if header_index is None:
                continue
            header = rows[header_index]
            name_col = find_col(header, ["姓名", "會員姓名", "病患姓名"])
            id_col = find_col(header, ["ID", "身分證", "身分證號", "身分證號碼", "身份證", "身份證號", "身份證號碼", "家醫收案會員ID"])
            bday_col = find_col(header, ["生日", "BIRTHDAY"])
            indicator_col = find_col(header, ["指標名稱"])
            ascvd_col = find_col(header, ["ASCVD", "ascvd"])
            disease_col = find_col(header, ["疾病樣態"])
            p4p_plan_col = find_col(header, ["P4P收案計畫"])
            p4p_status_col = find_col(header, ["收案狀態"])
            hba_val_col = find_col(header, ["最近一次HbA1c檢查結果(%)"])
            hba_dt_col = find_col(header, ["最近一次HbA1c檢查日期"])
            ldl_val_col = find_col(header, ["最近一次LDL檢查結果(mg/dL)"])
            ldl_dt_col = find_col(header, ["最近一次LDL檢查日期"])
            uacr_val_col = find_col(header, ["最近一次UACR檢查結果(mg/gm)", "最近一次UACR檢查結果"])
            uacr_dt_col = find_col(header, ["最近一次UACR檢查日期"])
            last_screen_col = find_col(header, ["最後篩檢日期"])
            if name_col is None or id_col is None:
                continue

            for row in rows[header_index + 1:]:
                if len(row) <= max(name_col, id_col):
                    continue
                name = row[name_col]
                pid = row[id_col]
                if not display_text(name) and not is_valid_id(pid):
                    continue
                bday = row[bday_col] if bday_col is not None and len(row) > bday_col else None
                rec = add_identity(
                    records,
                    name_to_ids,
                    chart_to_ids,
                    name=name,
                    pid=pid,
                    bday=bday,
                    source=path.name,
                )
                if rec is None:
                    continue
                if sheet_is_self_select:
                    rec.is_self_select = True
                if ascvd_col is not None and len(row) > ascvd_col and row[ascvd_col] not in (None, ""):
                    rec.ascvd = row[ascvd_col]
                if disease_col is not None and len(row) > disease_col and row[disease_col] not in (None, ""):
                    rec.dmk_raw = row[disease_col]
                if p4p_plan_col is not None and len(row) > p4p_plan_col:
                    rec.p4p["plan"] = row[p4p_plan_col]
                if p4p_status_col is not None and len(row) > p4p_status_col:
                    rec.p4p["status"] = row[p4p_status_col]
                if hba_val_col is not None and len(row) > hba_val_col:
                    rec.health["hba_val"] = row[hba_val_col]
                if hba_dt_col is not None and len(row) > hba_dt_col:
                    rec.health["hba_dt"] = parse_roc_or_ad_date(row[hba_dt_col])
                if ldl_val_col is not None and len(row) > ldl_val_col:
                    rec.health["ldl_val"] = row[ldl_val_col]
                if ldl_dt_col is not None and len(row) > ldl_dt_col:
                    rec.health["ldl_dt"] = parse_roc_or_ad_date(row[ldl_dt_col])
                if uacr_val_col is not None and len(row) > uacr_val_col:
                    rec.health["uacr_val"] = row[uacr_val_col]
                if uacr_dt_col is not None and len(row) > uacr_dt_col:
                    rec.health["uacr_dt"] = parse_roc_or_ad_date(row[uacr_dt_col])
                if indicator_col is not None and len(row) > indicator_col:
                    indicator = str(row[indicator_col] or "")
                    last_dt = row[last_screen_col] if last_screen_col is not None and len(row) > last_screen_col else None
                    if "成人預防保健" in indicator or "成人健檢" in indicator:
                        rec.screenings["成人健檢"] = parse_roc_or_ad_date(last_dt) or "-"
                    elif "子宮頸抹片" in indicator or "子宮抹片" in indicator:
                        rec.screenings["子宮抹片"] = parse_roc_or_ad_date(last_dt) or "-"
                    elif "流感" in indicator:
                        rec.screenings["老人流感"] = parse_roc_or_ad_date(last_dt) or "-"
                    elif "糞便潛血" in indicator or "潛血" in indicator:
                        rec.screenings["糞便潛血"] = parse_roc_or_ad_date(last_dt) or "-"
                    elif "肝炎" in indicator or "B、C肝" in indicator or "BC肝" in indicator.upper():
                        rec.screenings["BC肝炎"] = parse_roc_or_ad_date(last_dt) or "-"


def scan_r11440_files(
    source_dir: Path,
    records: Dict[str, MemberRecord],
    name_to_ids: Dict[str, Set[str]],
    chart_to_ids: Dict[str, Set[str]],
) -> None:
    report_dir = source_dir / "R11440"
    if not report_dir.is_dir():
        return
    for path in sorted(report_dir.iterdir()):
        if not path.is_file() or path.name.startswith("~$") or path.suffix.lower() != ".xlsx":
            continue
        for _sheet_name, rows in iter_workbook_rows(path):
            if not rows:
                continue
            header_index = find_header(rows, [["掛號證", "掛號証"], ["日期"], ["姓名"], ["身分證號", "身份證號"], ["次數"], ["總額"]])
            if header_index is None:
                continue
            header = rows[header_index]
            chart_col = find_col(header, ["掛號證", "掛號証", "病歷號", "病歷號碼"])
            date_col = find_col(header, ["日期", "看診日期"])
            name_col = find_col(header, ["姓名", "病患姓名"])
            id_col = find_col(header, ["身分證號", "身份證號", "身分證", "身份證", "ID"])
            bday_col = find_col(header, ["生日", "出生日期", "出生年月日"])
            count_col = find_col(header, ["次數", "看診次數", "就診次數"])
            amount_col = find_col(header, ["總額", "申報總額", "實際申報總額"])
            if None in (chart_col, date_col, name_col, id_col, count_col, amount_col):
                continue
            for row in rows[header_index + 1:]:
                if len(row) <= max(chart_col, date_col, name_col, id_col, count_col, amount_col):
                    continue
                visit_date = parse_roc_yyyymmdd(row[date_col])
                if visit_date is None:
                    continue
                rec = add_identity(
                    records,
                    name_to_ids,
                    chart_to_ids,
                    chart=row[chart_col],
                    name=row[name_col],
                    pid=row[id_col],
                    bday=row[bday_col] if bday_col is not None and len(row) > bday_col else None,
                    source=f"R11440/{path.name}",
                )
                if rec is None:
                    continue
                count = parse_number(row[count_col])
                amount = parse_number(row[amount_col])
                year_bucket = visit_date.year - 1911
                month = visit_date.month
                if year_bucket == 114:
                    rec.count_114 += count
                    rec.amount_114 += amount
                    if month <= 4:
                        rec.count_114_q1 += count
                        rec.amount_114_q1 += amount
                elif year_bucket == 115:
                    rec.count_115 += count
                    rec.amount_115 += amount
                    if month <= 4:
                        rec.count_115_q1 += count
                        rec.amount_115_q1 += amount
                if rec.last_visit is None or visit_date > rec.last_visit:
                    rec.last_visit = visit_date
                rec.notes.add("R11440月報表")


def scan_count_files(
    source_dir: Path,
    records: Dict[str, MemberRecord],
    name_to_ids: Dict[str, Set[str]],
    chart_to_ids: Dict[str, Set[str]],
) -> None:
    count_dir = source_dir / "次數"
    if not count_dir.is_dir():
        return
    for path in sorted(count_dir.glob("*.xls")):
        month_match = re.fullmatch(r"1(14|15)(\d{2})", path.stem)
        if not month_match:
            continue
        year_bucket = int(path.stem[:3])
        month = int(path.stem[3:5])
        rows = read_html_table(path)
        header_index = find_header(rows, [["病歷號"], ["身分證", "身分證號", "身份證號"], ["看診次數", "次數"]])
        if header_index is None:
            continue
        header = rows[header_index]
        chart_col = find_col(header, ["病歷號", "病歷號碼"])
        name_col = find_col(header, ["姓名", "病患姓名"])
        id_col = find_col(header, ["身分證", "身分證號", "身份證號", "ID"])
        bday_col = find_col(header, ["生日"])
        count_col = find_col(header, ["看診次數", "次數", "就診次數"])
        if None in (chart_col, name_col, id_col, count_col):
            continue
        for row in rows[header_index + 1:]:
            if len(row) <= max(chart_col, name_col, id_col, count_col):
                continue
            rec = add_identity(
                records,
                name_to_ids,
                chart_to_ids,
                chart=row[chart_col],
                name=row[name_col],
                pid=row[id_col],
                bday=row[bday_col] if bday_col is not None and len(row) > bday_col else None,
                source=f"次數/{path.name}",
            )
            if rec is None:
                continue
            count = parse_number(row[count_col])
            if year_bucket == 114:
                rec.count_114 += count
                if month <= 4:
                    rec.count_114_q1 += count
            elif year_bucket == 115:
                rec.count_115 += count
                if month <= 4:
                    rec.count_115_q1 += count
            rec.notes.add("病歷號對到ID")


def attach_fee_record(
    records: Dict[str, MemberRecord],
    name_to_ids: Dict[str, Set[str]],
    chart_to_ids: Dict[str, Set[str]],
    *,
    chart: str,
    name: str,
    visit_date: Optional[dt.date],
    amount: float,
    source: str,
) -> MemberRecord:
    chart_ids = chart_to_ids.get(chart, set())
    rec: Optional[MemberRecord] = None
    if len(chart_ids) == 1:
        rec = get_or_create_by_id(records, next(iter(chart_ids)))
        if rec:
            rec.notes.add("病歷號對到ID")
    elif len(chart_ids) > 1:
        rec = get_or_create_by_name(records, name)
        if rec:
            rec.notes.add("病歷號對到多個ID")
    else:
        name_ids = name_to_ids.get(normalize_name(name), set())
        if len(name_ids) == 1:
            rec = get_or_create_by_id(records, next(iter(name_ids)))
            if rec:
                rec.notes.add("姓名唯一對到ID")
        elif len(name_ids) > 1:
            rec = get_or_create_by_name(records, name)
            if rec:
                rec.notes.add("姓名同名多人，未自動帶ID")
        else:
            rec = get_or_create_by_name(records, name)
            if rec:
                rec.notes.add("未對到ID")
    if rec is None:
        rec = MemberRecord(name=name)
        records[f"FEE:{source}:{chart}:{name}:{len(records)}"] = rec
        rec.notes.add("未對到ID")

    rec.name = rec.name or name
    if chart:
        rec.charts.add(chart)
    if visit_date and (rec.last_visit is None or visit_date > rec.last_visit):
        rec.last_visit = visit_date
    if visit_date:
        year_bucket = visit_date.year - 1911
        month = visit_date.month
        if year_bucket == 114:
            rec.amount_114 += amount
            if month <= 4:
                rec.amount_114_q1 += amount
        elif year_bucket == 115:
            rec.amount_115 += amount
            if month <= 4:
                rec.amount_115_q1 += amount
    rec.sources.add(source)
    return rec


def scan_fee_files(
    source_dir: Path,
    records: Dict[str, MemberRecord],
    name_to_ids: Dict[str, Set[str]],
    chart_to_ids: Dict[str, Set[str]],
) -> None:
    fee_dir = source_dir / "費用"
    if not fee_dir.is_dir():
        return
    for path in sorted(fee_dir.glob("*.xls")):
        rows = read_html_table(path)
        header_index = find_header(rows, [["病歷號"], ["姓名"], ["掛帳費"]])
        if header_index is None:
            continue
        header = rows[header_index]
        chart_col = find_col(header, ["病歷號", "病歷號碼"])
        name_col = find_col(header, ["姓名", "病患姓名"])
        date_col = find_col(header, ["日期"])
        amount_col = find_col(header, ["掛帳費"])
        if None in (chart_col, name_col, date_col, amount_col):
            continue
        for row in rows[header_index + 1:]:
            if len(row) <= max(chart_col, name_col, date_col, amount_col):
                continue
            raw_date = str(row[date_col] or "").strip()
            if not re.fullmatch(r"1(?:14|15)\d{4}", raw_date):
                continue
            chart = normalize_chart(row[chart_col])
            name = display_text(row[name_col])
            visit_date = parse_roc_yyyymmdd(raw_date)
            amount = parse_number(row[amount_col])
            attach_fee_record(
                records,
                name_to_ids,
                chart_to_ids,
                chart=chart,
                name=name,
                visit_date=visit_date,
                amount=amount,
                source=f"費用/{path.name}",
            )


def collect_data(source_dir: Path) -> Dict[str, MemberRecord]:
    records: Dict[str, MemberRecord] = {}
    name_to_ids: Dict[str, Set[str]] = defaultdict(set)
    chart_to_ids: Dict[str, Set[str]] = defaultdict(set)
    scan_root_files(source_dir, records, name_to_ids, chart_to_ids)
    scan_r11440_files(source_dir, records, name_to_ids, chart_to_ids)
    scan_count_files(source_dir, records, name_to_ids, chart_to_ids)
    scan_fee_files(source_dir, records, name_to_ids, chart_to_ids)
    merge_name_records(records, name_to_ids)
    for rec in records.values():
        if not rec.pid:
            rec.notes.add("ID欄填-")
        if not rec.bday:
            rec.notes.add("缺生日")
        if rec.pid:
            rec.sex = infer_gender(rec.pid)
    return records


def is_self_select_source(text: Any) -> bool:
    compact = re.sub(r"[\s\-_()（）\[\]{}]+", "", str(text or "").strip()).lower()
    if not compact:
        return False
    if "115x" in compact or "不選" in compact or "不要" in compact:
        return False
    return any(token in compact for token in ("自選名單", "自選會員", "115自選", "a115"))


def is_excluded_source(text: Any) -> bool:
    compact = re.sub(r"[\s\-_()（）\[\]{}]+", "", str(text or "").strip()).lower()
    return bool(compact and ("115x" in compact or "不選" in compact or "不要" in compact))


def style_like_previous_row(ws, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy.copy(src._style)
        dst.number_format = src.number_format
        dst.font = copy.copy(src.font)
        dst.fill = copy.copy(src.fill)
        dst.border = copy.copy(src.border)
        dst.alignment = copy.copy(src.alignment)
        dst.protection = copy.copy(src.protection)


def style_member_total_row(ws, target_row: int, max_col: int) -> None:
    style_like_previous_row(ws, 2, target_row, max_col)
    for col in range(1, max_col + 1):
        ws.cell(target_row, col).fill = PatternFill(fill_type=None)
    generic._apply_member_row_style(ws, target_row, max_col)


def ensure_extra_columns(ws) -> Tuple[int, int, int]:
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    compact = {compact_header(value): index + 1 for index, value in enumerate(headers)}
    max_col = ws.max_column

    chart_col = compact.get(compact_header("病歷號"))
    if chart_col is None:
        max_col += 1
        chart_col = max_col
        ws.cell(1, chart_col).value = "病歷號"
        ws.cell(2, chart_col).value = ""

    note_col = compact.get(compact_header("備註"))
    if note_col is None:
        max_col += 1
        note_col = max_col
        ws.cell(1, note_col).value = "備註"
        ws.cell(2, note_col).value = ""

    self_select_col = compact.get(compact_header("是否為自選會員"))
    if self_select_col is None:
        max_col += 1
        self_select_col = max_col
        ws.cell(1, self_select_col).value = "是否為自選會員"
        ws.cell(2, self_select_col).value = ""

    for col in (chart_col, note_col, self_select_col):
        src = ws.cell(1, ws.max_column if ws.max_column < col else col - 1)
        ws.cell(1, col).font = Font(bold=True)
        ws.cell(1, col).fill = PatternFill("solid", fgColor="D9EAF7")
        ws.cell(1, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = 28 if col == note_col else 18
    return chart_col, note_col, self_select_col


def find_output_col(ws, aliases: Sequence[str]) -> Optional[int]:
    for row in (1, 2):
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row, col).value
            if any(compact_header(value) == compact_header(alias) for alias in aliases):
                return col
    return None


def set_if_col(ws, row: int, col: Optional[int], value: Any) -> None:
    if col:
        ws.cell(row, col).value = value


def number_or_blank(value: float) -> Any:
    return int(value) if value else None


def date_or_blank(value: Optional[dt.date]) -> Any:
    return value if value else None


def screening_value(rec: MemberRecord, key: str) -> Any:
    value = rec.screenings.get(key)
    if value in (None, "", "-"):
        return None
    return value


def disease_code_output(value: Any) -> Any:
    text = generic.normalize_text(value).upper()
    if not text:
        return value
    if "DKD" in text or "糖尿病腎" in text:
        return 3
    if "CKD" in text or "腎臟病" in text:
        return 2
    if "DM" in text or "糖尿病" in text:
        return 1
    if "ASCVD" in text:
        return 4
    return value


def load_doctor_reference_sheet():
    if not DOCTOR_FORMAT_REFERENCE.exists():
        return None
    wb = openpyxl.load_workbook(DOCTOR_FORMAT_REFERENCE)
    for name in wb.sheetnames:
        if "醫生看" in name:
            return wb[name]
    return None


def apply_doctor_reference_layout(ws, ref_ws) -> None:
    if ref_ws is None:
        return
    max_col = min(ref_ws.max_column, ws.max_column)
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row <= 3:
            ws.unmerge_cells(str(merged_range))
    for merged_range in ref_ws.merged_cells.ranges:
        if merged_range.min_row <= 3 and merged_range.max_col <= max_col:
            ws.merge_cells(str(merged_range))
    for row in range(1, 4):
        ws.row_dimensions[row].height = ref_ws.row_dimensions[row].height
        for col in range(1, max_col + 1):
            src = ref_ws.cell(row, col)
            dst = ws.cell(row, col)
            if dst.__class__.__name__ == "MergedCell":
                continue
            dst.value = src.value
            if src.has_style:
                dst._style = copy.copy(src._style)
            dst.number_format = src.number_format
            dst.font = copy.copy(src.font)
            dst.fill = copy.copy(src.fill)
            dst.border = copy.copy(src.border)
            dst.alignment = copy.copy(src.alignment)
            dst.protection = copy.copy(src.protection)
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = ref_ws.column_dimensions[letter].width
        ws.column_dimensions[letter].hidden = ref_ws.column_dimensions[letter].hidden


def write_doctor_row_values(ws, row_index: int, values_by_col: Dict[int, Any], ref_ws) -> None:
    max_col = max(ws.max_column, max(values_by_col.keys(), default=1))
    if ref_ws is not None:
        for col in range(1, max_col + 1):
            src = ref_ws.cell(4, col) if col <= ref_ws.max_column else ref_ws.cell(4, ref_ws.max_column)
            dst = ws.cell(row_index, col)
            if src.has_style:
                dst._style = copy.copy(src._style)
            dst.number_format = src.number_format
            dst.font = copy.copy(src.font)
            dst.fill = copy.copy(src.fill)
            dst.border = copy.copy(src.border)
            dst.alignment = copy.copy(src.alignment)
            dst.protection = copy.copy(src.protection)
    else:
        style_like_previous_row(ws, 3, row_index, max_col)
    for col, value in values_by_col.items():
        ws.cell(row_index, col).value = value


def parse_lab_float(value: Any) -> Optional[float]:
    try:
        if value in (None, "", "-"):
            return None
        return float(str(value).strip())
    except Exception:
        return None


def screening_display_value(rec: MemberRecord, key: str, today: dt.date) -> Any:
    value = rec.screenings.get(key)
    age = age_at_today(rec.bday, today)
    if value not in (None, "", "-"):
        parsed = parse_roc_or_ad_date(value)
        if parsed and parsed.year == today.year:
            return parsed
        if parsed:
            return "過期需受檢"
        return value

    if key == "成人健檢" and (age is not None and age < 40):
        return "不需受檢"
    if key == "子宮抹片" and (rec.sex == "男" or (age is not None and age < 30)):
        return "不需受檢"
    if key == "老人流感" and (age is not None and age < 65):
        return "不需受檢"
    if key == "糞便潛血" and (age is not None and (age < 50 or age > 74)):
        return "不需受檢"
    return "待受檢"


def lab_result_display(value: Any, date_value: Any, threshold: float) -> Any:
    parsed_value = parse_lab_float(value)
    if parsed_value is None:
        return None
    parsed_date = parse_roc_or_ad_date(date_value)
    if parsed_date and parsed_date.year == 2026 and parsed_value > threshold:
        return f"{parsed_value:g}\n(已受檢未達控制)"
    if (not parsed_date or parsed_date.year != 2026) and parsed_value > threshold:
        return f"{parsed_value:g}\n(2026需受檢)"
    return parsed_value


def lab_date_display(value: Any, date_value: Any, threshold: float) -> Any:
    parsed_date = parse_roc_or_ad_date(date_value)
    if not parsed_date:
        return None
    parsed_value = parse_lab_float(value)
    if parsed_date.year == 2026:
        return parsed_date
    if parsed_value is not None and parsed_value <= threshold:
        return f"{parsed_date.isoformat()}\n(不需受檢)"
    return parsed_date


def apply_doctor_status_styles(ws, row_index: int) -> None:
    pending_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")
    done_fill = PatternFill(fill_type="solid", fgColor="C6E0B4")
    no_fill = PatternFill(fill_type=None)
    for col in range(16, 21):
        cell = ws.cell(row_index, col)
        text = str(cell.value or "")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if "待受檢" in text or "過期需受檢" in text:
            cell.fill = pending_fill
        elif isinstance(cell.value, (dt.date, dt.datetime)):
            cell.fill = done_fill
        else:
            cell.fill = no_fill
    for value_col, date_col in ((21, 22), (23, 24)):
        value_cell = ws.cell(row_index, value_col)
        date_cell = ws.cell(row_index, date_col)
        if "未達控制" in str(value_cell.value or "") or "2026需受檢" in str(value_cell.value or ""):
            value_cell.font = copy.copy(value_cell.font)
            value_cell.font = Font(
                name=value_cell.font.name,
                sz=value_cell.font.sz,
                bold=value_cell.font.bold,
                italic=value_cell.font.italic,
                color="FF0000",
            )
            value_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if isinstance(date_cell.value, (dt.date, dt.datetime)) and date_cell.value.year == 2026:
            date_cell.fill = done_fill
        else:
            date_cell.fill = no_fill


def apply_doctor_prevention_fills(ws) -> None:
    pending_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")
    done_fill = PatternFill(fill_type="solid", fgColor="C6E0B4")
    no_fill = PatternFill(fill_type=None)
    for row in range(4, ws.max_row + 1):
        for col in range(16, 21):
            cell = ws.cell(row, col)
            text = str(cell.value or "")
            if "待受檢" in text or "過期需受檢" in text or "不確定" in text or "年齡未知" in text:
                cell.fill = pending_fill
            elif isinstance(cell.value, (dt.date, dt.datetime)):
                cell.fill = done_fill
            else:
                cell.fill = no_fill


def display_pid(rec: MemberRecord) -> str:
    return rec.pid or "-"


def avg_amount(total: float, months: int) -> Any:
    if not total or months <= 0:
        return None
    return round(total / months, 2)


def is_roc_year(value: Any, roc_year: int) -> bool:
    parsed = parse_roc_or_ad_date(value)
    return bool(parsed and parsed.year - 1911 == roc_year)


def has_screening_done(rec: MemberRecord, key: str) -> bool:
    value = rec.screenings.get(key)
    return value not in (None, "", "-")


def score_record(rec: MemberRecord) -> Tuple[int, str, str, Tuple[int, int, int, int]]:
    visit_score = 10 if (rec.count_115 >= 4 or rec.count_114 >= 6) else 0

    fee_score = 0
    avg_114_q1 = rec.amount_114_q1 / 4 if rec.amount_114_q1 else 0
    avg_115_q1 = rec.amount_115_q1 / 4 if rec.amount_115_q1 else 0
    if avg_114_q1 and avg_115_q1 and avg_115_q1 < avg_114_q1:
        fee_score = 6

    hba_2026 = is_roc_year(rec.health.get("hba_dt"), 115)
    ldl_2026 = is_roc_year(rec.health.get("ldl_dt"), 115)
    uacr_2026 = is_roc_year(rec.health.get("uacr_dt"), 115)
    exam_score = 10 if (hba_2026 or ldl_2026 or uacr_2026) else 0

    prevention_items = [
        ("成人預防保健", 6),
        ("子宮抹片", 6),
        ("老人流感", 4),
        ("糞便潛血", 6),
        ("BC肝炎", 6),
    ]
    prevention_score = sum(points for key, points in prevention_items if has_screening_done(rec, key))
    missing = [key for key, _ in prevention_items if not has_screening_done(rec, key)]
    reminder = "；".join(f"今年需檢測{key}" for key in missing) if missing else "-"

    total = visit_score + fee_score + exam_score + prevention_score
    breakdown = "\n".join([
        f"1. 固定就診次數：{visit_score} 分",
        f"2. 醫療費用：{fee_score} 分",
        f"3. 糖心腎管理：{exam_score} 分",
        f"4. 預防保健：{prevention_score} 分",
    ])
    return total, breakdown, reminder, (visit_score, fee_score, exam_score, prevention_score)


def clear_sheet_data(ws, start_row: int) -> None:
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def write_row_values(ws, row_index: int, values_by_col: Dict[int, Any], style_row: int) -> None:
    max_col = max(ws.max_column, max(values_by_col.keys(), default=1))
    style_like_previous_row(ws, style_row, row_index, max_col)
    for col, value in values_by_col.items():
        ws.cell(row_index, col).value = value


def build_screening_member_ids(records: Iterable[MemberRecord]) -> Dict[str, Set[str]]:
    mapping = {
        "adult": "成人健檢",
        "pap": "子宮抹片",
        "flu": "老人流感",
        "fit": "糞便潛血",
        "hep": "BC肝炎",
    }
    result: Dict[str, Set[str]] = {key: set() for key in mapping}
    for rec in records:
        if not rec.pid:
            continue
        for generic_key, local_key in mapping.items():
            if has_screening_done(rec, local_key):
                result[generic_key].add(rec.pid)
    return result


def apply_generic_derived_outputs(
    wb,
    ws,
    row_records: Dict[int, MemberRecord],
    data_start: int,
    last_row: int,
    today: dt.date,
) -> Tuple[Dict[str, Optional[int]], Tuple[str, str, str, str]]:
    cols = generic.detect_template_columns(ws, data_start)
    meta: Dict[int, generic.MemberMeta] = {}
    for row, rec in row_records.items():
        age = age_at_today(rec.bday, today)
        meta[row] = generic.MemberMeta(
            row=row,
            pid=rec.pid,
            bday=rec.bday,
            age=age if age is not None else -1,
            e_code=generic.parse_disease_code(disease_code_output(rec.dmk_raw)),
            ascvd=generic.parse_ascvd(rec.ascvd),
        )

    hba_candidates = generic._collect_hba_candidates(ws, cols, data_start, last_row)
    ldl_candidates = generic._collect_ldl_candidates(ws, cols, data_start, last_row)
    kpi_marks = generic.collect_kpi_mark_sets(
        ws,
        cols,
        data_start,
        last_row,
        hba_candidates=hba_candidates,
        ldl_candidates=ldl_candidates,
    )
    generic._compute_all_derived(
        ws,
        cols,
        meta,
        data_start,
        last_row,
        today,
        kpi_marks=kpi_marks,
        screening_member_ids=build_screening_member_ids(row_records.values()),
    )
    hba_main, hba_target = generic.calc_hba_kpi_ay_az(
        ws, cols, data_start, last_row, hba_candidates=hba_candidates
    )
    ldl_main, ldl_target = generic.calc_ldl_percentiles(
        ws, cols, data_start, last_row, ldl_candidates=ldl_candidates
    )
    generic._write_legacy_kpi_summary_cells(
        ws,
        hba_main_summary=hba_main,
        hba_target_summary=hba_target,
        ldl_main_summary=ldl_main,
        ldl_target_summary=ldl_target,
    )
    return cols, (hba_main, hba_target, ldl_main, ldl_target)


def populate_doctor_sheet(wb, records: List[MemberRecord], today: dt.date) -> None:
    if DOCTOR_SHEET_NAME not in wb.sheetnames:
        return
    ws = wb[DOCTOR_SHEET_NAME]
    ref_ws = load_doctor_reference_sheet()
    apply_doctor_reference_layout(ws, ref_ws)
    clear_sheet_data(ws, 4)
    for offset, rec in enumerate(records, start=4):
        total_score, breakdown, reminder, scores = score_record(rec)
        write_doctor_row_values(
            ws,
            offset,
            {
                1: display_pid(rec),
                2: rec.name or "-",
                3: date_or_blank(rec.bday),
                4: age_at_today(rec.bday, today),
                8: rec.dmk_raw,
                9: rec.ascvd,
                10: date_or_blank(rec.last_visit),
                11: number_or_blank(rec.count_114),
                12: number_or_blank(rec.count_114_q1),
                13: number_or_blank(rec.count_115),
                14: avg_amount(rec.amount_114, 12),
                15: avg_amount(rec.amount_115, 4),
                16: screening_display_value(rec, "成人健檢", today),
                17: screening_display_value(rec, "子宮抹片", today),
                18: screening_display_value(rec, "老人流感", today),
                19: screening_display_value(rec, "糞便潛血", today),
                20: screening_display_value(rec, "BC肝炎", today),
                21: lab_result_display(rec.health.get("hba_val"), rec.health.get("hba_dt"), 7.0),
                22: lab_date_display(rec.health.get("hba_val"), rec.health.get("hba_dt"), 7.0),
                23: lab_result_display(rec.health.get("ldl_val"), rec.health.get("ldl_dt"), 120.0),
                24: lab_date_display(rec.health.get("ldl_val"), rec.health.get("ldl_dt"), 120.0),
                25: rec.health.get("uacr_val"),
                26: rec.health.get("uacr_dt"),
                28: rec.p4p.get("status"),
                35: total_score,
                36: scores[0],
                37: scores[1],
                38: scores[2],
                39: scores[3],
                40: breakdown,
                41: reminder,
            },
            ref_ws,
        )
        apply_doctor_status_styles(ws, offset)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A4"


def populate_self_select_sheet(wb, records: List[MemberRecord]) -> None:
    if SELF_SELECT_SHEET_NAME not in wb.sheetnames:
        return
    ws = wb[SELF_SELECT_SHEET_NAME]
    score_headers = {
        22: "總分",
        23: "固定就診次數分",
        24: "醫療費用分",
        25: "糖心腎管理分",
        26: "預防保健分",
    }
    for col, header in score_headers.items():
        ws.cell(1, col).value = header
        if ws.cell(1, 21).has_style:
            ws.cell(1, col)._style = copy.copy(ws.cell(1, 21)._style)
        if ws.cell(2, 21).has_style:
            ws.cell(2, col)._style = copy.copy(ws.cell(2, 21)._style)
        ws.column_dimensions[get_column_letter(col)].hidden = True
    clear_sheet_data(ws, 3)
    selected_records = [rec for rec in records if rec.is_self_select]
    for offset, rec in enumerate(selected_records, start=3):
        total_score, _breakdown, _reminder, scores = score_record(rec)
        write_row_values(
            ws,
            offset,
            {
                1: rec.name or "-",
                2: display_pid(rec),
                3: number_or_blank(rec.count_114),
                4: number_or_blank(rec.count_114_q1),
                5: number_or_blank(rec.count_115),
                6: number_or_blank(rec.amount_114),
                7: number_or_blank(rec.amount_115),
                8: screening_value(rec, "成人健檢"),
                9: screening_value(rec, "子宮抹片"),
                10: screening_value(rec, "老人流感"),
                11: screening_value(rec, "糞便潛血"),
                12: screening_value(rec, "BC肝炎"),
                13: rec.health.get("hba_val"),
                14: rec.health.get("hba_dt"),
                15: rec.health.get("ldl_val"),
                16: rec.health.get("ldl_dt"),
                17: rec.p4p.get("status"),
                21: "✔" if rec.ascvd or rec.dmk_raw else None,
                22: total_score,
                23: scores[0],
                24: scores[1],
                25: scores[2],
                26: scores[3],
            },
            style_row=2,
        )
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A3"


def parse_lab_number(value: Any) -> Optional[float]:
    if value in (None, "", "-"):
        return None
    try:
        return float(str(value).strip())
    except Exception:
        return None


def populate_percentile_sheet(
    wb,
    cols: Dict[str, Optional[int]],
    data_start: int,
    last_row: int,
    summaries: Tuple[str, str, str, str],
) -> None:
    hba_main, hba_target, ldl_main, ldl_target = summaries
    generic.populate_percentile_sheet(
        wb,
        hba_main,
        hba_target,
        ldl_main,
        ldl_target,
        cols,
        data_start,
        last_row,
    )
    if PERCENTILE_SHEET_NAME in wb.sheetnames:
        wb[PERCENTILE_SHEET_NAME].sheet_view.showGridLines = True
        generic._finalize_percentile_sheet_alignment(wb[PERCENTILE_SHEET_NAME])


def create_output(source_dir: Path, output_dir: Optional[Path] = None) -> Path:
    records = collect_data(source_dir)
    template_path = find_template()
    wb = openpyxl.load_workbook(template_path)
    ws = wb[TARGET_SHEET]
    generic.prepare_template_layout(ws)
    chart_col, note_col, self_select_col = ensure_extra_columns(ws)

    cols = {
        "name": find_output_col(ws, ["姓名"]),
        "member": find_output_col(ws, ["會員"]),
        "id": find_output_col(ws, ["身份證號碼", "身分證號碼", "ID"]),
        "sex": find_output_col(ws, ["性別"]),
        "bday": find_output_col(ws, ["生日"]),
        "age": find_output_col(ws, ["年齡"]),
        "last_visit": find_output_col(ws, ["最後就診日"]),
        "count_114": find_output_col(ws, ["114年就診次數"]),
        "amount_114": find_output_col(ws, ["114年實際申報總額", "114年申報金額/月"]),
        "count_115": find_output_col(ws, ["115年就診次數"]),
        "amount_115": find_output_col(ws, ["115年實際申報總額", "115年申報金額/月"]),
        "count_114_q1": find_output_col(ws, ["114年1-4月就診次數"]),
        "count_115_q1": find_output_col(ws, ["115年1-4月就診次數"]),
        "avg_amount_114_hidden": find_output_col(ws, ["114年月平均"]),
        "avg_amount_115_hidden": find_output_col(ws, ["115年月平均"]),
        "ascvd": find_output_col(ws, ["ASCVD"]),
        "dmk": find_output_col(ws, ["疾病樣態分類(7類)", "疾病樣態編號"]),
        "hba_val": find_output_col(ws, ["最近一次HbA1c檢查結果(%)"]),
        "hba_dt": find_output_col(ws, ["最近一次HbA1c檢查日期"]),
        "ldl_val": find_output_col(ws, ["最近一次LDL\n檢查結果(mg/dL)", "最近一次LDL檢查結果(mg/dL)"]),
        "ldl_dt": find_output_col(ws, ["最近一次LDL\n檢查日期", "最近一次LDL檢查日期"]),
        "uacr_val": find_output_col(ws, ["最近一次UACR檢查結果", "最近一次UACR檢查結果(mg/gm)"]),
        "uacr_dt": find_output_col(ws, ["最近一次UACR\n檢查日期", "最近一次UACR檢查日期"]),
        "p4p_plan": find_output_col(ws, ["P4P收案計畫"]),
        "p4p_status": find_output_col(ws, ["收案狀態"]),
        "score": find_output_col(ws, ["分數"]),
        "breakdown": find_output_col(ws, ["分數說明"]),
        "prevention_note": find_output_col(ws, ["預防保健提醒", "備註"]),
        "is_self_select": self_select_col,
    }
    screening_cols = {key: find_output_col(ws, [header]) for key, header in SCREENING_COLUMNS.items()}

    data_start = 3
    if ws.max_row >= data_start:
        ws.delete_rows(data_start, ws.max_row - data_start + 1)

    today = dt.date.today()

    def sort_key(item: MemberRecord) -> Tuple[int, str, str]:
        return (0 if item.pid else 1, normalize_name(item.name), item.pid)

    sorted_records = sorted(records.values(), key=sort_key)

    row_index = data_start
    max_col = ws.max_column
    row_records: Dict[int, MemberRecord] = {}
    for rec in sorted_records:
        row_records[row_index] = rec
        style_member_total_row(ws, row_index, max_col)
        age = age_at_today(rec.bday, today)
        total_score, breakdown, reminder, _scores = score_record(rec)
        set_if_col(ws, row_index, cols["name"], rec.name or "-")
        set_if_col(ws, row_index, cols["member"], None)
        set_if_col(ws, row_index, cols["id"], rec.pid or "-")
        set_if_col(ws, row_index, cols["sex"], rec.sex or None)
        set_if_col(ws, row_index, cols["bday"], rec.bday)
        set_if_col(ws, row_index, cols["age"], age)
        set_if_col(ws, row_index, cols["last_visit"], rec.last_visit)
        set_if_col(ws, row_index, cols["count_114"], number_or_blank(rec.count_114))
        set_if_col(ws, row_index, cols["amount_114"], number_or_blank(rec.amount_114))
        set_if_col(ws, row_index, cols["count_115"], number_or_blank(rec.count_115))
        set_if_col(ws, row_index, cols["amount_115"], number_or_blank(rec.amount_115))
        set_if_col(ws, row_index, cols["count_114_q1"], number_or_blank(rec.count_114_q1))
        set_if_col(ws, row_index, cols["count_115_q1"], number_or_blank(rec.count_115_q1))
        set_if_col(ws, row_index, cols["avg_amount_114_hidden"], avg_amount(rec.amount_114_q1, 4))
        set_if_col(ws, row_index, cols["avg_amount_115_hidden"], avg_amount(rec.amount_115_q1, 4))
        set_if_col(ws, row_index, cols["ascvd"], rec.ascvd)
        set_if_col(ws, row_index, cols["dmk"], disease_code_output(rec.dmk_raw))
        set_if_col(ws, row_index, cols["hba_val"], rec.health.get("hba_val"))
        set_if_col(ws, row_index, cols["hba_dt"], rec.health.get("hba_dt"))
        set_if_col(ws, row_index, cols["ldl_val"], rec.health.get("ldl_val"))
        set_if_col(ws, row_index, cols["ldl_dt"], rec.health.get("ldl_dt"))
        set_if_col(ws, row_index, cols["uacr_val"], rec.health.get("uacr_val"))
        set_if_col(ws, row_index, cols["uacr_dt"], rec.health.get("uacr_dt"))
        set_if_col(ws, row_index, cols["p4p_plan"], rec.p4p.get("plan"))
        set_if_col(ws, row_index, cols["p4p_status"], rec.p4p.get("status"))
        set_if_col(ws, row_index, cols["score"], total_score)
        set_if_col(ws, row_index, cols["breakdown"], breakdown)
        set_if_col(ws, row_index, cols["prevention_note"], reminder)
        set_if_col(ws, row_index, cols["is_self_select"], "✔" if rec.is_self_select else None)
        for key, col in screening_cols.items():
            value = rec.screenings.get(key)
            set_if_col(ws, row_index, col, value if value not in (None, "") else None)
        ws.cell(row_index, chart_col).value = "、".join(sorted(rec.charts)) if rec.charts else "-"
        ws.cell(row_index, note_col).value = "；".join(sorted(rec.notes)) if rec.notes else "-"
        row_index += 1

    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        if col in (chart_col, note_col):
            continue
        if ws.column_dimensions[letter].width is None:
            ws.column_dimensions[letter].width = 12

    for row in ws.iter_rows(min_row=data_start, max_row=row_index - 1):
        for cell in row:
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal,
                vertical="center",
                wrap_text=True,
            )
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = True
    last_row = row_index - 1
    generic_cols, summaries = apply_generic_derived_outputs(wb, ws, row_records, data_start, last_row, today)
    generic.populate_doctor_sheet(wb, ws, generic_cols, data_start, last_row, today)
    if DOCTOR_SHEET_NAME in wb.sheetnames:
        apply_doctor_prevention_fills(wb[DOCTOR_SHEET_NAME])
        generic._finalize_doctor_sheet_alignment(wb[DOCTOR_SHEET_NAME])
    populate_self_select_sheet(wb, sorted_records)
    populate_percentile_sheet(wb, generic_cols, data_start, last_row, summaries)

    if output_dir is None:
        output_dir = source_dir.parent
    output_dir.mkdir(parents=True, exist_ok=True)
    clinic_name = extract_clinic_name(source_dir) or source_dir.name
    timestamp = dt.datetime.now().strftime("%m%d_%H%M")
    out_path = output_dir / f"{sanitize_filename(clinic_name)}展望雲端選會員_{timestamp}.xlsx"
    wb.save(out_path)
    return out_path


def extract_clinic_name(source_dir: Path) -> str:
    text = source_dir.name
    match = re.search(r"\d{10}(.+?)(?:雲端)?展望?$", text)
    if match:
        return match.group(1)
    return ""


def sanitize_filename(text: str) -> str:
    text = re.sub(r'[\\/:*?"<>|]+', "_", text).strip()
    return text or "展望雲端"


def choose_folder_gui() -> Optional[str]:
    try:
        import tkinter as tk
        from tkinter import filedialog
    except Exception:
        return None
    root = tk.Tk()
    root.withdraw()
    folder = filedialog.askdirectory(title="請選擇展望雲端資料夾")
    root.destroy()
    return folder or None


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="展望雲端專用選會員產檔")
    parser.add_argument("source_dir", nargs="?", help="展望雲端資料夾")
    parser.add_argument("-o", "--output-dir", help="輸出資料夾；預設為來源底下的展望雲端整理")
    args = parser.parse_args(argv)

    source = args.source_dir or choose_folder_gui()
    if not source:
        print("未選擇資料夾。")
        return 1
    source_dir = Path(source).expanduser().resolve()
    if not source_dir.is_dir():
        print(f"找不到資料夾：{source_dir}")
        return 1
    output_dir = Path(args.output_dir).expanduser().resolve() if args.output_dir else None
    out_path = create_output(source_dir, output_dir)
    print(f"已輸出：{out_path}")
    return 0


def gui_main() -> None:
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()

    source = filedialog.askdirectory(title="選擇來源資料夾（展望雲端資料夾）")
    if not source:
        return

    try:
        out_path = create_output(Path(source).expanduser().resolve())
        messagebox.showinfo("完成", f"已輸出：\n{out_path}")
        generic.open_file_cross_platform(str(out_path))
    except Exception as exc:
        messagebox.showerror("錯誤", str(exc))


if __name__ == "__main__":
    if len(sys.argv) > 1:
        raise SystemExit(main())
    gui_main()
