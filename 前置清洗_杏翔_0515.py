# -*- coding: utf-8 -*-
"""
杏翔前置清洗 + 通用主程式包裝

用途：
- 將杏翔的「次數.xlsx」年度總表轉成通用版可辨識的標準月份 CSV
- 將杏翔「次數/11401.xlsx」這類月次數檔轉成通用版可辨識的 HISB 次數 CSV
- 再呼叫 run_merge_通用（自動偵測最新版）完成輸出
"""

from __future__ import annotations

import csv
import datetime
import importlib.util
import os
import re
import shutil
import sys
import tempfile
from pathlib import Path
from typing import Any, List, Optional

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent


def _find_generic_script(script_dir: Path) -> Path:
    candidates = sorted(script_dir.glob("run_merge_通用_*.py"), reverse=True)
    if not candidates:
        raise RuntimeError("找不到通用程式 run_merge_通用_*.py")
    return candidates[0]

GENERIC_SCRIPT = _find_generic_script(SCRIPT_DIR)


def _load_generic_module():
    spec = importlib.util.spec_from_file_location("run_merge_generic", GENERIC_SCRIPT)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"無法載入通用程式：{GENERIC_SCRIPT}")
    mod = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = mod
    spec.loader.exec_module(mod)
    return mod


def _find_annual_count_file(source_dir: Path) -> Optional[Path]:
    for path in sorted(source_dir.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        if "次數" in path.stem:
            return path
    return None


def _header_index_map(header: List[object]) -> dict[str, int]:
    out: dict[str, int] = {}
    for idx, value in enumerate(header):
        key = str(value).strip() if value is not None else ""
        if key:
            out[key] = idx
    return out


def _pick(row: List[object], mapping: dict[str, int], *candidates: str) -> str:
    for candidate in candidates:
        idx = mapping.get(candidate)
        if idx is not None and idx < len(row):
            value = row[idx]
            if value is not None:
                text = str(value).strip()
                if text:
                    return text
    return ""


def _to_roc_yyyymmdd(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        value = value.date()
    if isinstance(value, datetime.date):
        roc_year = value.year - 1911
        return f"{roc_year:03d}{value.month:02d}{value.day:02d}" if roc_year > 0 else ""

    s = str(value).strip().lstrip("'")
    if not s:
        return ""
    s = re.sub(r"\.0$", "", s)

    m = re.fullmatch(r"(\d{4})(\d{2})(\d{2})", s)
    if m:
        year, month, day = map(int, m.groups())
        roc_year = year - 1911
        if roc_year <= 0:
            return ""
        try:
            datetime.date(year, month, day)
        except ValueError:
            return ""
        return f"{roc_year:03d}{month:02d}{day:02d}"

    m = re.fullmatch(r"(\d{2,3})(\d{2})(\d{2})", s)
    if m:
        roc_year, month, day = map(int, m.groups())
        try:
            datetime.date(roc_year + 1911, month, day)
        except ValueError:
            return ""
        return f"{roc_year:03d}{month:02d}{day:02d}"

    return ""


def _convert_annual_count_file(xlsx_path: Path, output_dir: Path) -> Optional[Path]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        if ws.max_row < 2:
            return None

        row_iter = ws.iter_rows(values_only=True)
        header = list(next(row_iter, ()))
        mapping = _header_index_map(header)
        required_id = any(k in mapping for k in ("ID", "身份證號", "身份證號碼", "身分證號", "身分證號碼"))
        required_cnt = any(k in mapping for k in ("次數", "就診次數", "門診次數"))
        if not required_id or not required_cnt:
            return None

        out_rows: List[List[object]] = [["ID", "次數", "申請金額"]]
        for row in row_iter:
            pid = _pick(row, mapping, "ID", "身份證號", "身份證號碼", "身分證號", "身分證號碼")
            cnt = _pick(row, mapping, "次數", "就診次數", "門診次數")
            if not pid or not cnt:
                continue
            out_rows.append([pid, cnt, 0])

        if len(out_rows) <= 1:
            return None

        # 年度總表沒有月份明細；固定轉成 11501 標準月檔讓通用版可接手統計次數。
        out_path = output_dir / "11501.csv"
        with out_path.open("w", encoding="utf-8-sig", newline="") as f:
            csv.writer(f).writerows(out_rows)
        return out_path
    finally:
        wb.close()


def _preclean_annual_counts(source_dir: Path) -> List[Path]:
    annual_file = _find_annual_count_file(source_dir)
    if annual_file is None:
        return []

    converted = _convert_annual_count_file(annual_file, source_dir)
    return [converted] if converted is not None else []


def _find_monthly_count_files(source_dir: Path) -> List[Path]:
    count_dir = source_dir / "次數"
    if not count_dir.is_dir():
        return []
    return [
        path
        for path in sorted(count_dir.glob("*.xlsx"))
        if not path.name.startswith("~$") and re.fullmatch(r"1(14|15)\d{2}", path.stem)
    ]


def _convert_monthly_count_file(xlsx_path: Path, output_dir: Path) -> Optional[Path]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        ws = wb[xlsx_path.stem] if xlsx_path.stem in wb.sheetnames else wb[wb.sheetnames[0]]
        if ws.max_row < 2:
            return None

        row_iter = ws.iter_rows(values_only=True)
        header = list(next(row_iter, ()))
        mapping = _header_index_map(header)
        required_name = any(k in mapping for k in ("姓名", "病患姓名", "患者姓名", "會員姓名"))
        required_bday = any(k in mapping for k in ("生日", "出生日期", "出生年月日"))
        required_cnt = any(k in mapping for k in ("次數", "就診次數", "門診次數", "來診次數"))
        if not (required_name and required_bday and required_cnt):
            return None

        out_rows: List[List[object]] = [["病歷號", "杏翔姓名", "杏翔生日", "電話", "次數", "地址"]]
        for row in row_iter:
            chart_no = _pick(row, mapping, "病歷號", "病歷號碼")
            name = _pick(row, mapping, "姓名", "病患姓名", "患者姓名", "會員姓名")
            bday_raw = _pick(row, mapping, "生日", "出生日期", "出生年月日")
            cnt = _pick(row, mapping, "次數", "就診次數", "門診次數", "來診次數")
            bday = _to_roc_yyyymmdd(bday_raw)
            if not name or not bday or not cnt:
                continue
            out_rows.append([chart_no, name, bday, "", cnt, ""])

        if len(out_rows) <= 1:
            return None

        out_path = output_dir / f"{xlsx_path.stem}.csv"
        with out_path.open("w", encoding="utf-8-sig", newline="") as f:
            csv.writer(f).writerows(out_rows)
        return out_path
    finally:
        wb.close()


def _preclean_monthly_counts(source_dir: Path) -> List[Path]:
    generated: List[Path] = []
    for monthly_file in _find_monthly_count_files(source_dir):
        converted = _convert_monthly_count_file(monthly_file, source_dir)
        if converted is not None:
            generated.append(converted)
    return generated


def _build_xing_profile(generic):
    class XingProfile(generic.ProcessingProfile):
        def scan_hisb_count_sheet(self, sheet_name: str, sheet: Any):
            year_bucket = self.sheet_year_bucket(sheet_name)
            month = self.sheet_month(sheet_name)
            if year_bucket not in (114, 115) or month is None:
                return super().scan_hisb_count_sheet(sheet_name, sheet)

            for header_row in range(1, min(10, sheet.max_row) + 1):
                hmap = generic.build_header_map(sheet, header_row)
                chart_col = generic.find_column_exact(hmap, ["病歷號", "病歷號碼"])
                name_col = generic.find_column_exact(hmap, ["杏翔姓名"])
                bday_col = generic.find_column_exact(hmap, ["杏翔生日"])
                count_col = generic.find_column_exact(hmap, ["次數", "就診次數", "門診次數", "來診次數"])
                if chart_col and name_col and bday_col and count_col:
                    return generic.HisbCountSheetScan(
                        sheet_name=sheet_name,
                        header_row=header_row,
                        id_col=None,
                        name_col=name_col,
                        bday_col=bday_col,
                        count_col=count_col,
                        date_col=None,
                        amount_col=None,
                        year_bucket=year_bucket,
                        month=month,
                    )
            return super().scan_hisb_count_sheet(sheet_name, sheet)

    return XingProfile()


def process_excel(source_path: str, template_path: Optional[str] = None) -> str:
    generic = _load_generic_module()
    source_dir = Path(source_path).resolve()
    if not source_dir.is_dir():
        raise ValueError("請選擇資料夾。")

    template = template_path or generic._find_template(str(SCRIPT_DIR))
    if not os.path.exists(template):
        raise ValueError(f"找不到模板檔：{template}")

    temp_root = Path(tempfile.mkdtemp(prefix="xing_clean_"))
    temp_source = temp_root / source_dir.name

    try:
        shutil.copytree(source_dir, temp_source)
        generated_annual = _preclean_annual_counts(temp_source)
        generated_monthly = _preclean_monthly_counts(temp_source)
        print(f"已產生清洗後年度次數檔 {len(generated_annual)} 個", flush=True)
        print(f"已產生清洗後月份次數檔 {len(generated_monthly)} 個", flush=True)

        temp_output = Path(generic.process_excel(str(temp_source), template, profile=_build_xing_profile(generic)))
        final_output = source_dir.parent / temp_output.name
        if final_output.exists():
            final_output.unlink()
        shutil.move(str(temp_output), str(final_output))
        return str(final_output)
    finally:
        shutil.rmtree(temp_root, ignore_errors=True)


def main() -> None:
    generic = _load_generic_module()
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()

    src = filedialog.askdirectory(title="選擇杏翔來源資料夾")
    if not src:
        return

    template = generic._find_template(str(SCRIPT_DIR))

    try:
        out = process_excel(src, template)
        messagebox.showinfo("完成", f"已輸出：\n{out}")
        generic.open_file_cross_platform(out)
    except Exception as e:
        messagebox.showerror("錯誤", str(e))


if __name__ == "__main__":
    main()
