# -*- coding: utf-8 -*-
"""
宏誠前置清洗 + 通用主程式包裝

用途：
- 將「費用/*.xlsx」與「次數/*.pdf」合成標準月份 CSV
- 再呼叫 run_merge_通用（自動偵測最新版）完成輸出
"""

from __future__ import annotations

import csv
import importlib.util
import os
import re
import shutil
import sys
import tempfile
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover
    PdfReader = None


SCRIPT_DIR = Path(__file__).resolve().parent


def _find_generic_script(script_dir: Path) -> Path:
    candidates = sorted(script_dir.glob("run_merge_通用_*.py"), reverse=True)
    if not candidates:
        raise RuntimeError("找不到通用程式 run_merge_通用_*.py")
    return candidates[0]

GENERIC_SCRIPT = _find_generic_script(SCRIPT_DIR)


def _load_generic_module():
    spec = importlib.util.spec_from_file_location("run_merge_generic", GENERIC_SCRIPT)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"無法載入通用程式：{GENERIC_SCRIPT}")
    mod = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = mod
    spec.loader.exec_module(mod)
    return mod


def _extract_month_code(name: str) -> Optional[str]:
    match = re.search(r"(?<!\d)(1(?:14|15)\d{2})(?!\d)", name)
    return match.group(1) if match else None


def _normalize_name(value: object) -> str:
    return "".join(str(value or "").split()).upper()


def _rewrite_csv_to_utf8(csv_path: Path) -> bool:
    for enc in ("utf-16", "cp950", "utf-8-sig", "utf-8"):
        try:
            text = csv_path.read_text(encoding=enc)
            csv_path.write_text(text, encoding="utf-8-sig")
            return True
        except UnicodeError:
            continue
        except Exception:
            continue
    return False


def _iter_fee_files(source_dir: Path) -> List[Tuple[str, Path]]:
    fee_dir = source_dir / "費用"
    if not fee_dir.is_dir():
        return []
    result: List[Tuple[str, Path]] = []
    for path in sorted(fee_dir.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        code = _extract_month_code(path.stem)
        if code:
            result.append((code, path))
    return result


def _iter_pdf_files(source_dir: Path) -> List[Tuple[str, Path]]:
    pdf_dir = source_dir / "次數"
    if not pdf_dir.is_dir():
        return []
    result: List[Tuple[str, Path]] = []
    for path in sorted(pdf_dir.glob("*.pdf")):
        code = _extract_month_code(path.stem)
        if code:
            result.append((code, path))
    return result


def _extract_pdf_name_totals(pdf_path: Path) -> Dict[str, float]:
    if PdfReader is None:
        return {}

    try:
        reader = PdfReader(str(pdf_path))
    except Exception:
        return {}

    totals: Dict[str, float] = {}
    for page in reader.pages:
        text = page.extract_text() or ""
        for raw_line in text.splitlines():
            line = str(raw_line).strip()
            if not line:
                continue
            parts = line.split()
            if len(parts) < 4:
                continue
            if not parts[0].isdigit() or not parts[-1].isdigit() or not parts[-2].isdigit():
                continue
            name_key = _normalize_name("".join(parts[1:-2]))
            if not name_key:
                continue
            try:
                cnt = float(parts[-2])
            except Exception:
                continue
            totals[name_key] = totals.get(name_key, 0.0) + cnt
    return totals


def _collect_all_members(generic, source_dir: Path) -> Dict[str, Dict[str, object]]:
    wb_src = generic._merge_source_folder(str(source_dir))
    cache = generic._scan_source_sheets(wb_src)
    return generic.collect_all_members(wb_src, partial_maps=cache.partial_maps)


def _build_unique_name_to_pid(all_members: Dict[str, Dict[str, object]]) -> Dict[str, str]:
    grouped: Dict[str, List[str]] = {}
    for pid, info in all_members.items():
        name_key = _normalize_name(info.get("name"))
        if name_key:
            grouped.setdefault(name_key, []).append(pid)
    return {name: ids[0] for name, ids in grouped.items() if len(ids) == 1}


def _aggregate_fee_month(xlsx_path: Path) -> Dict[str, Dict[str, object]]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        mapping = {str(v).strip(): idx for idx, v in enumerate(header) if v is not None}
        id_idx = mapping.get("身份證號")
        name_idx = mapping.get("姓名")
        date_idx = mapping.get("看診日期")
        amount_idx = mapping.get("申請金額")
        if None in (id_idx, name_idx, date_idx, amount_idx):
            return {}

        out: Dict[str, Dict[str, object]] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            pid = str(row[id_idx] or "").strip().upper()
            name = str(row[name_idx] or "").strip()
            dt = row[date_idx]
            amt = row[amount_idx]
            if not pid:
                continue
            rec = out.setdefault(pid, {"name": name, "date": dt, "amount": 0.0, "count": 0.0})
            if name and not rec.get("name"):
                rec["name"] = name
            if dt is not None:
                old_dt = rec.get("date")
                if old_dt is None or (hasattr(dt, "toordinal") and hasattr(old_dt, "toordinal") and dt > old_dt):
                    rec["date"] = dt
            try:
                rec["amount"] = float(rec.get("amount", 0.0)) + float(amt or 0)
            except Exception:
                pass
        return out
    finally:
        wb.close()


def _write_standard_month_csv(output_dir: Path, month_code: str, rows: Dict[str, Dict[str, object]]) -> Optional[Path]:
    out_rows: List[List[object]] = [["ID", "姓名", "日期", "次數", "申請金額"]]
    for pid in sorted(rows):
        rec = rows[pid]
        dt = rec.get("date")
        if hasattr(dt, "strftime"):
            dt_val = dt.strftime("%Y-%m-%d")
        else:
            dt_val = ""
        out_rows.append([
            pid,
            rec.get("name") or "",
            dt_val,
            rec.get("count", 0.0) or 0,
            rec.get("amount", 0.0) or 0,
        ])

    if len(out_rows) <= 1:
        return None

    out_path = output_dir / f"{month_code}.csv"
    with out_path.open("w", encoding="utf-8-sig", newline="") as f:
        csv.writer(f).writerows(out_rows)
    return out_path


def _preclean_pdf_and_fee(source_dir: Path) -> List[Path]:
    generic = _load_generic_module()
    for csv_path in source_dir.glob("*.csv"):
        _rewrite_csv_to_utf8(csv_path)
    all_members = _collect_all_members(generic, source_dir)
    unique_name_to_pid = _build_unique_name_to_pid(all_members)
    pdf_name_totals = {code: _extract_pdf_name_totals(path) for code, path in _iter_pdf_files(source_dir)}

    generated: List[Path] = []
    for month_code, fee_path in _iter_fee_files(source_dir):
        rows = _aggregate_fee_month(fee_path)
        if not rows:
            continue

        name_totals = pdf_name_totals.get(month_code, {})
        for name_key, total_cnt in name_totals.items():
            pid = unique_name_to_pid.get(name_key)
            if pid and pid in rows:
                rows[pid]["count"] = total_cnt

        out_path = _write_standard_month_csv(source_dir, month_code, rows)
        if out_path is not None:
            generated.append(out_path)

    return generated


def process_excel(source_path: str, template_path: Optional[str] = None) -> str:
    generic = _load_generic_module()
    source_dir = Path(source_path).resolve()
    if not source_dir.is_dir():
        raise ValueError("請選擇資料夾。")

    template = template_path or generic._find_template(str(SCRIPT_DIR))
    if not os.path.exists(template):
        raise ValueError(f"找不到模板檔：{template}")

    temp_root = Path(tempfile.mkdtemp(prefix="hong_clean_"))
    temp_source = temp_root / source_dir.name

    try:
        shutil.copytree(source_dir, temp_source)
        generated = _preclean_pdf_and_fee(temp_source)
        print(f"已產生清洗後月份檔 {len(generated)} 個", flush=True)

        for folder_name in ("費用", "次數"):
            raw_dir = temp_source / folder_name
            if raw_dir.exists():
                shutil.rmtree(raw_dir, ignore_errors=True)

        temp_output = Path(generic.process_excel(str(temp_source), template))
        final_output = source_dir.parent / temp_output.name
        if final_output.exists():
            final_output.unlink()
        shutil.move(str(temp_output), str(final_output))
        return str(final_output)
    finally:
        shutil.rmtree(temp_root, ignore_errors=True)


def main() -> None:
    generic = _load_generic_module()
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()

    src = filedialog.askdirectory(title="選擇宏誠來源資料夾")
    if not src:
        return

    template = generic._find_template(str(SCRIPT_DIR))

    try:
        out = process_excel(src, template)
        messagebox.showinfo("完成", f"已輸出：\n{out}")
        generic.open_file_cross_platform(out)
    except Exception as e:
        messagebox.showerror("錯誤", str(e))


if __name__ == "__main__":
    main()
