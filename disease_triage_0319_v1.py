# -*- coding: utf-8 -*-
"""
disease_triage_gui_v5_2_refactor.py

在不改變既有功能與分流/備註邏輯的前提下，整理成較簡潔、較可靠版本：
- 保留原本 GUI 選檔、輸出檔名、自動開檔
- 保留 7 分流 + 分析表 + 無疾病(4,0)
- 保留 Excel/ODS 容錯、前導單引號、民國日期解析
- 保留原本 remark_code / remark_text / LDL 目標邏輯
- 只做結構化與重複碼精簡，避免功能改動
"""

import os
import re
import sys
import subprocess
import traceback
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


BASE_YEAR = 2026
SHEET_ORDER = ["DM", "CKD", "ASCVD", "DKD", "CKD+ASCVD", "DM+ASCVD", "DKD+ASCVD"]

SCREENING_EXPORT_COLS = [
    "成人預防保健(6分)",
    "子宮頸抹片檢查率(6分)",
    "65歲以上老人流感注射率(4分)",
    "糞便潛血檢查率(6分)",
    "BC肝炎篩檢(6分)",
]

EXPORT_COLS = [
    "ID", "疾病樣態", "ASCVD", "姓名", "生日",
    "最近一次HbA1c檢查結果(%)", "最近一次HbA1c檢查日期",
    "最近一次LDL檢查結果(mg/dL)", "最近一次LDL檢查日期",
    "最近一次UACR檢查結果(mg/gm)", "最近一次UACR檢查日期",
    "備註",
] + SCREENING_EXPORT_COLS

NO_DISEASE_SHEET = "無疾病(4,0)"
NO_DISEASE_EXPORT_COLS = ["ID", "疾病樣態", "ASCVD", "姓名", "生日"] + SCREENING_EXPORT_COLS

NA_STRINGS = {"NaT", "nan", "None", "<NA>", "-", ""}
ZERO_STRINGS = {"0", "0.0"}
THIN_BLACK = Side(style="thin", color="000000")
BLACK_BORDER = Border(left=THIN_BLACK, right=THIN_BLACK, top=THIN_BLACK, bottom=THIN_BLACK)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT_WRAP_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
NORMAL_FONT = Font(size=12, bold=False)
HEADER_FONT = Font(size=12, bold=True)

FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")
FILL_PINK = PatternFill("solid", fgColor="FF6699")
FILL_YELLOW = PatternFill("solid", fgColor="FFFF00")
FILL_ORANGE = PatternFill("solid", fgColor="FFC000")
FILL_PURPLE = PatternFill("solid", fgColor="DDD9EE")

SCREENING_LOOKUP_SPECS = [
    ("成人健檢", "成人預防保健(6分)"),
    ("子宮抹片", "子宮頸抹片檢查率(6分)"),
    ("老人流感", "65歲以上老人流感注射率(4分)"),
    ("糞便潛血", "糞便潛血檢查率(6分)"),
    ("肝炎篩檢", "BC肝炎篩檢(6分)"),
]

SCREENING_SHEET_ALIASES: Dict[str, List[str]] = {
    "成人健檢": ["成人健檢", "成人健檢名單", "成人預防保健", "成人預防保健名單"],
    "子宮抹片": ["子宮抹片", "子宮頸抹片", "子宮抹片名單", "子宮頸抹片名單"],
    "老人流感": ["老人流感", "老人流感名單", "老人流感注射", "65歲以上老人流感", "65歲以上老人流感注射率"],
    "糞便潛血": ["糞便潛血", "糞便潛血名單", "糞便潛血檢查", "糞便潛血檢查名單"],
    "肝炎篩檢": ["肝炎篩檢", "BC肝炎篩檢", "B肝炎篩檢", "C肝炎篩檢", "BC肝炎篩檢名單", "肝炎篩檢名單"],
}


# ------------------------- utils -------------------------
def open_file(path: Path) -> None:
    p = str(path)
    if sys.platform.startswith("darwin"):
        subprocess.run(["open", p], check=False)
    elif os.name == "nt":
        os.startfile(p)  # type: ignore[attr-defined]
    else:
        subprocess.run(["xdg-open", p], check=False)


def _strip_excel_apostrophe(x: object) -> str:
    if x is None:
        return ""
    return str(x).lstrip("'")


def _norm_colname(x: object) -> str:
    return (
        str(x).strip().lower()
        .replace("\n", "")
        .replace("\r", "")
        .replace("\\n", "")
        .replace("\\r", "")
        .replace(" ", "")
        .replace("\u3000", "")
    )


def clean_string_series(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.strip()
    s = s.str.lstrip("'")
    s = s.str.replace("\u3000", "", regex=False)
    s = s.str.replace(r"\s+", "", regex=True)
    return s


def normalize_missing_strings(series: pd.Series, extra_na: Optional[Iterable[str]] = None) -> pd.Series:
    na_values = set(NA_STRINGS)
    if extra_na:
        na_values.update(extra_na)
    s = series.astype(str).str.strip()
    mask = s.isin(na_values)
    out = series.copy()
    out.loc[mask] = pd.NA
    return out


def replace_zero_like_with_dash(df: pd.DataFrame, columns: Iterable[str]) -> pd.DataFrame:
    out = df.copy()
    for col in [c for c in columns if c in out.columns]:
        s = out[col].astype(str).str.strip().str.lstrip("'")
        mask = out[col].notna() & s.isin(ZERO_STRINGS)
        if mask.any():
            out[col] = out[col].astype("object")
            out.loc[mask, col] = "-"
    return out


def header_index_map(ws) -> Dict[str, int]:
    return {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}


def get_last_data_row(ws, key_col: int) -> int:
    for r in range(ws.max_row, 1, -1):
        v = ws.cell(r, key_col).value
        if v is not None and str(v).strip() != "":
            return r
    return 1


def apply_fill_with_border(cell, fill: Optional[PatternFill]) -> None:
    if fill is not None:
        cell.fill = fill
    cell.border = BLACK_BORDER


def year_fill(year: Optional[int], base_year: int) -> Optional[PatternFill]:
    if year is None:
        return None
    if year == base_year:
        return FILL_GREEN
    if year == base_year - 1:
        return FILL_PINK
    return FILL_YELLOW


def norm_id(series: pd.Series) -> pd.Series:
    s = clean_string_series(series)
    s = s.str.replace(r"\.0$", "", regex=True)
    return s.replace({"nan": pd.NA, "None": pd.NA, "": pd.NA})


def pick_col_exact(cols: List[str], candidates: List[str]) -> Optional[str]:
    norm_map = {_norm_colname(c): c for c in cols}
    for cand in candidates:
        found = norm_map.get(_norm_colname(cand))
        if found:
            return found
    return None


def pick_col_contains(cols: List[str], keywords: List[str]) -> Optional[str]:
    norm_cols = [(_norm_colname(c), c) for c in cols]
    for kw in keywords:
        nkw = _norm_colname(kw)
        for nc, orig in norm_cols:
            if nkw and nkw in nc:
                return orig
    return None


def to_dt(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.strip().replace({"nan": pd.NA, "None": pd.NA, "": pd.NA})

    def normalize_one(x: object) -> Optional[str]:
        if x is None or pd.isna(x):
            return None
        raw = _strip_excel_apostrophe(x).strip()
        if not raw or raw.lower() in ("nan", "none"):
            return None

        x2 = raw.replace(".", "/").replace("-", "/")

        m = re.match(r"^(\d{2,3})/(\d{1,2})/(\d{1,2})$", x2)
        if m:
            y, mo, d = map(int, m.groups())
            return f"{y + 1911:04d}-{mo:02d}-{d:02d}"

        m = re.match(r"^(\d{2,3})(\d{2})(\d{2})$", raw)
        if m:
            y, mo, d = map(int, m.groups())
            return f"{y + 1911:04d}-{mo:02d}-{d:02d}"

        m = re.match(r"^(\d{4})/(\d{1,2})/(\d{1,2})$", x2)
        if m:
            y, mo, d = map(int, m.groups())
            return f"{y:04d}-{mo:02d}-{d:02d}"

        return raw

    return pd.to_datetime(s.map(normalize_one), format="%Y-%m-%d", errors="coerce")


def parse_mixed_date(v):
    import datetime as dt

    if v is None or pd.isna(v):
        return pd.NaT
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime()
    if isinstance(v, dt.datetime):
        return v
    if isinstance(v, dt.date):
        return dt.datetime(v.year, v.month, v.day)

    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            fv = float(v)
            if 20000 <= fv <= 60000:
                return dt.datetime(1899, 12, 30) + dt.timedelta(days=fv)
        except Exception:
            pass

    s = _strip_excel_apostrophe(v).strip()
    if not s or s.lower() in ("nan", "none", "nat") or s in {"-", "0", "0.0"}:
        return pd.NaT

    s2 = (
        s.replace("年", "/")
        .replace("月", "/")
        .replace("日", "")
        .replace(".", "/")
        .replace("\\", "/")
        .replace("-", "/")
    )
    s2 = re.sub(r"/+", "/", s2).strip("/")

    for fmt, val in (("%Y/%m/%d", s2), ("%Y%m%d", s2.replace("/", ""))):
        try:
            return dt.datetime.strptime(val, fmt)
        except Exception:
            pass

    m = re.match(r"^(\d{2,3})/(\d{1,2})/(\d{1,2})$", s2)
    if m:
        y, mo, d = map(int, m.groups())
        return dt.datetime(y + 1911, mo, d)

    m = re.match(r"^(\d{2,3})(\d{2})(\d{2})$", s2.replace("/", ""))
    if m:
        y, mo, d = map(int, m.groups())
        return dt.datetime(y + 1911, mo, d)

    return pd.NaT


def age_on(bday: pd.Timestamp, ref_date: pd.Timestamp):
    if pd.isna(bday):
        return None
    return ref_date.year - bday.year - ((ref_date.month, ref_date.day) < (bday.month, bday.day))


def ascvd_norm(x) -> str:
    v = _strip_excel_apostrophe(x).strip().lower()
    return v if v in ("a", "b") else ""


def _excel_engine_for(path: Path) -> Optional[str]:
    return "odf" if path.suffix.lower() == ".ods" else None


def read_table(path: Path, **kwargs) -> pd.DataFrame:
    eng = _excel_engine_for(path)
    if eng:
        kwargs.setdefault("engine", eng)
    return pd.read_excel(path, **kwargs)


def open_workbook(path: Path) -> pd.ExcelFile:
    eng = _excel_engine_for(path)
    return pd.ExcelFile(path, engine=eng) if eng else pd.ExcelFile(path)


def find_sheet_and_header_for_b(path_b: Path) -> Tuple[str, int]:
    xl = open_workbook(path_b)

    def row_has_required(tokens) -> bool:
        norm_tokens = [_norm_colname(t) for t in tokens if str(t).strip()]
        return (
            any(t == "id" for t in norm_tokens)
            and any("疾病樣態" in t for t in norm_tokens)
            and any("ascvd" in t for t in norm_tokens)
        )

    for sh in xl.sheet_names:
        probe = read_table(path_b, sheet_name=sh, header=None, nrows=25, dtype=str)
        for r in range(len(probe)):
            if row_has_required(probe.iloc[r].tolist()):
                return sh, r

    return xl.sheet_names[0], 5


# ------------------------- rules (LOGIC NOT CHANGED) -------------------------
def required_items(sheet_name: str) -> List[str]:
    if sheet_name in ("DM", "DM+ASCVD"):
        return ["HbA1c", "LDL"]
    if sheet_name in ("CKD", "CKD+ASCVD"):
        return ["LDL", "UACR"]
    if sheet_name in ("DKD", "DKD+ASCVD"):
        return ["HbA1c", "LDL", "UACR"]
    if sheet_name == "ASCVD":
        return ["LDL"]
    raise ValueError(sheet_name)


def ldl_target(row: pd.Series, sheet_name: str) -> float:
    a = ascvd_norm(row.get("ASCVD", ""))
    if a in ("a", "b") and sheet_name in ("ASCVD", "DM+ASCVD", "CKD+ASCVD", "DKD+ASCVD"):
        return 55.0 if a == "a" else 70.0
    if sheet_name == "CKD":
        return 130.0
    return 100.0


def remark_code(row: pd.Series, sheet_name: str, base_year: int, _ref: Optional[pd.Timestamp] = None) -> int:
    items = required_items(sheet_name)
    ref = _ref if _ref is not None else pd.Timestamp(f"{base_year}-06-30")
    age = age_on(row["_生日_dt"], ref)
    h_target = 8.0 if (age is not None and age >= 80) else 7.0
    l_target = ldl_target(row, sheet_name)
    u_target = 30.0

    tested: Dict[str, bool] = {}
    bad: Dict[str, bool] = {}
    years: List[int] = []

    for it in items:
        dt = row.get(f"_{it}_dt", pd.NaT)
        val = row.get(f"_{it}_num", pd.NA)
        is_y = (not pd.isna(dt)) and (dt.year == base_year)
        tested[it] = is_y

        if not pd.isna(dt):
            years.append(int(dt.year))

        if is_y and not pd.isna(val):
            if it == "HbA1c":
                bad[it] = val >= h_target
            elif it == "LDL":
                bad[it] = val >= l_target
            elif it == "UACR":
                bad[it] = val >= u_target
            else:
                bad[it] = False
        else:
            bad[it] = False

    all_tested = all(tested.values())
    all_missing = not any(tested.values())
    any_bad = any(bad.values())
    last_year = max(years) if years else None

    if all_tested and not any_bad:
        return 0
    if (not all_tested) and any_bad:
        return 3
    if (not all_tested) and (not all_missing):
        return 2
    if all_tested and any_bad:
        return 1

    if all_missing:
        any_prev = False
        for it in items:
            dt = row.get(f"_{it}_dt", pd.NaT)
            if not pd.isna(dt) and dt.year == (base_year - 1):
                any_prev = True
                break
        if any_prev:
            return 4
        if last_year is None:
            return 5
        if (base_year - last_year) >= 2:
            return 5
        return 2

    return 2


def remark_text(code: int, x_items: int, base_year: int) -> str:
    if code == 0:
        return f"0.{base_year}年{x_items}項前測控制良好，符合收案標準,建議收為會員。"
    if code == 1:
        return f"1.{base_year}年已受檢未達控制良好"
    if code == 2:
        return f"2.{base_year}年漏檢,請安排回診"
    if code == 3:
        return f"3.{base_year}年漏檢+未達控制良好,回診複查"
    if code == 4:
        return f"4.{base_year - 1}已檢：提醒{base_year}年度回診"
    if code == 5:
        return "5.逾兩年未檢：暫緩收案"
    return ""


# ------------------------- build df -------------------------
def build_master_df(path_a: Path, path_b: Path) -> pd.DataFrame:
    a0 = read_table(path_a, dtype=str)
    a0.columns = a0.columns.astype(str).str.strip()
    cols_a = list(a0.columns)

    col_id_a = pick_col_exact(cols_a, ["ID", "Id", "id"]) or pick_col_contains(cols_a, ["id"])
    if not col_id_a:
        raise ValueError("A 檔找不到欄位：ID")

    col_name_a = pick_col_exact(cols_a, ["姓名", "欄1", "name"]) or pick_col_contains(cols_a, ["姓名", "欄1"])
    if not col_name_a:
        raise ValueError("A 檔找不到欄位：姓名（或 欄1）")

    col_bday_a = pick_col_exact(cols_a, ["生日", "BIRTHDAY", "出生日期"]) or pick_col_contains(cols_a, ["生日", "birthday", "出生"])
    col_h_val = pick_col_exact(cols_a, ["最近一次HbA1c檢查結果(%)", "最近一次HbA1c檢查結果 (%)"]) or pick_col_contains(cols_a, ["最近一次hba1c檢查結果"])
    col_h_dt = pick_col_exact(cols_a, ["最近一次HbA1c檢查日期"]) or pick_col_contains(cols_a, ["最近一次hba1c檢查日期"])
    col_l_val = pick_col_exact(cols_a, ["最近一次LDL檢查結果(mg/dL)", "最近一次LDL檢查結果 (mg/dL)"]) or pick_col_contains(cols_a, ["最近一次ldl檢查結果"])
    col_l_dt = pick_col_exact(cols_a, ["最近一次LDL檢查日期"]) or pick_col_contains(cols_a, ["最近一次ldl檢查日期"])
    col_u_val = pick_col_exact(cols_a, ["最近一次UACR檢查結果(mg/gm)", "最近一次UACR檢查結果 (mg/gm)"]) or pick_col_contains(cols_a, ["最近一次uacr檢查結果"])
    col_u_dt = pick_col_exact(cols_a, ["最近一次UACR檢查日期"]) or pick_col_contains(cols_a, ["最近一次uacr檢查日期"])

    keep_a = [c for c in [col_id_a, col_name_a, col_bday_a, col_h_val, col_h_dt, col_l_val, col_l_dt, col_u_val, col_u_dt] if c]
    a = a0[keep_a].copy().rename(columns={col_id_a: "ID", col_name_a: "姓名"})
    a["生日"] = a0[col_bday_a] if col_bday_a else pd.NA
    a["ID"] = norm_id(a["ID"])

    b_sheet, b_header = find_sheet_and_header_for_b(path_b)
    b0 = read_table(path_b, sheet_name=b_sheet, header=b_header, dtype=str)
    b0.columns = b0.columns.astype(str).str.strip()
    cols_b = list(b0.columns)

    col_id_b = pick_col_exact(cols_b, ["ID", "Id", "id"]) or pick_col_contains(cols_b, ["id"])
    col_dis_b = pick_col_exact(cols_b, ["疾病樣態"]) or pick_col_contains(cols_b, ["疾病樣態"])
    col_ascvd_b = pick_col_exact(cols_b, ["ASCVD", "ascvd"]) or pick_col_contains(cols_b, ["ascvd"])
    if not col_id_b or not col_dis_b or not col_ascvd_b:
        raise ValueError("B 檔欄位找不到（需要 ID/疾病樣態/ASCVD）")

    b = b0[[col_id_b, col_dis_b, col_ascvd_b]].copy().rename(
        columns={col_id_b: "ID", col_dis_b: "疾病樣態", col_ascvd_b: "ASCVD"}
    )
    b["ID"] = norm_id(b["ID"])

    merged = a.merge(b, on="ID", how="left", indicator=True)
    dis_clean = clean_string_series(merged["疾病樣態"])

    df = pd.DataFrame({
        "ID": merged["ID"],
        "疾病樣態": pd.to_numeric(dis_clean, errors="coerce"),
        "ASCVD": merged["ASCVD"],
        "姓名": merged["姓名"],
        "生日": merged["生日"],
        "最近一次HbA1c檢查結果(%)": merged[col_h_val] if col_h_val else pd.NA,
        "最近一次HbA1c檢查日期": merged[col_h_dt] if col_h_dt else pd.NA,
        "最近一次LDL檢查結果(mg/dL)": merged[col_l_val] if col_l_val else pd.NA,
        "最近一次LDL檢查日期": merged[col_l_dt] if col_l_dt else pd.NA,
        "最近一次UACR檢查結果(mg/gm)": merged[col_u_val] if col_u_val else pd.NA,
        "最近一次UACR檢查日期": merged[col_u_dt] if col_u_dt else pd.NA,
    })
    df["_merge_flag"] = merged["_merge"]

    date_src_map = {
        "_生日_dt": "生日",
        "_HbA1c_dt": "最近一次HbA1c檢查日期",
        "_LDL_dt": "最近一次LDL檢查日期",
        "_UACR_dt": "最近一次UACR檢查日期",
    }
    for dst, src in date_src_map.items():
        df[dst] = to_dt(df[src])

    num_src_map = {
        "_HbA1c_num": "最近一次HbA1c檢查結果(%)",
        "_LDL_num": "最近一次LDL檢查結果(mg/dL)",
        "_UACR_num": "最近一次UACR檢查結果(mg/gm)",
    }
    for dst, src in num_src_map.items():
        df[dst] = pd.to_numeric(df[src].astype(str).str.lstrip("'"), errors="coerce")

    df["生日"] = df["_生日_dt"].dt.strftime("%Y-%m-%d")
    df["最近一次HbA1c檢查日期"] = df["_HbA1c_dt"].dt.strftime("%Y-%m-%d")
    df["最近一次LDL檢查日期"] = df["_LDL_dt"].dt.strftime("%Y-%m-%d")
    df["最近一次UACR檢查日期"] = df["_UACR_dt"].dt.strftime("%Y-%m-%d")
    df["ASCVD"] = df["ASCVD"].apply(ascvd_norm)

    miss_count = int(df["疾病樣態"].isna().sum())
    if miss_count:
        print(f"  [WARN] {miss_count} 筆在 B 檔找不到疾病樣態/ASCVD，將不分流")
    return df


def resolve_screening_sheet_name(path_a: Path, target_name: str, norm_to_real: Optional[Dict[str, str]] = None) -> Optional[str]:
    if norm_to_real is None:
        xl = open_workbook(path_a)
        norm_to_real = {_norm_colname(sh): sh for sh in xl.sheet_names}

    aliases = SCREENING_SHEET_ALIASES.get(target_name, [target_name])
    for alias in aliases:
        real = norm_to_real.get(_norm_colname(alias))
        if real:
            return real

    for alias in aliases:
        n_alias = _norm_colname(alias)
        for n_sheet, real in norm_to_real.items():
            if n_alias and (n_alias in n_sheet or n_sheet in n_alias):
                return real
    return None


def build_screening_lookup(path_a: Path, sheet_name: str, norm_to_real: Optional[Dict[str, str]] = None) -> pd.DataFrame:
    resolved_name = resolve_screening_sheet_name(path_a, sheet_name, norm_to_real)
    if not resolved_name:
        print(f"[WARN] 找不到 sheet：{sheet_name} → 已跳過")
        return pd.DataFrame(columns=["ID", "篩檢日期"])

    raw = read_table(path_a, sheet_name=resolved_name, header=None, dtype=object)
    if raw.empty:
        print(f"[WARN] sheet「{resolved_name}」是空的 → 已跳過")
        return pd.DataFrame(columns=["ID", "篩檢日期"])

    date_keywords = ["最後篩檢日期", "最後檢查日期", "篩檢日期", "檢查日期", "受檢日期", "最後一次檢查日期", "最後一次篩檢日期", "日期"]
    id_keywords = ["ID", "id", "Id", "身份證", "身分證", "身份證號", "身分證號"]

    header_row = None
    for r in range(min(len(raw), 15)):
        vals = [_norm_colname(v) for v in raw.iloc[r].tolist()]
        has_id = any(any(_norm_colname(k) == v or _norm_colname(k) in v for k in id_keywords) for v in vals if v)
        has_date = any(any(_norm_colname(k) in v for k in date_keywords) for v in vals if v)
        if has_id and has_date:
            header_row = r
            break
    if header_row is None:
        header_row = 0

    header_vals = raw.iloc[header_row].tolist()
    data = raw.iloc[header_row + 1:].copy().reset_index(drop=True)
    data.columns = [str(c).strip() if c is not None and not pd.isna(c) else f"col_{i}" for i, c in enumerate(header_vals)]
    cols = list(data.columns)

    def find_col(candidates_exact=None, candidates_kw=None):
        candidates_exact = candidates_exact or []
        candidates_kw = candidates_kw or []
        for c in cols:
            if any(_norm_colname(c) == _norm_colname(name) for name in candidates_exact):
                return c
        for c in cols:
            nc = _norm_colname(c)
            for kw in candidates_kw:
                nkw = _norm_colname(kw)
                if nkw and nkw in nc:
                    return c
        return None

    col_id = find_col(["ID", "id", "Id"], ["id", "身份證", "身分證", "身份證號", "身分證號"])
    col_dt = find_col(date_keywords, date_keywords)
    if not col_id:
        print(f"[WARN] sheet「{resolved_name}」找不到 ID 欄位 → 已跳過")
        return pd.DataFrame(columns=["ID", "篩檢日期"])

    out = data[[col_id]].copy().rename(columns={col_id: "ID"})
    out["ID"] = norm_id(out["ID"])
    out = out[out["ID"].notna()].copy()

    if not col_dt:
        out = out.drop_duplicates(subset=["ID"], keep="first")
        out["篩檢日期"] = pd.NA
        return out[["ID", "篩檢日期"]]

    out["篩檢日期"] = data[col_dt].apply(parse_mixed_date)
    latest = out.groupby("ID", as_index=False)["篩檢日期"].max()
    latest["篩檢日期"] = latest["篩檢日期"].apply(lambda x: x.strftime("%Y-%m-%d") if pd.notna(x) else pd.NA)
    latest["篩檢日期"] = normalize_missing_strings(latest["篩檢日期"])
    return latest[["ID", "篩檢日期"]]


def build_all_screening_lookup(path_a: Path) -> pd.DataFrame:
    xl = open_workbook(path_a)
    norm_to_real = {_norm_colname(sh): sh for sh in xl.sheet_names}

    lookups: Dict[str, pd.DataFrame] = {}
    all_ids: set = set()
    for logical_sheet_name, out_col in SCREENING_LOOKUP_SPECS:
        lkp = build_screening_lookup(path_a, logical_sheet_name, norm_to_real)
        if lkp.empty:
            print(f"  [INFO] 篩檢 sheet 無資料或找不到：{logical_sheet_name}")
            continue
        lkp = lkp.rename(columns={"篩檢日期": out_col})
        lookups[out_col] = lkp
        all_ids.update(lkp["ID"].dropna().tolist())

    if not all_ids:
        return pd.DataFrame(columns=["ID"] + SCREENING_EXPORT_COLS)

    result = pd.DataFrame({"ID": sorted(all_ids)})
    for out_col in SCREENING_EXPORT_COLS:
        result = result.merge(lookups[out_col], on="ID", how="left") if out_col in lookups else result.assign(**{out_col: pd.NA})

    result["ID"] = norm_id(result["ID"])
    return result[result["ID"].notna()][["ID"] + SCREENING_EXPORT_COLS].reset_index(drop=True)


def attach_screening_columns(df: pd.DataFrame, path_a: Path) -> pd.DataFrame:
    out = df.merge(build_all_screening_lookup(path_a), on="ID", how="left")
    for col in SCREENING_EXPORT_COLS:
        if col not in out.columns:
            out[col] = pd.NA
        out[col] = normalize_missing_strings(out[col])
    return out


def build_no_disease_df(df: pd.DataFrame, path_a: Optional[Path] = None) -> pd.DataFrame:
    base = df[(~df["疾病樣態"].isin([1, 2, 3])) & (~df["ASCVD"].isin(["a", "b"]))].copy()
    if base.empty:
        return pd.DataFrame(columns=NO_DISEASE_EXPORT_COLS)

    for col in NO_DISEASE_EXPORT_COLS:
        if col not in base.columns:
            base[col] = pd.NA
    for col in SCREENING_EXPORT_COLS:
        base[col] = normalize_missing_strings(base[col])

    return base[NO_DISEASE_EXPORT_COLS].sort_values("ID", ascending=True, kind="mergesort").reset_index(drop=True)


# ------------------------- split -------------------------
def split_into_sheets(df: pd.DataFrame, base_year: int) -> Tuple[Dict[str, pd.DataFrame], Dict[str, Dict[str, int]]]:
    is_dm = df["疾病樣態"] == 1
    is_ckd = df["疾病樣態"] == 2
    is_dkd = df["疾病樣態"] == 3
    has_ascvd = df["ASCVD"].isin(["a", "b"])

    sheets = {
        "DM": df[is_dm & (~has_ascvd)].copy(),
        "CKD": df[is_ckd & (~has_ascvd)].copy(),
        "DKD": df[is_dkd & (~has_ascvd)].copy(),
        "DM+ASCVD": df[is_dm & has_ascvd].copy(),
        "CKD+ASCVD": df[is_ckd & has_ascvd].copy(),
        "DKD+ASCVD": df[is_dkd & has_ascvd].copy(),
        "ASCVD": df[has_ascvd & (~(is_dm | is_ckd | is_dkd))].copy(),
    }

    analysis_counts: Dict[str, Dict[str, int]] = {}
    ref_ts = pd.Timestamp(f"{base_year}-06-30")

    for name, d in sheets.items():
        if len(d) > 0:
            x = len(required_items(name))
            codes = d.apply(lambda r: remark_code(r, name, base_year, ref_ts), axis=1)
            d["備註"] = [remark_text(int(c), x, base_year) for c in codes.tolist()]
            d = d.sort_values("ID", ascending=True, kind="mergesort")
            vc = d["備註"].astype(str).str.extract(r"^(\d)\.")[0].value_counts()
        else:
            d = pd.DataFrame(columns=EXPORT_COLS)
            vc = pd.Series(dtype="int64")

        analysis_counts[name] = {str(i): int(vc.get(str(i), 0)) for i in range(6)}
        sheets[name] = d[EXPORT_COLS].copy()

    return sheets, analysis_counts


# ------------------------- formatting helpers -------------------------
def autosize_columns(ws, max_width: int = 45) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len("" if cell.value is None else str(cell.value)) for cell in col), default=0)
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _parse_year(v) -> Optional[int]:
    if v is None:
        return None
    s = _strip_excel_apostrophe(v).strip()
    if not s or s in {"-", "0", "0.0"}:
        return None
    m = re.search(r"(\d{4})", s)
    return int(m.group(1)) if m else None


def _parse_float(v) -> Optional[float]:
    if v is None:
        return None
    s = _strip_excel_apostrophe(v).strip()
    if not s or s in {"-", "0", "0.0"}:
        return None
    try:
        return float(s)
    except Exception:
        return None


def style_triage_sheet(ws, sheet_name: str) -> None:
    if ws.max_row < 2:
        return

    idx = header_index_map(ws)
    c_id = idx.get("ID", 1)
    c_asc = idx.get("ASCVD")
    c_bday = idx.get("生日")
    c_remark = idx.get("備註")
    c_h_val = idx.get("最近一次HbA1c檢查結果(%)")
    c_h_dt = idx.get("最近一次HbA1c檢查日期")
    c_l_val = idx.get("最近一次LDL檢查結果(mg/dL)")
    c_l_dt = idx.get("最近一次LDL檢查日期")
    c_u_val = idx.get("最近一次UACR檢查結果(mg/gm)")
    c_u_dt = idx.get("最近一次UACR檢查日期")

    screening_date_cols = [idx.get(name) for name in SCREENING_EXPORT_COLS]
    screening_date_cols = [c for c in screening_date_cols if c]
    member_cols = [idx.get(name) for name in ["ID", "疾病樣態", "ASCVD", "姓名", "生日"]]
    member_cols = [c for c in member_cols if c]

    last_data_row = get_last_data_row(ws, c_id)
    if last_data_row == 1:
        return

    def paint_date(row_r: int, col_dt: Optional[int]) -> None:
        if not col_dt:
            return
        apply_fill_with_border(ws.cell(row_r, col_dt), year_fill(_parse_year(ws.cell(row_r, col_dt).value), BASE_YEAR))

    def paint_value_if_bad(row_r: int, col_val: Optional[int], col_dt: Optional[int], threshold: float) -> None:
        if (not col_val) or (not col_dt):
            return
        y = _parse_year(ws.cell(row_r, col_dt).value)
        v = _parse_float(ws.cell(row_r, col_val).value)
        if y == BASE_YEAR and (v is not None) and v >= threshold:
            apply_fill_with_border(ws.cell(row_r, col_val), FILL_ORANGE)

    items = required_items(sheet_name)
    style_ref = pd.Timestamp(f"{BASE_YEAR}-06-30")

    for r in range(2, last_data_row + 1):
        ws.row_dimensions[r].height = 20
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.font = NORMAL_FONT
            cell.alignment = LEFT_WRAP_ALIGN if (c_remark and c == c_remark) else CENTER_ALIGN
            cell.border = BLACK_BORDER

        code = ""
        if c_remark:
            rv = ws.cell(r, c_remark).value
            s = str(rv).strip() if rv else ""
            if s and s[0].isdigit():
                code = s[0]

        if code == "4":
            for c in member_cols:
                apply_fill_with_border(ws.cell(r, c), FILL_PURPLE)
        elif code == "5":
            for c in member_cols:
                apply_fill_with_border(ws.cell(r, c), FILL_YELLOW)

        row_ascvd = ascvd_norm(ws.cell(r, c_asc).value if c_asc else "")
        bday_str = ws.cell(r, c_bday).value if c_bday else None
        bday_dt = pd.to_datetime(_strip_excel_apostrophe(bday_str), errors="coerce") if bday_str not in (None, "-", "0", "0.0") else pd.NaT
        age = age_on(bday_dt, style_ref)
        h_target = 8.0 if (age is not None and age >= 80) else 7.0
        if row_ascvd in ("a", "b") and sheet_name in ("ASCVD", "DM+ASCVD", "CKD+ASCVD", "DKD+ASCVD"):
            l_target = 55.0 if row_ascvd == "a" else 70.0
        elif sheet_name == "CKD":
            l_target = 130.0
        else:
            l_target = 100.0

        if "HbA1c" in items:
            paint_date(r, c_h_dt)
            paint_value_if_bad(r, c_h_val, c_h_dt, h_target)
        if "LDL" in items:
            paint_date(r, c_l_dt)
            paint_value_if_bad(r, c_l_val, c_l_dt, l_target)
        if "UACR" in items:
            paint_date(r, c_u_dt)
            paint_value_if_bad(r, c_u_val, c_u_dt, 30.0)
        for c in screening_date_cols:
            paint_date(r, c)


# ------------------------- analysis sheet -------------------------
def build_analysis_sheet(ws, analysis_counts: Dict[str, Dict[str, int]], base_year: int) -> None:
    order = ["DM", "CKD", "DKD", "ASCVD", "CKD+ASCVD", "DM+ASCVD", "DKD+ASCVD"]
    group_fills = {
        "DM": PatternFill("solid", fgColor="D9E8EE"),
        "CKD": PatternFill("solid", fgColor="F3DED3"),
        "DKD": PatternFill("solid", fgColor="DFF0D8"),
        "ASCVD": PatternFill("solid", fgColor="D9E8EE"),
        "CKD+ASCVD": PatternFill("solid", fgColor="E6E6E6"),
        "DM+ASCVD": PatternFill("solid", fgColor="E5D8EE"),
        "DKD+ASCVD": PatternFill("solid", fgColor="D8E6EE"),
    }

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    title_font = Font(name="標楷體", size=20, bold=True)
    group_font = Font(name="Times New Roman", size=14, bold=False)
    analysis_header_font = Font(name="標楷體", size=14, bold=True)
    text_font = Font(name="標楷體", size=12, bold=False)
    red_font = Font(name="標楷體", size=12, color="FF0000", bold=True)
    num_font = Font(name="Times New Roman", size=14, bold=False)

    ws.column_dimensions["A"].width = 12
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 22

    ws.merge_cells("A1:G1")
    ws["A1"].value = "115年收案會員選案建議評估"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 40

    def remark_texts_for_group(g: str) -> List[str]:
        x = len(required_items(g))
        return [
            f"0.{base_year}年{x}項前測控制良好，符合收案標準,建議收為會員",
            f"1.{base_year}年已受檢未達控制良好",
            f"2.{base_year}年漏檢,請安排回診",
            f"3.{base_year}年漏檢+未達控制良好,回診複查",
            f"4.{base_year - 1}已檢：提醒{base_year}年度回診",
            "5.逾兩年未檢：暫緩收案",
        ]

    r = 2
    base_h = 33
    remark_h = base_h * 1.2

    for g in order:
        fill = group_fills.get(g, PatternFill("solid", fgColor="EEEEEE"))
        ws.row_dimensions[r].height = base_h
        ws.row_dimensions[r + 1].height = remark_h
        ws.row_dimensions[r + 2].height = base_h

        ws.merge_cells(f"A{r}:A{r + 2}")
        cA = ws[f"A{r}"]
        cA.value = g
        cA.font = group_font
        cA.alignment = center
        cA.fill = fill
        cA.border = BLACK_BORDER

        ws.merge_cells(f"B{r}:G{r}")
        cTot = ws[f"B{r}"]
        cTot.value = "總計"
        cTot.font = analysis_header_font
        cTot.alignment = center
        cTot.fill = fill
        cTot.border = BLACK_BORDER

        texts = remark_texts_for_group(g)
        counts = analysis_counts.get(g, {str(i): 0 for i in range(6)})

        for i, col in enumerate("BCDEFG"):
            text_cell = ws[f"{col}{r + 1}"]
            text_cell.value = texts[i]
            text_cell.alignment = left_wrap
            text_cell.fill = fill
            text_cell.border = BLACK_BORDER
            text_cell.font = red_font if any(k in texts[i] for k in ["建議收", "未達控制", "漏檢", "提醒"]) else text_font

            num_cell = ws[f"{col}{r + 2}"]
            num_cell.value = int(counts.get(str(i), 0))
            num_cell.font = num_font
            num_cell.alignment = Alignment(horizontal="right", vertical="center")
            num_cell.fill = PatternFill("solid", fgColor="FFFFFF")
            num_cell.border = BLACK_BORDER

            top_cell = ws[f"{col}{r}"]
            top_cell.border = BLACK_BORDER
            top_cell.fill = fill

        for rr in (r + 1, r + 2):
            ws[f"A{rr}"].fill = fill
            ws[f"A{rr}"].border = BLACK_BORDER

        sep_row = r + 3
        ws.row_dimensions[sep_row].height = 15
        r = sep_row + 1

    ws.sheet_view.showGridLines = False


def style_no_disease_sheet(ws) -> None:
    if ws.max_row < 1:
        return

    idx = header_index_map(ws)
    check_cols = [idx.get(name) for name in SCREENING_EXPORT_COLS]
    check_cols = [c for c in check_cols if c]

    for c in range(1, ws.max_column + 1):
        cell = ws.cell(1, c)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.font = HEADER_FONT
        cell.border = BLACK_BORDER

    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 20
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.alignment = CENTER_ALIGN
            cell.font = NORMAL_FONT
            cell.border = BLACK_BORDER

        for c in check_cols:
            cell = ws.cell(r, c)
            fill = year_fill(_parse_year(cell.value), BASE_YEAR)
            if fill is not None:
                cell.fill = fill

    autosize_columns(ws)
    c_name = idx.get("姓名")
    if c_name:
        ws.column_dimensions[get_column_letter(c_name)].width = 10
    c_birth = idx.get("生日")
    if c_birth:
        birth_col = get_column_letter(c_birth)
        if ws.column_dimensions[birth_col].width < 12:
            ws.column_dimensions[birth_col].width = 12
    ws.sheet_view.showGridLines = False


# ------------------------- write output -------------------------
def _excel_safe(v):
    if v is None:
        return None
    if isinstance(v, float) and pd.isna(v):
        return None
    try:
        if pd.isna(v):
            return None
    except Exception:
        pass
    return v


def write_df_to_sheet(ws, df: pd.DataFrame, columns: List[str]) -> None:
    for c, h in enumerate(columns, start=1):
        ws.cell(row=1, column=c, value=_excel_safe(h))
    for r_idx, row in enumerate(df.reindex(columns=columns).itertuples(index=False, name=None), start=2):
        for c_idx, v in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=_excel_safe(v))


def write_output(out_path: Path, sheets: Dict[str, pd.DataFrame], analysis_counts: Dict[str, Dict[str, int]], no_disease_df: Optional[pd.DataFrame] = None) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    ws_analysis = wb.create_sheet("分析表", 0)
    build_analysis_sheet(ws_analysis, analysis_counts, BASE_YEAR)

    fix_columns = set(
        ["ID", "疾病樣態", "ASCVD", "姓名", "生日", "備註"]
        + ["最近一次HbA1c檢查日期", "最近一次LDL檢查日期", "最近一次UACR檢查日期"]
        + SCREENING_EXPORT_COLS
        + ["最近一次HbA1c檢查結果(%)", "最近一次LDL檢查結果(mg/dL)", "最近一次UACR檢查結果(mg/gm)"]
    )

    for name in SHEET_ORDER:
        ws = wb.create_sheet(title=name)
        df = sheets.get(name, pd.DataFrame(columns=EXPORT_COLS)).copy()
        df = replace_zero_like_with_dash(df, [c for c in df.columns if c in fix_columns])
        write_df_to_sheet(ws, df, list(df.columns))
        autosize_columns(ws)

        idx = header_index_map(ws)
        c_name = idx.get("姓名")
        if c_name:
            ws.column_dimensions[get_column_letter(c_name)].width = 10
        c_remark = idx.get("備註")
        if c_remark:
            ws.column_dimensions[get_column_letter(c_remark)].width = 30

        style_triage_sheet(ws, sheet_name=name)

    ws_no = wb.create_sheet(title=NO_DISEASE_SHEET)
    no_df = no_disease_df.copy() if no_disease_df is not None else pd.DataFrame(columns=NO_DISEASE_EXPORT_COLS)
    no_df = replace_zero_like_with_dash(no_df, [c for c in NO_DISEASE_EXPORT_COLS if c in no_df.columns])
    write_df_to_sheet(ws_no, no_df, NO_DISEASE_EXPORT_COLS)
    style_no_disease_sheet(ws_no)

    wb.save(out_path)


# ------------------------- GUI main -------------------------
def main() -> None:
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()

    try:
        messagebox.showinfo("步驟1", "請選擇 A 檔（個案健康管理）")
        path_a = filedialog.askopenfilename(
            title="選擇 A 檔（姓名 + 檢驗結果）",
            filetypes=[("Spreadsheet files", "*.xlsx *.xls *.ods")],
        )
        if not path_a:
            return

        messagebox.showinfo("步驟2", "請選擇 B 檔（照護名單，含ASCVD）")
        path_b = filedialog.askopenfilename(
            title="選擇 B 檔（疾病樣態 + ASCVD）",
            filetypes=[("Spreadsheet files", "*.xlsx *.xls *.ods")],
        )
        if not path_b:
            return

        out_dir = Path(path_b).parent
        ts = datetime.now().strftime("%m%d_%H%M")
        out_path = out_dir / f"未達控制{ts}.xlsx"

        print("▶ 讀取並合併 A/B 檔...")
        df = build_master_df(Path(path_a), Path(path_b))
        print(f"  ✓ 合併完成，共 {len(df)} 筆")

        print("▶ 附加篩檢欄位...")
        df = attach_screening_columns(df, Path(path_a))
        print("  ✓ 篩檢欄位附加完成")

        print("▶ 分流中...")
        sheets, analysis_counts = split_into_sheets(df, BASE_YEAR)
        for sname, sdf in sheets.items():
            print(f"  {sname}: {len(sdf)} 筆")

        print("▶ 建立無疾病分頁...")
        no_disease_df = build_no_disease_df(df, Path(path_a))
        print(f"  ✓ 無疾病共 {len(no_disease_df)} 筆")

        print("▶ 寫出 Excel...")
        write_output(out_path, sheets, analysis_counts, no_disease_df)
        print(f"  ✓ 輸出完成：{out_path}")

        messagebox.showinfo("完成", f"報表完成（年度固定 {BASE_YEAR}）：\n{out_path}")
        open_file(out_path)

    except Exception:
        tb = traceback.format_exc()
        print(f"\n❌ 發生錯誤：\n{tb}")
        messagebox.showerror("錯誤(Traceback)", tb)


if __name__ == "__main__":
    main()
