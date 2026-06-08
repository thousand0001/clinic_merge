#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""鄭骨科耳鼻喉科診所選會員整合程式。

資料整理邏輯已內建；原始資料夾只讀，完成後直接產生正式 Excel。
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import importlib.util
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import traceback
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl
import xlrd


PROJECT_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = PROJECT_DIR / "選會員模板0430.xlsx"

CLINICS = {
    "zheng": {
        "code": "3531060324",
        "name": "鄭骨科耳鼻喉科診所",
        "folder": "3531060324鄭骨科展望",
        "system": "展望",
    },
    "youxiao": {
        "code": "3531026671",
        "name": "友孝診所",
        "folder": "3531026671友孝展望",
        "system": "展望",
    },
    "datong": {
        "code": "3531112707",
        "name": "大同耳鼻喉科診所",
        "folder": "3531112707大同耳鼻喉耀聖",
        "system": "耀聖",
    },
}

STANDARD_FILES = {
    "designated": "家醫名單_標準.xlsx",
    "self_select": "自選會員_標準.xlsx",
    "exclude_select": "不要自選會員_標準.xlsx",
    "claims": "門診次數費用_標準.xlsx",
    "health": "檢驗資料_標準.xlsx",
    "p4p_case": "P4P收案_標準.xlsx",
    "p4p_track": "P4P追蹤_標準.xlsx",
    "screening": "篩檢資料_標準.xlsx",
}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def normalize_key(value: Any) -> str:
    return re.sub(r"\s+", "", normalize_text(value)).upper()


def normalize_id(value: Any) -> str:
    text = normalize_text(value).strip("'").upper()
    if text.endswith(".0"):
        text = text[:-2]
    return text.replace(" ", "")


def is_valid_id(value: Any) -> bool:
    text = normalize_id(value)
    return bool(
        re.fullmatch(r"[A-Z][12]\d{8}", text)
        or re.fullmatch(r"[A-Z][A-Z0-9]\d{8}", text)
        or re.fullmatch(r"[A-Z]{1,2}\d{8,10}", text)
    )


def normalize_name(value: Any) -> str:
    return re.sub(r"\s+", "", normalize_text(value))


def normalize_chart(value: Any, width: int = 7) -> str:
    text = normalize_text(value)
    return text.zfill(width) if text.isdigit() else text


def is_missing(value: Any) -> bool:
    return normalize_text(value) in {"", "-", "—", "–", "None", "NaN"}


def parse_number(value: Any) -> float:
    text = normalize_text(value).replace(",", "")
    if text in {"", "-", "—", "–"}:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def parse_date(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    text = normalize_text(value)
    if text in {"", "-", "—", "–"}:
        return ""
    text = text.strip("'").split()[0]
    digits = re.sub(r"\D", "", text)
    if re.fullmatch(r"\d{7}", digits):
        year = int(digits[:3]) + 1911
        month = int(digits[3:5])
        day = int(digits[5:7])
    elif re.fullmatch(r"\d{8}", digits):
        year = int(digits[:4])
        if year < 1911:
            year += 1911
        month = int(digits[4:6])
        day = int(digits[6:8])
    elif re.fullmatch(r"\d{6}", digits):
        year = int(digits[:2]) + 1911
        month = int(digits[2:4])
        day = int(digits[4:6])
    else:
        match = re.fullmatch(r"(\d{2,4})[./-](\d{1,2})[./-](\d{1,2})", text)
        if not match:
            return text
        year = int(match.group(1))
        if year < 1911:
            year += 1911
        month = int(match.group(2))
        day = int(match.group(3))
    try:
        return dt.date(year, month, day).isoformat()
    except ValueError:
        return ""


def roc_date_for_claim(value: Any) -> str:
    parsed = parse_date(value)
    if not parsed:
        return normalize_text(value)
    y, m, d = parsed.split("-")
    return f"{int(y) - 1911:03d}{m}{d}"


def read_csv_rows(path: Path) -> List[List[str]]:
    data = path.read_bytes()
    for enc in ("utf-8-sig", "big5", "cp950"):
        try:
            text = data.decode(enc)
        except UnicodeDecodeError:
            continue
        if "病歷號" in text or "醫事機構代碼" in text or "身分證號" in text:
            return list(csv.reader(text.splitlines()))
    text = data.decode("cp950", errors="replace")
    return list(csv.reader(text.splitlines()))


def read_xlsx_sheet(path: Path, sheet: Optional[str] = None) -> Iterable[Tuple[str, List[List[Any]]]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        names = [sheet] if sheet else wb.sheetnames
        for name in names:
            ws = wb[name]
            rows = [list(row) for row in ws.iter_rows(values_only=True)]
            yield name, rows
    finally:
        wb.close()


def read_xls_rows(path: Path) -> List[List[Any]]:
    book = xlrd.open_workbook(str(path))
    rows: List[List[Any]] = []
    for sheet in book.sheets():
        for r_idx in range(sheet.nrows):
            rows.append([sheet.cell_value(r_idx, c_idx) for c_idx in range(sheet.ncols)])
    return rows


def compact_header(value: Any) -> str:
    return re.sub(r"[\s　]+", "", normalize_text(value)).lower()


def find_header(rows: Sequence[Sequence[Any]], required_groups: Sequence[Sequence[str]]) -> Optional[int]:
    for idx, row in enumerate(rows):
        values = {compact_header(value) for value in row}
        ok = True
        for group in required_groups:
            if not any(compact_header(item) in values for item in group):
                ok = False
                break
        if ok:
            return idx
    return None


def find_col(header: Sequence[Any], aliases: Sequence[str]) -> Optional[int]:
    compact = {compact_header(value): idx for idx, value in enumerate(header)}
    for alias in aliases:
        idx = compact.get(compact_header(alias))
        if idx is not None:
            return idx
    return None


def cell(row: Sequence[Any], idx: Optional[int]) -> Any:
    if idx is None or idx >= len(row):
        return ""
    return row[idx]


def write_table(path: Path, sheet_name: str, headers: Sequence[str], rows: Sequence[Dict[str, Any]]) -> None:
    if not rows:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(list(headers))
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max(len(header) + 4, 12), 28)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()


def copy_standard_xlsx(src: Path, dst: Path) -> None:
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)


def standard_claim_row(
    *,
    chart: Any = "",
    date: Any,
    name: Any,
    pid: Any,
    birth: Any = "",
    address: Any = "",
    phone: Any = "",
    count: Any = 0,
    amount: Any = 0,
    dx1: Any = "",
    dx23: Any = "",
) -> Dict[str, Any]:
    return {
        "掛號證": normalize_chart(chart),
        "日期": roc_date_for_claim(date),
        "姓名": normalize_text(name),
        "身分證號": normalize_id(pid),
        "生日": roc_date_for_claim(birth) if birth else "",
        "地址": normalize_text(address),
        "電話": normalize_text(phone),
        "病1": normalize_text(dx1),
        "病23": normalize_text(dx23),
        "次數": parse_number(count),
        "總額": parse_number(amount),
    }


CLAIM_HEADERS = ["掛號證", "日期", "姓名", "身分證號", "生日", "地址", "電話", "病1", "病23", "次數", "總額"]


def build_identity_map(source_dir: Path) -> Dict[Tuple[str, str], str]:
    mapping: Dict[Tuple[str, str], str] = {}
    for path in sorted(source_dir.iterdir()):
        if path.name.startswith("~$") or path.suffix.lower() != ".xlsx":
            continue
        for _sheet, rows in read_xlsx_sheet(path):
            h = find_header(rows, [["姓名", "會員姓名"], ["ID", "身分證號", "身份證號碼", "家醫收案會員ID"]])
            if h is None:
                continue
            header = rows[h]
            name_col = find_col(header, ["姓名", "會員姓名"])
            id_col = find_col(header, ["ID", "身分證號", "身份證號碼", "家醫收案會員ID"])
            birth_col = find_col(header, ["生日", "BIRTHDAY"])
            for row in rows[h + 1:]:
                pid = normalize_id(cell(row, id_col))
                name = normalize_name(cell(row, name_col))
                birth = parse_date(cell(row, birth_col))
                if pid and name and birth:
                    mapping[(name, birth)] = pid
    return mapping


def clean_designated_and_root_xlsx(source_dir: Path, out_dir: Path) -> None:
    for path in sorted(source_dir.iterdir()):
        if path.name.startswith("~$") or path.suffix.lower() != ".xlsx":
            continue
        name = path.name
        lower = name.lower()
        if "家醫" in name and ("追蹤日報表" in name or "收案追蹤報表" in name):
            copy_standard_xlsx(path, out_dir / STANDARD_FILES["designated"])
        elif "cliassay" in lower:
            copy_standard_xlsx(path, out_dir / STANDARD_FILES["health"])
        elif "clip4pcase" in lower:
            copy_standard_xlsx(path, out_dir / STANDARD_FILES["p4p_case"])
        elif "clip4ptrack" in lower:
            copy_standard_xlsx(path, out_dir / STANDARD_FILES["p4p_track"])
        elif "cliscores" in lower or "b、c肝" in lower or "bc肝" in lower:
            dst = out_dir / "篩檢資料"
            copy_standard_xlsx(path, dst / name)
        elif "自選會員" in name and "不要" not in name and "115x" not in lower:
            copy_standard_xlsx(path, out_dir / STANDARD_FILES["self_select"])
        elif "不要會員" in name or "115x" in lower:
            copy_standard_xlsx(path, out_dir / STANDARD_FILES["exclude_select"])


def clean_r11440_folder(source_dir: Path) -> List[Dict[str, Any]]:
    claims: List[Dict[str, Any]] = []
    report_dir = source_dir / "R11440"
    if not report_dir.is_dir():
        return claims
    for path in sorted(path for path in report_dir.iterdir() if path.is_file() and path.suffix.lower() == ".xlsx"):
        for _sheet, rows in read_xlsx_sheet(path):
            claims.extend(extract_claim_rows(rows))
    return claims


def clean_r11440_workbook(source_dir: Path) -> List[Dict[str, Any]]:
    claims: List[Dict[str, Any]] = []
    path = source_dir / "R11440月報表.xlsx"
    if not path.exists():
        return claims
    for _sheet, rows in read_xlsx_sheet(path):
        claims.extend(extract_claim_rows(rows))
    return claims


def extract_claim_rows(rows: Sequence[Sequence[Any]]) -> List[Dict[str, Any]]:
    claims: List[Dict[str, Any]] = []
    h = find_header(rows, [["日期"], ["姓名"], ["身分證號", "身份證號"], ["次數"], ["總額"]])
    if h is None:
        return claims
    header = rows[h]
    chart_col = find_col(header, ["掛號證", "掛號証", "病歷號"])
    date_col = find_col(header, ["日期"])
    name_col = find_col(header, ["姓名"])
    id_col = find_col(header, ["身分證號", "身份證號"])
    birth_col = find_col(header, ["生日"])
    address_col = find_col(header, ["地址"])
    phone_col = find_col(header, ["電話"])
    count_col = find_col(header, ["次數", "看診次數", "就診次數"])
    amount_col = find_col(header, ["總額", "申報總額"])
    dx1_col = find_col(header, ["病1"])
    dx23_col = find_col(header, ["病23"])
    for row in rows[h + 1:]:
        pid = normalize_id(cell(row, id_col))
        date_value = roc_date_for_claim(cell(row, date_col))
        if not is_valid_id(pid) or not re.match(r"1(?:14|15)\d{4}$", date_value):
            continue
        claims.append(standard_claim_row(
            chart=cell(row, chart_col),
            date=date_value,
            name=cell(row, name_col),
            pid=pid,
            birth=cell(row, birth_col),
            address=cell(row, address_col),
            phone=cell(row, phone_col),
            count=cell(row, count_col),
            amount=cell(row, amount_col),
            dx1=cell(row, dx1_col),
            dx23=cell(row, dx23_col),
        ))
    return claims


def clean_datong_count_csv(source_dir: Path, id_map: Dict[Tuple[str, str], str]) -> List[Dict[str, Any]]:
    claims: List[Dict[str, Any]] = []
    count_dir = source_dir / "次數"
    if not count_dir.is_dir():
        return claims
    for path in sorted(count_dir.glob("*.CSV")) + sorted(count_dir.glob("*.csv")):
        month_match = re.fullmatch(r"1(?:14|15)\d{2}", path.stem)
        if not month_match:
            continue
        month = path.stem
        rows = read_csv_rows(path)
        h = find_header(rows, [["病歷號"], ["姓名"], ["生日"], ["次數"]])
        if h is None:
            continue
        header = rows[h]
        chart_col = find_col(header, ["病歷號"])
        name_col = find_col(header, ["姓名", "姓    名"])
        birth_col = find_col(header, ["生日", "生   日"])
        phone_col = find_col(header, ["電話"])
        count_col = find_col(header, ["次數"])
        address_col = find_col(header, ["地址"])
        for row in rows[h + 1:]:
            name = normalize_text(cell(row, name_col))
            birth_iso = parse_date(cell(row, birth_col))
            pid = id_map.get((normalize_name(name), birth_iso), "")
            if not pid:
                continue
            claims.append(standard_claim_row(
                chart=cell(row, chart_col),
                date=f"{month}01",
                name=name,
                pid=pid,
                birth=cell(row, birth_col),
                address=cell(row, address_col),
                phone=cell(row, phone_col),
                count=cell(row, count_col),
                amount=0,
            ))
    return claims


def clean_datong_last_visit(source_dir: Path, id_map: Dict[Tuple[str, str], str]) -> List[Dict[str, Any]]:
    path = source_dir / "最後就診日.CSV"
    if not path.exists():
        return []
    rows = read_csv_rows(path)
    h = find_header(rows, [["病歷號"], ["姓名"], ["生日"], ["最後回診日"]])
    if h is None:
        return []
    header = rows[h]
    chart_col = find_col(header, ["病歷號"])
    name_col = find_col(header, ["姓名"])
    birth_col = find_col(header, ["生日"])
    phone_col = find_col(header, ["電話"])
    last_col = find_col(header, ["最後回診日"])
    dx_cols = [find_col(header, ["DX1"]), find_col(header, ["DX2"]), find_col(header, ["DX3"])]
    latest: Dict[str, Dict[str, Any]] = {}
    for row in rows[h + 1:]:
        name = normalize_text(cell(row, name_col))
        birth_iso = parse_date(cell(row, birth_col))
        pid = id_map.get((normalize_name(name), birth_iso), "")
        last_visit = roc_date_for_claim(cell(row, last_col))
        if not pid or not re.match(r"1(?:14|15)\d{4}$", last_visit):
            continue
        existing = latest.get(pid)
        if existing and normalize_text(existing["日期"]) >= last_visit:
            continue
        latest[pid] = standard_claim_row(
            chart=cell(row, chart_col),
            date=last_visit,
            name=name,
            pid=pid,
            birth=cell(row, birth_col),
            phone=cell(row, phone_col),
            count=0,
            amount=0,
            dx1=cell(row, dx_cols[0]),
            dx23=",".join(normalize_text(cell(row, idx)) for idx in dx_cols[1:] if idx is not None),
        )
    return list(latest.values())


def summarize_claims(claims: Sequence[Dict[str, Any]]) -> Dict[str, float]:
    totals = {
        "114_count": 0.0,
        "114_amount": 0.0,
        "115_count": 0.0,
        "115_amount": 0.0,
    }
    for row in claims:
        date_text = normalize_text(row.get("日期"))
        if date_text.startswith("114"):
            totals["114_count"] += parse_number(row.get("次數"))
            totals["114_amount"] += parse_number(row.get("總額"))
        elif date_text.startswith("115"):
            totals["115_count"] += parse_number(row.get("次數"))
            totals["115_amount"] += parse_number(row.get("總額"))
    return totals


def prepare_clinic_data(key: str, source_dir: Path, tmp_root: Path) -> Dict[str, Any]:
    info = CLINICS[key]
    source_dir = source_dir.expanduser().resolve()
    if not source_dir.is_dir():
        raise FileNotFoundError(f"找不到來源資料夾：{source_dir}")
    run_tag = dt.datetime.now().strftime("clean_%Y%m%d%H%M%S%f")
    out_dir = tmp_root / f"{info['code']}{info['name']}_{run_tag}"
    out_dir.mkdir(parents=True, exist_ok=True)
    data_dir = out_dir / run_tag

    clean_designated_and_root_xlsx(source_dir, data_dir)
    claims: List[Dict[str, Any]] = []
    claims.extend(clean_r11440_folder(source_dir))
    claims.extend(clean_r11440_workbook(source_dir))
    if key == "datong":
        id_map = build_identity_map(source_dir)
        claims.extend(clean_datong_count_csv(source_dir, id_map))
        claims.extend(clean_datong_last_visit(source_dir, id_map))

    write_table(data_dir / STANDARD_FILES["claims"], "門診次數費用", CLAIM_HEADERS, claims)
    summary = {
        "clinic_key": key,
        "clinic_code": info["code"],
        "clinic_name": info["name"],
        "source_dir": str(source_dir),
        "clean_dir": str(out_dir),
        "claim_rows": len(claims),
        "claim_totals": summarize_claims(claims),
        "standard_files": sorted(str(path.relative_to(out_dir)) for path in out_dir.rglob("*.xlsx")),
    }
    (out_dir / "clean_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def run_common_output(clean_dir: Path, output_dir: Path) -> Path:
    generic_script = PROJECT_DIR / "選會員_共用核心_0605.py"
    if not generic_script.is_file():
        raise RuntimeError(f"找不到通用程式：{generic_script}")
    spec = importlib.util.spec_from_file_location("run_merge_generic_0605", generic_script)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"無法載入通用程式：{generic_script}")
    generic = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = generic
    spec.loader.exec_module(generic)

    temp_output = Path(generic.process_excel(str(clean_dir), str(TEMPLATE_PATH)))
    output_dir.mkdir(parents=True, exist_ok=True)
    final_output = output_dir / temp_output.name
    if final_output.exists():
        stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        final_output = output_dir / f"{temp_output.stem}_{stamp}{temp_output.suffix}"
    shutil.move(str(temp_output), str(final_output))
    return final_output


def clean_all(tmp_root: Path) -> None:
    if tmp_root.exists():
        shutil.rmtree(tmp_root)


def choose_source_dir(clinic_name: str) -> Optional[Path]:
    try:
        import tkinter as tk
        from tkinter import filedialog
    except ImportError as exc:
        raise RuntimeError("無法開啟資料夾選擇視窗，請改用命令列傳入來源資料夾。") from exc

    root = tk.Tk()
    root.withdraw()
    try:
        selected = filedialog.askdirectory(title=f"選擇{clinic_name}來源資料夾")
    finally:
        root.destroy()
    return Path(selected) if selected else None


def open_file_cross_platform(path: str) -> None:
    try:
        if sys.platform.startswith("darwin"):
            subprocess.call(("open", path))
        elif os.name == "nt":
            os.startfile(path)  # type: ignore[attr-defined]
        else:
            subprocess.call(("xdg-open", path))
    except Exception:
        pass


def main(argv: Optional[Sequence[str]] = None, gui_root: Any = None) -> int:
    parser = argparse.ArgumentParser(description="診所前置清洗並產生正式選會員 Excel")
    parser.add_argument("clinic", choices=sorted(CLINC for CLINC in CLINICS) + ["all"])
    parser.add_argument("source_dir", nargs="?", help="診所來源資料夾；未指定時開啟選擇視窗")
    parser.add_argument("--tmp-root", help="暫存資料夾；未指定時使用系統暫存目錄")
    parser.add_argument("--no-run-output", action="store_false", dest="run_output", help="只執行前置清洗")
    parser.set_defaults(run_output=True)
    parser.add_argument("--output-dir", help="正式 Excel 輸出資料夾，預設為來源資料夾上一層")
    parser.add_argument("--cleanup", action="store_true", help="完成後刪除暫存清洗資料")
    args = parser.parse_args(argv)

    keys = list(CLINICS) if args.clinic == "all" else [args.clinic]
    if args.clinic == "all" and not args.source_dir:
        parser.error("all 模式必須透過命令列指定共同來源資料夾")

    if args.source_dir:
        source_dir = Path(args.source_dir)
    elif gui_root is not None:
        from tkinter import filedialog

        selected = filedialog.askdirectory(
            title=f"選擇{CLINICS[keys[0]]['name']}來源資料夾",
            mustexist=True,
        )
        source_dir = Path(selected) if selected else None
    else:
        source_dir = choose_source_dir(CLINICS[keys[0]]["name"])
    if source_dir is None:
        print("未選擇來源資料夾，程式已取消。")
        return 0
    source_dir = source_dir.expanduser().resolve()
    tmp_root = (
        Path(args.tmp_root).expanduser().resolve()
        if args.tmp_root
        else Path(tempfile.mkdtemp(prefix="clinic_merge_preclean_"))
    )
    output_dir = Path(args.output_dir).expanduser().resolve() if args.output_dir else source_dir.parent

    summaries = []
    outputs = []
    for key in keys:
        clinic_source = source_dir / CLINICS[key]["folder"] if args.clinic == "all" else source_dir
        summary = prepare_clinic_data(key, clinic_source, tmp_root)
        summaries.append(summary)
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        if args.run_output:
            info = CLINICS[key]
            output = run_common_output(Path(summary["clean_dir"]), output_dir)
            outputs.append(str(output))
            print(f"正式輸出：{output}")
    if args.cleanup:
        clean_all(tmp_root)
        print(f"已刪除暫存：{tmp_root}")
    if outputs:
        print("正式輸出清單：")
        for output in outputs:
            print(output)
        if gui_root is not None:
            from tkinter import messagebox

            messagebox.showinfo("完成", "已輸出：\n" + "\n".join(outputs))
            open_file_cross_platform(outputs[0])
    return 0


def gui_main(clinic_key: str) -> int:
    import tkinter as tk
    from tkinter import messagebox

    root = tk.Tk()
    root.withdraw()
    try:
        return main([clinic_key], gui_root=root)
    except (ValueError, KeyError, OSError) as exc:
        messagebox.showerror("錯誤", str(exc))
        return 1
    except Exception as exc:
        traceback.print_exc()
        messagebox.showerror("錯誤", f"未預期錯誤：{exc}")
        return 1
    finally:
        root.destroy()


if __name__ == "__main__":
    if len(sys.argv) == 1:
        raise SystemExit(gui_main("zheng"))
    raise SystemExit(main(["zheng", *sys.argv[1:]]))
