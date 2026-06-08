# -*- coding: utf-8 -*-
"""
嘉齡 / 家齡 / 益民 / 一珊 0608 前置清洗 + 通用主程式包裝。

原則：
- 不修改 run_merge_通用_0430_1.py。
- 不修改選會員樣板，只指定它作為輸出樣板。
- 原始診所資料不動；清洗資料只產生在暫存資料夾。
- L/M/N/O 核對以清洗階段同步累計的 totals 為準，輸出後只讀會員總表 L-O 欄。
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import importlib.util
import os
import re
import shutil
import sys
import tempfile
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, DefaultDict, Dict, Iterable, List, Optional, Tuple

import xlrd
from openpyxl import Workbook, load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
TEMPLATE_NAME = "選會員模板0526.xlsx"
TW_ID_RE = re.compile(r"^[A-Z][1289ABCD]\d{8}$")
TW_ID_FIND_RE = re.compile(r"[A-Z][1289ABCD]\d{8}")
ILLEGAL_XLSX_CHARS_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")

SCREENING_SHEET_NAMES = {
    "adult": "成人健檢",
    "pap": "子宮抹片",
    "flu": "老人流感",
    "fit": "糞便潛血",
    "hep": "肝炎篩檢",
}

SCREENING_CODE_MAP = {
    "IC3E": ("adult",),
    "IC23": ("adult",),
    "IC24": ("adult",),
    "IC29": ("hep",),
    "IC31": ("pap",),
    "IC37": ("pap",),
    "IC85": ("fit",),
    "IC01": ("flu",),
    "ICL1001": ("adult", "hep"),
    "ICL1002": ("adult", "hep"),
}

LAB_CODE_MAP = {
    "09006C": "hba",
    "09044C": "ldl",
}


@dataclass
class CleanResult:
    totals: Dict[str, float] = field(default_factory=lambda: {"L": 0.0, "M": 0.0, "N": 0.0, "O": 0.0})
    stats: Dict[str, int] = field(default_factory=dict)
    internal_ids: set = field(default_factory=set)


def _empty_health_record(name: Any = "", bday: Any = None) -> Dict[str, Any]:
    return {
        "name": _display_text(name),
        "bday": _parse_date(bday),
        "hba": "",
        "hba_dt": None,
        "ldl": "",
        "ldl_dt": None,
        "uacr": "",
        "uacr_dt": None,
    }


def _find_generic_script() -> Path:
    candidates = sorted(SCRIPT_DIR.glob("run_merge_通用_*.py"), reverse=True)
    if not candidates:
        raise RuntimeError("找不到通用程式 run_merge_通用_*.py")
    return candidates[0]


def _load_generic_module():
    generic_script = _find_generic_script()
    spec = importlib.util.spec_from_file_location("run_merge_generic_0608", generic_script)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"無法載入通用程式：{generic_script}")
    mod = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = mod
    spec.loader.exec_module(mod)
    return mod


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and abs(value - round(value)) < 1e-9:
        value = int(round(value))
    s = str(value).strip()
    if s.endswith(".0") and re.fullmatch(r"-?\d+\.0", s):
        s = s[:-2]
    return re.sub(r"\s+", "", s)


def _sanitize_filename_component(name: str) -> str:
    text = _clean_text(name)
    return re.sub(r'[\\/:*?"<>|]+', "_", text).strip()


def _display_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and abs(value - round(value)) < 1e-9:
        value = int(round(value))
    s = str(value).strip()
    if s.endswith(".0") and re.fullmatch(r"-?\d+\.0", s):
        s = s[:-2]
    return ILLEGAL_XLSX_CHARS_RE.sub("", s)


def _normalize_id(value: Any) -> str:
    s = _clean_text(value).upper()
    if s.endswith(".0"):
        s = s[:-2]
    return s.replace(" ", "")


def _is_valid_id(value: Any) -> bool:
    return bool(TW_ID_RE.fullmatch(_normalize_id(value)))


def _extract_last_tw_id(value: Any) -> str:
    matches = TW_ID_FIND_RE.findall(_normalize_id(value))
    return matches[-1] if matches else ""


def _extract_birth_after_id(value: Any, pid: str) -> Optional[dt.date]:
    text = _normalize_id(value)
    pid = _normalize_id(pid)
    if not text or not pid:
        return None
    pos = text.rfind(pid)
    if pos < 0:
        return None
    tail = text[pos + len(pid):]
    match = re.match(r"(\d{8})", tail)
    return _parse_date(match.group(1)) if match else None


def _normalize_phone(value: Any) -> str:
    raw = _display_text(value)
    if not raw:
        return ""
    digits = re.sub(r"\D+", "", raw)
    if len(digits) == 9 and digits.startswith("9"):
        digits = "0" + digits
    return digits


def _to_float(value: Any) -> float:
    s = _display_text(value).replace(",", "")
    if not s or s in ("-", "—", "–"):
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0


def _parse_date(value: Any) -> Optional[dt.date]:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value

    s = _display_text(value)
    if not s:
        return None

    m = re.fullmatch(r"(\d{2,4})[./-](\d{1,2})[./-](\d{1,2})", s)
    if m:
        y, mo, day = map(int, m.groups())
        if y < 1911:
            y += 1911
        try:
            return dt.date(y, mo, day)
        except ValueError:
            return None

    digits = re.sub(r"\D+", "", s)
    if len(digits) == 8:
        y = int(digits[:4])
        if 1900 <= y <= 2100:
            try:
                return dt.date(y, int(digits[4:6]), int(digits[6:8]))
            except ValueError:
                return None
    if len(digits) == 7:
        y = int(digits[:3]) + 1911
        try:
            return dt.date(y, int(digits[3:5]), int(digits[5:7]))
        except ValueError:
            return None
    if len(digits) == 6:
        y = int(digits[:2]) + 1911
        try:
            return dt.date(y, int(digits[2:4]), int(digits[4:6]))
        except ValueError:
            return None
    return None


def _date_to_month_code(value: Any) -> Optional[str]:
    parsed = _parse_date(value)
    if parsed is None:
        return None
    roc_year = parsed.year - 1911
    if roc_year not in (114, 115):
        return None
    return f"{roc_year}{parsed.month:02d}"


def _month_code_to_date(month_code: str) -> dt.date:
    return dt.date(int(month_code[:3]) + 1911, int(month_code[3:5]), 1)


def _internal_id(seed: str) -> str:
    # 通用程式的月份統計需要合法身分證格式 key；輸出後會依 internal_ids 精準清空。
    digest = hashlib.sha1(seed.encode("utf-8")).hexdigest()
    number = int(digest[:12], 16) % 100000000
    return f"Z2{number:08d}"


def _update_totals(totals: Dict[str, float], month_code: str, count: float, amount: float) -> None:
    if month_code.startswith("114"):
        totals["L"] += count
        totals["M"] += amount
    elif month_code.startswith("115"):
        totals["N"] += count
        totals["O"] += amount


def _safe(value: Any) -> Any:
    if isinstance(value, str):
        return ILLEGAL_XLSX_CHARS_RE.sub("", value)
    return value


def _write_workbook(path: Path, sheets: Dict[str, List[List[Any]]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for title, rows in sheets.items():
        ws = wb.create_sheet(title[:31])
        for row in rows:
            ws.append([_safe(v) for v in row])
    wb.save(path)


def _clinic_code_from_source(source_dir: Path) -> str:
    match = re.match(r"^([A-Za-z0-9]{10})", source_dir.name)
    return match.group(1) if match else ""


def _load_clinic_name_lookup() -> Dict[str, str]:
    lookup_path = SCRIPT_DIR / "醫療群_衛福部資料.xlsx"
    if not lookup_path.exists():
        return {}
    try:
        wb = load_workbook(lookup_path, read_only=True, data_only=True)
    except Exception:
        return {}
    try:
        for ws in wb.worksheets:
            if ws.max_row < 2:
                continue
            headers = [_clean_text(cell.value) for cell in ws[2]]
            try:
                code_idx = headers.index("醫事機構代碼")
            except ValueError:
                continue
            name_idx = None
            for candidate in ("官方名稱", "醫事機構名稱"):
                if candidate in headers:
                    name_idx = headers.index(candidate)
                    break
            if name_idx is None:
                continue
            result: Dict[str, str] = {}
            for row in ws.iter_rows(min_row=3, values_only=True):
                code = _clean_text(row[code_idx] if code_idx < len(row) else "")
                name = _display_text(row[name_idx] if name_idx < len(row) else "")
                if code and name:
                    result[code] = name
            if result:
                return result
    finally:
        wb.close()
    return {}


def _add_screening(
    screening: Dict[str, Dict[str, Tuple[dt.date, str, Optional[dt.date]]]],
    kind: str,
    pid: str,
    date_value: Any,
    name: Any = "",
    birth: Any = "",
) -> bool:
    pid = _normalize_id(pid)
    date = _parse_date(date_value)
    if not _is_valid_id(pid) or date is None:
        return False
    name_text = _display_text(name)
    bday = _parse_date(birth)
    current = screening[kind].get(pid)
    if current is None or date > current[0]:
        screening[kind][pid] = (date, name_text, bday)
    return True


def _add_screening_by_code(
    screening: Dict[str, Dict[str, Tuple[dt.date, str, Optional[dt.date]]]],
    code: Any,
    pid: Any,
    date_value: Any,
    name: Any = "",
    birth: Any = "",
) -> int:
    code_text = _clean_text(code).upper()
    kinds = SCREENING_CODE_MAP.get(code_text, ())
    written = 0
    for kind in kinds:
        if _add_screening(screening, kind, _normalize_id(pid), date_value, name, birth):
            written += 1
    return written


def _lab_kind_from_item(lab_name: Any, lab_code: Any) -> Optional[str]:
    code = _clean_text(lab_code).upper()
    if code in LAB_CODE_MAP:
        return LAB_CODE_MAP[code]
    name = _clean_text(lab_name).upper()
    if "HBA1C" in name or "醣化血紅素" in name or "糖化血紅素" in name:
        return "hba"
    if "LDL" in name or "低密度膽固醇" in name:
        return "ldl"
    if "UACR" in name or "尿液白蛋白肌酸酐" in name or "白蛋白肌酸酐比" in name:
        return "uacr"
    return None


def _add_lab_date(
    health_records: Dict[str, Dict[str, Any]],
    pid: Any,
    lab_kind: Optional[str],
    date_value: Any,
    name: Any = "",
    bday: Any = None,
) -> bool:
    pid_text = _normalize_id(pid)
    lab_date = _parse_date(date_value)
    if not _is_valid_id(pid_text) or lab_kind not in ("hba", "ldl", "uacr") or lab_date is None:
        return False
    rec = health_records.setdefault(pid_text, _empty_health_record(name, bday))
    if not rec.get("name"):
        rec["name"] = _display_text(name)
    if rec.get("bday") is None:
        rec["bday"] = _parse_date(bday)
    dt_key = f"{lab_kind}_dt"
    current = rec.get(dt_key)
    if current is None or lab_date > current:
        rec[dt_key] = lab_date
    return True


def _save_common_required_workbooks(clean_dir: Path, health_records: Optional[Dict[str, Dict[str, Any]]] = None) -> None:
    health_headers = [
        "家醫收案會員ID",
        "姓名",
        "生日",
        "最近一次HbA1c檢查結果(%)",
        "最近一次HbA1c檢查日期",
        "最近一次LDL檢查結果(mg/dL)",
        "最近一次LDL檢查日期",
        "最近一次UACR檢查結果(mg/gm)",
        "最近一次UACR檢查日期",
    ]
    rows = [health_headers]
    for pid, rec in sorted((health_records or {}).items()):
        rows.append([
            pid,
            rec.get("name") or "",
            rec.get("bday"),
            rec.get("hba") or "",
            rec.get("hba_dt"),
            rec.get("ldl") or "",
            rec.get("ldl_dt"),
            rec.get("uacr") or "",
            rec.get("uacr_dt"),
        ])
    _write_workbook(clean_dir / "003_清洗_HealthCase.xlsx", {"HealthCase": rows})


def _save_screening_workbook(
    clean_dir: Path,
    screening: Dict[str, Dict[str, Tuple[dt.date, str, Optional[dt.date]]]],
) -> None:
    sheets: Dict[str, List[List[Any]]] = {}
    for key, sheet_name in SCREENING_SHEET_NAMES.items():
        rows = [["ID", "姓名", "生日", "最後篩檢日期"]]
        for pid, (last_date, name, bday) in sorted(screening[key].items()):
            rows.append([pid, name, bday, last_date])
        sheets[sheet_name] = rows
    _write_workbook(clean_dir / "002_清洗_預防保健.xlsx", sheets)


def _save_contact_workbook(clean_dir: Path, contacts: Dict[str, Tuple[str, str]]) -> None:
    if not contacts:
        return
    rows = [["ID", "電話", "手機"]]
    for pid, (phone, mobile) in sorted(contacts.items()):
        rows.append([pid, phone, mobile])
    _write_workbook(clean_dir / "004_清洗_行動電話.xlsx", {"行動電話": rows})


def _read_first_sheet_rows(path: Path) -> Iterable[Tuple[Any, ...]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        for row in ws.iter_rows(values_only=True):
            yield tuple(row)
    finally:
        wb.close()


def _prepare_115x_roster(source_dir: Path, clean_dir: Path, result: CleanResult) -> set[str]:
    candidates = sorted(
        path
        for path in source_dir.glob("*.xlsx")
        if "115x" in path.name.lower() or "不選" in path.name or "不要" in path.name
    )
    x115_names: set[str] = set()
    row_count = 0
    for path in candidates:
        shutil.copy2(path, clean_dir / path.name)
        for idx, row in enumerate(_read_first_sheet_rows(path), start=1):
            if idx == 1:
                continue
            name = _clean_text(row[0] if row else "")
            if not name:
                continue
            x115_names.add(name)
            row_count += 1
    result.stats["115X來源筆數"] = row_count
    result.stats["115X不重複姓名數"] = len(x115_names)
    return x115_names


def _parse_fangding_self_select(path: Path) -> Tuple[List[List[Any]], Dict[str, Tuple[str, Optional[dt.date], str, str]], Dict[str, str]]:
    rows = [["ID", "姓名", "生日", "電話", "地址"]]
    by_name: DefaultDict[str, List[Tuple[str, Optional[dt.date], str, str]]] = defaultdict(list)
    id_to_name: Dict[str, str] = {}

    for idx, row in enumerate(_read_first_sheet_rows(path), start=1):
        if idx == 1:
            continue
        raw_key = _display_text(row[0] if len(row) > 0 else "")
        pid = _extract_last_tw_id(raw_key)
        name = _display_text(row[1] if len(row) > 1 else "")
        address = _display_text(row[2] if len(row) > 2 else "")
        phone = _normalize_phone(row[3] if len(row) > 3 else "")
        bday = _extract_birth_after_id(raw_key, pid)
        if not _is_valid_id(pid) or not name:
            continue
        rows.append([pid, name, bday, phone, address])
        by_name[_clean_text(name)].append((pid, bday, phone, address))
        id_to_name[pid] = name

    unique_by_name = {
        name: values[0]
        for name, values in by_name.items()
        if len({value[0] for value in values}) == 1
    }
    return rows, unique_by_name, id_to_name


def _clean_fangding(source_dir: Path, clean_dir: Path, with_screening: bool) -> CleanResult:
    result = CleanResult()
    screening: Dict[str, Dict[str, Tuple[dt.date, str, Optional[dt.date]]]] = {
        key: {} for key in SCREENING_SHEET_NAMES
    }
    contacts: Dict[str, Tuple[str, str]] = {}

    roster_path = source_dir / "115自選會員.xlsx"
    if not roster_path.exists():
        raise FileNotFoundError(f"找不到自選會員檔：{roster_path}")
    roster_rows, unique_by_name, id_to_name = _parse_fangding_self_select(roster_path)
    _write_workbook(clean_dir / "001_清洗_自選名單.xlsx", {"自選名單": roster_rows})
    result.stats["自選名單筆數"] = len(roster_rows) - 1

    month_rows: DefaultDict[str, List[List[Any]]] = defaultdict(
        lambda: [["ID", "姓名", "生日", "日期", "次數", "申請金額"]]
    )
    matched = 0
    skipped = 0
    fee_files = sorted(source_dir.glob("*健保醫療費用診次.xlsx"))
    for fee_path in fee_files:
        wb = load_workbook(fee_path, read_only=True, data_only=True)
        try:
            ws = wb.worksheets[0]
            for row in ws.iter_rows(min_row=2, values_only=True):
                visit_date = _parse_date(row[1] if len(row) > 1 else None)
                month_code = _date_to_month_code(visit_date)
                name_key = _clean_text(row[3] if len(row) > 3 else "")
                if not month_code or not name_key:
                    skipped += 1
                    continue
                matched_member = unique_by_name.get(name_key)
                if not matched_member:
                    skipped += 1
                    continue
                pid, bday, _phone, _address = matched_member
                amount = _to_float(row[9] if len(row) > 9 else 0)
                month_rows[month_code].append([pid, id_to_name.get(pid, name_key), bday, visit_date, 1, amount])
                _update_totals(result.totals, month_code, 1, amount)
                matched += 1
        finally:
            wb.close()
    if month_rows:
        _write_workbook(clean_dir / "000_清洗_月份統計.xlsx", dict(month_rows))
    result.stats["費用診次已匹配筆數"] = matched
    result.stats["費用診次未匹配筆數"] = skipped

    if with_screening:
        screening_path = source_dir / "成健名單.xlsx"
        if screening_path.exists():
            wb = load_workbook(screening_path, read_only=True, data_only=True)
            try:
                ws = wb["追蹤查詢明細表"] if "追蹤查詢明細表" in wb.sheetnames else wb.worksheets[0]
                written = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    name = row[4] if len(row) > 4 else ""
                    date_value = row[2] if len(row) > 2 else None
                    bday = row[14] if len(row) > 14 else None
                    pid = row[15] if len(row) > 15 else ""
                    code = row[16] if len(row) > 16 else ""
                    phone = _normalize_phone(row[19] if len(row) > 19 else "")
                    mobile = _normalize_phone(row[21] if len(row) > 21 else "")
                    pid_text = _normalize_id(pid)
                    if _is_valid_id(pid_text) and (phone or mobile):
                        old_phone, old_mobile = contacts.get(pid_text, ("", ""))
                        contacts[pid_text] = (old_phone or phone, old_mobile or mobile)
                    written += _add_screening_by_code(screening, code, pid, date_value, name, bday)
                result.stats["預防保健註記筆數"] = written
                result.stats["成健聯絡資料筆數"] = len(contacts)
            finally:
                wb.close()

    _save_screening_workbook(clean_dir, screening)
    _save_contact_workbook(clean_dir, contacts)
    _save_common_required_workbooks(clean_dir)
    return result


def _parse_yimin_roster(path: Path) -> Tuple[List[List[Any]], Dict[str, Tuple[str, Optional[dt.date]]]]:
    rows = [["ID", "姓名", "生日", "電話", "地址"]]
    by_name: DefaultDict[str, List[Tuple[str, Optional[dt.date]]]] = defaultdict(list)
    for idx, row in enumerate(_read_first_sheet_rows(path), start=1):
        if idx == 1:
            continue
        pid = _normalize_id(row[1] if len(row) > 1 else "")
        bday = _parse_date(row[2] if len(row) > 2 else None)
        name = _display_text(row[3] if len(row) > 3 else "")
        address = _display_text(row[5] if len(row) > 5 else "")
        phone = _normalize_phone(row[6] if len(row) > 6 else "")
        if not _is_valid_id(pid) or not name:
            continue
        rows.append([pid, name, bday, phone, address])
        by_name[_clean_text(name)].append((pid, bday))
    unique_by_name = {
        name: values[0]
        for name, values in by_name.items()
        if len({value[0] for value in values}) == 1
    }
    return rows, unique_by_name


def _clean_yimin(source_dir: Path, clean_dir: Path) -> CleanResult:
    result = CleanResult()
    screening: Dict[str, Dict[str, Tuple[dt.date, str, Optional[dt.date]]]] = {
        key: {} for key in SCREENING_SHEET_NAMES
    }

    roster_rows, unique_by_name = _parse_yimin_roster(source_dir / "115自選會員.xlsx")
    _write_workbook(clean_dir / "001_清洗_自選名單.xlsx", {"自選名單": roster_rows})
    result.stats["自選名單筆數"] = len(roster_rows) - 1

    month_rows: DefaultDict[str, List[List[Any]]] = defaultdict(
        lambda: [["ID", "姓名", "生日", "日期", "次數", "申請金額", "病歷號"]]
    )
    matched = 0
    skipped = 0
    for path in sorted((source_dir / "次數").glob("*.xlsx")):
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb["門診人次統計-資料明細"] if "門診人次統計-資料明細" in wb.sheetnames else wb.worksheets[0]
            for row in ws.iter_rows(min_row=2, values_only=True):
                visit_date = _parse_date(row[2] if len(row) > 2 else None)
                month_code = _date_to_month_code(visit_date) or _date_to_month_code(path.stem)
                name_key = _clean_text(row[6] if len(row) > 6 else "")
                if not month_code or not name_key:
                    skipped += 1
                    continue
                mapped = unique_by_name.get(name_key)
                if not mapped:
                    skipped += 1
                    continue
                pid, bday = mapped
                chart_no = _display_text(row[5] if len(row) > 5 else "")
                month_rows[month_code].append([pid, name_key, bday, visit_date or _month_code_to_date(month_code), 1, 0, chart_no])
                _update_totals(result.totals, month_code, 1, 0)
                matched += 1
        finally:
            wb.close()
    if month_rows:
        _write_workbook(clean_dir / "000_清洗_月份統計.xlsx", dict(month_rows))
    result.stats["次數已匹配筆數"] = matched
    result.stats["次數未匹配筆數"] = skipped

    prevention_path = source_dir / "預防保健.xlsx"
    if prevention_path.exists():
        wb = load_workbook(prevention_path, read_only=True, data_only=True)
        try:
            seen = set()
            written = 0
            for ws in wb.worksheets:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    date_value = row[0] if len(row) > 0 else None
                    name = _display_text(row[3] if len(row) > 3 else "")
                    bday = _parse_date(row[4] if len(row) > 4 else None)
                    code = row[6] if len(row) > 6 else ""
                    key = (_clean_text(name), bday, _parse_date(date_value), _clean_text(code).upper())
                    if key in seen:
                        continue
                    seen.add(key)
                    mapped = unique_by_name.get(_clean_text(name))
                    if not mapped:
                        continue
                    pid, roster_bday = mapped
                    written += _add_screening_by_code(screening, code, pid, date_value, name, bday or roster_bday)
            result.stats["預防保健註記筆數"] = written
        finally:
            wb.close()

    _save_screening_workbook(clean_dir, screening)
    _save_common_required_workbooks(clean_dir)
    return result


def _clean_zhou_yishan(source_dir: Path, clean_dir: Path) -> CleanResult:
    result = CleanResult()
    x115_names = _prepare_115x_roster(source_dir, clean_dir, result)
    screening: Dict[str, Dict[str, Tuple[dt.date, str, Optional[dt.date]]]] = {
        key: {} for key in SCREENING_SHEET_NAMES
    }
    health_records: Dict[str, Dict[str, Any]] = {}
    id_by_name_bday: DefaultDict[Tuple[str, str], set] = defaultdict(set)

    prevention_path = source_dir / "預防保健.xls"
    if prevention_path.exists():
        book = xlrd.open_workbook(str(prevention_path))
        sh = book.sheet_by_index(0)
        prevention_written = 0
        lab_written = 0
        for r in range(1, sh.nrows):
            row = [sh.cell_value(r, c) for c in range(sh.ncols)]
            date_value = row[0] if len(row) > 0 else None
            name = row[1] if len(row) > 1 else ""
            bday = _parse_date(row[2] if len(row) > 2 else None)
            pid = _normalize_id(row[3] if len(row) > 3 else "")
            code = row[4] if len(row) > 4 else ""
            lab_name = row[9] if len(row) > 9 else ""
            lab_code = row[10] if len(row) > 10 else ""
            if _is_valid_id(pid) and bday is not None:
                id_by_name_bday[(_clean_text(name), bday.isoformat())].add(pid)
            prevention_written += _add_screening_by_code(screening, code, pid, date_value, name, bday)
            if _add_lab_date(health_records, pid, _lab_kind_from_item(lab_name, lab_code), date_value, name, bday):
                lab_written += 1
        result.stats["預防保健註記筆數"] = prevention_written
        result.stats["檢驗日期註記筆數"] = lab_written

    roster_rows = [["姓名", "電話", "地址"]]
    overlap_115x_count = 0
    roster_path = source_dir / "115自選會員.xlsx"
    if roster_path.exists():
        for idx, row in enumerate(_read_first_sheet_rows(roster_path), start=1):
            if idx == 1:
                continue
            name = _display_text(row[0] if len(row) > 0 else "")
            phone = _normalize_phone(row[1] if len(row) > 1 else "")
            address = _display_text(row[2] if len(row) > 2 else "")
            if _clean_text(name) in x115_names:
                overlap_115x_count += 1
            if name:
                roster_rows.append([name, phone, address])
    _write_workbook(clean_dir / "001_清洗_自選名單.xlsx", {"自選名單": roster_rows})
    result.stats["自選名單筆數"] = len(roster_rows) - 1
    result.stats["自選名單與115X重疊筆數"] = overlap_115x_count

    month_rows: DefaultDict[str, List[List[Any]]] = defaultdict(
        lambda: [["ID", "病歷號", "姓名", "生日", "日期", "次數", "申請金額", "地址", "電話"]]
    )
    written = 0
    skipped = 0
    internal_count = 0
    real_id_count = 0
    for path in sorted((source_dir / "次數").glob("*.xlsx")):
        month_code = re.search(r"(1(?:14|15)\d{2})", path.stem)
        if not month_code:
            continue
        code = month_code.group(1)
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.worksheets[0]
            visit_date = _month_code_to_date(code)
            for row in ws.iter_rows(min_row=2, values_only=True):
                name = _display_text(row[1] if len(row) > 1 else "")
                bday = _parse_date(row[2] if len(row) > 2 else None)
                count = _to_float(row[4] if len(row) > 4 else 0)
                if not name or bday is None or count == 0:
                    skipped += 1
                    continue
                chart_no = _display_text(row[0] if len(row) > 0 else "")
                phone = _normalize_phone(row[3] if len(row) > 3 else "")
                address = _display_text(row[5] if len(row) > 5 else "")
                matched_ids = id_by_name_bday.get((_clean_text(name), bday.isoformat()), set())
                if len(matched_ids) == 1:
                    pid = next(iter(matched_ids))
                    real_id_count += 1
                else:
                    pid = _internal_id(f"周一珊|{chart_no}|{_clean_text(name)}|{bday.isoformat()}")
                    result.internal_ids.add(pid)
                    internal_count += 1
                month_rows[code].append([pid, chart_no, name, bday, visit_date, count, 0, address, phone])
                _update_totals(result.totals, code, count, 0)
                written += 1
        finally:
            wb.close()
    if month_rows:
        _write_workbook(clean_dir / "000_清洗_月份統計.xlsx", dict(month_rows))
    result.stats["次數清洗筆數"] = written
    result.stats["次數略過筆數"] = skipped
    result.stats["次數以預防保健ID匹配筆數"] = real_id_count
    result.stats["次數使用內部ID筆數"] = internal_count

    _save_screening_workbook(clean_dir, screening)
    _save_common_required_workbooks(clean_dir, health_records)
    return result


def _clean_one_clinic(source_dir: Path, clean_dir: Path) -> CleanResult:
    clean_dir.mkdir(parents=True, exist_ok=True)
    name = source_dir.name
    if "嘉齡" in name:
        return _clean_fangding(source_dir, clean_dir, with_screening=False)
    if "家齡" in name:
        return _clean_fangding(source_dir, clean_dir, with_screening=True)
    if "益民" in name:
        return _clean_yimin(source_dir, clean_dir)
    if "周一珊" in name:
        return _clean_zhou_yishan(source_dir, clean_dir)
    raise ValueError(f"不支援的診所資料夾：{source_dir}")


def _sum_output_member_totals(output_path: Path) -> Dict[str, float]:
    wb = load_workbook(output_path, read_only=True, data_only=True)
    try:
        if "會員總表" not in wb.sheetnames:
            raise ValueError("輸出檔缺少會員總表，無法核對 L/M/N/O")
        ws = wb["會員總表"]
        totals = {"L": 0.0, "M": 0.0, "N": 0.0, "O": 0.0}
        for row in ws.iter_rows(min_row=3, min_col=12, max_col=15, values_only=True):
            for key, value in zip(("L", "M", "N", "O"), row):
                totals[key] += _to_float(value)
        return totals
    finally:
        wb.close()


def _validate_output_totals(output_path: Path, expected: Dict[str, float]) -> None:
    actual = _sum_output_member_totals(output_path)
    labels = {"L": "114年次數", "M": "114年費用", "N": "115年次數", "O": "115年費用"}
    mismatches = []
    for key in ("L", "M", "N", "O"):
        if round(actual.get(key, 0.0), 2) != round(expected.get(key, 0.0), 2):
            mismatches.append(
                f"{key}欄{labels[key]}：清洗 {expected.get(key, 0.0):,.0f} / 輸出 {actual.get(key, 0.0):,.0f}"
            )
    if mismatches:
        raise ValueError("會員總表 L/M/N/O 核對不一致：\n" + "\n".join(mismatches))


def _prioritize_115x_for_overlaps(output_path: Path, expected_count: int) -> int:
    wb = load_workbook(output_path)
    try:
        if "會員總表" not in wb.sheetnames:
            raise ValueError("輸出檔缺少會員總表，無法處理選／不選重疊會員")
        self_select_sheet_name = "自選名單(從會員指標內容Key過來)"
        if self_select_sheet_name not in wb.sheetnames:
            raise ValueError("輸出檔缺少自選名單工作表，無法排除選／不選重疊會員")

        ws = wb["會員總表"]
        columns: Dict[str, int] = {}
        category_aliases = {"會員", "會員別"}
        id_aliases = {"身份證號碼", "身分證號碼", "身份證號", "身分證號", "ID"}
        targets = {"姓名", "是否為自選會員", "是否為115X"}
        for row in (1, 2):
            for col in range(1, ws.max_column + 1):
                header = _clean_text(ws.cell(row, col).value)
                if header in targets and header not in columns:
                    columns[header] = col
                elif header in category_aliases and "會員分類" not in columns:
                    columns["會員分類"] = col
                elif header in id_aliases and "身分證號" not in columns:
                    columns["身分證號"] = col
        missing = targets.difference(columns)
        if "會員分類" not in columns:
            missing.add("會員／會員別")
        if "身分證號" not in columns:
            missing.add("身分證號")
        if missing:
            raise ValueError("會員總表缺少欄位：" + "、".join(sorted(missing)))

        overlap_keys: set[Tuple[str, str]] = set()
        overlap_names: set[str] = set()
        for row in range(3, ws.max_row + 1):
            is_self = bool(_clean_text(ws.cell(row, columns["是否為自選會員"]).value))
            is_115x = bool(_clean_text(ws.cell(row, columns["是否為115X"]).value))
            if not (is_self and is_115x):
                continue
            category = _clean_text(ws.cell(row, columns["會員分類"]).value)
            if category != "E1/E2":
                raise ValueError(f"選／不選重疊會員分類不是 E1/E2：第 {row} 列為 {category or '空白'}")
            name = _clean_text(ws.cell(row, columns["姓名"]).value)
            pid = _normalize_id(ws.cell(row, columns["身分證號"]).value)
            overlap_keys.add((name, pid))
            overlap_names.add(name)

        if len(overlap_keys) != expected_count:
            raise ValueError(
                "選／不選重疊會員核對不一致："
                f"來源重疊 {expected_count} 筆 / 會員總表 E1/E2 {len(overlap_keys)} 筆"
            )

        ws_self = wb[self_select_sheet_name]
        removed = 0
        for row in range(ws_self.max_row, 2, -1):
            name = _clean_text(ws_self.cell(row, 1).value)
            pid = _normalize_id(ws_self.cell(row, 2).value)
            if (name, pid) in overlap_keys or name in overlap_names:
                ws_self.delete_rows(row, 1)
                removed += 1

        if removed != expected_count:
            raise ValueError(
                "自選名單排除重疊會員核對不一致："
                f"預期排除 {expected_count} 筆 / 實際排除 {removed} 筆"
            )
        if removed:
            wb.save(output_path)
        return removed
    finally:
        wb.close()


def _blank_internal_ids(output_path: Path, internal_ids: set) -> None:
    if not internal_ids:
        return
    wb = load_workbook(output_path)
    try:
        if "會員總表" not in wb.sheetnames:
            return
        ws = wb["會員總表"]
        id_col = None
        for row in (1, 2):
            for col in range(1, ws.max_column + 1):
                header = _clean_text(ws.cell(row, col).value)
                if header in ("身份證號碼", "身分證號碼", "身份證號", "身分證號", "ID"):
                    id_col = col
                    break
            if id_col:
                break
        if not id_col:
            return
        changed = False
        for row in range(3, ws.max_row + 1):
            value = _normalize_id(ws.cell(row, id_col).value)
            if value in internal_ids:
                ws.cell(row, id_col).value = None
                changed = True
        if changed:
            wb.save(output_path)
    finally:
        wb.close()


def _final_output_path(source_dir: Path, temp_output: Path) -> Path:
    timestamp = dt.datetime.now().strftime("%m%d_%H%M")
    clinic_code = _clinic_code_from_source(source_dir)
    clinic_name = _load_clinic_name_lookup().get(clinic_code, "")
    if not clinic_name:
        clinic_name = re.sub(r"^\d+", "", source_dir.name) or source_dir.name
    clinic_name = _sanitize_filename_component(clinic_name)
    filename = f"{clinic_name}選會員_{timestamp}.xlsx"
    candidate = source_dir.parent / filename
    if not candidate.exists():
        return candidate
    timestamp = dt.datetime.now().strftime("%m%d_%H%M_%S")
    return source_dir.parent / f"{clinic_name}選會員_{timestamp}.xlsx"


def process_one_clinic(source_dir: Path, template_path: Path) -> Tuple[Path, CleanResult]:
    source_dir = source_dir.resolve()
    template_path = template_path.resolve()
    if not source_dir.is_dir():
        raise ValueError(f"來源不是資料夾：{source_dir}")
    if not template_path.exists():
        raise ValueError(f"找不到樣板：{template_path}")

    generic = _load_generic_module()
    temp_root = Path(tempfile.mkdtemp(prefix="clinic_0608_clean_"))
    clean_dir = temp_root / source_dir.name
    try:
        result = _clean_one_clinic(source_dir, clean_dir)
        temp_output = Path(generic.process_excel(str(clean_dir), str(template_path)))
        _blank_internal_ids(temp_output, result.internal_ids)
        result.stats["選與不選重疊自選名單排除筆數"] = _prioritize_115x_for_overlaps(
            temp_output,
            result.stats.get("自選名單與115X重疊筆數", 0),
        )
        _validate_output_totals(temp_output, result.totals)
        final_path = _final_output_path(source_dir, temp_output)
        shutil.move(str(temp_output), str(final_path))
        return final_path, result
    finally:
        shutil.rmtree(str(temp_root), ignore_errors=True)


def _is_target_clinic_folder(path: Path) -> bool:
    return any(token in path.name for token in ("嘉齡", "家齡", "益民", "周一珊"))


def _choose_source_folder() -> Optional[Path]:
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    selected = filedialog.askdirectory(
        title="選擇嘉齡 / 家齡 / 益民 / 一珊來源資料夾",
    )
    root.destroy()
    return Path(selected) if selected else None


def _show_gui_message(title: str, message: str, error: bool = False) -> None:
    import tkinter as tk
    from tkinter import messagebox

    root = tk.Tk()
    root.withdraw()
    if error:
        messagebox.showerror(title, message)
    else:
        messagebox.showinfo(title, message)
    root.destroy()


def main() -> None:
    parser = argparse.ArgumentParser(description="嘉齡 / 家齡 / 益民 / 一珊 0608 前置清洗並呼叫通用程式")
    parser.add_argument("source", nargs="?", help="單一診所資料夾；未提供時會跳出資料夾選擇視窗")
    parser.add_argument("--template", default=str(SCRIPT_DIR / TEMPLATE_NAME), help="輸出樣板，預設選會員模板0430.xlsx")
    args = parser.parse_args()

    gui_mode = args.source is None
    if gui_mode:
        selected = _choose_source_folder()
        if selected is None:
            return
        source = selected.expanduser()
    else:
        source = Path(args.source).expanduser()

    template = Path(args.template).expanduser()
    if not _is_target_clinic_folder(source):
        message = f"請選擇嘉齡、家齡、益民或周一珊的單一診所來源資料夾：{source}"
        if gui_mode:
            _show_gui_message("錯誤", message, error=True)
            return
        raise ValueError(message)

    print(f"使用樣板：{template}", flush=True)
    outputs: List[str] = []
    try:
        print(f"\n開始處理：{source.name}", flush=True)
        output, result = process_one_clinic(source, template)
        outputs.append(str(output))
        print(f"成功輸出：{output}", flush=True)
        print(
            "會員總表 L/M/N/O 核對相符："
            f"L={result.totals['L']:,.0f}, M={result.totals['M']:,.0f}, "
            f"N={result.totals['N']:,.0f}, O={result.totals['O']:,.0f}",
            flush=True,
        )
        if result.stats:
            print("清洗統計：" + "；".join(f"{k}={v}" for k, v in result.stats.items()), flush=True)
    except Exception as exc:
        if gui_mode:
            _show_gui_message("錯誤", str(exc), error=True)
            return
        raise

    if gui_mode:
        _show_gui_message("完成", "已輸出：\n" + "\n".join(outputs))
        generic = _load_generic_module()
        for output in outputs:
            generic.open_file_cross_platform(output)


if __name__ == "__main__":
    main()
