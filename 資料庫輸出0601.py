# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import copy
import csv
import datetime as dt
import hashlib
import json
import os
import re
import shutil
import socket
import subprocess
import sys
import tempfile
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill


csv.field_size_limit(sys.maxsize)

PROJECT_DIR = Path(__file__).resolve().parent
DEFAULT_TEMPLATE = PROJECT_DIR / "115年6月指定名單格式" / "選會員模板0526.xlsx"
TZ = dt.timezone(dt.timedelta(hours=8))
EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
SKIP_PREFIXES = ("~$", ".")
FILL_PENDING = PatternFill("solid", fgColor="F4CCCC")
FILL_DONE_DATE = PatternFill("solid", fgColor="C6E0B4")
FONT_RED = Font(color="FF0000")
FONT_BLACK = Font(color="000000")


def normalize_id(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().upper()
    if text.endswith(".0"):
        text = text[:-2]
    return text.replace(" ", "")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def to_number(value: Any) -> float:
    text = normalize_text(value).replace(",", "")
    if text in {"", "-", "—", "–", "None", "NaN"}:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def clean_missing(value: Any) -> str:
    text = normalize_text(value)
    if text in {"", "-", "—", "–", "None", "NaN"}:
        return ""
    return text


def roc_month_from_text(value: Any) -> str:
    text = re.sub(r"\D", "", normalize_text(value))
    if len(text) >= 5:
        return text[:5]
    return ""


def roc_date_to_iso(value: Any) -> str:
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    text = normalize_text(value)
    if not text or text in {"-", "—", "–"}:
        return ""
    digits = re.sub(r"\D", "", text)
    if len(digits) == 7 and digits[:3].isdigit():
        year = int(digits[:3]) + 1911
        return f"{year:04d}-{digits[3:5]}-{digits[5:7]}"
    if len(digits) == 8 and digits[:4].isdigit():
        return f"{digits[:4]}-{digits[4:6]}-{digits[6:8]}"
    return text


def birth_to_iso(value: Any) -> str:
    text = normalize_text(value)
    if not text or text in {"-", "—", "–"}:
        return ""
    if re.match(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}$", text):
        return roc_date_to_iso(text)
    digits = re.sub(r"\D", "", text)
    if len(digits) == 7:
        return roc_date_to_iso(digits)
    if len(digits) == 6:
        year = int(digits[:3]) + 1911
        return f"{year:04d}-{digits[3:5]}-{digits[5:6].zfill(2)}"
    return text


def iso_to_date(value: Any) -> Optional[dt.date]:
    text = roc_date_to_iso(value)
    if not text:
        return None
    try:
        return dt.date.fromisoformat(text[:10])
    except ValueError:
        return None


def calculate_age(birth: Any, today: Optional[dt.date] = None) -> Optional[int]:
    bday = iso_to_date(birth)
    if not bday:
        return None
    today = today or dt.datetime.now(TZ).date()
    return today.year - bday.year - ((today.month, today.day) < (bday.month, bday.day))


def split_phone(phone: Any, mobile: Any = "") -> Tuple[str, str]:
    phone_text = normalize_text(phone)
    mobile_text = normalize_text(mobile)
    candidates = [mobile_text, phone_text]
    mobile_out = ""
    phone_out = ""
    for value in candidates:
        digits = re.sub(r"\D", "", value)
        if not value:
            continue
        if digits.startswith("09") and len(digits) >= 10 and not mobile_out:
            mobile_out = value
        elif not phone_out:
            phone_out = value
    return phone_out, mobile_out


def sex_from_id(pid: Any) -> str:
    text = normalize_id(pid)
    if len(text) >= 2:
        if text[1] in {"1", "8"}:
            return "男"
        if text[1] in {"2", "9"}:
            return "女"
    return ""


def disease_code_text(value: Any) -> str:
    text = normalize_text(value).upper()
    if text in {"1", "DM", "糖尿病_DM"}:
        return "DM"
    if text in {"2", "CKD", "初期慢性腎臟病_CKD"}:
        return "CKD"
    if text in {"3", "DKD"} or "DKD" in text:
        return "DKD"
    if text in {"4", "0", "NONE"}:
        return "None"
    return normalize_text(value)


def disease_class_text(disease: Any, ascvd_value: Any) -> str:
    dmk = disease_code_text(disease)
    has_ascvd = normalize_text(ascvd_value).lower() in {"1", "a", "b", "ascvd-a", "ascvd-b"}
    if dmk == "DM" and has_ascvd:
        return "DM+ASCVD"
    if dmk == "CKD" and has_ascvd:
        return "CKD+ASCVD"
    if dmk == "DKD" and has_ascvd:
        return "DKD+ASCVD"
    if has_ascvd:
        return "ASCVD"
    return "" if dmk in {"", "None"} else dmk


def doctor_ascvd_text(value: Any) -> str:
    text = normalize_text(value).lower()
    if text in {"1", "a"}:
        return "ASCVD-a"
    if text == "b":
        return "ASCVD-b"
    if text == "0":
        return "0"
    return normalize_text(value)


def has_active_ascvd(value: Any) -> bool:
    return normalize_text(value).lower() in {"1", "a", "b", "ascvd-a", "ascvd-b"}


def normalize_icd_code(value: Any) -> str:
    text = normalize_text(value).upper().replace(" ", "")
    if not text:
        return ""
    return re.sub(r"^([A-Z])(\d)(?=\.|$)", r"\g<1>0\g<2>", text)


def extract_icd_codes(value: Any) -> List[str]:
    text = normalize_text(value).upper()
    found = re.findall(r"\b[A-Z]\d{2,3}(?:\.\d+)?\b|\b[A-Z]\d[A-Z](?:\.\d+)?\b", text)
    result: List[str] = []
    for raw in found:
        code = normalize_icd_code(raw)
        if code and code not in result:
            result.append(code)
    return result


def compact_diagnosis_codes(*values: Any) -> str:
    codes: List[str] = []
    for value in values:
        for token in extract_icd_codes(value):
            if token not in codes:
                codes.append(token)
    return ",".join(codes)


def merge_diagnosis_codes(existing: Any, new_codes: Any) -> str:
    return compact_diagnosis_codes(existing, new_codes)


def classify_main_sub(codes_text: Any) -> str:
    text = normalize_text(codes_text).upper()
    codes = extract_icd_codes(text)
    has_dm = any(re.match(r"^E(08|09|10|11|12|13)", code) for code in codes)
    has_ckd = any(
        code in {"N18.1", "N18.2", "N18.3"}
        or re.match(r"^E(08|09|10|11|12|13)\.2", code)
        for code in codes
    )
    if has_dm and has_ckd:
        return "DM+CKD"
    if has_dm:
        return "DM"
    if has_ckd:
        return "CKD"
    return ""


def screening_status(kind: str, member: Dict[str, Any]) -> str:
    age = calculate_age(member.get("birth")) or 0
    sex = normalize_text(member.get("sex"))
    dmk = disease_code_text(member.get("disease_code"))
    date_value = iso_to_date(member.get(kind))
    has_list_row = kind in member

    if kind == "adult":
        if dmk in {"DM", "CKD", "DKD"}:
            return "不需受檢"
        if age < 30:
            return "不需受檢"
        if not has_list_row:
            return "不確定(主動確認+補做機會)"
        if date_value and date_value.year >= 2024:
            return "不需受檢"
        if date_value:
            return "過期需受檢"
        return "待受檢"

    if kind == "pap":
        if sex != "女" or age < 25:
            return "不需受檢"
        if date_value and date_value.year >= 2026:
            return date_value.isoformat()
        if date_value:
            return "過期需受檢"
        return "待受檢" if has_list_row else "不確定(主動確認+補做機會)"

    if kind == "flu":
        if age < 65:
            return "不需受檢"
        if date_value and date_value.year >= 2025:
            return "不需受檢"
        return "待受檢" if has_list_row else "不確定(主動確認+補做機會)"

    if kind == "fit":
        if not (45 <= age <= 75):
            return "不需受檢"
        if date_value and date_value.year >= 2026:
            return date_value.isoformat()
        if date_value:
            return "過期需受檢"
        return "待受檢" if has_list_row else "不確定(主動確認+補做機會)"

    if kind == "bc":
        if not (45 <= age < 80):
            return "不需受檢"
        if date_value:
            return date_value.isoformat()
        return "待受檢" if has_list_row else "不確定(主動確認+補做機會)"
    return ""


def lab_display(kind: str, value: Any, date_value: Any, member: Dict[str, Any]) -> str:
    text = clean_missing(value)
    if not text:
        return ""
    result_date = iso_to_date(date_value)
    number = to_number(text)
    suffix = ""
    dmk = disease_code_text(member.get("disease_code"))
    active_ascvd = has_active_ascvd(member.get("ASCVD"))
    if kind == "hba1c":
        if dmk not in {"DM", "DKD"}:
            return text
        if not result_date or result_date.year < 2026:
            suffix = "2026需受檢"
        elif number >= 7:
            suffix = "已受檢未達控制"
        if suffix == "2026需受檢" and re.fullmatch(r"\d+", text):
            text = f"{text}.0"
    elif kind == "ldl":
        if dmk not in {"DM", "CKD", "DKD"} and not active_ascvd:
            return text
        if not result_date or result_date.year < 2026:
            suffix = "2026需受檢"
        elif number >= 100:
            suffix = "已受檢未達控制"
        if suffix == "2026需受檢" and re.fullmatch(r"\d+", text):
            text = f"{text}.0"
    elif kind == "uacr":
        if dmk not in {"DM", "CKD", "DKD"}:
            return text
        if result_date and result_date.year >= 2026 and number >= 30:
            suffix = "已受檢未達控制"
    return f"{text}\n({suffix})" if suffix else text


def valid_lab_year(date_value: Any) -> bool:
    d = iso_to_date(date_value)
    return bool(d and d.year in {2025, 2026})


def followup_note(member: Dict[str, Any], days: int = 28) -> str:
    today = dt.date(2026, 6, 1)
    lines: List[str] = []
    dmk = disease_code_text(member.get("disease_code"))
    needs_hba = dmk in {"DM", "DKD"}
    needs_ldl = dmk in {"DM", "CKD", "DKD"} or has_active_ascvd(member.get("ASCVD"))
    if not needs_hba and not needs_ldl:
        return "非疾病不需回診" if days == 28 else ""
    for label, key, needed in [("HbA1c", "hba1c_date", needs_hba), ("LDL", "ldl_date", needs_ldl)]:
        if not needed:
            continue
        last = iso_to_date(member.get(key))
        if not last or last.year not in {2025, 2026}:
            lines.append(f"{label}:超過2年未檢查")
            continue
        due = last + dt.timedelta(days=days)
        if days == 28 and due <= today:
            lines.append(f"{label}:可立刻通知回診")
        else:
            base = today if days in {56, 84} and due <= today else last
            final_due = base + dt.timedelta(days=days)
            lines.append(f"{label}:{final_due:%Y-%m-%d}需回診")
    return "\n".join(lines)


def followup_after_immediate(member: Dict[str, Any], offset_days: int) -> str:
    today = dt.date(2026, 6, 1)
    lines = []
    dmk = disease_code_text(member.get("disease_code"))
    needs_hba = dmk in {"DM", "DKD"}
    needs_ldl = dmk in {"DM", "CKD", "DKD"} or has_active_ascvd(member.get("ASCVD"))
    if not needs_hba and not needs_ldl:
        return ""
    for label, key, needed in [("HbA1c", "hba1c_date", needs_hba), ("LDL", "ldl_date", needs_ldl)]:
        if not needed:
            continue
        last = iso_to_date(member.get(key))
        if last and last.year in {2025, 2026}:
            due = last + dt.timedelta(days=28)
            base = today if due <= today else due
        else:
            base = today
        lines.append(f"{label}:{base + dt.timedelta(days=offset_days):%Y-%m-%d}需回診")
    return "\n".join(lines)


def prevention_note(member: Dict[str, Any]) -> str:
    labels = [
        ("bc", "BC肝"),
        ("fit", "糞便"),
        ("pap", "子抹"),
        ("flu", "老流"),
        ("adult", "成人健檢"),
    ]
    lines = []
    for key, label in labels:
        status = screening_status(key, member)
        if status in {"待受檢", "過期需受檢", "不確定(主動確認+補做機會)"}:
            lines.append(f"今年需檢測{label}")
    return "\n".join(lines)


def visit_score(member: Dict[str, Any]) -> int:
    n115 = int(to_number(member.get("115_count")))
    if n115 >= 4:
        return 10
    if n115 >= 2:
        return 8
    if n115 == 1:
        return 6
    n114 = int(to_number(member.get("114_count")))
    if 8 <= n114 <= 20:
        return 4
    if 4 <= n114 <= 7:
        return 2
    return 1 if n114 > 0 else 0


def fee_score(member: Dict[str, Any]) -> int:
    avg114 = member.get("114_avg_amount")
    avg115 = member.get("115_avg_amount")
    if avg114 is None or avg115 is None:
        return 0
    return 6 if float(avg115) < float(avg114) else 0


def exam_score(member: Dict[str, Any]) -> int:
    dmk = disease_code_text(member.get("disease_code"))
    hba = to_number(member.get("hba1c"))
    ldl = to_number(member.get("ldl"))
    hba_date = iso_to_date(member.get("hba1c_date"))
    ldl_date = iso_to_date(member.get("ldl_date"))
    score = 0
    if dmk in {"DM", "DKD"} and hba_date and hba_date.year >= 2026 and hba < 7:
        score += 5
    if dmk in {"DM", "CKD", "DKD"} and ldl_date and ldl_date.year >= 2026 and ldl < 100:
        score += 5
    return score


def prevention_score(member: Dict[str, Any]) -> int:
    score = 28
    point_map = {"adult": 6, "pap": 6, "flu": 4, "fit": 6, "bc": 6}
    for key, points in point_map.items():
        if screening_status(key, member) in {"待受檢", "過期需受檢", "不確定(主動確認+補做機會)"}:
            score -= points
    return max(score, 0)


def score_parts(member: Dict[str, Any]) -> Tuple[int, int, int, int, int, str]:
    v = visit_score(member)
    f = fee_score(member)
    e = exam_score(member)
    p = prevention_score(member)
    total = v + f + e + p
    breakdown = "\n".join([
        f"1. 固定就診次數：{v} 分",
        f"2. 醫療費用：{f} 分",
        f"3. 糖心腎管理：{e} 分",
        f"4. 預防保健：{p} 分",
    ])
    return total, v, f, e, p, breakdown


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _find_psql() -> str:
    """跨平台尋找 psql 執行檔路徑。優先使用環境變數 CLINIC_PSQL，其次在 PATH 搜尋。"""
    import shutil
    env_val = os.getenv("CLINIC_PSQL")
    if env_val:
        return env_val
    found = shutil.which("psql")
    if found:
        return found
    # Windows 常見預設安裝路徑
    win_default = r"C:\Program Files\PostgreSQL\16\bin\psql.exe"
    if os.path.isfile(win_default):
        return win_default
    return "psql"  # 最後嘗試直接呼叫，讓系統 PATH 決定


def run_psql(args: Sequence[str], input_text: Optional[str] = None) -> str:
    cmd = [
        _find_psql(),
        "-h",
        os.getenv("CLINIC_DB_HOST", "localhost"),
        "-p",
        os.getenv("CLINIC_DB_PORT", "5432"),
        "-U",
        os.getenv("CLINIC_DB_USER", "thousand0001"),
        "-d",
        os.getenv("CLINIC_DB_NAME", "clinic_merge"),
        "-v",
        "ON_ERROR_STOP=1",
        *args,
    ]
    env = os.environ.copy()
    env.setdefault("PGPASSWORD", "0937930981")
    proc = subprocess.run(
        cmd,
        input=input_text,
        text=True,
        capture_output=True,
        env=env,
    )
    if proc.returncode != 0:
        raise RuntimeError(
            "PostgreSQL 指令失敗：\n"
            + (proc.stderr.strip() or proc.stdout.strip() or str(cmd))
        )
    return proc.stdout


def ensure_database_objects() -> None:
    sql = """
CREATE SCHEMA IF NOT EXISTS raw;

CREATE TABLE IF NOT EXISTS raw.uploaded_rows (
  uploaded_row_id BIGSERIAL PRIMARY KEY,
  batch_id UUID NOT NULL REFERENCES meta.import_batches(batch_id),
  clinic_id BIGINT NOT NULL REFERENCES meta.clinics(clinic_id),
  source_file_id BIGINT REFERENCES meta.source_files(source_file_id),
  file_name TEXT NOT NULL,
  sheet_name TEXT NOT NULL,
  row_no INTEGER NOT NULL,
  row_data JSONB NOT NULL,
  row_hash TEXT,
  created_at TIMESTAMPTZ NOT NULL DEFAULT now()
);

CREATE INDEX IF NOT EXISTS idx_uploaded_rows_batch
  ON raw.uploaded_rows(batch_id);

CREATE INDEX IF NOT EXISTS idx_uploaded_rows_clinic_sheet
  ON raw.uploaded_rows(clinic_id, sheet_name);

CREATE INDEX IF NOT EXISTS idx_uploaded_rows_row_data_gin
  ON raw.uploaded_rows USING GIN(row_data);

CREATE TABLE IF NOT EXISTS raw.current_uploaded_rows (
  current_row_id BIGSERIAL PRIMARY KEY,
  clinic_id BIGINT NOT NULL REFERENCES meta.clinics(clinic_id),
  source_file_id BIGINT REFERENCES meta.source_files(source_file_id),
  batch_id UUID NOT NULL REFERENCES meta.import_batches(batch_id),
  file_name TEXT NOT NULL,
  sheet_name TEXT NOT NULL,
  row_no INTEGER NOT NULL,
  row_data JSONB NOT NULL,
  row_hash TEXT,
  activated_at TIMESTAMPTZ NOT NULL DEFAULT now(),
  UNIQUE (clinic_id, file_name, sheet_name, row_no)
);

CREATE INDEX IF NOT EXISTS idx_current_uploaded_rows_clinic
  ON raw.current_uploaded_rows(clinic_id);

CREATE INDEX IF NOT EXISTS idx_current_uploaded_rows_file
  ON raw.current_uploaded_rows(clinic_id, file_name);

CREATE INDEX IF NOT EXISTS idx_current_uploaded_rows_data_gin
  ON raw.current_uploaded_rows USING GIN(row_data);

CREATE TABLE IF NOT EXISTS meta.generated_outputs (
  generated_output_id BIGSERIAL PRIMARY KEY,
  batch_id UUID REFERENCES meta.import_batches(batch_id),
  clinic_id BIGINT REFERENCES meta.clinics(clinic_id),
  output_path TEXT NOT NULL,
  template_path TEXT,
  row_count INTEGER NOT NULL DEFAULT 0,
  created_by TEXT,
  created_at TIMESTAMPTZ NOT NULL DEFAULT now()
);
"""
    run_psql(["-c", sql])


def detect_clinic_code(source_dir: Path, explicit_code: Optional[str]) -> str:
    if explicit_code:
        return explicit_code
    match = re.search(r"\d{10}", source_dir.name)
    if match:
        return match.group(0)
    match = re.search(r"\d{10}", str(source_dir))
    if match:
        return match.group(0)
    raise ValueError("無法從資料夾判斷醫事機構代碼，請用 --clinic-code 指定")


def get_clinic(clinic_code: str, fallback_name: str) -> Tuple[int, str]:
    sql = f"""
INSERT INTO meta.clinics (clinic_code, clinic_name, source_system)
VALUES ('{clinic_code}', '{fallback_name}', '通用輸出')
ON CONFLICT (clinic_code) DO UPDATE
SET updated_at = now()
RETURNING clinic_id, COALESCE(NULLIF(official_name, ''), clinic_name);
"""
    out = run_psql(["-At", "-F", "\t", "-c", sql]).strip()
    if not out:
        raise RuntimeError("無法取得診所資料")
    clinic_id, clinic_name = out.split("\t", 1)
    return int(clinic_id), clinic_name


def create_batch(clinic_id: int, source_dir: Path, actor: str) -> str:
    scope = json.dumps({"source_dir": str(source_dir)}, ensure_ascii=False)
    sql = f"""
INSERT INTO meta.import_batches (
  clinic_id, source_system, import_mode, scope, status, source_root, requested_by, message
)
VALUES (
  {clinic_id}, '通用輸出0601', 'replace_scope', '{scope}'::jsonb,
  'staged', '{source_dir}', '{actor}', '原始 Excel rows 已上傳 raw.uploaded_rows'
)
RETURNING batch_id;
"""
    out = run_psql(["-At", "-c", sql])
    match = re.search(r"[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}", out)
    if not match:
        raise RuntimeError("建立匯入批次失敗，沒有取得 batch_id")
    return match.group(0)


def mark_batch_failed(batch_id: str, message: str) -> None:
    safe_message = message.replace("'", "''")[:1000]
    sql = f"""
UPDATE meta.import_batches
SET status = 'failed',
    rejected_at = now(),
    message = COALESCE(message, '') || '；失敗：{safe_message}'
WHERE batch_id = '{batch_id}'::uuid
  AND status = 'staged';
"""
    run_psql(["-c", sql])


def excel_files(source_dir: Path) -> List[Path]:
    result: List[Path] = []
    for path in source_dir.rglob("*"):
        if not path.is_file():
            continue
        if path.name.startswith(SKIP_PREFIXES):
            continue
        if path.suffix.lower() in EXCEL_SUFFIXES:
            result.append(path)
    def sort_key(path: Path) -> Tuple[int, str]:
        month_match = re.search(r"(11[45]\d{2})", path.name)
        if month_match:
            return (0, month_match.group(1))
        return (1, path.name)

    return sorted(result, key=sort_key)


def make_headers(row: Sequence[Any]) -> List[str]:
    seen: Dict[str, int] = defaultdict(int)
    headers: List[str] = []
    for idx, value in enumerate(row, start=1):
        name = normalize_text(value) or f"欄{idx}"
        seen[name] += 1
        if seen[name] > 1:
            name = f"{name}_{seen[name]}"
        headers.append(name)
    return headers


def row_to_json(headers: Sequence[str], row: Sequence[Any]) -> Dict[str, str]:
    data: Dict[str, str] = {}
    for idx, header in enumerate(headers):
        value = row[idx] if idx < len(row) else None
        text = normalize_text(value)
        if text:
            data[header] = text
    return data


def header_score(row: Sequence[Any]) -> int:
    values = {normalize_text(value) for value in row if normalize_text(value)}
    score = 0
    header_sets = [
        {"院所ID", "ID", "BIRTHDAY", "個案類別"},
        {"身分證號", "姓名", "日期", "總額"},
        {"家醫收案會員ID", "姓名", "生日"},
        {"指標名稱", "ID", "生日", "姓名", "最後篩檢日期"},
        {"掛號証", "姓名", "來診日", "身份證號碼"},
    ]
    for headers in header_sets:
        matches = len(values & headers)
        if matches >= 2:
            score = max(score, matches * 10)
    nonblank = sum(1 for value in row if normalize_text(value))
    if nonblank >= 3:
        score += min(nonblank, 10)
    return score


def choose_header_index(rows: Sequence[Sequence[Any]]) -> int:
    best_index = 0
    best_score = -1
    for idx, row in enumerate(rows):
        score = header_score(row)
        if score >= best_score:
            best_score = score
            best_index = idx
    return best_index


def export_rows_to_csv(files: Sequence[Path], source_dir: Path, csv_path: Path) -> Tuple[int, Dict[str, int]]:
    counts: Dict[str, int] = defaultdict(int)
    total = 0
    with csv_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(["file_path", "file_name", "sheet_name", "row_no", "row_data", "row_hash"])
        for file_path in files:
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            except Exception as exc:
                print(f"略過無法讀取檔案：{file_path} ({exc})")
                continue
            try:
                rel_path = str(file_path.relative_to(source_dir))
                for ws in wb.worksheets:
                    row_iter = ws.iter_rows(values_only=True)
                    probe_rows: List[Sequence[Any]] = []
                    for _ in range(20):
                        try:
                            probe_rows.append(next(row_iter))
                        except StopIteration:
                            break
                    if not probe_rows:
                        continue
                    header_index = choose_header_index(probe_rows)
                    header_row = probe_rows[header_index]
                    headers = make_headers(header_row)
                    pending_rows: List[Tuple[int, Sequence[Any]]] = [
                        (idx + 1, row)
                        for idx, row in enumerate(probe_rows)
                        if idx > header_index
                    ]
                    for idx, row in enumerate(row_iter, start=len(probe_rows) + 1):
                        pending_rows.append((idx, row))
                    for row_no, row in pending_rows:
                        data = row_to_json(headers, row)
                        if not data:
                            continue
                        raw_json = json.dumps(data, ensure_ascii=False, sort_keys=True)
                        row_hash = hashlib.sha256(raw_json.encode("utf-8")).hexdigest()
                        writer.writerow([rel_path, file_path.name, ws.title, row_no, raw_json, row_hash])
                        total += 1
                        counts[file_path.name] += 1
            finally:
                wb.close()
    return total, dict(counts)


def upload_files_and_rows(
    clinic_id: int,
    batch_id: str,
    source_dir: Path,
    files: Sequence[Path],
    rows_csv: Path,
) -> None:
    files_csv = rows_csv.with_name("source_files.csv")
    with files_csv.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(["file_path", "file_name", "file_size", "sha256", "file_mtime"])
        for path in files:
            stat = path.stat()
            mtime = dt.datetime.fromtimestamp(stat.st_mtime, TZ).isoformat()
            writer.writerow([
                str(path.relative_to(source_dir)),
                path.name,
                stat.st_size,
                sha256_file(path),
                mtime,
            ])

    sql_path = rows_csv.with_suffix(".sql")
    sql_path.write_text(
        f"""
SET clinic.actor = '通用輸出0601';

CREATE TEMP TABLE tmp_source_files (
  file_path TEXT,
  file_name TEXT,
  file_size BIGINT,
  sha256 TEXT,
  file_mtime TIMESTAMPTZ
);
\\copy tmp_source_files FROM '{files_csv}' WITH (FORMAT csv, HEADER true)

INSERT INTO meta.source_files (
  clinic_id, batch_id, file_path, file_name, file_size, sha256, file_mtime, data_type
)
SELECT {clinic_id}, '{batch_id}'::uuid, file_path, file_name, file_size, sha256, file_mtime, 'excel'
FROM tmp_source_files
ON CONFLICT (clinic_id, file_path, COALESCE(sha256, '')) DO UPDATE
SET batch_id = EXCLUDED.batch_id,
    file_size = EXCLUDED.file_size,
    file_mtime = EXCLUDED.file_mtime,
    data_type = EXCLUDED.data_type;

CREATE TEMP TABLE tmp_uploaded_rows (
  file_path TEXT,
  file_name TEXT,
  sheet_name TEXT,
  row_no INTEGER,
  row_data JSONB,
  row_hash TEXT
);
\\copy tmp_uploaded_rows FROM '{rows_csv}' WITH (FORMAT csv, HEADER true)

DELETE FROM raw.uploaded_rows WHERE batch_id = '{batch_id}'::uuid;

INSERT INTO raw.uploaded_rows (
  batch_id, clinic_id, source_file_id, file_name, sheet_name, row_no, row_data, row_hash
)
SELECT
  '{batch_id}'::uuid,
  {clinic_id},
  sf.source_file_id,
  r.file_name,
  r.sheet_name,
  r.row_no,
  r.row_data,
  r.row_hash
FROM tmp_uploaded_rows r
LEFT JOIN meta.source_files sf
  ON sf.clinic_id = {clinic_id}
 AND sf.file_path = r.file_path;

DELETE FROM raw.current_uploaded_rows cur
USING (
  SELECT DISTINCT file_name
  FROM tmp_uploaded_rows
) touched
WHERE cur.clinic_id = {clinic_id}
  AND cur.file_name = touched.file_name;

INSERT INTO raw.current_uploaded_rows (
  clinic_id, source_file_id, batch_id, file_name, sheet_name, row_no, row_data, row_hash
)
SELECT
  {clinic_id},
  sf.source_file_id,
  '{batch_id}'::uuid,
  r.file_name,
  r.sheet_name,
  r.row_no,
  r.row_data,
  r.row_hash
FROM tmp_uploaded_rows r
LEFT JOIN meta.source_files sf
  ON sf.clinic_id = {clinic_id}
 AND sf.file_path = r.file_path
ON CONFLICT (clinic_id, file_name, sheet_name, row_no) DO UPDATE
SET source_file_id = EXCLUDED.source_file_id,
    batch_id = EXCLUDED.batch_id,
    row_data = EXCLUDED.row_data,
    row_hash = EXCLUDED.row_hash,
    activated_at = now();

INSERT INTO audit.operation_logs (
  actor, app_actor, action, schema_name, table_name, clinic_id, batch_id, after_data
)
VALUES (
  current_user, '通用輸出0601', 'upload_raw_rows', 'raw', 'uploaded_rows',
  {clinic_id}, '{batch_id}'::uuid,
  jsonb_build_object(
    'source_root', '{source_dir}',
    'file_count', {len(files)},
    'overwrite_rule', 'uploaded file_names replace current rows; untouched file_names are kept'
  )
);
""",
        encoding="utf-8",
    )
    run_psql(["-f", str(sql_path)])


def download_rows(batch_id: str, csv_path: Path) -> None:
    sql_path = csv_path.with_suffix(".sql")
    sql_path.write_text(
        "\\copy ("
        " SELECT file_name, sheet_name, row_no, row_data::text"
        " FROM raw.uploaded_rows"
        f" WHERE batch_id = '{batch_id}'::uuid"
        " ORDER BY file_name, sheet_name, row_no"
        f" ) TO '{csv_path}' WITH (FORMAT csv, HEADER true)\n",
        encoding="utf-8",
    )
    run_psql(["-f", str(sql_path)])


def download_current_rows(clinic_id: int, csv_path: Path) -> None:
    sql_path = csv_path.with_suffix(".sql")
    sql_path.write_text(
        "\\copy ("
        " SELECT file_name, sheet_name, row_no, row_data::text"
        " FROM raw.current_uploaded_rows"
        f" WHERE clinic_id = {clinic_id}"
        " ORDER BY file_name, sheet_name, row_no"
        f" ) TO '{csv_path}' WITH (FORMAT csv, HEADER true)\n",
        encoding="utf-8",
    )
    run_psql(["-f", str(sql_path)])


def load_db_rows(csv_path: Path) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with csv_path.open("r", newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            rows.append({
                "file_name": row["file_name"],
                "sheet_name": row["sheet_name"],
                "row_no": int(row["row_no"]),
                "row_data": json.loads(row["row_data"]),
            })

    def row_sort_key(item: Dict[str, Any]) -> Tuple[int, str, str, int]:
        file_name = item["file_name"]
        month_match = re.search(r"(11[45]\d{2})", file_name)
        if month_match:
            return (1, month_match.group(1), file_name, item["row_no"])
        if "家醫計畫" in file_name:
            priority = 0
        elif "不要自選會員" in file_name:
            priority = 8
        elif "自選會員" in file_name:
            priority = 9
        else:
            priority = 5
        return (priority, file_name, item["sheet_name"], item["row_no"])

    return sorted(rows, key=row_sort_key)


def infer_record_type(file_name: str, sheet_name: str, data: Dict[str, Any]) -> str:
    keys = set(data.keys())
    if {"院所ID", "ID", "BIRTHDAY", "個案類別"}.issubset(keys):
        return "designated"
    if {"身分證號", "姓名", "日期", "總額"}.issubset(keys):
        return "claim"
    if {"家醫收案會員ID", "最近一次HbA1c檢查結果(%)"}.issubset(keys):
        return "health"
    if {"家醫收案會員ID", "P4P收案計畫", "收案狀態"}.issubset(keys):
        return "p4p_case"
    if {"家醫收案會員ID", "P4P收案計畫", "收案日期"}.issubset(keys):
        return "p4p_track"
    if {"指標名稱", "ID", "最後篩檢日期"}.issubset(keys):
        return "screening"
    if "不要自選會員" in file_name and {"身份證號碼", "姓名"}.issubset(keys):
        return "exclude_select"
    if "自選會員" in file_name and {"身份證號碼", "姓名"}.issubset(keys):
        return "self_select"
    return "raw"


def merge_member(base: Dict[str, Any], new_values: Dict[str, Any]) -> None:
    for key, value in new_values.items():
        if value not in (None, "") and not base.get(key):
            base[key] = value


def build_members(rows: Sequence[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    members: Dict[str, Dict[str, Any]] = {}
    claims = defaultdict(lambda: {
        "114_count": 0.0,
        "114_count_full": 0.0,
        "115_count": 0.0,
        "114_amount_q1": 0.0,
        "115_amount_q1": 0.0,
        "114_amount_total": 0.0,
        "115_amount_total": 0.0,
        "last_visit": "",
    })
    claim_months_115: set = set()
    screenings: Dict[str, Dict[str, str]] = defaultdict(dict)

    for item in rows:
        data = item["row_data"]
        record_type = infer_record_type(item["file_name"], item["sheet_name"], data)
        if record_type == "designated":
            pid = normalize_id(data.get("ID"))
            if not pid or pid == "ID":
                continue
            member = members.setdefault(pid, {"id": pid})
            merge_member(member, {
                "birth": birth_to_iso(data.get("BIRTHDAY")),
                "designated": "✔",
                "is_114_member": "✔",
                "個案類別": data.get("個案類別"),
                "論質名單": data.get("論質名單"),
                "65歲以上多重慢性病註記": data.get("65歲以上多重慢性病註記"),
                "高診次註記": data.get("高診次註記"),
                "慢性病註記": data.get("慢性病註記"),
                "非慢性病註記": data.get("非慢性病註記"),
                "與前一年家醫收案診所相同": data.get("與前一年家醫收案診所相同"),
                "疾病樣態": data.get("疾病樣態"),
                "ASCVD": data.get("ASCVD"),
                "三高": data.get("三高"),
                "高血壓": data.get("高血壓"),
                "高血脂": data.get("高血脂"),
                "高血糖": data.get("高血糖"),
            })
            if member.get("designated"):
                disease_code = disease_code_text(data.get("疾病樣態"))
                member["disease_code"] = disease_code
                if disease_code == "None":
                    member["ASCVD"] = data.get("ASCVD")
                    member["doctor_ascvd"] = data.get("ASCVD")
                else:
                    member["ASCVD"] = "1"
                    member["doctor_ascvd"] = "ASCVD-a"
                member["disease_class"] = disease_class_text(member.get("disease_code"), member.get("ASCVD"))
        elif record_type == "claim":
            pid = normalize_id(data.get("身分證號"))
            if not pid:
                continue
            member = members.setdefault(pid, {"id": pid})
            phone_value, mobile_value = split_phone(data.get("電話"), data.get("手機號碼"))
            merge_member(member, {
                "name": data.get("姓名"),
                "birth": birth_to_iso(data.get("生日")),
                "phone": phone_value,
                "mobile": mobile_value,
                "address": data.get("地址"),
                "sex": data.get("性別") or data.get("性") or sex_from_id(pid),
            })
            dx = compact_diagnosis_codes(data.get("病1"), data.get("病23"))
            dx_class = classify_main_sub(dx)
            if dx:
                member["main_sub_dx"] = merge_diagnosis_codes(member.get("main_sub_dx"), dx)
                member["main_sub_class"] = classify_main_sub(member.get("main_sub_dx")) or dx_class
            service_date = roc_date_to_iso(data.get("日期"))
            if service_date and service_date > claims[pid]["last_visit"]:
                claims[pid]["last_visit"] = service_date
            month = roc_month_from_text(data.get("日期"))
            count = to_number(data.get("次數"))
            amount = to_number(data.get("總額"))
            if month.startswith("114"):
                claims[pid]["114_count_full"] += count
                claims[pid]["114_amount_total"] += amount
                if month <= "11404":
                    claims[pid]["114_count"] += count
                    claims[pid]["114_amount_q1"] += amount
            elif month.startswith("115"):
                claim_months_115.add(month)
                claims[pid]["115_count"] += count
                claims[pid]["115_amount_total"] += amount
                if month <= "11504":
                    claims[pid]["115_amount_q1"] += amount
        elif record_type == "screening":
            pid = normalize_id(data.get("ID"))
            if not pid:
                continue
            indicator = normalize_text(data.get("指標名稱"))
            date_value = roc_date_to_iso(data.get("最後篩檢日期"))
            if "成人" in indicator:
                screenings[pid]["adult"] = date_value
            elif "子宮" in indicator:
                screenings[pid]["pap"] = date_value
            elif "流感" in indicator:
                screenings[pid]["flu"] = date_value
            elif "糞便" in indicator:
                screenings[pid]["fit"] = date_value
            elif "肝" in indicator:
                screenings[pid]["bc"] = date_value
        elif record_type == "health":
            pid = normalize_id(data.get("家醫收案會員ID"))
            if not pid:
                continue
            member = members.setdefault(pid, {"id": pid})
            merge_member(member, {
                "name": data.get("姓名"),
                "birth": birth_to_iso(data.get("生日")),
                "hba1c": clean_missing(data.get("最近一次HbA1c檢查結果(%)")),
                "hba1c_date": roc_date_to_iso(data.get("最近一次HbA1c檢查日期")),
                "ldl": clean_missing(data.get("最近一次LDL檢查結果(mg/dL)")),
                "ldl_date": roc_date_to_iso(data.get("最近一次LDL檢查日期")),
                "uacr": clean_missing(data.get("最近一次UACR檢查結果(mg/gm)")),
                "uacr_date": roc_date_to_iso(data.get("最近一次UACR檢查日期")),
            })
        elif record_type in {"p4p_case", "p4p_track"}:
            pid = normalize_id(data.get("家醫收案會員ID"))
            if not pid:
                continue
            member = members.setdefault(pid, {"id": pid})
            merge_member(member, {
                "name": data.get("姓名"),
                "birth": birth_to_iso(data.get("生日")),
                "disease_text": disease_code_text(data.get("疾病樣態")),
                "p4p_plan": data.get("P4P收案計畫"),
                "p4p_status": data.get("收案狀態"),
                "p4p_enroll_date": roc_date_to_iso(data.get("收案日期")),
                "p4p_last_track": roc_date_to_iso(data.get("最後追蹤日")),
                "p4p_next_track": roc_date_to_iso(data.get("下次應追蹤日")),
                "p4p_overdue": data.get("逾期未追蹤"),
            })
            if not member.get("disease_code"):
                member["disease_code"] = disease_code_text(data.get("疾病樣態"))
        elif record_type in {"self_select", "exclude_select"}:
            pid = normalize_id(data.get("身份證號碼"))
            if not pid:
                continue
            member = members.setdefault(pid, {"id": pid})
            merge_member(member, {
                "name": data.get("姓名"),
                "birth": birth_to_iso(data.get("生日")),
                "address": data.get("地址"),
                "sex": data.get("性別") or sex_from_id(pid),
            })
            member["self_select" if record_type == "self_select" else "exclude_select"] = "✔"

    month_count_115 = max(len(claim_months_115), 1)
    for pid, sums in claims.items():
        member = members.setdefault(pid, {"id": pid})
        member["114_count_q1"] = sums["114_count"]
        member["114_count"] = sums["114_count_full"]
        member["114_amount"] = sums["114_amount_total"]
        member["115_count"] = sums["115_count"]
        member["115_amount"] = sums["115_amount_total"]
        member["last_visit"] = sums["last_visit"]
        member["114_avg_amount"] = round(sums["114_amount_total"] / 12, 2) if sums["114_amount_total"] else None
        member["115_avg_amount"] = round(sums["115_amount_q1"] / month_count_115, 2) if sums["115_amount_q1"] else None
    for pid, values in screenings.items():
        member = members.setdefault(pid, {"id": pid})
        member.update(values)
    for pid, member in members.items():
        member.setdefault("sex", sex_from_id(pid))
        if member.get("disease_code") and not member.get("disease_class"):
            member["disease_class"] = disease_class_text(member.get("disease_code"), member.get("ASCVD"))
    return members


def copy_template(template_path: Path, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)


def apply_body_style(ws, row_idx: int, start_col: int, end_col: int) -> None:
    ws.row_dimensions[row_idx].height = None


def apply_doctor_qz_styles(ws, row_idx: int) -> None:
    for col_idx in range(17, 27):
        cell = ws.cell(row_idx, col_idx)
        text = normalize_text(cell.value)
        cell.fill = PatternFill(fill_type=None)
        cell.font = copy.copy(FONT_BLACK)
        if col_idx in {17, 18, 19, 20} and text in {"待受檢", "過期需受檢"}:
            cell.fill = copy.copy(FILL_PENDING)
        if col_idx in {22, 24, 26}:
            date_value = iso_to_date(text)
            if date_value and date_value.year >= 2026:
                cell.fill = copy.copy(FILL_DONE_DATE)
        if col_idx in {21, 23, 25} and ("2026需受檢" in text or "已受檢未達控制" in text):
            cell.font = copy.copy(FONT_RED)


def excel_date(value: Any) -> Any:
    date_value = iso_to_date(value)
    if not date_value:
        return None if not normalize_text(value) else value
    return dt.datetime(date_value.year, date_value.month, date_value.day)


def clear_sheet_values(ws, start_row: int, end_col: int) -> None:
    for row in range(start_row, ws.max_row + 1):
        ws.row_dimensions[row].height = None
        for col in range(1, end_col + 1):
            ws.cell(row, col).value = None


def write_self_select_sheet(wb, ordered: Sequence[Dict[str, Any]]) -> None:
    sheet_name = "自選名單(從會員指標內容Key過來)"
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    clear_sheet_values(ws, 3, 26)
    for col in range(22, 27):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].hidden = True
    hidden_headers = ["分數", "固定就診次數", "醫療費用", "糖心腎管理", "預防保健"]
    for idx, header in enumerate(hidden_headers, start=22):
        ws.cell(2, idx).value = header

    out_row = 3
    for member in ordered:
        if not member.get("self_select"):
            continue
        total_score, v_score, f_score, e_score, p_score, _ = score_parts(member)
        ws.row_dimensions[out_row].height = None
        values = [
            member.get("name"),
            member.get("id"),
            member.get("114_count") or None,
            member.get("114_count_q1") or None,
            member.get("115_count") or None,
            member.get("114_amount") or None,
            member.get("115_amount") or None,
            excel_date(member.get("adult")),
            excel_date(member.get("pap")),
            excel_date(member.get("flu")),
            excel_date(member.get("fit")),
            excel_date(member.get("bc")),
            member.get("hba1c") or None,
            excel_date(member.get("hba1c_date")),
            member.get("ldl") or None,
            excel_date(member.get("ldl_date")),
            member.get("p4p_status") or None,
            excel_date(member.get("p4p_enroll_date")),
            excel_date(member.get("p4p_last_track")),
            excel_date(member.get("p4p_next_track")),
            member.get("is_114_member"),
            total_score,
            v_score,
            f_score,
            e_score,
            p_score,
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(out_row, col).value = value
        out_row += 1


def percentile_record(member: Dict[str, Any], metric: str, fill_kind: str) -> Dict[str, Any]:
    total_score, _, _, _, _, _ = score_parts(member)
    if metric == "ldl":
        return {
            "name": member.get("name"),
            "birth": excel_date(member.get("birth")),
            "id": member.get("id"),
            "score": total_score,
            "note": prevention_note(member),
            "value": member.get("ldl"),
            "date": excel_date(member.get("ldl_date")),
            "first": "\n".join([line for line in followup_note(member, 28).splitlines() if line.startswith("LDL:")]) or None,
            "second": "\n".join([line for line in followup_after_immediate(member, 28).splitlines() if line.startswith("LDL:")]) or None,
            "third": "\n".join([line for line in followup_after_immediate(member, 56).splitlines() if line.startswith("LDL:")]) or None,
            "fill": fill_kind,
        }
    return {
        "name": member.get("name"),
        "birth": excel_date(member.get("birth")),
        "id": member.get("id"),
        "score": total_score,
        "note": prevention_note(member),
        "value": member.get("hba1c"),
        "date": excel_date(member.get("hba1c_date")),
        "first": "\n".join([line for line in followup_note(member, 28).splitlines() if line.startswith("HbA1c:")]) or None,
        "second": "\n".join([line for line in followup_after_immediate(member, 28).splitlines() if line.startswith("HbA1c:")]) or None,
        "third": "\n".join([line for line in followup_after_immediate(member, 56).splitlines() if line.startswith("HbA1c:")]) or None,
        "fill": fill_kind,
    }


def write_percentile_side(ws, row_idx: int, start_col: int, record: Optional[Dict[str, Any]]) -> None:
    for offset in range(10):
        ws.cell(row_idx, start_col + offset).value = None
    if not record:
        return
    values = [
        record["name"], record["birth"], record["id"], record["score"], record["note"],
        record["value"], record["date"], record["first"], record["second"], record["third"],
    ]
    for offset, value in enumerate(values):
        ws.cell(row_idx, start_col + offset).value = value
    name_cell = ws.cell(row_idx, start_col)
    if record.get("fill") == "pink":
        name_cell.fill = PatternFill("solid", fgColor="EAC0C0")
    elif record.get("fill") == "blue":
        name_cell.fill = PatternFill("solid", fgColor="A9C2D9")


def write_percentile_sheet(wb, ordered: Sequence[Dict[str, Any]]) -> None:
    sheet_name = "百分位名單"
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    clear_sheet_values(ws, 5, 24)

    ldl_primary: List[Dict[str, Any]] = []
    ldl_secondary: List[Dict[str, Any]] = []
    hba_primary: List[Dict[str, Any]] = []
    hba_secondary: List[Dict[str, Any]] = []
    for member in ordered:
        dmk = disease_code_text(member.get("disease_code"))
        ldl_needed = dmk in {"DM", "CKD", "DKD"} or has_active_ascvd(member.get("ASCVD"))
        hba_needed = dmk in {"DM", "DKD"}
        ldl = to_number(member.get("ldl"))
        hba = to_number(member.get("hba1c"))
        if ldl_needed and ldl:
            if ldl <= 100:
                ldl_primary.append(percentile_record(member, "ldl", "pink"))
            elif ldl <= 91:
                ldl_secondary.append(percentile_record(member, "ldl", "blue"))
        if hba_needed and hba:
            if hba < 7:
                hba_primary.append(percentile_record(member, "hba", "pink"))
            elif hba <= 7:
                hba_secondary.append(percentile_record(member, "hba", "blue"))

    ldl_records = ldl_primary + ldl_secondary
    hba_records = hba_primary + hba_secondary
    ws["A1"] = f"LDL百分位(<=100，{len(ldl_primary)}人)、(<=91，{len(ldl_secondary)}人)"
    ws["N1"] = f"HBA1C百分位(<7，{len(hba_primary)}人)、(<=7，{len(hba_secondary)}人)"
    ws["A2"] = "紅色：達到標準、藍色：達到73.8%"
    ws["N2"] = "紅色：達到標準、藍色：達到73.8%"

    data_rows = max(len(ldl_records), len(hba_records), 1)
    for idx in range(data_rows):
        row_idx = 5 + idx
        ws.row_dimensions[row_idx].height = 66
        write_percentile_side(ws, row_idx, 1, ldl_records[idx] if idx < len(ldl_records) else None)
        write_percentile_side(ws, row_idx, 14, hba_records[idx] if idx < len(hba_records) else None)


def write_output(template_path: Path, output_path: Path, clinic_name: str, members: Dict[str, Dict[str, Any]]) -> int:
    copy_template(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb["會員總表"]
    ws_doc = wb["醫生看(從會員指標內容Key過來)"]
    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = True

    ordered = sorted(
        members.values(),
        key=lambda m: (
            0 if normalize_text(m.get("name")) else 1,
            0 if m.get("designated") else 1,
            str(m.get("name") or ""),
            str(m.get("id") or ""),
        ),
    )

    for row_idx, member in enumerate(ordered, start=3):
        total_score, v_score, f_score, e_score, p_score, breakdown = score_parts(member)
        apply_body_style(ws, row_idx, 1, 50)
        ws.cell(row_idx, 1).value = member.get("name")
        base_membership = "B" if disease_code_text(member.get("disease_code")) == "None" else "A"
        if member.get("exclude_select"):
            membership = f"{base_membership}/E2"
        elif member.get("self_select"):
            membership = f"{base_membership}/E1"
        elif member.get("designated"):
            membership = base_membership
        else:
            membership = None
        ws.cell(row_idx, 2).value = membership
        ws.cell(row_idx, 3).value = normalize_text(member.get("疾病樣態")) or member.get("disease_code")
        ws.cell(row_idx, 4).value = member.get("ASCVD")
        ws.cell(row_idx, 5).value = member.get("id")
        ws.cell(row_idx, 6).value = member.get("sex")
        ws.cell(row_idx, 7).value = member.get("birth")
        ws.cell(row_idx, 8).value = calculate_age(member.get("birth"))
        ws.cell(row_idx, 9).value = member.get("phone")
        ws.cell(row_idx, 10).value = member.get("mobile")
        ws.cell(row_idx, 11).value = member.get("last_visit")
        ws.cell(row_idx, 12).value = member.get("114_count")
        ws.cell(row_idx, 13).value = member.get("114_amount")
        ws.cell(row_idx, 14).value = member.get("115_count") or None
        ws.cell(row_idx, 15).value = member.get("115_amount") or None
        ws.cell(row_idx, 16).value = member.get("main_sub_dx")
        ws.cell(row_idx, 17).value = member.get("adult")
        ws.cell(row_idx, 18).value = member.get("pap")
        ws.cell(row_idx, 19).value = member.get("flu")
        ws.cell(row_idx, 20).value = member.get("fit")
        ws.cell(row_idx, 21).value = member.get("bc")
        ws.cell(row_idx, 22).value = member.get("disease_class")
        ws.cell(row_idx, 23).value = member.get("main_sub_class")
        ws.cell(row_idx, 24).value = member.get("hba1c")
        ws.cell(row_idx, 25).value = member.get("hba1c_date")
        ws.cell(row_idx, 26).value = member.get("ldl")
        ws.cell(row_idx, 27).value = member.get("ldl_date")
        ws.cell(row_idx, 28).value = member.get("uacr")
        ws.cell(row_idx, 29).value = member.get("uacr_date")
        ws.cell(row_idx, 30).value = member.get("p4p_plan")
        ws.cell(row_idx, 31).value = member.get("p4p_status")
        ws.cell(row_idx, 32).value = member.get("p4p_enroll_date")
        ws.cell(row_idx, 33).value = member.get("p4p_last_track")
        ws.cell(row_idx, 34).value = member.get("p4p_next_track")
        ws.cell(row_idx, 35).value = member.get("p4p_overdue")
        ws.cell(row_idx, 39).value = followup_note(member, 28)
        ws.cell(row_idx, 40).value = followup_after_immediate(member, 28)
        ws.cell(row_idx, 41).value = followup_after_immediate(member, 56)
        if to_number(member.get("hba1c")) and to_number(member.get("hba1c")) < 7:
            ws.cell(row_idx, 36).value = "✔"
        if to_number(member.get("ldl")) and to_number(member.get("ldl")) <= 100:
            ws.cell(row_idx, 37).value = "✔"
            ws.cell(row_idx, 45).value = "✔"
        if to_number(member.get("ldl")) and to_number(member.get("ldl")) <= 91:
            ws.cell(row_idx, 46).value = "✔"
        ws.cell(row_idx, 47).value = total_score
        ws.cell(row_idx, 48).value = breakdown
        ws.cell(row_idx, 49).value = prevention_note(member)

        doc_row = row_idx + 1
        apply_body_style(ws_doc, doc_row, 1, 62)
        ws_doc.cell(doc_row, 1).value = member.get("id")
        ws_doc.cell(doc_row, 2).value = member.get("name")
        ws_doc.cell(doc_row, 3).value = member.get("birth")
        ws_doc.cell(doc_row, 4).value = calculate_age(member.get("birth"))
        ws_doc.cell(doc_row, 5).value = member.get("phone")
        ws_doc.cell(doc_row, 6).value = member.get("mobile")
        ws_doc.cell(doc_row, 7).value = member.get("main_sub_dx")
        ws_doc.cell(doc_row, 8).value = member.get("disease_text") or member.get("disease_code") or disease_code_text(member.get("疾病樣態"))
        ws_doc.cell(doc_row, 9).value = member.get("doctor_ascvd") or doctor_ascvd_text(member.get("ASCVD"))
        ws_doc.cell(doc_row, 10).value = member.get("last_visit")
        ws_doc.cell(doc_row, 11).value = member.get("114_count")
        ws_doc.cell(doc_row, 12).value = member.get("114_count_q1") or None
        ws_doc.cell(doc_row, 13).value = member.get("115_count") or None
        ws_doc.cell(doc_row, 14).value = int(round(member.get("114_avg_amount"))) if member.get("114_avg_amount") is not None else None
        ws_doc.cell(doc_row, 15).value = int(round(member.get("115_avg_amount"))) if member.get("115_avg_amount") is not None else None
        ws_doc.cell(doc_row, 16).value = screening_status("adult", member)
        ws_doc.cell(doc_row, 17).value = screening_status("pap", member)
        ws_doc.cell(doc_row, 18).value = screening_status("flu", member)
        ws_doc.cell(doc_row, 19).value = screening_status("fit", member)
        ws_doc.cell(doc_row, 20).value = screening_status("bc", member)
        ws_doc.cell(doc_row, 21).value = lab_display("hba1c", member.get("hba1c"), member.get("hba1c_date"), member)
        ws_doc.cell(doc_row, 22).value = member.get("hba1c_date")
        ws_doc.cell(doc_row, 23).value = lab_display("ldl", member.get("ldl"), member.get("ldl_date"), member)
        ws_doc.cell(doc_row, 24).value = member.get("ldl_date")
        ws_doc.cell(doc_row, 25).value = lab_display("uacr", member.get("uacr"), member.get("uacr_date"), member)
        ws_doc.cell(doc_row, 26).value = member.get("uacr_date")
        p4p_status = member.get("p4p_status")
        p4p_plan = member.get("p4p_plan")
        if p4p_status and p4p_plan:
            p4p_status = f"{p4p_status}({p4p_plan})"
        ws_doc.cell(doc_row, 27).value = member.get("main_sub_class")
        ws_doc.cell(doc_row, 28).value = p4p_status
        ws_doc.cell(doc_row, 29).value = member.get("p4p_enroll_date")
        ws_doc.cell(doc_row, 30).value = member.get("p4p_last_track")
        ws_doc.cell(doc_row, 31).value = member.get("p4p_next_track")
        ws_doc.cell(doc_row, 32).value = member.get("exclude_select")
        ws_doc.cell(doc_row, 33).value = member.get("self_select")
        ws_doc.cell(doc_row, 34).value = member.get("is_114_member")
        ws_doc.cell(doc_row, 35).value = total_score
        ws_doc.cell(doc_row, 36).value = v_score
        ws_doc.cell(doc_row, 37).value = f_score
        ws_doc.cell(doc_row, 38).value = e_score
        ws_doc.cell(doc_row, 39).value = p_score
        ws_doc.cell(doc_row, 40).value = breakdown
        ws_doc.cell(doc_row, 41).value = prevention_note(member)
        ws_doc.cell(doc_row, 43).value = None
        apply_doctor_qz_styles(ws_doc, doc_row)
        designated_fields = [
            "個案類別", "論質名單", "65歲以上多重慢性病註記", "高診次註記",
            "慢性病註記", "非慢性病註記", "與前一年家醫收案診所相同", "疾病樣態",
            "ASCVD", "三高", "高血壓", "高血脂", "高血糖",
        ]
        for offset, field in enumerate(designated_fields, start=49):
            ws_doc.cell(doc_row, offset).value = member.get(field)
        if normalize_text(member.get("與前一年家醫收案診所相同")) == "1" and normalize_text(member.get("高血壓")) == "1":
            ws_doc.cell(doc_row, 62).value = "✔"

    write_percentile_sheet(wb, ordered)
    write_self_select_sheet(wb, ordered)

    wb.properties.title = f"{clinic_name}會員資料庫輸出"
    wb.save(output_path)
    wb.close()
    return len(ordered)


def record_output(batch_id: str, clinic_id: int, output_path: Path, template_path: Path, row_count: int, actor: str) -> None:
    sql = f"""
INSERT INTO meta.generated_outputs (
  batch_id, clinic_id, output_path, template_path, row_count, created_by
)
VALUES (
  '{batch_id}'::uuid, {clinic_id}, '{output_path}', '{template_path}', {row_count}, '{actor}'
);

UPDATE meta.import_batches
SET status = 'published',
    published_at = now(),
    published_by = '{actor}',
    message = COALESCE(message, '') || '；已產生選會員 Excel'
WHERE batch_id = '{batch_id}'::uuid;

INSERT INTO audit.operation_logs (
  actor, app_actor, action, schema_name, table_name, clinic_id, batch_id, after_data
)
VALUES (
  current_user, '通用輸出0601', 'generate_member_excel', 'meta', 'generated_outputs',
  {clinic_id}, '{batch_id}'::uuid,
  jsonb_build_object('output_path', '{output_path}', 'row_count', {row_count})
);
"""
    run_psql(["-c", sql])


def rollback_last_upload(clinic_code: str, actor: Optional[str] = None) -> str:
    actor = actor or os.getenv("USER") or "clinic_user"
    ensure_database_objects()
    clinic_id, clinic_name = get_clinic(clinic_code, clinic_code)
    sql = f"""
WITH last_batch AS (
  SELECT b.batch_id, b.started_at
  FROM meta.import_batches b
  WHERE b.clinic_id = {clinic_id}
    AND b.source_system = '通用輸出0601'
    AND b.status IN ('staged', 'published')
  ORDER BY b.started_at DESC
  LIMIT 1
),
touched AS (
  SELECT DISTINCT u.file_name
  FROM raw.uploaded_rows u
  JOIN last_batch lb ON lb.batch_id = u.batch_id
),
deleted_current AS (
  DELETE FROM raw.current_uploaded_rows cur
  USING touched t
  WHERE cur.clinic_id = {clinic_id}
    AND cur.file_name = t.file_name
  RETURNING cur.file_name
),
previous_rows AS (
  SELECT DISTINCT ON (u.file_name, u.sheet_name, u.row_no)
    u.clinic_id,
    u.source_file_id,
    u.batch_id,
    u.file_name,
    u.sheet_name,
    u.row_no,
    u.row_data,
    u.row_hash
  FROM raw.uploaded_rows u
  JOIN touched t ON t.file_name = u.file_name
  JOIN meta.import_batches b ON b.batch_id = u.batch_id
  JOIN last_batch lb ON true
  WHERE u.clinic_id = {clinic_id}
    AND b.started_at < lb.started_at
    AND b.status IN ('staged', 'published', 'superseded')
  ORDER BY u.file_name, u.sheet_name, u.row_no, b.started_at DESC
),
restored AS (
  INSERT INTO raw.current_uploaded_rows (
    clinic_id, source_file_id, batch_id, file_name, sheet_name, row_no, row_data, row_hash
  )
  SELECT clinic_id, source_file_id, batch_id, file_name, sheet_name, row_no, row_data, row_hash
  FROM previous_rows
  ON CONFLICT (clinic_id, file_name, sheet_name, row_no) DO UPDATE
  SET source_file_id = EXCLUDED.source_file_id,
      batch_id = EXCLUDED.batch_id,
      row_data = EXCLUDED.row_data,
      row_hash = EXCLUDED.row_hash,
      activated_at = now()
  RETURNING file_name
),
marked AS (
  UPDATE meta.import_batches b
  SET status = 'rejected',
      rejected_at = now(),
      message = COALESCE(message, '') || '；已 rollback，目前有效資料回到上一版'
  FROM last_batch lb
  WHERE b.batch_id = lb.batch_id
  RETURNING b.batch_id
),
logged AS (
  INSERT INTO audit.operation_logs (
    actor, app_actor, action, schema_name, table_name, clinic_id, batch_id, after_data
  )
  SELECT
    current_user,
    '通用輸出0601',
    'rollback_last_upload',
    'raw',
    'current_uploaded_rows',
    {clinic_id},
    marked.batch_id,
    jsonb_build_object(
      'clinic_code', '{clinic_code}',
      'clinic_name', '{clinic_name}',
      'touched_files', (SELECT jsonb_agg(file_name) FROM touched),
      'restored_rows', (SELECT count(*) FROM restored),
      'actor', '{actor}'
    )
  FROM marked
)
SELECT
  COALESCE((SELECT batch_id::text FROM marked), ''),
  COALESCE((SELECT count(*)::text FROM touched), '0'),
  COALESCE((SELECT count(*)::text FROM restored), '0');
"""
    out = run_psql(["-At", "-F", "\t", "-c", sql]).strip()
    if not out:
        raise RuntimeError("沒有可 rollback 的上傳批次")
    batch_id, touched_count, restored_count = out.split("\t")
    if not batch_id:
        raise RuntimeError(f"{clinic_code} 沒有可 rollback 的上傳批次")
    message = f"已 rollback {clinic_name}：批次 {batch_id}，影響檔案 {touched_count} 個，還原列 {restored_count} 列"
    print(message)
    return message


def timestamp() -> str:
    return dt.datetime.now(TZ).strftime("%m%d%H%M")


def process(
    source_dir: Path,
    clinic_code: Optional[str] = None,
    clinic_name: Optional[str] = None,
    output_dir: Optional[Path] = None,
    template_path: Path = DEFAULT_TEMPLATE,
    actor: Optional[str] = None,
) -> Path:
    if not source_dir.is_dir():
        raise FileNotFoundError(f"找不到來源資料夾：{source_dir}")
    if not template_path.is_file():
        raise FileNotFoundError(f"找不到模板：{template_path}")

    actor = actor or os.getenv("USER") or "clinic_user"
    ensure_database_objects()
    code = detect_clinic_code(source_dir, clinic_code)
    fallback_name = clinic_name or re.sub(r"^\d{10}", "", source_dir.name).replace("_x", "") or code
    clinic_id, db_clinic_name = get_clinic(code, fallback_name)
    final_clinic_name = clinic_name or db_clinic_name or fallback_name
    batch_id = create_batch(clinic_id, source_dir, actor)
    try:
        files = excel_files(source_dir)
        if not files:
            raise ValueError(f"來源資料夾沒有 xlsx/xlsm：{source_dir}")

        with tempfile.TemporaryDirectory(prefix="clinic_db_upload_") as tmp:
            tmp_path = Path(tmp)
            rows_csv = tmp_path / "uploaded_rows.csv"
            total_rows, counts = export_rows_to_csv(files, source_dir, rows_csv)
            upload_files_and_rows(clinic_id, batch_id, source_dir, files, rows_csv)
            db_rows_csv = tmp_path / "db_current_rows.csv"
            download_current_rows(clinic_id, db_rows_csv)
            db_rows = load_db_rows(db_rows_csv)
            members = build_members(db_rows)

        target_dir = output_dir or source_dir.parent
        safe_name = final_clinic_name.replace("診所", "")
        output_path = target_dir / f"{safe_name}會員{timestamp()}.xlsx"
        row_count = write_output(template_path, output_path, final_clinic_name, members)
        record_output(batch_id, clinic_id, output_path, template_path, row_count, actor)
    except Exception as exc:
        mark_batch_failed(batch_id, str(exc))
        raise

    print(f"批次：{batch_id}")
    print(f"上傳檔案：{len(files)} 個")
    print(f"上傳原始列：{total_rows} 列")
    print(f"建立會員列：{row_count} 列")
    for file_name, count in sorted(counts.items()):
        print(f"  {file_name}: {count}")
    print(f"輸出：{output_path}")
    return output_path


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="從 PostgreSQL 資料庫產生通用輸出，必要時先匯入診所 Excel")
    parser.add_argument("--rollback-last", action="store_true", help="回復指定診所最後一次資料庫上傳")
    parser.add_argument("source_dir", help="診所來源資料夾")
    parser.add_argument("--clinic-code", help="醫事機構代碼，未指定時從資料夾名稱抓 10 碼")
    parser.add_argument("--clinic-name", help="輸出檔使用的診所名稱")
    parser.add_argument("--output-dir", help="輸出資料夾，預設來源資料夾上一層")
    parser.add_argument("--template", default=str(DEFAULT_TEMPLATE), help="選會員模板路徑")
    parser.add_argument("--actor", default=os.getenv("USER") or socket.gethostname(), help="操作者")
    args = parser.parse_args(argv)

    if args.rollback_last:
        code = args.clinic_code or detect_clinic_code(Path(args.source_dir), None)
        rollback_last_upload(code, actor=args.actor)
        return 0

    process(
        Path(args.source_dir),
        clinic_code=args.clinic_code,
        clinic_name=args.clinic_name,
        output_dir=Path(args.output_dir) if args.output_dir else None,
        template_path=Path(args.template),
        actor=args.actor,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
