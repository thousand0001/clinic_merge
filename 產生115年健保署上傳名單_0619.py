# -*- coding: utf-8 -*-
"""
115年健保署會員名單上傳工具

從選會員 xlsx（醫生看分頁 AQ/AG 欄勾選）產生健保署上傳 FM.txt。
名單優先順序：AF 不選汰除 > AQ 指定會員 > AG 自選會員。

每位醫事人員各產生一份 FM.txt（流水號 01, 02, ...）。

輸出路徑：Desktop/健保署115年上傳名單/{院所代碼}{院所名稱}/自選會員/
格式：CP950 編碼，定長 208 bytes（家庭醫師整合性照護計畫名單格式）
每檔最多 9999 筆，超出自動拆檔（流水號遞增）
"""
from __future__ import annotations

import datetime
import re
import sys
import tkinter as tk
import unicodedata
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

import openpyxl

# ── 常數 ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR    = Path(__file__).resolve().parent
DOCTOR_KW     = "醫生看"
DATA_ROW      = 4           # 醫生看資料從第 4 列開始
CHECK_VALS    = {"✔", "✓", "v", "V"}

# 醫生看欄位（0-based index）
IDX_PID         = 0   # A: 身份證號碼
IDX_NAME        = 1   # B: 姓名
IDX_BIRTH       = 2   # C: 生日
IDX_TEL         = 4   # E: 電話
IDX_MOBILE      = 5   # F: 手機號碼
IDX_115X        = 31  # AF: 不選（115X）
IDX_SELF        = 32  # AG: 自選會員
IDX_DESIGNATED  = 42  # AQ: 115年指定會員

# 會員總表欄位（0-based index）
MASTER_IDX_PID  = 4   # E: 身份證號碼
MASTER_IDX_TEL  = 8   # I: 電話
MASTER_IDX_MOB  = 9   # J: 手機號碼
MASTER_IDX_ADDR = 57  # BF: 地址（0-based；A=0, ..., BF=57）
MASTER_SHEET_KW = "會員總表"

BRANCH       = 1      # 業務組別（台北=1）
PLAN_NO      = "17"   # 2026年計劃期別
MAX_PER_FILE = 9999

# 個案類別（case_type）
CASE_SELF       = "B"   # 自選名單
CASE_DESIGNATED = "A"   # 指定名單
CASE_QCARE      = "C"   # 論質名單
TZ_TW = datetime.timezone(datetime.timedelta(hours=8))

DESKTOP = Path.home() / "Desktop"
OUTPUT_ROOT = DESKTOP / "健保署115年上傳名單"

SELF_SUBFOLDER       = "自選會員"
DESIGNATED_SUBFOLDER = "指定會員"


# ── 工具函式 ──────────────────────────────────────────────────────────────────
def normalize_id(value) -> str:
    if value is None:
        return ""
    s = str(value).strip().upper()
    if s.endswith(".0"):
        s = s[:-2]
    return s.replace(" ", "")


def id_format_error(value, allow_legacy_resident: bool = False) -> str:
    pid = normalize_id(value)
    if not pid:
        return "身分證號為空白"
    valid_patterns = [r"[A-Z][1289]\d{8}"]
    if allow_legacy_resident:
        valid_patterns.append(r"[A-Z][A-D]\d{8}")
    if not any(re.fullmatch(pattern, pid) for pattern in valid_patterns):
        if allow_legacy_resident:
            return (
                "身分證號格式不符（須為英文字母 + 1/2/8/9 + 8 碼數字，"
                "或舊式統一證號 2 碼英文字母 + 8 碼數字）"
            )
        return "身分證號格式不符（須為英文字母 + 1/2/8/9 + 8 碼數字）"
    return ""


def to_yyyymmdd(v) -> str:
    """各種日期格式 → YYYYMMDD 字串；無法辨識回傳空字串。"""
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y%m%d")
    s = str(v).strip()
    if not s:
        return ""

    year = month = day = None
    if s.isdigit() and len(s) in (7, 8):
        if len(s) == 7:
            year, month, day = int(s[:3]) + 1911, int(s[3:5]), int(s[5:7])
        else:
            year, month, day = int(s[:4]), int(s[4:6]), int(s[6:8])
    else:
        match = re.fullmatch(r"(\d{3,4})[/-](\d{1,2})[/-](\d{1,2})", s)
        if match:
            year, month, day = map(int, match.groups())
            if year < 1000:
                year += 1911

    if year is not None:
        try:
            return datetime.date(year, month, day).strftime("%Y%m%d")
        except ValueError:
            return ""
    return ""


def sex_from_id(pid: str) -> str:
    if not pid or len(pid) < 2:
        return " "
    g = str(pid)[1].upper()
    if g in ("1", "8", "A", "C"):
        return "1"
    if g in ("2", "9", "B", "D"):
        return "2"
    return " "



def is_checked(val) -> bool:
    return str(val).strip() in CHECK_VALS if val is not None else False


def fb(value, byte_len: int) -> bytes:
    """固定長度 bytes 欄位（CP950，不足補空白，超出截斷）。"""
    encoded = bytearray()
    for char in ("" if value is None else str(value)).strip():
        char_bytes = char.encode("cp950", errors="replace")
        if len(encoded) + len(char_bytes) > byte_len:
            break
        encoded.extend(char_bytes)
    return bytes(encoded) + b" " * (byte_len - len(encoded))


def normalize_tel(raw: str) -> str:
    """清理電話號碼格式。"""
    digits = re.sub(r"\D", "", raw)
    if len(digits) == 9:
        if digits[0] in ("9", "2"):
            digits = "0" + digits
    elif len(digits) == 8:
        digits = "02" + digits
    return digits


def taiwan_today() -> datetime.date:
    return datetime.datetime.now(TZ_TW).date()


def format_member_errors(
    errors: list[dict],
    limit: int = 20,
    title: str = "選取資料含無法輸出的欄位：",
) -> str:
    lines = [title]
    for item in errors[:limit]:
        lines.append(
            f"第 {item['row']} 列｜姓名：{item['name'] or '空白'}｜"
            f"原值：{item['value'] or '空白'}｜原因：{item['reason']}"
        )
    if len(errors) > limit:
        lines.append(f"另有 {len(errors) - limit} 筆未列出。")
    return "\n".join(lines)


def garbled_name_reason(value) -> str:
    """回傳姓名中確定不適合 CP950 上傳的原因；正常時回傳空字串。"""
    name = "" if value is None else str(value).strip()
    if not name:
        return "姓名為空白"
    problems: list[str] = []
    for char in name:
        codepoint = f"U+{ord(char):04X}"
        if char in {"�", "?", "？"}:
            problems.append(f"疑似替代字元 {char!r}（{codepoint}）")
            continue
        if unicodedata.category(char) in {"Co", "Cs"}:
            problems.append(f"私用區字元（{codepoint}）")
            continue
        try:
            char.encode("cp950")
        except UnicodeEncodeError:
            problems.append(f"CP950 無法編碼字元 {char!r}（{codepoint}）")
    return "、".join(dict.fromkeys(problems))


def filter_garbled_names(members: list[dict]) -> tuple[list[dict], list[dict]]:
    """排除姓名含確定亂碼的會員，並保留摘要所需資訊。"""
    accepted: list[dict] = []
    excluded: list[dict] = []
    for member in members:
        reason = garbled_name_reason(member.get("name"))
        if reason:
            excluded.append({
                "pid": member.get("pid", ""),
                "name": member.get("name", ""),
                "reason": reason,
            })
        else:
            accepted.append(member)
    return accepted, excluded


def format_garbled_summary(errors: list[dict], limit: int = 10) -> list[str]:
    lines = [f"亂碼姓名排除：{len(errors)} 筆"]
    for item in errors[:limit]:
        lines.append(
            f"{item['pid'] or 'ID空白'}｜{item['name'] or '姓名空白'}｜{item['reason']}"
        )
    if len(errors) > limit:
        lines.append(f"另有 {len(errors) - limit} 筆未列出。")
    return lines


def make_record(
    hosp_id: str, pid: str, birthday: str, name: str, sex: str,
    addr: str, tel: str, prsn_id: str, case_type: str, case_date: str,
) -> bytes:
    rec = (
        b"A"
        + PLAN_NO.encode()
        + str(BRANCH).encode()
        + fb(hosp_id, 10)
        + fb(pid, 10)
        + fb(birthday, 8)
        + fb(name, 12)
        + sex.encode()
        + fb(addr, 120)
        + fb(tel, 15)
        + fb(prsn_id, 10)
        + case_type.encode()
        + fb(case_date, 8)
        + fb("", 8)   # 結案日期（空白）
        + b" "         # 結案原因（空白）
    )
    assert len(rec) == 208, f"記錄長度錯誤：{len(rec)}"
    return rec


# ── 醫療群資料查詢 ─────────────────────────────────────────────────────────────
def _strip_float_suffix(s: str) -> str:
    """移除 Excel 數值轉字串後的 '.0' 浮點後綴，不影響其他末尾數字。"""
    return s[:-2] if s.endswith(".0") else s


def load_clinic_lookup() -> dict:
    """讀取 醫療群_衛福部資料.xlsx，回傳 {名稱: {...}, 代碼: {...}} 雙鍵 dict。"""
    xls = SCRIPT_DIR / "醫療群_衛福部資料.xlsx"
    if not xls.exists():
        return {}
    wb = openpyxl.load_workbook(xls, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    result: dict = {}
    for row in ws.iter_rows(min_row=3, values_only=True):  # 從第3列（跳標題）
        code    = _strip_float_suffix(str(row[1]).strip()) if row[1] else ""
        name    = str(row[3]).strip() if row[3] else ""
        addr    = str(row[4]).strip() if row[4] else ""
        area    = _strip_float_suffix(str(row[5]).strip()).lstrip("0") if row[5] else ""
        tel_num = _strip_float_suffix(str(row[6]).strip()) if row[6] else ""
        tel     = f"0{area}{tel_num}" if area and tel_num else ""
        if code and code != "None" and name and name != "None":
            entry = {"code": code, "name": name, "addr": addr, "tel": tel}
            result[name] = entry
            result[code] = entry
    wb.close()
    return result


def find_clinic_info(filename_stem: str, lookup: dict) -> tuple:
    """
    從檔名推斷診所名稱，查詢代碼/地址/電話。
    回傳 (hosp_code, hosp_name, addr, tel)
    """
    name = re.sub(r"選會員.*$", "", filename_stem).strip()
    # 完整比對
    if name in lookup:
        e = lookup[name]
        return e["code"], e["name"], e["addr"], e["tel"]
    # 部分比對（最長鍵匹配）
    best = ""
    for k, v in lookup.items():
        if not k.isdigit() and k in name and len(k) > len(best):
            best = k
    if best:
        e = lookup[best]
        return e["code"], e["name"], e["addr"], e["tel"]
    return "", name, "", ""


# ── Excel 讀取 ────────────────────────────────────────────────────────────────
def _read_master_lookup(wb) -> dict:
    """
    從會員總表讀取 {pid: {addr, tel}} 補充資料。
    地址：BF（index 56）；電話：手機 J（index 9）優先，電話 I（index 8）次之。
    """
    master_ws = next((wb[s] for s in wb.sheetnames if MASTER_SHEET_KW in s), None)
    if master_ws is None:
        return {}
    lookup: dict = {}
    for row in master_ws.iter_rows(min_row=3, values_only=True):  # 前兩列為標題
        pid = normalize_id(row[MASTER_IDX_PID] if len(row) > MASTER_IDX_PID else None)
        if not pid:
            continue
        addr_raw = str(row[MASTER_IDX_ADDR]).strip() if len(row) > MASTER_IDX_ADDR and row[MASTER_IDX_ADDR] else ""
        tel_raw = ""
        if len(row) > MASTER_IDX_MOB and row[MASTER_IDX_MOB]:
            tel_raw = str(row[MASTER_IDX_MOB]).strip()
        elif len(row) > MASTER_IDX_TEL and row[MASTER_IDX_TEL]:
            tel_raw = str(row[MASTER_IDX_TEL]).strip()
        if pid not in lookup:
            lookup[pid] = {
                "addr": addr_raw,
                "tel":  normalize_tel(tel_raw) if tel_raw else "",
            }
    return lookup


def _member_from_doctor_row(row: tuple, excel_row: int, pid: str, name: str, master: dict) -> dict:
    """將醫生看單列轉為 FM 輸出需要的會員資料。"""
    tel_raw = ""
    if len(row) > IDX_MOBILE and row[IDX_MOBILE]:
        tel_raw = str(row[IDX_MOBILE]).strip()
    elif len(row) > IDX_TEL and row[IDX_TEL]:
        tel_raw = str(row[IDX_TEL]).strip()
    member_tel = normalize_tel(tel_raw) if tel_raw else master.get(pid, {}).get("tel", "")
    member_addr = master.get(pid, {}).get("addr", "")
    raw_birthday = row[IDX_BIRTH] if len(row) > IDX_BIRTH else None
    birthday = to_yyyymmdd(raw_birthday)
    birthday_error = None
    if raw_birthday is not None and str(raw_birthday).strip() and not birthday:
        birthday_error = {
            "row": excel_row,
            "name": name,
            "value": raw_birthday,
            "reason": "生日格式或日期無效",
        }
    return {
        "pid": pid,
        "name": name,
        "birthday": birthday,
        "tel": member_tel,
        "addr": member_addr,
        "_birthday_error": birthday_error,
    }


def read_prioritized_members(input_path: Path) -> tuple:
    """
    讀取醫生看 sheet，套用 AF 不選 > AQ 指定 > AG 自選的優先順序。
    回傳 (self_members, designated_members, stats, skipped_id_errors)。
    """
    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    doctor_ws = next((wb[s] for s in wb.sheetnames if DOCTOR_KW in s), None)
    if doctor_ws is None:
        wb.close()
        raise ValueError(f"找不到含「{DOCTOR_KW}」的分頁：{input_path.name}")

    master = _read_master_lookup(wb)

    self_ids: set = set()
    designated_ids: set = set()
    excluded_ids: set = set()
    self_order: list[str] = []
    designated_order: list[str] = []
    self_rows: dict[str, dict] = {}
    designated_rows: dict[str, dict] = {}
    skipped_id_errors: list[dict] = []

    for excel_row, row in enumerate(
        doctor_ws.iter_rows(min_row=DATA_ROW, values_only=True),
        start=DATA_ROW,
    ):
        excluded_checked = is_checked(row[IDX_115X] if len(row) > IDX_115X else None)
        self_checked = is_checked(row[IDX_SELF] if len(row) > IDX_SELF else None)
        designated_checked = is_checked(row[IDX_DESIGNATED] if len(row) > IDX_DESIGNATED else None)
        if not excluded_checked and not self_checked and not designated_checked:
            continue

        raw_pid = row[IDX_PID] if len(row) > IDX_PID else None
        pid = normalize_id(raw_pid)
        name = str(row[IDX_NAME]).strip() if len(row) > IDX_NAME and row[IDX_NAME] else ""
        pid_error = id_format_error(raw_pid, allow_legacy_resident=True)
        if pid_error:
            if excluded_checked:
                continue
            skipped_id_errors.append({
                "row": excel_row, "name": name, "value": raw_pid, "reason": pid_error,
            })
            continue

        if excluded_checked:
            excluded_ids.add(pid)
            continue

        if self_checked:
            if pid not in self_ids:
                self_order.append(pid)
                self_rows[pid] = _member_from_doctor_row(row, excel_row, pid, name, master)
            self_ids.add(pid)

        if designated_checked:
            if pid not in designated_ids:
                designated_order.append(pid)
                designated_rows[pid] = _member_from_doctor_row(row, excel_row, pid, name, master)
            designated_ids.add(pid)

    wb.close()

    effective_designated = [
        pid for pid in designated_order
        if pid in designated_ids and pid not in excluded_ids
    ]
    effective_self = [
        pid for pid in self_order
        if pid in self_ids and pid not in excluded_ids and pid not in designated_ids
    ]

    birthday_errors: list[dict] = []

    def finalize(order: list[str], rows_by_pid: dict[str, dict]) -> list[dict]:
        members: list[dict] = []
        for pid in order:
            member = rows_by_pid[pid].copy()
            if member["_birthday_error"]:
                birthday_errors.append(member["_birthday_error"])
            member.pop("_birthday_error", None)
            members.append(member)
        return members

    designated_members = finalize(effective_designated, designated_rows)
    self_members = finalize(effective_self, self_rows)
    if birthday_errors:
        raise ValueError(format_member_errors(birthday_errors))

    stats = {
        "self_checked": len(self_ids),
        "designated_checked": len(designated_ids),
        "excluded_checked": len(excluded_ids),
        "self_removed_by_excluded": len(self_ids & excluded_ids),
        "designated_removed_by_excluded": len(designated_ids & excluded_ids),
        "self_removed_by_designated": len((self_ids & designated_ids) - excluded_ids),
    }
    return self_members, designated_members, stats, skipped_id_errors


# ── FM.txt 輸出 ───────────────────────────────────────────────────────────────
def predict_output_paths(dest_dir: Path, hosp_id: str,
                         prsn_ids: list, member_count: int) -> list:
    """預測將產生的所有 FM.txt 路徑（用於覆蓋衝突檢查）。"""
    month = taiwan_today().strftime("%m")
    valid_ids = [p.strip() for p in prsn_ids if p.strip()]
    paths = []
    sn = 1
    for _ in valid_ids:
        n_chunks = max(1, (member_count + MAX_PER_FILE - 1) // MAX_PER_FILE)
        for _ in range(n_chunks):
            paths.append(dest_dir / f"{BRANCH}{hosp_id}{month}{sn:02d}FM.txt")
            sn += 1
    return paths


def write_fm_files(
    members: list, dest_dir: Path, hosp_id: str,
    prsn_ids: list, clinic_addr: str = "", clinic_tel: str = "",
    case_type: str = CASE_SELF,
) -> list:
    """
    為每個有效的 prsn_id 產生一份 FM.txt（流水號連續）。
    地址/電話：會員自身優先，無資料才採用診所的。
    回傳所有輸出路徑。
    """
    today     = taiwan_today()
    case_date = today.strftime("%Y%m%d")
    month     = today.strftime("%m")
    dest_dir.mkdir(parents=True, exist_ok=True)

    valid_ids = [p.strip() for p in prsn_ids if p.strip()]
    all_out: list = []
    sn = 1

    for prsn_id in valid_ids:
        chunks = [members[i:i + MAX_PER_FILE] for i in range(0, len(members), MAX_PER_FILE)]
        for chunk in chunks:
            fname    = f"{BRANCH}{hosp_id}{month}{sn:02d}FM.txt"
            out_path = dest_dir / fname
            with open(out_path, "wb") as fh:
                for m in chunk:
                    rec = make_record(
                        hosp_id   = hosp_id,
                        pid       = m["pid"],
                        birthday  = m["birthday"] or "        ",
                        name      = m["name"],
                        sex       = sex_from_id(m["pid"]),
                        addr      = m.get("addr") or clinic_addr,
                        tel       = m.get("tel") or clinic_tel,
                        prsn_id   = prsn_id,
                        case_type = case_type,
                        case_date = case_date,
                    )
                    fh.write(rec + b"\r\n")
            print(f"  ✓ {fname}  醫事人員：{prsn_id}  筆數：{len(chunk)}")
            all_out.append(out_path)
            sn += 1

    return all_out


# ── GUI ───────────────────────────────────────────────────────────────────────
class UploadApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("115年健保署會員名單上傳工具")
        self.root.resizable(False, False)

        self._clinic_lookup = load_clinic_lookup()

        self.src_var   = tk.StringVar()
        self.code_var  = tk.StringVar()
        self.name_var  = tk.StringVar()
        self.addr_var  = tk.StringVar()
        self.tel_var   = tk.StringVar()
        self.prsn_vars = [tk.StringVar() for _ in range(5)]

        self._build()

    def _build(self):
        root = self.root
        PAD = 10

        # ── 來源檔案 ─────────────────────────────────────────────────────────
        f_src = ttk.LabelFrame(root, text="會員來源檔案", padding=PAD)
        f_src.grid(row=0, column=0, sticky="ew", padx=PAD, pady=(PAD, 4))

        ttk.Entry(f_src, textvariable=self.src_var, width=52, state="readonly").grid(
            row=0, column=0, padx=(0, 6))
        ttk.Button(f_src, text="選擇…", command=self._pick_file).grid(row=0, column=1)

        # ── 診所資訊（自動帶入） ──────────────────────────────────────────────
        f_clinic = ttk.LabelFrame(root, text="診所資訊（自動帶入）", padding=PAD)
        f_clinic.grid(row=1, column=0, sticky="ew", padx=PAD, pady=4)

        ttk.Label(f_clinic, text="醫事機構代碼").grid(row=0, column=0, sticky="w")
        code_entry = ttk.Entry(f_clinic, textvariable=self.code_var, width=14)
        code_entry.grid(row=0, column=1, sticky="w", padx=(4, 16))

        ttk.Label(f_clinic, text="院所名稱").grid(row=0, column=2, sticky="w")
        ttk.Entry(f_clinic, textvariable=self.name_var, width=30).grid(
            row=0, column=3, sticky="w", padx=4)

        ttk.Label(f_clinic, text="機構地址").grid(row=1, column=0, sticky="w", pady=(4, 0))
        ttk.Entry(f_clinic, textvariable=self.addr_var, width=50).grid(
            row=1, column=1, columnspan=3, sticky="w", padx=4, pady=(4, 0))

        ttk.Label(f_clinic, text="診所電話").grid(row=2, column=0, sticky="w", pady=(4, 0))
        ttk.Entry(f_clinic, textvariable=self.tel_var, width=20).grid(
            row=2, column=1, sticky="w", padx=4, pady=(4, 0))

        # ── 醫事人員身分證號 ──────────────────────────────────────────────────
        f_prsn = ttk.LabelFrame(root, text="醫事人員身分證號（可貼上）", padding=PAD)
        f_prsn.grid(row=2, column=0, sticky="ew", padx=PAD, pady=4)

        for i, var in enumerate(self.prsn_vars, start=1):
            ttk.Label(f_prsn, text=f"醫事人員 {i}").grid(
                row=i - 1, column=0, sticky="w", pady=2)
            ttk.Entry(f_prsn, textvariable=var, width=16).grid(
                row=i - 1, column=1, sticky="w", padx=(6, 0), pady=2)
            if i == 1:
                ttk.Label(f_prsn, text="（至少填 1 位）",
                          foreground="gray").grid(row=i - 1, column=2, sticky="w", padx=6)

        # ── 輸出說明 ──────────────────────────────────────────────────────────
        f_out = ttk.LabelFrame(root, text="輸出位置", padding=PAD)
        f_out.grid(row=3, column=0, sticky="ew", padx=PAD, pady=4)
        out_hint = str(OUTPUT_ROOT / "{醫事機構代碼}{院所名稱}" / "{自選會員 | 指定會員}")
        ttk.Label(f_out, text=out_hint, foreground="#555").pack(anchor="w")

        # ── 執行按鈕 ──────────────────────────────────────────────────────────
        f_btn = ttk.Frame(root)
        f_btn.grid(row=4, column=0, pady=(6, PAD))
        ttk.Button(f_btn, text="輸出自選會員 B",
                   command=self._run_self,       width=20).grid(row=0, column=0, padx=8)
        ttk.Button(f_btn, text="輸出指定會員 A",
                   command=self._run_designated, width=20).grid(row=0, column=1, padx=8)

        root.columnconfigure(0, weight=1)

    def _pick_file(self):
        path = filedialog.askopenfilename(
            title="選擇選會員 Excel 檔案",
            filetypes=[("Excel 檔案", "*.xlsx *.xlsm"), ("所有檔案", "*.*")],
        )
        if not path:
            return
        self.src_var.set(path)
        self._auto_fill_clinic(Path(path))

    def _auto_fill_clinic(self, path: Path):
        code, name, addr, tel = find_clinic_info(path.stem, self._clinic_lookup)
        self.code_var.set(code)
        self.name_var.set(name)
        self.addr_var.set(addr)
        self.tel_var.set(tel)

    def _run_self(self):
        self._run_core("self")

    def _run_designated(self):
        self._run_core("designated")

    def _run_core(self, mode: str):
        """mode='self' → 自選會員B；mode='designated' → 指定會員A"""
        src_str = self.src_var.get().strip()
        if not src_str:
            messagebox.showwarning("提示", "請先選擇會員來源檔案。")
            return

        hosp_id   = self.code_var.get().strip()
        hosp_name = self.name_var.get().strip()
        if not hosp_id:
            messagebox.showwarning("提示", "醫事機構代碼不可為空，請手動填入。")
            return
        if not re.fullmatch(r"\d{10}", hosp_id):
            messagebox.showwarning("提示", "醫事機構代碼必須是 10 碼數字。")
            return

        prsn_ids = [normalize_id(v.get()) for v in self.prsn_vars]
        if not any(p.strip() for p in prsn_ids):
            messagebox.showwarning("提示", "請至少填入一位醫事人員身分證號。")
            return
        invalid_prsn_ids = [p for p in prsn_ids if p and id_format_error(p)]
        if invalid_prsn_ids:
            messagebox.showwarning(
                "提示",
                "下列醫事人員身分證號格式不符：\n" + "\n".join(invalid_prsn_ids),
            )
            return

        addr        = self.addr_var.get().strip()
        clinic_tel  = self.tel_var.get().strip()
        input_path  = Path(src_str)

        # ── 先讀 Excel，GUI 維持可見 ──────────────────────────────────────────
        try:
            self_members, designated_members, priority_stats, skipped_id_errors = read_prioritized_members(input_path)
        except (ValueError, OSError) as exc:
            messagebox.showerror("讀取失敗", str(exc))
            return
        except Exception as exc:
            import traceback; traceback.print_exc()
            messagebox.showerror("未預期錯誤", str(exc))
            return

        if mode == "self":
            subfolder  = SELF_SUBFOLDER
            case_type  = CASE_SELF
            label_type = "自選會員"
            members = self_members
            garbled_name_errors: list[dict] = []
        else:
            subfolder  = DESIGNATED_SUBFOLDER
            case_type  = CASE_DESIGNATED
            label_type = "指定會員"
            members, garbled_name_errors = filter_garbled_names(designated_members)

        if skipped_id_errors:
            warning = format_member_errors(
                skipped_id_errors,
                title=(
                    f"以下 {len(skipped_id_errors)} 筆身分證號不符合格式，"
                    "將跳過並繼續產生名單："
                ),
            )
            print(f"[{label_type}] {warning}")
            messagebox.showwarning("身分證號格式警示", warning)

        if not members:
            if mode == "self":
                msg = "自選會員（AG打勾，扣除不選與指定優先）為 0 筆，未產生輸出。"
            else:
                msg = "指定會員（AQ打勾且非不選，並排除亂碼姓名）為 0 筆，未產生輸出。"
                if garbled_name_errors:
                    msg += "\n\n" + "\n".join(format_garbled_summary(garbled_name_errors))
            print(f"[{label_type}] {msg}")
            messagebox.showwarning("無資料", msg)
            return

        # ── 預先確認覆蓋（GUI 仍可見，讓對話框有父視窗）────────────────────
        folder_name  = f"{hosp_id}{hosp_name}"
        dest_dir     = OUTPUT_ROOT / folder_name / subfolder
        would_create = predict_output_paths(dest_dir, hosp_id, prsn_ids, len(members))
        conflicts    = [p for p in would_create if p.exists()]
        if conflicts:
            names = "\n".join(f"  {p.name}" for p in conflicts)
            if not messagebox.askyesno(
                "檔案已存在",
                f"以下 {len(conflicts)} 個檔案已存在，確定要覆蓋嗎？\n\n{names}",
                default="no",
            ):
                return

        # ── 確認後隱藏 GUI，輸出至終端機 ──────────────────────────────────────
        self.root.withdraw()

        print("=" * 60)
        print(f"115年健保署會員名單上傳工具 ── {label_type}")
        print("=" * 60)
        print(f"來源檔案：{input_path.name}")
        print(f"醫事機構代碼：{hosp_id}　院所名稱：{hosp_name}")
        print(f"醫事人員：{[p.strip() for p in prsn_ids if p.strip()]}")
        print()
        if mode == "self":
            print(f"  AG 自選勾選：{priority_stats['self_checked']} 筆")
            print(f"  AF 不選汰除自選：{priority_stats['self_removed_by_excluded']} 筆")
            print(f"  AQ 指定優先排除自選：{priority_stats['self_removed_by_designated']} 筆")
        else:
            print(f"  AQ 指定勾選：{priority_stats['designated_checked']} 筆")
            print(f"  AF 不選汰除指定：{priority_stats['designated_removed_by_excluded']} 筆")
            print(f"  亂碼姓名排除：{len(garbled_name_errors)} 筆")
            for item in garbled_name_errors:
                print(f"    {item['pid']}｜{item['name']}｜{item['reason']}")
        print(f"  AF 不選總數：{priority_stats['excluded_checked']} 筆")
        print(f"  身分證格式不符跳過：{len(skipped_id_errors)} 筆")
        print(f"  實際輸出：{len(members)} 筆")
        print()
        print(f"輸出資料夾：{dest_dir}")
        print()

        try:
            out_paths = write_fm_files(
                members     = members,
                dest_dir    = dest_dir,
                hosp_id     = hosp_id,
                prsn_ids    = prsn_ids,
                clinic_addr = addr,
                clinic_tel  = clinic_tel,
                case_type   = case_type,
            )

            valid_prsn = [p.strip() for p in prsn_ids if p.strip()]
            print()
            print("=" * 60)
            print("輸出完成摘要")
            print("=" * 60)
            print(f"院所：{hosp_id} {hosp_name}")
            print(f"{label_type}筆數：{len(members)}")
            if mode == "self":
                print(f"AF 不選汰除自選：{priority_stats['self_removed_by_excluded']} 筆")
                print(f"AQ 指定優先排除自選：{priority_stats['self_removed_by_designated']} 筆")
            else:
                print(f"AF 不選汰除指定：{priority_stats['designated_removed_by_excluded']} 筆")
                print(f"亂碼姓名排除：{len(garbled_name_errors)} 筆")
            print(f"身分證格式不符跳過：{len(skipped_id_errors)} 筆")
            print(f"醫事人員數：{len(valid_prsn)}")
            print(f"輸出檔案數：{len(out_paths)}")
            print(f"輸出資料夾：{dest_dir}")
            print("輸出檔案：")
            for p in out_paths:
                print(f"  {p.name}")
            print("=" * 60)

            summary_lines = [
                f"院所：{hosp_id} {hosp_name}",
                f"{label_type}：{len(members)} 筆",
            ]
            if mode == "self":
                summary_lines.append(f"AF 不選汰除自選：{priority_stats['self_removed_by_excluded']} 筆")
                summary_lines.append(f"AQ 指定優先排除自選：{priority_stats['self_removed_by_designated']} 筆")
            else:
                summary_lines.append(f"AF 不選汰除指定：{priority_stats['designated_removed_by_excluded']} 筆")
                summary_lines.extend(format_garbled_summary(garbled_name_errors))
            summary_lines.append(f"身分證格式不符跳過：{len(skipped_id_errors)} 筆")
            summary_lines += [
                f"輸出 {len(out_paths)} 份 FM.txt",
                "",
                f"輸出資料夾：\n{dest_dir}",
            ]
            messagebox.showinfo("完成", "\n".join(summary_lines))

            import subprocess
            if sys.platform == "win32":
                subprocess.Popen(["explorer", str(dest_dir)])
            else:
                subprocess.Popen(["open", str(dest_dir)])

        except (ValueError, OSError, AssertionError) as exc:
            print(f"錯誤：{exc}")
            messagebox.showerror("錯誤", str(exc))
            self.root.deiconify()
            return
        except Exception as exc:
            import traceback
            traceback.print_exc()
            messagebox.showerror("未預期錯誤", str(exc))
            self.root.deiconify()
            return

        self.root.destroy()

    def run(self):
        self.root.mainloop()


# ── 進入點 ────────────────────────────────────────────────────────────────────
def main():
    app = UploadApp()
    app.run()


if __name__ == "__main__":
    main()
