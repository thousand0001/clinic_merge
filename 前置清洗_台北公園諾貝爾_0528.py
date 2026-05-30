# -*- coding: utf-8 -*-
"""
台北公園諾貝爾展望雲端前置清洗 + 通用主程式包裝

用途：
- 先用展望讀檔邏輯收斂原始資料。
- 轉成通用版可辨識的標準來源資料夾。
- 再呼叫 run_merge_通用，避免輸出後硬修格式。

重要規則：
- 百分位名單仍以通用版規則為準：需有疾病樣態才納入糖心腎 KPI 候選。
- 不修改、不刪除原始資料。
"""

from __future__ import annotations

import importlib.util
import datetime as dt
import os
import re
import shutil
import sys
import tempfile
from collections import defaultdict
from copy import copy
from dataclasses import dataclass, field
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple

import openpyxl
from openpyxl import Workbook


SCRIPT_DIR = Path(__file__).resolve().parent


def _find_generic_script(script_dir: Path) -> Path:
    candidates = sorted(script_dir.glob("run_merge_通用_*.py"), reverse=True)
    if not candidates:
        raise RuntimeError("找不到通用程式 run_merge_通用_*.py")
    return candidates[0]


GENERIC_SCRIPT = _find_generic_script(SCRIPT_DIR)


def _load_module(path: Path, module_name: str):
    spec = importlib.util.spec_from_file_location(module_name, path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"無法載入程式：{path}")
    mod = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = mod
    spec.loader.exec_module(mod)
    return mod


GENERIC = _load_module(GENERIC_SCRIPT, "run_merge_generic")


class TableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.rows: List[List[str]] = []
        self._row: Optional[List[str]] = None
        self._cell: Optional[List[str]] = None
        self._in_cell = False

    def handle_starttag(self, tag: str, attrs: Sequence[Tuple[str, Optional[str]]]) -> None:
        tag = tag.lower()
        if tag == "tr":
            self._row = []
        elif tag in ("td", "th") and self._row is not None:
            self._cell = []
            self._in_cell = True

    def handle_data(self, data: str) -> None:
        if self._in_cell and self._cell is not None:
            self._cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag in ("td", "th") and self._in_cell:
            text = " ".join("".join(self._cell or []).split())
            if self._row is not None:
                self._row.append(text)
            self._cell = None
            self._in_cell = False
        elif tag == "tr" and self._row is not None:
            if any(str(value).strip() for value in self._row):
                self.rows.append(self._row)
            self._row = None


@dataclass
class MemberRecord:
    name: str = ""
    pid: str = ""
    bday: Optional[dt.date] = None
    charts: Set[str] = field(default_factory=set)
    sex: str = ""
    clinic: str = ""
    ascvd: Any = None
    dmk_raw: Any = None
    health: Dict[str, Any] = field(default_factory=dict)
    p4p: Dict[str, Any] = field(default_factory=dict)
    screenings: Dict[str, Any] = field(default_factory=dict)
    count_114: float = 0.0
    count_114_q1: float = 0.0
    count_115: float = 0.0
    count_115_q1: float = 0.0
    amount_114: float = 0.0
    amount_114_q1: float = 0.0
    amount_115: float = 0.0
    amount_115_q1: float = 0.0
    last_visit: Optional[dt.date] = None
    notes: Set[str] = field(default_factory=set)
    sources: Set[str] = field(default_factory=set)
    is_self_select: bool = False


def normalize_name(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).upper()


def display_text(value: Any) -> str:
    return str(value or "").strip()


def normalize_id(value: Any) -> str:
    return str(value or "").strip().upper().replace(" ", "")


def is_valid_id(value: Any) -> bool:
    text = normalize_id(value)
    return bool(
        re.fullmatch(r"[A-Z][12]\d{8}", text)
        or re.fullmatch(r"[A-Z][A-Z0-9]\d{8}", text)
        or re.fullmatch(r"[A-Z]{1,2}\d{8,10}", text)
    )


def normalize_chart(value: Any) -> str:
    text = str(value or "").strip()
    return text.zfill(7) if text.isdigit() else text


def parse_number(value: Any) -> float:
    try:
        return float(str(value or "0").replace(",", "").strip() or 0)
    except Exception:
        return 0.0


def parse_compact_roc_or_ad_date(value: Any) -> Optional[dt.date]:
    text = str(value or "").strip()
    if not re.fullmatch(r"\d{6,8}", text):
        return None
    if len(text) == 7:
        year = int(text[:3]) + 1911
        month = int(text[3:5])
        day = int(text[5:7])
    elif len(text) == 8:
        year = int(text[:4])
        month = int(text[4:6])
        day = int(text[6:8])
        if year < 1911:
            year += 1911
    else:
        year = int(text[:2]) + 1911
        month = int(text[2:4])
        day = int(text[4:6])
    try:
        return dt.date(year, month, day)
    except ValueError:
        return None


def parse_roc_or_ad_date(value: Any) -> Optional[dt.date]:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = str(value).strip()
    if not text or text == "-":
        return None
    compact = re.sub(r"\D", "", text)
    if re.fullmatch(r"\d{6,8}", compact):
        parsed = parse_compact_roc_or_ad_date(compact)
        if parsed:
            return parsed
    text = text.split(" ")[0].replace("-", "/")
    match = re.fullmatch(r"(\d{2,4})/(\d{1,2})/(\d{1,2})", text)
    if not match:
        return None
    year = int(match.group(1))
    if year < 1911:
        year += 1911
    try:
        return dt.date(year, int(match.group(2)), int(match.group(3)))
    except ValueError:
        return None


def parse_roc_yyyymmdd(value: Any) -> Optional[dt.date]:
    return parse_compact_roc_or_ad_date(value)


def read_html_table(path: Path) -> List[List[str]]:
    parser = TableParser()
    parser.feed(path.read_text(encoding="utf-8-sig", errors="replace"))
    return parser.rows


def iter_workbook_rows(path: Path) -> Iterable[Tuple[str, List[List[Any]]]]:
    if path.suffix.lower() == ".ods":
        wb = GENERIC._load_ods_as_workbook(str(path))
    else:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    for ws in wb.worksheets:
        rows: List[List[Any]] = []
        for row in ws.iter_rows(values_only=True):
            values = list(row)
            if any(value is not None and str(value).strip() != "" for value in values):
                rows.append(values)
        yield ws.title, rows


def compact_header(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).upper()


def find_col(header: Sequence[Any], aliases: Sequence[str]) -> Optional[int]:
    targets = {compact_header(alias) for alias in aliases}
    for index, value in enumerate(header):
        if compact_header(value) in targets:
            return index
    return None


def find_header(rows: Sequence[Sequence[Any]], required_groups: Sequence[Sequence[str]], scan_rows: int = 30) -> Optional[int]:
    for index, row in enumerate(rows[:scan_rows]):
        if all(find_col(row, group) is not None for group in required_groups):
            return index
    return None


def infer_gender(pid: str) -> str:
    text = normalize_id(pid)
    if len(text) >= 2:
        if text[1] == "1":
            return "男"
        if text[1] == "2":
            return "女"
    return ""


def get_or_create_by_id(records: Dict[str, MemberRecord], pid: str) -> Optional[MemberRecord]:
    pid = normalize_id(pid)
    if not is_valid_id(pid):
        return None
    key = f"ID:{pid}"
    rec = records.setdefault(key, MemberRecord(pid=pid))
    rec.pid = pid
    return rec


def get_or_create_by_name(records: Dict[str, MemberRecord], name: str) -> Optional[MemberRecord]:
    norm = normalize_name(name)
    if not norm:
        return None
    key = f"NAME:{norm}"
    rec = records.setdefault(key, MemberRecord(name=display_text(name)))
    if not rec.name:
        rec.name = display_text(name)
    return rec


def add_identity(
    records: Dict[str, MemberRecord],
    name_to_ids: Dict[str, Set[str]],
    chart_to_ids: Dict[str, Set[str]],
    *,
    name: Any = "",
    pid: Any = "",
    chart: Any = "",
    bday: Any = "",
    source: str = "",
) -> Optional[MemberRecord]:
    text_pid = normalize_id(pid)
    rec = get_or_create_by_id(records, text_pid) if is_valid_id(text_pid) else get_or_create_by_name(records, display_text(name))
    if rec is None:
        return None
    if display_text(name) and not rec.name:
        rec.name = display_text(name)
    if is_valid_id(text_pid):
        rec.pid = text_pid
        if normalize_name(name):
            name_to_ids[normalize_name(name)].add(text_pid)
    parsed_bday = parse_roc_or_ad_date(bday)
    if parsed_bday and not rec.bday:
        rec.bday = parsed_bday
    chart_text = normalize_chart(chart)
    if chart_text:
        rec.charts.add(chart_text)
        if is_valid_id(text_pid):
            chart_to_ids[chart_text].add(text_pid)
    if source:
        rec.sources.add(source)
    return rec


def merge_name_records(records: Dict[str, MemberRecord], name_to_ids: Dict[str, Set[str]]) -> None:
    for key in [key for key in records if key.startswith("NAME:")]:
        rec = records[key]
        ids = name_to_ids.get(normalize_name(rec.name), set())
        if len(ids) != 1:
            continue
        target = get_or_create_by_id(records, next(iter(ids)))
        if target is None or target is rec:
            continue
        if not target.name:
            target.name = rec.name
        if not target.bday:
            target.bday = rec.bday
        target.charts.update(rec.charts)
        target.screenings.update({k: v for k, v in rec.screenings.items() if k not in target.screenings})
        target.health.update({k: v for k, v in rec.health.items() if k not in target.health})
        target.p4p.update({k: v for k, v in rec.p4p.items() if k not in target.p4p})
        target.count_114 += rec.count_114
        target.count_114_q1 += rec.count_114_q1
        target.count_115 += rec.count_115
        target.count_115_q1 += rec.count_115_q1
        target.amount_114 += rec.amount_114
        target.amount_114_q1 += rec.amount_114_q1
        target.amount_115 += rec.amount_115
        target.amount_115_q1 += rec.amount_115_q1
        if rec.last_visit and (target.last_visit is None or rec.last_visit > target.last_visit):
            target.last_visit = rec.last_visit
        target.notes.update(rec.notes)
        target.sources.update(rec.sources)
        del records[key]


def is_self_select_source(text: Any) -> bool:
    compact = re.sub(r"[\s\-_()（）\[\]{}]+", "", str(text or "").strip()).lower()
    if not compact:
        return False
    if "115x" in compact or "不選" in compact or "不要" in compact:
        return False
    return any(token in compact for token in ("自選名單", "自選會員", "115自選", "a115"))


def is_excluded_source(text: Any) -> bool:
    compact = re.sub(r"[\s\-_()（）\[\]{}]+", "", str(text or "").strip()).lower()
    return bool(compact and ("115x" in compact or "不選" in compact or "不要" in compact))


def scan_root_files(
    source_dir: Path,
    records: Dict[str, MemberRecord],
    name_to_ids: Dict[str, Set[str]],
    chart_to_ids: Dict[str, Set[str]],
) -> None:
    for path in sorted(source_dir.iterdir()):
        if not path.is_file() or path.name.startswith("~$") or path.suffix.lower() not in (".xlsx", ".ods"):
            continue
        if is_excluded_source(path.stem):
            continue
        source_is_self_select = is_self_select_source(path.stem)
        for sheet_name, rows in iter_workbook_rows(path):
            if is_excluded_source(sheet_name) or not rows:
                continue
            sheet_is_self_select = source_is_self_select or is_self_select_source(sheet_name)
            header_index = find_header(rows, [["姓名", "會員姓名", "病患姓名"], ["ID", "身分證", "身分證號", "身分證號碼", "身份證號", "身份證號碼", "家醫收案會員ID"]])
            if header_index is None:
                continue
            header = rows[header_index]
            name_col = find_col(header, ["姓名", "會員姓名", "病患姓名"])
            id_col = find_col(header, ["ID", "身分證", "身分證號", "身分證號碼", "身份證", "身份證號", "身份證號碼", "家醫收案會員ID"])
            bday_col = find_col(header, ["生日", "BIRTHDAY"])
            indicator_col = find_col(header, ["指標名稱"])
            ascvd_col = find_col(header, ["ASCVD", "ascvd"])
            disease_col = find_col(header, ["疾病樣態"])
            p4p_plan_col = find_col(header, ["P4P收案計畫"])
            p4p_status_col = find_col(header, ["收案狀態"])
            hba_val_col = find_col(header, ["最近一次HbA1c檢查結果(%)"])
            hba_dt_col = find_col(header, ["最近一次HbA1c檢查日期"])
            ldl_val_col = find_col(header, ["最近一次LDL檢查結果(mg/dL)"])
            ldl_dt_col = find_col(header, ["最近一次LDL檢查日期"])
            uacr_val_col = find_col(header, ["最近一次UACR檢查結果(mg/gm)", "最近一次UACR檢查結果"])
            uacr_dt_col = find_col(header, ["最近一次UACR檢查日期"])
            last_screen_col = find_col(header, ["最後篩檢日期"])
            if name_col is None or id_col is None:
                continue
            for row in rows[header_index + 1:]:
                if len(row) <= max(name_col, id_col):
                    continue
                name = row[name_col]
                pid = row[id_col]
                if not display_text(name) and not is_valid_id(pid):
                    continue
                bday = row[bday_col] if bday_col is not None and len(row) > bday_col else None
                rec = add_identity(records, name_to_ids, chart_to_ids, name=name, pid=pid, bday=bday, source=path.name)
                if rec is None:
                    continue
                if sheet_is_self_select:
                    rec.is_self_select = True
                if ascvd_col is not None and len(row) > ascvd_col and row[ascvd_col] not in (None, ""):
                    rec.ascvd = row[ascvd_col]
                if disease_col is not None and len(row) > disease_col and row[disease_col] not in (None, ""):
                    rec.dmk_raw = row[disease_col]
                if p4p_plan_col is not None and len(row) > p4p_plan_col:
                    rec.p4p["plan"] = row[p4p_plan_col]
                if p4p_status_col is not None and len(row) > p4p_status_col:
                    rec.p4p["status"] = row[p4p_status_col]
                if hba_val_col is not None and len(row) > hba_val_col:
                    rec.health["hba_val"] = row[hba_val_col]
                if hba_dt_col is not None and len(row) > hba_dt_col:
                    rec.health["hba_dt"] = parse_roc_or_ad_date(row[hba_dt_col])
                if ldl_val_col is not None and len(row) > ldl_val_col:
                    rec.health["ldl_val"] = row[ldl_val_col]
                if ldl_dt_col is not None and len(row) > ldl_dt_col:
                    rec.health["ldl_dt"] = parse_roc_or_ad_date(row[ldl_dt_col])
                if uacr_val_col is not None and len(row) > uacr_val_col:
                    rec.health["uacr_val"] = row[uacr_val_col]
                if uacr_dt_col is not None and len(row) > uacr_dt_col:
                    rec.health["uacr_dt"] = parse_roc_or_ad_date(row[uacr_dt_col])
                if indicator_col is not None and len(row) > indicator_col:
                    indicator = str(row[indicator_col] or "")
                    last_dt = row[last_screen_col] if last_screen_col is not None and len(row) > last_screen_col else None
                    if "成人預防保健" in indicator or "成人健檢" in indicator:
                        rec.screenings["成人健檢"] = parse_roc_or_ad_date(last_dt) or "-"
                    elif "子宮頸抹片" in indicator or "子宮抹片" in indicator:
                        rec.screenings["子宮抹片"] = parse_roc_or_ad_date(last_dt) or "-"
                    elif "流感" in indicator:
                        rec.screenings["老人流感"] = parse_roc_or_ad_date(last_dt) or "-"
                    elif "糞便潛血" in indicator or "潛血" in indicator:
                        rec.screenings["糞便潛血"] = parse_roc_or_ad_date(last_dt) or "-"
                    elif "肝炎" in indicator or "B、C肝" in indicator or "BC肝" in indicator.upper():
                        rec.screenings["BC肝炎"] = parse_roc_or_ad_date(last_dt) or "-"


def scan_r11440_files(
    source_dir: Path,
    records: Dict[str, MemberRecord],
    name_to_ids: Dict[str, Set[str]],
    chart_to_ids: Dict[str, Set[str]],
) -> None:
    report_dir = source_dir / "R11440"
    if not report_dir.is_dir():
        return
    for path in sorted(report_dir.iterdir()):
        if not path.is_file() or path.name.startswith("~$") or path.suffix.lower() != ".xlsx":
            continue
        for _sheet_name, rows in iter_workbook_rows(path):
            header_index = find_header(rows, [["掛號證", "掛號証"], ["日期"], ["姓名"], ["身分證號", "身份證號"], ["次數"], ["總額"]])
            if header_index is None:
                continue
            header = rows[header_index]
            chart_col = find_col(header, ["掛號證", "掛號証", "病歷號", "病歷號碼"])
            date_col = find_col(header, ["日期", "看診日期"])
            name_col = find_col(header, ["姓名", "病患姓名"])
            id_col = find_col(header, ["身分證號", "身份證號", "身分證", "身份證", "ID"])
            bday_col = find_col(header, ["生日", "出生日期", "出生年月日"])
            count_col = find_col(header, ["次數", "看診次數", "就診次數"])
            amount_col = find_col(header, ["總額", "申報總額", "實際申報總額"])
            if None in (chart_col, date_col, name_col, id_col, count_col, amount_col):
                continue
            for row in rows[header_index + 1:]:
                if len(row) <= max(chart_col, date_col, name_col, id_col, count_col, amount_col):
                    continue
                visit_date = parse_roc_yyyymmdd(row[date_col])
                if visit_date is None:
                    continue
                rec = add_identity(records, name_to_ids, chart_to_ids, chart=row[chart_col], name=row[name_col], pid=row[id_col], bday=row[bday_col] if bday_col is not None and len(row) > bday_col else None, source=f"R11440/{path.name}")
                if rec is None:
                    continue
                count = parse_number(row[count_col])
                amount = parse_number(row[amount_col])
                year_bucket = visit_date.year - 1911
                month = visit_date.month
                if year_bucket == 114:
                    rec.count_114 += count
                    rec.amount_114 += amount
                    if month <= 4:
                        rec.count_114_q1 += count
                        rec.amount_114_q1 += amount
                elif year_bucket == 115:
                    rec.count_115 += count
                    rec.amount_115 += amount
                    if month <= 4:
                        rec.count_115_q1 += count
                        rec.amount_115_q1 += amount
                if rec.last_visit is None or visit_date > rec.last_visit:
                    rec.last_visit = visit_date
                rec.notes.add("R11440月報表")


def scan_count_files(
    source_dir: Path,
    records: Dict[str, MemberRecord],
    name_to_ids: Dict[str, Set[str]],
    chart_to_ids: Dict[str, Set[str]],
) -> None:
    count_dir = source_dir / "次數"
    if not count_dir.is_dir():
        return
    for path in sorted(count_dir.glob("*.xls")):
        month_match = re.fullmatch(r"1(14|15)(\d{2})", path.stem)
        if not month_match:
            continue
        year_bucket = int(path.stem[:3])
        month = int(path.stem[3:5])
        rows = read_html_table(path)
        header_index = find_header(rows, [["病歷號"], ["身分證", "身分證號", "身份證號"], ["看診次數", "次數"]])
        if header_index is None:
            continue
        header = rows[header_index]
        chart_col = find_col(header, ["病歷號", "病歷號碼"])
        name_col = find_col(header, ["姓名", "病患姓名"])
        id_col = find_col(header, ["身分證", "身分證號", "身份證號", "ID"])
        bday_col = find_col(header, ["生日"])
        count_col = find_col(header, ["看診次數", "次數", "就診次數"])
        if None in (chart_col, name_col, id_col, count_col):
            continue
        for row in rows[header_index + 1:]:
            if len(row) <= max(chart_col, name_col, id_col, count_col):
                continue
            rec = add_identity(records, name_to_ids, chart_to_ids, chart=row[chart_col], name=row[name_col], pid=row[id_col], bday=row[bday_col] if bday_col is not None and len(row) > bday_col else None, source=f"次數/{path.name}")
            if rec is None:
                continue
            count = parse_number(row[count_col])
            if year_bucket == 114:
                rec.count_114 += count
                if month <= 4:
                    rec.count_114_q1 += count
            elif year_bucket == 115:
                rec.count_115 += count
                if month <= 4:
                    rec.count_115_q1 += count
            rec.notes.add("病歷號對到ID")


def collect_data(source_dir: Path) -> Dict[str, MemberRecord]:
    records: Dict[str, MemberRecord] = {}
    name_to_ids: Dict[str, Set[str]] = defaultdict(set)
    chart_to_ids: Dict[str, Set[str]] = defaultdict(set)
    scan_root_files(source_dir, records, name_to_ids, chart_to_ids)
    scan_r11440_files(source_dir, records, name_to_ids, chart_to_ids)
    scan_count_files(source_dir, records, name_to_ids, chart_to_ids)
    merge_name_records(records, name_to_ids)
    for rec in records.values():
        if not rec.pid:
            rec.notes.add("ID欄填-")
        if not rec.bday:
            rec.notes.add("缺生日")
        if rec.pid:
            rec.sex = infer_gender(rec.pid)
    return records


def _select_reader(source_dir: Path):
    return sys.modules[__name__]



def _normalize_disease(value: Any) -> Any:
    text = GENERIC.normalize_text(value).upper()
    if not text:
        return None
    if "DKD" in text or "糖尿病腎" in text:
        return 3
    if "CKD" in text or "腎臟病" in text:
        return 2
    if "DM" in text or "糖尿病" in text:
        return 1
    if "ASCVD" in text:
        return 4
    parsed = GENERIC.parse_disease_code(value)
    return parsed.value if parsed else None


def _date_value(value: Any) -> Any:
    parsed = GENERIC.parse_date(value)
    return parsed if parsed else value


def _number_or_blank(value: Any) -> Any:
    num = GENERIC.parse_float(value)
    if num is None or num == 0:
        return None
    return int(num) if float(num).is_integer() else num


def _is_x115_source(text: Any) -> bool:
    compact = re.sub(r"[\s\-_()（）\[\]{}]+", "", str(text or "").strip()).lower()
    return bool(compact and ("115x" in compact or "不選" in compact or "不要" in compact))


def _append_rows(ws, headers: Sequence[str], rows: Iterable[Sequence[Any]]) -> None:
    ws.append(list(headers))
    for row in rows:
        ws.append(list(row))


def _save_workbook(path: Path, sheets: Dict[str, tuple[Sequence[str], Iterable[Sequence[Any]]]]) -> None:
    wb = Workbook()
    first = True
    for title, (headers, rows) in sheets.items():
        ws = wb.active if first else wb.create_sheet()
        ws.title = title
        first = False
        _append_rows(ws, headers, rows)
    wb.save(path)


def _collect_x115_rows(source_dir: Path, reader) -> list[list[Any]]:
    rows_out: list[list[Any]] = []
    seen: set[str] = set()
    for path in sorted(source_dir.iterdir()):
        if not path.is_file() or path.name.startswith("~$") or path.suffix.lower() not in (".xlsx", ".ods"):
            continue
        if not _is_x115_source(path.stem):
            continue
        for sheet_name, rows in reader.iter_workbook_rows(path):
            if not _is_x115_source(sheet_name) and not _is_x115_source(path.stem):
                continue
            header_index = reader.find_header(
                rows,
                [["姓名", "會員姓名", "病患姓名"], ["ID", "身分證", "身分證號", "身分證號碼", "身份證號", "身份證號碼"]],
            )
            if header_index is None:
                continue
            header = rows[header_index]
            name_col = reader.find_col(header, ["姓名", "會員姓名", "病患姓名"])
            id_col = reader.find_col(header, ["ID", "身分證", "身分證號", "身分證號碼", "身份證", "身份證號", "身份證號碼"])
            bday_col = reader.find_col(header, ["生日", "BIRTHDAY", "出生日期", "出生年月日"])
            if name_col is None or id_col is None:
                continue
            for row in rows[header_index + 1:]:
                if len(row) <= max(name_col, id_col):
                    continue
                pid = GENERIC.normalize_id(row[id_col])
                name = GENERIC.normalize_text(row[name_col])
                if not pid or pid in seen:
                    continue
                bday = row[bday_col] if bday_col is not None and len(row) > bday_col else None
                rows_out.append([pid, name, _date_value(bday)])
                seen.add(pid)
    return rows_out


def _screening_rows(records, key: str):
    for rec in records:
        value = rec.screenings.get(key)
        parsed = GENERIC.parse_date(value)
        if rec.pid and parsed:
            yield [rec.pid, parsed]


def _month_code(path: Path) -> Optional[str]:
    match = re.fullmatch(r"1(?:14|15)\d{2}", path.stem)
    return match.group(0) if match else None


def _month_date(code: str) -> str:
    return f"{code[:3]}/{code[3:5]}/01"


def _record_lookup_maps(records, reader):
    chart_to_pid: Dict[str, set[str]] = {}
    name_to_pid: Dict[str, set[str]] = {}
    for rec in records:
        if not rec.pid:
            continue
        for chart in getattr(rec, "charts", set()):
            normalized_chart = reader.normalize_chart(chart)
            if normalized_chart:
                chart_to_pid.setdefault(normalized_chart, set()).add(rec.pid)
        name = reader.normalize_name(rec.name)
        if name:
            name_to_pid.setdefault(name, set()).add(rec.pid)
    return chart_to_pid, name_to_pid


def _resolve_pid(chart: Any, name: Any, chart_to_pid: Dict[str, set[str]], name_to_pid: Dict[str, set[str]], reader) -> str:
    chart_keys = [reader.normalize_chart(chart)]
    raw_chart = str(chart or "").strip()
    if raw_chart.isdigit():
        chart_keys.append(str(int(raw_chart)).zfill(7))
    for chart_key in dict.fromkeys(key for key in chart_keys if key):
        ids = chart_to_pid.get(chart_key, set())
        if len(ids) == 1:
            return next(iter(ids))
    name_key = reader.normalize_name(name)
    if name_key:
        ids = name_to_pid.get(name_key, set())
        if len(ids) == 1:
            return next(iter(ids))
    return ""


def _is_subtotal_row(value: Any) -> bool:
    text = GENERIC.normalize_text(value)
    return text in {"小計", "合計", "總計"}


def _monthly_member_key(name: Any, bday: Any, chart: Any, reader) -> Optional[Tuple[str, str]]:
    display_name = GENERIC.normalize_text(name)
    parsed_bday = reader.parse_roc_or_ad_date(bday)
    if not display_name or not parsed_bday:
        return None
    return (reader.normalize_name(display_name), parsed_bday.isoformat())


def _monthly_member_row_base(name: Any, bday: Any, chart: Any, reader) -> Optional[Dict[str, Any]]:
    key = _monthly_member_key(name, bday, chart, reader)
    if key is None:
        return None
    return {
        "chart": reader.normalize_chart(chart),
        "name": GENERIC.normalize_text(name),
        "bday": reader.parse_roc_or_ad_date(bday),
        "count": 0.0,
        "amount": 0.0,
        "date": None,
    }


def _keep_latest_date(current: Optional[dt.date], incoming: Any, fallback: str, reader) -> Optional[dt.date]:
    parsed = reader.parse_roc_or_ad_date(incoming) or reader.parse_roc_or_ad_date(fallback)
    if parsed is None:
        return current
    if current is None or parsed > current:
        return parsed
    return current


def _collect_monthly_claim_rows(source_dir: Path, records, reader) -> Dict[str, list[list[Any]]]:
    monthly: Dict[str, Dict[Tuple[str, str], Dict[str, Any]]] = {}

    count_dir = source_dir / "次數"
    if count_dir.is_dir():
        for path in sorted(count_dir.glob("*.xls")):
            code = _month_code(path)
            if not code:
                continue
            rows = reader.read_html_table(path)
            header_index = reader.find_header(
                rows,
                [["病歷號"], ["身分證", "身分證號", "身份證號"], ["看診次數", "次數"]],
            )
            if header_index is None:
                continue
            header = rows[header_index]
            chart_col = reader.find_col(header, ["病歷號", "病歷號碼"])
            name_col = reader.find_col(header, ["姓名", "病患姓名"])
            bday_col = reader.find_col(header, ["生日", "出生日期", "出生年月日"])
            count_col = reader.find_col(header, ["看診次數", "次數", "就診次數"])
            if None in (chart_col, name_col, bday_col, count_col):
                continue
            for row in rows[header_index + 1:]:
                if len(row) <= max(chart_col, name_col, bday_col, count_col):
                    continue
                if _is_subtotal_row(row[chart_col]) or _is_subtotal_row(row[name_col]):
                    continue
                count = reader.parse_number(row[count_col])
                if not count:
                    continue
                key = _monthly_member_key(row[name_col], row[bday_col], row[chart_col], reader)
                if key is None:
                    continue
                data = monthly.setdefault(code, {}).setdefault(
                    key,
                    _monthly_member_row_base(row[name_col], row[bday_col], row[chart_col], reader),
                )
                data["count"] += count
                data["date"] = _keep_latest_date(data.get("date"), None, _month_date(code), reader)

    fee_dir = source_dir / "費用"
    if fee_dir.is_dir():
        for path in sorted(fee_dir.glob("*.xls")):
            code = _month_code(path)
            if not code:
                continue
            rows = reader.read_html_table(path)
            header_index = reader.find_header(rows, [["病歷號"], ["姓名"], ["掛帳費", "費用小計", "總額", "申請金額"]])
            if header_index is None:
                continue
            header = rows[header_index]
            chart_col = reader.find_col(header, ["病歷號", "病歷號碼"])
            name_col = reader.find_col(header, ["姓名", "病患姓名"])
            bday_col = reader.find_col(header, ["生日", "出生日期", "出生年月日"])
            date_col = reader.find_col(header, ["日期", "看診日期", "就診日期"])
            amount_col = reader.find_col(header, ["掛帳費", "費用小計", "總額", "申請金額"])
            if None in (chart_col, name_col, bday_col, amount_col):
                continue
            for row in rows[header_index + 1:]:
                if len(row) <= max(chart_col, name_col, bday_col, amount_col):
                    continue
                if _is_subtotal_row(row[chart_col]) or _is_subtotal_row(row[name_col]):
                    continue
                amount = reader.parse_number(row[amount_col])
                if not amount:
                    continue
                key = _monthly_member_key(row[name_col], row[bday_col], row[chart_col], reader)
                if key is None:
                    continue
                data = monthly.setdefault(code, {}).setdefault(
                    key,
                    _monthly_member_row_base(row[name_col], row[bday_col], row[chart_col], reader),
                )
                data["amount"] += amount
                visit_date = row[date_col] if date_col is not None and len(row) > date_col else None
                data["date"] = _keep_latest_date(data.get("date"), visit_date, _month_date(code), reader)

    return {
        code: [
            [
                values["chart"],
                values["name"],
                values["bday"],
                values.get("date") or _date_value(_month_date(code)),
                _number_or_blank(values["count"]),
                _number_or_blank(values["amount"]),
            ]
            for _key, values in sorted(member_map.items())
            if values["count"] or values["amount"]
        ]
        for code, member_map in sorted(monthly.items())
    }


def _fallback_claim_rows(records, year: int, q1_only: bool = False):
    for rec in records:
        if not rec.pid:
            continue
        if year == 114:
            count = rec.count_114_q1 if q1_only else max(rec.count_114 - rec.count_114_q1, 0)
            amount = rec.amount_114_q1 if q1_only else max(rec.amount_114 - rec.amount_114_q1, 0)
        else:
            count = rec.count_115_q1 if q1_only else max(rec.count_115 - rec.count_115_q1, 0)
            amount = rec.amount_115_q1 if q1_only else max(rec.amount_115 - rec.amount_115_q1, 0)
        if count or amount:
            month = 1 if q1_only else 5
            yield [rec.pid, f"{year}/{month:02d}/01", _number_or_blank(count), _number_or_blank(amount)]


def _write_clean_source(source_dir: Path, clean_dir: Path) -> None:
    reader = _select_reader(source_dir)
    records = sorted(
        reader.collect_data(source_dir).values(),
        key=lambda rec: (0 if rec.pid else 1, reader.normalize_name(rec.name), rec.pid),
    )

    member_rows = []
    health_rows = []
    ascvd_rows = []
    self_select_rows = []
    x115_rows = []
    for rec in records:
        disease = _normalize_disease(rec.dmk_raw)
        ascvd = rec.ascvd if rec.ascvd not in (None, "") else None
        member_rows.append([
            rec.pid,
            rec.name,
            _date_value(rec.bday),
            rec.sex,
            disease,
            rec.last_visit,
            "A" if rec.pid else None,
        ])
        health_rows.append([
            rec.pid,
            _number_or_blank(rec.health.get("hba_val")),
            _date_value(rec.health.get("hba_dt")),
            _number_or_blank(rec.health.get("ldl_val")),
            _date_value(rec.health.get("ldl_dt")),
            _number_or_blank(rec.health.get("uacr_val")),
            _date_value(rec.health.get("uacr_dt")),
        ])
        if ascvd is not None:
            ascvd_rows.append([rec.pid, ascvd, rec.last_visit])
        if getattr(rec, "is_self_select", False):
            self_select_rows.append([rec.pid, rec.name, _date_value(rec.bday)])
        if getattr(rec, "is_115x", False):
            x115_rows.append([rec.pid, rec.name, _date_value(rec.bday)])
    for row in _collect_x115_rows(source_dir, reader):
        if row[0] not in {existing[0] for existing in x115_rows}:
            x115_rows.append(row)

    _save_workbook(
        clean_dir / "展望清洗_主檔.xlsx",
        {
            "會員名單": (
                ["ID", "姓名", "生日", "性別", "疾病樣態", "最後就診日", "會員別"],
                member_rows,
            ),
            "HealthCase": (
                [
                    "家醫收案會員ID",
                    "最近一次HbA1c檢查結果(%)",
                    "最近一次HbA1c檢查日期",
                    "最近一次LDL檢查結果(mg/dL)",
                    "最近一次LDL檢查日期",
                    "最近一次UACR檢查結果(mg/gm)",
                    "最近一次UACR檢查日期",
                ],
                health_rows,
            ),
            "ascvd": (["ID", "ASCVD", "最後就診日"], ascvd_rows),
            "自選會員": (["ID", "姓名", "生日"], self_select_rows),
            "115X": (["ID", "姓名", "生日"], x115_rows),
        },
    )

    screening_map = {
        "成人健檢": "成人健檢",
        "子宮抹片": "子宮抹片",
        "老人流感": "老人流感",
        "糞便潛血": "糞便潛血",
        "肝炎篩檢": "BC肝炎",
    }
    _save_workbook(
        clean_dir / "展望清洗_篩檢.xlsx",
        {
            title: (["ID", "最後篩檢日期"], list(_screening_rows(records, key)))
            for title, key in screening_map.items()
        },
    )

    monthly_claim_rows = _collect_monthly_claim_rows(source_dir, records, reader)
    if not monthly_claim_rows:
        monthly_claim_rows = {
            "11401": list(_fallback_claim_rows(records, 114, q1_only=True)),
            "11404": list(_fallback_claim_rows(records, 114, q1_only=False)),
            "11501": list(_fallback_claim_rows(records, 115, q1_only=True)),
            "11504": list(_fallback_claim_rows(records, 115, q1_only=False)),
        }
    _save_workbook(
        clean_dir / "展望清洗_申報統計.xlsx",
        {
            code: (["病歷號", "姓名", "生日", "日期", "次數", "申請金額"], rows)
            for code, rows in monthly_claim_rows.items()
        },
    )


def _postprocess_key(name: Any, bday: Any) -> Optional[Tuple[str, str]]:
    parsed = GENERIC.parse_date(bday)
    text = GENERIC.normalize_text(name)
    if not text or not parsed:
        return None
    return (normalize_name(text), parsed.isoformat())


def _empty_month_bucket() -> Dict[str, Any]:
    return {
        "name": "",
        "bday": None,
        "last_visit": None,
        "114_cnt": 0.0,
        "114_cnt_q1": 0.0,
        "115_cnt": 0.0,
        "114_amt": 0.0,
        "115_amt": 0.0,
    }


def _expected_month_buckets(source_dir: Path) -> Dict[Tuple[str, str], Dict[str, Any]]:
    reader = _select_reader(source_dir)
    records = list(reader.collect_data(source_dir).values())
    monthly_rows = _collect_monthly_claim_rows(source_dir, records, reader)
    buckets: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for code, rows in monthly_rows.items():
        year = int(code[:3])
        month = int(code[3:5])
        for row in rows:
            _chart, name, bday, visit_date, count, amount = row
            key = _postprocess_key(name, bday)
            if key is None:
                continue
            bucket = buckets.setdefault(key, _empty_month_bucket())
            bucket["name"] = GENERIC.normalize_text(name)
            bucket["bday"] = GENERIC.parse_date(bday)
            parsed_visit = GENERIC.parse_date(visit_date)
            if parsed_visit and (bucket["last_visit"] is None or parsed_visit > bucket["last_visit"]):
                bucket["last_visit"] = parsed_visit
            count_value = GENERIC.parse_float(count) or 0.0
            amount_value = GENERIC.parse_float(amount) or 0.0
            if year == 114:
                bucket["114_cnt"] += count_value
                bucket["114_amt"] += amount_value
                if month <= 4:
                    bucket["114_cnt_q1"] += count_value
            elif year == 115:
                bucket["115_cnt"] += count_value
                bucket["115_amt"] += amount_value
    return buckets


def _copy_row_style(ws, src_row: int, dst_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.border:
            dst.border = copy(src.border)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.font:
            dst.font = copy(src.font)
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def _sheet_key_rows(ws, name_col: int, bday_col: int, data_start: int = 3) -> Dict[Tuple[str, str], int]:
    result: Dict[Tuple[str, str], int] = {}
    for row in range(data_start, ws.max_row + 1):
        key = _postprocess_key(ws.cell(row, name_col).value, ws.cell(row, bday_col).value)
        if key and key not in result:
            result[key] = row
    return result


def _num_or_blank(value: float) -> Any:
    if not value:
        return None
    return int(round(value)) if abs(value - round(value)) < 1e-9 else value


def _write_member_month_values(ws, row: int, data: Dict[str, Any]) -> None:
    ws.cell(row, 1).value = data["name"]
    ws.cell(row, 2).value = ws.cell(row, 2).value or "D"
    ws.cell(row, 7).value = data["bday"]
    ws.cell(row, 11).value = data["last_visit"]
    ws.cell(row, 12).value = _num_or_blank(data["114_cnt"])
    ws.cell(row, 13).value = _num_or_blank(data["114_amt"])
    ws.cell(row, 14).value = _num_or_blank(data["115_cnt"])
    ws.cell(row, 15).value = _num_or_blank(data["115_amt"])


def _write_doctor_month_values(ws, row: int, data: Dict[str, Any], months_115: int) -> None:
    ws.cell(row, 2).value = data["name"]
    ws.cell(row, 3).value = data["bday"]
    ws.cell(row, 10).value = data["last_visit"]
    ws.cell(row, 11).value = _num_or_blank(data["114_cnt"])
    ws.cell(row, 12).value = _num_or_blank(data["114_cnt_q1"])
    ws.cell(row, 13).value = _num_or_blank(data["115_cnt"])
    ws.cell(row, 14).value = int(round(data["114_amt"] / 12.0)) if data["114_amt"] else None
    ws.cell(row, 15).value = int(round(data["115_amt"] / float(max(months_115, 1)))) if data["115_amt"] else None


def _postprocess_output(output_path: Path, source_dir: Path) -> None:
    expected = _expected_month_buckets(source_dir)
    if not expected:
        return

    months_115 = len(
        {
            int(code[3:5])
            for code in _collect_monthly_claim_rows(source_dir, list(collect_data(source_dir).values()), _select_reader(source_dir))
            if code.startswith("115")
        }
    ) or 1

    wb = openpyxl.load_workbook(output_path)
    ws_member = wb["會員總表"]
    doctor_name = next((name for name in wb.sheetnames if name.startswith("醫生看")), None)
    if not doctor_name:
        wb.save(output_path)
        return
    ws_doctor = wb[doctor_name]

    member_rows = _sheet_key_rows(ws_member, 1, 7)
    doctor_rows = _sheet_key_rows(ws_doctor, 2, 3)
    member_template_row = ws_member.max_row
    doctor_template_row = ws_doctor.max_row

    for key, data in expected.items():
        member_row = member_rows.get(key)
        if member_row is None:
            member_row = ws_member.max_row + 1
            _copy_row_style(ws_member, member_template_row, member_row)
            member_rows[key] = member_row
        _write_member_month_values(ws_member, member_row, data)

        doctor_row = doctor_rows.get(key)
        if doctor_row is None:
            doctor_row = ws_doctor.max_row + 1
            _copy_row_style(ws_doctor, doctor_template_row, doctor_row)
            doctor_rows[key] = doctor_row
        _write_doctor_month_values(ws_doctor, doctor_row, data, months_115)

    wb.save(output_path)


def process_excel(source_path: str, template_path: Optional[str] = None) -> str:
    source_dir = Path(source_path).expanduser().resolve()
    if not source_dir.is_dir():
        raise ValueError("請選擇台北公園諾貝爾展望雲端資料夾。")

    template = template_path or GENERIC._find_template(str(SCRIPT_DIR))
    if not os.path.exists(template):
        raise ValueError(f"找不到模板檔：{template}")

    temp_root = Path(tempfile.mkdtemp(prefix="zhanwang_clean_"))
    clean_dir = temp_root / source_dir.name
    clean_dir.mkdir(parents=True, exist_ok=True)
    try:
        _write_clean_source(source_dir, clean_dir)
        temp_output = Path(GENERIC.process_excel(str(clean_dir), template))
        _postprocess_output(temp_output, source_dir)
        final_output = source_dir.parent / temp_output.name
        if final_output.exists():
            final_output.unlink()
        shutil.move(str(temp_output), str(final_output))
        return str(final_output)
    finally:
        shutil.rmtree(temp_root, ignore_errors=True)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = list(argv if argv is not None else sys.argv[1:])
    if args:
        out = process_excel(args[0], args[1] if len(args) > 1 else None)
        print(f"已輸出：{out}")
        return 0

    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()
    src = filedialog.askdirectory(title="選擇台北公園諾貝爾展望雲端資料夾")
    if not src:
        return 1
    try:
        out = process_excel(src)
        messagebox.showinfo("完成", f"已輸出：\n{out}")
        GENERIC.open_file_cross_platform(out)
    except Exception as exc:
        messagebox.showerror("錯誤", str(exc))
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
